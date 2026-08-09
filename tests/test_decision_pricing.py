"""Operator-decision pricing → Lines(SO) unit price, across all three paths.

Inclusion rule (per spec):
  * INCLUDE  = "Include (their CP)"  -> vendor CP lands in Lines(SO)
  * OVERRIDE = "Include (our CP)"    -> operator CP lands in Lines(SO)
  * EXCLUDE                           -> line dropped from Lines(SO)

Vendor CP ("their CP") is basis-aware: fob_price on a cost basis, else
ref_fob_price. Covers lines_store.build_lines (DB record), Processor._vendor_cp
+ _apply_decisions (D365 package) and Processor._finalize_lines_so (Completed
workbook). No engine, no DB — fakes only.
"""

from types import SimpleNamespace

import openpyxl

from online_b2b.services import lines_store
from online_b2b.services.engine_bridge import Processor


def _so(**kw):
    base = dict(
        po_number='PO1', source_location='Loc A', item_no='IT1', ean='890123',
        description='EpiSense-like', qty=10, gst_code='GST18', forced_unit_price=None,
        cost_price_ref=100.0,        # our CP
        fob_price=120.0,             # vendor value on the marketplace price column
        ref_fob_price=110.0,         # reference vendor value
        applied_margin_pct=None, mrp=200.0, vendor_mrp=200.0, diffn=20.0,
        validation_status='MISMATCH', exception_label='',
    )
    base.update(kw)
    return SimpleNamespace(**base)


def _result(rows, basis='cost'):
    return SimpleNamespace(rows=rows, marketplace='Blink', margin_pct=0.70,
                           compare_basis=basis, output_type='so')


def _proc(result):
    p = Processor.__new__(Processor)       # skip __init__ — methods only need result
    p.result = result
    return p


KEY = 'PO1|IT1|890123'

# ── lines_store.build_lines (DB record unit_price) ──────────────────────────

def test_include_records_their_cp_cost_basis():
    res = _result([_so()], basis='cost')             # their CP = fob_price = 120
    ln = lines_store.build_lines(res, run_id=1,
                                 actions={KEY: {'action': 'INCLUDE', 'override_cp': ''}})[0]
    assert ln['unit_price'] == 120.0


def test_include_records_their_cp_landing_basis():
    res = _result([_so()], basis='landing')          # their CP = ref_fob_price = 110
    ln = lines_store.build_lines(res, run_id=1,
                                 actions={KEY: {'action': 'INCLUDE', 'override_cp': ''}})[0]
    assert ln['unit_price'] == 110.0


def test_override_records_our_cp():
    res = _result([_so()], basis='cost')
    ln = lines_store.build_lines(res, run_id=1,
                                 actions={KEY: {'action': 'OVERRIDE', 'override_cp': '95.5'}})[0]
    assert ln['unit_price'] == 95.5


def test_no_decision_records_engine_price():
    res = _result([_so()], basis='cost')
    ln = lines_store.build_lines(res, run_id=1)[0]
    assert ln['unit_price'] == 100.0                 # cost_price_ref (engine)


# ── Processor._vendor_cp (basis-aware "their CP") ───────────────────────────

def test_vendor_cp_basis_aware():
    so = _so()
    assert _proc(_result([so], basis='cost'))._vendor_cp(so) == 120.0
    assert _proc(_result([so], basis='landing'))._vendor_cp(so) == 110.0


# ── Processor._apply_decisions (D365 package via forced_unit_price) ──────────

def test_apply_decisions_include_override_exclude():
    rows = [
        _so(po_number='PO1'),                                    # INCLUDE
        _so(po_number='PO2'),                                    # OVERRIDE
        _so(po_number='PO3'),                                    # EXCLUDE
        _so(po_number='PO4', validation_status='OK'),            # untouched
    ]
    actions = {
        'PO1|IT1|890123': {'action': 'INCLUDE', 'override_cp': ''},
        'PO2|IT1|890123': {'action': 'OVERRIDE', 'override_cp': '95.5'},
        'PO3|IT1|890123': {'action': 'EXCLUDE'},
    }
    out = _proc(_result(rows, basis='cost'))._apply_decisions(actions)
    by_po = {so.po_number: so for so in out.rows}
    assert 'PO3' not in by_po                                    # excluded → dropped
    assert by_po['PO1'].forced_unit_price == 120.0              # their CP
    assert by_po['PO2'].forced_unit_price == 95.5              # our CP
    assert getattr(by_po['PO4'], 'forced_unit_price', None) is None  # untouched


# ── Processor._finalize_lines_so (Completed workbook Lines(SO)) ──────────────

def _make_lines_wb(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lines (SO)'
    ws.append(['Document Type', 'Document No.', 'Line No.', 'Type', 'No.',
               'Location Code', 'Quantity', 'Unit Price'])
    for i, (po, item, up) in enumerate(rows, 1):
        ws.append(['Order', po, i * 10000, 'Item', item, 'PICK', 10, up])
    wb.save(path)
    wb.close()


def test_finalize_lines_so_reprices_and_drops(tmp_path):
    path = str(tmp_path / 'completed.xlsx')
    _make_lines_wb(path, [('PO1', 'IT1', 100.0),      # INCLUDE → 120 (their CP)
                          ('PO2', 'IT1', 100.0),      # OVERRIDE → 95.5 (our CP)
                          ('PO3', 'IT1', 100.0)])     # EXCLUDE → dropped
    rows = [_so(po_number='PO1'), _so(po_number='PO2'), _so(po_number='PO3')]
    actions = {
        'PO1|IT1|890123': {'action': 'INCLUDE', 'override_cp': ''},
        'PO2|IT1|890123': {'action': 'OVERRIDE', 'override_cp': '95.5'},
        'PO3|IT1|890123': {'action': 'EXCLUDE'},
    }
    _proc(_result(rows, basis='cost'))._finalize_lines_so(path, actions)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Lines (SO)']
    got = {ws.cell(r, 2).value: ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}
    wb.close()
    assert 'PO3' not in got                # excluded dropped
    assert got['PO1'] == 120.0             # their CP
    assert got['PO2'] == 95.5             # our CP
