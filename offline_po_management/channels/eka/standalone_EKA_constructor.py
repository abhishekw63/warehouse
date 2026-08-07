"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                  RENEE PO PROCESSOR — EKA Script (v1.5.8)                    ║
║                  Simple Tkinter GUI Desktop Application                       ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  Author  : Agami AI / Vishal                                                 ║
║  Version : 1.5.8 (PWP_EAN_MAP single-EAN lookup layer)                       ║
║  Stack   : Python 3.13, Tkinter, pandas, openpyxl                           ║
╚═══════════════════════════════════════════════════════════════════════════════╝

═══════════════════════════════════════════════════════════════════════════════
  CHANGELOG
═══════════════════════════════════════════════════════════════════════════════

  v1.5.8 — PWP single-EAN mapping layer (this version)
  ────────────────────────────────────────────────────────────────
    ┌───────────────────────────┬───────────────────────────────────────┐
    │  CHANGE                   │  IMPACT                                │
    ├───────────────────────────┼───────────────────────────────────────┤
    │  New PWP_EAN_MAP dict +   │  PWP rows like 'RENEE BOLLYWOOD FILTER│
    │  whitespace normalizer    │  BLURRING SETTING SPRAY 5ML (NFS)' now│
    │  (_norm_pwp_name)         │  resolve to a real Item No (201432)   │
    │                           │  instead of dumping the long name into│
    │                           │  the Item No column. Adding a new     │
    │                           │  single-EAN PWP entry = one line in   │
    │                           │  PWP_EAN_MAP. Non-breaking spaces     │
    │                           │  (\\xa0) in source files are handled   │
    │                           │  automatically.                        │
    ├───────────────────────────┼───────────────────────────────────────┤
    │  Actionable warning on    │  When a PWP name doesn't match any    │
    │  unknown PWP names        │  list, the log + Processing Log sheet │
    │                           │  now shows the EXACT line to paste    │
    │                           │  into PWP_EAN_MAP: 'NAME': '<EAN>'.   │
    │                           │  No more hunting for the right spelling│
    │                           │  when whitespace is involved.          │
    ├───────────────────────────┼───────────────────────────────────────┤
    │  Validation includes      │  PWP names that ARE in PWP_EAN_MAP no │
    │  PWP_EAN_MAP keys in the  │  longer get flagged as Unknown during │
    │  'known' set              │  pre-processing validation. Cleaner    │
    │                           │  audit output.                         │
    └───────────────────────────┴───────────────────────────────────────┘

  v1.5.7 — Multicolor paper bag Non-Stock entries
  ────────────────────────────────────────────────────────────────
    Added 'Renee Multicolor Bag (Big)' → Item 300074 (EAN 8904473105984)
    and '... (Small)' → Item 300075 (EAN 8904473105991) to
    NON_STOCK_EAN_MAP.

  v1.3 — Simplified GUI
  v1.2 — Auto-fill + D365 TO Export + Tester TO Uniqueness
  v1.1 — Special Order Engine + Full Documentation
  v1.0 — Initial release (23 PO files tested)

═══════════════════════════════════════════════════════════════════════════════
  AUTO-LOAD CONVENTION
═══════════════════════════════════════════════════════════════════════════════

  At startup the GUI looks for two files in a folder named
  ``Calculation Data/`` next to this script:

      Calculation Data/
          Items_March.xlsx   ← product master (auto-loaded)
          EKA_DATA.xlsx      ← location registry (auto-loaded)

  When found, the relevant picker shows ``✓ <name> (auto-loaded)``
  and the user just adds PO files (Standalone) or selects the
  Special Order file (Special Order) and clicks Generate.

Requirements:
    pip install pandas openpyxl

Run:
    python renee_po_processor.py
"""

# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

import io
import os
import re
import shutil
import sys
import threading
import time
import warnings
import zipfile
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# v1.5.2: openpyxl emits a UserWarning when reading XLSX files that
# contain Conditional Formatting extensions (color rules, data bars,
# etc.). The data still loads fine — only the conditional-formatting
# metadata is dropped. The warning otherwise looks alarming on the
# terminal and made one user think the app had crashed. Suppress it.
warnings.filterwarnings(
    'ignore', category=UserWarning, module='openpyxl',
)

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

APP_VERSION = "1.5.8"
APP_TITLE = f"RENEE PO Processor v{APP_VERSION}"

# Bundled-file location convention. Looked up RELATIVE to the script
# directory so the .exe / .py both resolve correctly.
BUNDLED_DATA_FOLDER = "Calculation_Data_EKA"
BUNDLED_MASTER_NAME = "Items_March.xlsx"
BUNDLED_EKA_NAME = "EKA_DATA.xlsx"

# Shared Item Master source — the OneDrive "Online_B2B_Dump_Compilation"
# workbook (its 'Item Master' sheet), so EKA uses the SAME master as the
# Online PO tool instead of a separate Items_March copy. Tried in order;
# falls back to the bundled Items_March.xlsx when none are present.
ONLINEB2B_MASTER_PATHS = [
    r"D:\OneDrive - RENEE COSMETICS PRIVATE LIMITED\Online_B2B_Dump_Compilation.xlsx",
]


def get_script_dir() -> Path:
    """
    Return the directory containing this script.

    Resolves correctly for both Python source runs and PyInstaller-frozen
    executables (where ``__file__`` may not exist).
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def get_bundled_master_path() -> Optional[Path]:
    """
    Return the Item Master path. Prefers the shared OnlineB2B dump
    compilation (its 'Item Master' sheet — read by ``load_master``); falls
    back to the bundled ``Items_March.xlsx`` when that file isn't present.
    """
    for src in ONLINEB2B_MASTER_PATHS:
        if os.path.exists(src):
            return Path(src)
    p = get_script_dir() / BUNDLED_DATA_FOLDER / BUNDLED_MASTER_NAME
    return p if p.exists() else None


def get_bundled_eka_path() -> Optional[Path]:
    """Return the path to the bundled EKA_DATA, or None if absent."""
    p = get_script_dir() / BUNDLED_DATA_FOLDER / BUNDLED_EKA_NAME
    return p if p.exists() else None


def get_bundled_folder(create: bool = False) -> Path:
    """Return the bundled-data folder path; create on demand."""
    p = get_script_dir() / BUNDLED_DATA_FOLDER
    if create:
        p.mkdir(parents=True, exist_ok=True)
    return p

# ── Shared engine (extracted to a Tkinter-free module; single source of truth).
#    Bootstrap the repo root so this script still runs directly from source.
try:
    _REPO_ROOT = Path(__file__).resolve().parents[3]
    if str(_REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(_REPO_ROOT))
except Exception:
    pass
from offline.services.eka_engine import (  # noqa: E402
    read_file_bytes_shared,
    OutputRow, LocationResult, POEngine, ExcelWriter, SpecialOrderEngine,
    D365TOExporter, D365SOExporter,
    EKA_SEGMENT, EKA_MARKETPLACE, build_eka_order_rows,
)


class UpdateDialog:
    """Modal dialog asking which bundled file(s) to update."""

    def __init__(self, parent, folder: Path):
        self.parent = parent
        self.folder = folder
        self.result: Optional[str] = None

        self.top = tk.Toplevel(parent)
        self.top.title("Update Bundled Files")
        self.top.resizable(False, False)
        self.top.transient(parent)
        self.top.grab_set()

        self._build_ui()

        self.top.update_idletasks()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        w = self.top.winfo_width()
        h = self.top.winfo_height()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.top.geometry(f"+{max(0, x)}+{max(0, y)}")

    def _build_ui(self) -> None:
        tk.Label(
            self.top, text="Which file would you like to update?",
            font=("Arial", 11, "bold"),
        ).pack(padx=20, pady=(16, 6))

        tk.Label(
            self.top,
            text=f"Files will be saved to:\n{self.folder}",
            font=("Arial", 9), fg='gray', justify='center',
        ).pack(padx=20, pady=(0, 12))

        self.choice_var = tk.StringVar(value='master')

        radio_frame = tk.Frame(self.top)
        radio_frame.pack(padx=30, pady=(0, 12), anchor='w')

        tk.Radiobutton(
            radio_frame, text="Items Master  (Items_March.xlsx)",
            variable=self.choice_var, value='master',
            font=("Arial", 10),
        ).pack(anchor='w', pady=2)
        tk.Radiobutton(
            radio_frame, text="EKA_DATA  (EKA_DATA.xlsx)",
            variable=self.choice_var, value='eka',
            font=("Arial", 10),
        ).pack(anchor='w', pady=2)
        tk.Radiobutton(
            radio_frame, text="Both  (pick Master first, then EKA_DATA)",
            variable=self.choice_var, value='both',
            font=("Arial", 10),
        ).pack(anchor='w', pady=2)

        btn_frame = tk.Frame(self.top)
        btn_frame.pack(pady=(0, 16))

        tk.Button(
            btn_frame, text="Cancel", width=10,
            command=self._on_cancel,
        ).pack(side='left', padx=6)
        tk.Button(
            btn_frame, text="Update", width=10,
            font=("Arial", 10, "bold"),
            bg="#00C853", fg='white',
            command=self._on_update,
        ).pack(side='left', padx=6)

        self.top.bind('<Escape>', lambda _e: self._on_cancel())
        self.top.protocol('WM_DELETE_WINDOW', self._on_cancel)

    def _on_update(self) -> None:
        self.result = self.choice_var.get()
        self.top.destroy()

    def _on_cancel(self) -> None:
        self.result = None
        self.top.destroy()

    def show(self) -> Optional[str]:
        self.parent.wait_window(self.top)
        return self.result


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED HISTORY DB RECORDING  (Push to DB)
# ═══════════════════════════════════════════════════════════════════════════════
#
# EKA orders are recorded into the SAME shared history DB the Online tool
# (and the other offline channels) write to, so one consolidated tracker
# spans everything. Mapping:
#     segment           = 'Offline'
#     marketplace       = 'EKA'
#     marketplace_label = 'EKA'
#     po                = the order number (TO/.. or SO/..) — one DB row per
#                         distinct order number across all locations/types
#     order_type        = 'TO' or 'SO' (from the number prefix)
# Recording is a deliberate, separate step (the "Push to DB" button): the
# operator generates + verifies the output first, then pushes. If the DB
# layer can't be located/reached, the push soft-fails with a message and
# never affects the generated output.

# EKA_SEGMENT / EKA_MARKETPLACE + build_eka_order_rows now live in the shared
# eka_engine module (imported above) — single source of truth for desktop + web.


def _find_online_history_db():
    """Locate + import the Online tool's ``history_db`` module by walking up
    to ``online_po_management``. Returns the module or None."""
    import sys as _sys
    here = get_script_dir().resolve()
    for base in [here, *here.parents]:
        cand = base / 'online_po_management'
        if (cand / 'online_po_processor' / 'auto' / 'history_db.py').exists():
            if str(cand) not in _sys.path:
                _sys.path.insert(0, str(cand))
            try:
                import online_po_processor.auto.history_db as H
                return H
            except Exception:
                return None
    return None


def _load_eka_type_map() -> dict:
    """
    {code → Type} from EKA_DATA, where Type is the segment label shown in
    the DB (Airport / EBO / Kiosk). Keyed by Location, Transfer Code AND
    Short Code (all upper-cased) so an order resolves whether we match on
    its transfer-to code or its SO/TO short code. {} if EKA_DATA missing.
    """
    import io
    import pandas as pd
    p = get_bundled_eka_path()
    out: dict = {}
    if not p or not Path(p).exists():
        return out
    try:
        df = pd.read_excel(io.BytesIO(read_file_bytes_shared(str(p))),
                           dtype=str)
    except Exception:
        return out
    cols = {str(c).strip().lower(): c for c in df.columns}
    tcol = cols.get('type')
    if not tcol:
        return out
    key_cols = [cols[k] for k in ('location', 'transfer code', 'short code')
                if k in cols]
    for _, r in df.iterrows():
        t = str(r[tcol] or '').strip()
        if not t:
            continue
        for kc in key_cols:
            v = str(r[kc] or '').strip()
            if v:
                out[v.upper()] = t
    return out


# build_eka_order_rows moved to offline.services.eka_engine (imported above).


def eka_sql_preview(rows: list) -> list:
    """Human-readable INSERT statements for what will be written — shown in
    the EKA log so the operator can see the SQL going to the DB."""
    cols = ['segment', 'marketplace', 'marketplace_label', 'po', 'location',
            'warehouse', 'po_date', 'exp_date', 'order_type', 'items', 'qty',
            'order_value']
    out = []
    for o in rows:
        vals = ', '.join(
            (repr(o[c]) if isinstance(o[c], str) else str(o[c]))
            for c in cols)
        out.append(f"INSERT INTO order_headers ({', '.join(cols)}) "
                   f"VALUES ({vals});")
    return out


def record_eka_batch(results, output_file: str = '') -> dict:
    """Record the last generated EKA batch into the shared history DB
    (Offline segment). New order numbers only; soft-fails (never raises)."""
    H = _find_online_history_db()
    if H is None:
        return {'recorded': False, 'reason': 'history_db module not found'}
    from datetime import date
    rows = build_eka_order_rows(
        results, output_file, type_map=_load_eka_type_map(),
        po_date=date.today().isoformat(), warehouse='AHD')
    if not rows:
        return {'recorded': False, 'reason': 'no orders to record'}
    try:
        from datetime import datetime
        db_path = H.default_history_db_path()
        store = H.get_history_store(db_path)
        try:
            existing = store.existing_pos()
        finally:
            store.close()
        new_rows = [r for r in rows
                    if (r['marketplace'], r['po']) not in existing]
        skipped = len(rows) - len(new_rows)
        if not new_rows:
            return {'recorded': False, 'reason': 'all orders already recorded',
                    'skipped': skipped}
        run_meta = {
            'run_ts': datetime.now().isoformat(timespec='seconds'),
            'mode': 'MANUAL',
            'online_root': (f'OFFLINE EKA: {output_file}'
                            if output_file else 'OFFLINE EKA'),
            'marketplaces': 1,
            'total_pos':   len(new_rows),
            'total_items': sum(r['items'] for r in new_rows),
            'total_qty':   sum(r['qty'] for r in new_rows),
            'total_value': sum(r['order_value'] for r in new_rows),
            'consolidated_path': '',
            'tracker_path': '',
        }
        res = H._record(new_rows, run_meta, db_path, skipped=skipped)
        return {'recorded': True, 'skipped': skipped, **res}
    except Exception as e:   # noqa: BLE001 — never block on the DB
        return {'recorded': False, 'reason': f'DB error: {e}'}


# ═══════════════════════════════════════════════════════════════════════════════
#  GUI APPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

class ReneePOApp:
    """Simple Tkinter app for RENEE PO Processor."""

    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("560x780")
        self.root.resizable(False, False)

        self.engine = POEngine()
        self.eka_locations: List[Dict] = []

        self.master_path: Optional[str] = None
        self.eka_path: Optional[str] = None
        self.po_files: List[str] = []
        self.so_path: Optional[str] = None

        self.last_output: Optional[Path] = None
        self.last_results: List[LocationResult] = []
        self.is_running = False

        self.master_is_bundled = False
        self.eka_is_bundled = False

        self.mode_var: tk.StringVar
        self.master_var: tk.StringVar
        self.eka_var: tk.StringVar
        self.po_files_var: tk.StringVar
        self.so_var: tk.StringVar
        self.status_var: tk.StringVar
        self.po_listbox: tk.Listbox
        self.log_text: tk.Text
        self.gen_btn: tk.Button
        self.open_btn: tk.Button
        self.d365_btn: tk.Button
        self.po_frame: tk.LabelFrame
        self.so_frame: tk.LabelFrame

        self._build_ui()

        get_bundled_folder(create=True)

        self._auto_load_bundled()

        self.root.update()
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after(500, lambda: self.root.attributes('-topmost', False))

    def _build_ui(self) -> None:
        tk.Label(self.root, text="RENEE PO Processor",
                  font=("Arial", 14, "bold")).pack(pady=(12, 2))
        tk.Label(self.root, text=f"v{APP_VERSION}  •  EBO / Kiosk / Airport PO → SO/TO",
                  font=("Arial", 9), fg='gray').pack(pady=(0, 10))

        mode_frame = tk.Frame(self.root)
        mode_frame.pack(fill='x', padx=20, pady=(0, 8))
        tk.Label(mode_frame, text="Mode:", font=("Arial", 10, "bold")).pack(side='left')

        self.mode_var = tk.StringVar(value='standalone')
        tk.Radiobutton(
            mode_frame, text="Standalone PO files",
            variable=self.mode_var, value='standalone',
            command=self._on_mode_change,
            font=("Arial", 10),
        ).pack(side='left', padx=(8, 4))
        tk.Radiobutton(
            mode_frame, text="Special Order broadcast",
            variable=self.mode_var, value='special',
            command=self._on_mode_change,
            font=("Arial", 10),
        ).pack(side='left')

        files_frame = tk.LabelFrame(
            self.root, text="Input Files",
            font=("Arial", 10, "bold"), padx=10, pady=8,
        )
        files_frame.pack(fill='x', padx=20, pady=(0, 8))

        self.master_var = tk.StringVar(value="Not selected")
        self._build_file_row(
            files_frame, "Items Master:",
            self.master_var, self._select_master,
        )

        self.eka_var = tk.StringVar(value="Not selected")
        self._build_file_row(
            files_frame, "EKA Data:",
            self.eka_var, self._select_eka,
        )

        self.po_frame = tk.LabelFrame(
            self.root, text="PO Files (Standalone)",
            font=("Arial", 10, "bold"), padx=10, pady=8,
        )
        self.po_frame.pack(fill='both', expand=False, padx=20, pady=(0, 8))

        po_btn_row = tk.Frame(self.po_frame)
        po_btn_row.pack(fill='x', pady=(0, 4))
        tk.Button(po_btn_row, text="+ Add Files",
                   command=self._add_files, width=12).pack(side='left', padx=(0, 4))
        tk.Button(po_btn_row, text="✕ Clear All",
                   command=self._clear_files, width=10).pack(side='left')

        self.po_files_var = tk.StringVar(value="0 files added")
        tk.Label(po_btn_row, textvariable=self.po_files_var,
                  font=("Arial", 9), fg='gray').pack(side='right')

        list_frame = tk.Frame(self.po_frame)
        list_frame.pack(fill='both', expand=True)
        list_scroll = ttk.Scrollbar(list_frame, orient='vertical')
        list_scroll.pack(side='right', fill='y')
        self.po_listbox = tk.Listbox(
            list_frame, height=5, font=("Consolas", 9),
            yscrollcommand=list_scroll.set,
        )
        self.po_listbox.pack(fill='both', expand=True)
        list_scroll.config(command=self.po_listbox.yview)

        self.so_frame = tk.LabelFrame(
            self.root, text="Special Order File",
            font=("Arial", 10, "bold"), padx=10, pady=8,
        )

        self.so_var = tk.StringVar(value="Not selected")
        self._build_file_row(
            self.so_frame, "Special Order:",
            self.so_var, self._select_so,
        )

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=(4, 4))

        self.gen_btn = tk.Button(
            btn_frame, text="▶  Generate Output", width=24,
            font=("Arial", 10, "bold"),
            bg="#00C853", fg='white', command=self._run,
        )
        self.gen_btn.pack(pady=3)

        self.open_btn = tk.Button(
            btn_frame, text="📂  Open Last Output", width=24,
            state=tk.DISABLED, command=self._open_last,
        )
        self.open_btn.pack(pady=3)

        self.d365_btn = tk.Button(
            btn_frame, text="📤  Export D365 TO Package", width=24,
            state=tk.DISABLED, command=self._export_d365,
        )
        self.d365_btn.pack(pady=3)

        # Push to DB — record the generated orders into the shared history
        # DB (Offline / EKA). Enabled after a successful Generate; a
        # deliberate separate step (verify the output first, then push).
        self.push_db_btn = tk.Button(
            btn_frame, text="⤓  Push to DB", width=24,
            bg="#2563eb", fg='white', font=("Arial", 10, "bold"),
            state=tk.DISABLED, command=self._push_to_db,
        )
        self.push_db_btn.pack(pady=3)

        tmpl_btn = tk.Menubutton(
            btn_frame, text="📋  Download Templates ▾",
            width=22, relief='raised',
        )
        tmpl_menu = tk.Menu(tmpl_btn, tearoff=0)
        tmpl_btn['menu'] = tmpl_menu
        tmpl_menu.add_command(label="Blank PO Template",
                                command=self._download_po_template)
        tmpl_menu.add_command(label="EKA_DATA Template",
                                command=self._download_eka_template)
        tmpl_menu.add_command(label="Special Order Template",
                                command=self._download_so_template)
        tmpl_btn.pack(pady=3)

        tk.Button(btn_frame, text="📁  Update Bundled Files", width=24,
                   command=self._update_bundled).pack(pady=3)

        self.status_var = tk.StringVar(value="Ready")
        tk.Label(self.root, textvariable=self.status_var,
                  font=("Arial", 10), fg='gray',
                  wraplength=520).pack(pady=4)

        log_frame = tk.LabelFrame(self.root, text="Log", font=("Arial", 9))
        log_frame.pack(fill='both', expand=True, padx=20, pady=(0, 12))

        log_scroll = ttk.Scrollbar(log_frame, orient='vertical')
        log_scroll.pack(side='right', fill='y')
        self.log_text = tk.Text(
            log_frame, height=8, font=("Consolas", 9),
            state='disabled', wrap='word',
            yscrollcommand=log_scroll.set,
        )
        self.log_text.pack(fill='both', expand=True)
        log_scroll.config(command=self.log_text.yview)

        self.log_text.tag_config('ok', foreground='#00A651')
        self.log_text.tag_config('err', foreground='#D32F2F')
        self.log_text.tag_config('warn', foreground='#E65100')
        self.log_text.tag_config('inf', foreground='#0077B6')
        self.log_text.tag_config('dim', foreground='#5F6368')

    def _build_file_row(self, parent, label_text: str,
                          var: tk.StringVar, browse_cmd) -> None:
        row = tk.Frame(parent)
        row.pack(fill='x', pady=2)

        tk.Label(
            row, text=label_text,
            font=("Arial", 10), width=15, anchor='w',
        ).pack(side='left')

        tk.Button(
            row, text="Browse", width=8, command=browse_cmd,
        ).pack(side='right', padx=(4, 0))

        tk.Label(
            row, textvariable=var,
            font=("Arial", 10), fg='#1A237E', anchor='w',
            wraplength=300, justify='left',
        ).pack(side='left', fill='x', expand=True)

    def _auto_load_bundled(self) -> None:
        m_path = get_bundled_master_path()
        if m_path:
            try:
                count = self.engine.load_master(str(m_path))
                self.master_path = str(m_path)
                self.master_is_bundled = True
                self.master_var.set(f"✓ {m_path.name} (auto-loaded)")
                self._log(f"Auto-loaded master from "
                          f"{BUNDLED_DATA_FOLDER}/{m_path.name} → {count:,} items", 'ok')
            except Exception as e:
                self._log(f"Master auto-load failed: {e}", 'err')
        else:
            self._log(f"No bundled master at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_MASTER_NAME} "
                      f"— Browse to pick one or use 'Update Bundled Files'", 'dim')

        e_path = get_bundled_eka_path()
        if e_path:
            self._load_eka_file(str(e_path), is_bundled=True)
        else:
            self._log(f"No bundled EKA_DATA at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_EKA_NAME} "
                      f"— Browse to pick one or use 'Update Bundled Files'", 'dim')

    def _load_eka_file(self, path: str, is_bundled: bool) -> None:
        try:
            so_engine = SpecialOrderEngine(self.engine.master or {})
            logs = []
            count = so_engine.load_eka_data(path, logs)
            self.eka_locations = so_engine.locations
            self.eka_path = path
            self.eka_is_bundled = is_bundled

            label = (f"✓ {os.path.basename(path)} (auto-loaded)"
                     if is_bundled
                     else os.path.basename(path))
            self.eka_var.set(label)

            tag = 'ok' if is_bundled else 'inf'
            self._log(f"EKA_DATA loaded: {os.path.basename(path)} "
                      f"→ {count} active locations", tag)

            for level, msg in logs:
                if level != 'info':
                    log_tag = {'warn': 'warn', 'error': 'err',
                               'alert': 'warn'}.get(level, 'dim')
                    self._log(f"  {msg}", log_tag)
        except Exception as e:
            self.eka_var.set("ERROR — see log")
            self._log(f"EKA load failed: {e}", 'err')
            self._log(f"  Path attempted: {path}", 'dim')
            if not is_bundled:
                messagebox.showerror(
                    "EKA Load Failed",
                    f"Could not load EKA_DATA:\n\n{e}\n\n"
                    f"File: {os.path.basename(path)}",
                )

    def _log(self, msg: str, tag: str = 'dim') -> None:
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n", tag)
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _set_status(self, msg: str, color: str = 'gray') -> None:
        self.status_var.set(msg)

    def _on_mode_change(self) -> None:
        mode = self.mode_var.get()
        if mode == 'standalone':
            self.so_frame.pack_forget()
            self.po_frame.pack(
                fill='both', expand=False, padx=20, pady=(0, 8),
                before=self.gen_btn.master,
            )
            self._log("Mode: Standalone (per-location PO files)", 'inf')
        else:
            self.po_frame.pack_forget()
            self.so_frame.pack(
                fill='both', expand=False, padx=20, pady=(0, 8),
                before=self.gen_btn.master,
            )
            self._log("Mode: Special Order (broadcast across all locations)", 'inf')

    def _select_master(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Items Master file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            count = self.engine.load_master(path)
            self.master_path = path
            self.master_var.set(os.path.basename(path))
            self._log(f"Master loaded: {os.path.basename(path)} → "
                      f"{count:,} items", 'ok')

            try:
                target = get_bundled_folder(create=True) / BUNDLED_MASTER_NAME
                if Path(path).resolve() != target.resolve():
                    shutil.copy2(path, str(target))
                    self.master_path = str(target)
                    self.master_is_bundled = True
                    self.master_var.set(f"✓ {target.name} (auto-loaded)")
                    self._log(f"Saved to bundled folder → "
                              f"{BUNDLED_DATA_FOLDER}/{target.name} "
                              f"(will auto-load next time)", 'ok')
                else:
                    self.master_is_bundled = True
            except Exception as copy_err:
                self.master_is_bundled = False
                self._log(f"Could not save to bundled folder: {copy_err}",
                          'warn')
                self._log("  (Master is loaded for this session only — "
                          "use 'Update Bundled Files' to retry)", 'dim')

        except Exception as e:
            self.master_var.set("ERROR — see log")
            self._log(f"Master load failed: {e}", 'err')
            self._log(f"  Path attempted: {path}", 'dim')
            messagebox.showerror(
                "Master Load Failed",
                f"Could not load Items Master:\n\n{e}\n\n"
                f"File: {os.path.basename(path)}",
            )

    def _select_eka(self) -> None:
        path = filedialog.askopenfilename(
            title="Select EKA_DATA file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return

        self._load_eka_file(path, is_bundled=False)

        if self.eka_locations:
            try:
                target = get_bundled_folder(create=True) / BUNDLED_EKA_NAME
                if Path(path).resolve() != target.resolve():
                    shutil.copy2(path, str(target))
                    self.eka_path = str(target)
                    self.eka_is_bundled = True
                    self.eka_var.set(f"✓ {target.name} (auto-loaded)")
                    self._log(f"Saved to bundled folder → "
                              f"{BUNDLED_DATA_FOLDER}/{target.name} "
                              f"(will auto-load next time)", 'ok')
            except Exception as copy_err:
                self._log(f"Could not save to bundled folder: {copy_err}",
                          'warn')

    def _select_so(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Special Order file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        self.so_path = path
        self.so_var.set(os.path.basename(path))
        self._log(f"Special Order: {os.path.basename(path)}", 'inf')

    def _add_files(self) -> None:
        files = filedialog.askopenfilenames(
            title="Select PO Excel Files",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        added = 0
        for f in files:
            if (f not in self.po_files
                    and 'Items_March' not in f
                    and 'EKA_DATA' not in f
                    and 'PO_Output' not in f):
                self.po_files.append(f)
                added += 1
        self._refresh_po_list()
        if added:
            self._log(f"Added {added} PO file(s)", 'inf')

    def _clear_files(self) -> None:
        self.po_files.clear()
        self._refresh_po_list()
        self._log("PO file list cleared", 'dim')

    def _refresh_po_list(self) -> None:
        self.po_listbox.delete(0, 'end')
        for i, p in enumerate(self.po_files, 1):
            name = os.path.basename(p)
            display = name if len(name) <= 60 else name[:57] + '...'
            self.po_listbox.insert('end', f"  {i}. {display}")
        n = len(self.po_files)
        self.po_files_var.set(f"{n} file{'s' if n != 1 else ''} added")

    def _update_bundled(self) -> None:
        target = get_bundled_folder(create=True)
        dialog = UpdateDialog(self.root, folder=target)
        choice = dialog.show()
        if choice is None:
            return

        updated_any = False

        if choice in ('master', 'both'):
            updated_any |= self._do_update_one(
                kind='Items Master',
                title='Select new Items Master file',
                target_path=target / BUNDLED_MASTER_NAME,
            )

        if choice in ('eka', 'both'):
            updated_any |= self._do_update_one(
                kind='EKA_DATA',
                title='Select new EKA_DATA file',
                target_path=target / BUNDLED_EKA_NAME,
            )

        if updated_any:
            self._auto_load_bundled()
            messagebox.showinfo(
                "Bundled Files Updated",
                f"Bundled files updated in:\n{target}\n\n"
                f"Future runs will auto-load the new version.",
            )

    def _do_update_one(self, kind: str, title: str,
                        target_path: Path) -> bool:
        src = filedialog.askopenfilename(
            title=title,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not src:
            self._log(f"Update cancelled for {kind}", 'dim')
            return False
        try:
            shutil.copy2(src, str(target_path))
            self._log(f"Bundled {kind} updated → {target_path}", 'ok')
            return True
        except Exception as e:
            self._log(f"Update failed for {kind}: {e}", 'err')
            messagebox.showerror(
                "Update Failed",
                f"Could not copy {kind}:\n{e}",
            )
            return False

    def _lookup_location_from_filename(
        self, filepath: str,
    ) -> Tuple[Optional[Dict], int]:
        if not self.eka_locations:
            return (None, 0)

        fname = os.path.basename(filepath)
        loc_code = (fname.replace('.xlsx', '')
                          .replace('.xlsm', '')
                          .replace('.xls', ''))

        for loc in self.eka_locations:
            if loc.get('location', '') == loc_code:
                return (loc, 0)

        suffix_match = re.match(r'^(.+)_(\d+)$', loc_code)
        if suffix_match:
            base_code = suffix_match.group(1)
            suffix_idx = int(suffix_match.group(2))
            for loc in self.eka_locations:
                if loc.get('location', '') == base_code:
                    return (loc, suffix_idx)

        for loc in self.eka_locations:
            loc_val = loc.get('location', '')
            if loc_val and loc_code.startswith(loc_val):
                return (loc, 0)

        return (None, 0)

    def _run(self) -> None:
        if self.is_running:
            return
        if not self.master_path or not self.engine.master:
            messagebox.showwarning(
                "No Master",
                "Items Master is not loaded. Pick one via Browse or use "
                "'Update Bundled Files'.",
            )
            return

        mode = self.mode_var.get()
        if mode == 'special':
            self._run_special()
        else:
            self._run_standalone()

    def _run_standalone(self) -> None:
        if not self.po_files:
            messagebox.showwarning("No Files", "Please add at least one PO file.")
            return

        if not self.eka_locations:
            proceed = messagebox.askyesno(
                "EKA_DATA Not Loaded",
                "EKA_DATA is not loaded.\n\n"
                "Without EKA_DATA, these fields will be EMPTY in output:\n"
                "  • TO number\n"
                "  • Transfer-to Code\n"
                "  • Gen. Bus. Posting Group\n"
                "  • Headers (TO) sheet\n\n"
                "Continue anyway?",
            )
            if not proceed:
                return

        output_dir = (Path(self.po_files[0]).parent / 'eka_output')
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = time.strftime("%d%m%Y_%H%M%S")
        output_path = output_dir / f"PO_Output_{timestamp}.xlsx"

        self.is_running = True
        self.gen_btn.config(state=tk.DISABLED)
        self._set_status("Processing...")
        self.root.update()

        threading.Thread(
            target=self._standalone_worker,
            args=(str(output_path),),
            daemon=True,
        ).start()

    def _run_special(self) -> None:
        if not self.eka_path or not self.eka_locations:
            messagebox.showwarning(
                "No EKA_DATA",
                "EKA_DATA is required for Special Order mode.",
            )
            return
        if not self.so_path:
            messagebox.showwarning(
                "No Special Order File",
                "Please select the Special Order file.",
            )
            return

        output_dir = Path(self.so_path).parent / 'eka_output'
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = time.strftime("%d%m%Y_%H%M%S")
        output_path = output_dir / f"SO_Output_{timestamp}.xlsx"

        self.is_running = True
        self.gen_btn.config(state=tk.DISABLED)
        self._set_status("Processing Special Order...")
        self.root.update()

        threading.Thread(
            target=self._special_worker,
            args=(str(output_path),),
            daemon=True,
        ).start()

    def _standalone_worker(self, output_path: str) -> None:
        total = len(self.po_files)

        self.root.after(0, self._log, "─── PHASE 1: VALIDATING ───", 'inf')
        self.root.after(0, self._set_status, "Validating files...")

        if not self.eka_locations:
            self.root.after(
                0, self._log,
                "⚠ EKA_DATA not loaded — TO/Transfer/Posting will be empty.",
                'warn',
            )

        validation_results = {}
        files_with_errors = []
        alert_messages = []

        processing_log: List[Dict] = []

        for i, po_path in enumerate(self.po_files):
            fname = os.path.basename(po_path)
            loc = fname.replace('.xlsx', '')
            self.root.after(0, self._set_status,
                             f"Validating {i+1}/{total}: {fname[:45]}")

            log_entry = {
                'filename': fname,
                'location': loc,
                'status': 'OK',
                'issues': [],
                'actions': [],
                'to_number': '',
                'tt_number': '',
            }
            processing_log.append(log_entry)

            try:
                vlogs = self.engine.validate_file(po_path)
                has_err = any(level == 'error' for level, _ in vlogs)
                validation_results[po_path] = (vlogs, has_err)

                for level, msg in vlogs:
                    tag = {'info': 'inf', 'warn': 'warn',
                           'error': 'err', 'alert': 'warn'}.get(level, 'dim')
                    self.root.after(0, self._log, f"  [{loc}] {msg}", tag)
                    if level == 'alert':
                        alert_messages.append(f"• {loc}: {msg}")
                        log_entry['actions'].append(msg)
                        if log_entry['status'] == 'OK':
                            log_entry['status'] = 'AUTO_FIXED'
                    elif level == 'warn':
                        log_entry['issues'].append(msg)
                        if log_entry['status'] == 'OK':
                            log_entry['status'] = 'WARNING'
                    elif level == 'error':
                        log_entry['issues'].append(msg)
                        log_entry['status'] = 'FAILED'

                if has_err:
                    files_with_errors.append(fname)
                    self.root.after(0, self._log,
                                    f"✗ {loc} → VALIDATION FAILED", 'err')
                else:
                    self.root.after(0, self._log, f"✓ {loc} → OK", 'ok')
            except Exception as e:
                validation_results[po_path] = ([('error', str(e))], True)
                files_with_errors.append(fname)
                self.root.after(0, self._log, f"✗ {loc} → {e}", 'err')
                log_entry['status'] = 'FAILED'
                log_entry['issues'].append(f"Validation crashed: {e}")

        if alert_messages:
            preview = "\n".join(alert_messages[:10])
            self.root.after(0, lambda p=preview: messagebox.showinfo(
                "Auto-Fix Applied",
                f"Column names were auto-fixed:\n\n{p}",
            ))

        processable = [f for f in self.po_files
                        if not validation_results.get(f, ([], True))[1]]
        if not processable:
            self.root.after(0, self._done, None,
                             f"ABORTED — all {total} files failed validation")
            return

        skipped = total - len(processable)
        if skipped > 0:
            self.root.after(0, self._log,
                             f"⚠ Skipping {skipped} file(s) with errors", 'warn')

        self.root.after(0, self._log,
                         f"─── PHASE 2: EXTRACTING ({len(processable)} files) ───",
                         'inf')

        results: List[LocationResult] = []
        total_po = total_tester = total_pwp = total_gwp = total_ns = 0
        total_unmatched = 0

        counter = SpecialOrderEngine.get_today_date_code()

        for i, po_path in enumerate(processable):
            fname = os.path.basename(po_path)
            loc = fname.replace('.xlsx', '')
            self.root.after(0, self._set_status,
                             f"Extracting {i+1}/{len(processable)}: {fname[:45]}")

            try:
                res = self.engine.process_file(po_path)

                log_entry = next(
                    (e for e in processing_log if e['filename'] == fname),
                    None,
                )

                eka_loc, suffix_idx = self._lookup_location_from_filename(po_path)
                if eka_loc:
                    # Document split (per ops requirement):
                    #   Regular doc (month segment, e.g. /06/) = ONLY real
                    #     finished-goods PO items (regular_orders), priced
                    #     at calculated cost.
                    #   Tester doc (/TT/ segment) = every flat ₹0.54 line —
                    #     testers, PWP, GWP and Non-Stock.
                    # PWP used to ride on the regular doc; it now belongs on
                    # the TT doc so the regular order stays purely FG. See the
                    # matching item.to assignment + Summary fallbacks below.
                    has_regular = bool(res.regular_orders)
                    has_tester = bool(res.tester_orders or res.pwp_orders
                                       or res.gwp_orders or res.nonstock_orders)

                    short_code = eka_loc['short_code']
                    if suffix_idx > 0:
                        short_code = f"{short_code}_{suffix_idx + 1}"

                    to_regular = ''
                    to_tester = ''
                    if has_regular:
                        to_regular = SpecialOrderEngine.generate_to_number(
                            eka_loc['prefix'], short_code,
                            is_tester=False, date_code=counter)
                        counter += 1
                    if has_tester:
                        to_tester = SpecialOrderEngine.generate_to_number(
                            eka_loc['prefix'], short_code,
                            is_tester=True, date_code=counter)
                        counter += 1

                    tc = eka_loc['transfer_code']
                    pg = eka_loc['posting_group']

                    res._so_bill_to = eka_loc.get('bill_to', '')
                    res._so_ship_to = eka_loc.get('ship_to', '')

                    for item in res.regular_orders:
                        item.to = to_regular
                        item.transfer_to = tc
                        item.posting_group = pg
                    # PWP is a ₹0.54 line → goes on the TT (tester) doc,
                    # not the regular doc (ops requirement: regular = FG only).
                    for item in res.pwp_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg
                    for item in res.tester_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg
                    for item in res.gwp_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg
                    for item in res.nonstock_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg

                    res.logs.append(('info',
                        f"EKA auto-fill: {eka_loc['location']} → "
                        f"TO:{to_regular or '(none)'} / "
                        f"TT:{to_tester or '(none)'} / "
                        f"Transfer:{tc} / Posting:{pg}"))

                    if log_entry is not None:
                        log_entry['to_number'] = to_regular
                        log_entry['tt_number'] = to_tester
                        if suffix_idx > 0:
                            log_entry['actions'].append(
                                f"Same location as another file — "
                                f"indexed as #{suffix_idx + 1} "
                                f"(short_code: {short_code})")
                            if log_entry['status'] == 'OK':
                                log_entry['status'] = 'AUTO_FIXED'
                else:
                    if self.eka_locations:
                        res.logs.append(('warn',
                            f"EKA: '{loc}' not found in Location column — "
                            f"TO/Transfer/Posting left empty"))
                        if log_entry is not None:
                            log_entry['issues'].append(
                                f"Location '{loc}' not in EKA_DATA — "
                                f"TO/Transfer/Posting left empty")
                            if log_entry['status'] == 'OK':
                                log_entry['status'] = 'WARNING'

                results.append(res)

                if log_entry is not None and res.unmatched:
                    log_entry['issues'].append(
                        f"{len(res.unmatched)} EAN(s) not found in master "
                        f"— see Unmatched EANs sheet")
                    if log_entry['status'] == 'OK':
                        log_entry['status'] = 'WARNING'

                po_q = sum(r.qty for r in res.regular_orders)
                tt_q = sum(r.qty for r in res.tester_orders)
                pw_q = sum(r.qty for r in res.pwp_orders)
                gw_q = sum(r.qty for r in res.gwp_orders)
                ns_q = sum(r.qty for r in res.nonstock_orders)

                total_po += po_q
                total_tester += tt_q
                total_pwp += pw_q
                total_gwp += gw_q
                total_ns += ns_q
                total_unmatched += len(res.unmatched)

                for level, msg in res.logs:
                    tag = {'info': 'inf', 'warn': 'warn',
                           'error': 'err'}.get(level, 'dim')
                    self.root.after(0, self._log, f"  [{loc}] {msg}", tag)

                parts = [f"PO:{po_q}"]
                if tt_q: parts.append(f"T:{tt_q}")
                if pw_q: parts.append(f"PWP:{pw_q}")
                if gw_q: parts.append(f"GWP:{gw_q}")
                if ns_q: parts.append(f"NS:{ns_q}")
                self.root.after(0, self._log,
                                 f"✓ {loc} → {' | '.join(parts)}", 'ok')
            except Exception as e:
                self.root.after(0, self._log, f"✗ {fname} → {e}", 'err')
                log_entry = next(
                    (e2 for e2 in processing_log if e2['filename'] == fname),
                    None,
                )
                if log_entry is not None:
                    log_entry['status'] = 'FAILED'
                    log_entry['issues'].append(f"Extraction crashed: {e}")

        if not results:
            self.root.after(0, self._done, None, "ERROR: No data processed")
            return

        try:
            self.root.after(0, self._set_status, "Writing Excel...")
            self.last_results = results
            ExcelWriter.write(results, output_path,
                               processing_log=processing_log)

            grand = total_po + total_tester + total_pwp + total_gwp + total_ns

            stats_line = (
                f"Locations:{len(results)} | PO Qty:{total_po:,} | "
                f"PO Items:{sum(len(r.regular_orders) for r in results)} | "
                f"Tester:{total_tester:,} | PWP:{total_pwp:,} | "
                f"GWP:{total_gwp:,} | NS:{total_ns:,} | "
                f"GRAND TOTAL:{grand:,} | Unmatched:{total_unmatched}"
            )
            self.root.after(0, self._log, stats_line, 'ok')

            self.root.after(0, self._done, output_path,
                             f"Done — {len(results)} locations, {grand:,} total qty")
        except Exception as e:
            self.root.after(0, self._done, None, f"ERROR: {e}")

    def _special_worker(self, output_path: str) -> None:
        self.root.after(0, self._log, "═══ SPECIAL ORDER ═══", 'inf')

        so_engine = SpecialOrderEngine(self.engine.master)

        logs = []
        loc_count = so_engine.load_eka_data(self.eka_path, logs)
        for level, msg in logs:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err', 'alert': 'warn'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)

        if loc_count == 0:
            self.root.after(0, self._done, None, "ABORTED — EKA_DATA load failed")
            return

        self.root.after(0, self._set_status, "Loading Special Order...")
        logs2 = []
        alert_messages = []
        prod_count = so_engine.load_special_order(self.so_path, logs2)
        for level, msg in logs2:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err', 'alert': 'warn'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)
            if level == 'alert':
                alert_messages.append(msg)

        if prod_count == 0:
            self.root.after(0, self._done, None,
                             "ABORTED — Special Order load failed")
            return

        if alert_messages:
            preview = "\n".join(f"• {m}" for m in alert_messages)
            self.root.after(0, lambda p=preview: messagebox.showinfo(
                "Auto-Fix Applied",
                f"Column names auto-fixed:\n\n{p}",
            ))

        self.root.after(0, self._set_status, "Validating...")
        logs3 = []
        if not so_engine.validate(logs3):
            for level, msg in logs3:
                tag = {'info': 'inf', 'warn': 'warn',
                       'error': 'err'}.get(level, 'dim')
                self.root.after(0, self._log, f"  {msg}", tag)
            self.root.after(0, self._done, None, "ABORTED — Validation failed")
            return

        for level, msg in logs3:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)

        self.root.after(0, self._log, "─── GENERATING OUTPUT ───", 'inf')
        self.root.after(0, self._set_status, "Generating output...")

        logs4 = []
        results = so_engine.process(logs4)
        for level, msg in logs4:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)

        if not results:
            self.root.after(0, self._done, None, "ERROR — No results generated")
            return

        total_po = total_tester = total_unmatched = 0

        processing_log: List[Dict] = []

        for res in results:
            po_q = sum(r.qty for r in res.regular_orders)
            tt_q = sum(r.qty for r in res.tester_orders)
            total_po += po_q
            total_tester += tt_q
            total_unmatched += len(res.unmatched)

            to_reg = res.regular_orders[0].to if res.regular_orders else ''
            to_tt = res.tester_orders[0].to if res.tester_orders else ''
            self.root.after(0, self._log,
                             f"  ✓ {res.filename:<22} PO:{po_q} T:{tt_q}  [{to_reg}]",
                             'ok')

            entry = {
                'filename': res.filename,
                'location': res.filename,
                'status': 'OK',
                'issues': [],
                'actions': [],
                'to_number': to_reg,
                'tt_number': to_tt,
            }
            if res.unmatched:
                entry['issues'].append(
                    f"{len(res.unmatched)} EAN(s) not found in master "
                    f"— see Unmatched EANs sheet")
                entry['status'] = 'WARNING'
            processing_log.append(entry)

        try:
            self.root.after(0, self._set_status, "Writing Excel...")
            self.last_results = results
            ExcelWriter.write(
                results, output_path,
                eka_locations=so_engine.locations,
                master=self.engine.master,
                so_products=so_engine.products,
                processing_log=processing_log,
            )

            grand = total_po + total_tester
            stats_line = (
                f"Locations:{len(results)} | PO Qty:{total_po:,} | "
                f"PO Items:{sum(len(r.regular_orders) for r in results)} | "
                f"Tester:{total_tester:,} | PWP:0 | GWP:0 | NS:0 | "
                f"GRAND TOTAL:{grand:,} | Unmatched:{total_unmatched}"
            )
            self.root.after(0, self._log, stats_line, 'ok')

            self.root.after(
                0, self._done, output_path,
                f"Done — {len(results)} locations × {prod_count} products, "
                f"{grand:,} total qty",
            )
        except Exception as e:
            self.root.after(0, self._done, None, f"ERROR: {e}")

    def _done(self, output_path: Optional[str], status_msg: str) -> None:
        self.is_running = False
        self.gen_btn.config(state=tk.NORMAL)
        self._set_status(status_msg)

        if output_path:
            self.last_output = Path(output_path)
            self.open_btn.config(state=tk.NORMAL)
            self.d365_btn.config(state=tk.NORMAL)
            self.push_db_btn.config(state=tk.NORMAL)
            self._log(f"Saved → {output_path}", 'inf')

            if messagebox.askyesno(
                    "Done!",
                    f"Processing complete!\n\n{status_msg}\n\nOpen output file?"):
                self._open_last()

    def _open_last(self) -> None:
        if not self.last_output or not self.last_output.exists():
            messagebox.showwarning("Not Found", "No output file yet.")
            return
        try:
            if os.name == 'nt':
                os.startfile(str(self.last_output))
            elif os.name == 'posix':
                import subprocess
                opener = ('open' if hasattr(os, 'uname') and
                          os.uname().sysname == 'Darwin' else 'xdg-open')
                subprocess.Popen([opener, str(self.last_output)])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}")

    def _push_to_db(self) -> None:
        """Record the last generated EKA batch into the shared history DB
        (Offline / EKA). One row per order number (TO/SO); new only."""
        if not self.last_results:
            messagebox.showinfo(
                "Push to DB",
                "Generate output first, verify it, then push.")
            return
        from datetime import date
        preview = build_eka_order_rows(
            self.last_results, type_map=_load_eka_type_map(),
            po_date=date.today().isoformat(), warehouse='AHD')
        if not preview:
            messagebox.showinfo("Push to DB",
                                "No order numbers (TO/SO) to record.")
            return
        # Show the SQL that will run, in the log, so the operator can see it.
        self._log("─── Push to DB — SQL preview ───", 'inf')
        for stmt in eka_sql_preview(preview):
            self._log(stmt, 'dim')
        n_to = sum(1 for r in preview if r['order_type'] == 'TO')
        n_so = len(preview) - n_to
        labels = sorted({r['marketplace_label'] for r in preview})
        if not messagebox.askyesno(
                "Push to DB",
                f"Push {len(preview)} order(s) to the history DB?\n\n"
                f"  Transfer Orders: {n_to}\n"
                f"  Sales Orders:    {n_so}\n"
                f"  Segments:        {', '.join(labels)}\n"
                f"  Total qty:       {sum(r['qty'] for r in preview)}\n\n"
                f"Segment 'Offline' / marketplace 'EKA' (label = segment).\n"
                f"Re-pushing the same order numbers is skipped."):
            return
        self.push_db_btn.config(state=tk.DISABLED, text="⏳  Pushing...")
        self.root.update()
        out_name = self.last_output.name if self.last_output else ''
        try:
            rec = record_eka_batch(self.last_results, out_name)
        except Exception as e:   # noqa: BLE001
            rec = {'recorded': False, 'reason': str(e)}

        if rec.get('recorded'):
            msg = f"Recorded {rec.get('new_orders', 0)} order(s) [Offline / EKA]."
            if rec.get('skipped'):
                msg += f"\nSkipped {rec['skipped']} already in DB."
            self._log(f"Push to DB: {msg}", 'ok')
            messagebox.showinfo("Push to DB", msg)
            self.push_db_btn.config(state=tk.DISABLED, text="⤓  Push to DB")
        else:
            reason = rec.get('reason', 'unknown')
            self._log(f"Push to DB failed: {reason}", 'err')
            messagebox.showerror(
                "Push to DB",
                f"Not recorded:\n{reason}\n\nFix and try again.")
            self.push_db_btn.config(state=tk.NORMAL, text="⤓  Push to DB")

    def _export_d365(self) -> None:
        if not self.last_results:
            messagebox.showwarning(
                "No Data",
                "Run processing first to generate data for D365 export.")
            return

        has_to = False
        has_so = False
        for res in self.last_results:
            for item in (res.regular_orders + res.tester_orders +
                         res.pwp_orders + res.gwp_orders +
                         res.nonstock_orders):
                if item.to:
                    if item.to.startswith('TO/'):
                        has_to = True
                    elif item.to.startswith('SO/'):
                        has_so = True
                if has_to and has_so:
                    break
            if has_to and has_so:
                break

        if not has_to and not has_so:
            self._log("D365 export: no TO or SO numbers in last batch — "
                      "nothing to export.", 'warn')
            messagebox.showwarning(
                "Nothing To Export",
                "Last batch has no TO or SO numbers.\n\n"
                "TO numbers come from EKA_DATA rows where Prefix = 'TO'.\n"
                "SO numbers come from rows where Prefix = 'SO'.\n\n"
                "Check that EKA_DATA was loaded and that the matching "
                "filenames or location types resolved correctly.")
            return

        if self.po_files:
            out_dir = Path(self.po_files[0]).parent / 'eka_output'
        elif self.so_path:
            out_dir = Path(self.so_path).parent / 'eka_output'
        else:
            out_dir = get_script_dir() / 'eka_output'
        out_dir.mkdir(parents=True, exist_ok=True)

        ts = time.strftime('%d%m%Y_%H%M%S')

        to_template: Optional[str] = None
        if has_to:
            to_template = filedialog.askopenfilename(
                title="Select D365 TO Template (EKA_Sample_Package_TO.xlsx)",
                filetypes=[("Excel files", "*.xlsx")],
            )
            if not to_template:
                self._log("D365 export cancelled — no TO template selected.",
                          'warn')
                return

        so_template: Optional[str] = None
        if has_so:
            so_template = filedialog.askopenfilename(
                title="Select D365 SO Template (EKA_Sample_Package_SO.xlsx)",
                filetypes=[("Excel files", "*.xlsx")],
            )
            if not so_template:
                self._log("D365 export: SO template skipped — only TO will "
                          "be generated.", 'warn')
                so_template = None

        produced: List[Path] = []

        if has_to and to_template:
            try:
                to_path = out_dir / f"D365_TO_Package_{ts}.xlsx"
                D365TOExporter.export(
                    self.last_results, to_template, str(to_path))
                self._log(f"D365 TO package saved → {to_path}", 'ok')
                produced.append(to_path)
            except Exception as e:
                self._log(f"D365 TO export failed: {e}", 'err')
                messagebox.showerror(
                    "D365 TO Export Failed",
                    f"Could not produce the TO package:\n\n{e}")

        if has_so and so_template:
            try:
                so_path = out_dir / f"D365_SO_Package_{ts}.xlsx"
                D365SOExporter.export(
                    self.last_results, so_template, str(so_path))
                self._log(f"D365 SO package saved → {so_path}", 'ok')
                produced.append(so_path)
            except Exception as e:
                self._log(f"D365 SO export failed: {e}", 'err')
                messagebox.showerror(
                    "D365 SO Export Failed",
                    f"Could not produce the SO package:\n\n{e}")

        if not produced:
            return

        files_list = "\n".join(f"  • {p.name}" for p in produced)
        if messagebox.askyesno(
                "D365 Exported",
                f"Generated {len(produced)} D365 package(s):\n\n"
                f"{files_list}\n\n"
                f"Open the {'first ' if len(produced) > 1 else ''}file?"):
            target = produced[0]
            if os.name == 'nt':
                os.startfile(str(target))
            elif os.name == 'posix':
                import subprocess
                opener = ('open' if hasattr(os, 'uname') and
                          os.uname().sysname == 'Darwin' else 'xdg-open')
                subprocess.Popen([opener, str(target)])

    def _download_po_template(self) -> None:
        save_path = filedialog.asksaveasfilename(
            title="Save Blank PO Template",
            defaultextension=".xlsx",
            initialfile="PO_Template_Blank.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return
        try:
            wb = Workbook()
            wb.remove(wb.active)
            hdr_fill = PatternFill('solid', fgColor='1A237E')
            hdr_font = Font(bold=True, color='FFFFFF',
                             name='Aptos Display', size=11)

            def make_header(ws, headers, widths=None):
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c, value=h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal='center')
                if widths:
                    for c, w in enumerate(widths, 1):
                        ws.column_dimensions[get_column_letter(c)].width = w
                ws.freeze_panes = 'A2'

            ws = wb.create_sheet('PO')
            make_header(ws,
                ['Rank', 'Category', 'EAN', 'SKU Code', 'Product Name',
                 'Brand', 'MRP', 'Available', 'Order Qty', 'Tester Qty'],
                [8, 20, 18, 14, 50, 14, 10, 12, 12, 12])
            sample = [1, 'Eyes', '8906121646979', '06D19087',
                      'SAMPLE PRODUCT NAME', 'RENEE', 450, 0, '', '']
            for c, v in enumerate(sample, 1):
                cell = ws.cell(row=2, column=c, value=v)
                cell.font = Font(name='Aptos Display', size=11,
                                  color='999999', italic=True)
            ws.cell(row=3, column=1,
                value='← Delete sample row. Fill EAN, Order Qty, '
                      'Tester Qty.').font = Font(
                          name='Aptos Display', size=11,
                          color='FF6600', italic=True)

            ws = wb.create_sheet('PWP')
            make_header(ws, ['Sr. No.', 'Product Name', 'Avail.Qty', 'Req.Qty'],
                          [10, 30, 12, 12])
            # v1.5.8: include the bollywood spray as a pre-filled sample
            # so stores know the exact name format to use. PWP_EAN_MAP
            # uses whitespace normalization so non-breaking spaces in
            # the source file still match.
            for row_data in [
                (1, 'Stay With Me - Mini', '', ''),
                (2, 'Perfume', '', ''),
                (3, 'Crème Mini', '', ''),
                (4, 'RENEE BOLLYWOOD FILTER BLURRING SETTING SPRAY 5ML (NFS)',
                 '', ''),
            ]:
                ws.append(row_data)
            ws.append(('Total', None, 0, 0))

            ws = wb.create_sheet('GWP')
            make_header(ws,
                ['Sr. No.', 'EAN', 'Product Name', 'Avail.Qty', 'Req.Qty'],
                [10, 18, 45, 12, 12])
            gwp_items = [
                (1, 8904473101658, 'RENEE Lunar Luxe Trousseau box – Silver', '', ''),
                (2, 8904473101672, 'RENEE Red Velvet Trousseau Box - Red', '', ''),
                (3, 8904473101665, 'RENEE Rose Glow Trousseau Box - Pink', '', ''),
                (4, 8904473101009, 'RENEE Pink Puffer Pouch', '', ''),
                (5, 8904473101023, 'RENEE Red Puffer Pouch', '', ''),
                (6, 8904473101016, 'RENEE Silver Puffer Pouch', '', ''),
            ]
            for row_data in gwp_items:
                ws.append(row_data)
            ws.append(('Total', None, None, None, 0))

            ws = wb.create_sheet('Non Stock')
            make_header(ws, ['Sr. No.', 'Product Name', 'QTY'],
                          [10, 30, 10])
            ns_items = [
                (1, 'Cotton Rolls'), (2, 'Mirrors'), (3, 'Carry Bag (Small)'),
                (4, 'Carry Bag (Big)'),
                (5, 'Renee Multicolor Bag (Big)'),
                (6, 'Renee Multicolor Bag (Small)'),
                (7, 'Cleansers'), (8, 'Calculator'),
                (9, 'Blotters'), (10, 'Swabs'), (11, 'Bill Roll'),
                (12, 'Renee Notebook'), (13, 'Pen'),
            ]
            for sr, name in ns_items:
                ws.append((sr, name, ''))
            ws.append(('Total', None, 0))

            ws = wb.create_sheet('Summary')
            ws.cell(row=3, column=2, value='[Location Name]')
            for c, h in enumerate(
                ['PO', 'Tester', 'PWP', 'GWP',
                 'Non-Stock Requirement', 'Total'], 6):
                ws.cell(row=3, column=c, value=h)

            wb.save(save_path)
            self._log(f"PO template saved → {save_path}", 'ok')
            messagebox.showinfo(
                "Template Saved",
                f"Blank PO template saved to:\n{save_path}\n\n"
                "Sheets: PO, PWP, GWP, Non Stock, Summary",
            )
        except Exception as e:
            self._log(f"Template save failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def _download_eka_template(self) -> None:
        save_path = filedialog.asksaveasfilename(
            title="Save EKA_DATA Template", defaultextension=".xlsx",
            initialfile="EKA_DATA_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'EKA_DATA'

            headers = ['Desc', 'Bill to', 'Ship to', 'Location',
                        'Gen. Biz.  Posting Group', 'Short Name',
                        'Prefix', 'Short Code', 'Transfer Code', 'Type',
                        'Example Regular', 'Example Tester', 'Status']
            widths = [40, 10, 12, 15, 22, 22, 8, 12, 15, 10, 25, 25, 10]

            hdr_fill = PatternFill('solid', fgColor='E65100')
            hdr_font = Font(bold=True, color='FFFFFF',
                             name='Aptos Display', size=11)

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

            samples = [
                ('RENEE COSMETICS-ISCON ARCADE', '20329', '20329_1', 'EBO_AMD01',
                 'OFF-EBO', 'Ahmedabad EBO', 'TO', 'AHDEB', 'EBO_AMD01', 'EBO',
                 'TO/AHDEB/04/18426', 'TO/AHDEB/TT/18427', 'Active'),
                ('RENEE COSMETICS-CHENNAI AIRPORT', '20342', '20342_1', 'AP_CHEN01',
                 'OFF-AIRPORT', 'Chennai Airport', 'TO', 'CHNAP', 'AP_CHEN01', 'Airport',
                 'TO/CHNAP/04/18426', 'TO/CHNAP/TT/18427', 'Active'),
                ('OG BEAUTY PRIVATE LIMITED', '20395', '20395_1', 'EBO_PUNE02',
                 'OFF-EBO', 'Pune EBO', 'SO', 'PUNEB', '20395_1', 'EBO',
                 'SO/PUNEB/04/18426', 'SO/PUNEB/TT/18427', 'Active'),
            ]
            sample_font = Font(name='Aptos Display', size=11,
                                color='666666', italic=True)
            for r, row_data in enumerate(samples, 2):
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = sample_font

            ws.cell(row=len(samples) + 2, column=1,
                value='← Delete sample rows. Add one row per location. '
                      'Prefix: TO for Transfer Order, SO for Sales Order. '
                      'Status: Active or Inactive.').font = Font(
                          name='Aptos Display', size=11,
                          color='FF6600', italic=True)
            ws.freeze_panes = 'A2'

            wb.save(save_path)
            self._log(f"EKA template saved → {save_path}", 'ok')
            messagebox.showinfo(
                "Template Saved",
                f"EKA_DATA template saved to:\n{save_path}",
            )
        except Exception as e:
            self._log(f"EKA template save failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def _download_so_template(self) -> None:
        save_path = filedialog.asksaveasfilename(
            title="Save Special Order Template", defaultextension=".xlsx",
            initialfile="Special_Order_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Special Order'

            headers = ['Description', 'EAN', 'Item Category Code', 'MRP',
                        'EBO Qty', 'Airport Qty', 'Kiosk Qty', 'Tester Qty']
            widths = [45, 18, 18, 10, 12, 12, 12, 12]

            hdr_fill = PatternFill('solid', fgColor='1A237E')
            hdr_font = Font(bold=True, color='FFFFFF',
                             name='Aptos Display', size=11)

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

            samples = [
                ('RENEE PRO HD 3-IN-1 - AMANDE_9 GM', '8906121648515',
                 'POWDER', 650, 18, 18, 10, 1),
                ('RENEE PRO HD CONCEALER - BUFF_8 ML', '8906121648317',
                 'CONCEALER', 750, 18, 18, 10, 1),
                ('RENEE PAPER BAG BIG MULTICOLOR', '8904473105984',
                 'PAPER BAG', 750, '', '', '', 50),
            ]
            sample_font = Font(name='Aptos Display', size=11,
                                color='666666', italic=True)
            for r, row_data in enumerate(samples, 2):
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = sample_font

            ws.cell(row=len(samples) + 2, column=1,
                value='← Delete sample rows. EAN must match GTIN in '
                      'Items_March. Leave qty blank or 0 if not '
                      'applicable. Tester Qty goes to ALL locations '
                      'at ₹0.54.').font = Font(
                          name='Aptos Display', size=11,
                          color='FF6600', italic=True)
            ws.freeze_panes = 'A2'

            wb.save(save_path)
            self._log(f"SO template saved → {save_path}", 'ok')
            messagebox.showinfo(
                "Template Saved",
                f"Special Order template saved to:\n{save_path}",
            )
        except Exception as e:
            self._log(f"SO template save failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def run(self) -> None:
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    """Application entry point."""
    print(f"{APP_TITLE} starting...")
    print(f"  Script: {get_script_dir()}")
    print(f"  Bundled folder: {get_bundled_folder()}")
    print("  Initializing GUI window...")
    sys.stdout.flush()

    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            except Exception:
                ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

    try:
        app = ReneePOApp()
        print("  GUI window ready. If you don't see it, check Alt+Tab.")
        sys.stdout.flush()
        app.run()
    except Exception as e:
        print(f"\n  FATAL ERROR during startup: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        sys.stdout.flush()
        input("\n  Press Enter to close...")
        raise


if __name__ == '__main__':
    main()