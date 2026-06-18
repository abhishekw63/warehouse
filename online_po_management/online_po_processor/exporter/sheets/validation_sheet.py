"""
exporter.sheets.validation_sheet
================================

Writes the **Validation** sheet — per-item price check with clear
PASS/FAIL status for each line.

Column layout (13 columns, or 16 when HSN check is enabled)::

    1.  PO
    2.  Item No                    — resolved from master
    3.  EAN
    4.  Description                — from Items_March (readable product name)
    5.  GST Code
    6.  Vendor MRP                  — file's MRP (config ``mrp_col``; blank
                                      when the file has no MRP column)
    7.  Our MRP                     ─┐
    8.  Vendor Landing              │ "Our" cols (7/9/11) get GREEN headers
    9.  Our Landing (m%)            ├ — our master-calculated values, paired
    10. Vendor CP                   │   beside the vendor's stated value so
    11. Our CP                     ─┘   a mismatch in any single metric is
                                        obvious at a glance.
    12. Difference (<Label>)        — primary diff: vendor − our for the
                                      active compare_basis (fob − calc).
    13. Status                      — OK / MISMATCH / NOT_IN_MASTER / NO_PRICE

    — HSN cross-check columns (v1.6.0, conditional) —
    14. HSN (Marketplace)           — hsn_col value from the punch
    15. HSN (Master)                — HSN/SAC Code from Items_March
    16. HSN Check                   — OK / MISMATCH (amber alert)

``<Label>`` is the marketplace's ``compare_label`` from config (e.g.
"Landing Rate" for Myntra, "Cost" for RK).

Side-by-side Vendor vs Our (v2.3.1)
-----------------------------------
Each metric is a pair: the value the marketplace stated in its file
(Vendor) next to the value we computed from the Items_March master
(Our). When a vendor value is present and differs from ours by more
than a paisa, that vendor cell is amber-tinted — so MRP / Landing / CP
mismatches can be read straight off the columns. Vendor columns are
blank for metrics the file doesn't carry (e.g. Flipkart has no MRP or
CP column; only its landing rate is present).

Where vendor Landing / CP come from depends on ``compare_basis``:
* ``basis='landing'`` (Myntra, FirstCry, Flipkart): ``fob_col`` is the
  vendor LANDING; an optional ``ref_fob_col`` (FirstCry 'Base Cost') is
  the vendor COST PRICE.
* ``basis='cost'`` (RK, Blink): ``fob_col`` is the vendor COST PRICE;
  ``ref_fob_col`` (if set) is the vendor landing.
Our Landing = MRP × m% (pre-GST); Our CP = MRP × m% ÷ GST (post-GST).

HSN cross-check (v1.6.0)
------------------------
When the marketplace has ``hsn_col`` set in its config (currently
Reliance only), the engine compares the punch's HSN against the
master's HSN per row. The three trailing columns appear only when at
least one row has a non-empty ``hsn_check_status`` — otherwise this
sheet keeps its base 13-column layout.

Visual cues
-----------
* **Mismatch rows** get a pale-pink fill across the entire row, Status
  cell in bold red.
* **OK rows** get a green status pill only (the bulk of a clean batch,
  so we keep row fill neutral to reduce visual fatigue).
* **NOT_IN_MASTER rows** get a pale-orange fill so these are easy to
  spot and fix by adding the item to Items_March.
* **HSN mismatches** are an ALERT, not an error: the HSN Check cell
  gets an amber "verify" highlight (the same caution colour as fuzzy
  location matches), NOT the red price-mismatch pill. HSN never blocks
  the SO — the row's price validation_status is untouched and the
  mismatch is logged as a warning on the Warnings sheet. (v2.3.1 —
  downgraded from red so it doesn't read as a failure.)

The trailing info row records ``basis=... | Margin: m%`` so someone
reviewing the output three months later can tell at a glance what the
numbers mean.
"""

from __future__ import annotations

import pandas as pd
from openpyxl.comments import Comment

from online_po_processor.data.models import ProcessingResult
from online_po_processor.exporter._styles import (
    BOLD_DATA_FONT, CALC_FILL, HEADER_FILL, INFO_ITALIC_FONT,
    LOC_MISMATCH_FILL, MISMATCH_FILL,
    MISMATCH_TEXT_FONT, NO_MASTER_FILL, NOT_IN_MASTER_TEXT_FONT,
    STATUS_OK_FILL, STATUS_OK_FONT,
    auto_width, data_cell, hdr_cell,
)


# Calculated column indices (1-based). These get a green header instead
# of the default blue to visually separate "our math" from "their data".
# v2.3.1: the "our" half of each side-by-side pair (Our MRP / Our Landing
# / Our CP). v2.7: +1 each — a 'Qty' column was inserted after Description.
_CALC_COL_INDICES = {8, 10, 12}


def _margin_label(result: ProcessingResult) -> str:
    """Short, NON-misleading margin tag for the 'Our Landing' header and the
    info row. A single '66%' lies for rule-based marketplaces (Nykaa prices
    perfume at 69%, cosmetics at 66%) and GST-based ones (Reliance), so those
    get a descriptive tag instead of one number. Straight-margin marketplaces
    keep their honest single percentage."""
    cfg = getattr(result, 'resolved_config', None) or {}
    if cfg.get('margin_rules'):
        return 'per rule'
    if cfg.get('gst_margin_discount') is not None:
        return 'GST-based'
    return f"{int(result.margin_pct * 100)}%"


def write(wb, result: ProcessingResult) -> None:
    """
    Append the 'Validation' sheet to ``wb``.

    v2.1.4: TO marketplaces with master lookup go through the same
    validation pipeline as SO. Pre-v2.1.4, TO mode skipped this sheet
    entirely because Flipkart-TO read Item No / MRP / GST from the
    dump file directly — no master to validate against. With v2.1.4
    Flipkart-TO now uses ``item_resolution='from_ean'`` and master
    lookup, so we DO have data to validate, and a dedicated reference-
    only comparison mode (``compare_mode='reference_only'``) for
    surfacing diffs without flagging them as MISMATCH.

    The sheet is still skipped only when there's truly nothing to
    show — TO mode without a configured ``compare_basis`` (i.e. older
    TO marketplaces or future TO configs that opt out of validation).
    """
    if getattr(result, 'output_type', 'so') == 'to':
        # Only skip if the marketplace genuinely has nothing to
        # validate (no compare_basis configured). With v2.1.4
        # Flipkart-TO sets compare_basis='cost' so the sheet renders.
        if not getattr(result, 'compare_basis', None):
            return

    # v2.1.4: detect reference-only mode for the banner row.
    compare_mode = (result.resolved_config or {}).get('compare_mode')
    is_reference_only = (compare_mode == 'reference_only')

    ws = wb.create_sheet('Validation')

    label = result.compare_label or 'Price'
    margin_pct_int = int(result.margin_pct * 100)

    # v1.6.0: decide up front whether HSN columns belong on this run.
    # Only shown when the engine actually populated an hsn_check_status
    # on at least one row (which happens when the marketplace config
    # has ``hsn_col`` set). This keeps the sheet's width consistent
    # with previous versions for marketplaces that don't opt in.
    has_hsn_check = any(
        so_row.hsn_check_status for so_row in result.rows
    )

    # v2.3.1: side-by-side Vendor vs Our pairs for MRP / Landing / CP so a
    # mismatch in any single metric is obvious at a glance. "Vendor" = the
    # value stated in the marketplace's file; "Our" = computed from the
    # Items_March master. Vendor columns are blank when the file doesn't
    # carry that metric (e.g. Flipkart has no MRP or CP column).
    headers = [
        'PO', 'Item No', 'EAN', 'Description', 'Qty', 'GST Code',
        'Vendor MRP', 'Our MRP',
        'Vendor Landing', f'Our Landing ({_margin_label(result)})',
        'Vendor CP', 'Our CP',
        f'Difference ({label})',
        'Status',
    ]
    if has_hsn_check:
        headers.extend([
            'HSN (Marketplace)', 'HSN (Master)', 'HSN Check',
        ])

    # ── Header row ──────────────────────────────────────────────────────
    # v2.1.4: when compare_mode='reference_only', prepend a banner row
    # above the headers explaining that the 'Difference' column is
    # informational only — Transfer Prices in the output use the
    # engine's calculated values regardless of any diff. Without this
    # banner, an operator looking at the Validation sheet would
    # reasonably expect the diffs to be acted on (since that's how
    # all other marketplaces' Validation sheets work).
    #
    # The actual cell merge happens AFTER auto_width() at the bottom
    # of this function — auto_width can't iterate columns when row 1
    # has merged cells (MergedCell objects don't expose column_letter).
    # We write the banner text now so it's in place; the merge that
    # spans it across the header columns happens last.
    header_row_idx = 1
    banner_n_cols_for_merge = 0  # tracks merge width; 0 = no banner
    banner_parts = []
    if is_reference_only:
        banner_parts.append(
            f"\u2139 Reference-only comparison: '{label}' values from "
            f"the punch file are shown for audit only. Transfer Prices "
            f"in the output use ENGINE-CALCULATED values "
            f"({margin_pct_int}% margin from Items_March master) "
            f"regardless of any diff. Diffs above \u20b90.01 are also "
            f"logged as warnings."
        )
    # v2.7: GST-dependent pricing config (Reliance) \u2014 note the exact table
    # so the output records which pricing basis was followed.
    _gmd = (result.resolved_config or {}).get('gst_margin_discount')
    if _gmd is not None:
        banner_parts.append(
            f"\u2139 Pricing config (GST-dependent): margin {_gmd*100:.0f}% "
            f"pre-GST \u2192 cost = MRP \u00d7 keep%, keep% = 1 \u2212 "
            f"{_gmd*100:.0f}% \u00d7 (1+GST) = {(1-_gmd)*100:.2f}% / "
            f"{(1-_gmd*1.05)*100:.2f}% / {(1-_gmd*1.18)*100:.2f}% of MRP at "
            f"GST 0% / 5% / 18% (per the Reliance pricing config table)."
        )
    if banner_parts:
        banner_cell = ws.cell(row=1, column=1,
                              value='     |     '.join(banner_parts))
        banner_cell.font = INFO_ITALIC_FONT
        banner_n_cols_for_merge = len(headers)
        header_row_idx = 2

    for col_idx, header in enumerate(headers, start=1):
        fill = CALC_FILL if col_idx in _CALC_COL_INDICES else HEADER_FILL
        hdr_cell(ws, header_row_idx, col_idx, header, fill=fill)

    n_cols = len(headers)
    status_col = 14   # Price-validation status column (v2.7: +1 for 'Qty')
    # HSN columns, when present, occupy 15/16/17.
    hsn_punch_col = 15
    hsn_master_col = 16
    hsn_status_col = 17

    # ── Data rows ───────────────────────────────────────────────────────
    # Alignment policy:
    #   PO, Item No, EAN, GST Code, Status, HSN cols → center (identifiers/badges)
    #   Description                                  → left   (long prose)
    #   Vendor/Our MRP·Landing·CP, Difference        → right  (monetary)
    # v2.1.4: data starts at row 3 when reference-only banner is present
    # (banner=row 1, headers=row 2, data=row 3); otherwise row 2 as before.
    r = header_row_idx + 1
    mismatches = 0
    hsn_mismatches = 0
    basis = result.compare_basis or 'landing'

    def _money(v):
        """Round a price for display, '' when absent."""
        return (round(v, 2)
                if (v is not None and not pd.isna(v)) else '')

    for so_row in result.rows:
        # ── Vendor (file) vs Our (computed) values per metric ───────────
        # Vendor landing/cp come from the file's fob/ref-fob; which is
        # which depends on compare_basis:
        #   landing-basis → fob_price IS the vendor landing; ref_fob (when
        #                   configured, e.g. FirstCry 'Base Cost') is the
        #                   vendor cost price.
        #   cost-basis    → fob_price IS the vendor cost price; ref_fob is
        #                   the vendor landing.
        # Vendor MRP comes from the optional mrp_col (None when the file
        # has no MRP). Our values are computed from the master.
        v_mrp = so_row.vendor_mrp
        v_landing = (so_row.fob_price if basis == 'landing'
                     else so_row.ref_fob_price)
        v_cp = (so_row.fob_price if basis == 'cost'
                else so_row.ref_fob_price)
        o_mrp = so_row.mrp
        # v2.3.1: use the margin actually applied to THIS row (per-line
        # margin_rules can override the run margin, e.g. Nykaa perfumes),
        # falling back to the run margin for normal marketplaces.
        row_margin = (so_row.applied_margin_pct
                      if so_row.applied_margin_pct is not None
                      else result.margin_pct)
        o_landing = (float(so_row.mrp) * row_margin
                     if so_row.mrp and not pd.isna(so_row.mrp) else None)
        o_cp = so_row.cost_price_ref

        data_cell(ws, r, 1, so_row.po_number, align='center')
        data_cell(ws, r, 2, so_row.item_no, align='center')
        data_cell(ws, r, 3, so_row.ean, align='center')
        data_cell(ws, r, 4, so_row.description, align='left')
        # v2.7: Qty — so the operator can spot low-qty items to exclude
        # without working through them.
        data_cell(ws, r, 5, so_row.qty, align='center')
        data_cell(ws, r, 6, so_row.gst_code, align='center')

        # MRP pair | Landing pair | CP pair (Vendor then Our, side by side)
        data_cell(ws, r, 7,  _money(v_mrp),     '#,##0.00', align='right')
        data_cell(ws, r, 8,  _money(o_mrp),     '#,##0.00', align='right')
        data_cell(ws, r, 9,  _money(v_landing), '#,##0.00', align='right')
        data_cell(ws, r, 10, _money(o_landing), '#,##0.00', align='right')
        data_cell(ws, r, 11, _money(v_cp),      '#,##0.00', align='right')
        data_cell(ws, r, 12, _money(o_cp),      '#,##0.00', align='right')

        # Primary difference (active basis): vendor − our for that basis.
        data_cell(ws, r, 13,
                   round(so_row.diffn, 2) if so_row.diffn is not None else '',
                   '#,##0.00', align='right')

        data_cell(ws, r, status_col, so_row.validation_status, align='center')

        # ── Per-status row styling ──────────────────────────────────────
        if so_row.validation_status == 'MISMATCH':
            mismatches += 1
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = MISMATCH_FILL
            ws.cell(row=r, column=status_col).font = MISMATCH_TEXT_FONT

        elif so_row.validation_status == 'OK':
            ws.cell(row=r, column=status_col).fill = STATUS_OK_FILL
            ws.cell(row=r, column=status_col).font = STATUS_OK_FONT

        elif so_row.validation_status == 'NOT_IN_MASTER':
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = NO_MASTER_FILL
            ws.cell(row=r, column=status_col).font = NOT_IN_MASTER_TEXT_FONT

        # ── v2.4.1: per-row EXCEPTION highlight ─────────────────────────
        # When a Master Exception was applied to THIS row (vendor CP accepted,
        # price override, or EAN remap), amber-tint the whole row so the
        # operator can see at a glance that special handling was used — even
        # though the exception typically makes the row validate OK. The Status
        # pill keeps its green OK colour on top; a cell comment names the
        # exact exception. This is the Validation-sheet half of "highlight
        # each exception" (the Lines sheet amber-tints the forced Unit Price).
        exc_label = getattr(so_row, 'exception_label', '') or ''
        if exc_label and so_row.validation_status != 'MISMATCH':
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = LOC_MISMATCH_FILL
            sc = ws.cell(row=r, column=status_col)
            if so_row.validation_status == 'OK':
                sc.fill = STATUS_OK_FILL          # keep the green OK pill
                sc.font = STATUS_OK_FONT
            sc.comment = Comment(
                f"Exception applied: {exc_label}", "PO Engine")

        # ── v2.3.1: per-metric mismatch flag ────────────────────────────
        # Amber-tint the VENDOR cell of any pair whose vendor value differs
        # from ours by more than a paisa, so an MRP / Landing / CP mismatch
        # is spotted directly from the side-by-side columns. Applied AFTER
        # the row fill so it wins on that specific cell.
        for vcol, vval, oval in ((7, v_mrp, o_mrp),
                                  (9, v_landing, o_landing),
                                  (11, v_cp, o_cp)):
            if (vval is not None and oval is not None
                    and not pd.isna(vval) and not pd.isna(oval)
                    and abs(float(vval) - float(oval)) > 0.01):
                ws.cell(row=r, column=vcol).fill = LOC_MISMATCH_FILL
                ws.cell(row=r, column=vcol).font = BOLD_DATA_FONT

        # ── HSN columns (v1.6.0, only when marketplace opts in) ─────────
        if has_hsn_check:
            data_cell(ws, r, hsn_punch_col, so_row.hsn_punch, align='center')
            data_cell(ws, r, hsn_master_col, so_row.hsn_master, align='center')
            data_cell(ws, r, hsn_status_col, so_row.hsn_check_status,
                       align='center')

            # Mirror the price-status pill styling on the HSN Check
            # cell so the user can scan the column for red at a glance.
            hsn_cell = ws.cell(row=r, column=hsn_status_col)
            if so_row.hsn_check_status == 'OK':
                hsn_cell.fill = STATUS_OK_FILL
                hsn_cell.font = STATUS_OK_FONT
            elif so_row.hsn_check_status == 'MISMATCH':
                hsn_mismatches += 1
                # v2.3.1: HSN mismatch is an ALERT, not an error. It never
                # blocks the SO (the row's price validation_status is
                # unaffected), so render it as an amber "verify" highlight
                # — the same caution colour used for fuzzy location
                # matches — rather than the red error pill. The mismatch
                # is still logged as a warning on the Warnings sheet. If
                # it ever needs to hard-fail, escalate the styling then.
                hsn_cell.fill = LOC_MISMATCH_FILL
                hsn_cell.font = BOLD_DATA_FONT
            elif so_row.hsn_check_status == 'NOT_IN_MASTER':
                hsn_cell.font = NOT_IN_MASTER_TEXT_FONT

        r += 1

    # ── Footer summary ──────────────────────────────────────────────────
    r += 1
    total = len(result.rows)
    ok_count = sum(1 for so_row in result.rows
                    if so_row.validation_status == 'OK')
    basis_note = (f"basis={result.compare_basis} "
                  f"(compared against '{label}')")
    summary_parts = [
        f"Total: {total} items",
        f"OK: {ok_count}",
    ]
    # v2.1.4: in reference-only mode, "Mismatches: 0" is technically true
    # but misleading — the engine downgrades MISMATCH→OK on purpose. Show
    # the count of rows whose |diff| exceeds tolerance instead, which is
    # what an auditor actually cares about.
    if is_reference_only:
        ref_diff_count = sum(
            1 for so_row in result.rows
            if so_row.diffn is not None and abs(so_row.diffn) > 0.01
        )
        summary_parts.append(f"Diffs > \u20b90.01: {ref_diff_count}")
    else:
        summary_parts.append(f"Mismatches: {mismatches}")
    summary_parts.extend([
        f"Margin: {_margin_label(result)}",
        basis_note,
    ])
    if has_hsn_check:
        # Only mention HSN stats when the check ran; otherwise the
        # zero in "HSN mismatches: 0" reads as a claim when really
        # the feature wasn't active.
        summary_parts.insert(3, f"HSN mismatches: {hsn_mismatches}")

    ws.cell(
        row=r, column=1,
        value=" | ".join(summary_parts),
    ).font = INFO_ITALIC_FONT

    auto_width(ws)

    # v2.1.4: NOW it's safe to merge the banner across the header
    # columns. Doing it before auto_width breaks the column iterator
    # (the MergedCell objects in cols 2..N of row 1 don't have
    # column_letter attribute, which auto_width's col[0] indexing
    # depends on).
    if banner_n_cols_for_merge > 0:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=banner_n_cols_for_merge,
        )

    # v2.1.4: freeze below headers — A3 when banner present (row1=banner,
    # row2=headers), A2 otherwise. Without this the banner row scrolls
    # away with the data and operators lose the context note.
    ws.freeze_panes = f'A{header_row_idx + 1}'