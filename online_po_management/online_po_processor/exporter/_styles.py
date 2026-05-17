"""
exporter._styles
================

Shared formatting constants and small helper functions used by every
sheet writer.

Single source of truth for colours, fonts, borders, and the standard
``hdr_cell`` / ``data_cell`` / ``wrap_data_cell`` / ``auto_width``
patterns. Per-sheet modules import from here rather than re-defining
their own ``Font`` / ``Alignment`` objects so the workbook has a
consistent visual language across all sheets.

v2.1.0 alignment overhaul
-------------------------
``data_cell`` previously set no alignment at all, which meant Excel
fell back to its defaults: text → left, numbers → right. That produced
ragged-looking sheets where the same logical row had cells in mixed
horizontal positions (e.g. Summary's TOTAL row had right-aligned
numbers next to left-aligned 'TOTAL' text), and where ID-like numeric
columns (Item Code, EAN, HSN) drifted right when readers expect
center-aligned identifiers.

The fix is a smart-default alignment scheme:

    * **Numbers** (int / float / Decimal)  → ``horizontal='right'``
    * **Text**    (str / bool / None)      → ``horizontal='left'``
    * **Caller override** via ``align=``   → wins over auto-pick

Vertical is always ``'center'`` so multi-line cells (Warnings) and
wider rows (TOTAL) read consistently.

ID-style columns where the value is numeric but should look like an
identifier (PO number, EAN, Item Code, HSN, FSN) pass ``align='center'``
explicitly. Ordinary numeric values (MRP, Qty, Cost Price, Diff,
Total Amount) get the default right-alignment for easy column-wise
visual sum.

A new ``wrap_data_cell`` helper exists for free-form long-text cells
(currently the Warnings sheet's Warning column). It sets
``wrap_text=True`` and pairs naturally with a column-width cap on the
caller side — Excel then auto-grows the row height to fit.

A ``TOTAL_ROW_FILL`` + ``TOTAL_ROW_BORDER`` pair is added so a TOTAL
row reads as one continuous styled strip across all columns instead
of breaking visually wherever there's no value.

Color palette
-------------
::

    Hex        Role                                         Used by
    ─────────  ───────────────────────────────────────────  ─────────────────
    1A237E     Standard header (deep blue)                  most sheets
    E65100     Warning header (orange)                      Warnings sheet
    37474F     Raw passthrough header (slate)               Raw Data (left)
    1B5E20     Calculated/validation header (green)         Validation, Raw
    455A64     Reference column header (muted slate)        Raw Data ref Diffn
    00C853     OK status pill (green)                       Summary, Validation
    FF5252     UNMAPPED pill (red)                          Summary
    E8F5E9     Calc cells row tint (light green)            Raw Data
    FFEBEE     Validation MISMATCH row tint (pink)          Validation
    FFCDD2     Raw Data MISMATCH row tint (light red)       Raw Data
    FFF3E0     NOT_IN_MASTER row tint (light orange)        Validation
    ECEFF1     Reference Diffn row tint (light grey)        Raw Data
    FFF59D     Location mismatch tint (pale yellow)         Summary
    D32F2F     Mismatch text emphasis (red)                 Validation, Raw
    9E9E9E     Template "not read" header                   PO template only
    EEEEEE     TOTAL row fill (light grey)                  Summary (v2.1.0)
"""

from __future__ import annotations
from numbers import Number
from typing import Any, Dict, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Fills ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', fgColor='1A237E')   # deep blue
WARN_FILL = PatternFill('solid', fgColor='E65100')     # orange — Warnings sheet
RAW_HDR_FILL = PatternFill('solid', fgColor='37474F')  # slate — Raw Data passthrough headers
CALC_FILL = PatternFill('solid', fgColor='1B5E20')     # green — calc/validation headers
REF_FILL = PatternFill('solid', fgColor='455A64')      # muted slate — reference Diffn header

# Row-tint fills (applied to data cells, not headers)
OK_FILL = PatternFill('solid', fgColor='E8F5E9')
MISMATCH_FILL = PatternFill('solid', fgColor='FFEBEE')
NO_MASTER_FILL = PatternFill('solid', fgColor='FFF3E0')
CALC_BG = PatternFill('solid', fgColor='E8F5E9')
REF_BG = PatternFill('solid', fgColor='ECEFF1')
RAW_MISMATCH_BG = PatternFill('solid', fgColor='FFCDD2')
LOC_MISMATCH_FILL = PatternFill('solid', fgColor='FFF59D')

# Status pills (cell fill + bold font, applied per-cell)
STATUS_OK_FILL = PatternFill('solid', fgColor='00C853')
STATUS_BAD_FILL = PatternFill('solid', fgColor='FF5252')

# v2.1.0: TOTAL-row strip fill — light grey so the row reads as one
# continuous summary band even on cells that don't carry a value
# (e.g. Summary's Cust No / Ship-to / Status columns on the totals row).
TOTAL_ROW_FILL = PatternFill('solid', fgColor='EEEEEE')


# ── Fonts ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
DATA_FONT = Font(name='Aptos Display', size=11)

STATUS_OK_FONT = Font(name='Aptos Display', size=11, bold=True, color='000000')
STATUS_BAD_FONT = Font(name='Aptos Display', size=11, bold=True, color='FFFFFF')
MISMATCH_TEXT_FONT = Font(name='Aptos Display', size=11, bold=True,
                           color='D32F2F')
NOT_IN_MASTER_TEXT_FONT = Font(name='Aptos Display', size=11, bold=True,
                                color='E65100')

INFO_ITALIC_FONT = Font(name='Aptos Display', size=10, italic=True,
                         color='666666')
LEGEND_ITALIC_FONT = Font(name='Aptos Display', size=10, italic=True,
                           color='B7950B')

BOLD_DATA_FONT = Font(name='Aptos Display', size=11, bold=True)


# ── Borders ────────────────────────────────────────────────────────────────

THIN_SIDE = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                 top=THIN_SIDE, bottom=THIN_SIDE)

# v2.1.0: Thick top side for TOTAL rows so the row visibly separates
# from the data above it. Other three sides keep the standard thin
# border so the strip joins cleanly with the surrounding sheet.
_THICK_TOP = Side(style='medium', color='666666')
TOTAL_ROW_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                           top=_THICK_TOP, bottom=THIN_SIDE)


# ── Alignment helpers ──────────────────────────────────────────────────────
#
# Cached Alignment objects so we don't allocate a new one per cell.
# openpyxl shares Alignment instances across cells safely as long as the
# instance itself isn't mutated.

_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
_ALIGN_WRAP_LEFT = Alignment(horizontal='left', vertical='center',
                              wrap_text=True)


def _auto_align(value: Any) -> Alignment:
    """
    Pick the default horizontal alignment based on ``value`` type.

    Smart defaults:
        * Numbers (int, float, Decimal) → right-aligned. Matches Excel
          convention so column-wise totals read cleanly.
        * Everything else (str, bool, None, dates) → left-aligned.

    Booleans are explicitly NOT treated as numbers even though
    ``isinstance(True, int)`` returns True in Python — booleans render
    as ``True``/``False`` text in Excel and look better left-aligned.

    Args:
        value: The cell value.

    Returns:
        The pre-built ``Alignment`` for that value type.
    """
    if isinstance(value, bool):
        return _ALIGN_LEFT
    if isinstance(value, Number):
        return _ALIGN_RIGHT
    return _ALIGN_LEFT


# ── Cell-writing helpers ───────────────────────────────────────────────────
#
# Every sheet writer ends up doing the same three things over and over:
# write a styled header cell, write a styled data cell, and auto-fit the
# column widths. Centralising them here means a future style tweak is
# one-line wide instead of seven-files wide.

def hdr_cell(ws, row: int, col: int, value,
              fill: Optional[PatternFill] = None,
              font: Optional[Font] = None):
    """
    Write and style a header cell.

    Args:
        ws:    Target worksheet.
        row:   1-based row number.
        col:   1-based column number.
        value: Header text.
        fill:  Override the default ``HEADER_FILL`` (e.g. ``WARN_FILL``).
        font:  Override the default ``HEADER_FONT``.

    Returns:
        The created cell (for further per-call tweaks).
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or HEADER_FONT
    cell.fill = fill or HEADER_FILL
    cell.alignment = _ALIGN_CENTER
    cell.border = BORDER
    return cell


def data_cell(ws, row: int, col: int, value,
               number_format: Optional[str] = None,
               align: Optional[str] = None):
    """
    Write and style a data cell.

    v2.1.0: ``align`` parameter added. When omitted, alignment is
    auto-picked from the value type (numbers right, everything else
    left). Vertical is always ``center``.

    Args:
        ws:            Target worksheet.
        row:           1-based row number.
        col:           1-based column number.
        value:         Cell value (any openpyxl-acceptable type).
        number_format: Optional Excel number format string
                       (e.g. ``'#,##0.00'``).
        align:         Optional explicit horizontal alignment —
                       ``'left'``, ``'center'``, or ``'right'``. When
                       ``None`` (default), alignment is chosen from the
                       value type via :func:`_auto_align`. Use
                       ``'center'`` for ID-style numeric columns (PO,
                       EAN, Item Code, HSN, FSN) where the value is a
                       number but should be visually centered like an
                       identifier. Use ``'right'`` (or omit on numeric
                       values) for monetary / quantitative columns.

    Returns:
        The created cell.
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.border = BORDER
    if number_format:
        cell.number_format = number_format

    if align == 'left':
        cell.alignment = _ALIGN_LEFT
    elif align == 'center':
        cell.alignment = _ALIGN_CENTER
    elif align == 'right':
        cell.alignment = _ALIGN_RIGHT
    else:
        # No explicit override — pick by value type.
        cell.alignment = _auto_align(value)

    return cell


def wrap_data_cell(ws, row: int, col: int, value,
                    align: str = 'left'):
    """
    Write a data cell with text wrapping enabled.

    v2.1.0: introduced for the Warnings sheet so long warning messages
    don't render as squished single lines clipped at the column edge.
    Excel auto-grows the row height to fit wrapped content as long as
    the row has no explicit height set — so callers should NOT set
    ``ws.row_dimensions[r].height`` for rows containing wrap_data_cell.

    Pairs naturally with a column-width cap on the caller side
    (e.g. ``auto_width(ws, caps={'C': 80})``) — text then wraps at
    the cap width and the row grows to fit.

    Args:
        ws:    Target worksheet.
        row:   1-based row number.
        col:   1-based column number.
        value: Cell value (typically a long string).
        align: Horizontal alignment — defaults to ``'left'`` since
               wrapped paragraphs read most naturally left-aligned.

    Returns:
        The created cell.
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.border = BORDER
    if align == 'center':
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
    elif align == 'right':
        cell.alignment = Alignment(horizontal='right', vertical='center',
                                    wrap_text=True)
    else:
        cell.alignment = _ALIGN_WRAP_LEFT
    return cell


def auto_width(ws, max_width: int = 50,
                padding: int = 3,
                caps: Optional[Dict[str, int]] = None) -> None:
    """
    Auto-fit each column's width based on its longest value.

    v2.1.0: ``caps`` parameter added so callers can override the
    global ``max_width`` for specific columns. Used by the Warnings
    sheet to cap the Warning column at 80 chars so text wraps
    rather than producing a 200-char-wide column.

    Width calculation: counts characters in the rendered value via
    ``str(cell.value)``. For cells with a number format like the INR
    Indian-grouping format (which adds ₹ + commas to a raw number),
    the visible text is wider than ``str(raw_value)`` would suggest.
    For now we accept that under-estimate as a known limitation —
    it produces slightly snug INR columns but doesn't truncate them.
    The ``padding`` default of 3 absorbs most of the slack.

    Args:
        ws:        Target worksheet.
        max_width: Default cap on column width (Excel character units).
                   Prevents one extreme cell from blowing out the layout.
        padding:   Characters added to the longest value's length.
                   Default 3 leaves comfortable breathing room.
        caps:      Optional ``{column_letter: max_width}`` overrides.
                   E.g. ``{'C': 80}`` lets column C be wider than the
                   global max while keeping other columns at the default.
                   A cap of 0 means "no cap, use the longest value".
    """
    caps = caps or {}
    for col in ws.columns:
        letter = col[0].column_letter
        widest = max((len(str(c.value or '')) for c in col), default=8)
        # Per-column cap wins over global max_width when supplied.
        if letter in caps:
            cap = caps[letter]
            target = (widest + padding) if cap == 0 else min(widest + padding, cap)
        else:
            target = min(widest + padding, max_width)
        ws.column_dimensions[letter].width = target