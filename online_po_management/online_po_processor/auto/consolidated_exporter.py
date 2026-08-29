"""
auto.consolidated_exporter
===========================

AUTO mode (v2.4.0) — one combined workbook for a whole Auto run, so the
operator stops copy-pasting each marketplace's output before uploading to
D365.

Built from the in-memory :class:`~online_po_processor.auto.auto_runner.
MarketplaceRun` list the runner already returns (each holds its
``ProcessingResult``). Sheets:

1. **Overall Summary** — the run roll-up / "file mapping": one row per
   marketplace showing whether a file was present and how much was
   scraped (POs, items, qty, value), its warehouse, and warnings. The
   answer to "which marketplaces did we get data from today, and how
   much" — including the ones that had no file.
2. **Headers (SO)** / **Lines (SO)** — every SO marketplace's headers &
   lines concatenated, using the SAME column layout as the per-marketplace
   files (imported from the sheet writers so they never drift). Each
   marketplace's rows keep THEIR OWN dispatch warehouse (Auto mode allows
   AHD/BLR per marketplace), so the Location Code is correct per row.
3. **Headers (TO)** / **Lines (TO)** — same, for Transfer-Order
   marketplaces (Flipkart-TO, Meesho-TO). Only emitted when present, since
   SO and TO are separate D365 imports.
4. **Summary** — consolidated per-PO: Marketplace | PO | Raw/Mapped
   location | Items | Total Qty | Total Value.
5. **Validation** — consolidated per-item price check (Vendor vs Our MRP /
   Landing / CP + diff + status) so CP / landing-rate issues across ALL
   marketplaces are reviewable in one place.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook

from online_po_processor.config.marketplaces import MARKETPLACE_CONFIGS
from online_po_processor.data.master_loader import MasterLoader
from online_po_processor.exporter._styles import (
    BOLD_DATA_FONT, HEADER_FILL, INFO_ITALIC_FONT,
    LOC_MISMATCH_FILL, MISMATCH_FILL, MISMATCH_TEXT_FONT,
    NO_MASTER_FILL, NOT_IN_MASTER_TEXT_FONT,
    STATUS_BAD_FILL, STATUS_BAD_FONT, STATUS_OK_FILL, STATUS_OK_FONT,
    TOTAL_ROW_BORDER, TOTAL_ROW_FILL,
    auto_width, data_cell, hdr_cell,
)
# Import the exact column layouts from the per-marketplace sheet writers so
# the consolidated Headers/Lines stay in lockstep with the single-file ones.
from online_po_processor.exporter.sheets.headers_sheet import (
    _HEADERS as _SO_HDR_COLS,
    _TO_DIRECT_TRANSFER, _TO_HEADERS as _TO_HDR_COLS,
    _TO_IN_TRANSIT, _TO_TRANSFER_FROM,
)
from online_po_processor.exporter.sheets.lines_sheet import (
    _HEADERS as _SO_LINE_COLS, _LINE_NO_STEP, _TO_HEADERS as _TO_LINE_COLS,
)
from online_po_processor.exporter.sheets.tracker_sheet import (
    _HEADERS as _TRACKER_COLS, write_tracker_row,
)

_MONEY = '#,##0.00'
_INR = ('[>=10000000]"₹"##\\,##\\,##\\,##0.00;'
        '[>=100000]"₹"##\\,##\\,##0.00;'
        '"₹"##,##0.00')

_STATUS_LABEL = {
    'ok': 'Processed',
    'no_files': 'No file',
    'no_rows': 'No data extracted',
    'error': 'ERROR',
}


def _config_for(run) -> dict:
    """The marketplace config for a run (used for the GST flag/type)."""
    return MARKETPLACE_CONFIGS.get(run.marketplace, {})


def _row_value_incgst(so_row, config: dict) -> float:
    """
    GST-inclusive line value. For marketplaces whose ``amount_col`` is
    PRE-GST (flagged ``amount_is_pre_gst``, e.g. RK) the native amount is
    grossed up by the line's GST; every other marketplace's amount is
    already tax-inclusive, so it's taken as-is (no double counting).
    """
    amt = float(so_row.amount or 0.0)
    if config.get('amount_is_pre_gst'):
        return amt * MasterLoader.gst_divisor(so_row.gst_code)
    return amt


def _is_to(run) -> bool:
    return getattr(run.result, 'output_type', 'so') == 'to'


def _money(v):
    return round(v, 2) if (v is not None and not pd.isna(v)) else ''


# ── public entry point ─────────────────────────────────────────────────

def export_consolidated(runs: List, online_root: str) -> Path:
    """
    Write the consolidated workbook for an Auto run and return its path.

    Saved to ``<online_root>/_Consolidated/consolidated_<DD-MM-YYYY_HHMMSS>
    .xlsx`` so repeat runs never clobber each other.
    """
    out_dir = Path(online_root) / '_Consolidated'
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
    out_path = out_dir / f'consolidated_{stamp}.xlsx'

    ok = [r for r in runs if r.status == 'ok' and r.result is not None]
    so_runs = [r for r in ok if not _is_to(r)]
    to_runs = [r for r in ok if _is_to(r)]

    wb = Workbook()
    wb.remove(wb.active)

    # SO and TO are kept STRICTLY separate end to end — they are different
    # D365 imports and must never share a sheet. TO sheets are emitted
    # only when TO data is present, so an SO-only run stays uncluttered.
    # v2.4.0: the Tracker is no longer duplicated here — it's generated
    # once from the history DB (the single source of truth) via
    # ``export_tracker_from_db``. The consolidated workbook keeps the D365
    # import + review sheets only.
    _overall_summary(wb.create_sheet('Overall Summary'), runs)
    if so_runs:
        _headers(wb.create_sheet('Headers (SO)'), so_runs, _SO_HDR_COLS, 'so')
        _lines(wb.create_sheet('Lines (SO)'), so_runs, _SO_LINE_COLS, 'so')
        _consolidated_summary(wb.create_sheet('Summary (SO)'), so_runs)
        _consolidated_validation(wb.create_sheet('Validation (SO)'), so_runs)
    if to_runs:
        _headers(wb.create_sheet('Headers (TO)'), to_runs, _TO_HDR_COLS, 'to')
        _lines(wb.create_sheet('Lines (TO)'), to_runs, _TO_LINE_COLS, 'to')
        _consolidated_summary(wb.create_sheet('Summary (TO)'), to_runs)
        _consolidated_validation(wb.create_sheet('Validation (TO)'), to_runs)

    wb.save(str(out_path))
    return out_path


def export_tracker_from_db(run_id: int, dump_root) -> Path:
    """
    Write the standalone **internal Tracker** for one run **from the
    history DB** — the single source of truth (v2.4.0). No recomputation
    from the engine result; the tracker is a view of what was recorded.

    Saved (timestamped, versioned) to
    ``<Dump>/Tracker/Online/Online_Tracker_<DD-MM-YYYY_HHMMSS>.xlsx``. The DB
    holds only new POs, so the tracker naturally lists only newly-uploaded
    POs (already-uploaded ones were removed from output + DB by dedup).
    """
    from online_po_processor.auto.history_db import get_history_store

    dump = Path(dump_root)
    store = get_history_store(dump / 'Tracker' / 'history.db')
    try:
        orders = store.fetch_orders(run_id=run_id)
    finally:
        store.close()

    tracker_dir = dump / 'Tracker' / 'Online'
    tracker_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
    out_path = tracker_dir / f'Online_Tracker_{stamp}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tracker'
    for c, h in enumerate(_TRACKER_COLS, 1):
        hdr_cell(ws, 1, c, h)
    for i, o in enumerate(orders, start=2):
        write_tracker_row(ws, i, {
            'segment': o.get('segment') or 'OnlineB2B',
            'market_place': o['marketplace_label'],
            'po': o['po'],
            'location': o['location'] or '',
            # Pass the raw DB date (a date object on MySQL) straight through —
            # write_tracker_row writes it as a real Excel DD-MM-YYYY date.
            'po_date': o['po_date'],
            'exp_date': o['exp_date'],
            'aging': '',
            'order_value': o['order_value'],
            'order_qty': o['qty'],
        })
    auto_width(ws)
    ws.freeze_panes = 'A2'
    wb.save(str(out_path))
    return out_path


# ── 1. Overall Summary (file mapping / roll-up) ─────────────────────────

def _overall_summary(ws, runs: List) -> None:
    cols = ['Marketplace', 'Type', 'Status', 'Warehouse', 'Files', 'POs',
            'Items', 'Total Qty', 'Total Value (Inc GST)', 'Warnings']
    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 1, c, h)

    # Aggregate per marketplace (a marketplace can yield several batches).
    agg: Dict[str, dict] = {}
    order: List[str] = []
    for run in runs:
        mp = run.marketplace
        cfg = _config_for(run)
        if mp not in agg:
            agg[mp] = {'status': run.status, 'wh': run.warehouse,
                       'type': 'TO' if cfg.get('output_type') == 'to' else 'SO',
                       'files': [], 'pos': 0, 'items': 0, 'qty': 0,
                       'value': 0.0, 'warn': 0}
            order.append(mp)
        a = agg[mp]
        a['files'].extend(run.input_files)
        a['warn'] += run.warnings
        if run.status == 'ok':            # 'ok' wins if any batch succeeded
            a['status'] = 'ok'
            a['wh'] = run.warehouse or a['wh']
        a['pos'] += run.pos
        a['items'] += run.rows
        a['qty'] += run.qty
        if run.result is not None:
            a['value'] += sum(_row_value_incgst(s, cfg) for s in run.result.rows)

    # SO marketplaces listed first, then TO — never interleaved (stable
    # within each block by original encounter order).
    _pos = {mp: i for i, mp in enumerate(order)}
    order.sort(key=lambda mp: (agg[mp]['type'] == 'TO', _pos[mp]))

    r = 2
    tot_pos = tot_items = tot_qty = 0
    tot_value = 0.0
    for mp in order:
        a = agg[mp]
        processed = a['status'] == 'ok'
        data_cell(ws, r, 1, mp, align='left')
        data_cell(ws, r, 2, a['type'], align='center')
        scell = data_cell(ws, r, 3, _STATUS_LABEL.get(a['status'], a['status']),
                          align='center')
        data_cell(ws, r, 4, a['wh'] if processed else '', align='center')
        data_cell(ws, r, 5, len(a['files']), align='center')
        data_cell(ws, r, 6, a['pos'], align='center')
        data_cell(ws, r, 7, a['items'], align='center')
        data_cell(ws, r, 8, a['qty'], align='center')
        data_cell(ws, r, 9, round(a['value'], 2) if processed else '',
                  number_format=_INR, align='right')
        data_cell(ws, r, 10, a['warn'], align='center')

        if processed:
            scell.fill = STATUS_OK_FILL
            scell.font = STATUS_OK_FONT
        elif a['status'] == 'error':
            scell.fill = STATUS_BAD_FILL
            scell.font = STATUS_BAD_FONT

        tot_pos += a['pos']; tot_items += a['items']
        tot_qty += a['qty']; tot_value += a['value']
        r += 1

    # TOTAL strip
    totals = [(1, 'TOTAL'), (2, ''), (3, ''), (4, ''), (5, ''), (6, tot_pos),
              (7, tot_items), (8, tot_qty), (9, round(tot_value, 2)), (10, '')]
    for c, v in totals:
        cell = data_cell(ws, r, c, v,
                         number_format=_INR if c == 9 and v != '' else None,
                         align='center' if c != 9 else 'right')
        cell.fill = TOTAL_ROW_FILL
        cell.font = BOLD_DATA_FONT
        cell.border = TOTAL_ROW_BORDER

    r += 2
    ws.cell(row=r, column=1,
            value=(f"Auto run consolidated — generated "
                   f"{datetime.now().strftime('%d-%m-%Y %H:%M')} | "
                   f"SO and TO kept on separate sheets")
            ).font = INFO_ITALIC_FONT
    auto_width(ws)


# ── 2/3. Combined Headers / Lines (SO + TO) ─────────────────────────────

def _headers(ws, runs: List, cols: List[str], kind: str) -> None:
    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 1, c, h)
    today = datetime.now().strftime('%d-%m-%Y')
    seen: set = set()
    r = 1
    for run in runs:
        res = run.result
        loc = getattr(res, 'warehouse_code', '') or 'PICK'
        for so in res.rows:
            if so.po_number in seen:
                continue
            seen.add(so.po_number)
            r += 1
            if kind == 'so':
                data_cell(ws, r, 1, 'Order', align='center')
                data_cell(ws, r, 2, so.po_number, align='center')
                data_cell(ws, r, 3, so.cust_no, align='center')
                data_cell(ws, r, 4, so.ship_to, align='center')
                for col in range(5, 10):                 # 5 date columns
                    data_cell(ws, r, col, today, align='center')
                data_cell(ws, r, 10, so.po_number, align='center')
                data_cell(ws, r, 11, loc, align='center')
                data_cell(ws, r, 12, '', align='center')
                data_cell(ws, r, 13, 'B2B', align='center')
            else:  # TO
                try:
                    data_cell(ws, r, 1, int(so.po_number), align='center')
                except (ValueError, TypeError):
                    data_cell(ws, r, 1, so.po_number, align='center')
                data_cell(ws, r, 2, _TO_TRANSFER_FROM, align='center')
                data_cell(ws, r, 3, so.ship_to or '', align='center')
                data_cell(ws, r, 4, today, align='center')
                data_cell(ws, r, 5, _TO_IN_TRANSIT, align='center')
                data_cell(ws, r, 6, _TO_DIRECT_TRANSFER, align='center')
    auto_width(ws)


def _lines(ws, runs: List, cols: List[str], kind: str) -> None:
    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 1, c, h)
    # Flatten every run's rows (carrying each row's per-run loc/override) and
    # GROUP by PO so a PO's lines are contiguous and Line No. is 10000, 20000, …
    # unique within each PO. The old reset-on-PO-change logic broke when a punch
    # interleaved a PO's lines (e.g. Blink) OR when the same PO spanned runs —
    # both produced duplicate (Document No., Line No.) pairs that D365 OVERWRITES
    # on import. Grouping restores the correct per-PO numbering.
    from collections import OrderedDict
    groups: "OrderedDict[str, list]" = OrderedDict()
    for run in runs:
        res = run.result
        loc = getattr(res, 'warehouse_code', '') or 'PICK'
        override = bool(getattr(res, 'override_unit_price', False))
        for so in res.rows:
            groups.setdefault(so.po_number, []).append((so, loc, override))
    r = 1
    for _po, items in groups.items():
        line_no = 0
        for so, loc, override in items:
            line_no += _LINE_NO_STEP
            r += 1
            if kind == 'so':
                data_cell(ws, r, 1, 'Order', align='center')
                data_cell(ws, r, 2, so.po_number, align='center')
                data_cell(ws, r, 3, line_no, align='center')
                data_cell(ws, r, 4, 'Item', align='center')
                data_cell(ws, r, 5, so.item_no, align='center')
                data_cell(ws, r, 6, loc, align='center')
                data_cell(ws, r, 7, so.qty, align='center')
                if override and so.cost_price_ref is not None:
                    data_cell(ws, r, 8, round(so.cost_price_ref, 2),
                              number_format=_MONEY, align='right')
                else:
                    data_cell(ws, r, 8, '', align='right')
            else:  # TO
                try:
                    data_cell(ws, r, 1, int(so.po_number), align='center')
                except (ValueError, TypeError):
                    data_cell(ws, r, 1, so.po_number, align='center')
                data_cell(ws, r, 2, str(line_no), align='center')
                data_cell(ws, r, 3, so.item_no, align='center')
                data_cell(ws, r, 4, so.qty, align='center')
                for c in range(5, 9):
                    data_cell(ws, r, c, '', align='center')
                data_cell(ws, r, 9,
                          so.calc_price if so.calc_price is not None else '',
                          number_format='#,##0.0000', align='right')
    auto_width(ws)


# ── 4. Consolidated Summary (per PO, all marketplaces) ──────────────────

def _consolidated_summary(ws, runs: List) -> None:
    cols = ['Marketplace', 'PO', 'Location (Raw)', 'Location (Mapped)',
            'Items', 'Total Qty', 'Total Value (Inc GST)']
    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 1, c, h)

    r = 2
    tot_items = tot_qty = 0
    tot_value = 0.0
    for run in runs:
        res = run.result
        cfg = _config_for(run)
        groups: Dict[str, dict] = {}
        for so in res.rows:
            g = groups.get(so.po_number)
            if g is None:
                g = groups[so.po_number] = {
                    'raw': so.location, 'mapped': so.mapped_location,
                    'items': 0, 'qty': 0, 'value': 0.0}
            g['items'] += 1
            g['qty'] += so.qty
            g['value'] += _row_value_incgst(so, cfg)
        for po, g in groups.items():
            data_cell(ws, r, 1, run.marketplace, align='left')
            data_cell(ws, r, 2, po, align='center')
            data_cell(ws, r, 3, g['raw'], align='left')
            data_cell(ws, r, 4, g['mapped'], align='left')
            data_cell(ws, r, 5, g['items'], align='center')
            data_cell(ws, r, 6, g['qty'], align='center')
            data_cell(ws, r, 7, round(g['value'], 2),
                      number_format=_INR, align='right')
            # Amber both location cells on a fuzzy (raw≠mapped) match.
            raw_n = (g['raw'] or '').strip().lower()
            map_n = (g['mapped'] or '').strip().lower()
            if raw_n and map_n and raw_n != map_n:
                ws.cell(row=r, column=3).fill = LOC_MISMATCH_FILL
                ws.cell(row=r, column=4).fill = LOC_MISMATCH_FILL
            tot_items += g['items']; tot_qty += g['qty']
            tot_value += g['value']
            r += 1

    totals = [(1, 'TOTAL'), (2, ''), (3, ''), (4, ''), (5, tot_items),
              (6, tot_qty), (7, round(tot_value, 2))]
    for c, v in totals:
        cell = data_cell(ws, r, c, v,
                         number_format=_INR if c == 7 and v != '' else None,
                         align='center' if c != 7 else 'right')
        cell.fill = TOTAL_ROW_FILL
        cell.font = BOLD_DATA_FONT
        cell.border = TOTAL_ROW_BORDER
    auto_width(ws)


# ── 5. Consolidated Validation (per item, all marketplaces) ─────────────

def _consolidated_validation(ws, runs: List) -> None:
    cols = ['Marketplace', 'PO', 'Item No', 'EAN', 'Description', 'GST',
            'Vendor MRP', 'Our MRP', 'Vendor Landing', 'Our Landing',
            'Vendor CP', 'Our CP', 'Difference', 'Status']
    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 1, c, h)
    n_cols = len(cols)
    status_col = n_cols

    r = 2
    mismatches = 0
    for run in runs:
        res = run.result
        basis = res.compare_basis or 'landing'
        run_margin = res.margin_pct
        for so in res.rows:
            v_mrp = so.vendor_mrp
            v_landing = so.fob_price if basis == 'landing' else so.ref_fob_price
            v_cp = so.fob_price if basis == 'cost' else so.ref_fob_price
            o_mrp = so.mrp
            row_margin = (so.applied_margin_pct
                          if so.applied_margin_pct is not None else run_margin)
            o_landing = (float(so.mrp) * row_margin
                         if so.mrp and not pd.isna(so.mrp) else None)
            o_cp = so.cost_price_ref

            data_cell(ws, r, 1, run.marketplace, align='left')
            data_cell(ws, r, 2, so.po_number, align='center')
            data_cell(ws, r, 3, so.item_no, align='center')
            data_cell(ws, r, 4, so.ean, align='center')
            data_cell(ws, r, 5, so.description, align='left')
            data_cell(ws, r, 6, so.gst_code, align='center')
            data_cell(ws, r, 7, _money(v_mrp), _MONEY, align='right')
            data_cell(ws, r, 8, _money(o_mrp), _MONEY, align='right')
            data_cell(ws, r, 9, _money(v_landing), _MONEY, align='right')
            data_cell(ws, r, 10, _money(o_landing), _MONEY, align='right')
            data_cell(ws, r, 11, _money(v_cp), _MONEY, align='right')
            data_cell(ws, r, 12, _money(o_cp), _MONEY, align='right')
            data_cell(ws, r, 13,
                      round(so.diffn, 2) if so.diffn is not None else '',
                      _MONEY, align='right')
            data_cell(ws, r, status_col, so.validation_status, align='center')

            if so.validation_status == 'MISMATCH':
                mismatches += 1
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).fill = MISMATCH_FILL
                ws.cell(row=r, column=status_col).font = MISMATCH_TEXT_FONT
            elif so.validation_status == 'OK':
                ws.cell(row=r, column=status_col).fill = STATUS_OK_FILL
                ws.cell(row=r, column=status_col).font = STATUS_OK_FONT
            elif so.validation_status == 'NOT_IN_MASTER':
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).fill = NO_MASTER_FILL
                ws.cell(row=r, column=status_col).font = NOT_IN_MASTER_TEXT_FONT

            # Amber the vendor cell of any pair that differs from ours by
            # more than a paisa (direct CP / landing / MRP issue spotting).
            for vcol, vval, oval in ((7, v_mrp, o_mrp),
                                     (9, v_landing, o_landing),
                                     (11, v_cp, o_cp)):
                if (vval is not None and oval is not None
                        and not pd.isna(vval) and not pd.isna(oval)
                        and abs(float(vval) - float(oval)) > 0.01):
                    ws.cell(row=r, column=vcol).fill = LOC_MISMATCH_FILL
                    ws.cell(row=r, column=vcol).font = BOLD_DATA_FONT
            r += 1

    r += 1
    ws.cell(row=r, column=1,
            value=f"Total items: {r - 3} | Mismatches: {mismatches}"
            ).font = INFO_ITALIC_FONT
    auto_width(ws)
    ws.freeze_panes = 'A2'
