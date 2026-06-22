"""
gui.app_window
==============

Main Tkinter window — ``OnlinePOApp``.

Layout (520×720 px — bumped in v1.5.0 to accommodate D365 + Email buttons)::

    ┌─────────────────────────────────────────┐
    │         Online PO Processor              │  ← title
    │  Marketplace PO → ERP Sales Order Import │
    │                                          │
    │ Marketplace: [Myntra ▼]  Margin: [70]%   │  ← mkt + margin row
    │                                          │
    │ ┌─ Input Files ─────────────────────┐   │
    │ │ Items Master:    ✓ Items March... │   │
    │ │                  Updated: …        │   │
    │ │ Ship-To Mapping: ✓ Ship to B2B... │   │
    │ │                  Updated: …        │   │
    │ │ Marketplace PO:  Not selected      │   │
    │ └────────────────────────────────────┘   │
    │                                          │
    │        [▶ Generate SO]                   │  ← primary action
    │        [📂 Open Last Output]             │
    │        [📋 Download PO Template]         │
    │        [📁 Update Bundled Files]         │
    │        [📤 Export D365 Package]          │  ← NEW v1.5.0
    │        [📧 Send Email Report]            │  ← NEW v1.5.0
    │                                          │
    │ Status: ...                              │
    │ ┌─ Log ──────────────────────────────┐   │
    │ │ [time] message                     │   │
    │ └────────────────────────────────────┘   │
    └─────────────────────────────────────────┘

Responsibilities
----------------
* Wire the UI and state together.
* Auto-load bundled master/mapping on startup.
* Route user actions to the engine/exporter/template-writer.
* Surface progress in the Log panel and Status line.
* Gate D365 + Email actions on a successful SO generation (both need
  the ``ProcessingResult`` produced by ``generate()``).

The class is intentionally "procedural inside a class" — it holds Tk
widget references plus a few StringVars and path strings. Business
logic lives in ``engine`` / ``exporter`` / ``emailer`` modules; this
file is the thin layer on top.
"""

from __future__ import annotations
import logging
import os
import shutil
import time
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from online_po_processor.config.constants import (
    BUNDLED_DATA_FOLDER, BUNDLED_MAPPING_NAME, BUNDLED_MASTER_NAME,
)
from online_po_processor.config.email_config import get_email_config
from online_po_processor.config.marketplaces import (
    DEFAULT_WAREHOUSE, MARKETPLACE_CONFIGS, MARKETPLACE_NAMES,
    WAREHOUSE_CODES, WAREHOUSE_DISPLAY_NAMES,
)
from online_po_processor.config.paths import (
    get_bundled_data_folder, get_bundled_mapping_path,
    get_bundled_master_path, get_update_timestamp, record_update,
)
from online_po_processor.data.mapping_loader import MappingLoader
from online_po_processor.data.master_loader import MasterLoader
from online_po_processor.data.models import ProcessingResult
from online_po_processor.emailer import EmailSender
from online_po_processor.engine.marketplace_engine import MarketplaceEngine
from online_po_processor.exporter.d365_exporter import D365Exporter
from online_po_processor.exporter.so_exporter import SOExporter
from online_po_processor.gui._file_row import build_file_row
from online_po_processor.gui._update_dialog import UpdateDialog
from online_po_processor.utils.platform_open import open_file


# v2.3.1: bulk-consignment input flow. Originally Flipkart-TO-only;
# generalized so any marketplace (e.g. Meesho-TO) can opt in purely via
# config by declaring a ``consignment_mode`` block. The helper below is
# the single source of truth for "does this marketplace use the
# per-PO-file consignment flow?", so there are no scattered marketplace
# string literals in the GUI.
def _consignment_cfg(marketplace: str) -> Optional[dict]:
    """
    Return the marketplace's ``consignment_mode`` config dict when the
    bulk-consignment flow is enabled for it, else ``None``.

    Marketplaces opt in by setting ``consignment_mode.enabled = True``
    (Flipkart-TO, Meesho-TO). ``consolidated_option`` within that block
    further distinguishes:
      * True  → also offers the single consolidated-dump mode, so the GUI
                shows the two-way radio selector (Flipkart-TO).
      * False → bulk-consignment-only, no toggle needed (Meesho-TO).
    """
    cfg = MARKETPLACE_CONFIGS.get(marketplace, {}).get('consignment_mode')
    return cfg if cfg and cfg.get('enabled') else None


def _po_filetypes_for(marketplace: str) -> list:
    """
    Build a tkinter ``filetypes`` list for the PO file picker, based
    on the marketplace's ``source_format`` config (v2.2.0) and its
    optional ``accepted_extensions`` config (v2.3.0).

    Resolution order:

    1. If the marketplace's config has an explicit
       ``accepted_extensions: ['.csv', '.xlsx', '.xls']`` list, those
       are advertised in the dialog (in that order). This lets a
       marketplace opt into multi-format input — Blink uses this
       because its dashboard exports CSV now but the historical
       xlsx exports are still floating around on operators' disks.
    2. Otherwise, fall back to ``source_format``:
       * ``'pdf'``   → ``*.pdf`` filter (Dmart and future PDF
         marketplaces).
       * ``'excel'`` or unset → ``*.xlsx`` filter (the historical
         default that all other marketplaces still use).

    ``All files`` is always offered as a secondary option so the user
    can override if they have an oddly-named file. The engine
    auto-detects file type by extension regardless of what the user
    picked — so this filter is purely a UX convenience, never a
    correctness guarantee.
    """
    cfg = MARKETPLACE_CONFIGS.get(marketplace, {})

    # Path 1: explicit accepted_extensions list (most flexible)
    accepted = cfg.get('accepted_extensions')
    if accepted:
        # Build a single combined filter that accepts all listed
        # extensions, plus per-format filters so the user can narrow
        # down if they prefer. Tk needs the patterns space-separated
        # within ONE string for the combined filter.
        patterns = ' '.join(f'*{ext}' for ext in accepted)
        filters = [(f"{marketplace} PO files", patterns)]
        # Per-format filters for quick narrowing
        per_ext = {
            '.csv':  ('CSV files',   '*.csv'),
            '.xlsx': ('Excel 2007+', '*.xlsx'),
            '.xls':  ('Excel 97-03', '*.xls'),
            '.xlsm': ('Excel macro', '*.xlsm'),
            '.pdf':  ('PDF files',   '*.pdf'),
        }
        for ext in accepted:
            if ext in per_ext:
                filters.append(per_ext[ext])
        filters.append(('All files', '*.*'))
        return filters

    # Path 2: fall back to source_format
    source_format = cfg.get('source_format', 'excel')
    if source_format == 'pdf':
        return [
            (f"{marketplace} PO files", "*.pdf"),
            ("All files", "*.*"),
        ]
    # Default / 'excel'
    return [
        ("Excel files", "*.xlsx"),
        ("All files", "*.*"),
    ]


def _supports_multi_file(marketplace: str) -> bool:
    """
    True if this marketplace's GUI should offer a multi-file picker
    and its engine call should go through ``process_multi``.

    **PDF marketplaces** (Dmart, FirstCry, Reliance) qualify: every PDF
    PO is inherently one-PO-per-file (never a multi-PO container the way
    Blink's xlsx is), so the operator typically receives several PO PDFs
    and wants them combined into one SO batch. Marker:
    ``source_format == 'pdf'``.

    **Dual-format marketplaces** (Myntra, v2.4.1) also qualify: each
    Myntra PO arrives as its own PDF, so the operator picks several at
    once and wants them combined into one SO batch — exactly like the
    PDF-only marketplaces. Marker: a registered ``pdf_parser``. (When
    such a marketplace is instead fed Excel, ``process_multi`` still
    works — it processes each file via ``process()``, so a single
    multi-PO xlsx and a folder of one-PO PDFs both flow through.)

    **Flipkart** (v2.7.x) also qualifies: its new portal emits one
    ``purchase_order_<PO>.xlsx`` per PO, so it sets ``file_parser='flipkart'``
    and the operator drops all the day's PO files → one SO batch (replacing
    the old standalone dump generator). Marker: a registered ``file_parser``.

    Excel/CSV marketplaces that consolidate POs inside a single file
    (Blink, RK, Zepto, BlinkMP, Flipkart-TO) stay single-file — each upload
    already contains all the POs for that batch. (Flipkart-TO / Meesho-TO
    bulk-consignment mode is handled separately by ``_is_consignment_mode``.)
    """
    cfg = MARKETPLACE_CONFIGS.get(marketplace, {})
    return (cfg.get('source_format') == 'pdf'
            or bool(cfg.get('pdf_parser'))
            or bool(cfg.get('file_parser')))   # Big Basket: one .xlsx per PO


class OnlinePOApp:
    """GUI for Online Marketplace PO → SO generation."""

    # ── Construction ───────────────────────────────────────────────────

    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title("Online PO Processor — Marketplace SO Generator")
        # v2.3.1: +60px taller to fit the Flipkart-TO input-mode selector
        # without clipping the log/status area at the bottom.
        # v2.4.0: wider + shorter — the action buttons now sit in a
        # 2-column grid, so the old tall 520-wide window no longer fits the
        # content well. Allow resizing + a sensible minimum so nothing
        # clips on smaller displays.
        self.root.geometry("780x680")
        self.root.minsize(720, 600)
        self.root.resizable(True, True)

        # ── File paths (None until picked or auto-loaded) ───────────────
        self.master_path: Optional[str] = None
        self.mapping_path: Optional[str] = None
        self.po_path: Optional[str] = None
        # Multi-file upload support. Always a list — empty until the user
        # picks files, then one element for single-select marketplaces, or
        # N elements for batch-capable ones (PDF marketplaces, and the
        # Flipkart-TO / Meesho-TO bulk-consignment mode).
        self.po_paths: List[str] = []

        # v2.3.1: optional Consignment Visibility Report path — supplies
        # the PO → Warehouse Id (Location) lookup for Flipkart-TO bulk
        # consignment runs. None unless the operator picks one in that
        # mode; ignored by every other path.
        self.visibility_report_path: Optional[str] = None

        # ── Output tracking ─────────────────────────────────────────────
        self.last_output: Optional[Path] = None

        # v1.5.0: stash the full ProcessingResult from the last successful
        # generate() so the D365 and Email actions can reuse it without
        # re-running the engine. Reset to None on every new generate()
        # attempt; set to a real result only if the engine produced rows.
        self.last_result: Optional[ProcessingResult] = None

        # v2.7: the last generated result awaiting a manual "Push to DB"
        # (verify-then-confirm). {'result', 'output_path'} or None.
        self._pending_push: Optional[dict] = None

        # Track whether master/mapping came from the bundled folder (vs
        # user-picked). Used so the GUI can show "(auto-loaded)" and the
        # "Update Bundled Files" flow knows what's in use.
        self.master_is_bundled: bool = False
        self.mapping_is_bundled: bool = False

        # ── Engine-side state ───────────────────────────────────────────
        # MappingLoader is held on the app because it gets re-loaded each
        # run (when marketplace changes). MasterLoader is created fresh
        # in generate() — no state to carry between runs.
        self.mapping_loader = MappingLoader()
        self.exporter = SOExporter()

        # v1.5.0: D365 template filler — created once, reused across
        # multiple "Export D365 Package" clicks. Stateless so one
        # instance is fine.
        self.d365_exporter = D365Exporter()

        # ── Widget references populated by _build_ui ────────────────────
        self.marketplace_var: tk.StringVar
        self.marketplace_dropdown: ttk.Combobox
        self.margin_var: tk.StringVar
        self.margin_entry: tk.Entry
        self.master_var: tk.StringVar
        self.master_ts_var: tk.StringVar
        self.mapping_var: tk.StringVar
        self.mapping_ts_var: tk.StringVar
        self.po_var: tk.StringVar
        # v2.3.1: Flipkart-TO input-mode selector ('consolidated' or
        # 'consignments'). Only meaningful when Flipkart-TO is the
        # selected marketplace; disabled/ignored for all others.
        self.flipkart_to_mode_var: tk.StringVar
        self.open_btn: tk.Button
        # v1.5.0: D365 + email buttons start disabled; enabled after a
        # successful Generate SO run (same UX as open_btn).
        self.d365_btn: tk.Button
        self.email_btn: tk.Button
        self.status_var: tk.StringVar
        self.status_label: tk.Label
        self.log_text: tk.Text

        self._build_ui()

        # Auto-load AFTER the UI exists, so we can log and update
        # picker labels in one go.
        self._auto_load_bundled_files()

    # ── UI construction ────────────────────────────────────────────────

    def _build_ui(self) -> None:
        """Build the Tk widget tree."""

        # ── Title ───────────────────────────────────────────────────────
        tk.Label(
            self.root, text="Online PO Processor",
            font=("Arial", 14, "bold"),
        ).pack(pady=(12, 2))

        tk.Label(
            self.root, text="Marketplace PO → ERP Sales Order Import",
            font=("Arial", 9), fg='gray',
        ).pack(pady=(0, 10))

        # ── Marketplace selector + Margin input ─────────────────────────
        mkt_frame = tk.Frame(self.root)
        mkt_frame.pack(fill='x', padx=20, pady=(0, 8))

        tk.Label(
            mkt_frame, text="Marketplace:", font=("Arial", 10, "bold"),
        ).pack(side='left')

        self.marketplace_var = tk.StringVar(
            value=MARKETPLACE_NAMES[0] if MARKETPLACE_NAMES else ''
        )
        self.marketplace_dropdown = ttk.Combobox(
            mkt_frame, textvariable=self.marketplace_var,
            values=MARKETPLACE_NAMES, state='readonly', width=20,
        )
        self.marketplace_dropdown.pack(side='left', padx=8)
        self.marketplace_dropdown.bind(
            '<<ComboboxSelected>>', self._on_marketplace_change,
        )

        # Margin % — user can override per run (pre-filled from config)
        tk.Label(
            mkt_frame, text="Margin:", font=("Arial", 10, "bold"),
        ).pack(side='left', padx=(12, 0))
        self.margin_var = tk.StringVar(value=str(self._get_default_margin()))
        self.margin_entry = tk.Entry(
            mkt_frame, textvariable=self.margin_var, width=5,
            font=("Arial", 10), justify='center',
        )
        self.margin_entry.pack(side='left', padx=4)
        tk.Label(mkt_frame, text="%", font=("Arial", 10)).pack(side='left')
        tk.Label(
            mkt_frame, text="(Landing Cost)", font=("Arial", 8), fg='gray',
        ).pack(side='left', padx=4)

        # ── v1.9.0: Warehouse selector (own row — v1.9.2) ───────────────
        # Lives on its own row below Marketplace/Margin so the widgets
        # don't get clipped off the right edge of the 520px window.
        # Lets the user pick which RENEE warehouse fulfills this batch.
        # The friendly code (AHD/BLR/...) maps to the ERP location
        # string (PICK/DS_BL_OFF1/...) in config.WAREHOUSE_CODES and
        # gets stamped on every D365 Sales Header col K + Sales Line
        # col F. Default is AHD because that's the primary warehouse
        # for most batches today. Adding a warehouse = one line in
        # WAREHOUSE_CODES; no UI changes needed.
        wh_frame = tk.Frame(self.root)
        wh_frame.pack(fill='x', padx=20, pady=(0, 8))

        tk.Label(
            wh_frame, text="Warehouse:", font=("Arial", 10, "bold"),
        ).pack(side='left')
        self.warehouse_var = tk.StringVar(value=DEFAULT_WAREHOUSE)
        self.warehouse_combo = ttk.Combobox(
            wh_frame, textvariable=self.warehouse_var,
            values=WAREHOUSE_DISPLAY_NAMES, state='readonly', width=8,
            font=("Arial", 10),
        )
        self.warehouse_combo.pack(side='left', padx=8)
        # Live hint beside the dropdown — shows the ERP code for the
        # currently-selected warehouse (e.g. 'AHD → PICK'). Kept small
        # and gray so it reads as meta-info, not a primary field.
        self._warehouse_hint_var = tk.StringVar(
            value=f'→ {WAREHOUSE_CODES[DEFAULT_WAREHOUSE]}'
        )
        tk.Label(
            wh_frame, textvariable=self._warehouse_hint_var,
            font=("Arial", 9), fg='gray',
        ).pack(side='left', padx=4)
        self.warehouse_combo.bind(
            '<<ComboboxSelected>>', self._on_warehouse_change,
        )

        # ── v2.1.3: Override Unit Price toggle ──────────────────────────
        # Per-run checkbox controlling whether the engine-computed Cost
        # Price gets stamped into:
        #   * Lines (SO) col 8 (audit workbook)
        #   * D365 Sales Line col H (ERP import package)
        # Default state on marketplace change is taken from that
        # marketplace's ``override_unit_price`` config flag (currently
        # only BlinkMP has True). User can tick/untick either way per
        # run — this widget is the canonical runtime decision; the
        # config flag is only the default-state hint.
        ovr_frame = tk.Frame(self.root)
        ovr_frame.pack(fill='x', padx=20, pady=(0, 8))
        self.override_var = tk.BooleanVar(
            value=self._get_default_override(),
        )
        self.override_check = tk.Checkbutton(
            ovr_frame,
            text="Override Unit Price",
            variable=self.override_var, font=("Arial", 10, "bold"),
        )
        self.override_check.pack(side='left')
        # Hint label — explains what the toggle does at a glance.
        # Kept short; the Lines (SO) sheet itself carries the full
        # explanation as an info row when the toggle is on.
        tk.Label(
            ovr_frame,
            text="(stamps computed Cost Price into Sales Line Unit Price)",
            font=("Arial", 8), fg='gray',
        ).pack(side='left', padx=4)

        # ── v2.3.1: Transfer-Order input-mode selector ──────────────────
        # Shown for every marketplace but only LIVE for ones that declare
        # a ``consignment_mode`` (Flipkart-TO, Meesho-TO); greyed out
        # otherwise. Two shapes:
        #   * 'consolidated' — a single pre-merged dump → engine.process().
        #   * 'consignments' — raw per-PO files; the engine assembles the
        #     dump itself (PO from each filename) → process_consignments().
        # Marketplaces that ONLY support consignments (consolidated_option
        # = False, e.g. Meesho-TO) lock the selector to 'consignments' and
        # disable the consolidated radio. State is driven entirely by
        # config via _refresh_to_mode_widgets / _consignment_cfg.
        fkto_frame = tk.LabelFrame(
            self.root, text="Transfer Order Input Mode",
            font=("Arial", 9, "bold"), padx=10, pady=4,
        )
        fkto_frame.pack(fill='x', padx=20, pady=(0, 8))
        self._fkto_mode_frame = fkto_frame
        self.flipkart_to_mode_var = tk.StringVar(value='consolidated')
        self._fkto_radio_consolidated = tk.Radiobutton(
            fkto_frame, text="Consolidated dump (single file)",
            variable=self.flipkart_to_mode_var, value='consolidated',
            font=("Arial", 9), command=self._on_fkto_mode_change,
        )
        self._fkto_radio_consolidated.pack(anchor='w')
        self._fkto_radio_consignments = tk.Radiobutton(
            fkto_frame, text="Bulk consignment files (multi-file)",
            variable=self.flipkart_to_mode_var, value='consignments',
            font=("Arial", 9), command=self._on_fkto_mode_change,
        )
        self._fkto_radio_consignments.pack(anchor='w')

        # v2.3.1: optional location-report picker — only meaningful in
        # 'consignments' mode for marketplaces that define a location
        # report (Flipkart-TO's visibility report supplies each PO's
        # Warehouse Id). Hidden for marketplaces without one (Meesho-TO).
        # Shown/hidden by _refresh_fkto_report_row.
        self._fkto_report_row = tk.Frame(fkto_frame)
        tk.Button(
            self._fkto_report_row, text="Location Report (optional)…",
            font=("Arial", 8), command=self._select_visibility_report,
        ).pack(side='left')
        self.visibility_report_var = tk.StringVar(value="Not selected")
        tk.Label(
            self._fkto_report_row, textvariable=self.visibility_report_var,
            font=("Arial", 8), fg='gray',
        ).pack(side='left', padx=6)

        # Start in the correct state for whatever marketplace the dropdown
        # defaults to.
        self._refresh_to_mode_widgets()

        # ── File selectors ──────────────────────────────────────────────
        files_frame = tk.LabelFrame(
            self.root, text="Input Files", font=("Arial", 10, "bold"),
            padx=10, pady=8,
        )
        files_frame.pack(fill='x', padx=20, pady=(0, 8))

        # Items Master (with timestamp sub-line)
        self.master_var = tk.StringVar(value="Not selected")
        self.master_ts_var = tk.StringVar(value="")
        build_file_row(
            files_frame, "Items Master:", self.master_var,
            self._select_master, ts_var=self.master_ts_var,
        )

        # Ship-To Mapping (with timestamp sub-line)
        self.mapping_var = tk.StringVar(value="Not selected")
        self.mapping_ts_var = tk.StringVar(value="")
        build_file_row(
            files_frame, "Ship-To Mapping:", self.mapping_var,
            self._select_mapping, ts_var=self.mapping_ts_var,
        )

        # Marketplace PO (no timestamp sub-line — per-run input)
        self.po_var = tk.StringVar(value="Not selected")
        build_file_row(
            files_frame, "Marketplace PO:", self.po_var, self._select_po,
        )

        # ── Action buttons ──────────────────────────────────────────────
        # v2.4.0: laid out in a 2-column grid (instead of one tall column)
        # so the window is wider/shorter and every control is visible
        # without scrolling. The two primary actions (Generate SO / Auto
        # Mode) span the full width on top; secondary actions pair up below.
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=8)
        btn_frame.columnconfigure(0, weight=1, uniform='btn')
        btn_frame.columnconfigure(1, weight=1, uniform='btn')

        _BTN_W = 24

        # Row 0 — primary actions, side by side.
        tk.Button(
            btn_frame, text="▶  Generate SO", width=_BTN_W,
            font=("Arial", 10, "bold"),
            bg="#00C853", fg='white', command=self.generate,
        ).grid(row=0, column=0, padx=4, pady=4, sticky='ew')

        # Auto mode — headless batch window (process every Dump/Online/<mp>
        # folder). The single-file Generate flow above is unchanged.
        tk.Button(
            btn_frame, text="⚙  Auto Mode (all folders)", width=_BTN_W,
            font=("Arial", 10, "bold"),
            bg="#3949AB", fg='white', command=self._open_auto,
        ).grid(row=0, column=1, padx=4, pady=4, sticky='ew')

        # Rows 1-3 — secondary actions, two per row.
        self.open_btn = tk.Button(
            btn_frame, text="📂  Open Last Output", width=_BTN_W,
            state=tk.DISABLED, command=self.open_last,
        )
        self.open_btn.grid(row=1, column=0, padx=4, pady=4, sticky='ew')

        tk.Button(
            btn_frame, text="📜  View Order History", width=_BTN_W,
            command=self._view_history,
        ).grid(row=1, column=1, padx=4, pady=4, sticky='ew')

        tk.Button(
            btn_frame, text="📋  Download PO Template", width=_BTN_W,
            command=self._download_template,
        ).grid(row=2, column=0, padx=4, pady=4, sticky='ew')

        tk.Button(
            btn_frame, text="📁  Update Bundled Files", width=_BTN_W,
            command=self._update_bundled_files,
        ).grid(row=2, column=1, padx=4, pady=4, sticky='ew')

        # D365 + Email — disabled until a successful Generate (both need
        # ``self.last_result``).
        self.d365_btn = tk.Button(
            btn_frame, text="📤  Export D365 Package", width=_BTN_W,
            state=tk.DISABLED, command=self._export_d365,
        )
        self.d365_btn.grid(row=3, column=0, padx=4, pady=4, sticky='ew')

        self.email_btn = tk.Button(
            btn_frame, text="📧  Send Email Report", width=_BTN_W,
            state=tk.DISABLED, command=self._send_email,
        )
        self.email_btn.grid(row=3, column=1, padx=4, pady=4, sticky='ew')

        # Push to DB — a SEPARATE, deliberate confirm step. Generate only
        # writes + verifies the output (price diffs / address mismatches);
        # the operator records to the shared history DB only after checking.
        # Enabled after a successful Generate; disabled once pushed.
        self.push_db_btn = tk.Button(
            btn_frame, text="⤓  Push to DB (confirm after verifying)",
            width=_BTN_W, font=("Arial", 10, "bold"),
            bg="#2563eb", fg='white', activebackground="#1d4ed8",
            activeforeground='white', state=tk.DISABLED,
            command=self._push_to_db,
        )
        self.push_db_btn.grid(row=4, column=0, columnspan=2,
                               padx=4, pady=(4, 6), sticky='ew')

        # v2.4.6: Push Issues to DB — records ONLY the flagged lines (MISMATCH
        # / NOT_IN_MASTER) to the order_issue_lines audit table. Separate &
        # independent from the header "Push to DB"; append-with-guard (a
        # revised MRP/CP records a new dated snapshot, an identical re-push is
        # skipped). Enabled after a successful Generate.
        self.push_issues_btn = tk.Button(
            btn_frame, text="⚠  Push Issues to DB (mismatch / not-in-master)",
            width=_BTN_W, font=("Arial", 10, "bold"),
            bg="#b45309", fg='white', activebackground="#92400e",
            activeforeground='white', state=tk.DISABLED,
            command=self._push_issues_to_db,
        )
        self.push_issues_btn.grid(row=5, column=0, columnspan=2,
                                  padx=4, pady=(0, 6), sticky='ew')

        # ── Status line ─────────────────────────────────────────────────
        self.status_var = tk.StringVar(
            value="Status: Waiting — select files and generate"
        )
        self.status_label = tk.Label(
            self.root, textvariable=self.status_var,
            font=("Arial", 10), fg='gray', wraplength=460,
        )
        self.status_label.pack(pady=6)

        # ── Log panel ───────────────────────────────────────────────────
        log_frame = tk.LabelFrame(self.root, text="Log", font=("Arial", 9))
        log_frame.pack(fill='both', expand=True, padx=20, pady=(0, 12))

        scroll = ttk.Scrollbar(log_frame, orient='vertical')
        scroll.pack(side='right', fill='y')

        self.log_text = tk.Text(
            log_frame, height=6, font=("Consolas", 9),
            state='disabled', wrap='word',
            yscrollcommand=scroll.set,
        )
        self.log_text.pack(fill='both', expand=True)
        scroll.config(command=self.log_text.yview)

    # ── Logging helpers ────────────────────────────────────────────────

    def _log(self, msg: str) -> None:
        """Append a timestamped message to the log panel."""
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n")
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    # ── Margin helpers ─────────────────────────────────────────────────

    def _get_default_margin(self) -> int:
        """Default margin % for the currently selected marketplace."""
        mkt = (self.marketplace_var.get()
               if hasattr(self, 'marketplace_var') else '')
        if mkt and mkt in MARKETPLACE_CONFIGS:
            return MARKETPLACE_CONFIGS[mkt].get('default_margin', 70)
        return 70

    def _get_default_override(self) -> bool:
        """
        v2.1.3: Default state of the "Override Unit Price" checkbox
        for the currently-selected marketplace.

        Reads the marketplace's ``override_unit_price`` config flag —
        True means the box pre-checks on marketplace change, False
        means it pre-unchecks. The user can still toggle either way
        per run; this is just the convenience default.

        Currently only BlinkMP has the flag set to True (because BCPL
        is registered in BC at 70% but BlinkMP runs at 75% and the
        ERP would otherwise post wrong cost figures).
        """
        mkt = (self.marketplace_var.get()
               if hasattr(self, 'marketplace_var') else '')
        if mkt and mkt in MARKETPLACE_CONFIGS:
            return bool(MARKETPLACE_CONFIGS[mkt].get(
                'override_unit_price', False))
        return False

    def _on_marketplace_change(self, _event=None) -> None:
        """
        Reset margin and override-toggle to the newly-selected
        marketplace's defaults.

        v2.1.3: also refreshes the "Override Unit Price" checkbox state
        from the marketplace's ``override_unit_price`` config hint.
        Auto-checks for BlinkMP, auto-unchecks for everything else.
        Operator can still override either way before clicking Generate.
        """
        margin = self._get_default_margin()
        self.margin_var.set(str(margin))

        override_default = self._get_default_override()
        if hasattr(self, 'override_var'):
            self.override_var.set(override_default)

        # v2.3.1: sync the Transfer-Order input-mode widgets (radios +
        # optional report picker) to the newly-selected marketplace.
        self._refresh_to_mode_widgets()

        ovr_str = ' (Override Unit Price ON)' if override_default else ''
        self._log(f"Marketplace changed to {self.marketplace_var.get()}, "
                  f"margin set to {margin}%{ovr_str}")

    def _on_warehouse_change(self, _event=None) -> None:
        """
        Sync the gray hint label next to the warehouse dropdown so the
        user can see which ERP code their selection maps to.

        Example: switching dropdown from ``AHD`` to ``BLR`` updates
        the trailing hint from ``→ PICK`` to ``→ DS_BL_OFF1``. Useful
        when someone asks "why did BLR land on that weird code?" —
        the answer is visible right next to the dropdown instead of
        buried in config.
        """
        wh = self.warehouse_var.get()
        code = WAREHOUSE_CODES.get(wh, wh)
        self._warehouse_hint_var.set(f'→ {code}')
        self._log(f"Warehouse changed to {wh} (ERP code: {code})")

    # ── v2.3.1: Transfer-Order input-mode helpers (config-driven) ───────

    def _refresh_to_mode_widgets(self) -> None:
        """
        Sync the Transfer-Order input-mode radios + report row to the
        selected marketplace, entirely from config.

        Three cases, decided by :func:`_consignment_cfg`:
          * No consignment_mode → both radios greyed (the selector is
            inert; generate() ignores it for these marketplaces).
          * ``consolidated_option`` True (Flipkart-TO) → both radios live.
          * ``consolidated_option`` False (Meesho-TO) → bulk-consignment
            only: force 'consignments' and disable the consolidated radio.
        Then refresh the optional location-report row.
        """
        cfg = _consignment_cfg(self.marketplace_var.get())
        has_consignment = cfg is not None
        has_both = bool(cfg and cfg.get('consolidated_option', False))
        # Guard with hasattr so an early call during construction (before
        # the radios exist) is a harmless no-op.
        if hasattr(self, '_fkto_radio_consolidated'):
            if has_both:
                self._fkto_radio_consolidated.config(state=tk.NORMAL)
                self._fkto_radio_consignments.config(state=tk.NORMAL)
            elif has_consignment:
                # Consignment-only marketplace — lock to 'consignments'.
                self.flipkart_to_mode_var.set('consignments')
                self._fkto_radio_consolidated.config(state=tk.DISABLED)
                self._fkto_radio_consignments.config(state=tk.NORMAL)
            else:
                self._fkto_radio_consolidated.config(state=tk.DISABLED)
                self._fkto_radio_consignments.config(state=tk.DISABLED)
        self._refresh_fkto_report_row()

    def _refresh_fkto_report_row(self) -> None:
        """
        Show the optional location-report picker only when the run is in
        consignment mode AND the marketplace defines a location report.

        Flipkart-TO supplies a visibility report (``visibility_loc_col``
        in its consignment_mode); Meesho-TO has none yet, so the picker
        stays hidden for it. The report is also meaningless for the
        consolidated dump (which already carries a Location column).
        """
        if not hasattr(self, '_fkto_report_row'):
            return
        cfg = _consignment_cfg(self.marketplace_var.get())
        show = bool(
            self._is_consignment_mode()
            and cfg and cfg.get('visibility_loc_col')
        )
        if show:
            self._fkto_report_row.pack(anchor='w', pady=(2, 0))
        else:
            self._fkto_report_row.pack_forget()

    def _is_consignment_mode(self) -> bool:
        """
        True when the current run should use the bulk-consignment path.

        Driven by config:
          * Marketplace has no consignment_mode → False (unchanged
            single-file path).
          * consignment-only (``consolidated_option`` False, e.g.
            Meesho-TO) → always True.
          * offers both (Flipkart-TO) → True only when the radio is set
            to 'consignments'.

        Used by :meth:`_select_po` (multi-select CSV picker) and
        :meth:`generate` (routes to ``engine.process_consignments``).
        """
        cfg = _consignment_cfg(self.marketplace_var.get())
        if not cfg:
            return False
        if not cfg.get('consolidated_option', False):
            return True
        return (
            hasattr(self, 'flipkart_to_mode_var')
            and self.flipkart_to_mode_var.get() == 'consignments'
        )

    def _on_fkto_mode_change(self) -> None:
        """
        React to the operator flipping the Transfer-Order input-mode radios.

        The two modes use different pickers (single-file vs multi-file
        CSV), so a file picked under the old mode would be misleading
        under the new one — clear the current PO selection and prompt a
        fresh pick. Logged so the action is visible in the run log.
        """
        mode = self.flipkart_to_mode_var.get()
        # Reset any prior PO selection — the picker shape differs per mode.
        self.po_path = None
        self.po_paths = []
        if hasattr(self, 'po_var'):
            self.po_var.set("Not selected")
        # Show/hide the optional location-report picker for this mode.
        self._refresh_fkto_report_row()
        if mode == 'consignments':
            self._log(
                "Flipkart-TO mode: BULK CONSIGNMENTS — select the raw "
                "Consignment_Details_<PO>_<date>.csv files (multi-select). "
                "PO is read from each filename. Optionally add the "
                "Consignment Visibility Report so each PO's Warehouse Id "
                "fills the Location; without it, Location stays empty."
            )
        else:
            self._log(
                "Flipkart-TO mode: CONSOLIDATED — select the single "
                "pre-consolidated dump file."
            )

    def _select_visibility_report(self) -> None:
        """
        Pick the optional Consignment Visibility Report (CSV) for a
        Flipkart-TO bulk-consignment run.

        The report supplies each PO's Warehouse Id, which the engine
        writes as the row's Location (then resolved to a Transfer-to Code
        via the Ship-To B2B alias rows). Picking is optional — without it
        the run still works, just with empty Locations. A second pick
        replaces the first; there's no explicit "clear", which matches
        the other pickers' behaviour.
        """
        path = filedialog.askopenfilename(
            title="Select Flipkart Consignment Visibility Report (optional)",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            self.visibility_report_path = path
            self.visibility_report_var.set(os.path.basename(path))
            self._log(f"Location report: {os.path.basename(path)}")

    def _get_margin(self) -> float:
        """
        Current margin as a decimal (e.g. ``70`` → ``0.70``).

        Falls back to the marketplace default if the input field is
        empty or invalid. Valid range: 1..100 (inclusive).
        """
        try:
            val = float(self.margin_var.get().strip())
            if val <= 0 or val > 100:
                raise ValueError
            return val / 100.0
        except (ValueError, AttributeError):
            default = self._get_default_margin()
            self._log(f"Invalid margin input, using default {default}%")
            return default / 100.0

    # ── Bundled-file handling ─────────────────────────────────────────
    #
    # Items Master and Ship-To Mapping live in ``Calculation Data/``.
    # Startup auto-loads them so the user doesn't re-pick every run.
    # The "Update Bundled Files" button replaces what's in that folder.

    def _auto_load_bundled_files(self) -> None:
        """
        Look for Items Master + Ship-To Mapping in ``Calculation Data/``
        and pre-populate the picker labels if found.

        Does not abort startup on missing files — logs a hint and leaves
        the pickers in their default "Not selected" state.
        """
        master_p = get_bundled_master_path()
        mapping_p = get_bundled_mapping_path()

        if master_p:
            self.master_path = str(master_p)
            self.master_is_bundled = True
            self.master_var.set(f"✓ {master_p.name} (auto-loaded)")
            self._refresh_ts_label(self.master_ts_var, master_p.name)
            self._log(f"Auto-loaded master from "
                      f"{BUNDLED_DATA_FOLDER}/{master_p.name}")
        else:
            self._log(f"No bundled master at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_MASTER_NAME} "
                      f"— pick one manually or use 'Update Bundled Files'")

        if mapping_p:
            self.mapping_path = str(mapping_p)
            self.mapping_is_bundled = True
            self.mapping_var.set(f"✓ {mapping_p.name} (auto-loaded)")
            self._refresh_ts_label(self.mapping_ts_var, mapping_p.name)
            self._log(f"Auto-loaded mapping from "
                      f"{BUNDLED_DATA_FOLDER}/{mapping_p.name}")
        else:
            self._log(f"No bundled mapping at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_MAPPING_NAME} "
                      f"— pick one manually or use 'Update Bundled Files'")

    def _update_bundled_files(self) -> None:
        """
        Replace the bundled master and/or mapping in ``Calculation Data/``.

        Workflow:

        1. Ask which file(s) to update via :class:`UpdateDialog`.
        2. For each chosen kind, open a file picker and copy the picked
           file into the bundled folder under the canonical name.
        3. Refresh in-memory paths and picker labels.
        4. Log the outcome.
        """
        target_folder = get_bundled_data_folder(create=True)

        dialog = UpdateDialog(self.root, folder=target_folder)
        choice = dialog.show()
        if choice is None:
            return  # user cancelled

        updated_any = False

        if choice in ('master', 'both'):
            updated_any |= self._do_update_one_bundled(
                kind_label='Items Master',
                source_title='Select new Items Master file to bundle',
                target_path=target_folder / BUNDLED_MASTER_NAME,
                on_success=self._refresh_master_after_update,
            )

        if choice in ('mapping', 'both'):
            updated_any |= self._do_update_one_bundled(
                kind_label='Ship-To Mapping',
                source_title='Select new Ship-To Mapping file to bundle',
                target_path=target_folder / BUNDLED_MAPPING_NAME,
                on_success=self._refresh_mapping_after_update,
            )

        if updated_any:
            messagebox.showinfo(
                "Bundled Files Updated",
                f"Bundled files updated in:\n{target_folder}\n\n"
                f"Future runs will auto-load the new version.",
            )

    def _do_update_one_bundled(self, kind_label: str, source_title: str,
                                target_path: Path, on_success) -> bool:
        """
        Prompt for a source file and copy it to ``target_path``.

        Args:
            kind_label:   Display label (used in log/dialog text).
            source_title: Title of the file-picker dialog.
            target_path:  Destination (e.g.
                          ``Calculation Data/Items March.xlsx``).
            on_success:   Callback to refresh GUI state after a
                          successful copy.

        Returns:
            True if a copy was performed, False if the user cancelled
            or the copy failed.
        """
        src = filedialog.askopenfilename(
            title=source_title,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not src:
            self._log(f"Update cancelled for {kind_label}")
            return False

        try:
            shutil.copy2(src, str(target_path))
            # Stamp history BEFORE refresh so the sub-line shows the new
            # timestamp immediately.
            record_update(target_path.name)
            self._log(f"Bundled {kind_label} updated → {target_path}")
            on_success()
            return True
        except Exception as e:  # noqa: BLE001 — surface ANY copy error
            self._log(f"ERROR copying {kind_label}: {e}")
            messagebox.showerror(
                "Update Failed",
                f"Could not copy {kind_label}:\n{e}",
            )
            return False

    def _refresh_ts_label(self, ts_var: tk.StringVar, filename: str) -> None:
        """
        Refresh a timestamp StringVar from the in-app update history.

        Sets to ``"Updated: <date>"`` when there's a record, empty
        string otherwise (which renders as a blank sub-line).
        """
        ts = get_update_timestamp(filename)
        ts_var.set(f"Updated: {ts}" if ts else "")

    def _refresh_master_after_update(self) -> None:
        """Re-point in-memory master to the freshly bundled file."""
        p = get_bundled_master_path()
        if p:
            self.master_path = str(p)
            self.master_is_bundled = True
            self.master_var.set(f"✓ {p.name} (auto-loaded)")
            self._refresh_ts_label(self.master_ts_var, p.name)

    def _refresh_mapping_after_update(self) -> None:
        """Re-point in-memory mapping to the freshly bundled file."""
        p = get_bundled_mapping_path()
        if p:
            self.mapping_path = str(p)
            self.mapping_is_bundled = True
            self.mapping_var.set(f"✓ {p.name} (auto-loaded)")
            self._refresh_ts_label(self.mapping_ts_var, p.name)

    # ── Manual file pickers ────────────────────────────────────────────

    def _select_master(self) -> None:
        """
        Manually pick an Items Master file.

        Marks master as user-picked (not bundled); the bundled file in
        ``Calculation Data/`` is NOT touched — use "Update Bundled Files"
        for that.
        """
        path = filedialog.askopenfilename(
            title="Select Items Master file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.master_path = path
            self.master_is_bundled = False
            self.master_var.set(os.path.basename(path))
            # Clear the bundled timestamp — manual picks aren't tracked.
            self.master_ts_var.set("")
            self._log(f"Master (manual override): {os.path.basename(path)}")

    def _select_mapping(self) -> None:
        """
        Manually pick a Ship-To B2B mapping file. Bundled file untouched.
        """
        path = filedialog.askopenfilename(
            title="Select Mapping File (Ship-To B2B)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.mapping_path = path
            self.mapping_is_bundled = False
            self.mapping_var.set(os.path.basename(path))
            self.mapping_ts_var.set("")
            self._log(f"Mapping (manual override): {os.path.basename(path)}")

    def _select_po(self) -> None:
        """
        Pick the marketplace PO/punch file(s) for this run.

        v1.7.0: For Reliance, the dialog allows multi-file selection
        because each Reliance PO arrives as its own .xlsx file and
        users routinely process a batch of 5-10 POs at once. Other
        marketplaces stay single-select (Blink/Myntra/RK already
        consolidate multiple POs inside one file so there's no
        reason to complicate their upload flow).

        Result of the dialog is stored in two attributes:
            * ``self.po_path``  — first (or only) file path. Used by
              all the single-file display/log paths.
            * ``self.po_paths`` — list of all selected files.
              ``generate()`` uses this list for multi-file marketplaces
              (PDF-source: Dmart / FirstCry / Reliance).
        """
        marketplace = self.marketplace_var.get()
        # Centralized multi-file detection — PDF marketplaces
        # (source_format=pdf). See _supports_multi_file for the rationale.
        # v2.3.1: Flipkart-TO in 'consignments' mode is ALSO multi-file —
        # the operator picks many raw Consignment_Details CSVs at once.
        consignment = self._is_consignment_mode()
        supports_multi = _supports_multi_file(marketplace) or consignment

        if supports_multi:
            # v2.3.1: consignment mode is CSV-only (the raw exports are
            # always .csv); other multi-file marketplaces keep their
            # config-derived filters.
            filetypes = (
                [("Consignment CSV files", "*.csv"), ("All files", "*.*")]
                if consignment else _po_filetypes_for(marketplace)
            )
            title = (
                f"Select {marketplace} Consignment CSV files — "
                f"one per PO, pick many"
                if consignment else
                f"Select {marketplace} PO File(s) — pick one or many"
            )
            # Multi-select — returns a tuple (possibly empty).
            paths = filedialog.askopenfilenames(
                title=title,
                filetypes=filetypes,
            )
            if not paths:
                return  # user cancelled
            self.po_paths = list(paths)
            self.po_path = self.po_paths[0]
            n = len(self.po_paths)
            if n == 1:
                self.po_var.set(os.path.basename(self.po_path))
                self._log(f"PO file: {os.path.basename(self.po_path)}")
            else:
                self.po_var.set(f"{n} files selected")
                self._log(f"PO files: {n} selected for batch upload")
                for p in self.po_paths:
                    self._log(f"  • {os.path.basename(p)}")
        else:
            # Single-select — existing behavior.
            path = filedialog.askopenfilename(
                title="Select Marketplace PO File",
                filetypes=_po_filetypes_for(marketplace),
            )
            if path:
                self.po_path = path
                self.po_paths = [path]
                self.po_var.set(os.path.basename(path))
                self._log(f"PO file: {os.path.basename(path)}")

    # ── Main processing flow ───────────────────────────────────────────

    def generate(self) -> None:
        """Main action: load mapping → parse PO → generate output."""
        marketplace = self.marketplace_var.get()
        if not marketplace or marketplace not in MARKETPLACE_CONFIGS:
            messagebox.showwarning(
                "No Marketplace", "Please select a marketplace.",
            )
            return

        if not self.mapping_path:
            messagebox.showwarning(
                "No Mapping", "Please select the Ship-To mapping file.",
            )
            return

        if not self.po_path:
            messagebox.showwarning(
                "No PO File", "Please select the marketplace PO file.",
            )
            return

        config = MARKETPLACE_CONFIGS[marketplace]
        margin_pct = self._get_margin()
        start_time = time.time()

        self.status_var.set("Processing...")
        self.status_label.config(fg='blue')
        self.root.update()

        self._log(f"Marketplace: {marketplace} | "
                  f"Margin: {int(margin_pct * 100)}%")

        # ── Load mapping for this marketplace ───────────────────────────
        self._log(f"Loading mapping for '{marketplace}'...")
        warnings: List[Tuple[str, str, str]] = []
        loc_count = self.mapping_loader.load(
            self.mapping_path, config['party_name'], warnings,
        )

        if loc_count == 0:
            self._log("ERROR: No mapping locations found!")
            for _, _, msg in warnings:
                self._log(f"  {msg}")
            self.status_var.set("Failed — mapping load error")
            self.status_label.config(fg='red')
            return

        self._log(f"Loaded {loc_count} locations for {marketplace}")

        # v1.5.0: clear any previous run's stashed result BEFORE we do
        # anything else. If the new run fails anywhere below, the D365
        # and Email buttons will disable themselves — we don't want
        # them acting on stale data from the previous successful run.
        self.last_result = None
        self.d365_btn.config(state=tk.DISABLED)
        self.email_btn.config(state=tk.DISABLED)
        # A new generate supersedes any un-pushed result.
        self._pending_push = None
        self.push_db_btn.config(state=tk.DISABLED)
        self.push_issues_btn.config(state=tk.DISABLED)

        # ── Load Items_March (master) ───────────────────────────────────
        master_loader: Optional[MasterLoader] = None
        if self.master_path:
            self._log("Loading Items_March for validation...")
            master_loader = MasterLoader()
            try:
                item_count = master_loader.load(self.master_path)
                self._log(f"Loaded {item_count:,} items from master")
            except Exception as e:  # noqa: BLE001
                self._log(f"WARNING: Master load failed: {e} "
                          f"— skipping validation")
                master_loader = None

        # ── Engine run ──────────────────────────────────────────────────
        # Route PDF marketplaces (Dmart / FirstCry / Reliance) through
        # process_multi — each PO PDF is a separate file, so operators
        # picking several in one batch expect them combined into one SO
        # batch. Marker: ``_supports_multi_file(marketplace)``
        # (``source_format == 'pdf'``).
        engine = MarketplaceEngine(self.mapping_loader, master=master_loader)

        # v2.3.1: Flipkart-TO 'consignments' mode takes precedence — the
        # engine assembles the consolidated dump from the raw per-PO CSVs
        # (PO from each filename, Location left empty) then runs the
        # standard TO pipeline. All other paths are unchanged.
        if self._is_consignment_mode():
            rpt = self.visibility_report_path
            self._log(
                f"{marketplace} BULK CONSIGNMENTS — assembling dump from "
                f"{len(self.po_paths)} consignment file(s)"
                + (f"; location report: {os.path.basename(rpt)}"
                   if rpt else "; no location report (Locations will be "
                   "empty)")
                + "..."
            )
            result = engine.process_consignments(
                self.po_paths, config, margin_pct=margin_pct,
                visibility_report_path=rpt,
            )
        else:
            supports_multi = _supports_multi_file(marketplace)
            if supports_multi and len(self.po_paths) > 1:
                self._log(
                    f"Batch processing {len(self.po_paths)} "
                    f"{marketplace} files..."
                )
                result = engine.process_multi(
                    self.po_paths, config, margin_pct=margin_pct,
                )
            else:
                self._log(f"Processing {os.path.basename(self.po_path)}...")
                result = engine.process(
                    self.po_path, config, margin_pct=margin_pct,
                )
        result.margin_pct = margin_pct  # redundant but explicit

        # v1.9.0: stamp the GUI's warehouse selection onto the result
        # so D365Exporter uses the right Location Code (col K + F)
        # and the Summary footer + email banner can show which
        # warehouse fulfilled this batch.
        selected_wh = self.warehouse_var.get()
        result.warehouse_display = selected_wh
        result.warehouse_code = WAREHOUSE_CODES.get(selected_wh, 'PICK')
        self._log(
            f"Warehouse: {selected_wh} → ERP code {result.warehouse_code}"
        )

        # v2.1.3: stamp the runtime override flag from the GUI checkbox
        # onto the result. Consumed by:
        #   * lines_sheet.write — populates Lines (SO) col 8 + adds
        #     header tint and footer info-row when True
        #   * d365_exporter.export — populates Sales Line col H when True
        # Default False if the widget is missing (older code paths /
        # tests / direct API calls).
        result.override_unit_price = bool(
            getattr(self, 'override_var', None)
            and self.override_var.get()
        )
        if result.override_unit_price:
            self._log(
                "Override Unit Price: ON — Sales Line col H will carry "
                "computed Cost Price"
            )

        # ── v2.4.6: Flipkart Tracker from the optional header file ──────────
        # Flipkart's per-PO Order Value / Qty / dates live in the portal's
        # 'purchase-orders-*.csv' PO-list (the "header file"), not in the
        # individual PO xlsx. Offer to upload it; if provided, build the
        # Tracker rows (Market Place by location, locked mapping) and stamp
        # them on the result so the exporter writes a 'Tracker' sheet.
        if marketplace == 'Flipkart':
            self._maybe_load_flipkart_header(result)

        if not result.rows:
            self._log("ERROR: No valid rows extracted!")
            for _, _, msg in result.warnings:
                self._log(f"  WARNING: {msg}")
            self.status_var.set("Failed — no data extracted")
            self.status_label.config(fg='red')
            return

        # ── v2.4.0: dedup-skip — drop already-uploaded POs from output ──
        try:
            from online_po_processor.auto.history_db import apply_dedup
            _skipped = apply_dedup(result)
            if _skipped:
                self._log(
                    f"Dedup: {len(_skipped)} PO(s) already uploaded — removed "
                    f"from Headers/Lines (see 'Skipped POs' sheet)")
        except Exception as de:  # noqa: BLE001 — never block a generate
            self._log(f"WARNING: dedup check skipped ({type(de).__name__}: {de})")

        if not result.rows:
            # Every PO in this file was already uploaded — nothing new, but
            # still write the workbook so the Skipped sheet is the record.
            self._log("All POs already uploaded — nothing new to generate.")

        # ── Log summary ─────────────────────────────────────────────────
        unique_pos = {r.po_number for r in result.rows}
        total_qty = sum(r.qty for r in result.rows)

        self._log(f"Extracted: {len(result.rows)} items, "
                  f"{len(unique_pos)} PO(s), {total_qty} total qty")
        if result.warnings:
            self._log(f"Warnings: {len(result.warnings)}")
            for po, _loc, msg in result.warnings[:5]:
                self._log(f"  [{po}] {msg}")
            if len(result.warnings) > 5:
                self._log(f"  ... and {len(result.warnings) - 5} more "
                          f"(see Warnings sheet)")

        # ── Export ──────────────────────────────────────────────────────
        # v2.1.0: pass start_time so the exporter can stamp the full
        # pipeline elapsed time onto result.elapsed_seconds BEFORE
        # writing the Summary sheet — that's what makes the duration
        # visible in the file's footer. Without this, the Summary
        # writer sees elapsed_seconds=None and silently omits the
        # duration segment.
        self._log("Writing output...")
        output_path = self.exporter.export(result, start_time=start_time)

        elapsed = time.time() - start_time

        if output_path:
            self.last_output = output_path
            self.open_btn.config(state=tk.NORMAL)

            # v1.5.0: stash the result for the D365 + Email actions and
            # record elapsed time on the result itself (used by the
            # email report footer). Enable the downstream buttons.
            result.elapsed_seconds = elapsed
            self.last_result = result
            self.d365_btn.config(state=tk.NORMAL)
            self.email_btn.config(state=tk.NORMAL)

            status_msg = (f"Done — {len(result.rows)} items, "
                          f"{len(unique_pos)} PO(s), "
                          f"{total_qty} qty | {elapsed:.2f}s")
            if result.warnings:
                status_msg += f" | {len(result.warnings)} warning(s)"
                self.status_label.config(fg='orange')
            else:
                self.status_label.config(fg='darkgreen')

            self.status_var.set(status_msg)
            self._log(f"Saved: {output_path}")

            # v2.7: recording to the shared history DB is now a SEPARATE,
            # deliberate step. The operator VERIFIES the generated output
            # (price differences, address mismatches, warnings) and only
            # then clicks "Push to DB" to confirm. We stash the verified-
            # pending result here; the push happens in _push_to_db().
            self._pending_push = {'result': result,
                                   'output_path': str(output_path)}
            self.push_db_btn.config(state=tk.NORMAL)
            # v2.4.6: enable the Issues push only when there ARE flagged lines.
            _n_issues = sum(
                1 for r in result.rows
                if getattr(r, 'validation_status', '') in
                ('MISMATCH', 'NOT_IN_MASTER'))
            self.push_issues_btn.config(
                state=tk.NORMAL if _n_issues else tk.DISABLED)
            hist_line = ("\nHistory     : ⏳ NOT recorded yet — verify the "
                         "output, then click 'Push to DB'")

            answer = messagebox.askyesno(
                "SO Generated",
                f"Sales Order generated successfully!\n\n"
                f"Marketplace : {marketplace}\n"
                f"PO(s)       : {len(unique_pos)}\n"
                f"Items       : {len(result.rows)}\n"
                f"Total Qty   : {total_qty}\n"
                f"Warnings    : {len(result.warnings)}\n"
                f"Time        : {elapsed:.2f}s"
                f"{hist_line}\n\n"
                f"Do you want to open the output file?",
            )
            if answer:
                open_file(output_path)
        else:
            self.status_var.set("Failed — no output generated")
            self.status_label.config(fg='red')

    def _open_auto(self) -> None:
        """
        Open the Auto-mode batch window (v2.4.0).

        Imported lazily so app startup and the Manual flow carry no
        dependency on the Auto UI. Passes the master/mapping files the
        main window already loaded so Auto reuses them.
        """
        from online_po_processor.gui.auto_window import AutoWindow
        AutoWindow(self.root, self.master_path, self.mapping_path)

    def _view_history(self) -> None:
        """
        v2.4.0: open the shared order history (the same DB Manual + Auto
        write to) as a readable Excel grid. Gives Manual mode direct
        visibility into what's been recorded / uploaded.
        """
        from online_po_processor.auto.history_db import (
            default_history_db_path, get_history_store,
        )
        db_path = default_history_db_path()
        if not os.path.exists(db_path):
            messagebox.showinfo(
                "No History Yet",
                "Nothing recorded yet.\n\nGenerate at least one SO (or run "
                "Auto mode) and the order history will start filling in here.",
            )
            return
        out = os.path.join(os.path.dirname(str(db_path)), 'Order_History.xlsx')
        store = get_history_store(db_path)
        try:
            store.export_to_xlsx(out)
        finally:
            store.close()
        self._log(f"Order history exported: {out}")
        open_file(out)

    def open_last(self) -> None:
        """Open the last generated output file in the default app."""
        if self.last_output and self.last_output.exists():
            open_file(self.last_output)
        else:
            messagebox.showwarning("Not Found", "Output file not found.")

    def _push_to_db(self) -> None:
        """
        Record the LAST generated result into the shared history DB — a
        deliberate, separate step the operator runs ONLY after verifying
        the output (prices, addresses, mismatches). Logs the exact SQL
        being executed so it's visible in the Log panel.
        """
        pend = getattr(self, '_pending_push', None)
        if not pend:
            messagebox.showinfo(
                "Push to DB",
                "Generate an SO first, verify the output, then push.")
            return
        result = pend['result']
        output_path = pend['output_path']
        if not result.rows:
            messagebox.showinfo(
                "Push to DB",
                "Nothing new to record (all POs were already uploaded).")
            self._pending_push = None
            self.push_db_btn.config(state=tk.DISABLED)
            return

        n_pos = len({r.po_number for r in result.rows})
        if not messagebox.askyesno(
                "Push to DB",
                f"Record this output to the history DB?\n\n"
                f"Marketplace : {result.marketplace}\n"
                f"PO(s)       : {n_pos}\n"
                f"Items       : {len(result.rows)}\n\n"
                f"Do this ONLY after verifying prices / addresses."):
            return

        try:
            from online_po_processor.auto.history_db import (
                default_dump_root, record_manual, order_rows_from_result,
            )
            from online_po_processor.auto.consolidated_exporter import (
                export_tracker_from_db,
            )
            # Build the rows (same ones record_manual will insert) and log
            # the SQL so the operator can see exactly what hits the DB.
            wh = getattr(result, 'warehouse_display', '') or ''
            rows = order_rows_from_result(
                result, result.marketplace, wh, os.path.basename(output_path))
            self._log("─── Push to DB — SQL being executed ───")
            _cols = ['segment', 'marketplace', 'marketplace_label', 'po',
                     'location', 'warehouse', 'po_date', 'exp_date',
                     'order_type', 'items', 'qty', 'order_value', 'output_file']
            for o in rows:
                _vals = ', '.join(
                    (repr(o[c]) if isinstance(o[c], str) else str(o[c]))
                    for c in _cols)
                self._log(f"INSERT INTO order_headers "
                          f"({', '.join(_cols)}) VALUES ({_vals});")

            hinfo = record_manual(result, output_path)
            n_skip = hinfo['skipped']
            self._log(
                f"History: {hinfo['new_orders']} new PO(s) recorded "
                + (f"(run #{hinfo['run_id']})" if hinfo['run_id']
                   else "(nothing new)")
                + (f" — {n_skip} already-uploaded removed" if n_skip else ""))
            if hinfo['run_id']:
                try:
                    tpath = export_tracker_from_db(
                        hinfo['run_id'], default_dump_root())
                    self._log(f"Tracker (new POs): {tpath}")
                except Exception as te:  # noqa: BLE001
                    self._log(f"WARNING: tracker build failed: "
                              f"{type(te).__name__}: {te}")
            messagebox.showinfo(
                "Push to DB",
                f"Recorded {hinfo['new_orders']} new PO(s) to the history DB."
                + (f"\n{n_skip} already-uploaded removed."
                   if n_skip else ""))
            self._pending_push = None
            self.push_db_btn.config(state=tk.DISABLED)
        except Exception as e:  # noqa: BLE001 — surface, never crash
            self._log(f"ERROR: Push to DB failed: {type(e).__name__}: {e}")
            messagebox.showerror(
                "Push to DB",
                f"Not recorded:\n{type(e).__name__}: {e}\n\n"
                f"Fix the issue and try again.")

    def _push_issues_to_db(self) -> None:
        """v2.4.6: record ONLY the flagged lines (MISMATCH / NOT_IN_MASTER) of
        the last result to the ``order_issue_lines`` audit table. Append with
        the value-aware guard — a revised MRP/CP is a new dated snapshot, an
        identical re-push is skipped. Independent of the header Push to DB."""
        pend = getattr(self, '_pending_push', None)
        if not pend:
            messagebox.showinfo(
                "Push Issues to DB",
                "Generate an SO first, verify the output, then push issues.")
            return
        result = pend['result']
        output_path = pend['output_path']

        from online_po_processor.auto.history_db import (
            issue_lines_from_result, record_issue_lines_manual,
        )
        preview = issue_lines_from_result(result, os.path.basename(output_path))
        if not preview:
            messagebox.showinfo(
                "Push Issues to DB",
                "No flagged lines (MISMATCH / NOT_IN_MASTER) in this output — "
                "nothing to record.")
            return

        if not messagebox.askyesno(
                "Push Issues to DB",
                f"Record the FLAGGED lines to the issue-audit table?\n\n"
                f"Marketplace : {result.marketplace}\n"
                f"Flagged     : {len(preview)} line(s) "
                f"(mismatch / not-in-master)\n\n"
                f"Unchanged lines already pushed are skipped; revised values "
                f"are recorded as a new snapshot."):
            return

        try:
            self._log("─── Push Issues to DB — order_issue_lines ───")
            for ln in preview:
                self._log(
                    f"  [{ln['status']}] PO {ln['po']} item {ln['item_no']} "
                    f"vMRP={ln['vendor_mrp']} oMRP={ln['our_mrp']} "
                    f"vCP={ln['vendor_cp']} oCP={ln['our_cp']} "
                    f"diff={ln['diff']}")
            info = record_issue_lines_manual(result, output_path)
            self._log(
                f"Issues: {info['recorded']} recorded "
                f"(new/revised), {info['skipped']} unchanged skipped "
                f"(of {info['total_issues']} flagged).")
            messagebox.showinfo(
                "Push Issues to DB",
                f"Recorded {info['recorded']} flagged line(s) "
                f"(new/revised).\n"
                f"{info['skipped']} unchanged (already pushed) skipped.")
        except Exception as e:  # noqa: BLE001
            self._log(f"ERROR: Push Issues failed: {type(e).__name__}: {e}")
            messagebox.showerror(
                "Push Issues to DB",
                f"Not recorded:\n{type(e).__name__}: {e}")

    # ── v2.4.6: Flipkart Tracker header file ─────────────────────────────

    def _maybe_load_flipkart_header(self, result) -> None:
        """Ask whether to upload the Flipkart header file (PO-list CSV) and,
        if so, build the Tracker rows onto ``result.flipkart_tracker_rows``.

        Best-effort: any failure logs a warning and leaves the tracker empty
        (the rest of the output is unaffected)."""
        try:
            want = messagebox.askyesno(
                "Flipkart Tracker",
                "Upload the Flipkart header file (the portal "
                "'purchase-orders-*.csv' PO list) to generate the Tracker "
                "sheet?\n\nYes → pick the CSV.   No → skip the Tracker.",
            )
        except Exception:  # noqa: BLE001
            want = False
        if not want:
            self._log("Flipkart Tracker: header file not uploaded — skipped.")
            return

        path = filedialog.askopenfilename(
            title="Select Flipkart header file (purchase-orders-*.csv)",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            self._log("Flipkart Tracker: no file chosen — skipped.")
            return

        try:
            from online_po_processor.engine.flipkart_tracker import (
                build_flipkart_tracker, unknown_locations,
            )
            rows = build_flipkart_tracker(path)
        except Exception as e:  # noqa: BLE001
            self._log(f"Flipkart Tracker: could not read header file — {e}")
            messagebox.showwarning(
                "Flipkart Tracker",
                f"Couldn't build the Tracker from that file:\n\n{e}\n\n"
                f"Pick the portal 'purchase-orders-*.csv' export.",
            )
            return

        result.flipkart_tracker_rows = rows
        unk = unknown_locations(rows)
        self._log(f"Flipkart Tracker: {len(rows)} PO(s) from "
                  f"{os.path.basename(path)}"
                  + (f"; {len(unk)} unknown location(s) → 'FK (review)': "
                     f"{', '.join(unk)}" if unk else ""))

    # ── v1.5.0: D365 Package Export ──────────────────────────────────────

    def _export_d365(self) -> None:
        """
        Fill the BUNDLED D365 connector template with the last result.

        v2.4.3: no longer prompts for a template — the bound
        'Abhishek-Wagh' connector workbook ships with the app
        (online_po_processor/templates/) and is filled via the
        binding-preserving ZIP surgery (``d365_package.export_d365_package``),
        so the XML-map → D365 field binding survives and the file is
        ready to Publish. The main "Generate SO" flow already produces this
        automatically; this button just re-runs it for ``self.last_result``.

        Flow:
            1. Guard: ``self.last_result`` must be populated.
            2. Warn about any PO(s) whose Ship-To mapping failed (they
               export with empty Location Code).
            3. Pick the bundled SO/TO template by ``output_type`` — no
               file dialog.
            4. Fill via ``export_d365_package`` (binding preserved).
            5. Offer to open the resulting file.

        The output lands in the same ``output/`` folder as the main SO
        export so both artefacts sit side by side.
        """
        result = self.last_result
        if result is None or not result.rows:
            messagebox.showwarning(
                "No Data",
                "Generate an SO successfully first before exporting to "
                "D365.",
            )
            return

        # ── Step 2: warn about unmapped POs ─────────────────────────────
        # A PO is "unmapped" if any of its rows came back with
        # ``mapped=False`` from the engine — meaning the facility name
        # didn't match any entry in Ship-To B2B for this marketplace.
        unmapped_pos = sorted({
            r.po_number for r in result.rows if not r.mapped
        })

        if unmapped_pos:
            preview = "\n".join(f"  • {p}" for p in unmapped_pos[:10])
            more = (
                f"\n  … and {len(unmapped_pos) - 10} more"
                if len(unmapped_pos) > 10 else ""
            )
            proceed = messagebox.askyesno(
                "⚠️ Unmapped Ship-To Locations",
                f"{len(unmapped_pos)} PO(s) have no Ship-To mapping "
                f"and will export with EMPTY Location Code:\n\n"
                f"{preview}{more}\n\n"
                f"D365 import may reject or warn on these rows.\n\n"
                f"Continue with export anyway?",
            )
            if not proceed:
                self._log("D365 export cancelled by user (unmapped POs).")
                return

        # ── Step 3: use the BUNDLED connector template (NO manual pick) ──
        # v2.4.3: the bound 'Abhishek-Wagh' connector template ships with the
        # app (online_po_processor/templates/), so the operator never selects
        # a file. SO vs TO is chosen automatically by the result's
        # output_type. This is the SAME bound template the main "Generate SO"
        # flow now fills automatically — this button just re-runs it for the
        # last result on demand.
        is_to = getattr(result, 'output_type', 'so') == 'to'
        template_path = (SOExporter._D365_TO_TEMPLATE if is_to
                         else SOExporter._D365_SO_TEMPLATE)
        if not Path(template_path).exists():
            self.status_var.set("D365 template missing")
            self.status_label.config(fg='red')
            messagebox.showerror(
                "D365 Template Missing",
                f"The bundled D365 template is missing:\n\n{template_path}\n\n"
                f"Restore it under online_po_processor/templates/.",
            )
            return

        # ── Step 4: fill via the binding-preserving surgery ─────────────
        # Output mirrors where the main SO workbook landed
        # (``<punch_dir>/output/``) so every artefact lands together.
        from online_po_processor.exporter.d365_package import (
            export_d365_package,
        )
        punch_dir = Path(result.input_file_path).parent
        output_dir = punch_dir / 'output'
        output_dir.mkdir(parents=True, exist_ok=True)
        ts = time.strftime("%d-%m-%Y_%H%M%S")
        slug = (result.marketplace or 'online').lower().replace(' ', '_')
        kind = 'to' if is_to else 'so'
        d365_out = output_dir / f"{slug}_d365_{kind}_{ts}.xlsx"

        self._log("D365: filling bundled connector template "
                  f"({os.path.basename(str(template_path))})...")
        self.status_var.set("D365 export in progress...")
        self.status_label.config(fg='blue')
        self.root.update()

        try:
            d365_path = export_d365_package(result, template_path, d365_out)
        except Exception as e:  # noqa: BLE001
            logging.exception("D365 export crashed unexpectedly")
            self.status_var.set("D365 export failed")
            self.status_label.config(fg='red')
            messagebox.showerror(
                "D365 Export Failed",
                f"An unexpected error occurred:\n\n{e}",
            )
            return

        if d365_path is None:
            self.status_var.set("D365 export failed")
            self.status_label.config(fg='red')
            self._log("D365 export returned no file — check log for cause.")
            messagebox.showerror(
                "D365 Export Failed",
                "Could not produce the D365 import file. Check the log "
                "panel for details.",
            )
            return

        # ── Step 5: success popup ───────────────────────────────────────
        self._log(f"D365 file saved → {d365_path}")
        self.status_var.set(f"D365 export done — {d365_path.name}")
        self.status_label.config(fg='darkgreen')

        po_count = len({r.po_number for r in result.rows})
        item_count = len(result.rows)

        answer = messagebox.askyesno(
            "D365 Package Exported",
            f"D365 import file created successfully!\n\n"
            f"File  : {d365_path.name}\n"
            f"POs   : {po_count}\n"
            f"Items : {item_count}\n\n"
            f"Open the file now?",
        )
        if answer:
            open_file(d365_path)

    # ── v1.5.0: Email Report ─────────────────────────────────────────────

    def _send_email(self) -> None:
        """
        Send the HTML report email for the last generated result.

        Flow:
            1. Guard on ``self.last_result`` (defense-in-depth — the
               button is disabled when there's nothing to send).
            2. Freeze the UI with a "Sending..." status while SMTP is
               in flight.
            3. Ask :class:`EmailSender` to build + send the report.
            4. Show a success info box or a clear error box based on
               the return value.

        We don't disable the button during send — the root.update()
        freeze plus the status label are sufficient feedback, and
        an accidental double-click is handled fine by the stateless
        SMTP layer (it'll just send twice).
        """
        result = self.last_result
        if result is None or not result.rows:
            messagebox.showwarning(
                "No Data",
                "Generate an SO successfully first before emailing the "
                "report.",
            )
            return

        # ── Load effective config (defaults + optional JSON overrides)
        # Reloaded every send so edits to ``email_config.json`` take
        # effect without restarting the app.
        email_config = get_email_config()

        # ── Show "sending" state ────────────────────────────────────────
        self._log(
            f"Email: sending to {email_config['DEFAULT_RECIPIENT']}..."
        )
        self.status_var.set("Sending email...")
        self.status_label.config(fg='blue')
        self.root.update()

        # ── Dispatch ────────────────────────────────────────────────────
        ok, err = EmailSender.send(result, email_config)

        if ok:
            self.status_var.set("Email sent ✓")
            self.status_label.config(fg='darkgreen')
            self._log(
                f"Email sent OK → {email_config['DEFAULT_RECIPIENT']}"
                f" + {len(email_config.get('CC_RECIPIENTS', []))} CC"
            )

            cc_list = email_config.get('CC_RECIPIENTS', []) or []
            cc_display = ', '.join(cc_list) if cc_list else 'none'

            messagebox.showinfo(
                "Email Sent",
                f"Report sent successfully!\n\n"
                f"To : {email_config['DEFAULT_RECIPIENT']}\n"
                f"CC : {cc_display}",
            )
        else:
            self.status_var.set("Email failed")
            self.status_label.config(fg='red')
            self._log(f"Email failed: {err}")
            messagebox.showerror("Email Failed", err)

    # ── PO template download ──────────────────────────────────────────

    def _download_template(self) -> None:
        """
        v2.1.3: Generate a SINGLE workbook containing one sheet per
        marketplace — the master PO template covering every supported
        marketplace.

        Sheets are named after the marketplace dict keys, with one
        exception: BlinkMP gets two sheets ('BlinkMP (BLR)' and
        'BlinkMP (AHD)') because both formats are valid in production
        and the AHD/BLR ops teams use different column layouts. The
        engine accepts either format at runtime via list-aliased
        ``po_col`` / ``fob_col``; this template just shows both.

        Pre-v2.1.3 this method generated a single-sheet workbook for
        the currently-selected marketplace only, with a 'How this
        works' reference sheet attached. The reference sheet pattern
        doesn't scale to 9 marketplaces (would produce 18 sheets),
        so it's dropped — the colour-coded headers + legend rows on
        each marketplace sheet still convey the same information.

        Headers are colour-coded the same way as before:

        * **BLUE** (``#1A237E``) — Required. Script fails without these.
        * **GREEN** (``#1B5E20``) — Validation. Used for price check +
          master lookup.
        * **GREY** (``#9E9E9E``) — Not read by script. Kept only to
          mirror the marketplace's native file format.

        Each sheet's legend rows list which columns fall in each bucket.
        """
        save_path = filedialog.asksaveasfilename(
            title="Save All Marketplaces PO Template",
            defaultextension=".xlsx",
            initialfile="PO_Templates_All_Marketplaces.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            self._write_master_template_workbook(save_path)
            self._log(f"All-marketplaces template saved → {save_path}")
            n_sheets = sum(1 for cfg in MARKETPLACE_CONFIGS.values()
                            if 'template_headers_extra' in cfg) + len(MARKETPLACE_CONFIGS)
            messagebox.showinfo(
                "Templates Saved",
                f"Master PO template saved to:\n{save_path}\n\n"
                f"{n_sheets} sheets (one per marketplace; BlinkMP has "
                f"two for AHD and BLR formats).\n\n"
                f"Header colours:\n"
                f"  • Blue  = Required (must fill)\n"
                f"  • Green = Validation (recommended)\n"
                f"  • Grey  = Not read by script",
            )
        except Exception as e:  # noqa: BLE001
            self._log(f"Template save failed: {e}")
            messagebox.showerror(
                "Error", f"Failed to save template:\n{e}",
            )

    @classmethod
    def _write_master_template_workbook(cls, save_path: str) -> None:
        """
        v2.1.3: Build the multi-sheet master PO template workbook.

        Iterates over every marketplace in ``MARKETPLACE_CONFIGS`` and
        appends one sheet per marketplace (two for BlinkMP — AHD and
        BLR formats). Sheet names match dict keys, except BlinkMP
        which becomes 'BlinkMP (BLR)' + 'BlinkMP (AHD)'.

        Args:
            save_path: Where to save the .xlsx file.
        """
        wb = Workbook()
        # Workbook() auto-creates an empty 'Sheet' — remove it so the
        # final book contains only our named marketplace sheets.
        wb.remove(wb.active)

        for marketplace, config in MARKETPLACE_CONFIGS.items():
            extra = config.get('template_headers_extra')
            if extra:
                # BlinkMP today is the only marketplace with an extra
                # variant. Renders TWO sheets: the primary one (named
                # '<marketplace> (BLR)' for BlinkMP — the current canonical
                # format) and an extra one (named '<marketplace> (AHD)').
                # Hardcoding the suffix here because the only marketplace
                # using template_headers_extra is BlinkMP and its two
                # variants are AHD/BLR. If a future marketplace gains a
                # similar split, generalise the suffixing then.
                cls._append_marketplace_template_sheet(
                    wb, marketplace, config,
                    sheet_name=f'{marketplace} (BLR)',
                    headers_override=None,
                )
                cls._append_marketplace_template_sheet(
                    wb, marketplace, config,
                    sheet_name=f'{marketplace} (AHD)',
                    headers_override=extra,
                )
            else:
                cls._append_marketplace_template_sheet(
                    wb, marketplace, config,
                    sheet_name=marketplace,
                    headers_override=None,
                )

        wb.save(save_path)

    @staticmethod
    def _append_marketplace_template_sheet(wb, marketplace: str,
                                             config: dict,
                                             sheet_name: str,
                                             headers_override=None) -> None:
        """
        v2.1.3: Append one marketplace's template sheet to ``wb``.

        Refactored from the pre-v2.1.3 ``_write_template_workbook``
        method — same colour-coded header logic and legend rows, but
        operates on an existing workbook instead of creating a new one.
        Lets the master-template generator iterate over marketplaces
        without each call instantiating a fresh ``Workbook()``.

        Args:
            wb:                Target ``openpyxl.Workbook``.
            marketplace:       Marketplace dict key (e.g. ``'Blink'``).
                               Used in legend / instruction text.
            config:            Marketplace's config dict.
            sheet_name:        Tab name for the new sheet. Must be unique
                               within ``wb`` (caller ensures by handling
                               BlinkMP's AHD/BLR split).
            headers_override:  Optional alternate headers list (used for
                               BlinkMP's 'AHD' variant which uses
                               ``template_headers_extra`` instead of
                               ``template_headers``). When None, the
                               canonical ``template_headers`` is used.
        """
        ws = wb.create_sheet(sheet_name)

        # ── Determine required vs validation vs unused cols ─────────────
        # v1.5.5: column-config values can be either a scalar string
        # (most marketplaces) or a list of accepted aliases (Myntra's
        # ``po_col = ['PO', 'PO Number']``). For the template we need
        # ALL possible rendered header names to color-code correctly:
        #   * each alias itself (in case someone customises the
        #     template_headers to use one of the literal names)
        #   * the slash-joined combined label (e.g. 'PO/PO Number')
        #     which is what we put in template_headers so the user
        #     sees both options and knows to rename before uploading.
        item_resolution = config.get('item_resolution', 'from_column')

        def _normalize(val):
            """Turn a config column value into the set of names that,
            if seen as a template header, should be coloured as this
            column. Handles both scalar strings and list aliases."""
            if val is None:
                return set()
            if isinstance(val, list):
                names = set(val)
                # Also include the 'A/B' slash label so templates that
                # combine aliases as a single column header (the
                # recommended pattern for user-facing templates) get
                # coloured correctly.
                names.add('/'.join(val))
                return names
            return {val}

        required_cols: set = set()
        required_cols |= _normalize(config.get('po_col'))
        required_cols |= _normalize(config.get('loc_col'))
        required_cols |= _normalize(config.get('qty_col'))

        if item_resolution == 'from_ean':
            # EAN is the required identifier when Item No is resolved
            # from it — promote from GREEN to BLUE.
            required_cols |= _normalize(config.get('ean_col'))
        else:  # 'from_column'
            required_cols |= _normalize(config.get('item_col'))

        validation_cols: set = set()
        ean_names = _normalize(config.get('ean_col'))
        if ean_names and not (ean_names & required_cols):
            validation_cols |= ean_names
        validation_cols |= _normalize(config.get('fob_col'))

        # ── Build column list ───────────────────────────────────────────
        # v2.1.3: when headers_override is supplied (BlinkMP AHD case),
        # use it instead of config['template_headers']. Both lists are
        # equally valid template forms; the override mechanism just
        # lets us render multiple variants from one config entry.
        def _primary_label(val):
            if val is None:
                return None
            if isinstance(val, list):
                return '/'.join(val)
            return val

        if headers_override is not None:
            headers = list(headers_override)
        else:
            headers = config.get('template_headers')
            if not headers:
                headers = [
                    _primary_label(config.get('po_col')),
                    _primary_label(config.get('loc_col')),
                ]
                if (item_resolution == 'from_column'
                        and config.get('item_col')):
                    headers.append(_primary_label(config.get('item_col')))
                headers.append(_primary_label(config.get('qty_col')))
                ean_label = _primary_label(config.get('ean_col'))
                if ean_label and ean_label not in headers:
                    headers.append(ean_label)
                fob_label = _primary_label(config.get('fob_col'))
                if fob_label and fob_label not in headers:
                    headers.append(fob_label)
                # Filter out None in case any of the above weren't
                # configured — shouldn't happen for a valid marketplace
                # config but belt-and-braces.
                headers = [h for h in headers if h]

        # ── Styles per role ─────────────────────────────────────────────
        required_fill = PatternFill('solid', fgColor='1A237E')   # blue
        validation_fill = PatternFill('solid', fgColor='1B5E20')  # green
        unused_fill = PatternFill('solid', fgColor='9E9E9E')       # grey

        hdr_font_white = Font(bold=True, color='FFFFFF',
                               name='Aptos Display', size=11)
        hdr_font_dim = Font(bold=True, color='EEEEEE',
                             name='Aptos Display', size=11, italic=True)

        # ── Header row (colour-coded) ───────────────────────────────────
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            if h in required_cols:
                cell.fill = required_fill
                cell.font = hdr_font_white
            elif h in validation_cols:
                cell.fill = validation_fill
                cell.font = hdr_font_white
            else:
                cell.fill = unused_fill
                cell.font = hdr_font_dim
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(c)].width = max(
                len(h) + 4, 12,
            )

        # ── Legend rows (3-5) ───────────────────────────────────────────
        header_set = set(headers)
        required_labels = sorted(required_cols & header_set)
        validation_labels = sorted(validation_cols & header_set)

        legend_row = 3
        legend_items = [
            ('1A237E', 'FFFFFF', 'REQUIRED',
             f'Script fails without these — fill them in: '
             f'{", ".join(required_labels)}'),
            ('1B5E20', 'FFFFFF', 'VALIDATION',
             f'Used for price check & master lookup: '
             f'{", ".join(validation_labels) or "(none)"}'),
            ('9E9E9E', 'FFFFFF', 'NOT READ',
             'Optional — kept only to match marketplace file format; '
             'can stay blank'),
        ]
        for fg, fc, label, desc in legend_items:
            tag = ws.cell(row=legend_row, column=1, value=label)
            tag.fill = PatternFill('solid', fgColor=fg)
            tag.font = Font(bold=True, color=fc,
                             name='Aptos Display', size=10)
            tag.alignment = Alignment(horizontal='center')

            desc_cell = ws.cell(row=legend_row, column=2, value=desc)
            desc_cell.font = Font(name='Aptos Display', size=10,
                                    color='333333', italic=True)
            ws.merge_cells(start_row=legend_row, start_column=2,
                            end_row=legend_row,
                            end_column=min(8, len(headers)))
            legend_row += 1

        # v1.5.5: extra instruction row for marketplaces that use
        # slash-joined column labels (e.g. 'PO/PO Number'). Tells the
        # user to rename the header to one of the options before
        # uploading, so the engine's alias matcher can find it.
        slash_labels = [h for h in headers if '/' in h]
        if slash_labels:
            msg = (
                "⚠ Rename '" + "', '".join(slash_labels) + "' "
                "column(s) to ONE of the listed options (e.g. 'PO' OR "
                "'PO Number') before uploading. The script accepts "
                "either name but the header must be a single choice."
            )
            rename_cell = ws.cell(
                row=legend_row, column=1, value=msg,
            )
            rename_cell.font = Font(
                name='Aptos Display', size=10,
                color='C62828', italic=True, bold=True,
            )
            ws.merge_cells(
                start_row=legend_row, start_column=1,
                end_row=legend_row,
                end_column=min(8, len(headers)),
            )
            legend_row += 1

        # ── Final orange instruction row ────────────────────────────────
        ws.cell(
            row=legend_row + 1, column=1,
            value=(f'← {sheet_name} PO template. Fill data rows below '
                   f'the header. Only the BLUE & GREEN columns are read '
                   f'by the script.'),
        ).font = Font(name='Aptos Display', size=10,
                      color='FF6600', italic=True)

        ws.freeze_panes = 'A2'

    # ── Run the app ────────────────────────────────────────────────────

    def run(self) -> None:
        """Start the Tkinter main loop. Blocks until the window closes."""
        self.root.mainloop()