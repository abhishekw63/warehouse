"""
grn.views
=========

The **GRN** department — upload marketplace GRN PDFs (Blinkit today), parse them
with :mod:`grn.services.grn_parser`, review the per-PO summary + line items, and
download a formatted Excel. Thin views over the fat parser service.
"""
from __future__ import annotations

import uuid
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_POST

from .services import grn_parser as gp

_MEDIA = Path(settings.MEDIA_ROOT) if getattr(settings, 'MEDIA_ROOT', '') else Path('media')
_UPLOADS = _MEDIA / 'grn_uploads'


def _token_dir(token: str) -> Path:
    base = _UPLOADS.resolve()
    d = (_UPLOADS / token).resolve()
    if d != base and base not in d.parents:
        raise Http404()
    return d


@login_required
def index(request):
    """Landing page — upload one or more Blinkit GRN PDFs."""
    return render(request, 'grn/index.html', {'marketplaces': list(gp.REGISTRY)})


@login_required
@require_POST
def upload(request):
    """Stash the uploaded PDFs under a token → redirect to the result page."""
    files = request.FILES.getlist('pdfs')
    files = [f for f in files if f and f.name.lower().endswith('.pdf')]
    if not files:
        messages.error(request, 'Choose at least one GRN PDF.')
        return redirect('grn_index')
    token = uuid.uuid4().hex[:12]
    d = _UPLOADS / token
    d.mkdir(parents=True, exist_ok=True)
    for f in files:
        safe = Path(f.name).name
        with open(d / safe, 'wb') as out:
            for chunk in f.chunks():
                out.write(chunk)
    return redirect('grn_result', token=token)


def _parse_token(token: str) -> dict:
    d = _token_dir(token)
    if not d.exists():
        raise Http404('Upload not found or expired.')
    paths = sorted(str(p) for p in d.glob('*.pdf'))
    if not paths:
        raise Http404('No PDFs in this upload.')
    return gp.parse_many(paths)


@login_required
def result(request, token):
    """Parse the stashed PDFs and render the summary + line items."""
    data = _parse_token(token)
    return render(request, 'grn/result.html', {'token': token, 'd': data})


@login_required
def export(request, token):
    """Download the parsed GRN as a formatted .xlsx (PO Summary + Line Items)."""
    import datetime as _dt
    import io as _io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    data = _parse_token(token)
    items = data.get('items', [])
    icols = data.get('item_columns', [])
    scols = data.get('summary_columns', [])
    stat_col = data.get('status_column', '')
    scolors = data.get('status_colors', {})

    wb = Workbook()
    hf = Font(bold=True, color='FFFFFF')
    navy = PatternFill('solid', fgColor='1A237E')

    def _sheet(ws, cols, rows):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(1, c, h); cell.font = hf; cell.fill = navy
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        for r, row in enumerate(rows, 2):
            for c, k in enumerate(cols, 1):
                ws.cell(r, c, row.get(k))
            if stat_col and stat_col in cols:
                colr = scolors.get(row.get(stat_col))
                if colr:
                    ws.cell(r, cols.index(stat_col) + 1).fill = PatternFill(
                        'solid', fgColor=colr)
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            L = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[L].width = min(w + 2, 52)

    ws1 = wb.active; ws1.title = 'PO Summary'
    _sheet(ws1, scols, data.get('po_summaries', []))
    ws2 = wb.create_sheet('GRN Line Items')
    _sheet(ws2, icols, items)

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    stamp = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    mkt = (data.get('marketplace') or 'GRN')
    fname = f"{mkt}_GRN_{data['totals']['pos']}PO_{stamp}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
