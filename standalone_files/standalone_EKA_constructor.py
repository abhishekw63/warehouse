"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                  RENEE PO PROCESSOR — EKA Script (v1.5.8)                    ║
║                  Simple Tkinter GUI Desktop Application                       ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  Author  : Agami AI / Vishal                                                 ║
║  Version : 1.5.8 (PWP_EAN_MAP single-EAN lookup layer)                       ║
║  Stack   : Python 3.13, Tkinter, pandas, openpyxl                           ║
╚═══════════════════════════════════════════════════════════════════════════════╝

═══════════════════════════════════════════════════════════════════════════════
  CHANGELOG
═══════════════════════════════════════════════════════════════════════════════

  v1.5.8 — PWP single-EAN mapping layer (this version)
  ────────────────────────────────────────────────────────────────
    ┌───────────────────────────┬───────────────────────────────────────┐
    │  CHANGE                   │  IMPACT                                │
    ├───────────────────────────┼───────────────────────────────────────┤
    │  New PWP_EAN_MAP dict +   │  PWP rows like 'RENEE BOLLYWOOD FILTER│
    │  whitespace normalizer    │  BLURRING SETTING SPRAY 5ML (NFS)' now│
    │  (_norm_pwp_name)         │  resolve to a real Item No (201432)   │
    │                           │  instead of dumping the long name into│
    │                           │  the Item No column. Adding a new     │
    │                           │  single-EAN PWP entry = one line in   │
    │                           │  PWP_EAN_MAP. Non-breaking spaces     │
    │                           │  (\\xa0) in source files are handled   │
    │                           │  automatically.                        │
    ├───────────────────────────┼───────────────────────────────────────┤
    │  Actionable warning on    │  When a PWP name doesn't match any    │
    │  unknown PWP names        │  list, the log + Processing Log sheet │
    │                           │  now shows the EXACT line to paste    │
    │                           │  into PWP_EAN_MAP: 'NAME': '<EAN>'.   │
    │                           │  No more hunting for the right spelling│
    │                           │  when whitespace is involved.          │
    ├───────────────────────────┼───────────────────────────────────────┤
    │  Validation includes      │  PWP names that ARE in PWP_EAN_MAP no │
    │  PWP_EAN_MAP keys in the  │  longer get flagged as Unknown during │
    │  'known' set              │  pre-processing validation. Cleaner    │
    │                           │  audit output.                         │
    └───────────────────────────┴───────────────────────────────────────┘

  v1.5.7 — Multicolor paper bag Non-Stock entries
  ────────────────────────────────────────────────────────────────
    Added 'Renee Multicolor Bag (Big)' → Item 300074 (EAN 8904473105984)
    and '... (Small)' → Item 300075 (EAN 8904473105991) to
    NON_STOCK_EAN_MAP.

  v1.3 — Simplified GUI
  v1.2 — Auto-fill + D365 TO Export + Tester TO Uniqueness
  v1.1 — Special Order Engine + Full Documentation
  v1.0 — Initial release (23 PO files tested)

═══════════════════════════════════════════════════════════════════════════════
  AUTO-LOAD CONVENTION
═══════════════════════════════════════════════════════════════════════════════

  At startup the GUI looks for two files in a folder named
  ``Calculation Data/`` next to this script:

      Calculation Data/
          Items_March.xlsx   ← product master (auto-loaded)
          EKA_DATA.xlsx      ← location registry (auto-loaded)

  When found, the relevant picker shows ``✓ <name> (auto-loaded)``
  and the user just adds PO files (Standalone) or selects the
  Special Order file (Special Order) and clicks Generate.

Requirements:
    pip install pandas openpyxl

Run:
    python renee_po_processor.py
"""

# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

import os
import re
import shutil
import sys
import threading
import time
import warnings
import zipfile
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# v1.5.2: openpyxl emits a UserWarning when reading XLSX files that
# contain Conditional Formatting extensions (color rules, data bars,
# etc.). The data still loads fine — only the conditional-formatting
# metadata is dropped. The warning otherwise looks alarming on the
# terminal and made one user think the app had crashed. Suppress it.
warnings.filterwarnings(
    'ignore', category=UserWarning, module='openpyxl',
)

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

APP_VERSION = "1.5.8"
APP_TITLE = f"RENEE PO Processor v{APP_VERSION}"

# Bundled-file location convention. Looked up RELATIVE to the script
# directory so the .exe / .py both resolve correctly.
BUNDLED_DATA_FOLDER = "Calculation_Data_EKA"
BUNDLED_MASTER_NAME = "Items_March.xlsx"
BUNDLED_EKA_NAME = "EKA_DATA.xlsx"


def get_script_dir() -> Path:
    """
    Return the directory containing this script.

    Resolves correctly for both Python source runs and PyInstaller-frozen
    executables (where ``__file__`` may not exist).
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def get_bundled_master_path() -> Optional[Path]:
    """Return the path to the bundled Items Master, or None if absent."""
    p = get_script_dir() / BUNDLED_DATA_FOLDER / BUNDLED_MASTER_NAME
    return p if p.exists() else None


def get_bundled_eka_path() -> Optional[Path]:
    """Return the path to the bundled EKA_DATA, or None if absent."""
    p = get_script_dir() / BUNDLED_DATA_FOLDER / BUNDLED_EKA_NAME
    return p if p.exists() else None


def get_bundled_folder(create: bool = False) -> Path:
    """Return the bundled-data folder path; create on demand."""
    p = get_script_dir() / BUNDLED_DATA_FOLDER
    if create:
        p.mkdir(parents=True, exist_ok=True)
    return p



# ═══════════════════════════════════════════════════════════════════════════════
#  CORE DATA STRUCTURES
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class OutputRow:
    """
    Single row in the Final Data output.

    Fields:
        to             : Transfer Order / Sales Order number
                          (auto-filled from EKA_DATA in standalone mode)
        item_no        : Item No from Items_March, or product name if unresolved
        qty            : Quantity (order or tester)
        unit_price     : Calculated cost price (PO) or ₹0.54 (testers)
        transfer_to    : Transfer-to Code / Location Code
                          (auto-filled from EKA_DATA)
        posting_group  : Gen. Bus. Posting Group
                          (auto-filled from EKA_DATA)
        source         : 'PO', 'TESTER', 'PWP', 'GWP', 'NON_STOCK'
        ean            : Original EAN/barcode
        product_name   : Product description (for reference)
        lookup_status  : 'OK', 'NOT_FOUND', 'UNKNOWN', 'NO_MAP'
    """
    to: str = ''
    item_no: Any = ''
    qty: int = 0
    unit_price: float = 0.0
    transfer_to: str = ''
    posting_group: str = ''
    source: str = ''
    ean: str = ''
    product_name: str = ''
    lookup_status: str = ''


@dataclass
class LocationResult:
    """
    Complete processing result for one PO file (one location).

    Contains separate lists for each order type, plus unmatched EANs
    and processing logs for the GUI log panel.
    """
    filename: str
    regular_orders: List[OutputRow] = field(default_factory=list)   # PO orders
    tester_orders: List[OutputRow] = field(default_factory=list)    # Testers (₹0.54)
    pwp_orders: List[OutputRow] = field(default_factory=list)       # PWP (₹0.54)
    gwp_orders: List[OutputRow] = field(default_factory=list)       # GWP (₹0.54)
    nonstock_orders: List[OutputRow] = field(default_factory=list)  # Non-stock (₹0.54)
    unmatched: List[Dict] = field(default_factory=list)             # EANs not in master
    logs: List[tuple] = field(default_factory=list)                 # (level, message)


# ═══════════════════════════════════════════════════════════════════════════════
#  PO ENGINE — Standalone PO processing
# ═══════════════════════════════════════════════════════════════════════════════
#
# Loads Items_March master, validates PO file structure, processes
# PO/PWP/GWP/Non-Stock sheets, and applies GST-aware cost calculations.
#
# Cost formula:
#     Landing Cost = MRP × 60%
#     Cost Price   = Landing Cost ÷ (1 + GST rate)
#
# Testers / PWP / GWP / Non-Stock items: flat ₹0.54.
#
# PWP rules:
#     'Stay With Me - Mini' → IGNORED
#     'Crème Mini'          → IGNORED
#     'Perfume'             → SPLIT equally across 4 perfume EANs
#                              (remainder goes to first EANs)
#     v1.5.8: PWP_EAN_MAP   → single-EAN lookup (e.g. setting spray)

class POEngine:
    """Core PO processing logic — no GUI dependency."""

    # ┌────────────────────────────────────────────────────────────────────┐
    # │ NON-STOCK NAME → EAN/CODE                                          │
    # │ ⚠ Names MUST match EXACTLY what's in the Excel Non Stock sheet.    │
    # │ ⚠ EAN/codes MUST exist in Items_March (GTIN or No. column).        │
    # │   Add new non-stock items HERE.                                    │
    # └────────────────────────────────────────────────────────────────────┘
    NON_STOCK_EAN_MAP = {
        'Cotton Rolls':      'OPM-RSK-CR500-RE',        # → Item 400039
        'Mirrors':           'OPM-RSK-PU-LMS-RE',       # → Item 400037
        'Carry Bag (Small)': '8904473106011',            # → Item 300077
        'Carry Bag (Big)':   '8904473106004',            # → Item 300076
        # v1.5.7: paper bags. Distinct from the plastic 'Carry Bag'
        # entries above. Two SKUs maintained separately so stores can
        # request either size on the Non-Stock sheet without ambiguity.
        'Renee Multicolor Bag (Big)':   '8904473105984',  # → Item 300074
        'Renee Multicolor Bag (Small)': '8904473105991',  # → Item 300075
        'Cleansers':         '8906121643572',            # → Item 200101
        'Calculator':        'OPM-CAL-SK-RE',            # → Item 400111
        'Blotters':          'RCPL_PB',                  # → Item 400060
        'Swabs':             'OPM-NMS-OT-P100-SWB',     # → Item 400057
        'Bill Roll':         'OPM-TPR-VL-TSC-100-150',  # → Item 400088
        'Renee Notebook':    'RCPL_NOTEPAD',             # → Item 400059
        'Pen':               'RCPL_PEN',                 # → Item 400061
    }

    # ┌────────────────────────────────────────────────────────────────────┐
    # │ PERFUME PWP — 4 EANs to split demand equally                       │
    # │ Example: qty=10 → 3+3+2+2 (first EANs absorb the remainder)        │
    # └────────────────────────────────────────────────────────────────────┘
    PERFUME_EANS = [
        '8906121642674',  # RENEE BLOOM 8ML NFS
        '8906121647495',  # RENEE FLIRT 8ML NFS
        '8906121647501',  # RENEE MADAME 8ML NFS
        '8906121645743',  # RENEE RED NOIR 8ML NFS
    ]

    # ┌────────────────────────────────────────────────────────────────────┐
    # │ PWP NAME → EAN  (single-EAN mappings) — v1.5.8                     │
    # │                                                                    │
    # │ Used for PWP items that aren't on the ignore list AND aren't the   │
    # │ 4-way Perfume split — i.e. one row in the PWP sheet → one EAN in   │
    # │ Items_March, at flat ₹0.54.                                        │
    # │                                                                    │
    # │ Lookup is whitespace-normalized: \xa0 (non-breaking space) and     │
    # │ multiple spaces are collapsed to single regular spaces before      │
    # │ matching. Keys here should use REGULAR spaces — the normalizer     │
    # │ handles the input drift automatically.                             │
    # │                                                                    │
    # │ ⚠ EAN MUST exist in Items_March (GTIN column).                     │
    # │ ⚠ When the log says "PWP: Unknown '<name>' — add to PWP_EAN_MAP    │
    # │   if this is a known product", paste the suggested line HERE and  │
    # │   fill in the EAN.                                                 │
    # └────────────────────────────────────────────────────────────────────┘
    PWP_EAN_MAP = {
        # v1.5.8: setting spray. Master has the FOR SALE SKU at this EAN;
        # PWP files ship the NFS-marked name. Both refer to the same
        # physical product; we route the NFS name to the FOR SALE Item No.
        'RENEE BOLLYWOOD FILTER BLURRING SETTING SPRAY 5ML (NFS)':
            '8904473106318',  # → Item 201432 (G-18-S, MRP 99)
    }

    PWP_IGNORE = {'Stay With Me - Mini', 'Crème Mini'}

    def __init__(self):
        self.master: Dict[str, Dict] = {}

    def load_master(self, path: str) -> int:
        """
        Load Items_March.xlsx and build lookup dictionary.

        Indexed by BOTH GTIN (EAN) and No. (item code) so PO items
        can be looked up by EAN, and non-stock items by internal code.

        Returns: number of rows loaded.
        """
        df = pd.read_excel(path, header=0)
        df['GTIN_str'] = df['GTIN'].astype(str).str.strip()
        self.master = {}

        gtin_idx = df.columns.get_loc('GTIN_str')
        desc_idx = df.columns.get_loc('Description') if 'Description' in df.columns else None
        no_idx = df.columns.get_loc('No.')
        mrp_idx = df.columns.get_loc('Mrp')
        gst_idx = df.columns.get_loc('GST Group Code') if 'GST Group Code' in df.columns else None

        for r_vals in df.values:
            desc = str(r_vals[desc_idx]) if desc_idx is not None and pd.notna(r_vals[desc_idx]) else ''
            gst_val = str(r_vals[gst_idx]) if gst_idx is not None and pd.notna(r_vals[gst_idx]) else ''

            # Primary index: by GTIN (EAN barcode)
            self.master[r_vals[gtin_idx]] = {
                'item_no': r_vals[no_idx],
                'mrp': r_vals[mrp_idx],
                'gst_code': gst_val,
                'description': desc,
            }

            # Secondary index: by No. (item code) — for non-stock
            item_code = str(r_vals[no_idx]).strip()
            if item_code not in self.master:
                self.master[item_code] = {
                    'item_no': r_vals[no_idx],
                    'mrp': r_vals[mrp_idx],
                    'gst_code': gst_val,
                    'description': desc,
                }

        return len(df)

    @staticmethod
    def calc_cost_price(mrp, gst_code: str) -> Optional[float]:
        """
        Calculate unit price for regular PO orders.

        Formula:
            Landing = MRP × 0.60
            Cost    = Landing ÷ (1 + GST)
        """
        if mrp is None or pd.isna(mrp):
            return None

        landing = float(mrp) * 0.60
        gst = str(gst_code).strip().upper()

        if gst in ('0-G', 'G-0', 'G-0-S', '0', '') or gst == 'NAN':
            return landing
        if gst in ('G-3', 'G-3-S'):
            return landing / 1.03
        if '5' in gst and '18' not in gst and '12' not in gst:
            return landing / 1.05
        if '12' in gst:
            return landing / 1.12
        if '18' in gst:
            return landing / 1.18
        return landing / 1.18  # unknown → fallback to 18%

    def _detect_po_columns(self, ws, logs: Optional[List] = None) -> Dict[str, int]:
        """
        Scan PO sheet's header row for the 3 required columns.

        v1.5.5: matches column names case-insensitively. Real-world
        files have ``Order Qty``, ``order Qty``, ``Order qty``,
        ``order qty``, ``ORDER QTY`` — they all mean the same thing.
        Same for ``EAN`` and the tester column (``Tester Qty``,
        ``Tester``, ``tester``, etc.).

        Standard names: 'EAN', 'Order Qty', 'Tester Qty'.
        Each fallback emits an 'alert' so the user can fix the source.
        """
        hmap: Dict[str, int] = {}
        normalized: Dict[str, Tuple[str, int]] = {}

        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            val = str(cell.value or '').strip()
            idx = cell.column - 1
            if val:
                key = val.lower()
                normalized[key] = (val, idx)

        # ── EAN ──
        if 'ean' in normalized:
            original, idx = normalized['ean']
            hmap['ean'] = idx
            if original != 'EAN' and logs is not None:
                logs.append(('alert',
                    f"Auto-fixed: '{original}' → 'EAN'. "
                    f"Please rename column to 'EAN'."))

        # ── Order Qty ──
        if 'order qty' in normalized:
            original, idx = normalized['order qty']
            hmap['order_qty'] = idx
            if original != 'Order Qty' and logs is not None:
                logs.append(('alert',
                    f"Auto-fixed: '{original}' → 'Order Qty'. "
                    f"Please rename column to 'Order Qty'."))

        # ── Tester Qty ── (with 'Tester' fallback)
        if 'tester qty' in normalized:
            original, idx = normalized['tester qty']
            hmap['tester_qty'] = idx
            if original != 'Tester Qty' and logs is not None:
                logs.append(('alert',
                    f"Auto-fixed: '{original}' → 'Tester Qty'. "
                    f"Please rename column to 'Tester Qty'."))
        elif 'tester' in normalized:
            original, idx = normalized['tester']
            hmap['tester_qty'] = idx
            if logs is not None:
                logs.append(('alert',
                    f"Auto-fixed: '{original}' → 'Tester Qty'. "
                    f"Please rename column to 'Tester Qty'."))

        return hmap

    def _safe_int(self, val) -> int:
        """Convert cell value to int. Returns 0 for None/empty/errors."""
        try:
            if val is None or str(val).strip() in ('', '#N/A', 'None'):
                return 0
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    def _ean_str(self, raw) -> str:
        """Convert raw EAN value to clean string. Handles float→int."""
        if raw is None:
            return ''
        return str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()

    @staticmethod
    def _row_get(row, idx: int):
        """
        Defensive row-cell read.

        v1.5.5: rows in PWP/GWP/Non-Stock sheets vary in width across
        different PO file templates. Some files have the standard
        4-column PWP (Sr.No, Name, Avail.Qty, Req.Qty) but others
        (Pune EBO) ship with only 3 columns (Sr.No, Name, Store Name).
        Returns None when the requested index is out of range so the
        caller can skip that row gracefully instead of crashing on
        ``IndexError: tuple index out of range``.
        """
        if idx < len(row):
            return row[idx].value
        return None

    @staticmethod
    def _find_qty_col(ws, qty_keywords: List[str], default: int) -> int:
        """
        Locate the REQUIRED quantity column in a PWP/GWP/Non-Stock
        sheet by scanning the header row.

        v1.5.5 (revised): two-pass priority match. The first pass
        looks for an exact ``req.qty``/``req qty``/``required qty``
        header — the only column whose value we should ever read.
        Only if that fails does the second pass fall back to a plain
        ``qty`` header, AFTER explicitly excluding any column whose
        name contains ``avail`` or ``available``.
        """
        headers = []
        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            val = str(cell.value or '').strip()
            headers.append((val.lower(), cell.column - 1))

        # Pass 1: required-qty headers (the column we WANT)
        REQUIRED_NAMES = {'req.qty', 'req qty', 'required qty', 'required.qty'}
        for val, idx in headers:
            if val in REQUIRED_NAMES:
                return idx

        # Pass 2: lone 'qty' header — accepted only when no available-
        # qty column shadows it.
        for val, idx in headers:
            if val and 'avail' not in val and (
                val == 'qty' or val == 'quantity'
            ):
                return idx

        return default

    @staticmethod
    def _norm_pwp_name(raw: str) -> str:
        """
        Normalize a PWP product name for exact-match lookup. (v1.5.8)

        Real-world PWP sheets ship with whitespace inconsistencies:
        non-breaking spaces (\\xa0), tabs, double spaces, trailing
        spaces. Comparing the raw string against our maps breaks on
        every new file the warehouse sends. This helper collapses
        every whitespace run (incl. \\xa0) to a single regular space
        and strips edges so the map keys can stay clean and readable.

        Example::

            'RENEE\\xa0BOLLYWOOD  FILTER\\xa0SPRAY  '
                      → 'RENEE BOLLYWOOD FILTER SPRAY'
        """
        if raw is None:
            return ''
        # \xa0 (non-breaking space) is NOT matched by \s in Python's
        # default mode for str.split — replace explicitly first, then
        # let split() handle the rest.
        cleaned = raw.replace('\xa0', ' ')
        return ' '.join(cleaned.split())

    def process_po_sheet(self, ws, col_map: Dict, logs: List
                          ) -> Tuple[List[OutputRow], List[OutputRow], List[Dict]]:
        """Process the PO sheet → (regular_orders, tester_orders, unmatched)."""
        regular, testers, unmatched = [], [], []

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            ean_raw = row[col_map['ean']].value
            if ean_raw is None:
                continue
            if any(c.value and str(c.value).upper() == 'TOTAL' for c in row):
                continue

            row_num = row[0].row
            ean = self._ean_str(ean_raw)
            order_qty = self._safe_int(row[col_map['order_qty']].value) if 'order_qty' in col_map else 0
            tester_qty = self._safe_int(row[col_map['tester_qty']].value) if 'tester_qty' in col_map else 0

            info = self.master.get(ean) or self.master.get(ean.lstrip('0'))

            if info:
                item_no = info['item_no']
                gst_code = info['gst_code']
                known_gst = {'0-G', 'G-0', 'G-0-S', 'G-3', 'G-3-S',
                             'G-5', 'G-5-S', 'G-12', 'G-12-S',
                             'G-18', 'G-18-S', ''}
                gst_upper = str(gst_code).strip().upper()
                if gst_upper not in known_gst and gst_upper != 'NAN':
                    logs.append(('warn',
                        f"PO row {row_num}: Unknown GST '{gst_code}' "
                        f"for Item {item_no} — defaulting to 18%"))

                cost = self.calc_cost_price(info['mrp'], gst_code)
                status = 'OK'
            else:
                item_no = f'?EAN:{ean}'
                cost = None
                status = 'NOT_FOUND'
                unmatched.append({
                    'ean': ean, 'product_name': '',
                    'order_qty': order_qty, 'tester_qty': tester_qty,
                })
                logs.append(('warn', f"PO row {row_num}: EAN {ean} not found in master"))

            if order_qty > 0:
                regular.append(OutputRow(
                    item_no=item_no, qty=order_qty,
                    unit_price=cost or 0, source='PO',
                    ean=ean, lookup_status=status,
                ))

            if tester_qty > 0:
                testers.append(OutputRow(
                    item_no=item_no, qty=tester_qty,
                    unit_price=0.54, source='TESTER',
                    ean=ean, lookup_status=status,
                ))

        return regular, testers, unmatched

    def process_pwp(self, ws, logs: List) -> List[OutputRow]:
        """
        Process PWP sheet.

        Resolution order for each non-Total row with qty > 0:
            1. Name in PWP_IGNORE  → skip (e.g. Stay With Me - Mini)
            2. 'perfume' substring → 4-way EAN split (PERFUME_EANS)
            3. Name in PWP_EAN_MAP → single-EAN lookup at ₹0.54
                                      (v1.5.8: new single-EAN layer)
            4. Otherwise           → fallback: emit row with name as
                                      item_no, status 'UNKNOWN' for
                                      manual review in D365 import.
                                      Logs an actionable warning with
                                      the EXACT line to paste into
                                      PWP_EAN_MAP.

        All matched names go through ``_norm_pwp_name`` first so
        whitespace drift in source files (non-breaking spaces, double
        spaces) doesn't kill the lookup.

        v1.5.5: detects the qty column by header name.
        v1.5.8: adds P
        layer + whitespace normalization +
                actionable warning on unknown names.
        """
        rows = []

        # Pre-normalize ignore list + map keys so lookups are O(1)
        # against the normalized incoming name.
        ignore_norm = {self._norm_pwp_name(n) for n in self.PWP_IGNORE}
        pwp_map_norm = {
            self._norm_pwp_name(k): v
            for k, v in self.PWP_EAN_MAP.items()
        }

        # v1.5.5: find qty column by header name. Default to legacy
        # index 3 if no header matched.
        qty_idx = self._find_qty_col(ws, ['req.qty', 'req qty', 'required qty', 'qty'], 3)

        for row in ws.iter_rows(min_row=2, max_row=20, values_only=False):
            a = self._row_get(row, 0)
            b = self._row_get(row, 1)
            d = self._row_get(row, qty_idx)
            if a is None or str(a).strip().upper() == 'TOTAL':
                continue

            qty = self._safe_int(d)
            name_raw = str(b or '').strip()
            name_norm = self._norm_pwp_name(name_raw)

            # ── 1. Ignore list ──
            if name_norm in ignore_norm:
                if qty > 0:
                    logs.append(('info', f"PWP: '{name_raw}' qty={qty} → skipped (ignore list)"))
                continue
            if qty <= 0:
                continue

            # ── 2. Perfume 4-way split (substring match on 'perfume') ──
            if 'perfume' in name_norm.lower():
                base_qty = qty // 4
                remainder = qty % 4
                logs.append(('info',
                    f"PWP: Perfume qty={qty} → split 4 EANs "
                    f"({base_qty}+{base_qty}+{base_qty}+{base_qty}, "
                    f"remainder={remainder})"))

                for i, ean in enumerate(self.PERFUME_EANS):
                    eq = base_qty + (1 if i < remainder else 0)
                    if eq <= 0:
                        continue

                    info = self.master.get(ean)
                    if info:
                        item_no = info['item_no']
                    else:
                        item_no = f'?EAN:{ean}'
                        logs.append(('warn', f"PWP: Perfume EAN {ean} not in master"))

                    rows.append(OutputRow(
                        item_no=item_no, qty=eq, unit_price=0.54,
                        source='PWP', ean=ean,
                        product_name=f'Perfume ({ean})',
                        lookup_status='OK' if info else 'NOT_FOUND',
                    ))
                continue

            # ── 3. Single-EAN PWP map (v1.5.8) ──
            if name_norm in pwp_map_norm:
                ean = pwp_map_norm[name_norm]
                info = self.master.get(ean)
                if info:
                    item_no = info['item_no']
                    status = 'OK'
                else:
                    # Map says this EAN exists, but Items_March doesn't
                    # have it. Surfaces a clear NOT_FOUND on the Final
                    # Data sheet so the operator sees the gap before
                    # importing to D365.
                    item_no = f'?EAN:{ean}'
                    status = 'NOT_FOUND'
                    logs.append(('warn',
                        f"PWP: '{name_raw}' mapped to EAN {ean} but "
                        f"EAN not in Items_March — outputting placeholder"))

                rows.append(OutputRow(
                    item_no=item_no, qty=qty, unit_price=0.54,
                    source='PWP', ean=ean, product_name=name_raw,
                    lookup_status=status,
                ))
                continue

            # ── 4. Fallback — actionable warning + name out as item_no ──
            # v1.5.8: the warning now tells the operator EXACTLY what
            # to do — add the name to PWP_EAN_MAP. The normalized name
            # is shown so they can see whether whitespace was involved.
            log_msg = (
                f"PWP: Unknown product '{name_raw}' qty={qty} → "
                f"outputting name as Item No (status UNKNOWN). "
                f"If this is a known product, add to PWP_EAN_MAP "
                f"as: '{name_norm}': '<EAN>',"
            )
            logs.append(('warn', log_msg))
            rows.append(OutputRow(
                item_no=name_raw, qty=qty, unit_price=0.54,
                source='PWP', product_name=name_raw, lookup_status='UNKNOWN',
            ))

        return rows

    def process_gwp(self, ws, logs: List) -> List[OutputRow]:
        """
        Process GWP sheet (every item has an EAN).

        v1.5.5: detects qty column by header name. Standard layout
        is 5 cols (Sr.No, EAN, Name, Avail.Qty, Req.Qty) with qty
        at index 4, but Pune EBO format uses 4 cols (Sr.No, EAN,
        Name, Qty) with qty at index 3. Header-name detection
        handles both.
        """
        rows = []

        qty_idx = self._find_qty_col(ws, ['req.qty', 'req qty', 'required qty', 'qty'], 4)

        for row in ws.iter_rows(min_row=2, max_row=20, values_only=False):
            a = self._row_get(row, 0)
            ean_raw = self._row_get(row, 1)
            name = self._row_get(row, 2)
            req_raw = self._row_get(row, qty_idx)
            if a is None or str(a).strip().upper() == 'TOTAL':
                continue

            qty = self._safe_int(req_raw)
            if qty > 0:
                ean = self._ean_str(ean_raw)
                name_str = str(name or '')
                info = self.master.get(ean)

                if info:
                    item_no = info['item_no']
                else:
                    item_no = name_str
                    logs.append(('warn',
                        f"GWP: EAN {ean} ({name_str}) not found → using name"))

                rows.append(OutputRow(
                    item_no=item_no, qty=qty, unit_price=0.54,
                    source='GWP', ean=ean, product_name=name_str,
                    lookup_status='OK' if info else 'NOT_FOUND',
                ))

        return rows

    def process_non_stock(self, ws, logs: List) -> List[OutputRow]:
        """
        Process Non Stock sheet (name → EAN/code lookup).

        v1.5.5: defensive row reads to handle short rows.
        """
        rows = []

        for row in ws.iter_rows(min_row=2, max_row=20, values_only=False):
            a = self._row_get(row, 0)
            b = self._row_get(row, 1)
            c = self._row_get(row, 2)
            if a is None or str(a).strip().upper() == 'TOTAL':
                continue

            qty = self._safe_int(c)
            if qty <= 0:
                continue

            name = str(b or '').strip()
            ean = self.NON_STOCK_EAN_MAP.get(name, '')

            if not ean:
                logs.append(('warn',
                    f"Non-Stock: '{name}' qty={qty} → not in map, "
                    f"outputting name directly"))
                rows.append(OutputRow(
                    item_no=name, qty=qty, unit_price=0.54,
                    source='NON_STOCK', ean='', product_name=name,
                    lookup_status='NO_MAP',
                ))
                continue

            info = self.master.get(ean)
            if info:
                item_no = info['item_no']
                status = 'OK'
            else:
                item_no = name
                status = 'NOT_FOUND'
                logs.append(('warn',
                    f"Non-Stock: '{name}' code={ean} → not in master"))

            rows.append(OutputRow(
                item_no=item_no, qty=qty, unit_price=0.54,
                source='NON_STOCK', ean=ean, product_name=name,
                lookup_status=status,
            ))

        return rows

    def validate_file(self, filepath: str) -> List[tuple]:
        """
        Pre-processing validation pass.

        Checks: required sheets, PO columns, PWP names, Non-Stock names,
        and GWP EANs against master.

        Returns list of (level, message) tuples.
        """
        logs = []
        has_blocking = False

        try:
            wb = load_workbook(filepath, data_only=True)
        except Exception as e:
            err_str = str(e)
            if 'must be a sequence' in err_str:
                return [('error',
                    "File has broken external links. "
                    "Open in Excel → Data → Edit Links → Break Link, "
                    "then save and retry."),
                    ('info', f"(Underlying error: {err_str})")]
            return [('error', f"Cannot open: {e}")]

        # 1. Sheet validation
        required = ['PO', 'PWP', 'GWP', 'Non Stock']
        for sheet in required:
            if sheet not in wb.sheetnames:
                logs.append(('error', f"Missing sheet: '{sheet}'"))
                has_blocking = True

        found = [s for s in required if s in wb.sheetnames]
        logs.append(('info', f"Sheets: {', '.join(found)} of {len(required)}"))

        # 2. PO columns
        if 'PO' in wb.sheetnames:
            ws_po = wb['PO']
            col_map = self._detect_po_columns(ws_po, logs)

            if 'ean' not in col_map:
                logs.append(('error', "PO: 'EAN' not found"))
                has_blocking = True
            if 'order_qty' not in col_map:
                logs.append(('error', "PO: 'Order Qty' not found"))
                has_blocking = True
            if 'tester_qty' not in col_map:
                logs.append(('error', "PO: 'Tester Qty' not found"))
                has_blocking = True

            if not has_blocking:
                data_rows = 0
                for row in ws_po.iter_rows(min_row=2, max_row=ws_po.max_row, values_only=False):
                    ean_raw = row[col_map['ean']].value
                    if ean_raw and not any(
                        c.value and str(c.value).upper() == 'TOTAL' for c in row
                    ):
                        data_rows += 1
                logs.append(('info', f"PO: {data_rows} data rows"))

        # 3. PWP names — v1.5.8: include PWP_EAN_MAP keys in the
        # "known" set and use whitespace-normalized comparison so
        # mapped products + non-breaking-space drift don't get
        # flagged as Unknown during validation.
        if 'PWP' in wb.sheetnames:
            known_pwp_norm = {
                self._norm_pwp_name(n)
                for n in (set(self.PWP_IGNORE)
                          | set(self.PWP_EAN_MAP.keys())
                          | {'Perfume', 'perfume'})
            }
            ws_pwp = wb['PWP']
            qty_idx = self._find_qty_col(
                ws_pwp, ['req.qty', 'req qty', 'required qty', 'qty'], 3)
            for row in ws_pwp.iter_rows(min_row=2, max_row=20, values_only=False):
                a = self._row_get(row, 0)
                b = self._row_get(row, 1)
                d = self._row_get(row, qty_idx)
                if a is None or str(a).strip().upper() == 'TOTAL':
                    continue
                name = str(b or '').strip()
                name_norm = self._norm_pwp_name(name)
                qty = self._safe_int(d)
                # 'perfume' is a substring trigger, so any normalized
                # name containing it is implicitly known.
                if (qty > 0 and name_norm not in known_pwp_norm
                        and 'perfume' not in name_norm.lower()):
                    # v1.5.8: actionable warning — shows the exact
                    # normalized name to add to PWP_EAN_MAP.
                    logs.append(('warn',
                        f"PWP: Unknown product '{name}' qty={qty}. "
                        f"Add to PWP_EAN_MAP as: '{name_norm}': '<EAN>',"))

        # 4. Non-Stock names
        if 'Non Stock' in wb.sheetnames:
            for row in wb['Non Stock'].iter_rows(min_row=2, max_row=20, values_only=False):
                a = self._row_get(row, 0)
                b = self._row_get(row, 1)
                c = self._row_get(row, 2)
                if a is None or str(a).strip().upper() == 'TOTAL':
                    continue
                name = str(b or '').strip()
                qty = self._safe_int(c)
                if qty > 0 and name not in self.NON_STOCK_EAN_MAP:
                    logs.append(('warn', f"Non-Stock: '{name}' qty={qty} — not in map"))

        # 5. GWP EANs
        if 'GWP' in wb.sheetnames and self.master:
            ws_gwp = wb['GWP']
            qty_idx = self._find_qty_col(
                ws_gwp, ['req.qty', 'req qty', 'required qty', 'qty'], 4)
            for row in ws_gwp.iter_rows(min_row=2, max_row=20, values_only=False):
                a = self._row_get(row, 0)
                ean_raw = self._row_get(row, 1)
                name = self._row_get(row, 2)
                req_raw = self._row_get(row, qty_idx)
                if a is None or str(a).strip().upper() == 'TOTAL':
                    continue
                qty = self._safe_int(req_raw)
                if qty > 0:
                    ean = self._ean_str(ean_raw)
                    if not self.master.get(ean):
                        logs.append(('warn',
                            f"GWP: EAN {ean} ({name}) qty={qty} — not in master"))

        return logs

    def process_file(self, filepath: str) -> LocationResult:
        """Process one PO file → LocationResult with all order types."""
        wb = load_workbook(filepath, data_only=True)
        res = LocationResult(filename=Path(filepath).name)

        for sheet in ['PO', 'PWP', 'GWP', 'Non Stock']:
            if sheet not in wb.sheetnames:
                res.logs.append(('error', f"Sheet '{sheet}' not found"))

        if 'PO' in wb.sheetnames:
            ws_po = wb['PO']
            col_map = self._detect_po_columns(ws_po, res.logs)

            if 'ean' not in col_map:
                res.logs.append(('error', "PO: 'EAN' not found"))
            if 'order_qty' not in col_map:
                res.logs.append(('error', "PO: 'Order Qty' not found"))
            if 'tester_qty' not in col_map:
                res.logs.append(('error', "PO: 'Tester Qty' not found"))

            if 'ean' in col_map:
                res.regular_orders, res.tester_orders, res.unmatched = \
                    self.process_po_sheet(ws_po, col_map, res.logs)
                res.logs.append(('info',
                    f"PO: {len(res.regular_orders)} orders, "
                    f"{len(res.tester_orders)} testers"))
            else:
                res.logs.append(('error', "PO: Skipping — no EAN column"))

        if 'PWP' in wb.sheetnames:
            res.pwp_orders = self.process_pwp(wb['PWP'], res.logs)

        if 'GWP' in wb.sheetnames:
            res.gwp_orders = self.process_gwp(wb['GWP'], res.logs)

        if 'Non Stock' in wb.sheetnames:
            res.nonstock_orders = self.process_non_stock(wb['Non Stock'], res.logs)

        return res


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER — produces 8+ sheet output workbook (unchanged from v1.5.7)
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelWriter:
    """Writes formatted Excel output with color-coded Source and Status."""

    HEADER_FILL = PatternFill('solid', fgColor='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
    THIN_SIDE = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    SOURCE_COLORS = {
        'PO': ('000000', 'FFFFFF'), 'TESTER': ('7B61FF', 'FFFFFF'),
        'PWP': ('FF6600', 'FFFFFF'), 'GWP': ('00BCD4', 'FFFFFF'),
        'NON_STOCK': ('795548', 'FFFFFF'),
    }
    STATUS_COLORS = {
        'OK': ('00C853', '000000'), 'NOT_FOUND': ('FF5252', 'FFFFFF'),
        'NEEDS_EAN': ('FFB300', '000000'), 'NEEDS_ITEM_NO': ('FFB300', '000000'),
    }

    @classmethod
    def write(cls, results, output_path, eka_locations=None,
              master=None, so_products=None, processing_log=None):
        """Write output Excel with all sheets."""
        wb = Workbook()
        wb.remove(wb.active)

        loc_lookup = {}
        if eka_locations:
            for loc in eka_locations:
                loc_lookup[loc['short_name']] = loc

        if processing_log:
            cls._write_processing_log(wb, processing_log)

        cls._write_lines_to(wb, results, loc_lookup)
        cls._write_lines_so(wb, results, loc_lookup)
        cls._write_headers_so(wb, results, loc_lookup)
        cls._write_headers_to(wb, results, loc_lookup)
        cls._write_final_data(wb, results)
        cls._write_summary(wb, results)
        cls._write_unmatched(wb, results)
        cls._write_tester_master(wb, results)

        if so_products and master:
            cls._write_so_reference(wb, so_products, master)

        wb.save(output_path)

    @classmethod
    def _hdr_cell(cls, ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = cls.HEADER_FONT
        cell.fill = cls.HEADER_FILL
        # wrap_text lets multi-word labels (e.g. 'Gen. Bus. Posting
        # Group') stack neatly instead of forcing a very wide column.
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        cell.border = cls.BORDER
        return cell

    @classmethod
    def _data_cell(cls, ws, row, col, value, fmt=None, align='center'):
        # Cosmetic-only: data cells are centre-aligned by default for a
        # clean ID-strip look. Callers pass align='right' for money or
        # align='left' for long prose (product names). No value/logic
        # change — purely presentation.
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name='Aptos Display', size=11)
        cell.border = cls.BORDER
        cell.alignment = Alignment(horizontal=align, vertical='center')
        if fmt:
            cell.number_format = fmt
        return cell

    @classmethod
    def _auto_width(cls, ws, max_w=50):
        for col in ws.columns:
            letter = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            # min 9 so centred short values (Qty, codes) aren't cramped.
            ws.column_dimensions[letter].width = max(min(w + 3, max_w), 9)
        # Keep the header row visible while scrolling, and give it height
        # for the wrapped labels.
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 30

    @classmethod
    def _write_headers_to(cls, wb, results, loc_lookup):
        ws = wb.create_sheet('Headers (TO)')
        headers = [
            'No.', 'Transfer-from Code', 'Transfer-to Code', 'Posting Date',
            'In-Transit Code', 'Direct Transfer', 'Gen. Bus. Posting Group',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Department Code (Dimension)',
            'Geography Code (Dimension)',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        today_str = time.strftime("%d-%m-%Y")
        seen_to = set()

        for res in results:
            all_rows = (res.regular_orders + res.tester_orders +
                       res.pwp_orders + res.gwp_orders + res.nonstock_orders)
            for item in all_rows:
                to_num = item.to
                if to_num and to_num.startswith('TO/') and to_num not in seen_to:
                    seen_to.add(to_num)
                    loc = loc_lookup.get(res.filename, {})
                    cls._data_cell(ws, r, 1, to_num)
                    cls._data_cell(ws, r, 2, 'PICK')
                    cls._data_cell(ws, r, 3, loc.get('transfer_code', item.transfer_to))
                    cls._data_cell(ws, r, 4, today_str)
                    cls._data_cell(ws, r, 5, 'IN TRANSIT')
                    cls._data_cell(ws, r, 6, 'false')
                    cls._data_cell(ws, r, 7, loc.get('posting_group', item.posting_group))
                    r += 1

        cls._auto_width(ws)

    @classmethod
    def _write_headers_so(cls, wb, results, loc_lookup):
        ws = wb.create_sheet('Headers (SO)')
        headers = [
            'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
            'Posting Date', 'Order Date', 'Document Date',
            'Invoice From Date', 'Invoice To Date',
            'External Document No.', 'Location Code', 'Dimension Set ID',
            'Supply Type', 'Voucher Narration',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Geography Code (Dimension)',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        today_str = time.strftime("%d-%m-%Y")
        seen_so = set()

        for res in results:
            all_rows = (res.regular_orders + res.tester_orders +
                       res.pwp_orders + res.gwp_orders + res.nonstock_orders)
            for item in all_rows:
                so_num = item.to
                if so_num and so_num.startswith('SO/') and so_num not in seen_so:
                    seen_so.add(so_num)
                    loc = loc_lookup.get(res.filename, {})
                    bill_to = loc.get('bill_to', '')
                    ship_to = loc.get('ship_to', '')

                    cls._data_cell(ws, r, 1, 'Order')
                    cls._data_cell(ws, r, 2, so_num)
                    cls._data_cell(ws, r, 3, bill_to)
                    cls._data_cell(ws, r, 4, ship_to)
                    for c in range(5, 10):
                        cls._data_cell(ws, r, c, today_str)
                    cls._data_cell(ws, r, 10, so_num)
                    cls._data_cell(ws, r, 11, 'PICK')
                    cls._data_cell(ws, r, 12, '')
                    cls._data_cell(ws, r, 13, 'B2B')
                    r += 1

        cls._auto_width(ws)

    @classmethod
    def _write_lines_to(cls, wb, results, loc_lookup):
        ws = wb.create_sheet('Lines (TO)')
        headers = [
            'Document No.', 'Line No.', 'Item No.', 'Quantity',
            'Unit of Measure', 'Qty. to Ship', 'Qty. to Receive',
            'Dimension Set ID', 'Transfer Price',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        current_to = None
        line_no = 0

        for res in results:
            all_rows = (res.regular_orders + res.pwp_orders +
                       res.tester_orders + res.gwp_orders + res.nonstock_orders)
            for item in all_rows:
                if not item.to or not item.to.startswith('TO/'):
                    continue
                if item.to != current_to:
                    current_to = item.to
                    line_no = 0
                line_no += 10000

                cls._data_cell(ws, r, 1, item.to)
                cls._data_cell(ws, r, 2, line_no)
                cls._data_cell(ws, r, 3, item.item_no)
                cls._data_cell(ws, r, 4, item.qty)
                cls._data_cell(ws, r, 5, 'Piece-1')
                cls._data_cell(ws, r, 6, '')
                cls._data_cell(ws, r, 7, '')
                cls._data_cell(ws, r, 8, '')
                cls._data_cell(ws, r, 9,
                    round(item.unit_price, 10) if item.unit_price else 0,
                    '#,##0.0000000000', align='right')
                r += 1

        cls._auto_width(ws)

    @classmethod
    def _write_lines_so(cls, wb, results, loc_lookup):
        ws = wb.create_sheet('Lines (SO)')
        headers = [
            'Document Type', 'Document No.', 'Line No.', 'Type',
            'No.', 'Location Code', 'Quantity', 'Unit Price',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        current_so = None
        line_no = 0

        for res in results:
            all_rows = (res.regular_orders + res.pwp_orders +
                       res.tester_orders + res.gwp_orders + res.nonstock_orders)
            for item in all_rows:
                if not item.to or not item.to.startswith('SO/'):
                    continue
                if item.to != current_so:
                    current_so = item.to
                    line_no = 0
                line_no += 10000

                cls._data_cell(ws, r, 1, 'Order')
                cls._data_cell(ws, r, 2, item.to)
                cls._data_cell(ws, r, 3, line_no)
                cls._data_cell(ws, r, 4, 'Item')
                cls._data_cell(ws, r, 5, item.item_no)
                cls._data_cell(ws, r, 6, 'PICK')
                cls._data_cell(ws, r, 7, item.qty)
                cls._data_cell(ws, r, 8,
                    round(item.unit_price, 10) if item.unit_price else 0,
                    '#,##0.0000000000', align='right')
                r += 1

        cls._auto_width(ws)

    @classmethod
    def _write_final_data(cls, wb, results):
        ws = wb.create_sheet('Final Data')
        headers = [
            'TO', 'Item', 'Qty', 'Unit Price', 'Transfer-to Code',
            'Gen. Bus. Posting Group', 'Source', 'Location',
            'EAN', 'Product Name', 'Lookup Status',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        for res in results:
            loc = res.filename.replace('.xlsx', '').replace('_NEW_PO', '').replace('_New_PO', '')

            def write_row(item, row_num):
                cls._data_cell(ws, row_num, 1, item.to)
                cls._data_cell(ws, row_num, 2, item.item_no)
                cls._data_cell(ws, row_num, 3, item.qty)
                cls._data_cell(ws, row_num, 4,
                    round(item.unit_price, 10) if item.unit_price else 0,
                    '#,##0.0000000000', align='right')
                cls._data_cell(ws, row_num, 5, item.transfer_to)
                cls._data_cell(ws, row_num, 6, item.posting_group)

                src_cell = cls._data_cell(ws, row_num, 7, item.source)
                sc = cls.SOURCE_COLORS.get(item.source, ('333333', 'FFFFFF'))
                src_cell.fill = PatternFill('solid', fgColor=sc[0])
                src_cell.font = Font(name='Aptos Display', size=11, bold=True, color=sc[1])
                src_cell.alignment = Alignment(horizontal='center')

                cls._data_cell(ws, row_num, 8, loc, align='left')
                cls._data_cell(ws, row_num, 9, item.ean)
                cls._data_cell(ws, row_num, 10, item.product_name, align='left')

                st_cell = cls._data_cell(ws, row_num, 11, item.lookup_status)
                stc = cls.STATUS_COLORS.get(item.lookup_status, ('666666', 'FFFFFF'))
                st_cell.fill = PatternFill('solid', fgColor=stc[0])
                st_cell.font = Font(name='Aptos Display', size=11, bold=True, color=stc[1])
                st_cell.alignment = Alignment(horizontal='center')

                return row_num + 1

            for item in res.regular_orders: r = write_row(item, r)
            for item in res.pwp_orders: r = write_row(item, r)
            for item in res.tester_orders: r = write_row(item, r)
            for item in res.gwp_orders: r = write_row(item, r)
            for item in res.nonstock_orders: r = write_row(item, r)
            r += 1

        cls._auto_width(ws)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    @classmethod
    def _write_summary(cls, wb, results):
        """Per-location qty breakdown with TOTALS row."""
        ws = wb.create_sheet('Summary')
        headers = [
            'Location', 'TO/SO Number', 'Tester TO/SO',
            'Transfer-to Code', 'Gen. Bus. Posting Group',
            'PO Qty', 'PO Items', 'Tester Qty', 'Tester Items',
            'PWP Qty', 'GWP Qty', 'Non-Stock Qty', 'Total Qty',
            'Unmatched EANs',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        for i, res in enumerate(results, 2):
            loc = res.filename.replace('.xlsx', '').replace('_NEW_PO', '').replace('_New_PO', '')

            # Regular doc = FG only (regular_orders). Tester doc carries
            # every ₹0.54 line — testers, PWP, GWP, Non-Stock — so PWP now
            # falls back into the Tester TO/SO column, not the regular one.
            to_regular = ''
            to_tester = ''
            if res.regular_orders and res.regular_orders[0].to:
                to_regular = res.regular_orders[0].to
            if res.tester_orders and res.tester_orders[0].to:
                to_tester = res.tester_orders[0].to
            elif res.pwp_orders and res.pwp_orders[0].to:
                to_tester = res.pwp_orders[0].to
            elif res.gwp_orders and res.gwp_orders[0].to:
                to_tester = res.gwp_orders[0].to
            elif res.nonstock_orders and res.nonstock_orders[0].to:
                to_tester = res.nonstock_orders[0].to

            transfer_dest = ''
            posting_group = ''
            for bucket in (res.regular_orders, res.tester_orders,
                            res.pwp_orders, res.gwp_orders,
                            res.nonstock_orders):
                if bucket:
                    posting_group = bucket[0].posting_group or ''
                    break

            if to_regular.startswith('SO/') or to_tester.startswith('SO/'):
                transfer_dest = getattr(res, '_so_ship_to', '') or ''
            else:
                for bucket in (res.regular_orders, res.tester_orders,
                                res.pwp_orders, res.gwp_orders,
                                res.nonstock_orders):
                    if bucket and bucket[0].transfer_to:
                        transfer_dest = bucket[0].transfer_to
                        break

            po_q = sum(r.qty for r in res.regular_orders)
            tt_q = sum(r.qty for r in res.tester_orders)
            pw_q = sum(r.qty for r in res.pwp_orders)
            gw_q = sum(r.qty for r in res.gwp_orders)
            ns_q = sum(r.qty for r in res.nonstock_orders)
            total = po_q + tt_q + pw_q + gw_q + ns_q

            cls._data_cell(ws, i, 1, loc, align='left')
            cls._data_cell(ws, i, 2, to_regular)
            cls._data_cell(ws, i, 3, to_tester)
            cls._data_cell(ws, i, 4, transfer_dest)
            cls._data_cell(ws, i, 5, posting_group)
            cls._data_cell(ws, i, 6, po_q)
            cls._data_cell(ws, i, 7, len(res.regular_orders))
            cls._data_cell(ws, i, 8, tt_q)
            cls._data_cell(ws, i, 9, len(res.tester_orders))
            cls._data_cell(ws, i, 10, pw_q)
            cls._data_cell(ws, i, 11, gw_q)
            cls._data_cell(ws, i, 12, ns_q)
            cls._data_cell(ws, i, 13, total)
            cls._data_cell(ws, i, 14, len(res.unmatched))

            if res.unmatched:
                ws.cell(row=i, column=14).fill = PatternFill(
                    'solid', fgColor='FF5252')
                ws.cell(row=i, column=14).font = Font(
                    name='Aptos Display', size=11,
                    bold=True, color='FFFFFF')

        tr = len(results) + 2
        cls._data_cell(ws, tr, 1, 'TOTAL', align='left')
        for c in range(6, 15):
            total = sum(ws.cell(row=r, column=c).value or 0
                        for r in range(2, tr))
            cls._data_cell(ws, tr, c, total)
        # Band the whole TOTAL row (incl. the empty 2-5) light-grey + bold
        # so it reads as a clear footer strip.
        total_fill = PatternFill('solid', fgColor='EEEEEE')
        total_font = Font(name='Aptos Display', size=11, bold=True)
        for c in range(1, 15):
            cell = ws.cell(row=tr, column=c)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = cls.BORDER

        cls._auto_width(ws)
        ws.freeze_panes = 'A2'

    @classmethod
    def _write_processing_log(cls, wb, processing_log):
        """Per-file status sheet."""
        ws = wb.create_sheet('Processing Log', 0)
        headers = [
            'File', 'Location', 'Status',
            'TO/SO Number', 'Tester TO/SO',
            'Issues', 'Actions Taken',
        ]
        for c, h in enumerate(headers, 1):
            cls._hdr_cell(ws, 1, c, h)

        STATUS_FILLS = {
            'OK':         ('00C853', '000000'),
            'AUTO_FIXED': ('FFB300', '000000'),
            'WARNING':    ('FFB300', '000000'),
            'FAILED':     ('FF5252', 'FFFFFF'),
        }
        STATUS_LABELS = {
            'OK':         '✓ OK',
            'AUTO_FIXED': '⚠ Auto-Fixed',
            'WARNING':    '⚠ Warning',
            'FAILED':     '✗ FAILED',
        }

        for r, entry in enumerate(processing_log, 2):
            cls._data_cell(ws, r, 1, entry.get('filename', ''))
            cls._data_cell(ws, r, 2, entry.get('location', ''))

            status = entry.get('status', 'OK')
            status_cell = cls._data_cell(ws, r, 3, STATUS_LABELS.get(status, status))
            fill_color, font_color = STATUS_FILLS.get(status, ('666666', 'FFFFFF'))
            status_cell.fill = PatternFill('solid', fgColor=fill_color)
            status_cell.font = Font(name='Aptos Display', size=11,
                                      bold=True, color=font_color)
            status_cell.alignment = Alignment(horizontal='center')

            cls._data_cell(ws, r, 4, entry.get('to_number', ''))
            cls._data_cell(ws, r, 5, entry.get('tt_number', ''))

            issues = entry.get('issues', [])
            actions = entry.get('actions', [])
            cls._data_cell(ws, r, 6, '\n'.join(issues) if issues else '')
            cls._data_cell(ws, r, 7, '\n'.join(actions) if actions else '')

            ws.cell(row=r, column=6).alignment = Alignment(
                wrap_text=True, vertical='top')
            ws.cell(row=r, column=7).alignment = Alignment(
                wrap_text=True, vertical='top')

        if processing_log:
            counts = {'OK': 0, 'AUTO_FIXED': 0, 'WARNING': 0, 'FAILED': 0}
            for e in processing_log:
                s = e.get('status', 'OK')
                counts[s] = counts.get(s, 0) + 1

            tr = len(processing_log) + 2
            cls._data_cell(ws, tr, 1, 'TOTAL')
            ws.cell(row=tr, column=1).font = Font(
                name='Aptos Display', size=11, bold=True)

            summary_parts = []
            if counts['OK']: summary_parts.append(f"{counts['OK']} OK")
            if counts['AUTO_FIXED']: summary_parts.append(f"{counts['AUTO_FIXED']} auto-fixed")
            if counts['WARNING']: summary_parts.append(f"{counts['WARNING']} warning")
            if counts['FAILED']: summary_parts.append(f"{counts['FAILED']} failed")

            cls._data_cell(ws, tr, 3, ' | '.join(summary_parts))
            ws.cell(row=tr, column=3).font = Font(
                name='Aptos Display', size=11, bold=True)

        cls._auto_width(ws, max_w=60)
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 50
        ws.freeze_panes = 'A2'

    @classmethod
    def _write_unmatched(cls, wb, results):
        ws = wb.create_sheet('Unmatched EANs')
        for c, h in enumerate(['Location', 'EAN', 'Product Name', 'Order Qty', 'Tester Qty'], 1):
            cls._hdr_cell(ws, 1, c, h)

        r = 2
        for res in results:
            loc = res.filename.replace('.xlsx', '')
            for u in res.unmatched:
                cls._data_cell(ws, r, 1, loc)
                cls._data_cell(ws, r, 2, u['ean'])
                cls._data_cell(ws, r, 3, u['product_name'])
                cls._data_cell(ws, r, 4, u['order_qty'])
                cls._data_cell(ws, r, 5, u['tester_qty'])
                r += 1

        if r == 2:
            cls._data_cell(ws, 2, 1, 'No unmatched EANs — all lookups resolved! ✓')
            ws.cell(row=2, column=1).font = Font(name='Aptos Display', size=11, color='00C853')
            ws.merge_cells('A2:E2')

        cls._auto_width(ws)

    @classmethod
    def _write_tester_master(cls, wb, results):
        ws = wb.create_sheet('Tester Items Master')
        for c, h in enumerate(['Type', 'Product Name', 'EAN', 'Item No', 'Status', 'Used In Locations'], 1):
            cls._hdr_cell(ws, 1, c, h)

        items_map = {}
        for res in results:
            loc = res.filename.replace('.xlsx', '').replace('_NEW_PO', '').replace('_New_PO', '')
            for item_list in [res.pwp_orders, res.gwp_orders, res.nonstock_orders]:
                for item in item_list:
                    key = (item.source, item.ean or item.product_name)
                    if key not in items_map:
                        items_map[key] = {
                            'source': item.source, 'name': item.product_name,
                            'ean': item.ean, 'item_no': item.item_no,
                            'status': item.lookup_status, 'locations': set(),
                        }
                    items_map[key]['locations'].add(loc)
                    if item.lookup_status == 'OK':
                        items_map[key]['status'] = 'OK'
                        items_map[key]['item_no'] = item.item_no

        TYPE_COLORS = {'PWP': 'FF6600', 'GWP': '00BCD4', 'NON_STOCK': '795548'}

        r = 2
        for key in sorted(items_map.keys(), key=lambda k: (k[0], k[1])):
            entry = items_map[key]
            is_ok = entry['status'] == 'OK'

            cls._data_cell(ws, r, 1, entry['source'])
            tc = TYPE_COLORS.get(entry['source'], '333333')
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=tc)
            ws.cell(row=r, column=1).font = Font(
                name='Aptos Display', size=11, bold=True, color='FFFFFF')

            cls._data_cell(ws, r, 2, entry['name'])
            cls._data_cell(ws, r, 3, entry['ean'])
            cls._data_cell(ws, r, 4, entry['item_no'] if is_ok else '')

            status_text = 'OK' if is_ok else entry['status']
            cls._data_cell(ws, r, 5, status_text)
            if is_ok:
                ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor='00C853')
                ws.cell(row=r, column=5).font = Font(
                    name='Aptos Display', size=11, bold=True, color='000000')
            else:
                ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor='FF5252')
                ws.cell(row=r, column=5).font = Font(
                    name='Aptos Display', size=11, bold=True, color='FFFFFF')

            cls._data_cell(ws, r, 6, ', '.join(sorted(entry['locations'])))
            r += 1

        if r == 2:
            cls._data_cell(ws, 2, 1, 'No PWP/GWP/Non-Stock items found')
            ws.merge_cells('A2:F2')

        cls._auto_width(ws)

    @classmethod
    def _write_so_reference(cls, wb, so_products, master):
        ws = wb.create_sheet('SO Reference')
        headers = [
            'Description', 'EAN', 'EBO Qty', 'Airport Qty', 'Kiosk Qty',
            'Tester Qty', 'Item No', 'MRP', 'GST Code',
            'Landing (×0.6)', 'Cost Price',
        ]
        calc_hdr_fill = PatternFill('solid', fgColor='1B5E20')
        for c, h in enumerate(headers, 1):
            cell = cls._hdr_cell(ws, 1, c, h)
            if c >= 7:
                cell.fill = calc_hdr_fill

        r = 2
        for prod in so_products:
            ean = prod['ean']
            info = master.get(ean) or master.get(ean.lstrip('0'))

            if info:
                item_no = info['item_no']
                mrp = info['mrp']
                gst_code = info.get('gst_code', '')
                description = info.get('description', '')
                landing = float(mrp) * 0.60 if mrp and not pd.isna(mrp) else 0
                cost_price = POEngine.calc_cost_price(mrp, gst_code) or 0
            else:
                item_no = f'?EAN:{ean}'
                mrp = ''
                gst_code = ''
                description = ''
                landing = cost_price = 0

            cls._data_cell(ws, r, 1, description)
            cls._data_cell(ws, r, 2, ean)
            cls._data_cell(ws, r, 3, prod.get('ebo_qty', 0))
            cls._data_cell(ws, r, 4, prod.get('airport_qty', 0))
            cls._data_cell(ws, r, 5, prod.get('kiosk_qty', 0))
            cls._data_cell(ws, r, 6, prod.get('tester_qty', 0))

            calc_fill = PatternFill('solid', fgColor='E8F5E9')
            for ci in range(7, 12):
                ws.cell(row=r, column=ci).fill = calc_fill

            cls._data_cell(ws, r, 7, item_no)
            cls._data_cell(ws, r, 8, mrp, '#,##0.00' if mrp else None)
            cls._data_cell(ws, r, 9, gst_code)
            cls._data_cell(ws, r, 10, round(landing, 2) if landing else 0, '#,##0.00')
            cls._data_cell(ws, r, 11,
                round(cost_price, 10) if cost_price else 0, '#,##0.0000000000')
            r += 1

        cls._auto_width(ws)
        ws.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════════════════════════════════
#  SPECIAL ORDER ENGINE — broadcast products to all locations
# ═══════════════════════════════════════════════════════════════════════════════

class SpecialOrderEngine:
    """Processes Special Order broadcasts across all EKA locations."""

    EKA_REQUIRED_COLS = [
        'Short Name', 'Prefix', 'Short Code', 'Transfer Code',
        'Type', 'Gen. Biz. Posting Group',
    ]

    SO_STANDARD_COLS = {
        'ean':         ('EAN', []),
        'ebo_qty':     ('EBO Qty', ['EBO']),
        'airport_qty': ('Airport Qty', ['Airport']),
        'kiosk_qty':   ('Kiosk Qty', ['Kiosk']),
        'tester_qty':  ('Tester Qty', ['Tester']),
    }

    TYPE_QTY_MAP = {
        'EBO': 'ebo_qty',
        'Airport': 'airport_qty',
        'Kiosk': 'kiosk_qty',
    }

    def __init__(self, master: Dict[str, Dict]):
        self.master = master
        self.locations: List[Dict] = []
        self.products: List[Dict] = []
        self.so_col_map: Dict[str, int] = {}

    def _safe_int(self, val) -> int:
        try:
            if val is None or str(val).strip() in ('', '#N/A', 'None'):
                return 0
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    def _ean_str(self, raw) -> str:
        if raw is None:
            return ''
        return str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()

    @staticmethod
    def get_today_date_code() -> int:
        """Return the date code for today: DD + M + YY."""
        today = date.today()
        dd = today.strftime('%d')
        m = str(today.month)
        yy = today.strftime('%y')
        return int(f"{dd}{m}{yy}")

    @staticmethod
    def get_today_month_str() -> str:
        """Return current month as a 2-digit string ('04' for April)."""
        return f"{date.today().month:02d}"

    @staticmethod
    def generate_to_number(prefix: str, short_code: str,
                            is_tester: bool,
                            date_code: Optional[int] = None) -> str:
        """Build a TO/SO doc number string."""
        if date_code is None:
            date_code = SpecialOrderEngine.get_today_date_code()

        segment = 'TT' if is_tester else SpecialOrderEngine.get_today_month_str()
        return f"{prefix}/{short_code}/{segment}/{date_code}"

    def load_eka_data(self, filepath: str, logs: List[tuple]) -> int:
        """Parse EKA_DATA.xlsx → self.locations. Returns active count."""
        wb = load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]

        header_map = {}
        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            val = str(cell.value or '').strip()
            if val:
                header_map[val] = cell.column - 1

        col_idx = {}
        for req in self.EKA_REQUIRED_COLS:
            if req in header_map:
                col_idx[req] = header_map[req]
            else:
                for h, idx in header_map.items():
                    if (req.lower().replace(' ', '').replace('.', '') in
                            h.lower().replace(' ', '').replace('.', '')):
                        col_idx[req] = idx
                        break
                else:
                    logs.append(('error',
                        f"EKA_DATA: Column '{req}' not found. "
                        f"Available: {list(header_map.keys())}"))

        if len(col_idx) < len(self.EKA_REQUIRED_COLS):
            return 0

        location_col_idx = None
        for h, idx in header_map.items():
            if h.strip().lower() == 'location':
                location_col_idx = idx
                break

        self.locations = []
        skipped_inactive = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            short_name = row[col_idx['Short Name']].value
            if not short_name or str(short_name).strip() == '':
                continue

            prefix = str(row[col_idx['Prefix']].value or 'TO').strip()
            short_code = str(row[col_idx['Short Code']].value or '').strip()
            transfer_code = str(row[col_idx['Transfer Code']].value or '').strip()
            loc_type = str(row[col_idx['Type']].value or '').strip()
            posting = str(row[col_idx['Gen. Biz. Posting Group']].value or '').strip()

            location_code = ''
            if location_col_idx is not None:
                location_code = str(row[location_col_idx].value or '').strip()

            bill_to = ''
            ship_to = ''
            status = 'Active'
            for h, idx in header_map.items():
                h_lower = h.lower().strip()
                if 'bill' in h_lower and 'to' in h_lower:
                    bill_to = str(row[idx].value or '').strip()
                elif 'ship' in h_lower and 'to' in h_lower:
                    ship_to = str(row[idx].value or '').strip()
                elif h_lower == 'status':
                    status = str(row[idx].value or 'Active').strip()

            if status.lower() == 'inactive':
                skipped_inactive += 1
                logs.append(('info',
                    f"EKA_DATA: '{str(short_name).strip()}' Inactive — skipping"))
                continue

            if not short_code:
                logs.append(('warn',
                    f"EKA_DATA: '{short_name}' has no Short Code — skipping"))
                continue

            self.locations.append({
                'short_name': str(short_name).strip(),
                'prefix': prefix,
                'short_code': short_code,
                'transfer_code': transfer_code,
                'location': location_code,
                'type': loc_type,
                'posting_group': posting,
                'bill_to': bill_to,
                'ship_to': ship_to,
            })

        if skipped_inactive:
            logs.append(('info',
                f"EKA_DATA: Skipped {skipped_inactive} Inactive location(s)"))
        logs.append(('info',
            f"EKA_DATA: Loaded {len(self.locations)} Active locations"))

        return len(self.locations)

    def _detect_so_columns(self, ws, logs: List[tuple]) -> Dict[str, int]:
        """Detect Special Order columns with fallbacks + alerts."""
        hmap = {}
        all_headers = {}

        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            val = str(cell.value or '').strip()
            idx = cell.column - 1
            if val:
                all_headers[val] = idx

        for key, (standard, fallbacks) in self.SO_STANDARD_COLS.items():
            if standard in all_headers:
                hmap[key] = all_headers[standard]
            else:
                for fb in fallbacks:
                    if fb in all_headers:
                        hmap[key] = all_headers[fb]
                        logs.append(('alert',
                            f"Auto-fixed: '{fb}' → '{standard}'. "
                            f"Please rename column to '{standard}'."))
                        break

        return hmap

    def load_special_order(self, filepath: str, logs: List[tuple]) -> int:
        """Parse Special_Order.xlsx → self.products. Returns count."""
        wb = load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]

        self.so_col_map = self._detect_so_columns(ws, logs)

        if 'ean' not in self.so_col_map:
            logs.append(('error', "Special Order: 'EAN' not found — cannot process"))
            return 0

        missing_qty = []
        for key in ('ebo_qty', 'airport_qty', 'kiosk_qty', 'tester_qty'):
            if key not in self.so_col_map:
                std_name = self.SO_STANDARD_COLS[key][0]
                missing_qty.append(std_name)

        if missing_qty:
            logs.append(('error',
                f"Special Order: Missing columns: {', '.join(missing_qty)}"))
            return 0

        self.products = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            ean_raw = row[self.so_col_map['ean']].value
            if ean_raw is None:
                continue

            ean = self._ean_str(ean_raw)
            ebo_qty = self._safe_int(row[self.so_col_map['ebo_qty']].value)
            airport_qty = self._safe_int(row[self.so_col_map['airport_qty']].value)
            kiosk_qty = self._safe_int(row[self.so_col_map['kiosk_qty']].value)
            tester_qty = self._safe_int(row[self.so_col_map['tester_qty']].value)

            self.products.append({
                'ean': ean,
                'ebo_qty': ebo_qty,
                'airport_qty': airport_qty,
                'kiosk_qty': kiosk_qty,
                'tester_qty': tester_qty,
            })

        logs.append(('info', f"Special Order: Loaded {len(self.products)} products"))
        return len(self.products)

    def validate(self, logs: List[tuple]) -> bool:
        """Validate EANs against master and location types."""
        if not self.locations:
            logs.append(('error', "No locations loaded from EKA_DATA"))
            return False
        if not self.products:
            logs.append(('error', "No products loaded from Special Order"))
            return False

        missing = 0
        for prod in self.products:
            ean = prod['ean']
            info = self.master.get(ean) or self.master.get(ean.lstrip('0'))
            if not info:
                logs.append(('warn', f"Special Order: EAN {ean} not found in master"))
                missing += 1

        if missing:
            logs.append(('warn',
                f"Special Order: {missing} EAN(s) not found — will output with ?EAN:"))
        else:
            logs.append(('info',
                f"Special Order: All {len(self.products)} EANs found in master ✓"))

        valid_types = set(self.TYPE_QTY_MAP.keys())
        for loc in self.locations:
            if loc['type'] not in valid_types:
                logs.append(('warn',
                    f"EKA_DATA: '{loc['short_name']}' has unknown Type "
                    f"'{loc['type']}' — will skip"))

        return True

    def process(self, logs: List[tuple]) -> List[LocationResult]:
        """Generate LocationResult per location."""
        results = []
        counter = self.get_today_date_code()

        for loc in self.locations:
            loc_type = loc['type']
            qty_key = self.TYPE_QTY_MAP.get(loc_type)
            if not qty_key:
                logs.append(('warn',
                    f"Skipping '{loc['short_name']}' — unknown Type '{loc_type}'"))
                continue

            has_regular = any(p[qty_key] > 0 for p in self.products)
            has_tester = any(p['tester_qty'] > 0 for p in self.products)

            to_regular = ''
            to_tester = ''
            if has_regular:
                to_regular = self.generate_to_number(
                    loc['prefix'], loc['short_code'],
                    is_tester=False, date_code=counter)
                counter += 1
            if has_tester:
                to_tester = self.generate_to_number(
                    loc['prefix'], loc['short_code'],
                    is_tester=True, date_code=counter)
                counter += 1

            res = LocationResult(filename=loc['short_name'])
            res._so_bill_to = loc.get('bill_to', '')
            res._so_ship_to = loc.get('ship_to', '')
            unmatched = []

            for prod in self.products:
                ean = prod['ean']
                regular_qty = prod[qty_key]
                tester_qty = prod['tester_qty']

                info = self.master.get(ean) or self.master.get(ean.lstrip('0'))

                if info:
                    item_no = info['item_no']
                    cost = POEngine.calc_cost_price(info['mrp'], info['gst_code'])
                    product_name = str(info.get('description', ''))
                    status = 'OK'
                else:
                    item_no = f'?EAN:{ean}'
                    cost = None
                    product_name = ''
                    status = 'NOT_FOUND'
                    unmatched.append({
                        'ean': ean, 'product_name': '',
                        'order_qty': regular_qty, 'tester_qty': tester_qty,
                    })

                if regular_qty > 0:
                    res.regular_orders.append(OutputRow(
                        to=to_regular,
                        item_no=item_no, qty=regular_qty,
                        unit_price=cost or 0,
                        transfer_to=loc['transfer_code'],
                        posting_group=loc['posting_group'],
                        source='PO', ean=ean,
                        product_name=product_name, lookup_status=status,
                    ))

                if tester_qty > 0:
                    res.tester_orders.append(OutputRow(
                        to=to_tester,
                        item_no=item_no, qty=tester_qty, unit_price=0.54,
                        transfer_to=loc['transfer_code'],
                        posting_group=loc['posting_group'],
                        source='TESTER', ean=ean,
                        product_name=product_name, lookup_status=status,
                    ))

            res.unmatched = unmatched
            po_q = sum(r.qty for r in res.regular_orders)
            tt_q = sum(r.qty for r in res.tester_orders)
            res.logs.append(('info',
                f"PO: {len(res.regular_orders)} items ({po_q} qty), "
                f"Testers: {len(res.tester_orders)} items ({tt_q} qty)"))

            results.append(res)

        logs.append(('info',
            f"Special Order: Generated {len(results)} location results"))
        return results


# ═══════════════════════════════════════════════════════════════════════════════
#  D365 XML HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _d365_detect_data_style(xml: str, data_start_row: int = 4,
                              fallback: str = '11') -> str:
    """Find the style id used by data cells in a template sheet."""
    match = re.search(
        rf'<c r="[A-Z]+{data_start_row}"[^>]*s="(\d+)"', xml,
    )
    return match.group(1) if match else fallback


def _d365_ensure_enough_rows(xml: str, needed: int,
                               data_start_row: int,
                               columns: List[str],
                               style_id: str) -> str:
    """Append empty pre-styled <row> elements when the template's
    pre-existing row capacity is less than `needed`."""
    def _phantom_replacer(m: 're.Match[str]') -> str:
        row_num = int(m.group(1))
        body = m.group(0)
        if row_num <= data_start_row:
            return body
        if f'<c r="A{row_num}"' in body:
            return body
        return ''

    xml = re.sub(
        r'<row r="(\d+)"[^>]*>.*?</row>',
        _phantom_replacer, xml, flags=re.DOTALL,
    )

    row_nums = [int(x) for x in re.findall(r'<row r="(\d+)"', xml)]
    max_existing_row = max(row_nums) if row_nums else 0
    existing_capacity = max(0, max_existing_row - data_start_row + 1)

    if needed <= existing_capacity:
        return xml

    new_row_count = needed - existing_capacity
    first_new_row = max_existing_row + 1

    new_rows: List[str] = []
    for offset in range(new_row_count):
        row_num = first_new_row + offset
        cells = ''.join(
            f'<c r="{col}{row_num}" s="{style_id}"/>' for col in columns
        )
        new_rows.append(
            f'<row r="{row_num}" spans="1:{len(columns)}" '
            f'x14ac:dyDescent="0.3">{cells}</row>'
        )

    return xml.replace('</sheetData>', ''.join(new_rows) + '</sheetData>')


def _d365_fill_cell(xml: str, col: str, row_num: int, value,
                     is_string: bool, string_map: Dict[str, int]) -> str:
    """Replace a single cell's content in a sheet XML string."""
    ref = f"{col}{row_num}"

    if is_string:
        idx = string_map.get(str(value), 0)
        empty_replacement = f'<c r="{ref}" s="\\1" t="s"><v>{idx}</v></c>'
    else:
        empty_replacement = f'<c r="{ref}" s="\\1"><v>{value}</v></c>'

    empty_pat = f'<c r="{ref}" s="(\\d+)"\\s*/>'
    new_xml, n = re.subn(empty_pat, empty_replacement, xml, count=1)
    if n > 0:
        return new_xml

    prefilled_pat = (
        rf'<c r="{ref}"[^>]*s="(\d+)"[^>]*'
        r'(?:/>|>.*?</c>)'
    )

    def _replace(match: 're.Match[str]') -> str:
        style = match.group(1)
        if is_string:
            idx = string_map.get(str(value), 0)
            return f'<c r="{ref}" s="{style}" t="s"><v>{idx}</v></c>'
        return f'<c r="{ref}" s="{style}"><v>{value}</v></c>'

    return re.sub(prefilled_pat, _replace, xml, count=1, flags=re.DOTALL)


def _d365_fill_inline_string(xml: str, col: str, row_num: int,
                               value: str) -> str:
    """Replace a cell with a literal inline-string value."""
    ref = f"{col}{row_num}"
    esc = (str(value).replace('&', '&amp;')
                       .replace('<', '&lt;')
                       .replace('>', '&gt;'))

    empty_pat = f'<c r="{ref}" s="(\\d+)"\\s*/>'
    replacement = f'<c r="{ref}" s="\\1" t="inlineStr"><is><t>{esc}</t></is></c>'
    new_xml, n = re.subn(empty_pat, replacement, xml, count=1)
    if n > 0:
        return new_xml

    prefilled_pat = (
        rf'<c r="{ref}"[^>]*s="(\d+)"[^>]*(?:/>|>.*?</c>)'
    )

    def _rep(match: 're.Match[str]') -> str:
        style = match.group(1)
        return (f'<c r="{ref}" s="{style}" t="inlineStr">'
                f'<is><t>{esc}</t></is></c>')

    return re.sub(prefilled_pat, _rep, xml, count=1, flags=re.DOTALL)


def _d365_remove_rows_beyond(xml: str, max_row: int) -> str:
    """Drop every <row r="N">...</row> element where N > max_row."""
    def _replacer(match: 're.Match[str]') -> str:
        row_num = int(match.group(1))
        return '' if row_num > max_row else match.group(0)

    return re.sub(
        r'<row r="(\d+)"[^>]*>.*?</row>',
        _replacer, xml, flags=re.DOTALL,
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  D365 TRANSFER ORDER EXPORTER
# ═══════════════════════════════════════════════════════════════════════════════

class D365TOExporter:
    """Fills a D365 Transfer Order template via ZIP/XML manipulation."""

    HDR_COLS = list('ABCDEFGHIJKL')
    LINE_COLS = list('ABCDEFGHI')

    @staticmethod
    def export(results: List[LocationResult], template_path: str,
               output_path: str) -> str:
        """Fill the D365 TO template with processed data."""
        shutil.copy2(template_path, output_path)
        today_str = time.strftime("%d-%m-%Y")

        unique_tos = []
        seen = set()
        for res in results:
            all_rows = (res.regular_orders + res.pwp_orders +
                        res.tester_orders + res.gwp_orders +
                        res.nonstock_orders)
            for item in all_rows:
                if (item.to and item.to.startswith('TO/')
                        and item.to not in seen):
                    seen.add(item.to)
                    unique_tos.append(
                        (item.to, item.transfer_to, item.posting_group))

        all_lines = []
        for res in results:
            all_rows = (res.regular_orders + res.pwp_orders +
                        res.tester_orders + res.gwp_orders +
                        res.nonstock_orders)
            for item in all_rows:
                if item.to and item.to.startswith('TO/'):
                    all_lines.append(
                        (item.to, item.item_no, item.qty, item.unit_price))

        if not unique_tos:
            return output_path

        zip_contents = {}
        with zipfile.ZipFile(output_path, 'r') as z:
            for zi in z.namelist():
                zip_contents[zi] = z.read(zi)

        ss_xml = zip_contents['xl/sharedStrings.xml'].decode('utf-8')
        existing = re.findall(r'<t[^>]*>([^<]*)</t>', ss_xml)
        string_map = {s: i for i, s in enumerate(existing)}

        new_strings = {'PICK', 'IN TRANSIT', 'false', today_str, 'Piece-1'}
        for to_num, tc, pg in unique_tos:
            new_strings.add(to_num)
            if tc:
                new_strings.add(tc)
            if pg:
                new_strings.add(pg)
        for _, item_no, _, _ in all_lines:
            try:
                int(str(item_no))
            except (ValueError, TypeError):
                new_strings.add(str(item_no))

        next_idx = len(existing)
        for s in sorted(new_strings):
            if s not in string_map:
                string_map[s] = next_idx
                next_idx += 1

        total_count = next_idx
        si_items = [''] * total_count
        for s, idx in string_map.items():
            esc = (s.replace('&', '&amp;')
                    .replace('<', '&lt;')
                    .replace('>', '&gt;'))
            si_items[idx] = f'<si><t>{esc}</t></si>'

        zip_contents['xl/sharedStrings.xml'] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{total_count}" uniqueCount="{total_count}">'
            + ''.join(si_items) + '</sst>'
        ).encode('utf-8')

        s1 = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        s1_style = _d365_detect_data_style(s1, data_start_row=4, fallback='11')
        s1 = _d365_ensure_enough_rows(
            s1, needed=len(unique_tos), data_start_row=4,
            columns=D365TOExporter.HDR_COLS, style_id=s1_style,
        )

        for i, (to_num, tc, pg) in enumerate(unique_tos):
            r = i + 4
            s1 = _d365_fill_cell(s1, 'A', r, to_num, True, string_map)
            s1 = _d365_fill_cell(s1, 'B', r, 'PICK', True, string_map)
            s1 = _d365_fill_cell(s1, 'C', r, tc or '', True, string_map)
            s1 = _d365_fill_cell(s1, 'D', r, today_str, True, string_map)
            s1 = _d365_fill_cell(s1, 'E', r, 'IN TRANSIT', True, string_map)
            s1 = _d365_fill_cell(s1, 'F', r, 'false', True, string_map)
            if pg:
                s1 = _d365_fill_cell(s1, 'G', r, pg, True, string_map)

        zip_contents['xl/worksheets/sheet1.xml'] = s1.encode('utf-8')

        s2 = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        s2_style = _d365_detect_data_style(s2, data_start_row=4, fallback='8')
        s2 = _d365_ensure_enough_rows(
            s2, needed=len(all_lines), data_start_row=4,
            columns=D365TOExporter.LINE_COLS, style_id=s2_style,
        )

        current_doc = None
        line_no = 0
        for i, (doc_no, item_no, qty, price) in enumerate(all_lines):
            if doc_no != current_doc:
                current_doc = doc_no
                line_no = 0
            line_no += 10000
            r = i + 4

            s2 = _d365_fill_cell(s2, 'A', r, doc_no, True, string_map)
            s2 = _d365_fill_cell(s2, 'B', r, line_no, False, string_map)

            try:
                item_int = int(str(item_no))
                s2 = _d365_fill_cell(s2, 'C', r, item_int, False, string_map)
            except (ValueError, TypeError):
                val = str(item_no)
                if val in string_map:
                    s2 = _d365_fill_cell(s2, 'C', r, val, True, string_map)
                else:
                    s2 = _d365_fill_inline_string(s2, 'C', r, val)

            s2 = _d365_fill_cell(s2, 'D', r, qty, False, string_map)
            s2 = _d365_fill_cell(s2, 'E', r, 'Piece-1', True, string_map)
            s2 = _d365_fill_cell(
                s2, 'I', r, price if price else 0, False, string_map)

        zip_contents['xl/worksheets/sheet2.xml'] = s2.encode('utf-8')

        last_hdr = 3 + len(unique_tos)
        last_line = 3 + len(all_lines)

        s1c = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        s1c = _d365_remove_rows_beyond(s1c, last_hdr)
        s1c = re.sub(r'<dimension ref="[^"]*"/>',
                      f'<dimension ref="A1:L{last_hdr}"/>', s1c)
        zip_contents['xl/worksheets/sheet1.xml'] = s1c.encode('utf-8')

        s2c = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        s2c = _d365_remove_rows_beyond(s2c, last_line)
        s2c = re.sub(r'<dimension ref="[^"]*"/>',
                      f'<dimension ref="A1:I{last_line}"/>', s2c)
        zip_contents['xl/worksheets/sheet2.xml'] = s2c.encode('utf-8')

        for tbl in ['xl/tables/table1.xml', 'xl/tables/table2.xml']:
            if tbl in zip_contents:
                t = zip_contents[tbl].decode('utf-8')
                if 'table1' in tbl:
                    t = re.sub(r'ref="A3:[A-Z]+\d+"',
                               f'ref="A3:L{last_hdr}"', t)
                else:
                    t = re.sub(r'ref="A3:[A-Z]+\d+"',
                               f'ref="A3:I{last_line}"', t)
                zip_contents[tbl] = t.encode('utf-8')

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zo:
            for name, data in zip_contents.items():
                zo.writestr(name, data)

        return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  D365 SALES ORDER EXPORTER
# ═══════════════════════════════════════════════════════════════════════════════

class D365SOExporter:
    """Fills a D365 Sales Order template via ZIP/XML manipulation."""

    HDR_COLS = list('ABCDEFGHIJKLMN')
    LINE_COLS = list('ABCDEFGH')

    @staticmethod
    def export(results: List[LocationResult], template_path: str,
               output_path: str) -> str:
        """Fill the D365 SO template with processed data."""
        shutil.copy2(template_path, output_path)
        today_str = time.strftime("%d-%m-%Y")

        unique_sos = []
        seen = set()
        for res in results:
            all_rows = (res.regular_orders + res.tester_orders +
                        res.pwp_orders + res.gwp_orders +
                        res.nonstock_orders)
            for item in all_rows:
                if (item.to and item.to.startswith('SO/')
                        and item.to not in seen):
                    seen.add(item.to)
                    bill = getattr(res, '_so_bill_to', '')
                    ship = getattr(res, '_so_ship_to', '')
                    unique_sos.append((item.to, bill, ship))

        all_lines = []
        for res in results:
            all_rows = (res.regular_orders + res.pwp_orders +
                        res.tester_orders + res.gwp_orders +
                        res.nonstock_orders)
            for item in all_rows:
                if item.to and item.to.startswith('SO/'):
                    all_lines.append(
                        (item.to, item.item_no, item.qty, item.unit_price))

        if not unique_sos:
            return output_path

        zip_contents = {}
        with zipfile.ZipFile(output_path, 'r') as z:
            for zi in z.namelist():
                zip_contents[zi] = z.read(zi)

        ss_xml = zip_contents['xl/sharedStrings.xml'].decode('utf-8')
        existing = re.findall(r'<t[^>]*>([^<]*)</t>', ss_xml)
        string_map = {s: i for i, s in enumerate(existing)}

        new_strings = {'Order', 'PICK', 'B2B', today_str, 'Item'}
        for so_num, bill, ship in unique_sos:
            new_strings.add(so_num)
            if bill:
                new_strings.add(bill)
            if ship:
                new_strings.add(ship)
        for _, item_no, _, _ in all_lines:
            try:
                int(str(item_no))
            except (ValueError, TypeError):
                new_strings.add(str(item_no))

        next_idx = len(existing)
        for s in sorted(new_strings):
            if s not in string_map:
                string_map[s] = next_idx
                next_idx += 1

        total_count = next_idx
        si_items = [''] * total_count
        for s, idx in string_map.items():
            esc = (s.replace('&', '&amp;')
                    .replace('<', '&lt;')
                    .replace('>', '&gt;'))
            si_items[idx] = f'<si><t>{esc}</t></si>'

        zip_contents['xl/sharedStrings.xml'] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{total_count}" uniqueCount="{total_count}">'
            + ''.join(si_items) + '</sst>'
        ).encode('utf-8')

        s1 = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        s1_style = _d365_detect_data_style(s1, data_start_row=4, fallback='5')
        s1 = _d365_ensure_enough_rows(
            s1, needed=len(unique_sos), data_start_row=4,
            columns=D365SOExporter.HDR_COLS, style_id=s1_style,
        )

        for i, (so_num, bill, ship) in enumerate(unique_sos):
            r = i + 4
            s1 = _d365_fill_cell(s1, 'A', r, 'Order', True, string_map)
            s1 = _d365_fill_cell(s1, 'B', r, so_num, True, string_map)
            if bill:
                s1 = _d365_fill_cell(s1, 'C', r, bill, True, string_map)
            if ship:
                s1 = _d365_fill_cell(s1, 'D', r, ship, True, string_map)
            for col in 'EFGHI':
                s1 = _d365_fill_cell(s1, col, r, today_str, True, string_map)
            s1 = _d365_fill_cell(s1, 'J', r, so_num, True, string_map)
            s1 = _d365_fill_cell(s1, 'K', r, 'PICK', True, string_map)
            s1 = _d365_fill_cell(s1, 'M', r, 'B2B', True, string_map)

        zip_contents['xl/worksheets/sheet1.xml'] = s1.encode('utf-8')

        s2 = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        s2_style = _d365_detect_data_style(s2, data_start_row=4, fallback='4')
        s2 = _d365_ensure_enough_rows(
            s2, needed=len(all_lines), data_start_row=4,
            columns=D365SOExporter.LINE_COLS, style_id=s2_style,
        )

        current_doc = None
        line_no = 0
        for i, (doc_no, item_no, qty, price) in enumerate(all_lines):
            if doc_no != current_doc:
                current_doc = doc_no
                line_no = 0
            line_no += 10000
            r = i + 4

            s2 = _d365_fill_cell(s2, 'A', r, 'Order', True, string_map)
            s2 = _d365_fill_cell(s2, 'B', r, doc_no, True, string_map)
            s2 = _d365_fill_cell(s2, 'C', r, line_no, False, string_map)
            s2 = _d365_fill_cell(s2, 'D', r, 'Item', True, string_map)

            try:
                item_int = int(str(item_no))
                s2 = _d365_fill_cell(s2, 'E', r, item_int, False, string_map)
            except (ValueError, TypeError):
                val = str(item_no)
                if val in string_map:
                    s2 = _d365_fill_cell(s2, 'E', r, val, True, string_map)
                else:
                    s2 = _d365_fill_inline_string(s2, 'E', r, val)

            s2 = _d365_fill_cell(s2, 'F', r, 'PICK', True, string_map)
            s2 = _d365_fill_cell(s2, 'G', r, qty, False, string_map)
            s2 = _d365_fill_cell(
                s2, 'H', r, price if price else 0, False, string_map)

        zip_contents['xl/worksheets/sheet2.xml'] = s2.encode('utf-8')

        last_hdr = 3 + len(unique_sos)
        last_line = 3 + len(all_lines)

        s1c = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        s1c = _d365_remove_rows_beyond(s1c, last_hdr)
        s1c = re.sub(r'<dimension ref="[^"]*"/>',
                      f'<dimension ref="A1:N{last_hdr}"/>', s1c)
        zip_contents['xl/worksheets/sheet1.xml'] = s1c.encode('utf-8')

        s2c = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        s2c = _d365_remove_rows_beyond(s2c, last_line)
        s2c = re.sub(r'<dimension ref="[^"]*"/>',
                      f'<dimension ref="A1:H{last_line}"/>', s2c)
        zip_contents['xl/worksheets/sheet2.xml'] = s2c.encode('utf-8')

        for tbl in ['xl/tables/table1.xml', 'xl/tables/table2.xml']:
            if tbl in zip_contents:
                t = zip_contents[tbl].decode('utf-8')
                if 'table1' in tbl:
                    t = re.sub(r'ref="A3:[A-Z]+\d+"',
                               f'ref="A3:N{last_hdr}"', t)
                else:
                    t = re.sub(r'ref="A3:[A-Z]+\d+"',
                               f'ref="A3:H{last_line}"', t)
                zip_contents[tbl] = t.encode('utf-8')

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zo:
            for name, data in zip_contents.items():
                zo.writestr(name, data)

        return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  UPDATE DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class UpdateDialog:
    """Modal dialog asking which bundled file(s) to update."""

    def __init__(self, parent, folder: Path):
        self.parent = parent
        self.folder = folder
        self.result: Optional[str] = None

        self.top = tk.Toplevel(parent)
        self.top.title("Update Bundled Files")
        self.top.resizable(False, False)
        self.top.transient(parent)
        self.top.grab_set()

        self._build_ui()

        self.top.update_idletasks()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        w = self.top.winfo_width()
        h = self.top.winfo_height()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.top.geometry(f"+{max(0, x)}+{max(0, y)}")

    def _build_ui(self) -> None:
        tk.Label(
            self.top, text="Which file would you like to update?",
            font=("Arial", 11, "bold"),
        ).pack(padx=20, pady=(16, 6))

        tk.Label(
            self.top,
            text=f"Files will be saved to:\n{self.folder}",
            font=("Arial", 9), fg='gray', justify='center',
        ).pack(padx=20, pady=(0, 12))

        self.choice_var = tk.StringVar(value='master')

        radio_frame = tk.Frame(self.top)
        radio_frame.pack(padx=30, pady=(0, 12), anchor='w')

        tk.Radiobutton(
            radio_frame, text="Items Master  (Items_March.xlsx)",
            variable=self.choice_var, value='master',
            font=("Arial", 10),
        ).pack(anchor='w', pady=2)
        tk.Radiobutton(
            radio_frame, text="EKA_DATA  (EKA_DATA.xlsx)",
            variable=self.choice_var, value='eka',
            font=("Arial", 10),
        ).pack(anchor='w', pady=2)
        tk.Radiobutton(
            radio_frame, text="Both  (pick Master first, then EKA_DATA)",
            variable=self.choice_var, value='both',
            font=("Arial", 10),
        ).pack(anchor='w', pady=2)

        btn_frame = tk.Frame(self.top)
        btn_frame.pack(pady=(0, 16))

        tk.Button(
            btn_frame, text="Cancel", width=10,
            command=self._on_cancel,
        ).pack(side='left', padx=6)
        tk.Button(
            btn_frame, text="Update", width=10,
            font=("Arial", 10, "bold"),
            bg="#00C853", fg='white',
            command=self._on_update,
        ).pack(side='left', padx=6)

        self.top.bind('<Escape>', lambda _e: self._on_cancel())
        self.top.protocol('WM_DELETE_WINDOW', self._on_cancel)

    def _on_update(self) -> None:
        self.result = self.choice_var.get()
        self.top.destroy()

    def _on_cancel(self) -> None:
        self.result = None
        self.top.destroy()

    def show(self) -> Optional[str]:
        self.parent.wait_window(self.top)
        return self.result


# ═══════════════════════════════════════════════════════════════════════════════
#  GUI APPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

class ReneePOApp:
    """Simple Tkinter app for RENEE PO Processor."""

    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("560x780")
        self.root.resizable(False, False)

        self.engine = POEngine()
        self.eka_locations: List[Dict] = []

        self.master_path: Optional[str] = None
        self.eka_path: Optional[str] = None
        self.po_files: List[str] = []
        self.so_path: Optional[str] = None

        self.last_output: Optional[Path] = None
        self.last_results: List[LocationResult] = []
        self.is_running = False

        self.master_is_bundled = False
        self.eka_is_bundled = False

        self.mode_var: tk.StringVar
        self.master_var: tk.StringVar
        self.eka_var: tk.StringVar
        self.po_files_var: tk.StringVar
        self.so_var: tk.StringVar
        self.status_var: tk.StringVar
        self.po_listbox: tk.Listbox
        self.log_text: tk.Text
        self.gen_btn: tk.Button
        self.open_btn: tk.Button
        self.d365_btn: tk.Button
        self.po_frame: tk.LabelFrame
        self.so_frame: tk.LabelFrame

        self._build_ui()

        get_bundled_folder(create=True)

        self._auto_load_bundled()

        self.root.update()
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after(500, lambda: self.root.attributes('-topmost', False))

    def _build_ui(self) -> None:
        tk.Label(self.root, text="RENEE PO Processor",
                  font=("Arial", 14, "bold")).pack(pady=(12, 2))
        tk.Label(self.root, text=f"v{APP_VERSION}  •  EBO / Kiosk / Airport PO → SO/TO",
                  font=("Arial", 9), fg='gray').pack(pady=(0, 10))

        mode_frame = tk.Frame(self.root)
        mode_frame.pack(fill='x', padx=20, pady=(0, 8))
        tk.Label(mode_frame, text="Mode:", font=("Arial", 10, "bold")).pack(side='left')

        self.mode_var = tk.StringVar(value='standalone')
        tk.Radiobutton(
            mode_frame, text="Standalone PO files",
            variable=self.mode_var, value='standalone',
            command=self._on_mode_change,
            font=("Arial", 10),
        ).pack(side='left', padx=(8, 4))
        tk.Radiobutton(
            mode_frame, text="Special Order broadcast",
            variable=self.mode_var, value='special',
            command=self._on_mode_change,
            font=("Arial", 10),
        ).pack(side='left')

        files_frame = tk.LabelFrame(
            self.root, text="Input Files",
            font=("Arial", 10, "bold"), padx=10, pady=8,
        )
        files_frame.pack(fill='x', padx=20, pady=(0, 8))

        self.master_var = tk.StringVar(value="Not selected")
        self._build_file_row(
            files_frame, "Items Master:",
            self.master_var, self._select_master,
        )

        self.eka_var = tk.StringVar(value="Not selected")
        self._build_file_row(
            files_frame, "EKA Data:",
            self.eka_var, self._select_eka,
        )

        self.po_frame = tk.LabelFrame(
            self.root, text="PO Files (Standalone)",
            font=("Arial", 10, "bold"), padx=10, pady=8,
        )
        self.po_frame.pack(fill='both', expand=False, padx=20, pady=(0, 8))

        po_btn_row = tk.Frame(self.po_frame)
        po_btn_row.pack(fill='x', pady=(0, 4))
        tk.Button(po_btn_row, text="+ Add Files",
                   command=self._add_files, width=12).pack(side='left', padx=(0, 4))
        tk.Button(po_btn_row, text="✕ Clear All",
                   command=self._clear_files, width=10).pack(side='left')

        self.po_files_var = tk.StringVar(value="0 files added")
        tk.Label(po_btn_row, textvariable=self.po_files_var,
                  font=("Arial", 9), fg='gray').pack(side='right')

        list_frame = tk.Frame(self.po_frame)
        list_frame.pack(fill='both', expand=True)
        list_scroll = ttk.Scrollbar(list_frame, orient='vertical')
        list_scroll.pack(side='right', fill='y')
        self.po_listbox = tk.Listbox(
            list_frame, height=5, font=("Consolas", 9),
            yscrollcommand=list_scroll.set,
        )
        self.po_listbox.pack(fill='both', expand=True)
        list_scroll.config(command=self.po_listbox.yview)

        self.so_frame = tk.LabelFrame(
            self.root, text="Special Order File",
            font=("Arial", 10, "bold"), padx=10, pady=8,
        )

        self.so_var = tk.StringVar(value="Not selected")
        self._build_file_row(
            self.so_frame, "Special Order:",
            self.so_var, self._select_so,
        )

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=(4, 4))

        self.gen_btn = tk.Button(
            btn_frame, text="▶  Generate Output", width=24,
            font=("Arial", 10, "bold"),
            bg="#00C853", fg='white', command=self._run,
        )
        self.gen_btn.pack(pady=3)

        self.open_btn = tk.Button(
            btn_frame, text="📂  Open Last Output", width=24,
            state=tk.DISABLED, command=self._open_last,
        )
        self.open_btn.pack(pady=3)

        self.d365_btn = tk.Button(
            btn_frame, text="📤  Export D365 TO Package", width=24,
            state=tk.DISABLED, command=self._export_d365,
        )
        self.d365_btn.pack(pady=3)

        tmpl_btn = tk.Menubutton(
            btn_frame, text="📋  Download Templates ▾",
            width=22, relief='raised',
        )
        tmpl_menu = tk.Menu(tmpl_btn, tearoff=0)
        tmpl_btn['menu'] = tmpl_menu
        tmpl_menu.add_command(label="Blank PO Template",
                                command=self._download_po_template)
        tmpl_menu.add_command(label="EKA_DATA Template",
                                command=self._download_eka_template)
        tmpl_menu.add_command(label="Special Order Template",
                                command=self._download_so_template)
        tmpl_btn.pack(pady=3)

        tk.Button(btn_frame, text="📁  Update Bundled Files", width=24,
                   command=self._update_bundled).pack(pady=3)

        self.status_var = tk.StringVar(value="Ready")
        tk.Label(self.root, textvariable=self.status_var,
                  font=("Arial", 10), fg='gray',
                  wraplength=520).pack(pady=4)

        log_frame = tk.LabelFrame(self.root, text="Log", font=("Arial", 9))
        log_frame.pack(fill='both', expand=True, padx=20, pady=(0, 12))

        log_scroll = ttk.Scrollbar(log_frame, orient='vertical')
        log_scroll.pack(side='right', fill='y')
        self.log_text = tk.Text(
            log_frame, height=8, font=("Consolas", 9),
            state='disabled', wrap='word',
            yscrollcommand=log_scroll.set,
        )
        self.log_text.pack(fill='both', expand=True)
        log_scroll.config(command=self.log_text.yview)

        self.log_text.tag_config('ok', foreground='#00A651')
        self.log_text.tag_config('err', foreground='#D32F2F')
        self.log_text.tag_config('warn', foreground='#E65100')
        self.log_text.tag_config('inf', foreground='#0077B6')
        self.log_text.tag_config('dim', foreground='#5F6368')

    def _build_file_row(self, parent, label_text: str,
                          var: tk.StringVar, browse_cmd) -> None:
        row = tk.Frame(parent)
        row.pack(fill='x', pady=2)

        tk.Label(
            row, text=label_text,
            font=("Arial", 10), width=15, anchor='w',
        ).pack(side='left')

        tk.Button(
            row, text="Browse", width=8, command=browse_cmd,
        ).pack(side='right', padx=(4, 0))

        tk.Label(
            row, textvariable=var,
            font=("Arial", 10), fg='#1A237E', anchor='w',
            wraplength=300, justify='left',
        ).pack(side='left', fill='x', expand=True)

    def _auto_load_bundled(self) -> None:
        m_path = get_bundled_master_path()
        if m_path:
            try:
                count = self.engine.load_master(str(m_path))
                self.master_path = str(m_path)
                self.master_is_bundled = True
                self.master_var.set(f"✓ {m_path.name} (auto-loaded)")
                self._log(f"Auto-loaded master from "
                          f"{BUNDLED_DATA_FOLDER}/{m_path.name} → {count:,} items", 'ok')
            except Exception as e:
                self._log(f"Master auto-load failed: {e}", 'err')
        else:
            self._log(f"No bundled master at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_MASTER_NAME} "
                      f"— Browse to pick one or use 'Update Bundled Files'", 'dim')

        e_path = get_bundled_eka_path()
        if e_path:
            self._load_eka_file(str(e_path), is_bundled=True)
        else:
            self._log(f"No bundled EKA_DATA at "
                      f"{BUNDLED_DATA_FOLDER}/{BUNDLED_EKA_NAME} "
                      f"— Browse to pick one or use 'Update Bundled Files'", 'dim')

    def _load_eka_file(self, path: str, is_bundled: bool) -> None:
        try:
            so_engine = SpecialOrderEngine(self.engine.master or {})
            logs = []
            count = so_engine.load_eka_data(path, logs)
            self.eka_locations = so_engine.locations
            self.eka_path = path
            self.eka_is_bundled = is_bundled

            label = (f"✓ {os.path.basename(path)} (auto-loaded)"
                     if is_bundled
                     else os.path.basename(path))
            self.eka_var.set(label)

            tag = 'ok' if is_bundled else 'inf'
            self._log(f"EKA_DATA loaded: {os.path.basename(path)} "
                      f"→ {count} active locations", tag)

            for level, msg in logs:
                if level != 'info':
                    log_tag = {'warn': 'warn', 'error': 'err',
                               'alert': 'warn'}.get(level, 'dim')
                    self._log(f"  {msg}", log_tag)
        except Exception as e:
            self.eka_var.set("ERROR — see log")
            self._log(f"EKA load failed: {e}", 'err')
            self._log(f"  Path attempted: {path}", 'dim')
            if not is_bundled:
                messagebox.showerror(
                    "EKA Load Failed",
                    f"Could not load EKA_DATA:\n\n{e}\n\n"
                    f"File: {os.path.basename(path)}",
                )

    def _log(self, msg: str, tag: str = 'dim') -> None:
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n", tag)
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _set_status(self, msg: str, color: str = 'gray') -> None:
        self.status_var.set(msg)

    def _on_mode_change(self) -> None:
        mode = self.mode_var.get()
        if mode == 'standalone':
            self.so_frame.pack_forget()
            self.po_frame.pack(
                fill='both', expand=False, padx=20, pady=(0, 8),
                before=self.gen_btn.master,
            )
            self._log("Mode: Standalone (per-location PO files)", 'inf')
        else:
            self.po_frame.pack_forget()
            self.so_frame.pack(
                fill='both', expand=False, padx=20, pady=(0, 8),
                before=self.gen_btn.master,
            )
            self._log("Mode: Special Order (broadcast across all locations)", 'inf')

    def _select_master(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Items Master file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            count = self.engine.load_master(path)
            self.master_path = path
            self.master_var.set(os.path.basename(path))
            self._log(f"Master loaded: {os.path.basename(path)} → "
                      f"{count:,} items", 'ok')

            try:
                target = get_bundled_folder(create=True) / BUNDLED_MASTER_NAME
                if Path(path).resolve() != target.resolve():
                    shutil.copy2(path, str(target))
                    self.master_path = str(target)
                    self.master_is_bundled = True
                    self.master_var.set(f"✓ {target.name} (auto-loaded)")
                    self._log(f"Saved to bundled folder → "
                              f"{BUNDLED_DATA_FOLDER}/{target.name} "
                              f"(will auto-load next time)", 'ok')
                else:
                    self.master_is_bundled = True
            except Exception as copy_err:
                self.master_is_bundled = False
                self._log(f"Could not save to bundled folder: {copy_err}",
                          'warn')
                self._log("  (Master is loaded for this session only — "
                          "use 'Update Bundled Files' to retry)", 'dim')

        except Exception as e:
            self.master_var.set("ERROR — see log")
            self._log(f"Master load failed: {e}", 'err')
            self._log(f"  Path attempted: {path}", 'dim')
            messagebox.showerror(
                "Master Load Failed",
                f"Could not load Items Master:\n\n{e}\n\n"
                f"File: {os.path.basename(path)}",
            )

    def _select_eka(self) -> None:
        path = filedialog.askopenfilename(
            title="Select EKA_DATA file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return

        self._load_eka_file(path, is_bundled=False)

        if self.eka_locations:
            try:
                target = get_bundled_folder(create=True) / BUNDLED_EKA_NAME
                if Path(path).resolve() != target.resolve():
                    shutil.copy2(path, str(target))
                    self.eka_path = str(target)
                    self.eka_is_bundled = True
                    self.eka_var.set(f"✓ {target.name} (auto-loaded)")
                    self._log(f"Saved to bundled folder → "
                              f"{BUNDLED_DATA_FOLDER}/{target.name} "
                              f"(will auto-load next time)", 'ok')
            except Exception as copy_err:
                self._log(f"Could not save to bundled folder: {copy_err}",
                          'warn')

    def _select_so(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Special Order file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        self.so_path = path
        self.so_var.set(os.path.basename(path))
        self._log(f"Special Order: {os.path.basename(path)}", 'inf')

    def _add_files(self) -> None:
        files = filedialog.askopenfilenames(
            title="Select PO Excel Files",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        added = 0
        for f in files:
            if (f not in self.po_files
                    and 'Items_March' not in f
                    and 'EKA_DATA' not in f
                    and 'PO_Output' not in f):
                self.po_files.append(f)
                added += 1
        self._refresh_po_list()
        if added:
            self._log(f"Added {added} PO file(s)", 'inf')

    def _clear_files(self) -> None:
        self.po_files.clear()
        self._refresh_po_list()
        self._log("PO file list cleared", 'dim')

    def _refresh_po_list(self) -> None:
        self.po_listbox.delete(0, 'end')
        for i, p in enumerate(self.po_files, 1):
            name = os.path.basename(p)
            display = name if len(name) <= 60 else name[:57] + '...'
            self.po_listbox.insert('end', f"  {i}. {display}")
        n = len(self.po_files)
        self.po_files_var.set(f"{n} file{'s' if n != 1 else ''} added")

    def _update_bundled(self) -> None:
        target = get_bundled_folder(create=True)
        dialog = UpdateDialog(self.root, folder=target)
        choice = dialog.show()
        if choice is None:
            return

        updated_any = False

        if choice in ('master', 'both'):
            updated_any |= self._do_update_one(
                kind='Items Master',
                title='Select new Items Master file',
                target_path=target / BUNDLED_MASTER_NAME,
            )

        if choice in ('eka', 'both'):
            updated_any |= self._do_update_one(
                kind='EKA_DATA',
                title='Select new EKA_DATA file',
                target_path=target / BUNDLED_EKA_NAME,
            )

        if updated_any:
            self._auto_load_bundled()
            messagebox.showinfo(
                "Bundled Files Updated",
                f"Bundled files updated in:\n{target}\n\n"
                f"Future runs will auto-load the new version.",
            )

    def _do_update_one(self, kind: str, title: str,
                        target_path: Path) -> bool:
        src = filedialog.askopenfilename(
            title=title,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not src:
            self._log(f"Update cancelled for {kind}", 'dim')
            return False
        try:
            shutil.copy2(src, str(target_path))
            self._log(f"Bundled {kind} updated → {target_path}", 'ok')
            return True
        except Exception as e:
            self._log(f"Update failed for {kind}: {e}", 'err')
            messagebox.showerror(
                "Update Failed",
                f"Could not copy {kind}:\n{e}",
            )
            return False

    def _lookup_location_from_filename(
        self, filepath: str,
    ) -> Tuple[Optional[Dict], int]:
        if not self.eka_locations:
            return (None, 0)

        fname = os.path.basename(filepath)
        loc_code = (fname.replace('.xlsx', '')
                          .replace('.xlsm', '')
                          .replace('.xls', ''))

        for loc in self.eka_locations:
            if loc.get('location', '') == loc_code:
                return (loc, 0)

        suffix_match = re.match(r'^(.+)_(\d+)$', loc_code)
        if suffix_match:
            base_code = suffix_match.group(1)
            suffix_idx = int(suffix_match.group(2))
            for loc in self.eka_locations:
                if loc.get('location', '') == base_code:
                    return (loc, suffix_idx)

        for loc in self.eka_locations:
            loc_val = loc.get('location', '')
            if loc_val and loc_code.startswith(loc_val):
                return (loc, 0)

        return (None, 0)

    def _run(self) -> None:
        if self.is_running:
            return
        if not self.master_path or not self.engine.master:
            messagebox.showwarning(
                "No Master",
                "Items Master is not loaded. Pick one via Browse or use "
                "'Update Bundled Files'.",
            )
            return

        mode = self.mode_var.get()
        if mode == 'special':
            self._run_special()
        else:
            self._run_standalone()

    def _run_standalone(self) -> None:
        if not self.po_files:
            messagebox.showwarning("No Files", "Please add at least one PO file.")
            return

        if not self.eka_locations:
            proceed = messagebox.askyesno(
                "EKA_DATA Not Loaded",
                "EKA_DATA is not loaded.\n\n"
                "Without EKA_DATA, these fields will be EMPTY in output:\n"
                "  • TO number\n"
                "  • Transfer-to Code\n"
                "  • Gen. Bus. Posting Group\n"
                "  • Headers (TO) sheet\n\n"
                "Continue anyway?",
            )
            if not proceed:
                return

        output_dir = (Path(self.po_files[0]).parent / 'eka_output')
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = time.strftime("%d%m%Y_%H%M%S")
        output_path = output_dir / f"PO_Output_{timestamp}.xlsx"

        self.is_running = True
        self.gen_btn.config(state=tk.DISABLED)
        self._set_status("Processing...")
        self.root.update()

        threading.Thread(
            target=self._standalone_worker,
            args=(str(output_path),),
            daemon=True,
        ).start()

    def _run_special(self) -> None:
        if not self.eka_path or not self.eka_locations:
            messagebox.showwarning(
                "No EKA_DATA",
                "EKA_DATA is required for Special Order mode.",
            )
            return
        if not self.so_path:
            messagebox.showwarning(
                "No Special Order File",
                "Please select the Special Order file.",
            )
            return

        output_dir = Path(self.so_path).parent / 'eka_output'
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = time.strftime("%d%m%Y_%H%M%S")
        output_path = output_dir / f"SO_Output_{timestamp}.xlsx"

        self.is_running = True
        self.gen_btn.config(state=tk.DISABLED)
        self._set_status("Processing Special Order...")
        self.root.update()

        threading.Thread(
            target=self._special_worker,
            args=(str(output_path),),
            daemon=True,
        ).start()

    def _standalone_worker(self, output_path: str) -> None:
        total = len(self.po_files)

        self.root.after(0, self._log, "─── PHASE 1: VALIDATING ───", 'inf')
        self.root.after(0, self._set_status, "Validating files...")

        if not self.eka_locations:
            self.root.after(
                0, self._log,
                "⚠ EKA_DATA not loaded — TO/Transfer/Posting will be empty.",
                'warn',
            )

        validation_results = {}
        files_with_errors = []
        alert_messages = []

        processing_log: List[Dict] = []

        for i, po_path in enumerate(self.po_files):
            fname = os.path.basename(po_path)
            loc = fname.replace('.xlsx', '')
            self.root.after(0, self._set_status,
                             f"Validating {i+1}/{total}: {fname[:45]}")

            log_entry = {
                'filename': fname,
                'location': loc,
                'status': 'OK',
                'issues': [],
                'actions': [],
                'to_number': '',
                'tt_number': '',
            }
            processing_log.append(log_entry)

            try:
                vlogs = self.engine.validate_file(po_path)
                has_err = any(level == 'error' for level, _ in vlogs)
                validation_results[po_path] = (vlogs, has_err)

                for level, msg in vlogs:
                    tag = {'info': 'inf', 'warn': 'warn',
                           'error': 'err', 'alert': 'warn'}.get(level, 'dim')
                    self.root.after(0, self._log, f"  [{loc}] {msg}", tag)
                    if level == 'alert':
                        alert_messages.append(f"• {loc}: {msg}")
                        log_entry['actions'].append(msg)
                        if log_entry['status'] == 'OK':
                            log_entry['status'] = 'AUTO_FIXED'
                    elif level == 'warn':
                        log_entry['issues'].append(msg)
                        if log_entry['status'] == 'OK':
                            log_entry['status'] = 'WARNING'
                    elif level == 'error':
                        log_entry['issues'].append(msg)
                        log_entry['status'] = 'FAILED'

                if has_err:
                    files_with_errors.append(fname)
                    self.root.after(0, self._log,
                                    f"✗ {loc} → VALIDATION FAILED", 'err')
                else:
                    self.root.after(0, self._log, f"✓ {loc} → OK", 'ok')
            except Exception as e:
                validation_results[po_path] = ([('error', str(e))], True)
                files_with_errors.append(fname)
                self.root.after(0, self._log, f"✗ {loc} → {e}", 'err')
                log_entry['status'] = 'FAILED'
                log_entry['issues'].append(f"Validation crashed: {e}")

        if alert_messages:
            preview = "\n".join(alert_messages[:10])
            self.root.after(0, lambda p=preview: messagebox.showinfo(
                "Auto-Fix Applied",
                f"Column names were auto-fixed:\n\n{p}",
            ))

        processable = [f for f in self.po_files
                        if not validation_results.get(f, ([], True))[1]]
        if not processable:
            self.root.after(0, self._done, None,
                             f"ABORTED — all {total} files failed validation")
            return

        skipped = total - len(processable)
        if skipped > 0:
            self.root.after(0, self._log,
                             f"⚠ Skipping {skipped} file(s) with errors", 'warn')

        self.root.after(0, self._log,
                         f"─── PHASE 2: EXTRACTING ({len(processable)} files) ───",
                         'inf')

        results: List[LocationResult] = []
        total_po = total_tester = total_pwp = total_gwp = total_ns = 0
        total_unmatched = 0

        counter = SpecialOrderEngine.get_today_date_code()

        for i, po_path in enumerate(processable):
            fname = os.path.basename(po_path)
            loc = fname.replace('.xlsx', '')
            self.root.after(0, self._set_status,
                             f"Extracting {i+1}/{len(processable)}: {fname[:45]}")

            try:
                res = self.engine.process_file(po_path)

                log_entry = next(
                    (e for e in processing_log if e['filename'] == fname),
                    None,
                )

                eka_loc, suffix_idx = self._lookup_location_from_filename(po_path)
                if eka_loc:
                    # Document split (per ops requirement):
                    #   Regular doc (month segment, e.g. /06/) = ONLY real
                    #     finished-goods PO items (regular_orders), priced
                    #     at calculated cost.
                    #   Tester doc (/TT/ segment) = every flat ₹0.54 line —
                    #     testers, PWP, GWP and Non-Stock.
                    # PWP used to ride on the regular doc; it now belongs on
                    # the TT doc so the regular order stays purely FG. See the
                    # matching item.to assignment + Summary fallbacks below.
                    has_regular = bool(res.regular_orders)
                    has_tester = bool(res.tester_orders or res.pwp_orders
                                       or res.gwp_orders or res.nonstock_orders)

                    short_code = eka_loc['short_code']
                    if suffix_idx > 0:
                        short_code = f"{short_code}_{suffix_idx + 1}"

                    to_regular = ''
                    to_tester = ''
                    if has_regular:
                        to_regular = SpecialOrderEngine.generate_to_number(
                            eka_loc['prefix'], short_code,
                            is_tester=False, date_code=counter)
                        counter += 1
                    if has_tester:
                        to_tester = SpecialOrderEngine.generate_to_number(
                            eka_loc['prefix'], short_code,
                            is_tester=True, date_code=counter)
                        counter += 1

                    tc = eka_loc['transfer_code']
                    pg = eka_loc['posting_group']

                    res._so_bill_to = eka_loc.get('bill_to', '')
                    res._so_ship_to = eka_loc.get('ship_to', '')

                    for item in res.regular_orders:
                        item.to = to_regular
                        item.transfer_to = tc
                        item.posting_group = pg
                    # PWP is a ₹0.54 line → goes on the TT (tester) doc,
                    # not the regular doc (ops requirement: regular = FG only).
                    for item in res.pwp_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg
                    for item in res.tester_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg
                    for item in res.gwp_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg
                    for item in res.nonstock_orders:
                        item.to = to_tester
                        item.transfer_to = tc
                        item.posting_group = pg

                    res.logs.append(('info',
                        f"EKA auto-fill: {eka_loc['location']} → "
                        f"TO:{to_regular or '(none)'} / "
                        f"TT:{to_tester or '(none)'} / "
                        f"Transfer:{tc} / Posting:{pg}"))

                    if log_entry is not None:
                        log_entry['to_number'] = to_regular
                        log_entry['tt_number'] = to_tester
                        if suffix_idx > 0:
                            log_entry['actions'].append(
                                f"Same location as another file — "
                                f"indexed as #{suffix_idx + 1} "
                                f"(short_code: {short_code})")
                            if log_entry['status'] == 'OK':
                                log_entry['status'] = 'AUTO_FIXED'
                else:
                    if self.eka_locations:
                        res.logs.append(('warn',
                            f"EKA: '{loc}' not found in Location column — "
                            f"TO/Transfer/Posting left empty"))
                        if log_entry is not None:
                            log_entry['issues'].append(
                                f"Location '{loc}' not in EKA_DATA — "
                                f"TO/Transfer/Posting left empty")
                            if log_entry['status'] == 'OK':
                                log_entry['status'] = 'WARNING'

                results.append(res)

                if log_entry is not None and res.unmatched:
                    log_entry['issues'].append(
                        f"{len(res.unmatched)} EAN(s) not found in master "
                        f"— see Unmatched EANs sheet")
                    if log_entry['status'] == 'OK':
                        log_entry['status'] = 'WARNING'

                po_q = sum(r.qty for r in res.regular_orders)
                tt_q = sum(r.qty for r in res.tester_orders)
                pw_q = sum(r.qty for r in res.pwp_orders)
                gw_q = sum(r.qty for r in res.gwp_orders)
                ns_q = sum(r.qty for r in res.nonstock_orders)

                total_po += po_q
                total_tester += tt_q
                total_pwp += pw_q
                total_gwp += gw_q
                total_ns += ns_q
                total_unmatched += len(res.unmatched)

                for level, msg in res.logs:
                    tag = {'info': 'inf', 'warn': 'warn',
                           'error': 'err'}.get(level, 'dim')
                    self.root.after(0, self._log, f"  [{loc}] {msg}", tag)

                parts = [f"PO:{po_q}"]
                if tt_q: parts.append(f"T:{tt_q}")
                if pw_q: parts.append(f"PWP:{pw_q}")
                if gw_q: parts.append(f"GWP:{gw_q}")
                if ns_q: parts.append(f"NS:{ns_q}")
                self.root.after(0, self._log,
                                 f"✓ {loc} → {' | '.join(parts)}", 'ok')
            except Exception as e:
                self.root.after(0, self._log, f"✗ {fname} → {e}", 'err')
                log_entry = next(
                    (e2 for e2 in processing_log if e2['filename'] == fname),
                    None,
                )
                if log_entry is not None:
                    log_entry['status'] = 'FAILED'
                    log_entry['issues'].append(f"Extraction crashed: {e}")

        if not results:
            self.root.after(0, self._done, None, "ERROR: No data processed")
            return

        try:
            self.root.after(0, self._set_status, "Writing Excel...")
            self.last_results = results
            ExcelWriter.write(results, output_path,
                               processing_log=processing_log)

            grand = total_po + total_tester + total_pwp + total_gwp + total_ns

            stats_line = (
                f"Locations:{len(results)} | PO Qty:{total_po:,} | "
                f"PO Items:{sum(len(r.regular_orders) for r in results)} | "
                f"Tester:{total_tester:,} | PWP:{total_pwp:,} | "
                f"GWP:{total_gwp:,} | NS:{total_ns:,} | "
                f"GRAND TOTAL:{grand:,} | Unmatched:{total_unmatched}"
            )
            self.root.after(0, self._log, stats_line, 'ok')

            self.root.after(0, self._done, output_path,
                             f"Done — {len(results)} locations, {grand:,} total qty")
        except Exception as e:
            self.root.after(0, self._done, None, f"ERROR: {e}")

    def _special_worker(self, output_path: str) -> None:
        self.root.after(0, self._log, "═══ SPECIAL ORDER ═══", 'inf')

        so_engine = SpecialOrderEngine(self.engine.master)

        logs = []
        loc_count = so_engine.load_eka_data(self.eka_path, logs)
        for level, msg in logs:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err', 'alert': 'warn'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)

        if loc_count == 0:
            self.root.after(0, self._done, None, "ABORTED — EKA_DATA load failed")
            return

        self.root.after(0, self._set_status, "Loading Special Order...")
        logs2 = []
        alert_messages = []
        prod_count = so_engine.load_special_order(self.so_path, logs2)
        for level, msg in logs2:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err', 'alert': 'warn'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)
            if level == 'alert':
                alert_messages.append(msg)

        if prod_count == 0:
            self.root.after(0, self._done, None,
                             "ABORTED — Special Order load failed")
            return

        if alert_messages:
            preview = "\n".join(f"• {m}" for m in alert_messages)
            self.root.after(0, lambda p=preview: messagebox.showinfo(
                "Auto-Fix Applied",
                f"Column names auto-fixed:\n\n{p}",
            ))

        self.root.after(0, self._set_status, "Validating...")
        logs3 = []
        if not so_engine.validate(logs3):
            for level, msg in logs3:
                tag = {'info': 'inf', 'warn': 'warn',
                       'error': 'err'}.get(level, 'dim')
                self.root.after(0, self._log, f"  {msg}", tag)
            self.root.after(0, self._done, None, "ABORTED — Validation failed")
            return

        for level, msg in logs3:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)

        self.root.after(0, self._log, "─── GENERATING OUTPUT ───", 'inf')
        self.root.after(0, self._set_status, "Generating output...")

        logs4 = []
        results = so_engine.process(logs4)
        for level, msg in logs4:
            tag = {'info': 'inf', 'warn': 'warn',
                   'error': 'err'}.get(level, 'dim')
            self.root.after(0, self._log, f"  {msg}", tag)

        if not results:
            self.root.after(0, self._done, None, "ERROR — No results generated")
            return

        total_po = total_tester = total_unmatched = 0

        processing_log: List[Dict] = []

        for res in results:
            po_q = sum(r.qty for r in res.regular_orders)
            tt_q = sum(r.qty for r in res.tester_orders)
            total_po += po_q
            total_tester += tt_q
            total_unmatched += len(res.unmatched)

            to_reg = res.regular_orders[0].to if res.regular_orders else ''
            to_tt = res.tester_orders[0].to if res.tester_orders else ''
            self.root.after(0, self._log,
                             f"  ✓ {res.filename:<22} PO:{po_q} T:{tt_q}  [{to_reg}]",
                             'ok')

            entry = {
                'filename': res.filename,
                'location': res.filename,
                'status': 'OK',
                'issues': [],
                'actions': [],
                'to_number': to_reg,
                'tt_number': to_tt,
            }
            if res.unmatched:
                entry['issues'].append(
                    f"{len(res.unmatched)} EAN(s) not found in master "
                    f"— see Unmatched EANs sheet")
                entry['status'] = 'WARNING'
            processing_log.append(entry)

        try:
            self.root.after(0, self._set_status, "Writing Excel...")
            self.last_results = results
            ExcelWriter.write(
                results, output_path,
                eka_locations=so_engine.locations,
                master=self.engine.master,
                so_products=so_engine.products,
                processing_log=processing_log,
            )

            grand = total_po + total_tester
            stats_line = (
                f"Locations:{len(results)} | PO Qty:{total_po:,} | "
                f"PO Items:{sum(len(r.regular_orders) for r in results)} | "
                f"Tester:{total_tester:,} | PWP:0 | GWP:0 | NS:0 | "
                f"GRAND TOTAL:{grand:,} | Unmatched:{total_unmatched}"
            )
            self.root.after(0, self._log, stats_line, 'ok')

            self.root.after(
                0, self._done, output_path,
                f"Done — {len(results)} locations × {prod_count} products, "
                f"{grand:,} total qty",
            )
        except Exception as e:
            self.root.after(0, self._done, None, f"ERROR: {e}")

    def _done(self, output_path: Optional[str], status_msg: str) -> None:
        self.is_running = False
        self.gen_btn.config(state=tk.NORMAL)
        self._set_status(status_msg)

        if output_path:
            self.last_output = Path(output_path)
            self.open_btn.config(state=tk.NORMAL)
            self.d365_btn.config(state=tk.NORMAL)
            self._log(f"Saved → {output_path}", 'inf')

            if messagebox.askyesno(
                    "Done!",
                    f"Processing complete!\n\n{status_msg}\n\nOpen output file?"):
                self._open_last()

    def _open_last(self) -> None:
        if not self.last_output or not self.last_output.exists():
            messagebox.showwarning("Not Found", "No output file yet.")
            return
        try:
            if os.name == 'nt':
                os.startfile(str(self.last_output))
            elif os.name == 'posix':
                import subprocess
                opener = ('open' if hasattr(os, 'uname') and
                          os.uname().sysname == 'Darwin' else 'xdg-open')
                subprocess.Popen([opener, str(self.last_output)])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}")

    def _export_d365(self) -> None:
        if not self.last_results:
            messagebox.showwarning(
                "No Data",
                "Run processing first to generate data for D365 export.")
            return

        has_to = False
        has_so = False
        for res in self.last_results:
            for item in (res.regular_orders + res.tester_orders +
                         res.pwp_orders + res.gwp_orders +
                         res.nonstock_orders):
                if item.to:
                    if item.to.startswith('TO/'):
                        has_to = True
                    elif item.to.startswith('SO/'):
                        has_so = True
                if has_to and has_so:
                    break
            if has_to and has_so:
                break

        if not has_to and not has_so:
            self._log("D365 export: no TO or SO numbers in last batch — "
                      "nothing to export.", 'warn')
            messagebox.showwarning(
                "Nothing To Export",
                "Last batch has no TO or SO numbers.\n\n"
                "TO numbers come from EKA_DATA rows where Prefix = 'TO'.\n"
                "SO numbers come from rows where Prefix = 'SO'.\n\n"
                "Check that EKA_DATA was loaded and that the matching "
                "filenames or location types resolved correctly.")
            return

        if self.po_files:
            out_dir = Path(self.po_files[0]).parent / 'eka_output'
        elif self.so_path:
            out_dir = Path(self.so_path).parent / 'eka_output'
        else:
            out_dir = get_script_dir() / 'eka_output'
        out_dir.mkdir(parents=True, exist_ok=True)

        ts = time.strftime('%d%m%Y_%H%M%S')

        to_template: Optional[str] = None
        if has_to:
            to_template = filedialog.askopenfilename(
                title="Select D365 TO Template (EKA_Sample_Package_TO.xlsx)",
                filetypes=[("Excel files", "*.xlsx")],
            )
            if not to_template:
                self._log("D365 export cancelled — no TO template selected.",
                          'warn')
                return

        so_template: Optional[str] = None
        if has_so:
            so_template = filedialog.askopenfilename(
                title="Select D365 SO Template (EKA_Sample_Package_SO.xlsx)",
                filetypes=[("Excel files", "*.xlsx")],
            )
            if not so_template:
                self._log("D365 export: SO template skipped — only TO will "
                          "be generated.", 'warn')
                so_template = None

        produced: List[Path] = []

        if has_to and to_template:
            try:
                to_path = out_dir / f"D365_TO_Package_{ts}.xlsx"
                D365TOExporter.export(
                    self.last_results, to_template, str(to_path))
                self._log(f"D365 TO package saved → {to_path}", 'ok')
                produced.append(to_path)
            except Exception as e:
                self._log(f"D365 TO export failed: {e}", 'err')
                messagebox.showerror(
                    "D365 TO Export Failed",
                    f"Could not produce the TO package:\n\n{e}")

        if has_so and so_template:
            try:
                so_path = out_dir / f"D365_SO_Package_{ts}.xlsx"
                D365SOExporter.export(
                    self.last_results, so_template, str(so_path))
                self._log(f"D365 SO package saved → {so_path}", 'ok')
                produced.append(so_path)
            except Exception as e:
                self._log(f"D365 SO export failed: {e}", 'err')
                messagebox.showerror(
                    "D365 SO Export Failed",
                    f"Could not produce the SO package:\n\n{e}")

        if not produced:
            return

        files_list = "\n".join(f"  • {p.name}" for p in produced)
        if messagebox.askyesno(
                "D365 Exported",
                f"Generated {len(produced)} D365 package(s):\n\n"
                f"{files_list}\n\n"
                f"Open the {'first ' if len(produced) > 1 else ''}file?"):
            target = produced[0]
            if os.name == 'nt':
                os.startfile(str(target))
            elif os.name == 'posix':
                import subprocess
                opener = ('open' if hasattr(os, 'uname') and
                          os.uname().sysname == 'Darwin' else 'xdg-open')
                subprocess.Popen([opener, str(target)])

    def _download_po_template(self) -> None:
        save_path = filedialog.asksaveasfilename(
            title="Save Blank PO Template",
            defaultextension=".xlsx",
            initialfile="PO_Template_Blank.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return
        try:
            wb = Workbook()
            wb.remove(wb.active)
            hdr_fill = PatternFill('solid', fgColor='1A237E')
            hdr_font = Font(bold=True, color='FFFFFF',
                             name='Aptos Display', size=11)

            def make_header(ws, headers, widths=None):
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c, value=h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal='center')
                if widths:
                    for c, w in enumerate(widths, 1):
                        ws.column_dimensions[get_column_letter(c)].width = w
                ws.freeze_panes = 'A2'

            ws = wb.create_sheet('PO')
            make_header(ws,
                ['Rank', 'Category', 'EAN', 'SKU Code', 'Product Name',
                 'Brand', 'MRP', 'Available', 'Order Qty', 'Tester Qty'],
                [8, 20, 18, 14, 50, 14, 10, 12, 12, 12])
            sample = [1, 'Eyes', '8906121646979', '06D19087',
                      'SAMPLE PRODUCT NAME', 'RENEE', 450, 0, '', '']
            for c, v in enumerate(sample, 1):
                cell = ws.cell(row=2, column=c, value=v)
                cell.font = Font(name='Aptos Display', size=11,
                                  color='999999', italic=True)
            ws.cell(row=3, column=1,
                value='← Delete sample row. Fill EAN, Order Qty, '
                      'Tester Qty.').font = Font(
                          name='Aptos Display', size=11,
                          color='FF6600', italic=True)

            ws = wb.create_sheet('PWP')
            make_header(ws, ['Sr. No.', 'Product Name', 'Avail.Qty', 'Req.Qty'],
                          [10, 30, 12, 12])
            # v1.5.8: include the bollywood spray as a pre-filled sample
            # so stores know the exact name format to use. PWP_EAN_MAP
            # uses whitespace normalization so non-breaking spaces in
            # the source file still match.
            for row_data in [
                (1, 'Stay With Me - Mini', '', ''),
                (2, 'Perfume', '', ''),
                (3, 'Crème Mini', '', ''),
                (4, 'RENEE BOLLYWOOD FILTER BLURRING SETTING SPRAY 5ML (NFS)',
                 '', ''),
            ]:
                ws.append(row_data)
            ws.append(('Total', None, 0, 0))

            ws = wb.create_sheet('GWP')
            make_header(ws,
                ['Sr. No.', 'EAN', 'Product Name', 'Avail.Qty', 'Req.Qty'],
                [10, 18, 45, 12, 12])
            gwp_items = [
                (1, 8904473101658, 'RENEE Lunar Luxe Trousseau box – Silver', '', ''),
                (2, 8904473101672, 'RENEE Red Velvet Trousseau Box - Red', '', ''),
                (3, 8904473101665, 'RENEE Rose Glow Trousseau Box - Pink', '', ''),
                (4, 8904473101009, 'RENEE Pink Puffer Pouch', '', ''),
                (5, 8904473101023, 'RENEE Red Puffer Pouch', '', ''),
                (6, 8904473101016, 'RENEE Silver Puffer Pouch', '', ''),
            ]
            for row_data in gwp_items:
                ws.append(row_data)
            ws.append(('Total', None, None, None, 0))

            ws = wb.create_sheet('Non Stock')
            make_header(ws, ['Sr. No.', 'Product Name', 'QTY'],
                          [10, 30, 10])
            ns_items = [
                (1, 'Cotton Rolls'), (2, 'Mirrors'), (3, 'Carry Bag (Small)'),
                (4, 'Carry Bag (Big)'),
                (5, 'Renee Multicolor Bag (Big)'),
                (6, 'Renee Multicolor Bag (Small)'),
                (7, 'Cleansers'), (8, 'Calculator'),
                (9, 'Blotters'), (10, 'Swabs'), (11, 'Bill Roll'),
                (12, 'Renee Notebook'), (13, 'Pen'),
            ]
            for sr, name in ns_items:
                ws.append((sr, name, ''))
            ws.append(('Total', None, 0))

            ws = wb.create_sheet('Summary')
            ws.cell(row=3, column=2, value='[Location Name]')
            for c, h in enumerate(
                ['PO', 'Tester', 'PWP', 'GWP',
                 'Non-Stock Requirement', 'Total'], 6):
                ws.cell(row=3, column=c, value=h)

            wb.save(save_path)
            self._log(f"PO template saved → {save_path}", 'ok')
            messagebox.showinfo(
                "Template Saved",
                f"Blank PO template saved to:\n{save_path}\n\n"
                "Sheets: PO, PWP, GWP, Non Stock, Summary",
            )
        except Exception as e:
            self._log(f"Template save failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def _download_eka_template(self) -> None:
        save_path = filedialog.asksaveasfilename(
            title="Save EKA_DATA Template", defaultextension=".xlsx",
            initialfile="EKA_DATA_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'EKA_DATA'

            headers = ['Desc', 'Bill to', 'Ship to', 'Location',
                        'Gen. Biz.  Posting Group', 'Short Name',
                        'Prefix', 'Short Code', 'Transfer Code', 'Type',
                        'Example Regular', 'Example Tester', 'Status']
            widths = [40, 10, 12, 15, 22, 22, 8, 12, 15, 10, 25, 25, 10]

            hdr_fill = PatternFill('solid', fgColor='E65100')
            hdr_font = Font(bold=True, color='FFFFFF',
                             name='Aptos Display', size=11)

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

            samples = [
                ('RENEE COSMETICS-ISCON ARCADE', '20329', '20329_1', 'EBO_AMD01',
                 'OFF-EBO', 'Ahmedabad EBO', 'TO', 'AHDEB', 'EBO_AMD01', 'EBO',
                 'TO/AHDEB/04/18426', 'TO/AHDEB/TT/18427', 'Active'),
                ('RENEE COSMETICS-CHENNAI AIRPORT', '20342', '20342_1', 'AP_CHEN01',
                 'OFF-AIRPORT', 'Chennai Airport', 'TO', 'CHNAP', 'AP_CHEN01', 'Airport',
                 'TO/CHNAP/04/18426', 'TO/CHNAP/TT/18427', 'Active'),
                ('OG BEAUTY PRIVATE LIMITED', '20395', '20395_1', 'EBO_PUNE02',
                 'OFF-EBO', 'Pune EBO', 'SO', 'PUNEB', '20395_1', 'EBO',
                 'SO/PUNEB/04/18426', 'SO/PUNEB/TT/18427', 'Active'),
            ]
            sample_font = Font(name='Aptos Display', size=11,
                                color='666666', italic=True)
            for r, row_data in enumerate(samples, 2):
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = sample_font

            ws.cell(row=len(samples) + 2, column=1,
                value='← Delete sample rows. Add one row per location. '
                      'Prefix: TO for Transfer Order, SO for Sales Order. '
                      'Status: Active or Inactive.').font = Font(
                          name='Aptos Display', size=11,
                          color='FF6600', italic=True)
            ws.freeze_panes = 'A2'

            wb.save(save_path)
            self._log(f"EKA template saved → {save_path}", 'ok')
            messagebox.showinfo(
                "Template Saved",
                f"EKA_DATA template saved to:\n{save_path}",
            )
        except Exception as e:
            self._log(f"EKA template save failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def _download_so_template(self) -> None:
        save_path = filedialog.asksaveasfilename(
            title="Save Special Order Template", defaultextension=".xlsx",
            initialfile="Special_Order_Template.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Special Order'

            headers = ['Description', 'EAN', 'Item Category Code', 'MRP',
                        'EBO Qty', 'Airport Qty', 'Kiosk Qty', 'Tester Qty']
            widths = [45, 18, 18, 10, 12, 12, 12, 12]

            hdr_fill = PatternFill('solid', fgColor='1A237E')
            hdr_font = Font(bold=True, color='FFFFFF',
                             name='Aptos Display', size=11)

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

            samples = [
                ('RENEE PRO HD 3-IN-1 - AMANDE_9 GM', '8906121648515',
                 'POWDER', 650, 18, 18, 10, 1),
                ('RENEE PRO HD CONCEALER - BUFF_8 ML', '8906121648317',
                 'CONCEALER', 750, 18, 18, 10, 1),
                ('RENEE PAPER BAG BIG MULTICOLOR', '8904473105984',
                 'PAPER BAG', 750, '', '', '', 50),
            ]
            sample_font = Font(name='Aptos Display', size=11,
                                color='666666', italic=True)
            for r, row_data in enumerate(samples, 2):
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = sample_font

            ws.cell(row=len(samples) + 2, column=1,
                value='← Delete sample rows. EAN must match GTIN in '
                      'Items_March. Leave qty blank or 0 if not '
                      'applicable. Tester Qty goes to ALL locations '
                      'at ₹0.54.').font = Font(
                          name='Aptos Display', size=11,
                          color='FF6600', italic=True)
            ws.freeze_panes = 'A2'

            wb.save(save_path)
            self._log(f"SO template saved → {save_path}", 'ok')
            messagebox.showinfo(
                "Template Saved",
                f"Special Order template saved to:\n{save_path}",
            )
        except Exception as e:
            self._log(f"SO template save failed: {e}", 'err')
            messagebox.showerror("Error", f"Failed:\n{e}")

    def run(self) -> None:
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    """Application entry point."""
    print(f"{APP_TITLE} starting...")
    print(f"  Script: {get_script_dir()}")
    print(f"  Bundled folder: {get_bundled_folder()}")
    print("  Initializing GUI window...")
    sys.stdout.flush()

    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            except Exception:
                ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

    try:
        app = ReneePOApp()
        print("  GUI window ready. If you don't see it, check Alt+Tab.")
        sys.stdout.flush()
        app.run()
    except Exception as e:
        print(f"\n  FATAL ERROR during startup: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        sys.stdout.flush()
        input("\n  Press Enter to close...")
        raise


if __name__ == '__main__':
    main()