"""
Blinkit GRN PDF Parser — Tkinter GUI
=====================================
A desktop app to extract SKU-level GRN data from Blinkit GRN PDFs.

Requirements:
    pip install pdfplumber pandas openpyxl

Run:
    python blinkit_grn_gui.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import glob
import re
import time

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── COLORS & FONTS ────────────────────────────────────────────────────────────
BG          = "#0F1117"
SURFACE     = "#1A1D27"
SURFACE2    = "#22263A"
ACCENT      = "#00D4FF"
ACCENT2     = "#7B61FF"
GREEN       = "#00E676"
RED         = "#FF5252"
AMBER       = "#FFB300"
TEXT        = "#E8EAF6"
TEXT_DIM    = "#6B7280"
BORDER      = "#2D3250"

FONT_TITLE  = ("Courier New", 18, "bold")
FONT_SUB    = ("Courier New", 10)
FONT_LABEL  = ("Courier New", 9, "bold")
FONT_MONO   = ("Courier New", 9)
FONT_BTN    = ("Courier New", 10, "bold")
FONT_BIG    = ("Courier New", 28, "bold")


# ─── PDF PARSING LOGIC ─────────────────────────────────────────────────────────

def clean_upc(raw):
    return re.sub(r'\s+', '', str(raw))

def clean_number(val):
    if val is None or str(val).strip() in ('-', '', 'None'):
        return 0.0
    try:
        return float(str(val).replace('\n', '').replace(',', '').strip())
    except:
        return 0.0

def extract_header_info(pdf):
    text = pdf.pages[0].extract_text() or ""
    po  = re.search(r'P\.O\.\s*Number\s*[:\s]+(\d+)', text)
    dt  = re.search(r'Date\s*[:\s]+([\w.]+\s+\d+,\s+\d{4})', text)
    fac = re.search(r'BCPL\s*-\s*(.+?)(?:\n|Contact)', text)
    return {
        'PO Number': po.group(1).strip()  if po  else 'UNKNOWN',
        'PO Date':   dt.group(1).strip()  if dt  else '',
        'Facility':  fac.group(1).strip() if fac else '',
    }

def extract_summary(all_text):
    def find(pattern):
        m = re.search(pattern, all_text, re.IGNORECASE)
        return clean_number(m.group(1)) if m else 0.0
    fr = re.search(r'Fill rate:\s*([\d.]+)%', all_text)
    return {
        'Total PO Qty':    find(r'Total Quantity in PO:\s*([\d,]+)'),
        'Total GRN Qty':   find(r'Total Quantity in GRN\(s\):\s*([\d,]+)'),
        'Fill Rate %':     float(fr.group(1)) if fr else 0.0,
        'Articles in PO':  find(r'Articles in PO:\s*([\d,]+)'),
        'Articles in GRN': find(r'Articles in GRN\(s\):\s*([\d,]+)'),
        'Total PO Amount': find(r'Total Amount in PO\s+([\d,\.]+)'),
        'Net GRN Amount':  find(r'Net amt\. by GRN\s+([\d,\.]+)'),
        'GMV Loss':        find(r'Potential GMV Loss \(in INR\)\s+([\d,\.]+)'),
    }

def is_item_row(row):
    if not row or row[0] is None:
        return False
    return bool(re.match(r'^\d+$', str(row[0]).strip()))

def parse_grn_pdf(pdf_path):
    rows = []
    all_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        header = extract_header_info(pdf)
        for page in pdf.pages:
            all_text += (page.extract_text() or "") + "\n"
            for table in page.extract_tables():
                for row in table:
                    if not is_item_row(row):
                        continue
                    try:
                        upc     = clean_upc(row[2] if len(row) > 2 else '')
                        desc    = str(row[3] or '').replace('\n', ' ').strip()
                        po_qty  = int(clean_number(row[8]))  if len(row) > 8  else 0
                        grn_qty = int(clean_number(row[9]))  if len(row) > 9  else 0
                        mrp     = clean_number(row[4])        if len(row) > 4  else 0.0
                        lr      = clean_number(row[6])        if len(row) > 6  else 0.0
                        fr_raw  = str(row[10] or '').strip()  if len(row) > 10 else '-'
                        fr      = clean_number(fr_raw) if fr_raw != '-' else 0.0
                        grn_amt = clean_number(row[11]) if len(row) > 11 else 0.0
                        gmv     = clean_number(row[12]) if len(row) > 12 else 0.0

                        if grn_qty == 0:
                            status = 'Not GRNed'
                        elif grn_qty < po_qty:
                            status = 'Partial GRN'
                        else:
                            status = 'Full GRN'

                        rows.append({
                            'PO Number':       header['PO Number'],
                            'PO Date':         header['PO Date'],
                            'Facility':        header['Facility'],
                            'Sr No':           int(row[0]),
                            'Item Code':       str(row[1] or '').strip(),
                            'UPC / GTIN':      upc,
                            'Description':     desc,
                            'MRP':             mrp,
                            'Landing Rate':    lr,
                            'PO Qty':          po_qty,
                            'GRN Qty':         grn_qty,
                            'Fill Rate %':     fr,
                            'GRN Amount':      grn_amt,
                            'GMV Loss':        gmv,
                            'Line GRN Status': status,
                            'PO<>EAN':         f"{header['PO Number']}<>{upc}",
                        })
                    except:
                        pass

    summary = extract_summary(all_text)
    header.update(summary)
    return header, pd.DataFrame(rows)

def format_excel(output_path):
    wb = load_workbook(output_path)
    status_colors = {'Full GRN': '00C853', 'Partial GRN': 'FFB300', 'Not GRNed': 'D50000'}
    hfill  = PatternFill('solid', start_color='1A237E')
    hfont  = Font(bold=True, color='FFFFFF', name='Courier New', size=9)
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb['GRN Line Items']
    sc = None
    for cell in ws[1]:
        if cell.value == 'Line GRN Status':
            sc = cell.column
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sv = row[sc - 1].value if sc else ''
        for cell in row:
            cell.border = border
            cell.font   = Font(name='Courier New', size=9)
            cell.alignment = Alignment(vertical='center')
            if sc and cell.column == sc:
                c = status_colors.get(str(sv), 'FFFFFF')
                cell.fill = PatternFill('solid', start_color=c)
                cell.font = Font(name='Courier New', size=9, bold=True,
                                 color='FFFFFF' if sv == 'Not GRNed' else '000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')

    widths = [16,10,8,10,14,16,52,10,14,10,10,12,14,12,16,28]
    for i, w in enumerate(widths[:ws.max_column], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb['PO Summary']
    for cell in ws2[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws2.row_dimensions[1].height = 28
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = border
            cell.font   = Font(name='Courier New', size=9)
    ws2.freeze_panes = 'A2'
    for i in range(1, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 20

    wb.save(output_path)


# ─── GUI APPLICATION ───────────────────────────────────────────────────────────

class BlinkitGRNApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Blinkit GRN Parser")
        self.geometry("920x680")
        self.resizable(True, True)
        self.configure(bg=BG)
        self.minsize(780, 560)

        self.pdf_files   = []
        self.last_output = None   # path of most recently saved file
        self.is_running  = False

        self._build_ui()

    # ── UI CONSTRUCTION ────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header bar ──
        hdr = tk.Frame(self, bg=SURFACE, height=60)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)

        tk.Label(hdr, text="▶ BLINKIT GRN PARSER", font=FONT_TITLE,
                 bg=SURFACE, fg=ACCENT).pack(side='left', padx=20, pady=12)
        tk.Label(hdr, text="PDF → Excel  //  SKU-level GRN extraction",
                 font=FONT_SUB, bg=SURFACE, fg=TEXT_DIM).pack(side='left', padx=4)

        # ── Main content (2 columns) ──
        body = tk.Frame(self, bg=BG)
        body.pack(fill='both', expand=True, padx=16, pady=12)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        left  = tk.Frame(body, bg=BG)
        right = tk.Frame(body, bg=BG)
        left.grid(row=0, column=0, sticky='nsew', padx=(0, 8))
        right.grid(row=0, column=1, sticky='nsew')

        self._build_left(left)
        self._build_right(right)

        # ── Bottom bar ──
        self._build_bottom()

    def _build_left(self, parent):
        # Section: PDF Files
        self._section(parent, "01  //  PDF FILES")

        # Drop zone
        dz = tk.Frame(parent, bg=SURFACE2, relief='flat', bd=0,
                      highlightthickness=1, highlightbackground=BORDER)
        dz.pack(fill='x', pady=(0, 8))
        dz_inner = tk.Frame(dz, bg=SURFACE2)
        dz_inner.pack(fill='x', padx=1, pady=1)

        tk.Label(dz_inner, text="DROP ZONE", font=("Courier New", 9, "bold"),
                 bg=SURFACE2, fg=ACCENT2).pack(pady=(10, 2))
        tk.Label(dz_inner, text="Add GRN PDFs from Blinkit portal",
                 font=FONT_MONO, bg=SURFACE2, fg=TEXT_DIM).pack(pady=(0, 10))

        btn_row = tk.Frame(dz_inner, bg=SURFACE2)
        btn_row.pack(pady=(0, 10))
        self._btn(btn_row, "+ ADD FILES", self._add_files, ACCENT).pack(side='left', padx=4)
        self._btn(btn_row, "+ ADD FOLDER", self._add_folder, ACCENT2).pack(side='left', padx=4)
        self._btn(btn_row, "✕ CLEAR ALL", self._clear_files, RED).pack(side='left', padx=4)

        # File list
        list_frame = tk.Frame(parent, bg=SURFACE,
                              highlightthickness=1, highlightbackground=BORDER)
        list_frame.pack(fill='both', expand=True, pady=(0, 8))

        list_hdr = tk.Frame(list_frame, bg=SURFACE2)
        list_hdr.pack(fill='x')
        tk.Label(list_hdr, text="  FILE", font=FONT_LABEL, bg=SURFACE2, fg=TEXT_DIM,
                 width=38, anchor='w').pack(side='left', padx=4, pady=4)
        tk.Label(list_hdr, text="STATUS", font=FONT_LABEL, bg=SURFACE2, fg=TEXT_DIM).pack(side='right', padx=12, pady=4)

        scroll_y = ttk.Scrollbar(list_frame, orient='vertical')
        scroll_y.pack(side='right', fill='y')
        self.file_list = tk.Listbox(list_frame, bg=SURFACE, fg=TEXT, font=FONT_MONO,
                                    selectbackground=SURFACE2, selectforeground=ACCENT,
                                    borderwidth=0, highlightthickness=0,
                                    yscrollcommand=scroll_y.set, activestyle='none')
        self.file_list.pack(fill='both', expand=True, padx=4, pady=4)
        scroll_y.config(command=self.file_list.yview)

        # File count badge
        self.file_count_var = tk.StringVar(value="0 files loaded")
        tk.Label(parent, textvariable=self.file_count_var,
                 font=FONT_MONO, bg=BG, fg=TEXT_DIM).pack(anchor='w')

    def _build_right(self, parent):
        # Section: Output
        self._section(parent, "02  //  OUTPUT")

        out_frame = tk.Frame(parent, bg=SURFACE,
                             highlightthickness=1, highlightbackground=BORDER)
        out_frame.pack(fill='x', pady=(0, 12))

        tk.Label(out_frame, text="Output folder:", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10, pady=(8, 2))
        tk.Label(out_frame, text="  output_grn/  (auto-created next to this script)",
                 font=FONT_MONO, bg=SURFACE, fg=ACCENT2).pack(anchor='w', padx=10)
        tk.Label(out_frame, text="Filename:", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10, pady=(6, 2))
        tk.Label(out_frame, text="  ddmmyyyy_hhmmss.xlsx  (stamped at run time)",
                 font=FONT_MONO, bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10)

        self.last_path_var = tk.StringVar(value="No run yet")
        tk.Label(out_frame, text="Last saved:", font=FONT_LABEL,
                 bg=SURFACE, fg=TEXT_DIM).pack(anchor='w', padx=10, pady=(6, 2))
        tk.Label(out_frame, textvariable=self.last_path_var,
                 font=FONT_MONO, bg=SURFACE, fg=GREEN,
                 wraplength=280, justify='left').pack(anchor='w', padx=10, pady=(0, 10))

        # Section: Stats
        self._section(parent, "03  //  LAST RUN STATS")
        stats_frame = tk.Frame(parent, bg=SURFACE,
                               highlightthickness=1, highlightbackground=BORDER)
        stats_frame.pack(fill='x', pady=(0, 12))

        self.stat_vars = {}
        stats = [
            ("POs Processed", "pos"),
            ("Total SKUs",    "skus"),
            ("Full GRN",      "full"),
            ("Partial GRN",   "partial"),
            ("Not GRNed",     "not_grn"),
            ("Total GRN Qty", "grn_qty"),
        ]
        stat_colors = {
            "full": GREEN, "partial": AMBER, "not_grn": RED,
            "pos": ACCENT, "skus": ACCENT2, "grn_qty": TEXT
        }
        for i, (label, key) in enumerate(stats):
            row = tk.Frame(stats_frame, bg=SURFACE2 if i % 2 == 0 else SURFACE)
            row.pack(fill='x')
            tk.Label(row, text=f"  {label}", font=FONT_MONO,
                     bg=row['bg'], fg=TEXT_DIM, width=18, anchor='w').pack(side='left', pady=4, padx=4)
            var = tk.StringVar(value="—")
            self.stat_vars[key] = var
            tk.Label(row, textvariable=var, font=("Courier New", 11, "bold"),
                     bg=row['bg'], fg=stat_colors.get(key, TEXT)).pack(side='right', padx=12, pady=4)

        # Section: Log
        self._section(parent, "04  //  LOG")
        log_frame = tk.Frame(parent, bg=SURFACE,
                             highlightthickness=1, highlightbackground=BORDER)
        log_frame.pack(fill='both', expand=True)
        scroll_log = ttk.Scrollbar(log_frame, orient='vertical')
        scroll_log.pack(side='right', fill='y')
        self.log_text = tk.Text(log_frame, bg=SURFACE, fg=TEXT_DIM, font=FONT_MONO,
                                height=6, wrap='word', state='disabled',
                                borderwidth=0, highlightthickness=0,
                                yscrollcommand=scroll_log.set)
        self.log_text.pack(fill='both', expand=True, padx=6, pady=6)
        scroll_log.config(command=self.log_text.yview)
        self.log_text.tag_config('ok',  foreground=GREEN)
        self.log_text.tag_config('err', foreground=RED)
        self.log_text.tag_config('inf', foreground=ACCENT)
        self.log_text.tag_config('dim', foreground=TEXT_DIM)

    def _build_bottom(self):
        bottom = tk.Frame(self, bg=SURFACE, height=64)
        bottom.pack(fill='x', side='bottom')
        bottom.pack_propagate(False)

        # Progress bar (custom)
        pb_frame = tk.Frame(bottom, bg=SURFACE)
        pb_frame.pack(fill='x', padx=16, pady=(8, 0))

        self.progress_canvas = tk.Canvas(pb_frame, height=4, bg=SURFACE2,
                                         highlightthickness=0)
        self.progress_canvas.pack(fill='x')
        self.progress_bar_id = None
        self.progress_value  = 0

        ctrl = tk.Frame(bottom, bg=SURFACE)
        ctrl.pack(fill='x', padx=16, pady=(4, 8))

        self.status_label = tk.Label(ctrl, text="READY  //  Add PDFs to begin",
                                     font=FONT_MONO, bg=SURFACE, fg=TEXT_DIM)
        self.status_label.pack(side='left')

        self.run_btn = self._btn(ctrl, "▶  EXTRACT GRN DATA", self._run, ACCENT, large=True)
        self.run_btn.pack(side='right')

        self._btn(ctrl, "📂 OPEN OUTPUT", self._open_output, TEXT_DIM).pack(side='right', padx=8)

    # ── HELPERS ────────────────────────────────────────────────────────────────

    def _section(self, parent, title):
        f = tk.Frame(parent, bg=BG)
        f.pack(fill='x', pady=(6, 4))
        tk.Label(f, text=title, font=FONT_LABEL, bg=BG, fg=ACCENT).pack(side='left')
        tk.Frame(f, bg=BORDER, height=1).pack(side='left', fill='x', expand=True, padx=8)

    def _btn(self, parent, text, cmd, color, large=False):
        font = ("Courier New", 10, "bold") if large else ("Courier New", 9, "bold")
        padx = 16 if large else 10
        pady = 6  if large else 4
        btn = tk.Label(parent, text=text, font=font, bg=SURFACE2, fg=color,
                       cursor='hand2', padx=padx, pady=pady,
                       relief='flat', bd=0,
                       highlightthickness=1, highlightbackground=color)
        btn.bind('<Button-1>', lambda e: cmd())
        btn.bind('<Enter>', lambda e: btn.config(bg=color, fg=BG))
        btn.bind('<Leave>', lambda e: btn.config(bg=SURFACE2, fg=color))
        return btn

    def _log(self, msg, tag='dim'):
        self.log_text.config(state='normal')
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert('end', f"[{ts}] {msg}\n", tag)
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _set_status(self, msg, color=TEXT_DIM):
        self.status_label.config(text=msg, fg=color)

    def _set_progress(self, pct):
        self.progress_canvas.update_idletasks()
        w = self.progress_canvas.winfo_width()
        h = 4
        self.progress_canvas.delete('all')
        self.progress_canvas.create_rectangle(0, 0, w, h, fill=SURFACE2, outline='')
        if pct > 0:
            bar_w = int(w * pct / 100)
            self.progress_canvas.create_rectangle(0, 0, bar_w, h, fill=ACCENT, outline='')

    def _refresh_file_list(self):
        self.file_list.delete(0, 'end')
        for p in self.pdf_files:
            name = os.path.basename(p)
            # Truncate long names
            display = name if len(name) <= 42 else name[:39] + '...'
            self.file_list.insert('end', f"  {display}")
        self.file_count_var.set(f"{len(self.pdf_files)} file(s) loaded")

    def _update_stats(self, df_all):
        if df_all is None or len(df_all) == 0:
            return
        vc = df_all['Line GRN Status'].value_counts()
        pos  = df_all['PO Number'].nunique()
        self.stat_vars['pos'].set(str(pos))
        self.stat_vars['skus'].set(str(len(df_all)))
        self.stat_vars['full'].set(str(vc.get('Full GRN', 0)))
        self.stat_vars['partial'].set(str(vc.get('Partial GRN', 0)))
        self.stat_vars['not_grn'].set(str(vc.get('Not GRNed', 0)))
        self.stat_vars['grn_qty'].set(f"{int(df_all['GRN Qty'].sum()):,}")

    # ── ACTIONS ────────────────────────────────────────────────────────────────

    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Select GRN PDFs",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        added = 0
        for f in files:
            if f not in self.pdf_files:
                self.pdf_files.append(f)
                added += 1
        if added:
            self._refresh_file_list()
            self._log(f"Added {added} file(s)", 'inf')

    def _add_folder(self):
        folder = filedialog.askdirectory(title="Select folder containing GRN PDFs")
        if not folder:
            return
        files = sorted(glob.glob(os.path.join(folder, "*.pdf")))
        added = 0
        for f in files:
            if f not in self.pdf_files:
                self.pdf_files.append(f)
                added += 1
        self._refresh_file_list()
        self._log(f"Scanned folder → added {added} PDF(s)", 'inf')

    def _clear_files(self):
        self.pdf_files.clear()
        self._refresh_file_list()
        self._log("File list cleared", 'dim')

    def _open_output(self):
        if self.last_output and os.path.exists(self.last_output):
            if os.name == 'nt':
                os.startfile(self.last_output)
            else:
                os.system(f'open "{self.last_output}"')
        else:
            # Fallback: open the output_grn folder if it exists
            folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output_grn")
            if os.path.exists(folder):
                if os.name == 'nt':
                    os.startfile(folder)
                else:
                    os.system(f'open "{folder}"')
            else:
                messagebox.showwarning("Not Found", "No output file yet. Run extraction first.")

    def _run(self):
        if self.is_running:
            return
        if not self.pdf_files:
            messagebox.showwarning("No Files", "Please add at least one GRN PDF.")
            return

        # Build output path: output_grn/ddmmyyyy_hhmmss.xlsx
        script_dir  = os.path.dirname(os.path.abspath(__file__))
        output_dir  = os.path.join(script_dir, "output_grn")
        os.makedirs(output_dir, exist_ok=True)
        timestamp   = time.strftime("%d%m%Y_%H%M%S")
        output      = os.path.join(output_dir, f"{timestamp}.xlsx")

        self.is_running = True
        self.run_btn.config(fg=TEXT_DIM)
        self._set_status("PROCESSING...", AMBER)
        threading.Thread(target=self._extract_worker, args=(output,), daemon=True).start()

    def _extract_worker(self, output):
        all_items     = []
        all_summaries = []
        total = len(self.pdf_files)

        for i, pdf_path in enumerate(self.pdf_files):
            fname = os.path.basename(pdf_path)
            self.after(0, self._set_status, f"Processing {i+1}/{total}:  {fname[:40]}", AMBER)
            self.after(0, self._set_progress, int((i / total) * 90))
            try:
                header, df = parse_grn_pdf(pdf_path)
                all_items.append(df)
                all_summaries.append(header)
                msg = f"✓ {fname}  →  PO {header['PO Number']}  |  {len(df)} SKUs  |  Fill {header['Fill Rate %']}%"
                self.after(0, self._log, msg, 'ok')
            except Exception as e:
                self.after(0, self._log, f"✗ {fname}  →  {e}", 'err')

        if not all_items:
            self.after(0, self._done, None, None, "ERROR: No data extracted", RED)
            return

        try:
            combined = pd.concat(all_items, ignore_index=True)
            summary  = pd.DataFrame(all_summaries)

            sc_cols = ['PO Number', 'PO Date', 'Facility',
                       'Total PO Qty', 'Total GRN Qty', 'Fill Rate %',
                       'Articles in PO', 'Articles in GRN',
                       'Total PO Amount', 'Net GRN Amount', 'GMV Loss']
            summary = summary[[c for c in sc_cols if c in summary.columns]]

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                combined.to_excel(writer, sheet_name='GRN Line Items', index=False)
                summary.to_excel(writer, sheet_name='PO Summary', index=False)

            self.after(0, self._set_progress, 95)
            format_excel(output)
            self.after(0, self._set_progress, 100)
            self.after(0, self._update_stats, combined)
            self.after(0, self._done, combined, output, f"DONE  //  {len(combined)} rows saved", GREEN)
        except Exception as e:
            self.after(0, self._done, None, None, f"ERROR: {e}", RED)

    def _done(self, df, output, msg, color):
        self.is_running = False
        self.run_btn.config(fg=ACCENT)
        self._set_status(msg, color)
        if df is not None:
            self.last_output = output
            self.last_path_var.set(os.path.basename(output))
            self._log(f"Saved → {output}", 'inf')
            if messagebox.askyesno("Done!", f"Extraction complete!\n\n{msg}\n\nOpen output file?"):
                self._open_output()


# ─── STYLE ─────────────────────────────────────────────────────────────────────

def apply_style():
    style = ttk.Style()
    style.theme_use('default')
    style.configure('Vertical.TScrollbar',
                    background=SURFACE2, troughcolor=SURFACE,
                    arrowcolor=TEXT_DIM, bordercolor=BORDER,
                    lightcolor=SURFACE2, darkcolor=SURFACE2)
    style.map('Vertical.TScrollbar',
              background=[('active', BORDER)])


# ─── ENTRY ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app = BlinkitGRNApp()
    apply_style()
    app.mainloop()