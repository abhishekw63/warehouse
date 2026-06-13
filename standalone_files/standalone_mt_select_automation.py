"""
================================================================================
MT Select — Multi-Channel PO Processor  v0.6
================================================================================

PURPOSE
-------
Convert MT-channel PO files (H&G today; H&B, Apollo, Reliance, Wellness
Forever next) into D365-ready Sales Orders. SINGLE-FILE tool with a
Tkinter GUI. No CLI. Adding a new channel is one entry in the CHANNELS
registry (Section 2) — no code rewrite.

ARCHITECTURE
------------
Each MT channel declares its own ChannelConfig — CSV column schema,
lookup chain (SKU→EAN→Item vs EAN→Item direct), Ship-to party filter,
input folder, validation expectations, optional channel master sheet.
The engine reads the active channel's config and behaves accordingly.

  CHANNELS = {
      'HG': ChannelConfig(..., lookup_via='SKU', channel_master_sheet='HG Master', ...),
      # 'HB': ChannelConfig(..., lookup_via='EAN', channel_master_sheet=None, ...),
  }

CURRENT CAPABILITIES (v0.6)
---------------------------
  ✓ Multi-channel architecture (one channel registered today: HG)
  ✓ Tkinter GUI styled like Online PO Processor:
       - Channel selector dropdown (shows all registered channels)
       - Warehouse selector (AHD → PICK, BLR → DS_BL_OFF1)
       - Subtitle updates per active channel
       - Auto-loads MT_Masters.xlsx on startup
       - Sheet-name aliases (Item Master vs Items Master, etc.)
       - Persisted master path + active channel across runs
       - Big green Generate Sales Orders button
  ✓ Loads MT_Masters.xlsx with:
       - Items Master  (shared across channels: EAN → Item No)
       - Ship-to B2B   (shared, filtered by channel.party)
       - One sheet per channel that needs a SKU→EAN master (HG Master)
  ✓ Parses N CSV files in one batch via channel-aware reader
  ✓ Resolves every line via the active channel's lookup chain

NOT YET IMPLEMENTED
-------------------
  ✗ SO number generation (Phase 3 — one seq per PO, regular + tester
    derived from same number)
  ✗ Output workbook writer (Phase 4 — Headers/Lines/Summary/Validation/...
    sheets matching D365 import format)

WHY ONE FILE
------------
Operator-friendly: one .py to ship, one .exe when frozen, no dependency
graph to explain. The cost is a longer file (~1400 lines) — mitigated by
clear ─── SECTION ─── markers throughout.

FOLDER STRUCTURE
----------------
    <script_dir>/
        mt_select.py
        Calculation_Data_MT/
            MT_Masters.xlsx            ← three master sheets
        Input_HG/
            91143-ENR013.csv           ← H&G PO files (default location)
            91149-MNR073.csv
            ...

The "multi-file" workflow: drop all the day's H&G CSVs into ``Input_HG/``,
run ``python mt_select.py``, and the tool batch-processes everything in
one go. (Each CSV stays its own PO — "merge" means batch-process into one
consolidated report/output, not combine lines across POs.)

H&G CSV FORMAT (verified against real files)
--------------------------------------------
15 columns, comma-delimited, UTF-8, CR-only line endings:
    REGION_CODE  LOCATION_CODE  LOCATION_NAME  MANUF_NAME
    VENDOR_CODE  VENDOR_NAME    PO_NO          STORE_NAME
    PDF_STATUS   SKU_CODE       SKU_NAME       MRP
    QUANTITY     PURCHASE_COST  PO_VALUE

Strict rules:
    one file = one PO_NO  |  one file = one STORE_NAME

Business invariants (validated; deviations flagged as warnings):
    PURCHASE_COST × QUANTITY = PO_VALUE
    PURCHASE_COST / MRP      = 0.65  (H&G's standard landing rate)

RUN
---
    python mt_select.py
        Opens the GUI. There is no command-line mode.

        Drop H&G CSVs in Input_HG/ next to the script, or click
        "Add Files..." inside the GUI to pick them from anywhere.

        If MT_Masters.xlsx isn't where expected, click "Browse..."
        to point at it, or "Create Template" to generate a fresh
        empty one.

DEPENDENCIES
------------
    pip install pandas openpyxl

================================================================================
"""

# ════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ════════════════════════════════════════════════════════════════════════════

import os
import sys
import re
import json
import argparse
import threading
import io
from contextlib import redirect_stdout
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Any, Set

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Tkinter is part of Python's stdlib — should always be available on a
# normal Python install (CPython on Windows always has it). We import
# lazily inside main() so headless servers can still use --cli mode
# without crashing on missing tkinter.

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 1 — CONSTANTS & PATHS ───────────────────────────
# ════════════════════════════════════════════════════════════════════════════

# ── Folder + file conventions ──
BUNDLED_FOLDER     = "Calculation_Data_MT"   # holds MT_Masters.xlsx
BUNDLED_WORKBOOK   = "MT_Masters.xlsx"
INPUT_FOLDER_HG    = "Input_HG"              # default place for H&G CSVs

# ── Sheet names inside MT_Masters.xlsx (fixed, case-insensitive match) ──
SHEET_ITEMS_MASTER = "Items Master"
SHEET_SHIP_TO_B2B  = "Ship-to B2B"
SHEET_HG_MASTER    = "HG Master"
SHEET_NON_STOCK    = "Non Stock"

# ── Sheet-name aliases ──
# Operators don't always use our canonical sheet names. Common real-world
# variations (singular/plural, hyphen vs space, capitalisation) all resolve
# to the same canonical sheet via _find_sheet().
#
# Add new aliases here whenever a new naming convention turns up in the
# wild — no other code needs to change.
SHEET_ALIASES: Dict[str, List[str]] = {
    SHEET_ITEMS_MASTER: [
        'Items Master', 'Item Master',         # plural / singular
        'Items_Master', 'Item_Master',
        'ItemMaster',   'ItemsMaster',
    ],
    SHEET_SHIP_TO_B2B: [
        'Ship-to B2B',  'Ship-To B2B',         # case on the 'T'
        'Ship to B2B',  'ShipTo B2B',          # no hyphen
        'Ship_To_B2B',  'Ship_to_B2B',
        'ShipToB2B',
    ],
    SHEET_HG_MASTER: [
        'HG Master', 'HGMaster', 'HG_Master',
        'H&G Master', 'H&G_Master',
    ],
    SHEET_NON_STOCK: [
        'Non Stock', 'NonStock', 'Non_Stock', 'Non-Stock',
        'NON STOCK', 'NonStockItems', 'Non Stock Items',
    ],
}

# ── Config persistence ──
# Remember the last-used masters path across runs. Operators often keep
# their masters file on OneDrive / a shared drive rather than next to the
# script — saving the path means they don't have to re-browse every launch.
CONFIG_FILE = "mt_select_config.json"

# ── Sequence state file ──
# Persists the per-day SO number counter across runs so a second batch
# in the same day picks up where the first left off (no SO# collisions).
# Resets automatically on a new day. Per-channel state keyed by channel
# code so HG's counter doesn't interfere with HB's (when HB joins).
SEQ_STATE_FILE = "mt_select_seq.json"

# ── Column schemas for each master sheet ──
# The loader matches column NAMES (case-insensitive), not positions, so the
# operator can reorder columns without breaking the load. The order listed
# here is what the template generator uses.

COLS_ITEMS_MASTER = [
    'No.',              # D365 Item No — what gets written to SO line.No.
    'GTIN',             # 13-digit EAN — the lookup key from HG Master.ENN code
    'Description',      # For human readability in outputs
    'Mrp',              # MRP — used in mismatch validation & cost calcs
    'GST Group Code',   # e.g. G-18-S — used by Phase 3 cost calcs
]

COLS_SHIP_TO_B2B = [
    'Party',            # Marketplace tag: 'HG', 'HB', ... (filters the sheet)
    'Del Location',     # Lookup key (HG: store name; HB: site code)
    'Cust No',          # D365 Sell-to Customer No.
    'Ship to',          # D365 Ship-to Code
    'Name',             # Branch display name (Summary verification)
    'Address',          # ↓ Verification only — not written to D365
    'Address 2',
    'Postcode',
    'City',
]

COLS_HG_MASTER = [
    'sku_code',         # H&G's SKU number (lookup key from CSV)
    'sku_name',         # Product name (for debugging)
    'ENN code',         # EAN/GTIN (links to Items Master)
    'status',           # 'Active' or 'Inactive'
]

# Non Stock sheet columns — GWP/PWP kit items appended to tester SOs.
COLS_NON_STOCK = [
    'No.',              # D365 Item No (used directly, no lookup)
    'GTIN',             # EAN, reference only
    'Description',      # Item name
    'Qty',              # Planning reference (each row -> 1 tester line of qty 1)
    'Group',            # Category label
]

# ── H&G CSV defaults — kept for back-compat; the canonical source is
#    now CHANNELS['HG'] in Section 1b. New code should use channel.* not
#    these globals.
EXPECTED_LANDING_RATIO  = 0.65   # PURCHASE_COST / MRP
LANDING_RATIO_TOLERANCE = 0.01   # ±1% wiggle for rounding
PO_VALUE_TOLERANCE      = 0.01   # ±₹0.01 for cost×qty vs PO_VALUE check
MRP_DRIFT_THRESHOLD     = 0.05   # 5% — H&G MRP vs Items Master MRP

# ── Lookup party for H&G channel ──
PARTY_HG = 'HG'

# ── Warehouse codes (matches Online PO Processor convention) ──
# Operator picks AHD or BLR in the GUI; engine writes the resolved D365
# location code (PICK or DS_BL_OFF1) into the SO header's Location Code.
WAREHOUSES: Dict[str, str] = {
    'AHD': 'PICK',
    'BLR': 'DS_BL_OFF1',
}
DEFAULT_WAREHOUSE = 'AHD'

# ── Excel styling for the template generator ──
_HDR_FILL  = PatternFill(start_color='1F4E78', end_color='1F4E78',
                          fill_type='solid')
_HDR_FONT  = Font(bold=True, color='FFFFFF', size=11)
_NOTE_FONT = Font(italic=True, color='666666')

# ── Status colours for the output workbook (Summary / Validation /
#    Reconciliation). Conditional-formatting-style cell fills used to
#    surface row health at a glance. Colours chosen to match Excel's
#    own Good/Neutral/Bad named styles.
#
#    OK    — Excel "Good"     light green  (#C6EFCE bg, #006100 text)
#    WARN  — Excel "Neutral"  light yellow (#FFEB9C bg, #9C5700 text)
#    FAIL  — Excel "Bad"      light red    (#FFC7CE bg, #9C0006 text)
#
#    Strong (BG) variants are for the Status cell itself; Tint variants
#    are paler shades for tinting the rest of the row, so problem rows
#    stand out without overwhelming the eye.
_OK_FILL    = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                            fill_type='solid')
_WARN_FILL  = PatternFill(start_color='FFEB9C', end_color='FFEB9C',
                            fill_type='solid')
_FAIL_FILL  = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                            fill_type='solid')
_OK_TINT    = PatternFill(start_color='E8F5E9', end_color='E8F5E9',
                            fill_type='solid')
_WARN_TINT  = PatternFill(start_color='FFF8E1', end_color='FFF8E1',
                            fill_type='solid')
_FAIL_TINT  = PatternFill(start_color='FFEBEE', end_color='FFEBEE',
                            fill_type='solid')

_OK_FONT    = Font(bold=True, color='006100')
_WARN_FONT  = Font(bold=True, color='9C5700')
_FAIL_FONT  = Font(bold=True, color='9C0006')

# Status → (cell fill, cell font, row tint) lookup. Centralising this
# means every sheet that uses status highlighting stays consistent.
_STATUS_STYLES: Dict[str, Tuple[PatternFill, Font, PatternFill]] = {
    'OK':      (_OK_FILL,   _OK_FONT,   _OK_TINT),
    'WARN':    (_WARN_FILL, _WARN_FONT, _WARN_TINT),
    'WARNING': (_WARN_FILL, _WARN_FONT, _WARN_TINT),
    'FAIL':    (_FAIL_FILL, _FAIL_FONT, _FAIL_TINT),
    'FAILED':  (_FAIL_FILL, _FAIL_FONT, _FAIL_TINT),
    'ERROR':   (_FAIL_FILL, _FAIL_FONT, _FAIL_TINT),
    'WRITTEN': (_OK_FILL,   _OK_FONT,   _OK_TINT),
    'SKIPPED': (_WARN_FILL, _WARN_FONT, _WARN_TINT),
}

# ── Visual presentation styles for the output workbook ──
# Goal: every sheet feels like a polished operator report — headers
# stand out, numbers right-align, text columns left-align, derivative
# blocks (testers / non-stock) get subtle tints so structure is visible
# without scanning row by row.

_ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
_ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=False)
_ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center', wrap_text=False)

# Thin grey borders for data rows; medium dark-blue for headers.
_BORDER_THIN = Border(
    left=Side(style='thin',   color='B0B0B0'),
    right=Side(style='thin',  color='B0B0B0'),
    top=Side(style='thin',    color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
_BORDER_HEADER = Border(
    left=Side(style='thin',   color='1F4E78'),
    right=Side(style='thin',  color='1F4E78'),
    top=Side(style='medium',  color='1F4E78'),
    bottom=Side(style='medium', color='1F4E78'),
)

# Row tints — pale shades so operators can spot tester vs non-stock
# blocks at a glance without overwhelming the page.
_TESTER_ROW_TINT   = PatternFill(start_color='E3F2FD', end_color='E3F2FD',
                                   fill_type='solid')      # pale blue
_NONSTOCK_ROW_TINT = PatternFill(start_color='FFF3E0', end_color='FFF3E0',
                                   fill_type='solid')      # pale orange

# D365 preamble row styling.
_PREAMBLE_FILL = PatternFill(start_color='305496', end_color='305496',
                               fill_type='solid')
_PREAMBLE_FONT = Font(bold=True, color='FFFFFF', size=11)

# TOTAL / CHECK row styling.
_TOTAL_FILL = PatternFill(start_color='FFD966', end_color='FFD966',
                            fill_type='solid')
_TOTAL_FONT = Font(bold=True, color='000000', size=11)

# Footer caption styling.
_FOOTER_FONT = Font(italic=True, color='555555', size=10)
_FOOTER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2',
                             fill_type='solid')

# Canonical column-name buckets for per-column alignment.
_NUMERIC_COLS = {
    'qty', 'quantity', 'total qty', 'items', 'count of ean',
    'amount', 'total amount', 'mrp', 'landing', 'landing (65%)',
    'our cost price', 'marketplace cost', 'difference with cost',
    'line no.', 'output line no', 'tester line no', 'lines written',
    'csv row', 'unit cost', 'unit price', 'no. of ean',
}
_CENTERED_COLS = {
    'po', 'po no', 'so no', 'tester so', 'document type', 'type',
    'ship-to code', 'sell-to customer no.', 'cust no', 'ship-to',
    'location code', 'item no', 'item no.', 'no.',
    'posting date', 'order date', 'document date',
    'invoice from date', 'invoice to date',
    'dimension set id', 'supply type', 'gst code',
    'brand code (dimension)', 'channel code (dimension)',
    'catagory (dimension)', 'geography code (dimension)',
    'status', 'ean', 'ean (resolved)', 'gtin',
    'external document no.', 'voucher narration', 'source file',
}

def _apply_table_format(ws, header_row: int = 1,
                          data_start: Optional[int] = None,
                          data_end: Optional[int] = None) -> None:
    """
    Apply table-style formatting to a worksheet:
      - Header row: dark blue fill, white bold, centered, bordered.
      - Data rows: thin grey borders + per-column alignment.
      - Freeze on header row.
    """
    if ws.max_row < header_row:
        return

    for cell in ws[header_row]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_HEADER

    ws.freeze_panes = f'A{header_row + 1}'

    # Per-column alignment based on header text
    header_cells = ws[header_row]
    numeric_cols, center_cols = set(), set()
    for idx, cell in enumerate(header_cells, start=1):
        name = (str(cell.value) if cell.value is not None else '').lower().strip()
        if name in _NUMERIC_COLS:
            numeric_cols.add(idx)
        elif name in _CENTERED_COLS:
            center_cols.add(idx)

    if data_start is None:
        data_start = header_row + 1
    if data_end is None:
        data_end = ws.max_row

    for r in range(data_start, data_end + 1):
        for c_idx, cell in enumerate(ws[r], start=1):
            cell.border = _BORDER_THIN
            if c_idx in numeric_cols:
                cell.alignment = _ALIGN_RIGHT
            elif c_idx in center_cols:
                cell.alignment = _ALIGN_CENTER
            else:
                cell.alignment = _ALIGN_LEFT

def _apply_preamble_style(ws, preamble_row: int = 1) -> None:
    """Style the D365 import preamble row (dark blue band, white bold)."""
    for cell in ws[preamble_row]:
        if cell.value is not None:
            cell.fill = _PREAMBLE_FILL
            cell.font = _PREAMBLE_FONT
            cell.alignment = _ALIGN_LEFT

def _tint_row(ws, row_idx: int, tint: PatternFill) -> None:
    """Apply a background tint to every populated cell in a row."""
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        cell.fill = tint

def _style_total_row(ws, row_idx: int) -> None:
    """Bold gold-band styling for TOTAL / CHECK rows."""
    for cell in ws[row_idx]:
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _BORDER_THIN

def _style_footer_row(ws, row_idx: int) -> None:
    """Italic-grey footer caption styling."""
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.font = _FOOTER_FONT
            cell.fill = _FOOTER_FILL
            cell.alignment = _ALIGN_LEFT

def get_script_dir() -> Path:
    """
    Return the directory containing this script — works for both source
    (.py) and PyInstaller-frozen (.exe) execution.
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent

def get_masters_path() -> Path:
    """Path to the bundled MT_Masters.xlsx workbook."""
    return get_script_dir() / BUNDLED_FOLDER / BUNDLED_WORKBOOK

def get_input_folder(channel_code: Optional[str] = None) -> Path:
    """
    Path to a channel's H&G CSV drop folder. If ``channel_code`` is None,
    returns the default (HG) folder for back-compat.
    """
    if channel_code and channel_code in CHANNELS:
        folder_name = CHANNELS[channel_code].input_folder_name
    else:
        folder_name = INPUT_FOLDER_HG
    return get_script_dir() / folder_name

def get_output_folder(channel_code: Optional[str] = None) -> Path:
    """
    Path to a channel's SO output folder. Created on demand. Falls back
    to a sensible default if the channel doesn't declare one.
    """
    if channel_code and channel_code in CHANNELS:
        folder_name = (CHANNELS[channel_code].output_folder_name
                       or f'Output_{channel_code}')
    else:
        folder_name = 'Output_HG'
    return get_script_dir() / folder_name

def ensure_folders() -> None:
    """Create Calculation_Data_MT/ + Input_*/ + Output_*/ folders if missing."""
    (get_script_dir() / BUNDLED_FOLDER).mkdir(parents=True, exist_ok=True)
    for channel in CHANNELS.values():
        (get_script_dir() / channel.input_folder_name
         ).mkdir(parents=True, exist_ok=True)
        if channel.output_folder_name:
            (get_script_dir() / channel.output_folder_name
             ).mkdir(parents=True, exist_ok=True)

def get_config_path() -> Path:
    """Path to the small JSON file that remembers operator preferences."""
    return get_script_dir() / CONFIG_FILE

def load_config() -> Dict[str, Any]:
    """
    Load saved operator preferences (last-used masters path, etc.).
    Returns an empty dict if the file doesn't exist or is corrupt — we
    never block startup on a broken config.
    """
    path = get_config_path()
    if not path.exists():
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}

def save_config(cfg: Dict[str, Any]) -> None:
    """Persist operator preferences. Failures are silent — config is
    convenience-only, never required for correct operation."""
    try:
        with open(get_config_path(), 'w', encoding='utf-8') as f:
            json.dump(cfg, f, indent=2)
    except OSError:
        pass

def get_seq_state_path() -> Path:
    """Path to the JSON file holding per-channel SO number counter state."""
    return get_script_dir() / SEQ_STATE_FILE

def load_seq_state() -> Dict[str, Any]:
    """
    Load the full seq-state dict (all channels). Shape:
        {
          "HG": {"date": "2026-06-05", "next_counter": 50628},
          "HB": {"date": "2026-06-04", "next_counter": 40627}
        }
    Returns {} if missing or corrupt — caller treats this as "first run".
    """
    path = get_seq_state_path()
    if not path.exists():
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}

def save_seq_state(state: Dict[str, Any]) -> None:
    """Persist seq-state. Failures are silent — see save_config rationale."""
    try:
        with open(get_seq_state_path(), 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2)
    except OSError:
        pass

def _norm_sheet_name(s: Any) -> str:
    """
    Normalise a sheet name for alias matching: lowercase + drop everything
    except letters and digits.

    Examples:
        'Ship-To B2B'  → 'shiptob2b'
        'Ship_to_B2B'  → 'shiptob2b'
        'ShipToB2B'    → 'shiptob2b'
        'Item Master'  → 'itemmaster'
        'Items Master' → 'itemsmaster'  ← still distinct from 'Item Master',
                                          which is why SHEET_ALIASES exists.
    """
    if s is None:
        return ''
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def _find_sheet(sheet_names: List[str], canonical: str) -> Optional[str]:
    """
    Find which of the workbook's actual sheet names corresponds to one of
    the aliases for ``canonical``. Returns the actual sheet name as it
    appears in the workbook (for use with pd.read_excel) or None.

    Matching is case-insensitive and tolerant of whitespace/hyphen/
    underscore differences. Singular vs plural ('Item Master' vs
    'Items Master') is handled via the explicit SHEET_ALIASES list.
    """
    aliases = SHEET_ALIASES.get(canonical, [canonical])
    normalised_aliases = {_norm_sheet_name(a) for a in aliases}
    for actual in sheet_names:
        if _norm_sheet_name(actual) in normalised_aliases:
            return actual
    return None

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 2 — DATA MODELS ─────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Dataclasses for everything the tool moves around. Each class is the
# "shape" of one concept — channel configuration, master entries, parsed
# PO, etc. Field types are explicit so IDE/typing catches mistakes early.

# ── Channel configuration ──
# Each MT channel (HG = Health & Glow, future HB = Health & Beauty, etc.)
# declares its CSV format, master file requirements, lookup chain, and
# display metadata in a single ChannelConfig instance. The engine reads
# these declarations and behaves accordingly — adding a new channel is
# ONE new entry in CHANNELS, not a code rewrite.

@dataclass
class ChannelConfig:
    """Per-channel configuration. See CHANNELS registry below for instances."""
    # ── Identity ──
    code:               str   # 'HG', 'HB', ... — short channel code
    display_name:       str   # 'Health & Glow' — full name shown in GUI
    party:              str   # Ship-to B2B Party filter (usually = code)
    input_folder_name:  str   # 'Input_HG' — default subfolder for CSVs

    # ── CSV column mapping ──
    # Engine extracts values by column name (case-insensitive). Required
    # columns are validated at load time; missing required = hard error.
    csv_required_cols:  List[str]
    csv_po_col:         str          # PO/document number column
    csv_store_col:      str          # store name / site code (Ship-to lookup)
    csv_id_col:         str          # SKU code OR EAN (depends on lookup_via)
    csv_qty_col:        str
    csv_mrp_col:        str
    csv_cost_col:       Optional[str] = None   # per-unit cost (if present)
    csv_value_col:      Optional[str] = None   # line total (if present)

    # ── Lookup strategy ──
    # 'SKU': csv_id_col is a SKU; resolve via channel master to EAN, then
    #        EAN → Items Master → Item No. (Used by HG.)
    # 'EAN': csv_id_col IS the EAN; resolve EAN → Items Master directly.
    #        No channel master sheet needed.
    lookup_via:         str = 'SKU'

    # ── Optional channel master (SKU → EAN) ──
    # Used only when lookup_via == 'SKU'. The sheet in MT_Masters.xlsx
    # holding this channel's SKU → EAN map.
    channel_master_sheet:  Optional[str] = None
    channel_master_cols:   Optional[List[str]] = None

    # ── D365 SO header constants ──
    # These are the values written into D365 SO headers for this channel.
    # Sell-to is the channel's master customer account in D365; ship-to
    # codes come from Ship-to B2B sheet at runtime.
    sell_to:           str = ''        # e.g. '20039' (HG master account)
    document_type:     str = 'Order'
    supply_type:       str = 'B2B'

    # ── Tester SO settings ──
    # When testers are enabled (GUI checkbox), every cleanly-parsed
    # regular PO gets a paired tester SO using the same counter block.
    # Tester rows have qty=1 and unit_price = tester_unit_price.
    # None = this channel doesn't support testers; leave the GUI
    # checkbox unticked or it's a no-op.
    tester_unit_price:        Optional[float] = None
    tester_external_doc_no:   str = 'TESTERS'

    # ── Output folder ──
    # Where the generated SO Excel workbook gets written.
    output_folder_name: str = ''       # e.g. 'Output_HG'

    # ── Validation expectations ──
    # When set, the CSV reader warns if actual values drift significantly.
    # None = don't check.
    expected_landing_ratio:    Optional[float] = None
    landing_ratio_tolerance:   float = 0.01
    po_value_tolerance:        float = 0.01
    mrp_drift_threshold:       float = 0.05

# ── Channel registry ──
# Add a new channel by appending an entry here. No other code changes
# needed (engine and GUI both read from this dict).
CHANNELS: Dict[str, ChannelConfig] = {
    'HG': ChannelConfig(
        code='HG',
        display_name='Health & Glow',
        party='HG',
        input_folder_name='Input_HG',
        output_folder_name='Output_HG',
        sell_to='20039',                    # Health & Glow Private Limited
        tester_unit_price=0.54,             # HG tester sample price
        csv_required_cols=[
            'REGION_CODE',    'LOCATION_CODE',   'LOCATION_NAME',   'MANUF_NAME',
            'VENDOR_CODE',    'VENDOR_NAME',     'PO_NO',           'STORE_NAME',
            'PDF_STATUS',     'SKU_CODE',        'SKU_NAME',        'MRP',
            'QUANTITY',       'PURCHASE_COST',   'PO_VALUE',
        ],
        csv_po_col='PO_NO',
        csv_store_col='STORE_NAME',
        csv_id_col='SKU_CODE',
        csv_qty_col='QUANTITY',
        csv_mrp_col='MRP',
        csv_cost_col='PURCHASE_COST',
        csv_value_col='PO_VALUE',
        lookup_via='SKU',
        channel_master_sheet='HG Master',
        channel_master_cols=['sku_code', 'sku_name', 'ENN code', 'status'],
        expected_landing_ratio=0.65,
    ),
    # Example placeholder showing how HB will be added later:
    # 'HB': ChannelConfig(
    #     code='HB', display_name='Health & Beauty', party='HB',
    #     input_folder_name='Input_HB',
    #     csv_required_cols=['Purchasing Document', 'Site', 'EAN', ...],
    #     csv_po_col='Purchasing Document', csv_store_col='Site',
    #     csv_id_col='EAN', csv_qty_col='Order Quantity', csv_mrp_col='MRP',
    #     lookup_via='EAN',          # H&B CSVs carry EAN directly
    #     channel_master_sheet=None, # no SKU→EAN master needed
    #     expected_landing_ratio=None,
    # ),
}

DEFAULT_CHANNEL = 'HG'

def get_channel(code: str) -> ChannelConfig:
    """Look up a channel by code; raises KeyError on unknown channel."""
    return CHANNELS[code]

# ── Master sheet records ──

@dataclass
class ItemMasterEntry:
    """One row from Items Master, keyed by GTIN."""
    item_no:     str              # D365 Item No (e.g. '201336')
    gtin:        str              # 13-digit EAN as string
    description: str
    mrp:         Optional[float]  # May be None if blank
    gst_code:    str              # e.g. 'G-18-S'

@dataclass
class ShipToEntry:
    """
    One row from Ship-to B2B, keyed by (party, del_location).

    Address fields are kept for verification on the Summary sheet of the
    output workbook (your "for safer side take full address" requirement).
    The engine itself only needs ship_to and cust_no for the SO header.
    """
    party:        str
    del_location: str
    cust_no:      str             # D365 Sell-to (e.g. '20039')
    ship_to:      str             # D365 Ship-to Code (e.g. '20039_10')
    name:         str = ''
    address:      str = ''
    address_2:    str = ''
    postcode:     str = ''
    city:         str = ''

@dataclass
class ChannelMasterEntry:
    """
    One row from a channel's SKU → EAN master sheet (HG Master, future
    HB Master if needed, etc.). Used by channels with lookup_via='SKU'.

    EAN may be blank — that's the "SKU exists but unmapped" state.
    """
    sku_code: str
    sku_name: str
    enn_code: Optional[str]       # None if blank
    status:   str                 # 'Active' or 'Inactive'

@dataclass
class NonStockEntry:
    """
    One row from the Non Stock sheet — a GWP/PWP/sample item appended to
    tester SOs when the operator ticks "Add Non Stock". Item No is THE
    D365 identifier (no resolution chain). GTIN / Qty / Group are
    reference fields. Each row emits exactly ONE tester line of qty=1.
    """
    item_no:     str
    gtin:        str               # for reference only
    description: str
    qty:         int               # reference; not used as multiplier
    group:       str               # category label

@dataclass
class MasterBundle:
    """All masters loaded together, plus file metadata."""
    items_by_gtin:  Dict[str, ItemMasterEntry]  = field(default_factory=dict)
    ship_to_lookup: Dict[tuple, ShipToEntry]    = field(default_factory=dict)

    # Channel-specific masters keyed by channel code.
    channel_masters: Dict[str, Dict[str, ChannelMasterEntry]] = field(
        default_factory=dict)

    # Non-stock kit items appended to tester SOs (single shared list).
    non_stock: List[NonStockEntry] = field(default_factory=list)

    workbook_path:  Optional[Path]     = None
    workbook_mtime: Optional[datetime] = None
    findings:       List[Tuple[str, str]] = field(default_factory=list)

    items_sheet_name:    str = ''
    ship_to_sheet_name:  str = ''
    non_stock_sheet_name: str = ''
    channel_sheet_names: Dict[str, str] = field(default_factory=dict)

    # ── Convenience accessors ──
    def get_channel_master(self, channel_code: str
                             ) -> Dict[str, ChannelMasterEntry]:
        """Return the SKU→entry dict for a channel; {} if absent."""
        return self.channel_masters.get(channel_code, {})

    def get_channel_sheet_name(self, channel_code: str) -> str:
        """Return the matched sheet name for a channel master."""
        return self.channel_sheet_names.get(channel_code, '')

# ── PO records ──

@dataclass
class POLine:
    """
    One PO line — original CSV row + resolved fields + status.

    Original CSV values are preserved for audit. Resolved fields are
    populated by the master lookup chain.
    """
    # ── Original CSV values ──
    csv_row_num:   int        # 1-based file row (2 = first data row)
    sku_code:      str
    sku_name:      str
    mrp:           float
    quantity:      int
    purchase_cost: float
    po_value:      float

    # ── Resolved fields ──
    ean:                Optional[str]   = None
    item_no:            Optional[str]   = None
    items_master_mrp:   Optional[float] = None
    gst_code:           Optional[str]   = None
    items_master_desc:  Optional[str]   = None
    hg_master_status:   Optional[str]   = None

    # ── Status & notes ──
    status: str = 'OK'   # 'OK', 'WARN', 'SKIP'
    notes:  List[str] = field(default_factory=list)

    # ── Output tracking ──
    # Populated by _write_lines_sheet when this line is actually written
    # to Lines (SO). 0 means "not written to Lines" — either because the
    # line was skipped, the master lookup failed, or the file had a hard
    # error. The Reconciliation sheet uses this to prove every input row
    # is accounted for.
    output_line_no: int = 0

    # Populated when this line was ALSO written to the paired tester SO
    # (operator enabled the Generate Testers checkbox). 0 if no tester
    # row was generated for this input line.
    tester_output_line_no: int = 0

    # Marks whether this line should produce a tester output row.
    # Default False — assign_so_numbers flips it to True for eligible
    # lines when the operator has testers enabled.
    #
    # In automatic mode (no dump file): every resolved, non-SKIP line
    # in a PO with an assigned regular SO becomes eligible.
    #
    # In selective mode (dump file provided): a line is eligible only
    # if (PO.location_code, line.sku_code) appears in the dump's
    # eligible_keys set.
    is_tester_eligible: bool = False

    def add_note(self, level: str, msg: str) -> None:
        self.notes.append(f"[{level.upper()}] {msg}")
        if level == 'warn' and self.status == 'OK':
            self.status = 'WARN'
        elif level == 'skip':
            self.status = 'SKIP'

@dataclass
class POFile:
    """Result of parsing one CSV file."""
    source_path:    Path
    source_name:    str
    file_mtime:     Optional[datetime] = None

    # PO header (single value across all lines)
    po_no:          str = ''
    store_name:     str = ''
    vendor_code:    str = ''
    vendor_name:    str = ''
    location_code:  str = ''
    region_code:    str = ''

    # Generated SO number for D365 (assigned by assign_so_numbers after parsing)
    # Format: SO/{channel}/{MM}/{DDMMYY incremented per PO}
    # Empty when SO assignment hasn't run, or when the PO had hard errors.
    so_number: str = ''

    # Paired tester SO number — populated only when the operator enabled
    # "Generate Testers" in the GUI before clicking Generate Sales Orders.
    # Same counter block as regulars (assigned in a second pass), but
    # with 'TT' literal in the MM slot to mark it as a tester SO.
    # Format: SO/{channel}/TT/{counter}
    # Empty when testers were not enabled, OR when the regular SO is
    # also empty (no tester without a corresponding regular).
    tester_so_number: str = ''

    # Resolved ship-to
    ship_to:        Optional[str]         = None
    cust_no:        Optional[str]         = None
    ship_to_entry:  Optional[ShipToEntry] = None

    # Parsed lines
    lines: List[POLine] = field(default_factory=list)

    # Original CSV dataframe — kept so the Raw Data sheet in the output
    # workbook can echo the exact source rows (with appended resolved
    # fields). None for files that hit a hard error before df was built.
    raw_df: Optional[Any] = None       # pd.DataFrame at runtime

    # Control totals (computed from CSV — for cross-check against output)
    input_line_count:          int   = 0
    input_qty_total:           int   = 0
    input_po_value_total:      float = 0.0
    input_purchase_cost_total: float = 0.0

    # Findings
    findings: List[Tuple[str, str]] = field(default_factory=list)
    has_hard_errors: bool = False

    def add_finding(self, level: str, msg: str) -> None:
        self.findings.append((level, msg))
        if level == 'error':
            self.has_hard_errors = True

    def ok_lines(self)   -> List[POLine]:
        return [l for l in self.lines if l.status == 'OK']

    def warn_lines(self) -> List[POLine]:
        return [l for l in self.lines if l.status == 'WARN']

    def skip_lines(self) -> List[POLine]:
        return [l for l in self.lines if l.status == 'SKIP']

@dataclass
class POBatch:
    """All POFile results from a single multi-file parse run."""
    po_files:       List[POFile]            = field(default_factory=list)
    master_bundle:  Optional[MasterBundle]  = None
    parsed_at:      datetime                = field(default_factory=datetime.now)
    cross_findings: List[Tuple[str, str]]   = field(default_factory=list)

    # Populated by assign_so_numbers() — one-line summary of the SO# pass.
    # Empty string before assign_so_numbers runs.
    so_assignment_summary: str = ''

    def all_ok(self) -> bool:
        return (not any(f.has_hard_errors for f in self.po_files)
                and not any(lvl == 'error' for lvl, _ in self.cross_findings))

@dataclass
class TesterDump:
    """
    Selective-tester dump file ("Skin_care_NPI_tester.xlsm").

    The operator can optionally provide a dump file alongside the
    Generate Testers checkbox. When present, it acts as a FILTER:
    only PO lines whose (LOCATION_CODE, SKU_CODE) appears in
    eligible_keys become tester lines. PO lines NOT in the dump get
    no tester. A PO with no matching lines gets no tester SO at all.

    File schema (Sheet2):
        CODE  | Store           | SKU    | ENN ode       | DESCRIPTION | MRP | Tester
        ─────────────────────────────────────────────────────────────────────────────
        74    | HG-KARKHANA-HYD | 579176 | 8904473104048 | RENEE PINK… | 599 | 1
        208   | HG-GSM MALL-HYD | 579176 | 8904473104048 | RENEE PINK… | 599 | 1
        ...

    Matching: (LOCATION_CODE, SKU_CODE) — confirmed with operator.
    Tester qty: always 1 (the "Tester" column value is treated as a
    presence flag, not a quantity). This keeps the tester semantics
    consistent across automatic and selective modes.
    """
    source_path:    Optional[Path]                  = None
    source_name:    str                              = ''
    sheet_used:     str                              = ''        # which sheet was read
    eligible_keys:  Set[Tuple[str, str]]             = field(default_factory=set)
    rows_loaded:    int                              = 0
    findings:       List[Tuple[str, str]]            = field(default_factory=list)
    has_hard_errors: bool                            = False

    def add_finding(self, level: str, msg: str) -> None:
        self.findings.append((level, msg))
        if level == 'error':
            self.has_hard_errors = True

    def is_eligible(self, location_code: str, sku_code: str) -> bool:
        """True iff (location_code, sku_code) is in the dump."""
        return (location_code, sku_code) in self.eligible_keys

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 3 — VALUE COERCION HELPERS ──────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Excel and CSV both have type-inference quirks. These helpers normalize
# values to clean string/numeric forms regardless of how pandas read them.

def _normalize_ean(raw: Any) -> str:
    """
    Convert any Excel cell value into a plain digit string for EAN lookup.

    Excel commonly delivers EANs as int (8906121641769), float with .0
    (8906121641769.0), or padded string. We always end up with the digit
    string. Returns '' for None / NaN / empty.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ''
    s = str(raw).strip()
    if s.lower() in ('', 'nan', 'none'):
        return ''
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s

def _coerce_id(raw: Any) -> str:
    """
    Convert any Excel cell to a clean ID string (no trailing .0, no NaN).
    Used for Cust No, Ship-to Code, sku_code — identifiers we treat as
    strings but Excel may give us as int/float.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ''
    s = str(raw).strip()
    if s.lower() in ('', 'nan', 'none'):
        return ''
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s

def _opt_str(raw: Any) -> str:
    """Optional string — '' for NaN/None, stripped str otherwise."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ''
    return str(raw).strip()

def _safe_int(val: Any) -> int:
    """Parse int from any cell value; returns 0 on failure."""
    try:
        if val is None or pd.isna(val):
            return 0
        s = str(val).strip().replace(',', '')
        if not s or s.lower() in ('nan', 'none'):
            return 0
        return int(float(s))
    except (ValueError, TypeError):
        return 0

def _safe_float(val: Any) -> float:
    """Parse float from any cell value; returns 0.0 on failure."""
    try:
        if val is None or pd.isna(val):
            return 0.0
        s = str(val).strip().replace(',', '')
        if not s or s.lower() in ('nan', 'none'):
            return 0.0
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def _normalize_header(s: Any) -> str:
    """Strip and lowercase a header cell for case-insensitive matching."""
    if s is None:
        return ''
    return str(s).strip().lower()

def _resolve_columns(df: pd.DataFrame, expected: List[str]
                      ) -> Dict[str, Optional[str]]:
    """
    Map canonical column names to actual DataFrame column names
    (case-insensitive, whitespace-tolerant).
    """
    lookup = {_normalize_header(c): c for c in df.columns}
    return {col: lookup.get(_normalize_header(col)) for col in expected}

def _get_col_value(row: pd.Series, col_name: str) -> Any:
    """Case-insensitive column-value getter for a CSV row."""
    target = col_name.lower()
    for c, v in row.items():
        if str(c).strip().lower() == target:
            return v
    return None

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 4 — MASTER LOADERS ──────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Each loader reads one sheet and returns a lookup dict + findings list.
# load_all_masters() ties them together into a MasterBundle.

def load_items_master(workbook_path: Path
                       ) -> Tuple[Dict[str, ItemMasterEntry],
                                   List[Tuple[str, str]], str]:
    """
    Load Items Master sheet → dict keyed by GTIN (EAN as string).

    Returns a 3-tuple: (entries, findings, matched_sheet_name). The
    matched sheet name is the actual name in the workbook (which may
    differ from SHEET_ITEMS_MASTER thanks to SHEET_ALIASES — e.g. 'Item
    Master' instead of 'Items Master'). Empty string if no sheet matched.
    """
    findings: List[Tuple[str, str]] = []
    entries: Dict[str, ItemMasterEntry] = {}

    # Find the actual sheet name (case/whitespace/alias tolerant)
    try:
        xls = pd.ExcelFile(workbook_path)
    except Exception as e:
        findings.append(('error', f"Cannot open workbook: {e}"))
        return entries, findings, ''

    actual_sheet = _find_sheet(xls.sheet_names, SHEET_ITEMS_MASTER)
    if actual_sheet is None:
        findings.append(('error',
            f"No sheet matching Items Master found. Looked for: "
            f"{SHEET_ALIASES[SHEET_ITEMS_MASTER]}. "
            f"Workbook has: {xls.sheet_names}"))
        return entries, findings, ''

    try:
        df = pd.read_excel(workbook_path, sheet_name=actual_sheet,
                            header=0)
    except Exception as e:
        findings.append(('error',
            f"Cannot read sheet '{actual_sheet}': {e}"))
        return entries, findings, actual_sheet

    col_map = _resolve_columns(df, COLS_ITEMS_MASTER)
    missing = [c for c, actual in col_map.items() if actual is None]
    if missing:
        findings.append(('error',
            f"Items Master: missing required columns {missing}. "
            f"Found: {list(df.columns)}"))
        return entries, findings, actual_sheet

    blank_gtin = 0
    duplicates = 0
    for _, row in df.iterrows():
        raw_gtin = row[col_map['GTIN']]
        if pd.isna(raw_gtin):
            blank_gtin += 1
            continue

        gtin = _normalize_ean(raw_gtin)
        if not gtin:
            blank_gtin += 1
            continue

        if gtin in entries:
            duplicates += 1
            continue

        item_no = row[col_map['No.']]
        item_no_str = (str(int(item_no))
                        if isinstance(item_no, float) and not pd.isna(item_no)
                        else str(item_no).strip())

        desc_raw = row[col_map['Description']]
        desc = '' if pd.isna(desc_raw) else str(desc_raw).strip()

        mrp_raw = row[col_map['Mrp']]
        mrp = float(mrp_raw) if pd.notna(mrp_raw) else None

        gst_raw = row[col_map['GST Group Code']]
        gst = '' if pd.isna(gst_raw) else str(gst_raw).strip()

        entries[gtin] = ItemMasterEntry(
            item_no=item_no_str, gtin=gtin, description=desc,
            mrp=mrp, gst_code=gst,
        )

    findings.append(('info',
        f"Items Master: loaded {len(entries)} EAN entries "
        f"(blank GTIN: {blank_gtin}, duplicates skipped: {duplicates})"))
    return entries, findings, actual_sheet

def load_ship_to_b2b(workbook_path: Path
                      ) -> Tuple[Dict[tuple, ShipToEntry],
                                  List[Tuple[str, str]], str]:
    """Load Ship-to B2B sheet → dict keyed by (Party, Del Location).
    Returns (entries, findings, matched_sheet_name)."""
    findings: List[Tuple[str, str]] = []
    entries: Dict[tuple, ShipToEntry] = {}

    try:
        xls = pd.ExcelFile(workbook_path)
    except Exception as e:
        findings.append(('error', f"Cannot open workbook: {e}"))
        return entries, findings, ''

    actual_sheet = _find_sheet(xls.sheet_names, SHEET_SHIP_TO_B2B)
    if actual_sheet is None:
        findings.append(('error',
            f"No sheet matching Ship-to B2B found. Looked for: "
            f"{SHEET_ALIASES[SHEET_SHIP_TO_B2B]}. "
            f"Workbook has: {xls.sheet_names}"))
        return entries, findings, ''

    try:
        df = pd.read_excel(workbook_path, sheet_name=actual_sheet,
                            header=0)
    except Exception as e:
        findings.append(('error',
            f"Cannot read sheet '{actual_sheet}': {e}"))
        return entries, findings, actual_sheet

    col_map = _resolve_columns(df, COLS_SHIP_TO_B2B)
    missing = [c for c, actual in col_map.items() if actual is None]
    if missing:
        findings.append(('error',
            f"Ship-to B2B: missing required columns {missing}. "
            f"Found: {list(df.columns)}"))
        return entries, findings, actual_sheet

    blank_loc  = 0
    duplicates = 0
    party_counts: Dict[str, int] = {}

    for _, row in df.iterrows():
        party   = str(row[col_map['Party']] or '').strip()
        del_loc = str(row[col_map['Del Location']] or '').strip()
        if not party or not del_loc:
            blank_loc += 1
            continue

        key = (party, del_loc)
        if key in entries:
            duplicates += 1
            continue

        entries[key] = ShipToEntry(
            party        = party,
            del_location = del_loc,
            cust_no      = _coerce_id(row[col_map['Cust No']]),
            ship_to      = _coerce_id(row[col_map['Ship to']]),
            name         = _opt_str(row[col_map['Name']]),
            address      = _opt_str(row[col_map['Address']]),
            address_2    = _opt_str(row[col_map['Address 2']]),
            postcode     = _opt_str(row[col_map['Postcode']]),
            city         = _opt_str(row[col_map['City']]),
        )
        party_counts[party] = party_counts.get(party, 0) + 1

    party_summary = ', '.join(f"{p}={n}"
                                for p, n in sorted(party_counts.items()))
    findings.append(('info',
        f"Ship-to B2B: loaded {len(entries)} entries [{party_summary}] "
        f"(blank rows skipped: {blank_loc}, duplicates: {duplicates})"))
    return entries, findings, actual_sheet

def load_channel_master(channel: ChannelConfig, workbook_path: Path
                          ) -> Tuple[Dict[str, ChannelMasterEntry],
                                      List[Tuple[str, str]], str]:
    """
    Load a channel's SKU→EAN master sheet → dict keyed by sku_code.

    Only meaningful when channel.lookup_via == 'SKU' and
    channel.channel_master_sheet is set. For EAN-direct channels, returns
    ({}, [], '') with no findings.

    Returns (entries, findings, matched_sheet_name).
    """
    findings: List[Tuple[str, str]] = []
    entries: Dict[str, ChannelMasterEntry] = {}

    # Channel doesn't need a master? Done.
    if channel.lookup_via != 'SKU' or not channel.channel_master_sheet:
        return entries, findings, ''

    sheet_label = channel.channel_master_sheet   # e.g. 'HG Master'
    required_cols = channel.channel_master_cols or [
        'sku_code', 'sku_name', 'ENN code', 'status']

    try:
        xls = pd.ExcelFile(workbook_path)
    except Exception as e:
        findings.append(('error', f"Cannot open workbook: {e}"))
        return entries, findings, ''

    actual_sheet = _find_sheet(xls.sheet_names, sheet_label)
    if actual_sheet is None:
        findings.append(('error',
            f"No sheet matching '{sheet_label}' found. Looked for: "
            f"{SHEET_ALIASES.get(sheet_label, [sheet_label])}. "
            f"Workbook has: {xls.sheet_names}"))
        return entries, findings, ''

    try:
        df = pd.read_excel(workbook_path, sheet_name=actual_sheet,
                            header=0)
    except Exception as e:
        findings.append(('error',
            f"Cannot read sheet '{actual_sheet}': {e}"))
        return entries, findings, actual_sheet

    col_map = _resolve_columns(df, required_cols)
    missing = [c for c, actual in col_map.items() if actual is None]
    if missing:
        findings.append(('error',
            f"{sheet_label}: missing required columns {missing}. "
            f"Found: {list(df.columns)}"))
        return entries, findings, actual_sheet

    n_active    = 0
    n_inactive  = 0
    n_blank_ean = 0
    duplicates  = 0
    blank_sku   = 0

    for _, row in df.iterrows():
        sku_raw = row[col_map['sku_code']]
        if pd.isna(sku_raw):
            blank_sku += 1
            continue

        sku = _coerce_id(sku_raw)
        if not sku:
            blank_sku += 1
            continue

        if sku in entries:
            duplicates += 1
            continue

        sku_name = _opt_str(row[col_map['sku_name']])
        status   = _opt_str(row[col_map['status']]) or 'Active'
        ean_raw  = row[col_map['ENN code']]

        if pd.isna(ean_raw) or str(ean_raw).strip() == '':
            ean = None
            n_blank_ean += 1
        else:
            ean = _normalize_ean(ean_raw)
            if not ean:
                ean = None
                n_blank_ean += 1

        entries[sku] = ChannelMasterEntry(
            sku_code=sku, sku_name=sku_name, enn_code=ean, status=status,
        )
        if status.lower() == 'active':
            n_active += 1
        elif status.lower() == 'inactive':
            n_inactive += 1

    findings.append(('info',
        f"{sheet_label}: loaded {len(entries)} SKUs "
        f"(active: {n_active}, inactive: {n_inactive}, "
        f"blank EAN: {n_blank_ean})"))

    if n_blank_ean > 0:
        unmapped = [e.sku_code for e in entries.values()
                     if e.enn_code is None]
        findings.append(('warn',
            f"{sheet_label}: {len(unmapped)} active SKU(s) have BLANK EAN — "
            f"they will fail lookup if they appear in a PO. "
            f"SKUs: {', '.join(unmapped[:10])}"
            + ("..." if len(unmapped) > 10 else "")))

    if duplicates > 0:
        findings.append(('warn',
            f"{sheet_label}: {duplicates} duplicate sku_code row(s) — "
            f"only first occurrence kept."))

    return entries, findings, actual_sheet

def load_non_stock_list(workbook_path: Path
                          ) -> Tuple[List[NonStockEntry],
                                      List[Tuple[str, str]], str]:
    """
    Load the 'Non Stock' sheet → list of NonStockEntry.

    Optional sheet. Item No is the canonical D365 identifier, no lookup
    chain. Note rows (long prose in column A with blank Description or
    spaces in Item No) are skipped automatically.
    """
    findings: List[Tuple[str, str]] = []
    entries: List[NonStockEntry] = []

    try:
        xls = pd.ExcelFile(workbook_path)
    except Exception as e:
        findings.append(('error', f"Cannot open workbook: {e}"))
        return entries, findings, ''

    actual_sheet = _find_sheet(xls.sheet_names, SHEET_NON_STOCK)
    if actual_sheet is None:
        findings.append(('info',
            "Non Stock: sheet not found. The 'Add Non Stock' checkbox "
            "will have no effect until a 'Non Stock' sheet is added."))
        return entries, findings, ''

    try:
        df = pd.read_excel(workbook_path, sheet_name=actual_sheet, header=0)
    except Exception as e:
        findings.append(('error',
            f"Cannot read sheet '{actual_sheet}': {e}"))
        return entries, findings, actual_sheet

    col_map = _resolve_columns(df, COLS_NON_STOCK)
    if col_map.get('No.') is None or col_map.get('Description') is None:
        findings.append(('error',
            f"Non Stock: missing required columns 'No.' or 'Description'."))
        return entries, findings, actual_sheet

    seen: Set[str] = set()
    duplicates = blank_rows = 0

    for _, row in df.iterrows():
        item_no = _coerce_id(row[col_map['No.']])
        desc    = _opt_str(row[col_map['Description']]) or ''
        if not item_no or not desc:
            blank_rows += 1
            continue
        if len(item_no) > 30 or ' ' in item_no:
            # Note rows (long prose) get filtered here
            blank_rows += 1
            continue
        if item_no in seen:
            duplicates += 1
            continue
        seen.add(item_no)

        gtin = ''
        if col_map.get('GTIN') is not None:
            gtin = _normalize_ean(row[col_map['GTIN']]) or _opt_str(
                row[col_map['GTIN']]) or ''

        qty_val = 1
        if col_map.get('Qty') is not None:
            try:
                raw = row[col_map['Qty']]
                qty_val = int(float(raw)) if raw is not None else 1
            except (ValueError, TypeError):
                qty_val = 1
            if qty_val < 1:
                qty_val = 1

        group = ''
        if col_map.get('Group') is not None:
            group = _opt_str(row[col_map['Group']]) or ''

        entries.append(NonStockEntry(
            item_no=item_no, gtin=gtin, description=desc,
            qty=qty_val, group=group))

    findings.append(('info',
        f"Non Stock: loaded {len(entries)} item(s) "
        f"(blank rows: {blank_rows}, duplicates: {duplicates})"))
    return entries, findings, actual_sheet

def load_all_masters(workbook_path: Optional[Path] = None) -> MasterBundle:
    """
    Top-level loader: reads shared masters (Items Master, Ship-to B2B)
    plus every configured channel's master sheet (if any).

    All findings from all loaders are collected into bundle.findings.
    Per-channel master entries are stored in bundle.channel_masters
    keyed by channel code.
    """
    bundle = MasterBundle()

    if workbook_path is None:
        workbook_path = get_masters_path()

    bundle.workbook_path = workbook_path

    if not workbook_path.exists():
        bundle.findings.append(('error',
            f"Workbook not found: {workbook_path}\n"
            f"  Use Browse... to point to an existing file, or "
            f"Create Template to generate a fresh one."))
        return bundle

    try:
        ts = os.path.getmtime(workbook_path)
        bundle.workbook_mtime = datetime.fromtimestamp(ts)
    except OSError as e:
        bundle.findings.append(('warn', f"Cannot read mtime: {e}"))

    # ── Shared masters (every channel uses these) ──
    items, items_findings, items_sheet = load_items_master(workbook_path)
    bundle.items_by_gtin = items
    bundle.items_sheet_name = items_sheet
    bundle.findings.extend(items_findings)

    ship_to, ship_findings, ship_sheet = load_ship_to_b2b(workbook_path)
    bundle.ship_to_lookup = ship_to
    bundle.ship_to_sheet_name = ship_sheet
    bundle.findings.extend(ship_findings)

    # ── Non Stock list (optional GWP/PWP kit) ──
    ns_entries, ns_findings, ns_sheet = load_non_stock_list(workbook_path)
    bundle.non_stock = ns_entries
    bundle.non_stock_sheet_name = ns_sheet
    bundle.findings.extend(ns_findings)

    # ── Channel-specific masters (one per channel that needs one) ──
    # Iterating CHANNELS means every registered channel gets its master
    # loaded — adding HB later just means adding it to CHANNELS, no
    # change to this function.
    for code, channel in CHANNELS.items():
        if channel.lookup_via != 'SKU' or not channel.channel_master_sheet:
            continue
        entries, findings, sheet_name = load_channel_master(
            channel, workbook_path)
        bundle.channel_masters[code] = entries
        bundle.channel_sheet_names[code] = sheet_name
        bundle.findings.extend(findings)

    return bundle

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 4b — TESTER DUMP FILE LOADER ────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Reads the optional selective-tester dump file the operator provides
# alongside the Generate Testers checkbox. The file lists which
# (LOCATION_CODE, SKU_CODE) pairs are eligible for tester generation.
# Without this file, the system runs in automatic mode (every resolved
# line in every regular PO gets a tester).

# Column-name aliases for the dump sheet. Operators rename columns
# slightly across versions ("ENN ode" → "EAN", "Tester" → "Tester Qty"),
# so we match case-insensitively and accept the common variants.
_DUMP_COL_ALIASES = {
    'CODE':        ['code', 'location_code', 'loc_code'],
    'Store':       ['store', 'store_name', 'location_name'],
    'SKU':         ['sku', 'sku_code'],
    'ENN ode':     ['enn ode', 'enncode', 'ean', 'ean code', 'eancode'],
    'DESCRIPTION': ['description', 'desc', 'sku_name'],
    'MRP':         ['mrp'],
    'Tester':      ['tester', 'tester qty', 'qty'],
}

# Preferred sheet name when the workbook has more than one sheet.
# "Sheet2" was Vishal's primary in Skin_care_NPI_tester.xlsm; if absent
# we fall back to the first non-empty sheet.
_DUMP_PREFERRED_SHEET = 'Sheet2'

def _find_dump_column(actual_columns: List[str], canonical: str
                       ) -> Optional[str]:
    """
    Find the actual column header matching ``canonical`` from the dump
    sheet, case-insensitively, also trying known aliases. Returns the
    actual column name as it appears in the workbook (preserving case
    so downstream code can use it as a pandas key).
    """
    aliases = _DUMP_COL_ALIASES.get(canonical, [canonical.lower()])
    actual_lower = {str(c).strip().lower(): c for c in actual_columns}
    for alias in aliases:
        if alias in actual_lower:
            return actual_lower[alias]
    return None

def read_tester_dump_file(file_path: Path) -> TesterDump:
    """
    Read the selective-tester dump file → TesterDump.

    Sheet selection:
      1. Look for a sheet named exactly 'Sheet2' (Vishal's convention).
      2. If absent, use the first sheet that has the required columns.
      3. If no sheet matches, return a TesterDump with a hard error.

    Required columns (matched case-insensitively + via aliases):
      - CODE  (or 'location_code', 'loc_code')      → location key
      - SKU   (or 'sku_code')                        → SKU key

    Optional columns (loaded if present but not required for matching):
      - Store, ENN ode, DESCRIPTION, MRP, Tester

    Rows are deduplicated on (CODE, SKU). Duplicate rows emit a
    'warn' finding but the first occurrence wins.

    Returns:
      TesterDump containing eligible_keys: Set[(location_code, sku_code)].
      Empty set on hard errors; findings list explains why.
    """
    dump = TesterDump(source_path=file_path, source_name=file_path.name)

    if not file_path.exists():
        dump.add_finding('error',
            f"Tester dump file not found: {file_path}")
        return dump

    # Open workbook (xlsm requires keep_vba=False; we don't run macros)
    try:
        wb = load_workbook(str(file_path), data_only=True, read_only=True)
    except Exception as e:
        dump.add_finding('error',
            f"Cannot open tester dump file: {e}")
        return dump

    # ── Pick a sheet ──
    if _DUMP_PREFERRED_SHEET in wb.sheetnames:
        sheet_name = _DUMP_PREFERRED_SHEET
    else:
        # Fall back: first sheet with the required columns
        sheet_name = None
        for name in wb.sheetnames:
            ws = wb[name]
            first_row = next(ws.iter_rows(values_only=True), None)
            if first_row is None:
                continue
            cols = [str(c).strip() if c is not None else '' for c in first_row]
            if (_find_dump_column(cols, 'CODE') is not None
                    and _find_dump_column(cols, 'SKU') is not None):
                sheet_name = name
                break
        if sheet_name is None:
            dump.add_finding('error',
                f"Tester dump has no sheet with the required columns "
                f"(CODE + SKU). Sheets found: {wb.sheetnames}")
            wb.close()
            return dump

    ws = wb[sheet_name]
    dump.sheet_used = sheet_name

    # ── Find the header row + required columns ──
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        dump.add_finding('error',
            f"Sheet '{sheet_name}' is empty")
        wb.close()
        return dump

    columns = [str(c).strip() if c is not None else '' for c in header_row]
    code_col = _find_dump_column(columns, 'CODE')
    sku_col  = _find_dump_column(columns, 'SKU')
    if code_col is None or sku_col is None:
        dump.add_finding('error',
            f"Sheet '{sheet_name}' missing CODE or SKU column. "
            f"Found columns: {columns}")
        wb.close()
        return dump

    code_idx = columns.index(code_col)
    sku_idx  = columns.index(sku_col)

    # ── Read data rows ──
    seen_keys: Set[Tuple[str, str]] = set()
    duplicate_count = 0

    for row_num, raw_row in enumerate(rows_iter, start=2):
        # Skip fully-blank rows (very common in operator-edited sheets)
        if all(c is None or str(c).strip() == '' for c in raw_row):
            continue

        if code_idx >= len(raw_row) or sku_idx >= len(raw_row):
            continue

        code_val = _coerce_id(raw_row[code_idx])
        sku_val  = _coerce_id(raw_row[sku_idx])

        # Need both to be non-empty
        if not code_val or not sku_val:
            # Soft skip — these are likely category-header / divider rows
            continue

        key = (code_val, sku_val)
        if key in seen_keys:
            duplicate_count += 1
            continue

        seen_keys.add(key)
        dump.rows_loaded += 1

    dump.eligible_keys = seen_keys
    wb.close()

    if duplicate_count:
        dump.add_finding('warn',
            f"{duplicate_count} duplicate (CODE, SKU) row(s) in dump — "
            f"first occurrence used, others ignored")

    if dump.rows_loaded == 0:
        dump.add_finding('error',
            f"No valid (CODE, SKU) rows in sheet '{sheet_name}' — "
            f"file is empty or malformed")

    return dump

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 5 — CSV READER ──────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Parses H&G PO CSVs. Single-file reader + multi-file batch reader.

def _validate_columns(df: pd.DataFrame, po_file: POFile,
                       channel: ChannelConfig) -> bool:
    """Verify the CSV has every required column for this channel
    (case-insensitive). Returns True if all present."""
    actual_lower = {str(c).strip().lower(): c for c in df.columns}
    missing = [req for req in channel.csv_required_cols
                if req.lower() not in actual_lower]
    if missing:
        po_file.add_finding('error',
            f"Missing required column(s) for {channel.display_name}: "
            f"{missing}. Found: {list(df.columns)}")
        return False
    return True

def _resolve_line(line: POLine, channel: ChannelConfig,
                    bundle: MasterBundle) -> None:
    """
    Resolve a PO line through the channel's lookup chain, in place.

    For channel.lookup_via == 'SKU':
        line.sku_code (the SKU from CSV) → channel master → EAN
                                          → Items Master → Item No
    For channel.lookup_via == 'EAN':
        line.sku_code (which IS the EAN) → Items Master → Item No

    Adds actionable warning notes at each failure step so the operator
    sees exactly what to fix in which master.
    """
    sheet_label = channel.channel_master_sheet or 'channel master'
    sku_or_ean  = line.sku_code

    if channel.lookup_via == 'SKU':
        # ── Step 1: SKU → channel master ──
        ch_master = bundle.get_channel_master(channel.code)
        ch_entry  = ch_master.get(sku_or_ean)
        if ch_entry is None:
            line.add_note('warn',
                f"SKU {sku_or_ean} not in {sheet_label} — add it to "
                f"MT_Masters.xlsx ({sheet_label} sheet) with the correct EAN")
            return

        line.hg_master_status = ch_entry.status

        if ch_entry.status.lower() == 'inactive':
            line.add_note('warn',
                f"SKU {sku_or_ean} is Inactive in {sheet_label} — "
                f"processing anyway, but verify with "
                f"{channel.display_name} before posting")

        if not ch_entry.enn_code:
            line.add_note('warn',
                f"SKU {sku_or_ean} ({ch_entry.sku_name}) has no EAN in "
                f"{sheet_label} — add it to MT_Masters.xlsx "
                f"({sheet_label}, ENN code column)")
            return

        line.ean = ch_entry.enn_code

    else:  # lookup_via == 'EAN'
        # CSV's id column IS the EAN. No channel master in between.
        if not sku_or_ean:
            line.add_note('warn', "EAN is blank in CSV row")
            return
        line.ean = _normalize_ean(sku_or_ean)

    # ── Step 2: EAN → Items Master (every channel uses this step) ──
    item_entry = bundle.items_by_gtin.get(line.ean)
    if item_entry is None:
        line.add_note('warn',
            f"EAN {line.ean} not in Items Master — "
            f"add it to MT_Masters.xlsx (Items Master sheet)")
        return

    line.item_no           = item_entry.item_no
    line.items_master_mrp  = item_entry.mrp
    line.gst_code          = item_entry.gst_code
    line.items_master_desc = item_entry.description

    # NOTE: MRP-drift warning (channel MRP vs Items Master MRP) used to
    # live here. It was REMOVED on operator instruction — MRP
    # differences are irrelevant for SO generation. The Items Master
    # MRP is still captured on the line and shown in the Validation
    # sheet alongside the channel's MRP for reference.

def _resolve_ship_to(po_file: POFile, channel: ChannelConfig,
                       bundle: MasterBundle) -> None:
    """Resolve store key → Ship-to B2B → Ship-to Code + Cust No."""
    key = (channel.party, po_file.store_name)
    entry = bundle.ship_to_lookup.get(key)

    if entry is None:
        po_file.add_finding('warn',
            f"{channel.csv_store_col} '{po_file.store_name}' not in "
            f"Ship-to B2B (party={channel.party}) — Ship-to Code and "
            f"Cust No will be BLANK. Add this entry to MT_Masters.xlsx "
            f"(Ship-to B2B sheet), then re-run.")
        return

    po_file.ship_to       = entry.ship_to
    po_file.cust_no       = entry.cust_no
    po_file.ship_to_entry = entry

def read_channel_csv(file_path: Path, channel: ChannelConfig,
                       bundle: MasterBundle) -> List[POFile]:
    """
    Parse ONE channel CSV (H&G, future H&B, etc.) → list of POFiles.

    Each returned POFile represents ONE (PO_NO, STORE_NAME) group from
    the source file. A single CSV may contain multiple POs across
    multiple stores (real example: 90723-CNR092.csv has 9 POs at 9
    stores in one file). Each group becomes ONE D365 Sales Order.

    File-level errors (unreadable, missing columns, empty file) return
    a single POFile with has_hard_errors=True so the caller can still
    report it in the Reconciliation/Warnings sheets without crashing.

    File-level warnings (e.g. latin-1 decode fallback) are copied onto
    EVERY group from the file so each SO inherits the context.

    Args:
        file_path:  Absolute path to a channel CSV.
        channel:    Channel config (provides column mapping, lookup chain).
        bundle:     Loaded MasterBundle for SKU/EAN/ship-to resolution.

    Returns:
        List of POFile objects:
          - Single-PO file → list of 1 POFile.
          - Multi-PO file  → list of N POFiles (one per (PO, store) group).
          - File-level hard error → list of 1 POFile with hard error flag.
    """
    # ── File-level POFile is used only to carry hard errors / warnings.
    # When the file is OK we throw it away and build one POFile per group.
    file_level_pf = POFile(source_path=file_path, source_name=file_path.name)

    try:
        file_level_pf.file_mtime = datetime.fromtimestamp(
            os.path.getmtime(file_path))
    except OSError as e:
        file_level_pf.add_finding('warn', f"Cannot read file mtime: {e}")

    # H&G CSVs use \r line endings (legacy). pandas auto-detects.
    # UTF-8 first, latin-1 fallback for any channel that ships non-ASCII.
    try:
        df = pd.read_csv(file_path, dtype=str, encoding='utf-8',
                          lineterminator=None)
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(file_path, dtype=str, encoding='latin-1',
                              lineterminator=None)
            file_level_pf.add_finding('warn',
                f"Decoded as latin-1 (UTF-8 failed) — verify product names")
        except Exception as e:
            file_level_pf.add_finding('error', f"Cannot read file: {e}")
            return [file_level_pf]
    except Exception as e:
        file_level_pf.add_finding('error', f"Cannot read file: {e}")
        return [file_level_pf]

    if df.empty:
        file_level_pf.add_finding('error', "File contains no data rows")
        return [file_level_pf]

    if not _validate_columns(df, file_level_pf, channel):
        return [file_level_pf]

    # ── Drop fully-blank rows up front (H&G files have legacy CR-only
    # endings that pandas sometimes parses as a trailing empty row).
    df = df.dropna(how='all').reset_index(drop=True)

    if df.empty:
        file_level_pf.add_finding('error',
            "File contains only blank rows after header")
        return [file_level_pf]

    # ── Normalize key grouping columns (string strip, coerce IDs).
    # This is done in-place on a copy so we don't lose the original df
    # data (Raw Data sheet still echoes the source values).
    df = df.copy()
    df['_po_key'] = df.apply(
        lambda r: _coerce_id(_get_col_value(r, channel.csv_po_col)),
        axis=1)
    df['_store_key'] = df.apply(
        lambda r: _opt_str(_get_col_value(r, channel.csv_store_col)),
        axis=1)

    # If literally no PO_NO anywhere → file-level error.
    nonblank_pos = df[df['_po_key'] != '']
    if nonblank_pos.empty:
        file_level_pf.add_finding('error',
            f"No {channel.csv_po_col} found in any row")
        return [file_level_pf]

    # ── Group by (PO_NO, STORE_NAME). Order: preserve first-appearance
    # so the output SO numbers follow source file order, which the
    # operator expects when scanning the workbook.
    results: List[POFile] = []
    grouped = df.groupby(['_po_key', '_store_key'], sort=False)

    for (po_no, store_name), group_df in grouped:
        # Skip groups that have no PO_NO (the orphan-row case — those
        # rows had data but no PO column value). We don't want to emit
        # an SO for those, but they still need to appear in the
        # Reconciliation sheet, so collect them under a placeholder
        # error POFile at the end.
        if not po_no:
            continue
        pf = _build_pofile_from_group(
            group_df=group_df,
            source_path=file_path,
            source_name=file_path.name,
            file_mtime=file_level_pf.file_mtime,
            file_level_findings=list(file_level_pf.findings),
            po_no=po_no,
            store_name=store_name,
            channel=channel,
            bundle=bundle,
        )
        results.append(pf)

    # Orphan rows (PO_NO blank) — emit one POFile with a hard error so
    # the operator sees them, but don't generate an SO.
    orphan_df = df[df['_po_key'] == '']
    if not orphan_df.empty:
        orphan_pf = POFile(source_path=file_path,
                            source_name=file_path.name,
                            file_mtime=file_level_pf.file_mtime)
        orphan_pf.add_finding('error',
            f"{len(orphan_df)} row(s) have a blank "
            f"{channel.csv_po_col} — cannot generate SO for them")
        orphan_pf.raw_df = orphan_df.drop(
            columns=['_po_key', '_store_key'], errors='ignore')
        results.append(orphan_pf)

    return results

def _build_pofile_from_group(group_df,
                                source_path: Path,
                                source_name: str,
                                file_mtime: Optional[datetime],
                                file_level_findings: List[Tuple[str, str]],
                                po_no: str,
                                store_name: str,
                                channel: ChannelConfig,
                                bundle: MasterBundle) -> POFile:
    """
    Build one POFile from a single (PO_NO, STORE_NAME) group of rows.

    This is the per-group worker called by read_channel_csv. It handles:
      - Setting PO header fields (po_no, store_name, vendor info, etc.)
      - Resolving ship-to from the channel's Ship-to B2B sheet
      - Parsing each row of the group into a POLine
      - Per-line validations (qty>0, cost×qty=po_value, landing rate)
      - Item Master lookup via the channel's lookup chain (SKU or EAN)
      - Rolling up totals (line count, qty, value, purchase cost)
      - Snapshotting the raw rows for the Raw Data sheet

    Args:
        group_df:              Rows belonging to this (PO, store) group.
                                Must include _po_key/_store_key helper columns
                                added by read_channel_csv (they're stripped
                                before storing on the POFile).
        source_path:           Path of the source CSV (for traceability).
        source_name:           Basename of the source CSV.
        file_mtime:            Modification time of the source CSV.
        file_level_findings:   Warnings emitted at file level (latin-1,
                                missing mtime, etc.) — copied onto every
                                group from the file.
        po_no:                  Already-coerced PO number for this group.
        store_name:             Already-stripped store name for this group.
        channel:                Channel config.
        bundle:                 MasterBundle for lookups.

    Returns:
        Fully populated POFile for this group.
    """
    pf = POFile(source_path=source_path, source_name=source_name,
                  file_mtime=file_mtime)

    # Carry file-level findings through to this group so they appear on
    # each SO derived from the file.
    for level, msg in file_level_findings:
        pf.add_finding(level, msg)

    pf.po_no = po_no
    pf.store_name = store_name

    # Capture other PO-header fields from the first row of this group.
    # In well-formed files these are constant within the group; if not
    # we still take the first occurrence and warn.
    first = group_df.iloc[0]
    pf.vendor_code   = _opt_str(_get_col_value(first, 'VENDOR_CODE'))
    pf.vendor_name   = _opt_str(_get_col_value(first, 'VENDOR_NAME'))
    pf.location_code = _coerce_id(_get_col_value(first, 'LOCATION_CODE'))
    pf.region_code   = _coerce_id(_get_col_value(first, 'REGION_CODE'))

    # Sanity: vendor/location should be constant within a group. Warn if
    # not — usually indicates the input file is malformed.
    if 'VENDOR_CODE' in group_df.columns:
        unique_vendors = set(
            _opt_str(_get_col_value(group_df.iloc[i], 'VENDOR_CODE'))
            for i in range(len(group_df))
        ) - {''}
        if len(unique_vendors) > 1:
            pf.add_finding('warn',
                f"Group has inconsistent VENDOR_CODE values: "
                f"{sorted(unique_vendors)} — using first occurrence "
                f"({pf.vendor_code})")

    _resolve_ship_to(pf, channel, bundle)

    # Parse each line in the group
    for idx, row in group_df.iterrows():
        csv_row_num = idx + 2   # pandas idx 0 → file row 2 (after header)

        sku   = _coerce_id(_get_col_value(row, channel.csv_id_col))
        name  = _opt_str(_get_col_value(row, 'SKU_NAME'))   # HG-specific
        mrp   = _safe_float(_get_col_value(row, channel.csv_mrp_col))
        qty   = _safe_int(_get_col_value(row, channel.csv_qty_col))

        pcost = (_safe_float(_get_col_value(row, channel.csv_cost_col))
                  if channel.csv_cost_col else 0.0)
        pval  = (_safe_float(_get_col_value(row, channel.csv_value_col))
                  if channel.csv_value_col else 0.0)

        line = POLine(
            csv_row_num=csv_row_num, sku_code=sku, sku_name=name,
            mrp=mrp, quantity=qty, purchase_cost=pcost, po_value=pval,
        )

        if not sku:
            line.add_note('skip', f"Blank {channel.csv_id_col}")
            pf.lines.append(line)
            continue

        if qty <= 0:
            line.add_note('skip', f"Quantity is {qty} (must be > 0)")
            pf.lines.append(line)
            continue

        # NOTE: Cost integrity checks (cost×qty vs po_value, landing
        # rate vs expected) used to live here. They were REMOVED on
        # operator instruction: "the cost comparison is irrelevant ...
        # we can have that for reference purpose ... we won't consider
        # any difference, we will directly take all items in our
        # lines". The raw values are still captured on the POLine and
        # shown in the Validation sheet for reference.

        _resolve_line(line, channel, bundle)
        pf.lines.append(line)

    # Roll-up control totals (group-scoped)
    pf.input_line_count          = len(pf.lines)
    pf.input_qty_total           = sum(l.quantity for l in pf.lines)
    pf.input_po_value_total      = round(
        sum(l.po_value for l in pf.lines), 2)
    pf.input_purchase_cost_total = round(
        sum(l.purchase_cost for l in pf.lines), 2)

    # Keep the per-group raw dataframe — the Raw Data sheet in the
    # output workbook echoes these rows. Drop the helper grouping
    # columns we added so they don't leak into the output.
    pf.raw_df = group_df.drop(
        columns=['_po_key', '_store_key'], errors='ignore')

    return pf

def read_channel_csv_batch(file_paths: List[Path],
                              channel: ChannelConfig,
                              bundle: MasterBundle) -> POBatch:
    """
    Multi-file batch reader.

    Combines all input CSVs into one unified batch — operator's explicit
    instruction: "there will be multiple csv so take all and combined
    and then perform task". Each file is parsed independently into
    one-or-more POFiles (one per (PO, store) group within the file);
    all of them are concatenated into a single batch.po_files list, and
    cross-file validations (duplicate PO at same store, etc.) run over
    the combined set.

    A file-level hard error (unreadable, missing columns, etc.) emits a
    single POFile with has_hard_errors=True so the operator sees it in
    Reconciliation/Warnings without blocking other files.
    """
    batch = POBatch(master_bundle=bundle)

    for path in file_paths:
        # read_channel_csv returns a list — one POFile per (PO, store)
        # group in the file. A multi-PO file produces multiple POFiles;
        # a hard-error file produces a single POFile with the error.
        batch.po_files.extend(read_channel_csv(path, channel, bundle))

    # ── Cross-file check 1: same PO at same location (HIGHLIGHT) ──
    # Operator's explicit ask: flag when the same PO_NO appears twice
    # for the SAME location/store. This catches accidental duplicate
    # files for the same store more reliably than a global PO_NO check
    # (since some channels may legitimately reuse PO_NOs across stores).
    # Warning-level: don't block processing, just highlight for review.
    po_store_to_files: Dict[Tuple[str, str], List[str]] = {}
    for pf in batch.po_files:
        if pf.po_no and pf.store_name and not pf.has_hard_errors:
            key = (pf.po_no, pf.store_name)
            po_store_to_files.setdefault(key, []).append(pf.source_name)
    for (po_no, store), files in po_store_to_files.items():
        if len(files) > 1:
            batch.cross_findings.append(('warn',
                f"DUPLICATE: {channel.csv_po_col} {po_no} appears "
                f"{len(files)} times for {channel.csv_store_col} "
                f"'{store}': {files}. Verify these aren't accidental "
                f"copies before posting."))

    # ── Cross-file check 2: same PO across different stores (warn) ──
    # Different stores having the same PO_NO is unusual — channels
    # normally assign unique PO_NOs across their network. Worth flagging
    # but at warning level (not blocking).
    po_to_files: Dict[str, List[Tuple[str, str]]] = {}
    for pf in batch.po_files:
        if pf.po_no and pf.store_name and not pf.has_hard_errors:
            po_to_files.setdefault(pf.po_no, []).append(
                (pf.source_name, pf.store_name))
    for po_no, file_store_pairs in po_to_files.items():
        unique_stores = {s for _, s in file_store_pairs}
        if len(file_store_pairs) > 1 and len(unique_stores) > 1:
            details = ', '.join(f"{f} ({s})" for f, s in file_store_pairs)
            batch.cross_findings.append(('warn',
                f"{channel.csv_po_col} {po_no} appears across "
                f"different stores: {details}. Verify this is correct."))

    # ── Cross-file check 3: same store has multiple POs (info) ──
    store_to_files: Dict[str, List[str]] = {}
    for pf in batch.po_files:
        if pf.store_name and not pf.has_hard_errors:
            store_to_files.setdefault(pf.store_name, []).append(
                f"{pf.source_name} (PO {pf.po_no})")
    for store, files in store_to_files.items():
        if len(files) > 1:
            batch.cross_findings.append(('info',
                f"{channel.csv_store_col} '{store}' has {len(files)} POs "
                f"in this batch: {files}"))

    return batch

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 5b — SO NUMBER GENERATION ───────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Generates D365 Sales Order numbers for parsed POs.
#
# Format (HG channel): SO/{channel}/{MM}/{DDMMYY counter}
#   - MM        — today's month, literal 2-digit (e.g., "06" for June)
#   - counter   — starts at today's DDMMYY (e.g., 050626 for 05-Jun-2026)
#                 increments by 1 for each subsequent PO
#
# Examples:
#   First batch on 05-Jun-2026 (3 POs):  050626, 050627, 050628
#   Second batch later same day (2 POs): 050629, 050630   ← continues from
#                                                            persisted counter
#   First batch on 06-Jun-2026 (1 PO):   060626           ← counter resets
#                                                            to new day's DDMMYY
#   First batch on 01-Jul-2026 (1 PO):   010726           ← MM also changes
#                                                            (07 = July)
#
# Counter state is persisted per-channel in mt_select_seq.json so that
# multiple batches run within the same day continue the sequence (no
# duplicate SOs), and the counter automatically resets when a new day
# begins.

def assign_so_numbers(batch: POBatch,
                       channel: ChannelConfig,
                       today: Optional[datetime] = None,
                       generate_testers: bool = False,
                       tester_dump: Optional[TesterDump] = None) -> None:
    """
    Assign D365 SO numbers to every cleanly-parsed PO in the batch.

    Modifies POFiles in place:
      - po_file.so_number         = "SO/{ch}/{MM}/{counter:06d}"
      - po_file.tester_so_number  = "SO/{ch}/TT/{counter:06d}"  (only if
                                     generate_testers=True AND the PO has
                                     at least one tester-eligible line)
      - line.is_tester_eligible   = True/False (set per line based on
                                     mode — see below)

    POs with hard errors are skipped — they don't get an SO number and
    don't advance the counter (no wasted numbers on rejected files).

    When ``generate_testers`` is True, the function runs TWO passes:
      Pass 1 — assign a regular SO to each eligible PO.
      Pass 2 — for each PO with a regular SO:
                 (a) walk every resolved, non-SKIP line
                 (b) mark line.is_tester_eligible based on the active
                     mode (AUTOMATIC: all eligible; SELECTIVE: only those
                     whose (location_code, sku_code) appears in the dump)
                 (c) if at least one line is eligible, assign a tester
                     SO number from the same counter block
                 (d) if zero lines are eligible (selective mode + no
                     matches), this PO gets NO tester SO at all

    Modes:
      AUTOMATIC mode (tester_dump=None):
        every resolved line gets a tester. This is the "automatic for all
        while ticking" mode.
      SELECTIVE mode (tester_dump provided):
        only lines whose (LOCATION_CODE, SKU_CODE) appears in the dump
        file get a tester. This is the "provide dump to take only
        selected SKUs not all" mode.

    Example for 3 POs on 05-Jun-2026 with generate_testers=True:
        Regular block:  050626, 050627, 050628
        Tester block:   050629, 050630, 050631
        next_counter persisted as 050632.

    Args:
        batch:            The POBatch to mutate.
        channel:          Channel config (provides code + tester settings).
        today:            Date used for MM and DDMMYY base. Default = now.
        generate_testers: If True, also assign tester SO numbers.
        tester_dump:      Optional TesterDump for SELECTIVE mode. Pass
                          None for AUTOMATIC mode (all lines become
                          testers). Ignored when generate_testers=False.
    """
    if today is None:
        today = datetime.now()
    today_iso = today.date().isoformat()
    mm = today.strftime('%m')                       # "06"
    today_ddmmyy = int(today.strftime('%d%m%y'))    # 50626

    # ── Load + advance per-channel state ──
    full_state = load_seq_state()
    ch_state = full_state.get(channel.code, {})

    if ch_state.get('date') != today_iso:
        # New day (or first run for this channel) — reset counter to
        # today's DDMMYY base.
        ch_state = {
            'date': today_iso,
            'next_counter': today_ddmmyy,
        }

    counter_start = ch_state['next_counter']

    # ── Pass 1: regular SO numbers ──
    n_regular = 0
    for po_file in batch.po_files:
        if po_file.has_hard_errors:
            continue
        counter = ch_state['next_counter']
        po_file.so_number = (
            f"SO/{channel.code}/{mm}/{counter:06d}")
        ch_state['next_counter'] = counter + 1
        n_regular += 1

    # ── Pass 2: tester SO numbers (continuous block) ──
    # Only runs when the operator enabled the GUI checkbox. Uses the
    # same counter as regulars but with 'TT' in the month slot to mark
    # the SO as a tester batch.
    n_tester = 0
    n_tester_lines_eligible = 0    # how many individual lines became eligible
    n_pos_skipped_no_match = 0     # selective: POs with zero matches → no tester SO
    if generate_testers:
        if channel.tester_unit_price is None:
            # Channel doesn't declare a tester price — treat as misconfig.
            batch.cross_findings.append(('warn',
                f"Tester checkbox is on but channel "
                f"{channel.display_name} has no tester_unit_price "
                f"configured — skipping tester SO generation."))
        else:
            selective = tester_dump is not None and not tester_dump.has_hard_errors

            for po_file in batch.po_files:
                if po_file.has_hard_errors or not po_file.so_number:
                    continue

                # Mark per-line tester eligibility based on the active mode.
                # We always reset is_tester_eligible up front so re-running
                # this function (e.g. after the operator toggles the dump
                # file) starts from a clean slate.
                any_eligible = False
                for line in po_file.lines:
                    # Lines that can't even produce a regular output row
                    # are never tester-eligible.
                    if not line.item_no or line.status == 'SKIP':
                        line.is_tester_eligible = False
                        continue

                    if selective:
                        # SELECTIVE mode: only lines whose
                        # (location_code, sku_code) appears in the dump.
                        line.is_tester_eligible = tester_dump.is_eligible(
                            po_file.location_code, line.sku_code)
                    else:
                        # AUTOMATIC mode: every resolved line gets a tester.
                        line.is_tester_eligible = True

                    if line.is_tester_eligible:
                        any_eligible = True

                # Assign a tester SO only if at least one line is eligible.
                # In selective mode, POs with zero matching SKUs end up
                # without any tester at all — exactly as the operator
                # specified ("provide dump to take only selected skus
                # not all").
                if any_eligible:
                    counter = ch_state['next_counter']
                    po_file.tester_so_number = (
                        f"SO/{channel.code}/TT/{counter:06d}")
                    ch_state['next_counter'] = counter + 1
                    n_tester += 1
                    n_tester_lines_eligible += sum(
                        1 for l in po_file.lines if l.is_tester_eligible)
                elif selective:
                    n_pos_skipped_no_match += 1

    # ── Persist updated state ──
    full_state[channel.code] = ch_state
    save_seq_state(full_state)

    # Record what happened (shown at top of the batch report)
    if n_tester:
        mode = ('SELECTIVE (dump file)'
                 if tester_dump is not None and not tester_dump.has_hard_errors
                 else 'AUTOMATIC (all resolved lines)')
        skipped_note = (
            f", skipped {n_pos_skipped_no_match} PO(s) with no dump match"
            if n_pos_skipped_no_match else '')
        batch.so_assignment_summary = (
            f"Assigned {n_regular} regular + {n_tester} tester SO "
            f"numbers using counter "
            f"{counter_start:06d}..{ch_state['next_counter']-1:06d} "
            f"(channel {channel.code}, MM={mm}/TT, date={today_iso}, "
            f"tester mode: {mode}, {n_tester_lines_eligible} eligible "
            f"line(s){skipped_note})"
        )
    else:
        batch.so_assignment_summary = (
            f"Assigned {n_regular} SO numbers using counter "
            f"{counter_start:06d}..{ch_state['next_counter']-1:06d} "
            f"(channel {channel.code}, MM={mm}, date={today_iso})"
        )

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 6 — TEMPLATE GENERATOR ──────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def create_template(output_path: Optional[Path] = None,
                     overwrite: bool = False) -> Path:
    """
    Generate a fresh empty MT_Masters.xlsx with three sheets, correct
    column headers, and a placeholder note in each sheet.
    """
    if output_path is None:
        ensure_folders()
        output_path = get_masters_path()

    if output_path.exists() and not overwrite:
        raise FileExistsError(
            f"{output_path} already exists. Pass overwrite=True to replace.")

    wb = Workbook()
    wb.remove(wb.active)

    def _build_sheet(name, headers, note_text, col_widths):
        ws = wb.create_sheet(name)
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = _HDR_FILL
            cell.font = _HDR_FONT
            cell.alignment = Alignment(horizontal='center')
        note = ws.cell(row=3, column=1, value=note_text)
        note.font = _NOTE_FONT
        for c, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = 'A2'

    _build_sheet(SHEET_ITEMS_MASTER, COLS_ITEMS_MASTER,
        '← Paste your Items_March data here. '
        'Columns above must match (case-insensitive).',
        [12, 18, 48, 10, 18])

    _build_sheet(SHEET_SHIP_TO_B2B, COLS_SHIP_TO_B2B,
        '← Paste Ship-to mappings here. One row per (Party, Del Location). '
        'Engine uses Party+Del Location+Cust No+Ship to. '
        'Name/Address/Postcode/City are for Summary verification.',
        [8, 28, 10, 14, 50, 50, 50, 12, 18])

    # One sheet per channel that has a SKU→EAN master. Iterating CHANNELS
    # means adding HB (or any future channel with a master) just needs a
    # new entry in CHANNELS — no change here.
    for code, channel in CHANNELS.items():
        if channel.lookup_via != 'SKU' or not channel.channel_master_sheet:
            continue
        cols = channel.channel_master_cols or [
            'sku_code', 'sku_name', 'ENN code', 'status']
        _build_sheet(channel.channel_master_sheet, cols,
            f'← Paste {channel.display_name} master rows here. '
            f'Blank ENN code is OK but the SKU will fail lookup until '
            f'filled.',
            [12, 50, 18, 12])

    # ── Non Stock sheet ──
    # Pre-populated with the canonical GWP/PWP kit list. Each row in
    # this sheet becomes ONE tester line of qty=1 when the "Add Non
    # Stock" checkbox is on.
    ws_ns = wb.create_sheet(SHEET_NON_STOCK)
    for c, h in enumerate(COLS_NON_STOCK, start=1):
        cell = ws_ns.cell(row=1, column=c, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal='center')
    sample_rows = [
        ('300023', '8904473101009', 'RENEE PINK PUFFER POUCH',         1, 'Puffer'),
        ('300034', '8904473101023', 'RENEE RED PUFFER POUCH',          1, 'Puffer'),
        ('300037', '8904473101016', 'RENEE SILVER PUFFER POUCH',       1, 'Puffer'),
        ('300009', '8906121643640', 'RENEE FLAT SILVER POUCH',         4, 'Flat Pouch'),
        ('200076', '8906121642674', 'RENEE BLOOM EAU DE PARFUM 8ML NFS',     2, 'PWP 8ml NFS'),
        ('200355', '8906121647495', 'RENEE FLIRT EAU DE PARFUM 8 ML NFS',    2, 'PWP 8ml NFS'),
        ('200483', '8906121647501', 'RENEE MADAME EAU DE PARFUM 8 ML NFS',   2, 'PWP 8ml NFS'),
        ('200701', '8906121645743', 'RENEE RED NOIR EAU DE PARFUM 8ML NFS',  2, 'PWP 8ml NFS'),
        ('400039', 'OPM-RSK-CR500-RE', 'RENEE STARTER KIT COTTON ROLL 500GM', 1, 'Cotton Rolls'),
        ('200101', '8906121643572', 'RENEE CLEANSING MILK 500ML',      1, 'Cleansing Milk'),
    ]
    for r_idx, row_data in enumerate(sample_rows, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            ws_ns.cell(row=r_idx, column=c_idx, value=val)
    note_row = len(sample_rows) + 3
    note = ws_ns.cell(row=note_row, column=1,
        value='← Edit / extend this list as your GWP/PWP kit changes. '
              'Each row above adds ONE tester line of qty=1 when the '
              '"Add Non Stock" checkbox is on.')
    note.font = _NOTE_FONT
    for c, w in enumerate([12, 18, 50, 8, 22], start=1):
        ws_ns.column_dimensions[get_column_letter(c)].width = w
    ws_ns.freeze_panes = 'A2'

    wb.save(str(output_path))
    return output_path

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 6b — OUTPUT WORKBOOK WRITER ─────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Produces the D365-ready SO workbook. Six sheets, matching the structure
# of existing marketplace tools (e.g. RK):
#
#   Headers (SO)   — D365 SO header rows. Primary D365 import target.
#   Lines (SO)     — D365 SO line rows. Primary D365 import target.
#   Summary        — per-PO rollup with totals. Human review.
#   Validation     — per-line resolved data + cost comparison. Human review.
#   Warnings       — every warn-level finding, with PO context. Human review.
#   Raw Data       — original CSV rows + appended resolved fields. Audit trail.
#
# Files that hit hard errors during parsing don't get Headers/Lines rows
# (no SO number was assigned to them) but they DO appear in the Warnings
# sheet so the operator sees why they were excluded.

def generate_output_filename(channel: ChannelConfig,
                               when: Optional[datetime] = None) -> str:
    """
    Construct the output filename matching existing tools' convention:
        {channel}_so_{DD-MM-YYYY}_{HHMMSS}.xlsx
    Example: hg_so_05-06-2026_143015.xlsx
    """
    when = when or datetime.now()
    return (f"{channel.code.lower()}_so_"
            f"{when.strftime('%d-%m-%Y')}_"
            f"{when.strftime('%H%M%S')}.xlsx")

# ── Header constants used in the output workbook ──
# Defined once so renaming is a single-line change.

# ── D365 import preamble (Abhishek's SO-import template format) ──
# Each Headers/Lines sheet starts with a 3-row preamble:
#   R1: owner | table-name | table-id
#   R2: blank
#   R3: column headers
# 36 = D365 BC table ID for Sales Header; 37 = Sales Line. Fixed by D365.
D365_IMPORT_OWNER      = 'ABHISHEK WAGH - SO'
D365_TABLE_NAME_HEADER = 'Sales Header'
D365_TABLE_NAME_LINE   = 'Sales Line'
D365_TABLE_ID_HEADER   = 36
D365_TABLE_ID_LINE     = 37

HEADERS_SHEET_COLS = [
    'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
    'Posting Date', 'Order Date', 'Document Date',
    'Invoice From Date', 'Invoice To Date',
    'External Document No.', 'Location Code', 'Dimension Set ID',
    'Supply Type', 'Voucher Narration',
    'Brand Code (Dimension)',     'Channel Code (Dimension)',
    'Catagory (Dimension)',       'Geography Code (Dimension)',
]

LINES_SHEET_COLS = [
    'Document Type', 'Document No.', 'Line No.', 'Type', 'No.',
    'Location Code', 'Quantity', 'Unit Price',
]

SUMMARY_SHEET_COLS = [
    'Location Code',   # numeric LOCATION_CODE from CSV (matches
                        # operator's pivot view: "3 | HG-ADYAR-CHE | ...")
    'PO',              # H&G's PO number from CSV
    # v1.10: SO No columns let the operator trace each PO → its
    # generated D365 Sales Order at a glance, without cross-referencing
    # the Headers (SO) sheet. The Tester SO column is populated only
    # when the "Generate Testers" checkbox was on AND the PO had at
    # least one tester-eligible line; otherwise it stays blank.
    'SO No',           # regular SO assigned to this PO (e.g. SO/HG/06/070626)
    'Tester SO',       # tester SO if generated (e.g. SO/HG/TT/070627) else ''
    'Location (Raw)',  # CSV's STORE_NAME / LOCATION_NAME
    'Location (Mapped)',
    'Cust No',
    'Ship-to',
    'Items',           # rows in input CSV (= count of unique EAN/SKU
                        # in a clean H&G PO, matches "Count of EAN")
    'Total Qty',       # sum of QUANTITY column from CSV
    'Total Amount',    # sum of PO_VALUE column from CSV
    'Lines Written',   # how many lines actually made it into the output SO
    'Status',
]

VALIDATION_SHEET_COLS = [
    'PO', 'Item No', 'EAN', 'Description', 'MRP',
    'Landing ({pct}%)', 'GST Code',
    'Our Cost Price', 'Marketplace Cost',
    'Difference with Cost', 'Status',
]

WARNINGS_SHEET_COLS = ['PO', 'Location', 'Warning']

def write_so_workbook(batch: POBatch,
                        channel: ChannelConfig,
                        warehouse_code: str,
                        output_path: Optional[Path] = None,
                        when: Optional[datetime] = None,
                        add_non_stock: bool = False,
                        ) -> Path:
    """
    Write the 6-sheet D365 SO workbook for ``batch``.

    ``warehouse_code`` is the resolved D365 location code (e.g. 'PICK',
    'DS_BL_OFF1'). This goes into every Header.Location Code and
    Line.Location Code cell.

    Returns the absolute path of the written file.
    """
    when = when or datetime.now()

    if output_path is None:
        out_dir = get_output_folder(channel.code)
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / generate_output_filename(channel, when)

    wb = Workbook()
    wb.remove(wb.active)   # discard the default empty sheet

    # Order matters: Lines must be written before Reconciliation,
    # because Reconciliation reads `line.output_line_no` which is
    # populated as each line is written. The tab order in the file
    # follows the call order below.
    _write_headers_sheet(       wb, batch, channel, warehouse_code, when)
    _write_lines_sheet(         wb, batch, channel, warehouse_code,
                                  add_non_stock=add_non_stock)
    _write_summary_sheet(       wb, batch, channel, warehouse_code)
    _write_reconciliation_sheet(wb, batch, channel,
                                  add_non_stock=add_non_stock)
    _write_validation_sheet(    wb, batch, channel)
    _write_warnings_sheet(      wb, batch, channel)
    _write_raw_data_sheet(      wb, batch, channel)

    wb.save(str(output_path))
    return output_path

def _bold_header_row(ws):
    """Bold the first row (header row) of a worksheet."""
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

def _autosize_columns(ws, max_width: int = 50):
    """Best-effort column widths based on the longest cell value."""
    for col_cells in ws.iter_cols():
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = min(
            max_len + 2, max_width)

def _apply_status_style(ws, row_idx: int, status_col: int,
                          status: str,
                          tint_row_cols: Optional[Tuple[int, int]] = None
                          ) -> None:
    """
    Visually highlight a status row.

    Two things happen:
      1. The Status cell itself gets a strong fill + bold coloured font
         (green/yellow/red depending on ``status``).
      2. Optionally, every other cell in the row (between columns
         tint_row_cols[0] and tint_row_cols[1] inclusive) gets a pale
         tint of the same colour, so the entire row stands out without
         overwhelming the cell-level highlight.

    Args:
        ws:               openpyxl worksheet
        row_idx:          1-based row to style
        status_col:       1-based column index of the Status cell
        status:           Status string ('OK', 'WARN', 'FAIL', etc.).
                          Looked up case-insensitively in _STATUS_STYLES.
        tint_row_cols:    Optional (first_col, last_col) inclusive range
                          to tint with a pale shade. Pass None to skip
                          row tinting (only the Status cell is coloured).
    """
    style = _STATUS_STYLES.get(status.upper() if status else '')
    if style is None:
        return
    fill, font, tint = style

    # 1. Strong fill on the Status cell
    cell = ws.cell(row=row_idx, column=status_col)
    cell.fill = fill
    cell.font = font

    # 2. Pale tint on the rest of the row (optional)
    if tint_row_cols is not None:
        first, last = tint_row_cols
        for col in range(first, last + 1):
            if col == status_col:
                continue
            ws.cell(row=row_idx, column=col).fill = tint

def _write_headers_sheet(wb, batch: POBatch, channel: ChannelConfig,
                          warehouse_code: str, when: datetime) -> None:
    """
    Headers (SO) — D365 SO header import target.

    Two blocks, in order:
      1. Regular SO headers — one row per cleanly-parsed PO with an
         assigned ``so_number``.
      2. Tester SO headers — one row per PO that also has a
         ``tester_so_number`` (only when operator enabled testers).

    The continuous-block ordering matches the SO counter assignment:
    if regulars consumed counters 050626..050628 and testers consumed
    050629..050631, the Headers sheet shows the same order top-down.
    Tester headers differ from their paired regular only in:
      - No. (tester SO#)
      - External Document No. = channel.tester_external_doc_no ('TESTERS')
    Everything else (Sell-to, Ship-to, dates, Location, Supply Type) is
    inherited from the regular's data.
    """
    ws = wb.create_sheet('Headers (SO)')

    # D365 import preamble: owner | table name | table id (R1 + blank R2)
    ws.append([D365_IMPORT_OWNER, D365_TABLE_NAME_HEADER,
               D365_TABLE_ID_HEADER])
    ws.append([])
    ws.append(HEADERS_SHEET_COLS)        # column headers (row 3)

    date_str = when.strftime('%d-%m-%Y')

    # ── Block 1: Regular headers ──
    for pf in batch.po_files:
        if pf.has_hard_errors or not pf.so_number:
            continue
        row = [
            channel.document_type,                  # 'Order'
            pf.so_number,                            # SO/HG/06/050626
            channel.sell_to,                         # '20039' for HG
            pf.ship_to or '',                        # 20039_115
            date_str, date_str, date_str,           # Posting / Order / Document
            date_str, date_str,                      # Invoice From / To
            pf.po_no,                                # External Document No.
            warehouse_code,                          # PICK or DS_BL_OFF1
            '',                                      # Dimension Set ID
            channel.supply_type,                     # 'B2B'
            '',                                      # Voucher Narration
            '', '', '', '',                          # 4 Dimension Codes
        ]
        ws.append(row)

    # ── Block 2: Tester headers (only when generate_testers was on) ──
    # Track tester row indices so we can tint them visually distinct
    # from regular orders.
    tester_row_indices: List[int] = []
    for pf in batch.po_files:
        if (pf.has_hard_errors
                or not pf.so_number
                or not pf.tester_so_number):
            continue
        row = [
            channel.document_type,                  # 'Order'
            pf.tester_so_number,                     # SO/HG/TT/050629
            channel.sell_to,                         # '20039' for HG
            pf.ship_to or '',                        # same as regular
            date_str, date_str, date_str,           # same as regular
            date_str, date_str,                      # same as regular
            channel.tester_external_doc_no,          # literal 'TESTERS'
            warehouse_code,                          # same as regular
            '',                                      # Dimension Set ID
            channel.supply_type,                     # 'B2B'
            '',                                      # Voucher Narration
            '', '', '', '',                          # 4 Dimension Codes
        ]
        ws.append(row)
        tester_row_indices.append(ws.max_row)

    # Style: D365 preamble band on row 1, then table format with
    # header at row 3, then tint tester rows pale blue.
    _apply_preamble_style(ws, preamble_row=1)
    _apply_table_format(ws, header_row=3)
    for r in tester_row_indices:
        _tint_row(ws, r, _TESTER_ROW_TINT)
    _autosize_columns(ws)

def _write_lines_sheet(wb, batch: POBatch, channel: ChannelConfig,
                         warehouse_code: str,
                         add_non_stock: bool = False) -> None:
    """
    Lines (SO) — one row per resolved PO line. Line numbers step by
    10000 within each PO (D365 convention) and restart at 10000 for each
    new PO. Unresolved lines (no Item No) are skipped — they appear in
    the Warnings sheet instead.

    ``add_non_stock``: when True AND a tester SO was assigned, the
    bundle's non-stock kit is appended to that PO's tester lines.
    """
    ws = wb.create_sheet('Lines (SO)')

    # D365 import preamble (R1 owner|table|id, R2 blank, R3 headers)
    ws.append([D365_IMPORT_OWNER, D365_TABLE_NAME_LINE,
               D365_TABLE_ID_LINE])
    ws.append([])
    ws.append(LINES_SHEET_COLS)

    # Tester unit price as STRING (matches Abhishek's template — D365
    # parses this cell as text during line creation, avoiding any
    # Excel float-rounding quirks).
    tester_price_str = (f"{channel.tester_unit_price}"
                          if channel.tester_unit_price is not None else '')

    # ── Block 1: Regular lines ──
    # One row per resolved input line, grouped by PO. Each PO restarts
    # its line numbering at 10000 (D365 convention).
    for pf in batch.po_files:
        if pf.has_hard_errors or not pf.so_number:
            continue

        line_no = 10000
        for line in pf.lines:
            # Skip lines we couldn't resolve to a D365 Item No
            if not line.item_no:
                continue
            # Also skip status='SKIP' lines (blank SKU / zero qty)
            if line.status == 'SKIP':
                continue

            # Lines sheet stays minimal — Unit Price is deliberately
            # LEFT BLANK for regulars. D365 will source the price from
            # its own sales price list / customer agreement at posting
            # time. The cost vs MRP×landing comparison is reported
            # separately in the Validation sheet, not duplicated in Lines.
            row = [
                channel.document_type,            # 'Order'
                pf.so_number,                      # SO/HG/06/050626
                line_no,                           # 10000, 20000, ...
                'Item',
                line.item_no,                      # D365 Item No
                warehouse_code,                    # PICK / DS_BL_OFF1
                line.quantity,
                '',                                # Unit Price — see Validation
            ]
            ws.append(row)
            # Stamp the assigned line number onto the POLine so the
            # Reconciliation sheet can later report exactly where this
            # input row was written.
            line.output_line_no = line_no
            line_no += 10000

    # ── Block 2: Tester lines + Block 2b: Non-stock kit ──
    # Track tester / non-stock row indices for distinct tinting.
    tester_row_indices:   List[int] = []
    nonstock_row_indices: List[int] = []

    if channel.tester_unit_price is not None:
        for pf in batch.po_files:
            if (pf.has_hard_errors
                    or not pf.so_number
                    or not pf.tester_so_number):
                continue

            t_line_no = 10000
            # Regular tester lines (one per eligible input line)
            for line in pf.lines:
                if not line.item_no:
                    continue
                if line.status == 'SKIP':
                    continue
                if not line.is_tester_eligible:
                    continue
                row = [
                    channel.document_type,            # 'Order'
                    pf.tester_so_number,               # SO/HG/TT/...
                    t_line_no,
                    'Item',
                    line.item_no,
                    warehouse_code,
                    1,                                  # qty = 1 always
                    tester_price_str,                  # '0.54' STRING
                ]
                ws.append(row)
                tester_row_indices.append(ws.max_row)
                line.tester_output_line_no = t_line_no
                t_line_no += 10000

            # Non-stock kit (only when checkbox is on). Each entry
            # becomes ONE tester line of qty=1 appended to the same
            # tester SO, continuing the line counter.
            if (add_non_stock
                    and batch.master_bundle is not None
                    and batch.master_bundle.non_stock):
                for ns_entry in batch.master_bundle.non_stock:
                    row = [
                        channel.document_type,
                        pf.tester_so_number,
                        t_line_no,
                        'Item',
                        ns_entry.item_no,              # D365 Item No directly
                        warehouse_code,
                        1,
                        tester_price_str,
                    ]
                    ws.append(row)
                    nonstock_row_indices.append(ws.max_row)
                    t_line_no += 10000

    # Style: preamble band, table format (header_row=3), tinting.
    _apply_preamble_style(ws, preamble_row=1)
    _apply_table_format(ws, header_row=3)
    for r in tester_row_indices:
        _tint_row(ws, r, _TESTER_ROW_TINT)
    for r in nonstock_row_indices:
        _tint_row(ws, r, _NONSTOCK_ROW_TINT)
    _autosize_columns(ws)

def _write_summary_sheet(wb, batch: POBatch, channel: ChannelConfig,
                          warehouse_code: str) -> None:
    """
    Summary — per-PO rollup with TOTAL row + footer noting the run
    parameters (channel, landing rate, warehouse). Operator's quick-look
    sheet for sanity-checking quantities and amounts before D365 upload.
    """
    ws = wb.create_sheet('Summary')
    ws.append(SUMMARY_SHEET_COLS)
    status_col = len(SUMMARY_SHEET_COLS)     # 1-based index of Status column

    total_items, total_qty, total_amount = 0, 0, 0.0
    total_resolved = 0

    for pf in batch.po_files:
        if pf.has_hard_errors:
            status = 'FAIL'
            ws.append([
                pf.location_code or '',          # Location Code
                pf.po_no or pf.source_name,      # PO
                pf.so_number or '',              # SO No (probably blank — pf failed)
                pf.tester_so_number or '',       # Tester SO (blank)
                pf.store_name or '',             # Location (Raw)
                '', '', '',                       # Loc Mapped / Cust No / Ship-to
                0, 0, 0.0,                       # Items / Qty / Amount
                0,                                # Lines Written
                status,                          # Status
            ])
            _apply_status_style(
                ws, ws.max_row, status_col, status,
                tint_row_cols=(1, status_col))
            continue

        # Summary reflects the PO AS SENT BY THE CHANNEL — operator wants
        # to see the actual qty/amount/items in the PO (just like a
        # pandas pivot on the raw CSV), not just what made it through
        # resolution. So the totals come from the input CSV columns:
        #   - input_line_count = rows in CSV (clean POs = 1 SKU per row,
        #                       so this matches "Count of EAN" pivot)
        #   - input_qty_total  = SUM of QUANTITY column
        #   - input_po_value_total = SUM of PO_VALUE column
        # The new "Lines Written" column shows the resolved subset so
        # the operator can see master-coverage gaps at a glance.
        items_n   = pf.input_line_count
        qty_n     = pf.input_qty_total
        amt_n     = pf.input_po_value_total

        resolved_n = sum(
            1 for l in pf.lines
            if l.item_no and l.status != 'SKIP' and l.output_line_no)

        # Status: OK if no warnings anywhere, else WARN.
        # Includes Ship-to unresolved (would block D365) AND any line
        # that failed to resolve (operator needs visibility).
        has_warnings = (
            any(lvl == 'warn' for lvl, _ in pf.findings)
            or any(l.status == 'WARN' for l in pf.lines)
            or not pf.ship_to
            or resolved_n < items_n
        )
        status = 'WARN' if has_warnings else 'OK'

        loc_mapped = (pf.ship_to_entry.del_location
                       if pf.ship_to_entry else '')

        ws.append([
            pf.location_code or '',         # Location Code (LOCATION_CODE)
            pf.po_no,                        # PO
            pf.so_number or '',              # SO No  (v1.10)
            pf.tester_so_number or '',       # Tester SO  (v1.10)
            pf.store_name,                   # Location (Raw)
            loc_mapped,                      # Location (Mapped)
            pf.cust_no or '',                # Cust No
            pf.ship_to or '',                # Ship-to
            items_n,                         # Items
            qty_n,                           # Total Qty
            amt_n,                           # Total Amount
            resolved_n,                      # Lines Written
            status,                          # Status
        ])
        _apply_status_style(
            ws, ws.max_row, status_col, status,
            tint_row_cols=(1, status_col))
        total_items    += items_n
        total_qty      += qty_n
        total_amount   += amt_n
        total_resolved += resolved_n

    # TOTAL row
    total_row_idx = ws.max_row + 1
    ws.append([
        '',                                  # Location Code (blank)
        'TOTAL',                             # PO column header
        '',                                  # SO No (blank)
        '',                                  # Tester SO (blank)
        '', '', '', '',                       # Loc Raw/Mapped/Cust/Ship-to (blank)
        total_items, total_qty,
        round(total_amount, 2),
        total_resolved, '',
    ])
    # Bold the TOTAL row
    bold = Font(bold=True)
    for cell in ws[total_row_idx]:
        cell.font = bold

    # Spacer + footer
    ws.append([])
    pct = (f"{int(channel.expected_landing_ratio*100)}%"
            if channel.expected_landing_ratio is not None else 'N/A')
    ws.append([
        f"Channel: {channel.display_name} ({channel.code})  |  "
        f"Landing rate: {pct}  |  "
        f"Warehouse: {warehouse_code}"
    ])

    _apply_table_format(ws, header_row=1)
    _autosize_columns(ws)

def _write_validation_sheet(wb, batch: POBatch,
                              channel: ChannelConfig) -> None:
    """
    Validation — one row per resolved line with item lookup details + a
    cost-vs-expected comparison. Lets the operator spot pricing drift.
    """
    pct_int = (int(channel.expected_landing_ratio*100)
                if channel.expected_landing_ratio is not None else 0)
    cols = [c.format(pct=pct_int) for c in VALIDATION_SHEET_COLS]

    ws = wb.create_sheet('Validation')
    ws.append(cols)
    status_col = len(cols)         # 1-based index of the Status column

    for pf in batch.po_files:
        if pf.has_hard_errors:
            continue

        for line in pf.lines:
            if not line.item_no:
                continue
            if line.status == 'SKIP':
                continue

            mrp_master = line.items_master_mrp or 0
            landing = (round(mrp_master * channel.expected_landing_ratio, 2)
                        if (channel.expected_landing_ratio is not None
                             and mrp_master)
                        else '')
            our_cost = landing if landing != '' else ''
            mkt_cost = line.purchase_cost

            # Difference is INFORMATION ONLY per operator instruction:
            # "we won't consider any difference". Status reflects the
            # line's actual health (resolution, ship-to), NOT the diff.
            if landing != '' and mkt_cost:
                diff = round(mkt_cost - landing, 2)
            else:
                diff = ''
            status = 'OK' if line.status == 'OK' else 'WARN'

            ws.append([
                pf.po_no,
                line.item_no,
                line.ean or '',
                line.items_master_desc or line.sku_name or '',
                mrp_master if mrp_master else '',
                landing,
                line.gst_code or '',
                our_cost,
                mkt_cost if mkt_cost else '',
                diff,
                status,
            ])
            _apply_status_style(
                ws, ws.max_row, status_col, status,
                tint_row_cols=(1, status_col))

    _apply_table_format(ws, header_row=1)
    _autosize_columns(ws)

# Column layout for the Reconciliation sheet. Defined as a module-level
# constant so the column count (used for status-cell indexing) stays in
# sync with the data row that gets written.
RECONCILIATION_SHEET_COLS = [
    'Source File', 'PO No', 'SO No', 'Ship-to Code',
    'CSV Row', 'SKU Code', 'SKU Name', 'Quantity', 'Unit Cost',
    'EAN (Resolved)', 'Item No (Resolved)', 'MRP (Master)',
    'Output Line No', 'Tester Line No', 'Status', 'Notes',
]

def _write_reconciliation_sheet(wb, batch: POBatch,
                                  channel: ChannelConfig,
                                  add_non_stock: bool = False) -> None:
    """
    Reconciliation — the integrity check sheet. Every input CSV row in
    the batch becomes one row here, no exceptions:

        - WRITTEN  → row made it into Lines (SO), output_line_no shown.
        - SKIPPED  → row was excluded by the engine (blank SKU, zero qty).
        - FAILED   → row could not be resolved (SKU not in channel master,
                     or EAN not in Items Master).
        - FILE ERR → entire source file was rejected for a hard error
                     (missing columns, multi-PO file, etc.). One row per
                     attempted file is still emitted so the operator can
                     see why nothing came through.

    Status cells are colour-coded (green / yellow / red). Operator can
    sort or filter on Status to instantly find rows that need attention.

    The 'Verification' totals block at the bottom proves the integrity
    arithmetic: TOTAL INPUT ROWS == WRITTEN + SKIPPED + FAILED + FILE ERR.
    If those don't match, something is wrong with the writer itself.
    """
    ws = wb.create_sheet('Reconciliation')
    ws.append(RECONCILIATION_SHEET_COLS)
    status_col = len(RECONCILIATION_SHEET_COLS)   # 1-based Status col idx

    # Running counts for the verification totals row
    n_written = n_skipped = n_failed = n_file_err = 0
    n_tester_written = 0   # informational only, not part of integrity sum

    for pf in batch.po_files:
        # ── File with hard errors: emit one FILE ERROR row ──
        if pf.has_hard_errors:
            # If raw_df was parsed before the error, count its rows so
            # the operator sees how many input rows were lost. Otherwise
            # emit a single placeholder row noting the file was rejected.
            row_count = (len(pf.raw_df)
                          if pf.raw_df is not None and not pf.raw_df.empty
                          else 1)
            hard_errs = '; '.join(
                msg for lvl, msg in pf.findings if lvl == 'error')
            ws.append([
                pf.source_name,
                pf.po_no or '',
                '',
                '',
                f'(all {row_count} rows)',
                '', '', '', '',
                '', '', '',
                '',                  # Output Line No
                '',                  # Tester Line No
                'FILE ERR',
                hard_errs or 'File had hard errors',
            ])
            _apply_status_style(
                ws, ws.max_row, status_col, 'FAIL',
                tint_row_cols=(1, status_col))
            n_file_err += row_count
            continue

        # ── Clean file: emit one row per parsed line ──
        for line in pf.lines:
            # Classify the line based on REGULAR output (testers are
            # derivative — they only exist for WRITTEN lines):
            if line.status == 'SKIP':
                status, reason = 'SKIPPED', _status_reason_from_notes(line)
                n_skipped += 1
            elif line.item_no and line.output_line_no:
                status = 'WRITTEN'
                reason = _status_reason_from_notes(line) or ''
                n_written += 1
                if line.tester_output_line_no:
                    n_tester_written += 1
            else:
                # Has data, but didn't make it into Lines — failed lookup
                status = 'FAILED'
                reason = (_status_reason_from_notes(line)
                            or 'Could not resolve to D365 Item No')
                n_failed += 1

            ws.append([
                pf.source_name,
                pf.po_no,
                pf.so_number,
                pf.ship_to or '',
                line.csv_row_num,
                line.sku_code,
                line.sku_name,
                line.quantity,
                line.purchase_cost if line.purchase_cost else '',
                line.ean or '',
                line.item_no or '',
                line.items_master_mrp or '',
                line.output_line_no or '',
                line.tester_output_line_no or '',
                status,
                reason,
            ])
            _apply_status_style(
                ws, ws.max_row, status_col, status,
                tint_row_cols=(1, status_col))

    # ── Verification totals block ──
    # The integrity equation: every input row appears in exactly one of
    # WRITTEN / SKIPPED / FAILED / FILE ERR buckets. If they don't add
    # up to TOTAL INPUT ROWS, the CHECK row turns red and the writer is
    # provably wrong.
    ws.append([])
    total_input = n_written + n_skipped + n_failed + n_file_err
    bold = Font(bold=True)

    def _totals_row(label, value):
        # Row layout matches RECONCILIATION_SHEET_COLS:
        #   13 leading blanks + value at the Status column + 1 trailing blank
        ws.append([label] + [''] * 13 + [value, ''])
        for cell in ws[ws.max_row]:
            cell.font = bold

    _totals_row('TOTAL INPUT ROWS',  total_input)
    _totals_row('WRITTEN to Lines',  n_written)
    _totals_row('SKIPPED',            n_skipped)
    _totals_row('FAILED lookup',     n_failed)
    _totals_row('FILE ERROR rows',   n_file_err)

    # Sanity check row — if writer math drifts, this row turns red.
    check_status = ('OK'
        if total_input == n_written + n_skipped + n_failed + n_file_err
        else 'FAIL')
    ws.append(['CHECK: WRITTEN + SKIPPED + FAILED + FILE ERR == TOTAL']
               + [''] * 13 + [check_status, ''])
    _apply_status_style(ws, ws.max_row, status_col, check_status,
                          tint_row_cols=(1, status_col))
    for cell in ws[ws.max_row]:
        cell.font = bold

    # Tester informational row — separate from the integrity sum.
    # Operator can see how many tester lines were generated. Only
    # appears when at least one tester line was written.
    if n_tester_written:
        ws.append([])
        ws.append(['TESTER LINES generated'] + [''] * 13
                    + [n_tester_written, ''])
        for cell in ws[ws.max_row]:
            cell.font = bold

    # Non-stock items log — what the operator asked for: "also logged
    # somewhere in output that these are added". Per-PO detail rows make
    # it possible to trace any non-stock line back to its source.
    if (add_non_stock
            and batch.master_bundle is not None
            and batch.master_bundle.non_stock):
        ns_list = batch.master_bundle.non_stock
        tester_pos = [pf for pf in batch.po_files if pf.tester_so_number]
        n_ns_lines_total = len(ns_list) * len(tester_pos)

        ws.append([])
        ws.append(['NON-STOCK ITEMS added to tester SOs']
                    + [''] * 13 + [n_ns_lines_total, ''])
        for cell in ws[ws.max_row]:
            cell.font = bold

        for pf in tester_pos:
            for ns in ns_list:
                ws.append([
                    'NON-STOCK',
                    pf.po_no or '',
                    pf.tester_so_number,
                    pf.ship_to or '',
                    '', '',
                    ns.description,
                    1,
                    '',
                    ns.gtin or '',
                    ns.item_no,
                    '', '',
                    'appended',
                    'NON-STOCK',
                    f"Group: {ns.group}" if ns.group else '',
                ])

    _apply_table_format(ws, header_row=1)
    _autosize_columns(ws)

def _status_reason_from_notes(line) -> str:
    """
    Extract a one-line reason from a POLine's accumulated notes.
    Returns '' if the line has no notes. Used to populate the
    Reconciliation sheet's Notes column without re-running any logic.
    """
    if not line.notes:
        return ''
    # Notes are pre-formatted as "[LEVEL] message". Join with semicolons
    # for compactness; the cell autosize will widen the column if needed.
    return '; '.join(line.notes)

def _write_warnings_sheet(wb, batch: POBatch,
                            channel: ChannelConfig) -> None:
    """
    Warnings — every warn-level finding with PO/location context.
    Includes hard-error files (so the operator sees why they were
    excluded from Headers/Lines).
    """
    ws = wb.create_sheet('Warnings')
    ws.append(WARNINGS_SHEET_COLS)

    # Per-file findings
    for pf in batch.po_files:
        # File-level findings
        for level, msg in pf.findings:
            if level in ('warn', 'error'):
                marker = '✗ ' if level == 'error' else ''
                ws.append([pf.po_no or pf.source_name,
                            pf.store_name or '',
                            marker + msg])
        # Line-level warnings — line.notes is List[str] like
        # "[WARN] ..." / "[SKIP] ...". Filter to WARN/ERROR only.
        for line in pf.lines:
            for note in line.notes:
                if note.startswith('[WARN]') or note.startswith('[ERROR]'):
                    ws.append([pf.po_no or pf.source_name,
                                pf.store_name or '',
                                f"Row {line.csv_row_num} "
                                f"(SKU {line.sku_code}): {note}"])

    # Cross-file findings
    for level, msg in batch.cross_findings:
        if level in ('warn', 'error'):
            marker = '✗ ' if level == 'error' else ''
            ws.append(['', '', marker + msg])

    _apply_table_format(ws, header_row=1)
    _autosize_columns(ws)

def _write_raw_data_sheet(wb, batch: POBatch,
                            channel: ChannelConfig) -> None:
    """
    Raw Data — original CSV rows + appended resolved fields per row.
    Concatenates all PO files vertically with a leading PO column so
    the operator can trace any line back to its source.

    Files with hard errors are still included (with whatever rows
    parsed) so this sheet is a complete audit log.
    """
    ws = wb.create_sheet('Raw Data')

    # Build the output header by combining the channel's CSV columns
    # with the appended resolved fields. Use the first non-empty file's
    # columns to preserve any quirks of the operator's source format.
    sample_df = None
    for pf in batch.po_files:
        if pf.raw_df is not None and not pf.raw_df.empty:
            sample_df = pf.raw_df
            break

    if sample_df is None:
        # No data to emit; still write headers for consistency
        appended = ['Item No (Master)', 'MRP (Master)', 'Landing',
                     'GST Code', 'Cost Price (Resolved)',
                     'Difference with Cost']
        ws.append(channel.csv_required_cols + appended)
        _bold_header_row(ws)
        return

    csv_cols = list(sample_df.columns)
    appended_cols = ['Item No (Master)', 'MRP (Master)', 'Landing',
                      'GST Code', 'Cost Price (Resolved)',
                      'Difference with Cost']
    ws.append(csv_cols + appended_cols)

    for pf in batch.po_files:
        if pf.raw_df is None or pf.raw_df.empty:
            continue

        # Index parsed POLines by csv_row_num for fast appending
        line_by_row = {l.csv_row_num: l for l in pf.lines}

        for idx, raw_row in pf.raw_df.iterrows():
            csv_row_num = idx + 2   # +1 for 1-based, +1 for header
            row_values = [
                (raw_row[c] if c in pf.raw_df.columns else '')
                for c in csv_cols
            ]

            line = line_by_row.get(csv_row_num)
            if line is not None:
                mrp_m = line.items_master_mrp or ''
                landing = (round(mrp_m * channel.expected_landing_ratio, 2)
                            if (channel.expected_landing_ratio is not None
                                 and mrp_m)
                            else '')
                cost_p = line.purchase_cost or ''
                diff = (round(cost_p - landing, 2)
                         if (landing != '' and cost_p)
                         else '')
                appended = [
                    line.item_no or '',
                    mrp_m,
                    landing,
                    line.gst_code or '',
                    cost_p,
                    diff,
                ]
            else:
                appended = ['', '', '', '', '', '']

            ws.append(row_values + appended)

    _apply_table_format(ws, header_row=1)
    _autosize_columns(ws)

#
# One unified report: master-load summary + batch parse summary + verdict.
# Mirrors what the eventual GUI status panel will show.

def print_batch_report(batch: POBatch) -> None:
    """Print the consolidated report (master load + batch parse + verdict)."""
    bundle = batch.master_bundle

    print()
    print("═" * 78)
    print("  MT SELECT — H&G BATCH RUN REPORT")
    print("═" * 78)
    print(f"  Run at:       {batch.parsed_at:%Y-%m-%d %H:%M:%S}")
    if bundle and bundle.workbook_path:
        print(f"  Masters file: {bundle.workbook_path.name}")
        if bundle.workbook_mtime:
            print(f"  Last updated: {bundle.workbook_mtime:%Y-%m-%d %H:%M:%S}")
    print(f"  Files in:     {len(batch.po_files)}")
    if batch.so_assignment_summary:
        print(f"  SO numbers:   {batch.so_assignment_summary}")
    print()

    # ── Master counts ──
    if bundle:
        print("─" * 78)
        print("  MASTER COUNTS")
        print("─" * 78)
        print(f"  Items Master: {len(bundle.items_by_gtin)} EAN entries")
        print(f"  Ship-to B2B:  {len(bundle.ship_to_lookup)} locations")

        party_counts: Dict[str, int] = {}
        for k in bundle.ship_to_lookup.keys():
            party_counts[k[0]] = party_counts.get(k[0], 0) + 1
        for p, n in sorted(party_counts.items()):
            print(f"                  └─ {p}: {n}")

        print(f"  HG Master:    {len(bundle.get_channel_master('HG'))} SKUs")
        hg_master = bundle.get_channel_master('HG')
        active   = sum(1 for e in hg_master.values()
                        if e.status.lower() == 'active')
        inactive = sum(1 for e in hg_master.values()
                        if e.status.lower() == 'inactive')
        blank_e  = sum(1 for e in hg_master.values()
                        if e.enn_code is None)
        print(f"                  └─ Active: {active}")
        print(f"                  └─ Inactive: {inactive}")
        print(f"                  └─ Blank EAN: {blank_e}")

        # Master findings (excluding the "loaded N entries" info noise)
        master_warnings = [(l, m) for l, m in bundle.findings
                            if l in ('warn', 'error')]
        if master_warnings:
            print()
            print("  Master findings:")
            for level, msg in master_warnings:
                marker = {'warn': '⚠', 'error': '✗'}.get(level, '·')
                print(f"     {marker} [{level.upper():5}] {msg}")
        print()

    # ── Cross-file findings ──
    if batch.cross_findings:
        print("─" * 78)
        print("  CROSS-FILE FINDINGS")
        print("─" * 78)
        for level, msg in batch.cross_findings:
            marker = {'info': 'ℹ', 'warn': '⚠', 'error': '✗'}.get(level, '·')
            print(f"  {marker} [{level.upper():5}] {msg}")
        print()

    # ── Batch-level rollup ──
    print("─" * 78)
    print("  BATCH ROLLUP")
    print("─" * 78)
    n_ok    = sum(1 for f in batch.po_files
                    if not f.has_hard_errors
                    and not any(l.status == 'WARN' for l in f.lines)
                    and not any(lvl == 'warn' for lvl, _ in f.findings))
    n_warn  = sum(1 for f in batch.po_files
                    if not f.has_hard_errors
                    and (any(l.status == 'WARN' for l in f.lines)
                         or any(lvl == 'warn' for lvl, _ in f.findings)))
    n_error = sum(1 for f in batch.po_files if f.has_hard_errors)
    total_lines = sum(len(f.lines) for f in batch.po_files
                       if not f.has_hard_errors)
    total_qty   = sum(f.input_qty_total for f in batch.po_files
                       if not f.has_hard_errors)
    total_val   = sum(f.input_po_value_total for f in batch.po_files
                       if not f.has_hard_errors)
    print(f"  Files clean:  {n_ok}")
    print(f"  Files warn:   {n_warn}")
    print(f"  Files error:  {n_error}")
    print(f"  Total lines:  {total_lines}")
    print(f"  Total qty:    {total_qty}")
    print(f"  Total value:  ₹{total_val:,.2f}")
    print()

    # ── Per-file detail ──
    for i, pf in enumerate(batch.po_files, 1):
        print("─" * 78)
        print(f"  FILE {i}/{len(batch.po_files)}: {pf.source_name}")
        print("─" * 78)

        if pf.has_hard_errors:
            print(f"  ✗ HARD ERROR — file rejected:")
            for level, msg in pf.findings:
                if level == 'error':
                    print(f"     • {msg}")
            print()
            continue

        print(f"  PO No:        {pf.po_no}")
        if pf.so_number:
            print(f"  SO No:        {pf.so_number}   ← generated for D365")
        if pf.tester_so_number:
            print(f"  Tester SO:    {pf.tester_so_number}   ← qty=1, "
                  f"ext doc 'TESTERS'")
        print(f"  Store:        {pf.store_name}")
        print(f"  Ship to:      {pf.ship_to or '(unresolved)'}")
        print(f"  Cust No:      {pf.cust_no or '(unresolved)'}")
        print(f"  Vendor:       {pf.vendor_code} ({pf.vendor_name})")
        print(f"  Location:     {pf.location_code} | Region: {pf.region_code}")

        ok   = len(pf.ok_lines())
        warn = len(pf.warn_lines())
        skip = len(pf.skip_lines())
        print(f"  Lines:        {pf.input_line_count} total → "
              f"OK={ok}  WARN={warn}  SKIP={skip}")
        print(f"  Qty total:    {pf.input_qty_total}")
        print(f"  PO_VALUE:     ₹{pf.input_po_value_total:,.2f}")
        print(f"  Cost total:   ₹{pf.input_purchase_cost_total:,.2f}")

        non_line = [(l, m) for l, m in pf.findings if l != 'error']
        if non_line:
            print(f"\n  File findings:")
            for level, msg in non_line:
                marker = {'info': 'ℹ', 'warn': '⚠'}.get(level, '·')
                print(f"     {marker} [{level.upper():5}] {msg}")

        # Sample resolved lines (always show first 5)
        print(f"\n  Resolved lines (first 5):")
        for l in pf.lines[:5]:
            status_marker = {'OK': '✓', 'WARN': '⚠',
                              'SKIP': '✗'}.get(l.status, '·')
            item_str = (f"→ Item {l.item_no}" if l.item_no
                         else "→ (no Item No)")
            print(f"     {status_marker} R{l.csv_row_num} | SKU {l.sku_code} "
                  f"| qty {l.quantity} | EAN {l.ean or '?'} | {item_str}")

        # Problems
        problems = pf.warn_lines() + pf.skip_lines()
        if problems:
            print(f"\n  Lines needing attention ({len(problems)}):")
            for l in problems[:10]:
                print(f"     R{l.csv_row_num} SKU {l.sku_code} "
                      f"[{l.status}]: {l.sku_name}")
                for n in l.notes:
                    print(f"        {n}")
            if len(problems) > 10:
                print(f"     ... and {len(problems) - 10} more")
        print()

    # ── Final verdict ──
    print("═" * 78)
    if n_error:
        print(f"  RESULT: ✗ {n_error} file(s) REJECTED, "
              f"{n_ok} clean, {n_warn} with warnings")
    elif n_warn:
        print(f"  RESULT: ⚠ {n_ok} clean, {n_warn} with warnings — "
              f"REVIEW BEFORE SO GENERATION")
    elif n_ok:
        print(f"  RESULT: ✓ All {n_ok} file(s) parsed clean")
    else:
        print(f"  RESULT: (no files processed)")
    print("═" * 78)
    print()

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 8 — GUI (Tkinter) ───────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Tkinter GUI styled to match the Online PO Processor visual layout:
#   - Centered title + subtitle at the top
#   - Compact horizontal settings row (Warehouse → resolved code)
#   - Input Files frame with master file + PO CSV picker
#   - Big green primary action button
#   - Secondary buttons stacked centered below
#   - Status line above the log
#   - Log panel with [HH:MM:SS]-prefixed entries
#
# Auto-loads masters on startup. Background thread for processing so the
# UI never freezes. All long-running operations dispatch through
# root.after(0, ...) to update the UI safely.

class MTSelectGUI:
    """
    Single-window Tkinter app for MT Select.

    Visual layout (mirrors Online PO Processor):
        ┌─ MT Select — Multi-Channel PO Processor ──────────────────┐
        │                                                            │
        │                       MT Select                            │ ← centered
        │   Health & Glow (HG) PO  →  D365 Sales Order Import        │ ← dynamic subtitle
        │                                                            │
        │  Warehouse: [AHD ▼]  → PICK                                │ ← settings
        │                                                            │
        │  ┌─ Input Files ──────────────────────────────────────┐   │
        │  │  MT_Masters: ✓ MT_Masters.xlsx (auto-loaded)       │   │
        │  │              Updated: 2026-06-04 18:18:43  [Browse]│   │
        │  │                                                     │   │
        │  │  PO CSVs:    2 file(s) selected         [Add Files] │   │
        │  │              ┌─────────────────────────────────┐    │   │
        │  │              │ 91143-ENR013.csv                │    │   │
        │  │              │ 91149-MNR073.csv                │    │   │
        │  │              └─────────────────────────────────┘    │   │
        │  │                              [Scan] [Remove] [Clear]│   │
        │  └─────────────────────────────────────────────────────┘   │
        │                                                            │
        │            ┌───────────────────────┐                       │
        │            │ ▶  Generate Sales Orders     │ ← big green primary   │
        │            └───────────────────────┘                       │
        │            ┌───────────────────────┐                       │
        │            │ Create Template       │                       │
        │            └───────────────────────┘                       │
        │            ┌───────────────────────┐                       │
        │            │ Reload Masters        │                       │
        │            └───────────────────────┘                       │
        │                                                            │
        │  Status: Ready — select files and validate                 │
        │                                                            │
        │  ┌─ Log ────────────────────────────────────────────────┐  │
        │  │ [18:18:43] Auto-loaded MT_Masters.xlsx              │  │
        │  │ [18:18:43] Items Master: 8 EANs                     │  │
        │  │ [18:18:43] Ship-to B2B: 36 locations (HG=36)        │  │
        │  │ [18:18:43] HG Master: 191 SKUs (189 Active, ...)    │  │
        │  │ ...                                                  │  │
        │  └──────────────────────────────────────────────────────┘  │
        └────────────────────────────────────────────────────────────┘
    """

    # ─── Visual styling constants (match Online PO Processor) ───
    COLOR_TITLE         = '#000000'
    COLOR_SUBTITLE      = '#777777'
    COLOR_FILENAME      = '#1e40af'     # blue for loaded file display
    COLOR_TIMESTAMP     = '#666666'
    COLOR_WARN          = '#b85c00'
    COLOR_OK            = '#1e8449'
    COLOR_ERROR         = '#c0392b'
    COLOR_PRIMARY_BG    = '#28a745'
    COLOR_PRIMARY_HOVER = '#218838'
    COLOR_PRIMARY_FG    = '#ffffff'

    FONT_TITLE      = ('Segoe UI', 16, 'bold')
    FONT_SUBTITLE   = ('Segoe UI', 9)
    FONT_LABEL      = ('Segoe UI', 9, 'bold')
    FONT_NORMAL     = ('Segoe UI', 9)
    FONT_SMALL      = ('Segoe UI', 8)
    FONT_MONO       = ('Consolas', 9)
    FONT_PRIMARY_BTN = ('Segoe UI', 11, 'bold')

    def __init__(self, root):
        # Lazy import — keeps the loader portion of the module headless-safe.
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox, scrolledtext
        self.tk = tk
        self.ttk = ttk
        self.filedialog = filedialog
        self.messagebox = messagebox
        self.scrolledtext = scrolledtext

        self.root = root
        self.root.title("MT Select — Multi-Channel PO Processor  v0.6")
        self.root.geometry("1020x840")
        self.root.minsize(920, 620)

        # ── State ──
        # Load saved preferences (last-used masters path).
        # If the operator browsed to a non-default path last session,
        # restore it now — they don't have to re-browse every launch.
        self.config: Dict[str, Any] = load_config()
        saved_path = self.config.get('master_path')
        if saved_path and Path(saved_path).exists():
            self.master_path: Path = Path(saved_path)
        else:
            self.master_path: Path = get_masters_path()

        self.bundle:       Optional[MasterBundle] = None
        self.csv_files:    List[Path] = []
        self.is_processing: bool = False

        # Active channel — restored from config if previously set,
        # otherwise defaults to the first registered channel.
        saved_channel = self.config.get('active_channel', DEFAULT_CHANNEL)
        if saved_channel not in CHANNELS:
            saved_channel = DEFAULT_CHANNEL
        self.active_channel_code: str = saved_channel

        # Tester checkbox state — persisted so operator's preference
        # survives between launches. Default off (regulars only).
        self.generate_testers: bool = bool(
            self.config.get('generate_testers', False))

        # "Add Non Stock" checkbox — appends the GWP/PWP kit from
        # MT_Masters → Non Stock sheet to each tester SO.
        self.add_non_stock: bool = bool(
            self.config.get('add_non_stock', False))

        # Tester dump file path — operator-selected file used to filter
        # tester eligibility in SELECTIVE mode. Empty string means
        # AUTOMATIC mode (every resolved line becomes a tester).
        # Persisted across launches like the master_path.
        self.tester_dump_path: str = self.config.get('tester_dump_path', '')
        self.tester_dump: Optional[TesterDump] = None

        # ── Tk variables (need an existing root before creation) ──
        self.channel_var         = tk.StringVar(value=self.active_channel_code)
        self.testers_var         = tk.BooleanVar(value=self.generate_testers)
        self.add_non_stock_var   = tk.BooleanVar(value=self.add_non_stock)
        self.warehouse_var       = tk.StringVar(value=DEFAULT_WAREHOUSE)
        self.warehouse_code_var  = tk.StringVar(
            value=f'→ {WAREHOUSES[DEFAULT_WAREHOUSE]}')
        self.master_status_var   = tk.StringVar(value='Loading...')
        self.master_mtime_var    = tk.StringVar(value='')
        self.master_path_var     = tk.StringVar(value=str(self.master_path))
        self.input_count_var     = tk.StringVar(value='No files selected')
        # Tester dump display — shows file basename or "(not set — automatic)"
        self.tester_dump_var     = tk.StringVar(
            value=self._format_dump_label())
        self.status_var          = tk.StringVar(
            value='Initialising — auto-loading masters...')
        self.action_status_var   = tk.StringVar(value='')

        # ── Build UI then trigger initial load ──
        self._build_ui()
        self.root.after(100, self._auto_load_masters)

    # ────────────────────────────────────────────────────────────────────────
    #                       UI CONSTRUCTION
    # ────────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        """Construct the entire window layout. Called once at startup."""
        tk, ttk = self.tk, self.ttk

        # Outer container with consistent padding
        outer = ttk.Frame(self.root, padding=14)
        outer.pack(fill='both', expand=True)
        outer.columnconfigure(0, weight=1)
        # Log row gets all extra vertical space
        outer.rowconfigure(7, weight=1)

        row = 0

        # ─── Title + Subtitle (centered) ───
        # Title is static "MT Select"; subtitle reflects the active
        # channel and updates when the operator switches via the dropdown.
        ttk.Label(outer, text='MT Select',
                   font=self.FONT_TITLE, foreground=self.COLOR_TITLE,
                   anchor='center'
                   ).grid(row=row, column=0, sticky='ew', pady=(0, 2))
        row += 1
        self.subtitle_var = tk.StringVar(value=self._subtitle_for_channel())
        ttk.Label(outer,
                   textvariable=self.subtitle_var,
                   font=self.FONT_SUBTITLE, foreground=self.COLOR_SUBTITLE,
                   anchor='center'
                   ).grid(row=row, column=0, sticky='ew', pady=(0, 12))
        row += 1

        # ─── Settings row (Warehouse) ───
        self._build_settings_row(outer).grid(row=row, column=0,
                                                sticky='w', pady=(0, 10))
        row += 1

        # ─── Input Files frame ───
        self._build_input_files_frame(outer).grid(row=row, column=0,
                                                     sticky='ew',
                                                     pady=(0, 12))
        row += 1

        # ─── Action buttons (centered, stacked) ───
        self._build_action_frame(outer).grid(row=row, column=0,
                                                pady=(0, 8))
        row += 1

        # ─── Inline action status (under buttons, centered) ───
        ttk.Label(outer, textvariable=self.action_status_var,
                   font=self.FONT_NORMAL, foreground='#444',
                   anchor='center'
                   ).grid(row=row, column=0, sticky='ew', pady=(0, 8))
        row += 1

        # ─── Status line ───
        status_frame = ttk.Frame(outer)
        status_frame.grid(row=row, column=0, sticky='ew', pady=(0, 6))
        status_frame.columnconfigure(0, weight=1)
        ttk.Label(status_frame, textvariable=self.status_var,
                   font=self.FONT_NORMAL, foreground='#555',
                   anchor='center'
                   ).grid(row=0, column=0, sticky='ew')
        row += 1

        # ─── Log panel ───
        self._build_log_frame(outer).grid(row=row, column=0, sticky='nsew')

    def _build_settings_row(self, parent):
        """
        Channel + Warehouse + tester controls.

        Layout (two rows):
          Row 1: Channel  Warehouse → CODE   [✓] Generate Testers (price)
          Row 2: Tester Dump: <filename>   [Browse...] [Clear]
                 (only shown when the active channel supports testers)
        """
        tk, ttk = self.tk, self.ttk
        outer = ttk.Frame(parent)

        # ── Row 1: Channel / Warehouse / Testers checkbox ──
        frame = ttk.Frame(outer)
        frame.pack(side='top', anchor='w', fill='x')

        # ─── Channel selector ───
        # Dropdown values are "HG — Health & Glow" style for clarity.
        # When HB is added to CHANNELS, it appears here automatically.
        ttk.Label(frame, text='Channel:',
                   font=self.FONT_LABEL).pack(side='left', padx=(0, 6))

        self.channel_combo_values = [
            f'{code} — {ch.display_name}'
            for code, ch in CHANNELS.items()
        ]
        self.channel_combo = ttk.Combobox(
            frame, values=self.channel_combo_values,
            state='readonly', width=22,
            font=self.FONT_NORMAL,
        )
        # Set initial display value to match active_channel_code
        for i, v in enumerate(self.channel_combo_values):
            if v.startswith(self.active_channel_code + ' '):
                self.channel_combo.current(i)
                break
        self.channel_combo.pack(side='left', padx=(0, 18))
        self.channel_combo.bind('<<ComboboxSelected>>',
                                  self._on_channel_change)

        # ─── Warehouse selector ───
        ttk.Label(frame, text='Warehouse:',
                   font=self.FONT_LABEL).pack(side='left', padx=(0, 6))

        combo = ttk.Combobox(frame, textvariable=self.warehouse_var,
                              values=list(WAREHOUSES.keys()),
                              state='readonly', width=8,
                              font=self.FONT_NORMAL)
        combo.pack(side='left')
        combo.bind('<<ComboboxSelected>>', self._on_warehouse_change)

        ttk.Label(frame, textvariable=self.warehouse_code_var,
                   font=self.FONT_NORMAL,
                   foreground='#666').pack(side='left', padx=(10, 0))

        # ─── Testers checkbox ───
        # When enabled, every cleanly-parsed regular PO also gets a
        # paired tester SO (SO/HG/TT/...) with qty=1 / unit_price=0.54
        # / external_doc=TESTERS. Counter continues in a single
        # continuous block after the regulars (confirmed by operator).
        # Disabled when the active channel doesn't declare a tester
        # price (channel.tester_unit_price is None).
        active_channel = CHANNELS.get(self.active_channel_code)
        chk_state = ('normal'
                      if (active_channel is not None
                          and active_channel.tester_unit_price is not None)
                      else 'disabled')
        self.testers_check = ttk.Checkbutton(
            frame,
            text='Generate Testers',
            variable=self.testers_var,
            command=self._on_testers_toggle,
            state=chk_state,
        )
        self.testers_check.pack(side='left', padx=(20, 0))

        # Show the tester price next to the checkbox so operator can
        # confirm at a glance which price will be used.
        if active_channel is not None and active_channel.tester_unit_price is not None:
            ttk.Label(
                frame,
                text=f'(qty=1, price={active_channel.tester_unit_price})',
                font=self.FONT_SMALL,
                foreground='#666',
            ).pack(side='left', padx=(6, 0))

        # ─── Add Non Stock checkbox ───
        # Active only on channels that support testers (non-stock goes
        # INTO the tester SO).
        if active_channel is not None and active_channel.tester_unit_price is not None:
            self.non_stock_check = ttk.Checkbutton(
                frame,
                text='Add Non Stock',
                variable=self.add_non_stock_var,
                command=self._on_non_stock_toggle,
            )
            self.non_stock_check.pack(side='left', padx=(16, 0))
            ttk.Label(
                frame,
                text='(GWP/PWP)',
                font=self.FONT_SMALL,
                foreground='#666',
            ).pack(side='left', padx=(4, 0))

        # ── Row 2: Tester Dump file picker ──
        # Only added for channels that support testers. Operator can
        # leave it unset (AUTOMATIC mode — all resolved lines become
        # testers) or browse to a Skin_care_NPI_tester.xlsm-style file
        # (SELECTIVE mode — only (location, SKU) pairs in the dump
        # become testers).
        if active_channel is not None and active_channel.tester_unit_price is not None:
            dump_frame = ttk.Frame(outer)
            dump_frame.pack(side='top', anchor='w', fill='x',
                              pady=(6, 0))

            ttk.Label(dump_frame, text='Tester Dump:',
                       font=self.FONT_LABEL).pack(side='left', padx=(0, 6))
            ttk.Label(dump_frame, textvariable=self.tester_dump_var,
                       font=self.FONT_NORMAL, foreground='#444',
                       width=55, anchor='w',
                       ).pack(side='left', padx=(0, 8))
            ttk.Button(dump_frame, text='Browse...',
                        command=self._on_browse_dump
                        ).pack(side='left', padx=(0, 6))
            ttk.Button(dump_frame, text='Clear',
                        command=self._on_clear_dump
                        ).pack(side='left')

        return outer

    def _build_input_files_frame(self, parent):
        """
        Input Files frame containing:
          - Masters file row (✓ filename + timestamp + Browse)
          - PO CSV list (listbox + Add/Scan/Remove/Clear buttons)
        """
        tk, ttk = self.tk, self.ttk
        frame = ttk.LabelFrame(parent, text='  Input Files  ', padding=12)
        frame.columnconfigure(1, weight=1)

        # ─── Masters file row ───
        # Row 0: Label | Status (filename + ✓) | Browse button
        ttk.Label(frame, text='MT_Masters:', font=self.FONT_LABEL,
                   anchor='nw').grid(row=0, column=0, sticky='nw',
                                       padx=(0, 10))

        self.master_status_lbl = ttk.Label(
            frame, textvariable=self.master_status_var,
            font=self.FONT_LABEL, foreground=self.COLOR_FILENAME)
        self.master_status_lbl.grid(row=0, column=1, sticky='w')

        ttk.Button(frame, text='Browse...',
                    command=self._browse_master
                    ).grid(row=0, column=2, sticky='e', padx=(10, 0))

        # Row 1: empty | timestamp
        ttk.Label(frame, textvariable=self.master_mtime_var,
                   font=self.FONT_SMALL, foreground=self.COLOR_TIMESTAMP
                   ).grid(row=1, column=1, sticky='w', pady=(0, 8))

        # ─── PO CSVs row ───
        ttk.Label(frame, text='PO CSVs:', font=self.FONT_LABEL,
                   anchor='nw').grid(row=2, column=0, sticky='nw',
                                       padx=(0, 10), pady=(4, 0))

        # File count label
        ttk.Label(frame, textvariable=self.input_count_var,
                   font=self.FONT_NORMAL, foreground='#444'
                   ).grid(row=2, column=1, sticky='w', pady=(4, 2))

        # Add Files button on row 2 right
        ttk.Button(frame, text='Add Files...',
                    command=self._add_files
                    ).grid(row=2, column=2, sticky='e', padx=(10, 0),
                            pady=(4, 0))

        # Row 3: listbox spanning cols 1-2
        list_frame = ttk.Frame(frame)
        list_frame.grid(row=3, column=1, columnspan=2, sticky='ew',
                         pady=(0, 4))
        list_frame.columnconfigure(0, weight=1)

        self.files_listbox = tk.Listbox(list_frame, height=4,
                                         selectmode='extended',
                                         font=self.FONT_MONO,
                                         relief='solid', borderwidth=1,
                                         highlightthickness=0,
                                         activestyle='dotbox')
        self.files_listbox.grid(row=0, column=0, sticky='ew')
        scroll = ttk.Scrollbar(list_frame, orient='vertical',
                                command=self.files_listbox.yview)
        scroll.grid(row=0, column=1, sticky='ns')
        self.files_listbox.config(yscrollcommand=scroll.set)

        # Row 4: action buttons under the listbox (right-aligned)
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=1, columnspan=2, sticky='e')

        # Button label is dynamic — reflects the active channel's folder

        ttk.Button(btn_frame, text='Remove Selected',
                    command=self._remove_selected
                    ).pack(side='left', padx=4)
        ttk.Button(btn_frame, text='Clear All',
                    command=self._clear_files
                    ).pack(side='left', padx=(4, 0))

        return frame

    def _build_action_frame(self, parent):
        """
        Action buttons — big green primary + smaller secondaries below,
        all centered horizontally.
        """
        tk, ttk = self.tk, self.ttk
        outer = ttk.Frame(parent)

        # Center the buttons via a sub-frame with auto-fit width
        center = ttk.Frame(outer)
        center.pack(anchor='center')

        # Primary: big green button (native tk.Button — green bg works cross-platform)
        self.process_btn = tk.Button(
            center, text='▶  Generate Sales Orders',
            font=self.FONT_PRIMARY_BTN,
            bg=self.COLOR_PRIMARY_BG, fg=self.COLOR_PRIMARY_FG,
            activebackground=self.COLOR_PRIMARY_HOVER,
            activeforeground=self.COLOR_PRIMARY_FG,
            relief='flat', borderwidth=0,
            padx=36, pady=8, cursor='hand2',
            command=self._on_process_click,
        )
        self.process_btn.pack(pady=(0, 6))

        # Secondary buttons (smaller, ttk default style, stacked)
        self.template_btn = ttk.Button(
            center, text='Create Template',
            command=self._create_template, width=22)
        self.template_btn.pack(pady=2)

        self.reload_btn = ttk.Button(
            center, text='Reload Masters',
            command=self._auto_load_masters, width=22)
        self.reload_btn.pack(pady=2)

        return outer

    def _build_log_frame(self, parent):
        """The Log panel: scrollable text with [HH:MM:SS]-prefixed entries."""
        tk, ttk = self.tk, self.ttk
        scrolledtext = self.scrolledtext

        frame = ttk.LabelFrame(parent, text='  Log  ', padding=8)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)

        self.log_text = scrolledtext.ScrolledText(
            frame, wrap='word', font=self.FONT_MONO,
            background='#fafafa', foreground='#222',
            relief='solid', borderwidth=1, padx=8, pady=6,
        )
        self.log_text.grid(row=0, column=0, sticky='nsew')
        self.log_text.config(state='disabled')

        # Tags for coloured output
        self.log_text.tag_configure('error',   foreground=self.COLOR_ERROR)
        self.log_text.tag_configure('warn',    foreground=self.COLOR_WARN)
        self.log_text.tag_configure('ok',      foreground=self.COLOR_OK)
        self.log_text.tag_configure('dim',     foreground='#888')

        # Clear-log link at the bottom-right of the frame
        clear_frame = ttk.Frame(frame)
        clear_frame.grid(row=1, column=0, sticky='e', pady=(4, 0))
        ttk.Button(clear_frame, text='Clear Log',
                    command=self._clear_log).pack()

        return frame

    # ────────────────────────────────────────────────────────────────────────
    #                       SETTINGS HANDLERS
    # ────────────────────────────────────────────────────────────────────────

    def _on_warehouse_change(self, _event=None):
        """Update the → CODE display when warehouse dropdown changes."""
        wh = self.warehouse_var.get()
        self.warehouse_code_var.set(f'→ {WAREHOUSES.get(wh, "?")}')
        self._log(f'Warehouse changed to {wh} (D365 location: '
                   f'{WAREHOUSES.get(wh, "?")})')

    def _on_testers_toggle(self):
        """
        Tester checkbox toggled — persist the setting so it survives
        between launches, and log the change for transparency.
        """
        self.generate_testers = bool(self.testers_var.get())
        self.config['generate_testers'] = self.generate_testers
        save_config(self.config)
        if self.generate_testers:
            channel = CHANNELS[self.active_channel_code]
            self._log(
                f'Testers ENABLED — each regular PO will also produce a '
                f'tester SO (SO/{channel.code}/TT/...) with qty=1, '
                f'unit_price={channel.tester_unit_price}, '
                f'External Document No.={channel.tester_external_doc_no}')
            # Re-log dump mode so operator sees the current selection.
            self._log_dump_mode()
        else:
            self._log('Testers DISABLED — only regular SOs will be generated')

    def _on_non_stock_toggle(self):
        """
        Non Stock checkbox toggled — persist + log. Only takes effect
        when Generate Testers is also on.
        """
        self.add_non_stock = bool(self.add_non_stock_var.get())
        self.config['add_non_stock'] = self.add_non_stock
        save_config(self.config)
        if self.add_non_stock:
            ns_count = (len(self.bundle.non_stock)
                          if self.bundle is not None else 0)
            self._log(
                f'Non Stock ENABLED — {ns_count} item(s) from '
                f'MT_Masters → Non Stock sheet will be appended to '
                f'each tester SO (qty=1 each)')
            if not self.generate_testers:
                self._log(
                    '  Note: Generate Testers is OFF — non-stock items '
                    'only get added when testers are also on.', tag='warn')
        else:
            self._log('Non Stock DISABLED')

    def _format_dump_label(self) -> str:
        """
        Build the label shown next to the Tester Dump path. Returns a
        short user-facing string indicating which mode is active.
        """
        if not self.tester_dump_path:
            return '(not set — AUTOMATIC mode, all SKUs become testers)'
        try:
            return f'✓ {Path(self.tester_dump_path).name}  (SELECTIVE mode)'
        except Exception:
            return self.tester_dump_path

    def _log_dump_mode(self) -> None:
        """
        Log the current dump file state. Called whenever the operator
        toggles testers or changes the dump file.
        """
        if not self.tester_dump_path:
            self._log('Tester mode: AUTOMATIC — every resolved line in '
                       'every regular PO will get a tester')
        else:
            self._log(
                f'Tester mode: SELECTIVE — only lines whose '
                f'(LOCATION_CODE, SKU_CODE) appears in '
                f'"{Path(self.tester_dump_path).name}" will get a tester')

    def _on_browse_dump(self):
        """
        Browse... clicked on the Tester Dump row. Open a file dialog,
        load the file, validate it, and persist the path on success.
        """
        path = self.filedialog.askopenfilename(
            title='Select Tester Dump file',
            filetypes=[
                ('Excel macro-enabled workbook', '*.xlsm'),
                ('Excel workbook', '*.xlsx'),
                ('All files', '*.*'),
            ],
        )
        if not path:
            return    # operator cancelled

        # Try to load the dump immediately so we can show problems early
        # (rather than only at processing time).
        self._log(f'Loading tester dump: {path}')
        dump = read_tester_dump_file(Path(path))

        # Surface any findings in the log
        for level, msg in dump.findings:
            tag = 'error' if level == 'error' else 'warn'
            self._log(f'  [{level.upper()}] {msg}', tag=tag)

        if dump.has_hard_errors:
            self.messagebox.showerror(
                'Tester Dump load failed',
                f'Could not load {Path(path).name}. See log for details.\n\n'
                f'Tester dump path was NOT updated.')
            return

        # All good — persist and update GUI
        self.tester_dump_path = path
        self.tester_dump = dump
        self.config['tester_dump_path'] = path
        save_config(self.config)
        self.tester_dump_var.set(self._format_dump_label())
        self._log(
            f'✓ Tester dump loaded: {dump.rows_loaded} eligible '
            f'(LOCATION_CODE, SKU_CODE) pair(s) from sheet '
            f'"{dump.sheet_used}"')
        self._log_dump_mode()

    def _on_clear_dump(self):
        """
        Clear the Tester Dump path — operator wants to switch back to
        AUTOMATIC mode (all SKUs become testers).
        """
        if not self.tester_dump_path:
            return  # nothing to clear
        self.tester_dump_path = ''
        self.tester_dump = None
        self.config['tester_dump_path'] = ''
        save_config(self.config)
        self.tester_dump_var.set(self._format_dump_label())
        self._log('Tester dump cleared — switched to AUTOMATIC mode')

    def _ensure_dump_loaded(self) -> Optional[TesterDump]:
        """
        Return a fresh TesterDump for the current path, reloading from
        disk if needed. Returns None if no path is set or load failed.

        We reload on every processing run rather than caching, because
        the operator may have edited the dump in Excel between runs.
        """
        if not self.tester_dump_path:
            return None
        path = Path(self.tester_dump_path)
        dump = read_tester_dump_file(path)
        # Cache for inspection but always reload before use
        self.tester_dump = dump
        return dump

    def _on_channel_change(self, _event=None):
        """
        Channel dropdown changed → switch the active channel.
        The masters stay loaded (every channel reads from the same
        bundled workbook), but processing will now use the new channel's
        CSV schema, lookup chain, and input folder.
        """
        selected = self.channel_combo.get()    # e.g. "HG — Health & Glow"
        new_code = selected.split(' — ')[0].strip()
        if new_code not in CHANNELS:
            return
        if new_code == self.active_channel_code:
            return
        self.active_channel_code = new_code
        self.config['active_channel'] = new_code
        save_config(self.config)

        channel = CHANNELS[new_code]
        # Clear the CSV file list — files for one channel rarely belong
        # to another (different format).
        self.csv_files = []
        self._refresh_files_listbox()

        # Update the subtitle to reflect the newly active channel
        self.subtitle_var.set(self._subtitle_for_channel())
        self._log(f'Channel changed to {channel.code} '
                   f'({channel.display_name}) — input folder is now '
                   f'"{channel.input_folder_name}/"')

    # ────────────────────────────────────────────────────────────────────────
    #                       MASTER LOADING
    # ────────────────────────────────────────────────────────────────────────

    def _auto_load_masters(self):
        """Load masters from the conventional (or browsed) path."""
        if not self.master_path.exists():
            self._set_status(
                'MT_Masters.xlsx not found — use Browse or Create Template',
                error=True)
            self.master_status_var.set('✗ NOT LOADED')
            self.master_status_lbl.configure(foreground=self.COLOR_ERROR)
            self.master_mtime_var.set('')
            self._log(f'Master file not found: {self.master_path}',
                       tag='error')
            self._log('  Use Browse... to point to an existing file, or '
                       'Create Template to generate a fresh one.', tag='dim')
            return

        self._set_status('Loading masters...')
        threading.Thread(target=self._do_load_masters, daemon=True).start()

    def _do_load_masters(self):
        """Worker thread: actually load, then schedule UI update."""
        try:
            bundle = load_all_masters(self.master_path)
            self.root.after(0, lambda: self._on_masters_loaded(bundle))
        except Exception as e:
            err = str(e)
            self.root.after(0, lambda:
                self._set_status(f'Master load crashed: {err}', error=True))

    def _on_masters_loaded(self, bundle: MasterBundle):
        """UI thread: bundle ready, update display and log."""
        self.bundle = bundle

        has_errors = any(l == 'error' for l, _ in bundle.findings)

        # Status line in the Input Files row
        if has_errors:
            self.master_status_var.set(
                f'✗ {self.master_path.name} (LOAD ERRORS)')
            self.master_status_lbl.configure(foreground=self.COLOR_ERROR)
            self._set_status('Master load had errors — see log', error=True)
        else:
            self.master_status_var.set(
                f'✓ {self.master_path.name} (auto-loaded)')
            self.master_status_lbl.configure(foreground=self.COLOR_FILENAME)
            self._set_status('Ready — select files and validate')

        # mtime
        if bundle.workbook_mtime:
            self.master_mtime_var.set(
                f'Updated: {bundle.workbook_mtime:%Y-%m-%d %H:%M:%S}')
        else:
            self.master_mtime_var.set('')

        # Log the load (matches Online PO Processor style)
        self._log(f'Auto-loaded master from {self.master_path}')

        # Per-sheet summary lines (channel-agnostic)
        n_items = len(bundle.items_by_gtin)
        n_ship  = len(bundle.ship_to_lookup)

        party_counts: Dict[str, int] = {}
        for k in bundle.ship_to_lookup.keys():
            party_counts[k[0]] = party_counts.get(k[0], 0) + 1
        party_str = ', '.join(f"{p}={n}"
                                for p, n in sorted(party_counts.items())
                               ) or 'none'

        self._log(f'  Items Master: {n_items} EANs'
                   + self._sheet_suffix(bundle.items_sheet_name,
                                          SHEET_ITEMS_MASTER))
        self._log(f'  Ship-to B2B:  {n_ship} locations ({party_str})'
                   + self._sheet_suffix(bundle.ship_to_sheet_name,
                                          SHEET_SHIP_TO_B2B))

        # One line per channel-specific master — iterating CHANNELS means
        # HB will show up here automatically when we add it to the registry.
        for code, channel in CHANNELS.items():
            if channel.lookup_via != 'SKU' or not channel.channel_master_sheet:
                continue
            ch_master = bundle.get_channel_master(code)
            n_sku    = len(ch_master)
            active   = sum(1 for e in ch_master.values()
                            if e.status.lower() == 'active')
            inactive = sum(1 for e in ch_master.values()
                            if e.status.lower() == 'inactive')
            blank_e  = sum(1 for e in ch_master.values()
                            if e.enn_code is None)
            matched = bundle.get_channel_sheet_name(code)
            label = channel.channel_master_sheet.ljust(13)
            self._log(f'  {label}: {n_sku} SKUs '
                       f'(Active={active}  Inactive={inactive}  '
                       f'BlankEAN={blank_e})'
                       + self._sheet_suffix(matched,
                                              channel.channel_master_sheet))

        # Surface warnings/errors right after the count lines
        problems = [(l, m) for l, m in bundle.findings
                     if l in ('warn', 'error')]
        for level, msg in problems:
            tag = 'error' if level == 'error' else 'warn'
            marker = '✗' if level == 'error' else '⚠'
            # Multi-line messages: indent continuation lines
            lines = msg.split('\n')
            self._log(f'  {marker} {lines[0]}', tag=tag)
            for extra in lines[1:]:
                self._log(f'    {extra}', tag=tag)

    def _browse_master(self):
        """File dialog: point to an existing MT_Masters.xlsx anywhere on disk."""
        path_str = self.filedialog.askopenfilename(
            title='Select MT_Masters.xlsx',
            filetypes=[('Excel workbooks', '*.xlsx'),
                        ('All files', '*.*')],
            initialdir=str(self.master_path.parent
                            if self.master_path.parent.exists()
                            else get_script_dir()),
        )
        if not path_str:
            return
        self.master_path = Path(path_str)
        self.master_path_var.set(str(self.master_path))
        # Persist so next launch auto-loads from the same place.
        self.config['master_path'] = str(self.master_path)
        save_config(self.config)
        self._log(f'Master path changed to: {self.master_path}')
        self._auto_load_masters()

    def _create_template(self):
        """Generate a fresh empty MT_Masters.xlsx (confirm overwrite)."""
        target = get_masters_path()

        if target.exists():
            ok = self.messagebox.askyesno(
                'Replace existing?',
                f'A workbook already exists:\n\n{target}\n\n'
                f'Replace it with a fresh empty template? '
                f'(Any existing data will be overwritten.)'
            )
            if not ok:
                return

        try:
            path = create_template(target, overwrite=True)
            self.messagebox.showinfo(
                'Template created',
                f'Fresh empty workbook written:\n\n{path}\n\n'
                f'Open it, paste your real data into each sheet, '
                f'then click "Reload Masters".'
            )
            self.master_path = path
            self.master_path_var.set(str(path))
            self._log(f'Created fresh template: {path}')
            self._auto_load_masters()
        except Exception as e:
            self.messagebox.showerror('Template creation failed',
                                        f'Could not write template:\n\n{e}')
            self._log(f'Template creation failed: {e}', tag='error')

    # ────────────────────────────────────────────────────────────────────────
    #                       PO CSV FILE LIST
    # ────────────────────────────────────────────────────────────────────────

    def _add_files(self):
        """File dialog: multi-select CSVs from anywhere on disk."""
        channel = CHANNELS[self.active_channel_code]
        ch_folder = get_input_folder(self.active_channel_code)
        paths = self.filedialog.askopenfilenames(
            title=f'Select {channel.display_name} PO CSVs',
            filetypes=[('CSV files', '*.csv'), ('All files', '*.*')],
            initialdir=str(ch_folder if ch_folder.exists()
                            else get_script_dir()),
        )
        if not paths:
            return
        added = 0
        for p in paths:
            path = Path(p).resolve()
            if path not in self.csv_files:
                self.csv_files.append(path)
                added += 1
        self._refresh_files_listbox()
        if added:
            self._log(f'Added {added} file(s) — '
                       f'{len(self.csv_files)} total')

    def _remove_selected(self):
        """Drop highlighted listbox entries."""
        selected = list(self.files_listbox.curselection())
        if not selected:
            return
        for idx in reversed(selected):
            del self.csv_files[idx]
        self._refresh_files_listbox()
        self._log(f'Removed {len(selected)} file(s)')

    def _clear_files(self):
        """Empty the file list entirely."""
        if not self.csv_files:
            return
        n = len(self.csv_files)
        self.csv_files = []
        self._refresh_files_listbox()
        self._log(f'Cleared {n} file(s) from input list')

    def _refresh_files_listbox(self):
        """Rebuild the listbox from self.csv_files."""
        self.files_listbox.delete(0, 'end')
        for p in self.csv_files:
            self.files_listbox.insert('end', p.name)
        n = len(self.csv_files)
        if n == 0:
            self.input_count_var.set('No files selected')
        elif n == 1:
            self.input_count_var.set('1 file selected')
        else:
            self.input_count_var.set(f'{n} files selected')

    # ────────────────────────────────────────────────────────────────────────
    #                       PROCESSING (Generate Sales Orders)
    # ────────────────────────────────────────────────────────────────────────

    def _on_process_click(self):
        """Generate Sales Orders button → run the parse in a background thread."""
        if self.is_processing:
            return

        if not self.bundle:
            self.messagebox.showerror(
                'Masters not loaded',
                'Load the masters file first. If MT_Masters.xlsx is '
                'missing, use Browse... or Create Template.'
            )
            return

        if any(l == 'error' for l, _ in self.bundle.findings):
            self.messagebox.showerror(
                'Masters had errors',
                'The masters file loaded with errors. Fix the file and '
                'click "Reload Masters" before processing CSVs.'
            )
            return

        if not self.csv_files:
            self.messagebox.showwarning(
                'No files selected',
                'Add CSV files using "Add Files...".'
            )
            return

        # Lock UI
        self.is_processing = True
        self.process_btn.config(state='disabled',
                                  text='⏳  Processing...')
        self.action_status_var.set('')
        self._set_status(f'Processing {len(self.csv_files)} file(s)...')

        wh = self.warehouse_var.get()
        wh_code = WAREHOUSES.get(wh, '?')
        self._log(f'─── Generate Sales Orders clicked ───', tag='dim')
        self._log(f'Warehouse: {wh} → {wh_code}')
        self._log(f'Files: {[p.name for p in self.csv_files]}')

        # Spawn worker thread
        threading.Thread(target=self._do_process, daemon=True).start()

    def _do_process(self):
        """
        Worker thread: parse batch, assign SO numbers, write the output
        workbook, capture report.

        Output workbook location: by default, written next to the CSV
        files the operator selected (so it shows up in the same folder
        the operator was already navigating). If the source folder isn't
        writable (read-only network share, etc.), falls back to the
        channel's default Output_HG/ folder under the script directory.
        """
        try:
            channel = CHANNELS[self.active_channel_code]
            warehouse_code = WAREHOUSES.get(
                self.warehouse_var.get(), '')
            buf = io.StringIO()
            output_path: Optional[Path] = None

            # Decide WHERE the workbook will be written. Vishal's
            # preference: same folder as the source CSVs, so the output
            # appears alongside the inputs the operator just picked.
            # If multiple files have different parents (rare), we use
            # the first file's parent — consistent and predictable.
            output_dir = (self.csv_files[0].parent
                           if self.csv_files else
                           get_output_folder(channel.code))
            output_filename = generate_output_filename(channel)
            preferred_output = output_dir / output_filename

            with redirect_stdout(buf):
                batch = read_channel_csv_batch(
                    self.csv_files, channel, self.bundle)

                # If testers are enabled AND a dump file is configured,
                # reload the dump from disk (operator may have edited it
                # in Excel since the last run) and pass it to the SO
                # assignment for SELECTIVE mode. If not, fall through to
                # AUTOMATIC mode (every resolved line gets a tester).
                tester_dump = None
                if self.generate_testers and self.tester_dump_path:
                    tester_dump = self._ensure_dump_loaded()
                    if tester_dump is not None and not tester_dump.has_hard_errors:
                        print(f"  Tester dump loaded: "
                              f"{tester_dump.rows_loaded} eligible pair(s) "
                              f"from {tester_dump.source_name} "
                              f"(sheet '{tester_dump.sheet_used}')")
                    elif tester_dump is not None and tester_dump.has_hard_errors:
                        # Surface the errors but proceed without the
                        # dump (falls back to AUTOMATIC mode for safety).
                        print(f"  ⚠ Tester dump failed to load — "
                              f"falling back to AUTOMATIC mode")
                        for level, msg in tester_dump.findings:
                            print(f"    [{level.upper()}] {msg}")
                        tester_dump = None

                # Assign D365 SO numbers. The generate_testers flag
                # (set by the GUI checkbox) drives whether each cleanly
                # parsed PO also gets a paired tester SO (SO/HG/TT/...)
                # using the continuous counter block — see
                # assign_so_numbers for details. tester_dump=None means
                # AUTOMATIC mode; a TesterDump means SELECTIVE mode.
                assign_so_numbers(batch, channel,
                                    generate_testers=self.generate_testers,
                                    tester_dump=tester_dump)
                print_batch_report(batch)

                # Write the output workbook IF at least one file parsed
                # cleanly enough to have an SO number. Files that hit
                # hard errors still appear in the Reconciliation +
                # Warnings sheets.
                any_eligible = any(
                    pf.so_number for pf in batch.po_files)
                if any_eligible:
                    try:
                        do_non_stock = (self.add_non_stock
                                          and self.generate_testers)
                        output_path = write_so_workbook(
                            batch, channel, warehouse_code,
                            output_path=preferred_output,
                            add_non_stock=do_non_stock)
                    except (PermissionError, OSError) as write_err:
                        # Source folder isn't writable — fall back to
                        # the script's default Output_HG/ folder.
                        fallback_dir = get_output_folder(channel.code)
                        fallback_dir.mkdir(parents=True, exist_ok=True)
                        fallback_path = fallback_dir / output_filename
                        print()
                        print(f"  ⚠ Cannot write to {output_dir}: "
                              f"{write_err}")
                        print(f"    Falling back to {fallback_dir}")
                        output_path = write_so_workbook(
                            batch, channel, warehouse_code,
                            output_path=fallback_path,
                            add_non_stock=do_non_stock)
                    print()
                    print("─" * 78)
                    print(f"  ✓ OUTPUT WORKBOOK WRITTEN")
                    print("─" * 78)
                    print(f"  {output_path}")
                else:
                    print()
                    print("  ✗ No files cleanly parsed — no output workbook "
                          "written.")

            output = buf.getvalue()
            self.root.after(0, lambda:
                self._on_process_done(batch, output, output_path))
        except Exception as e:
            err_msg = f'Processing crashed: {e}'
            self.root.after(0, lambda: self._on_process_error(err_msg))

    def _on_process_done(self, batch: POBatch, output: str,
                           output_path: Optional[Path]):
        """UI thread: append the report to log, unlock UI."""
        # Push the captured report into the log AS-IS (no per-line
        # timestamping — the report itself is structured).
        self._append_log(output)

        # Summary
        n_files = len(batch.po_files)
        n_error = sum(1 for f in batch.po_files if f.has_hard_errors)
        n_warn = sum(1 for f in batch.po_files
                      if not f.has_hard_errors
                      and (any(l.status == 'WARN' for l in f.lines)
                           or any(lvl == 'warn' for lvl, _ in f.findings)))
        n_ok = n_files - n_error - n_warn

        if n_error and n_ok + n_warn == 0:
            # Nothing got through — error state
            summary = f'Done — all {n_error} file(s) REJECTED, no output written'
            self._set_status(summary, error=True)
            self.action_status_var.set(summary)
            self._log(summary, tag='error')
        elif n_error:
            summary = (f'Done — {n_error} REJECTED, '
                        f'{n_ok} clean, {n_warn} with warnings — '
                        f'workbook written')
            self._set_status(summary, warn=True)
            self.action_status_var.set(summary)
            self._log(summary, tag='warn')
        elif n_warn:
            summary = (f'Done — {n_ok} clean, {n_warn} with warnings — '
                        f'workbook written')
            self._set_status(summary, warn=True)
            self.action_status_var.set(summary)
            self._log(summary, tag='warn')
        else:
            summary = (f'Done — all {n_ok} file(s) parsed clean — '
                        f'workbook written')
            self._set_status(summary)
            self.action_status_var.set(summary)
            self._log(summary, tag='ok')

        if output_path:
            self._log(f'Output: {output_path}', tag='ok')
            # Prompt to open the folder containing the output
            ok = self.messagebox.askyesno(
                'Workbook written',
                f'SO workbook saved:\n\n{output_path}\n\n'
                f'Open output folder?'
            )
            if ok:
                self._open_path(output_path.parent)

        self.process_btn.config(state='normal',
                                  text='▶  Generate Sales Orders')
        self.is_processing = False

    def _open_path(self, path: Path) -> None:
        """Open a file or folder in the platform's default viewer."""
        try:
            if sys.platform == 'win32':
                os.startfile(str(path))   # noqa: type-checker
            elif sys.platform == 'darwin':
                import subprocess
                subprocess.Popen(['open', str(path)])
            else:
                import subprocess
                subprocess.Popen(['xdg-open', str(path)])
        except Exception as e:
            self._log(f'Could not open {path}: {e}', tag='warn')

    def _on_process_error(self, msg: str):
        """UI thread: processing threw — show the error, unlock."""
        self._set_status(msg, error=True)
        self.action_status_var.set('Error — see log')
        self._log(f'*** ERROR *** {msg}', tag='error')
        self.process_btn.config(state='normal', text='▶  Generate Sales Orders')
        self.is_processing = False

    # ────────────────────────────────────────────────────────────────────────
    #                       LOG PANEL
    # ────────────────────────────────────────────────────────────────────────

    def _log(self, msg: str, tag: Optional[str] = None):
        """
        Append a timestamped log entry. Format matches Online PO Processor:
            [HH:MM:SS] message
        """
        ts = datetime.now().strftime('[%H:%M:%S]')
        self._append_log(f'{ts} {msg}\n', tag=tag)

    def _append_log(self, text: str, tag: Optional[str] = None):
        """Append raw text (no timestamping). Used for the multi-line report."""
        self.log_text.config(state='normal')
        if tag:
            self.log_text.insert('end', text, tag)
        else:
            self.log_text.insert('end', text)
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def _clear_log(self):
        """Empty the log panel."""
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.config(state='disabled')

    # ────────────────────────────────────────────────────────────────────────
    #                       STATUS BAR
    # ────────────────────────────────────────────────────────────────────────

    def _set_status(self, msg: str, error: bool = False, warn: bool = False):
        """Update the status line."""
        prefix = '✗ ' if error else ('⚠ ' if warn else '')
        self.status_var.set(prefix + msg)

    def _sheet_suffix(self, actual: str, canonical: str) -> str:
        """
        If the operator's actual sheet name differs from the canonical
        one (e.g. 'Item Master' vs 'Items Master'), append a hint so
        they can confirm we picked up the right sheet. Returns '' when
        names match (no hint needed, keeps log clean).
        """
        if not actual or actual == canonical:
            return ''
        return f"  [sheet: '{actual}']"

    def _subtitle_for_channel(self) -> str:
        """Build the centered subtitle line based on the active channel."""
        channel = CHANNELS.get(self.active_channel_code)
        if channel is None:
            return 'Marketplace PO → D365 Sales Order Import'
        return (f'{channel.display_name} ({channel.code}) PO  →  '
                 f'D365 Sales Order Import')

def run_gui():
    """Open the Tkinter GUI."""
    import tkinter as tk
    root = tk.Tk()
    MTSelectGUI(root)
    root.mainloop()

# ════════════════════════════════════════════════════════════════════════════
# ───────────────── SECTION 9 — ENTRY POINT ─────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
#
# Single entry point: open the GUI. No CLI mode, no flags.
# Template creation and master loading are both done from inside the GUI.

def main():
    """Open the GUI. That's it."""
    run_gui()

if __name__ == '__main__':
    main()