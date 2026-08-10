"""
GT Mass Dump Generator — Django Web Engine (full feature parity with standalone v2.4)
======================================================================================

Features:
    - TemplateValidator (strict: BC Code + Order Qty + PO Number required)
    - MetadataExtractor (Distributor, City, State, Location, SO Number)
    - Location Code mapping (AHD→PICK, BLR→DS_BL_OFF1)
    - 7-sheet Excel output (Headers, Lines, Sales Lines/Header, SKU, Mapping, Warnings)
    - D365 Package Export (fills D365 template with processed data)
    - Email Report (HTML summary via Gmail SMTP)
    - PO Template Download (blank template for the team)
    - ProcessResult tracks attempted_files for full traceability
    - Session-based result storage for multi-step workflows (generate → D365/email)
"""

from __future__ import annotations

import io
import json
import logging
import re
import smtplib
import zipfile
from email.message import EmailMessage
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

LOCATION_CODE_MAP: Dict[str, str] = {
    'AHD': 'PICK',
    'BLR': 'DS_BL_OFF1',
    'NORTH': 'NORTH WH-0',   # North region → North Warehouse (3rd inventory WH)
}

STATE_LIKE_VALUES = {
    "up", "mp", "ap", "hp", "uk", "jk", "wb", "tn", "kl", "ka",
    "gj", "rj", "hr", "pb", "br", "od", "as", "mh", "cg", "jh",
    "north", "south", "east", "west", "central",
    "uttar pradesh", "madhya pradesh", "rajasthan", "punjab",
    "maharashtra", "gujarat", "karnataka", "tamil nadu",
    "haryana", "delhi", "u.p", "u.p.", "m.p", "m.p.",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_str_val(row_vals, idx: Optional[int], as_int_str: bool = False) -> str:
    """Safely extract a string from row_vals[idx]. Returns '' if None/NaN."""
    if idx is None:
        return ''
    val = row_vals[idx]
    if pd.isna(val):
        return ''
    if as_int_str and isinstance(val, (int, float)):
        return str(int(val))
    return str(val).strip()


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class OrderRow:
    """Single ordered item extracted from a GT-Mass file."""
    so_number: str
    item_no: str
    ean: str
    category: str
    description: str
    qty: int
    tester_qty: int
    distributor: str
    city: str
    state: str
    location: str
    location_code: str
    source_file: str


@dataclass
class ProcessResult:
    """
    Aggregated result from processing all uploaded files.

    Fields:
        rows            : All OrderRow objects across all files
        failed_files    : [(filename, reason)] — files that couldn't be parsed
        warned_files    : [(filename, warning)] — non-fatal issues
        attempted_files : ALL filenames in upload order (for File → SO Mapping)
    """
    rows: List[OrderRow] = field(default_factory=list)
    failed_files: List[Tuple[str, str]] = field(default_factory=list)
    warned_files: List[Tuple[str, str]] = field(default_factory=list)
    attempted_files: List[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════════════
#  SO NUMBER FORMATTER
# ═══════════════════════════════════════════════════════════════════════════════

class SONumberFormatter:
    """Extracts SO number from filename digits (fallback)."""

    @staticmethod
    def from_filename(filename: str) -> Optional[str]:
        """
        Extract first digit sequence from filename.
        Example: "SOGTM6325.xlsx" → "SO/GTM/6325"
        """
        stem = Path(filename).stem
        match = re.search(r"\d+", stem)

        if not match:
            logger.warning(f"No digits in filename: {filename}")
            return None

        return f"SO/GTM/{match.group()}"


# ═══════════════════════════════════════════════════════════════════════════════
#  FILE READER — reads Django InMemoryUploadedFile into DataFrame
# ═══════════════════════════════════════════════════════════════════════════════

class FileReader:
    """
    Reads uploaded Excel files into raw DataFrames (no header).

    Handles Django's InMemoryUploadedFile by reading .read() bytes
    into a BytesIO wrapper for pandas.
    """

    @staticmethod
    def read(file_obj: Any, filename: str) -> pd.DataFrame:
        """
        Read an uploaded file into a raw DataFrame.

        Args:
            file_obj : Django InMemoryUploadedFile (has .read() and .name)
            filename : Display name for logging

        Returns:
            DataFrame with integer column indices and no header.

        Raises:
            RuntimeError: if file cannot be read.
        """
        try:
            file_obj.seek(0)
            data = file_obj.read()
            buf = io.BytesIO(data)

            ext = Path(filename).suffix.lower()

            if ext in (".xlsx", ".xlsm"):
                df = pd.read_excel(buf, header=None, engine="openpyxl")
            elif ext == ".xls":
                df = pd.read_excel(buf, header=None, engine="xlrd")
            else:
                raise RuntimeError(f"Unsupported format: '{ext}'")

            logger.info(f"{filename} — read ({len(df)} rows)")
            return df

        except RuntimeError:
            raise
        except Exception as e:
            raise RuntimeError(f"Cannot read '{filename}': {e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  TEMPLATE VALIDATOR
# ═══════════════════════════════════════════════════════════════════════════════

class TemplateValidator:
    """
    Validates that an uploaded file's first sheet matches the GT-Mass template.

    Hard rejections (file skipped entirely):
        1. No header row with 'BC Code' AND 'Order Qty'
        2. No 'PO Number' label with a value

    Soft checks (file still processes, warning logged):
        - Missing Location → MetadataExtractor fires ❌ CRITICAL warning
    """

    @staticmethod
    def validate(file_obj: Any, filename: str) -> Tuple[bool, Optional[str]]:
        """
        Run template compliance checks.

        Args:
            file_obj : Django InMemoryUploadedFile
            filename : Display name

        Returns:
            (is_valid, reason) — reason is None if valid.
        """
        try:
            raw_df = FileReader.read(file_obj, filename)
        except RuntimeError as e:
            return False, str(e)

        # Rule 1: Header row with BC Code + Order Qty
        header_row = TemplateValidator._find_header_row(raw_df)

        if header_row is None:
            return False, (
                "Template violation: header row not found. "
                "File must have a row with BOTH 'BC Code' AND 'Order Qty'."
            )

        # Rule 2: PO Number label with value
        meta_df = raw_df.iloc[:header_row]
        po_found, po_has_value = TemplateValidator._check_po_number(meta_df)

        if not po_found:
            return False, (
                "Template violation: missing 'PO Number' label in meta rows."
            )

        if not po_has_value:
            return False, (
                "Template violation: 'PO Number' label exists but value is empty."
            )

        return True, None

    @staticmethod
    def _find_header_row(raw_df: pd.DataFrame) -> Optional[int]:
        """Scan for the row containing both 'BC Code' and 'Order Qty'."""
        for i, row_vals in enumerate(raw_df.values):
            vals = [str(v).lower() for v in row_vals]

            if "bc code" in vals and any("order qty" in v for v in vals):
                return i

        return None

    @staticmethod
    def _check_po_number(meta_df: pd.DataFrame) -> Tuple[bool, bool]:
        """
        Check if 'PO Number' label exists with a non-empty adjacent value.

        Returns:
            (label_found, has_value)
        """
        for _, row in meta_df.iterrows():
            for col_idx in range(min(len(row) - 1, 10)):
                if pd.isna(row.iloc[col_idx]):
                    continue

                if str(row.iloc[col_idx]).strip().lower() == "po number":
                    # Check next 1-2 cells for a value
                    for offset in range(1, 3):
                        check = col_idx + offset

                        if check >= len(row):
                            break

                        val = row.iloc[check]

                        if pd.notna(val) and str(val).strip() and str(val).strip().lower() != 'nan':
                            return True, True

                    return True, False

        return False, False


# ═══════════════════════════════════════════════════════════════════════════════
#  METADATA EXTRACTOR
# ═══════════════════════════════════════════════════════════════════════════════

class MetadataExtractor:
    """Extracts meta fields (SO#, Distributor, City, State, Location) from header rows."""

    @staticmethod
    def extract(raw_df: pd.DataFrame, header_row: int) -> Tuple[dict, List[str]]:
        """
        Scan rows 0..header_row-1 for meta field labels and values.

        Returns:
            (meta_dict, warnings_list)
        """
        distributor = ""
        city = ""
        location = ""
        so_number = ""
        state_values: List[str] = []
        warnings: List[str] = []

        meta_df = raw_df.iloc[:header_row]

        for _, row in meta_df.iterrows():
            # LEFT SIDE: Col A (label) + Col B (value)
            label = ""
            if pd.notna(row.iloc[0]):
                label = str(row.iloc[0]).strip().lower()

            value = ""
            if pd.notna(row.iloc[1]):
                value = str(row.iloc[1]).strip()
                if value.lower() == "nan":
                    value = ""

            if label == "distributor name" and not distributor:
                distributor = value

            elif label == "city" and not city:
                city = value

            elif label == "state":
                state_values.append(value)

            # RIGHT SIDE: scan cols 0-9 for "PO Number" / "Location"
            for col_idx in range(min(len(row) - 1, 10)):
                if pd.isna(row.iloc[col_idx]):
                    continue

                cell_text = str(row.iloc[col_idx]).strip().lower()

                if cell_text == "location":
                    for vi in range(col_idx + 1, min(col_idx + 3, len(row))):
                        lv = row.iloc[vi]
                        if pd.notna(lv) and str(lv).strip() and str(lv).strip().lower() != 'nan':
                            location = str(lv).strip()
                            break

                elif cell_text == "po number" and not so_number:
                    for vi in range(col_idx + 1, min(col_idx + 3, len(row))):
                        pv = row.iloc[vi]
                        if pd.notna(pv) and str(pv).strip() and str(pv).strip().lower() != 'nan':
                            so_number = str(pv).strip()
                            break

        # Resolve state
        state = next((s for s in reversed(state_values) if s), "")

        # Map Location → ERP Location Code
        location_code = ""
        if location:
            location_code = LOCATION_CODE_MAP.get(location.upper().strip(), location)

        # Warnings
        if not distributor:
            warnings.append("Distributor Name is blank.")

        if not city:
            warnings.append("City is blank.")

        if not state:
            warnings.append("State is blank.")

        if not location_code:
            warnings.append(
                "❌ CRITICAL: Location Code is EMPTY — "
                "ERP import will fail without Location Code."
            )

        if distributor and distributor.strip().lower() in STATE_LIKE_VALUES:
            warnings.append(
                f"Distributor '{distributor}' looks like a state — verify."
            )

        return {
            "distributor": distributor,
            "city": city,
            "state": state,
            "location": location,
            "location_code": location_code,
            "so_number": so_number,
        }, warnings


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelParser:
    """Parses a single GT-Mass file into OrderRow objects."""

    BC_COLUMN = "bc code"
    QTY_COLUMN = "order qty"
    TESTER_COLUMN = "tester qty"

    def parse(self, file_obj: Any, filename: str) -> Tuple[List[OrderRow], List[str]]:
        """
        Parse one uploaded file end-to-end.

        Args:
            file_obj : Django InMemoryUploadedFile
            filename : Display name

        Returns:
            (list of OrderRow, list of warnings)

        Raises:
            RuntimeError: if file has broken structure.
        """
        logger.info(f"Parsing: {filename}")

        warnings: List[str] = []

        # Read raw
        raw_df = FileReader.read(file_obj, filename)

        # Find header row
        header_row = self._find_header_row(raw_df)

        # Extract meta
        meta, meta_warnings = MetadataExtractor.extract(raw_df, header_row)
        warnings.extend(meta_warnings)

        # Resolve SO number
        so_number, so_warnings = self._resolve_so_number(meta, filename)
        warnings.extend(so_warnings)

        # Build data table
        df = raw_df.iloc[header_row + 1:].copy()
        df.columns = raw_df.iloc[header_row].values
        df = df.reset_index(drop=True)

        # Extract rows
        rows, extract_warnings = self._extract_rows(df, so_number, meta, filename)
        warnings.extend(extract_warnings)

        return rows, warnings

    def _find_header_row(self, raw_df: pd.DataFrame) -> int:
        """Find the row containing 'BC Code' + 'Order Qty'."""
        for i, row_vals in enumerate(raw_df.values):
            vals = [str(v).lower() for v in row_vals]
            if "bc code" in vals and any("order qty" in v for v in vals):
                return i

        raise RuntimeError("Header row not found — no 'BC Code' + 'Order Qty'.")

    def _resolve_so_number(self, meta: dict, filename: str) -> Tuple[str, List[str]]:
        """Resolve SO: file PO Number → filename digits → UNKNOWN."""
        warnings: List[str] = []
        so = meta.get("so_number", "")

        if so:
            return so, warnings

        so = SONumberFormatter.from_filename(filename)
        if so:
            warnings.append(f"SO from filename: '{so}'. Fill PO Number field.")
            return so, warnings

        warnings.append("SO not found — using 'SO/GTM/UNKNOWN'.")
        return "SO/GTM/UNKNOWN", warnings

    def _extract_rows(
        self, df: pd.DataFrame, so_number: str,
        meta: dict, filename: str
    ) -> Tuple[List[OrderRow], List[str]]:
        """Build OrderRow list from the data table."""
        warnings: List[str] = []

        bc_col, qty_col, tester_col, ean_col, cat_col, desc_col = self._detect_columns(df)

        if bc_col is None:
            raise RuntimeError("'BC Code' column not found.")
        if qty_col is None:
            raise RuntimeError("'Order Qty' column not found.")
        if tester_col is None:
            warnings.append("'Tester Qty' not found — defaulting to 0.")

        bc_idx = df.columns.get_loc(bc_col)
        qty_idx = df.columns.get_loc(qty_col)
        tester_idx = df.columns.get_loc(tester_col) if tester_col else None
        ean_idx = df.columns.get_loc(ean_col) if ean_col else None
        cat_idx = df.columns.get_loc(cat_col) if cat_col else None
        desc_idx = df.columns.get_loc(desc_col) if desc_col else None

        rows: List[OrderRow] = []

        for rv in df.values:
            bc = rv[bc_idx]
            if pd.isna(bc):
                continue
            try:
                bc = int(bc)
            except (ValueError, TypeError):
                continue

            qty = self._clean_qty(rv[qty_idx])
            tqty = self._clean_qty(rv[tester_idx]) if tester_idx is not None else 0

            if qty <= 0 and tqty <= 0:
                continue

            rows.append(OrderRow(
                so_number=so_number,
                item_no=str(bc),
                ean=safe_str_val(rv, ean_idx, as_int_str=True),
                category=safe_str_val(rv, cat_idx),
                description=safe_str_val(rv, desc_idx),
                qty=qty,
                tester_qty=tqty,
                distributor=meta["distributor"],
                city=meta["city"],
                state=meta["state"],
                location=meta["location"],
                location_code=meta["location_code"],
                source_file=filename,
            ))

        if not rows:
            warnings.append("No ordered rows — all quantities are 0.")

        return rows, warnings

    def _detect_columns(self, df) -> Tuple[Optional[str], ...]:
        """Find BC Code, Order Qty, Tester Qty, EAN, Category, Description."""
        bc = qty = tester = ean = cat = desc = None

        for col in df.columns:
            name = str(col).strip().lower()
            if name == self.BC_COLUMN:
                bc = col
            if self.QTY_COLUMN in name:
                qty = col
            if self.TESTER_COLUMN in name:
                tester = col
            if name == 'ean' and not ean:
                ean = col
            if name == 'category' and not cat:
                cat = col
            if 'article description' in name:
                desc = col
            elif name == 'description' and not desc:
                desc = col

        return bc, qty, tester, ean, cat, desc

    @staticmethod
    def _clean_qty(value) -> int:
        """Clean quantity cell: NaN → 0, commas stripped, float→int."""
        if pd.isna(value):
            return 0
        value = str(value).strip()
        if value in ("", "-"):
            return 0
        value = value.replace(",", "")
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0


# ═══════════════════════════════════════════════════════════════════════════════
#  DUMP EXPORTER — writes 7-sheet Excel to BytesIO
# ═══════════════════════════════════════════════════════════════════════════════

class DumpExporter:
    """Writes the 7-sheet output Excel to memory (BytesIO) for HTTP response."""

    HEADER_FILL = PatternFill('solid', fgColor='1A237E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
    DATA_FONT = Font(name='Aptos Display', size=11)
    THIN_SIDE = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    def _hdr_cell(self, ws, row, col, value):
        """Create a formatted header cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.BORDER
        return cell

    def _data_cell(self, ws, row, col, value, fmt=None):
        """Create a formatted data cell."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.DATA_FONT
        cell.border = self.BORDER
        if fmt:
            cell.number_format = fmt
        return cell

    def _auto_width(self, ws, max_w=50):
        """Auto-fit column widths."""
        for col in ws.columns:
            letter = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[letter].width = min(w + 3, max_w)

    def export_to_memory(self, result: ProcessResult) -> Optional[io.BytesIO]:
        """
        Write all 7 sheets to an in-memory Excel file.

        Args:
            result: ProcessResult from the engine

        Returns:
            BytesIO buffer with the Excel file, or None if no data.
        """
        if not result.rows and not result.attempted_files:
            return None

        wb = Workbook()
        wb.remove(wb.active)

        if result.rows:
            self._write_headers_so(wb, result)
            self._write_lines_so(wb, result)
            self._write_sales_lines(wb, result)
            self._write_sales_header(wb, result)
            self._write_sku_summary(wb, result)

        self._write_file_so_mapping(wb, result)
        self._write_warnings(wb, result)

        if not wb.sheetnames:
            wb.create_sheet('Empty')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    # ─────────────────────────────────────────────────────────────────
    #  Sheet writers (same logic as standalone v2.4)
    # ─────────────────────────────────────────────────────────────────

    def _write_headers_so(self, wb, result: ProcessResult):
        """Sheet 1: Headers (SO) — one row per unique SO."""
        ws = wb.create_sheet('Headers (SO)')

        headers = [
            'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
            'Posting Date', 'Order Date', 'Document Date',
            'Invoice From Date', 'Invoice To Date',
            'External Document No.', 'Location Code', 'Dimension Set ID',
            'Supply Type', 'Voucher Narration',
            'Brand Code (Dimension)', 'Channel Code (Dimension)',
            'Catagory (Dimension)', 'Geography Code (Dimension)',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        today_str = datetime.now().strftime("%d-%m-%Y")

        seen = set()
        unique_sos = []
        for row in result.rows:
            if row.so_number not in seen:
                seen.add(row.so_number)
                unique_sos.append(row)

        r = 2
        for row in unique_sos:
            self._data_cell(ws, r, 1, 'Order')
            self._data_cell(ws, r, 2, row.so_number)
            self._data_cell(ws, r, 3, '')
            self._data_cell(ws, r, 4, '')

            for c in range(5, 10):
                self._data_cell(ws, r, c, today_str)

            self._data_cell(ws, r, 10, row.so_number)
            self._data_cell(ws, r, 11, row.location_code)
            self._data_cell(ws, r, 12, '')
            self._data_cell(ws, r, 13, 'B2B')
            r += 1

        self._auto_width(ws)

    def _write_lines_so(self, wb, result: ProcessResult):
        """Sheet 2: Lines (SO) — one row per item, 10K line increments."""
        ws = wb.create_sheet('Lines (SO)')

        headers = [
            'Document Type', 'Document No.', 'Line No.', 'Type',
            'No.', 'Location Code', 'Quantity', 'Unit Price',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        r = 2
        current_so = None
        line_no = 0

        for row in result.rows:
            if row.so_number != current_so:
                current_so = row.so_number
                line_no = 0
            line_no += 10000

            self._data_cell(ws, r, 1, 'Order')
            self._data_cell(ws, r, 2, row.so_number)
            self._data_cell(ws, r, 3, line_no)
            self._data_cell(ws, r, 4, 'Item')
            self._data_cell(ws, r, 5, row.item_no)
            self._data_cell(ws, r, 6, row.location_code)
            self._data_cell(ws, r, 7, row.qty)
            self._data_cell(ws, r, 8, '')
            r += 1

        self._auto_width(ws)

    def _write_sales_lines(self, wb, result: ProcessResult):
        """Sheet 3: Sales Lines — detailed flat list."""
        ws = wb.create_sheet('Sales Lines')

        headers = [
            'SO Number', 'EAN', 'BC Code', 'Category',
            'Article Description', 'Order Qty', 'Tester Qty',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        for r, row in enumerate(result.rows, 2):
            self._data_cell(ws, r, 1, row.so_number)
            self._data_cell(ws, r, 2, row.ean)
            self._data_cell(ws, r, 3, row.item_no)
            self._data_cell(ws, r, 4, row.category)
            self._data_cell(ws, r, 5, row.description)
            self._data_cell(ws, r, 6, row.qty)
            self._data_cell(ws, r, 7, row.tester_qty)

        self._auto_width(ws)

    def _write_sales_header(self, wb, result: ProcessResult):
        """Sheet 4: Sales Header — grouped summary per SO."""
        ws = wb.create_sheet('Sales Header')

        headers = [
            'SO Number', 'Order Qty', 'Tester Qty', 'Total Qty',
            'Distributor', 'City', 'State', 'Location',
        ]

        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        so_groups: Dict[str, dict] = {}
        for row in result.rows:
            if row.so_number not in so_groups:
                so_groups[row.so_number] = {
                    'oq': 0, 'tq': 0,
                    'd': row.distributor, 'c': row.city,
                    's': row.state, 'l': row.location,
                }
            so_groups[row.so_number]['oq'] += row.qty
            so_groups[row.so_number]['tq'] += row.tester_qty

        r = 2
        for so, info in so_groups.items():
            self._data_cell(ws, r, 1, so)
            self._data_cell(ws, r, 2, info['oq'])
            self._data_cell(ws, r, 3, info['tq'])
            self._data_cell(ws, r, 4, info['oq'] + info['tq'])
            self._data_cell(ws, r, 5, info['d'])
            self._data_cell(ws, r, 6, info['c'])
            self._data_cell(ws, r, 7, info['s'])
            self._data_cell(ws, r, 8, info['l'])
            r += 1

        self._auto_width(ws)

    def _write_sku_summary(self, wb, result: ProcessResult):
        """Sheet 5: SKU Summary — demand pivot per BC Code."""
        ws = wb.create_sheet('SKU Summary')

        headers = ['BC Code', 'Description', 'Category', 'Order Qty', 'Tester Qty', 'Total Qty']
        for c, h in enumerate(headers, 1):
            self._hdr_cell(ws, 1, c, h)

        sku: Dict[str, dict] = {}
        for row in result.rows:
            if row.item_no not in sku:
                sku[row.item_no] = {'d': row.description, 'c': row.category, 'oq': 0, 'tq': 0}
            sku[row.item_no]['oq'] += row.qty
            sku[row.item_no]['tq'] += row.tester_qty
            if not sku[row.item_no]['d'] and row.description:
                sku[row.item_no]['d'] = row.description
            if not sku[row.item_no]['c'] and row.category:
                sku[row.item_no]['c'] = row.category

        sorted_skus = sorted(sku.items(), key=lambda x: x[1]['oq'] + x[1]['tq'], reverse=True)

        r = 2
        go = gt = 0
        for item, info in sorted_skus:
            t = info['oq'] + info['tq']
            go += info['oq']
            gt += info['tq']
            self._data_cell(ws, r, 1, item)
            self._data_cell(ws, r, 2, info['d'])
            self._data_cell(ws, r, 3, info['c'])
            self._data_cell(ws, r, 4, info['oq'])
            self._data_cell(ws, r, 5, info['tq'])
            self._data_cell(ws, r, 6, t)
            r += 1

        bold = Font(name='Aptos Display', size=11, bold=True)
        ws.cell(row=r, column=1, value='GRAND TOTAL').font = bold
        ws.cell(row=r, column=2, value=f'{len(sorted_skus)} unique SKUs').font = bold
        ws.cell(row=r, column=4, value=go).font = bold
        ws.cell(row=r, column=5, value=gt).font = bold
        ws.cell(row=r, column=6, value=go + gt).font = bold
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = self.BORDER

        self._auto_width(ws)

    def _write_file_so_mapping(self, wb, result: ProcessResult):
        """
        Sheet 6: File → SO Mapping — every uploaded file gets an entry.

        Status:
            ✅ OK       — parsed cleanly
            ⚠️ WARNING  — parsed with warnings
            ❌ FAILED   — rejected at validation or parse stage
        """
        ws = wb.create_sheet('File → SO Mapping')

        for c, h in enumerate(['Sr No', 'Filename', 'SO Number', 'Status'], 1):
            self._hdr_cell(ws, 1, c, h)

        file_to_so = {}
        for row in result.rows:
            if row.source_file not in file_to_so:
                file_to_so[row.source_file] = row.so_number

        failed_map = {f: r for f, r in result.failed_files}
        warned_set = {f for f, _ in result.warned_files}

        red_fill = PatternFill('solid', fgColor='FFCDD2')
        red_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')
        yellow_fill = PatternFill('solid', fgColor='FFF9C4')

        sr = 1
        ok_count = warn_count = fail_count = 0

        for filename in result.attempted_files:
            r = sr + 1

            if filename in failed_map:
                self._data_cell(ws, r, 1, sr)
                self._data_cell(ws, r, 2, filename)
                self._data_cell(ws, r, 3, f"❌ FAILED: {failed_map[filename]}")
                self._data_cell(ws, r, 4, '❌ FAILED')
                for c in range(1, 5):
                    ws.cell(row=r, column=c).fill = red_fill
                    ws.cell(row=r, column=c).font = red_font
                fail_count += 1

            elif filename in file_to_so:
                so = file_to_so[filename]

                if filename in warned_set:
                    self._data_cell(ws, r, 1, sr)
                    self._data_cell(ws, r, 2, filename)
                    self._data_cell(ws, r, 3, so)
                    self._data_cell(ws, r, 4, '⚠️ WARNING')
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = yellow_fill
                    warn_count += 1
                else:
                    self._data_cell(ws, r, 1, sr)
                    self._data_cell(ws, r, 2, filename)
                    self._data_cell(ws, r, 3, so)
                    self._data_cell(ws, r, 4, '✅ OK')
                    ok_count += 1
            else:
                self._data_cell(ws, r, 1, sr)
                self._data_cell(ws, r, 2, filename)
                self._data_cell(ws, r, 3, '(no data)')
                self._data_cell(ws, r, 4, '❌ FAILED')
                for c in range(1, 5):
                    ws.cell(row=r, column=c).fill = red_fill
                    ws.cell(row=r, column=c).font = red_font
                fail_count += 1

            sr += 1

        # Summary row
        summary_r = sr + 1
        bold = Font(name='Aptos Display', size=11, bold=True)
        ws.cell(row=summary_r, column=1, value='TOTAL').font = bold
        ws.cell(row=summary_r, column=2,
                value=f'{len(result.attempted_files)} file(s) attempted').font = bold
        ws.cell(row=summary_r, column=3,
                value=f'✅ {ok_count} OK | ⚠️ {warn_count} warn | ❌ {fail_count} failed').font = bold
        for c in range(1, 5):
            ws.cell(row=summary_r, column=c).border = self.BORDER

        self._auto_width(ws)

    def _write_warnings(self, wb, result: ProcessResult):
        """Sheet 7: Warnings — failures first (red), then warnings."""
        if not result.warned_files and not result.failed_files:
            return

        ws = wb.create_sheet('Warnings')
        for c, h in enumerate(['File', 'Type', 'Message'], 1):
            self._hdr_cell(ws, 1, c, h)

        red_fill = PatternFill('solid', fgColor='FFCDD2')
        red_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')

        r = 2

        for fname, reason in result.failed_files:
            self._data_cell(ws, r, 1, fname)
            self._data_cell(ws, r, 2, '❌ FAILED')
            self._data_cell(ws, r, 3, reason)
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = red_fill
                ws.cell(row=r, column=c).font = red_font
            r += 1

        for fname, warning in result.warned_files:
            is_critical = '❌ CRITICAL' in warning
            self._data_cell(ws, r, 1, fname)
            self._data_cell(ws, r, 2, '❌ CRITICAL' if is_critical else '⚠️ WARNING')
            self._data_cell(ws, r, 3, warning)
            if is_critical:
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = red_fill
                    ws.cell(row=r, column=c).font = red_font
            r += 1

        self._auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN AUTOMATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class GTMassAutomation:
    """
    Orchestrates validation, parsing, and export for Django web context.

    Replaces the simple process_files → List[OrderRow] with a full
    ProcessResult that tracks attempted files, failures, and warnings.
    """

    def __init__(self):
        self.validator = TemplateValidator()
        self.parser = ExcelParser()
        self.exporter = DumpExporter()

    def process_files(self, file_objects: List[Any]) -> ProcessResult:
        """
        Process all uploaded files.

        Args:
            file_objects: List of Django InMemoryUploadedFile objects

        Returns:
            ProcessResult with rows, warnings, failures, attempted_files
        """
        result = ProcessResult()

        for file_obj in file_objects:
            fname = file_obj.name
            result.attempted_files.append(fname)

            # Validate template compliance
            is_valid, reason = self.validator.validate(file_obj, fname)

            if not is_valid:
                result.failed_files.append((fname, reason))
                logger.error(f"{fname} REJECTED: {reason}")
                continue

            # Reset file pointer after validation read it
            file_obj.seek(0)

            # Parse
            try:
                rows, warnings = self.parser.parse(file_obj, fname)
                result.rows.extend(rows)

                for w in warnings:
                    result.warned_files.append((fname, w))
                    logger.warning(f"{fname}: {w}")

            except RuntimeError as e:
                result.failed_files.append((fname, str(e)))
                logger.error(f"{fname} FAILED: {e}")

            except (ValueError, KeyError, TypeError) as e:
                result.failed_files.append((fname, f"Data error: {e}"))
                logger.error(f"{fname} DATA: {e}")

            except Exception as e:
                result.failed_files.append((fname, f"Unexpected: {e}"))
                logger.error(f"{fname} UNEXPECTED: {e}")

        logger.info(
            f"Done — {len(result.attempted_files)} attempted | "
            f"{len(result.rows)} rows | "
            f"{len({r.so_number for r in result.rows})} SOs | "
            f"{len(result.failed_files)} failed | "
            f"{len(result.warned_files)} warnings"
        )

        return result

# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

import os as _os  # noqa: E402 — local alias for env-sourced email creds


def _email_cred(key: str, default: str = '') -> str:
    """Email cred from the environment first (.env on a host), else the gitignored
    online_po_management/Calculation Data/email_config.json — NEVER hardcoded."""
    v = _os.environ.get(key)
    if v:
        return v
    try:
        import json as _json
        from pathlib import Path as _P
        here = _P(__file__).resolve()
        for base in [here, *here.parents]:
            j = base / 'online_po_management' / 'Calculation Data' / 'email_config.json'
            if j.exists():
                return _json.loads(j.read_text(encoding='utf-8-sig')).get(key, default)
    except Exception:  # noqa: BLE001
        pass
    return default


EMAIL_CONFIG = {
    # Credentials from the environment (.env on a host) or the gitignored
    # Calculation Data/email_config.json — never hardcoded / committed. Blank when
    # unset → email simply doesn't send (soft-fail).
    'EMAIL_SENDER': _email_cred('EMAIL_SENDER'),
    'EMAIL_PASSWORD': _email_cred('EMAIL_PASSWORD'),
    'SMTP_SERVER': _os.environ.get('SMTP_SERVER', 'smtp.gmail.com'),
    'SMTP_PORT': int(_os.environ.get('SMTP_PORT', '587')),
    'DEFAULT_RECIPIENT': _os.environ.get('EMAIL_DEFAULT_RECIPIENT',
                                         'abhishek.wagh@reneecosmetics.in'),
    'CC_RECIPIENTS': [
        'offlineb2b@reneecosmetics.in',
        'kirpalsinh.bihola@reneecosmetics.in',
        'aritra.barmanray@reneecosmetics.in',
        'milan.nayak@reneecosmetics.in',
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
#  COLORS — centralized palette for email HTML
# ═══════════════════════════════════════════════════════════════════════════════

class Colors:
    """Centralized color palette for email HTML."""
    NAVY   = '#1A237E'
    GREEN  = '#2E7D32'
    ORANGE = '#E65100'
    PURPLE = '#6A1B9A'
    GOLD   = '#FFD600'


# ═══════════════════════════════════════════════════════════════════════════════
#  SESSION SERIALIZATION — store/restore ProcessResult across requests
# ═══════════════════════════════════════════════════════════════════════════════

def format_indian(number) -> str:
    """Format number in Indian system: 1,23,456."""
    try:
        number = float(number)
    except (ValueError, TypeError):
        return str(number)
    sign = '-' if number < 0 else ''
    number = abs(number)
    if number == int(number):
        int_part, dec_part = str(int(number)), ''
    else:
        parts = f"{number:.2f}".split('.')
        int_part, dec_part = parts[0], '.' + parts[1]
    if len(int_part) <= 3:
        return sign + int_part + dec_part
    result = int_part[-3:]
    remaining = int_part[:-3]
    while remaining:
        result = remaining[-2:] + ',' + result
        remaining = remaining[:-2]
    return sign + result + dec_part


def result_to_session(result: ProcessResult) -> dict:
    """
    Serialize ProcessResult to a JSON-safe dict for Django session storage.

    Django sessions serialize to JSON — dataclasses and tuples need conversion.
    """
    return {
        'rows': [
            {
                'so_number': r.so_number,
                'item_no': r.item_no,
                'ean': r.ean,
                'category': r.category,
                'description': r.description,
                'qty': r.qty,
                'tester_qty': r.tester_qty,
                'distributor': r.distributor,
                'city': r.city,
                'state': r.state,
                'location': r.location,
                'location_code': r.location_code,
                'source_file': r.source_file,
            }
            for r in result.rows
        ],
        'failed_files': [[f, r] for f, r in result.failed_files],
        'warned_files': [[f, w] for f, w in result.warned_files],
        'attempted_files': result.attempted_files,
    }


def result_from_session(data: dict) -> ProcessResult:
    """
    Deserialize a session dict back into a ProcessResult.
    """
    return ProcessResult(
        rows=[OrderRow(**r) for r in data['rows']],
        failed_files=[(f, r) for f, r in data['failed_files']],
        warned_files=[(f, w) for f, w in data['warned_files']],
        attempted_files=data['attempted_files'],
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  D365 PACKAGE EXPORTER — fills D365 template with processed data
# ═══════════════════════════════════════════════════════════════════════════════

class D365Exporter:
    """
    Fills a D365 sample package template with processed data.

    The template has pre-formatted empty rows. We replace empty cells
    via regex in the underlying XML. If data exceeds template capacity,
    new <row> elements are injected before filling.

    Returns BytesIO for HTTP response.
    """

    @staticmethod
    def export(result: ProcessResult, template_file: Any) -> io.BytesIO:
        """
        Fill the D365 template with data from ProcessResult.

        Args:
            result        : ProcessResult with rows to export
            template_file : Django InMemoryUploadedFile (the D365 template)

        Returns:
            BytesIO with the filled Excel file.

        Raises:
            ValueError: if result has no rows.
            RuntimeError: if template format is invalid.
        """
        if not result.rows:
            raise ValueError("No data to export.")

        import re as re_mod

        # Read template bytes
        template_file.seek(0)
        template_bytes = template_file.read()
        buf = io.BytesIO(template_bytes)

        today_str = datetime.now().strftime("%d-%m-%Y")

        # Collect unique SOs in order
        seen = set()
        unique_sos = []
        for row in result.rows:
            if row.so_number not in seen:
                seen.add(row.so_number)
                unique_sos.append(row)

        # Read ZIP contents
        zip_contents = {}
        with zipfile.ZipFile(buf, 'r') as z:
            for item in z.namelist():
                zip_contents[item] = z.read(item)

        # ── Extend sharedStrings.xml ──
        ss_xml = zip_contents['xl/sharedStrings.xml'].decode('utf-8')
        existing = re_mod.findall(r'<t[^>]*>([^<]*)</t>', ss_xml)
        string_map = {s: i for i, s in enumerate(existing)}

        new_strings = {'Order', 'Item', 'B2B', today_str}
        for r in unique_sos:
            new_strings.add(r.so_number)
            if r.location_code:
                new_strings.add(r.location_code)
        for r in result.rows:
            new_strings.add(r.so_number)
            if r.location_code:
                new_strings.add(r.location_code)

        next_idx = len(existing)
        for s in sorted(new_strings):
            if s not in string_map:
                string_map[s] = next_idx
                next_idx += 1

        total_count = next_idx
        si_items = [''] * total_count
        for s, idx in string_map.items():
            esc = s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            si_items[idx] = f'<si><t>{esc}</t></si>'

        new_ss = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{total_count}" uniqueCount="{total_count}">'
            + ''.join(si_items) + '</sst>'
        )
        zip_contents['xl/sharedStrings.xml'] = new_ss.encode('utf-8')

        # ── Helpers ──
        def fill_cell(xml, col, row_num, value, is_string=True):
            ref = f"{col}{row_num}"
            pat = f'<c r="{ref}" s="(\\d+)"\\s*/>'
            if is_string:
                idx = string_map.get(str(value), 0)
                rep = f'<c r="{ref}" s="\\1" t="s"><v>{idx}</v></c>'
            else:
                rep = f'<c r="{ref}" s="\\1"><v>{value}</v></c>'
            return re_mod.sub(pat, rep, xml, count=1)

        def inject_row(xml, row_num, columns, style_id):
            cells = ''.join(f'<c r="{c}{row_num}" s="{style_id}"/>' for c in columns)
            new_row = (
                f'<row r="{row_num}" spans="1:{len(columns)}" '
                f'x14ac:dyDescent="0.3">{cells}</row>'
            )
            return xml.replace('</sheetData>', new_row + '</sheetData>')

        # ── Sheet 1: Sales Header ──
        s1 = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        s1_rows = len(re_mod.findall(r'<row r="(\d+)"', s1)) - 2
        hdr_cols = list('ABCDEFGHIJKLMNOPQR')

        if len(unique_sos) > s1_rows:
            for extra in range(s1_rows + 4, len(unique_sos) + 4):
                s1 = inject_row(s1, extra, hdr_cols, '11')

        for i, row in enumerate(unique_sos):
            r = i + 4
            s1 = fill_cell(s1, 'A', r, 'Order')
            s1 = fill_cell(s1, 'B', r, row.so_number)
            for c in 'EFGHI':
                s1 = fill_cell(s1, c, r, today_str)
            s1 = fill_cell(s1, 'J', r, row.so_number)
            if row.location_code:
                s1 = fill_cell(s1, 'K', r, row.location_code)
            s1 = fill_cell(s1, 'M', r, 'B2B')

        zip_contents['xl/worksheets/sheet1.xml'] = s1.encode('utf-8')

        # ── Sheet 2: Sales Line ──
        s2 = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        s2_rows = len(re_mod.findall(r'<row r="(\d+)"', s2)) - 3
        line_cols = list('ABCDEFGH')

        if len(result.rows) > s2_rows:
            for extra in range(s2_rows + 4, len(result.rows) + 4):
                s2 = inject_row(s2, extra, line_cols, '8')

        current_so = None
        line_no = 0
        for i, row in enumerate(result.rows):
            if row.so_number != current_so:
                current_so = row.so_number
                line_no = 0
            line_no += 10000
            r = i + 4

            s2 = fill_cell(s2, 'A', r, 'Order')
            s2 = fill_cell(s2, 'B', r, row.so_number)
            s2 = fill_cell(s2, 'C', r, line_no, is_string=False)
            s2 = fill_cell(s2, 'D', r, 'Item')
            try:
                s2 = fill_cell(s2, 'E', r, int(row.item_no), is_string=False)
            except (ValueError, TypeError):
                s2 = fill_cell(s2, 'E', r, row.item_no)
            if row.location_code:
                s2 = fill_cell(s2, 'F', r, row.location_code)
            s2 = fill_cell(s2, 'G', r, row.qty, is_string=False)

        zip_contents['xl/worksheets/sheet2.xml'] = s2.encode('utf-8')

        # ── Cleanup: remove empty rows, update dimensions/table refs ──
        last_hdr = 3 + len(unique_sos)
        last_line = 3 + len(result.rows)

        s1c = zip_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
        for r in range(last_hdr + 1, 100):
            s1c = re_mod.sub(rf'<row r="{r}"[^>]*>.*?</row>', '', s1c, flags=re_mod.DOTALL)
        s1c = re_mod.sub(r'<dimension ref="[^"]*"/>', f'<dimension ref="A1:R{last_hdr}"/>', s1c)
        zip_contents['xl/worksheets/sheet1.xml'] = s1c.encode('utf-8')

        s2c = zip_contents['xl/worksheets/sheet2.xml'].decode('utf-8')
        for r in range(last_line + 1, 1000):
            s2c = re_mod.sub(rf'<row r="{r}"[^>]*>.*?</row>', '', s2c, flags=re_mod.DOTALL)
        s2c = re_mod.sub(r'<dimension ref="[^"]*"/>', f'<dimension ref="A1:H{last_line}"/>', s2c)
        zip_contents['xl/worksheets/sheet2.xml'] = s2c.encode('utf-8')

        if 'xl/tables/table1.xml' in zip_contents:
            t1 = zip_contents['xl/tables/table1.xml'].decode('utf-8')
            t1 = re_mod.sub(r'ref="A3:[A-Z]+\d+"', f'ref="A3:R{last_hdr}"', t1)
            zip_contents['xl/tables/table1.xml'] = t1.encode('utf-8')

        if 'xl/tables/table2.xml' in zip_contents:
            t2 = zip_contents['xl/tables/table2.xml'].decode('utf-8')
            t2 = re_mod.sub(r'ref="A3:[A-Z]+\d+"', f'ref="A3:H{last_line}"', t2)
            zip_contents['xl/tables/table2.xml'] = t2.encode('utf-8')

        # ── Write final ZIP ──
        output = io.BytesIO()
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zo:
            for name, data in zip_contents.items():
                zo.writestr(name, data)

        output.seek(0)
        logger.info(f"D365 export: {len(unique_sos)} SOs, {len(result.rows)} items")
        return output


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL BUILDER + SENDER
# ═══════════════════════════════════════════════════════════════════════════════

class EmailBuilder:
    """Pure data → HTML transform. No network I/O."""

    @staticmethod
    def build_subject(result: ProcessResult) -> str:
        ts = datetime.now().strftime('%d-%m-%Y %H:%M')
        so_count = len({r.so_number for r in result.rows})
        return f"📊 GT Mass SO Report: {so_count} SOs, {len(result.rows)} Items — {ts}"

    @staticmethod
    def build_html(result: ProcessResult, elapsed_str: str = '') -> str:
        """Build the full HTML email body."""
        C = Colors
        unique_sos = list({r.so_number: r for r in result.rows}.values())
        total_order = sum(r.qty for r in result.rows)
        total_tester = sum(r.tester_qty for r in result.rows)
        total_qty = total_order + total_tester
        ts = datetime.now().strftime('%d-%m-%Y %H:%M:%S')

        # SO-level aggregation
        so_groups: Dict[str, dict] = {}
        for r in result.rows:
            if r.so_number not in so_groups:
                so_groups[r.so_number] = {'order': 0, 'tester': 0}
            so_groups[r.so_number]['order'] += r.qty
            so_groups[r.so_number]['tester'] += r.tester_qty

        # SKU aggregation
        sku_groups: Dict[str, dict] = {}
        for r in result.rows:
            if r.item_no not in sku_groups:
                sku_groups[r.item_no] = {'desc': r.description, 'cat': r.category, 'order': 0, 'tester': 0}
            sku_groups[r.item_no]['order'] += r.qty
            sku_groups[r.item_no]['tester'] += r.tester_qty
            if not sku_groups[r.item_no]['desc'] and r.description:
                sku_groups[r.item_no]['desc'] = r.description

        sorted_skus = sorted(sku_groups.items(), key=lambda x: x[1]['order']+x[1]['tester'], reverse=True)

        # ── Build HTML ──
        html = f'''<html><body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f0f2f5;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f2f5;">
<tr><td align="center" style="padding:20px 10px;">
<table width="800" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:8px;overflow:hidden;border:1px solid #ddd;">
<tr><td style="background:{C.NAVY};padding:25px 30px;text-align:center;">
    <p style="margin:0;font-size:22px;font-weight:bold;color:white;">📊 GT Mass — Sales Order Report</p>
    <p style="margin:8px 0 0;font-size:12px;color:#9fa8da;">Generated: {ts} | Processing: {elapsed_str}</p>
    <table style="margin:10px auto 0;"><tr><td style="background:#283593;padding:5px 15px;border-radius:15px;">
        <span style="font-size:10px;color:#9fa8da;letter-spacing:1px;">⚡ GT MASS DUMP GENERATOR v2.4</span>
    </td></tr></table>
</td></tr>
<tr><td style="height:4px;font-size:0;"><table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td width="25%" style="background:{C.ORANGE};height:4px;"></td>
    <td width="25%" style="background:{C.GOLD};height:4px;"></td>
    <td width="25%" style="background:#00E676;height:4px;"></td>
    <td width="25%" style="background:#2979FF;height:4px;"></td>
</tr></table></td></tr>
<tr><td style="padding:0;border-bottom:1px solid #eee;"><table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td width="25%" style="text-align:center;padding:20px 10px;border-right:1px solid #f0f0f0;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.NAVY};">{len(unique_sos)}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;">Sales Orders</p></td>
    <td width="25%" style="text-align:center;padding:20px 10px;border-right:1px solid #f0f0f0;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.GREEN};">{format_indian(len(result.rows))}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;">Line Items</p></td>
    <td width="25%" style="text-align:center;padding:20px 10px;border-right:1px solid #f0f0f0;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.ORANGE};">{format_indian(total_order)}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;">Order Qty</p></td>
    <td width="25%" style="text-align:center;padding:20px 10px;">
        <p style="margin:0;font-size:32px;font-weight:bold;color:{C.PURPLE};">{format_indian(total_tester)}</p>
        <p style="margin:5px 0 0;font-size:10px;color:#999;text-transform:uppercase;">Tester Qty</p></td>
</tr></table></td></tr>

<tr><td style="padding:14px 20px;font-weight:bold;font-size:14px;color:{C.NAVY};border-left:5px solid {C.NAVY};background:#E8EAF6;">📋 Sales Order Details</td></tr>
<tr><td style="padding:0;"><table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">
<tr>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">SO Number</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">Distributor</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">City</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">State</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">Location</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">Order</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">Tester</th>
    <th style="background:{C.NAVY};color:white;padding:10px 8px;font-size:11px;">Total</th>
</tr>'''

        for i, so_row in enumerate(unique_sos):
            si = so_groups.get(so_row.so_number, {'order':0,'tester':0})
            bg = '#f9f9f9' if i%2==1 else '#fff'
            html += f'''<tr style="background:{bg};">
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;font-weight:bold;">{so_row.so_number}</td>
    <td style="padding:9px 8px;text-align:left;font-size:12px;border-bottom:1px solid #eee;">{so_row.distributor or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{so_row.city or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{so_row.state or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{so_row.location_code or "—"}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(si['order'])}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(si['tester'])}</td>
    <td style="padding:9px 8px;text-align:center;font-size:12px;border-bottom:1px solid #eee;font-weight:bold;">{format_indian(si['order']+si['tester'])}</td>
</tr>'''

        html += f'''<tr style="background:#E8EAF6;font-weight:bold;">
    <td style="padding:10px 8px;text-align:center;font-size:12px;">TOTAL</td>
    <td colspan="4" style="padding:10px 8px;font-size:12px;">{len(unique_sos)} Sales Orders</td>
    <td style="padding:10px 8px;text-align:center;font-size:12px;">{format_indian(total_order)}</td>
    <td style="padding:10px 8px;text-align:center;font-size:12px;">{format_indian(total_tester)}</td>
    <td style="padding:10px 8px;text-align:center;font-size:12px;">{format_indian(total_qty)}</td>
</tr></table></td></tr>

<tr><td style="padding:14px 20px;font-weight:bold;font-size:14px;color:{C.GREEN};border-left:5px solid {C.GREEN};background:#E8F5E9;">📦 SKU Demand Summary</td></tr>
<tr><td style="padding:0;"><table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">
<tr>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">#</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">BC CODE</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">DESCRIPTION</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">CATEGORY</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">ORDER</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">TESTER</th>
    <th style="background:{C.GREEN};color:white;padding:10px 6px;font-size:11px;">TOTAL</th>
</tr>'''

        for rank, (item_no, info) in enumerate(sorted_skus, 1):
            total = info['order'] + info['tester']
            desc = info['desc'][:45] + '...' if len(info['desc']) > 45 else info['desc']
            bg = '#f1f8e9' if rank%2==0 else '#fff'
            html += f'''<tr style="background:{bg};">
    <td style="padding:8px 6px;text-align:center;font-size:12px;color:#999;border-bottom:1px solid #eee;">{rank}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;font-weight:bold;border-bottom:1px solid #eee;">{item_no}</td>
    <td style="padding:8px 6px;text-align:left;font-size:12px;border-bottom:1px solid #eee;">{desc or "—"}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{info["cat"] or "—"}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(info["order"])}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;border-bottom:1px solid #eee;">{format_indian(info["tester"])}</td>
    <td style="padding:8px 6px;text-align:center;font-size:12px;font-weight:bold;border-bottom:1px solid #eee;">{format_indian(total)}</td>
</tr>'''

        html += f'''<tr style="background:#E8F5E9;font-weight:bold;">
    <td style="padding:10px 6px;"></td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">GRAND TOTAL</td>
    <td style="padding:10px 6px;font-size:12px;">{len(sorted_skus)} unique SKUs</td>
    <td></td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">{format_indian(total_order)}</td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">{format_indian(total_tester)}</td>
    <td style="padding:10px 6px;text-align:center;font-size:12px;">{format_indian(total_qty)}</td>
</tr></table></td></tr>

<tr><td style="background:{C.NAVY};padding:30px;text-align:center;">
    <p style="margin:0 0 5px;font-size:16px;font-weight:bold;color:{C.GOLD};">⚡ GT MASS DUMP GENERATOR v2.4</p>
    <p style="margin:0 0 18px;font-size:11px;color:#7986CB;">Warehouse Automation Suite</p>
    <table style="margin:0 auto;max-width:400px;"><tr><td style="background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.15);padding:18px;border-radius:10px;text-align:center;">
        <p style="margin:0 0 3px;font-size:10px;color:#7986CB;text-transform:uppercase;letter-spacing:2px;">🚀 Engineered by</p>
        <p style="margin:0 0 5px;font-size:18px;font-weight:bold;color:white;">Abhishek Wagh</p>
        <p style="margin:0;font-size:11px;color:#9FA8DA;">Order Management and Automation</p>
    </td></tr></table>
    <p style="margin:18px 0 0;font-size:9px;color:#5C6BC0;">© 2026 RENEE Cosmetics Pvt. Ltd.</p>
</td></tr></table></td></tr></table></body></html>'''

        return html


class EmailSender:
    """Sends HTML email via SMTP."""

    @staticmethod
    def send_report(result: ProcessResult, elapsed_str: str = '') -> Tuple[bool, str]:
        """
        Build and send the email report.

        Returns:
            (success, error_message)
        """
        config = EMAIL_CONFIG

        if not config['EMAIL_SENDER'] or not config['DEFAULT_RECIPIENT']:
            return False, "Email not configured."

        try:
            html = EmailBuilder.build_html(result, elapsed_str)
            subject = EmailBuilder.build_subject(result)

            msg = EmailMessage()
            msg['From'] = config['EMAIL_SENDER']
            msg['To'] = config['DEFAULT_RECIPIENT']

            if config['CC_RECIPIENTS']:
                msg['Cc'] = ', '.join(config['CC_RECIPIENTS'])

            msg['Subject'] = subject
            msg.set_content("Please view in HTML-compatible client.")
            msg.add_alternative(html, subtype='html')

            server = smtplib.SMTP(config['SMTP_SERVER'], config['SMTP_PORT'])
            server.starttls()
            server.login(config['EMAIL_SENDER'], config['EMAIL_PASSWORD'])

            recipients = [config['DEFAULT_RECIPIENT']] + config['CC_RECIPIENTS']
            server.send_message(msg, to_addrs=recipients)
            server.quit()

            logger.info(f"Email sent to {config['DEFAULT_RECIPIENT']} + {len(config['CC_RECIPIENTS'])} CC")
            return True, ""

        except smtplib.SMTPAuthenticationError as e:
            return False, f"Auth failed: {e}"
        except smtplib.SMTPException as e:
            return False, f"SMTP error: {e}"
        except (ConnectionError, OSError) as e:
            return False, f"Network error: {e}"
        except (ValueError, KeyError) as e:
            return False, f"Config error: {e}"


# ═══════════════════════════════════════════════════════════════════════════════
#  PO TEMPLATE GENERATOR — blank template download
# ═══════════════════════════════════════════════════════════════════════════════

class TemplateGenerator:
    """Generates a blank GT-Mass PO template Excel in memory."""

    @staticmethod
    def generate() -> io.BytesIO:
        """
        Create the standard PO template and return as BytesIO.

        Returns:
            BytesIO with the template Excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'PO Template'

        # Fonts
        title_font = Font(name='Aptos Display', size=14, bold=True, color='1A237E')
        label_font = Font(name='Aptos Display', size=11, bold=True)
        value_font = Font(name='Aptos Display', size=11, color='0000CC')
        note_font = Font(name='Aptos Display', size=10, italic=True, color='FF6600')
        sample_font = Font(name='Aptos Display', size=11, color='888888', italic=True)
        hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)
        crit_font = Font(name='Aptos Display', size=11, bold=True, color='D32F2F')
        crit_hdr_font = Font(bold=True, color='FFFFFF', name='Aptos Display', size=11)

        # Fills
        hdr_fill = PatternFill('solid', fgColor='1A237E')
        meta_fill = PatternFill('solid', fgColor='E3F2FD')
        crit_fill = PatternFill('solid', fgColor='FFCDD2')
        crit_hdr_fill = PatternFill('solid', fgColor='D32F2F')

        # Row 1: Title
        ws.cell(row=1, column=1, value='Purchase Order GT-Mass (Template)').font = title_font

        # Meta rows
        for r, label, value in [
            (2, 'Distributor Name', '<Enter Distributor Name>'),
            (3, 'DB Code', '<DB Code>'),
            (5, 'City', '<City>'),
            (6, 'State', '<State>'),
        ]:
            ws.cell(row=r, column=1, value=label).font = label_font
            ws.cell(row=r, column=1).fill = meta_fill
            ws.cell(row=r, column=2, value=value).font = value_font

        # Right-side meta
        ws.cell(row=2, column=7, value='ASM').font = label_font
        ws.cell(row=2, column=7).fill = meta_fill
        ws.cell(row=2, column=9, value='<ASM Name>').font = value_font
        ws.cell(row=3, column=7, value='RSM').font = label_font
        ws.cell(row=3, column=7).fill = meta_fill
        ws.cell(row=3, column=9, value='<RSM Name>').font = value_font

        ws.cell(row=4, column=1, value='BDE Name').font = label_font
        ws.cell(row=4, column=1).fill = meta_fill
        ws.cell(row=4, column=2, value='<BDE Name>').font = value_font

        # Critical fields (red)
        ws.cell(row=4, column=7, value='PO Number').font = crit_font
        ws.cell(row=4, column=7).fill = crit_fill
        ws.cell(row=4, column=9, value='SO/GTM/0000').font = crit_font
        ws.cell(row=4, column=9).fill = crit_fill

        ws.cell(row=5, column=7, value='Date of PO').font = label_font
        ws.cell(row=5, column=7).fill = meta_fill
        ws.cell(row=5, column=9, value='DD.MM.YYYY').font = value_font

        ws.cell(row=6, column=7, value='Location').font = crit_font
        ws.cell(row=6, column=7).fill = crit_fill
        ws.cell(row=6, column=9, value='AHD').font = crit_font
        ws.cell(row=6, column=9).fill = crit_fill

        # Data header (row 7)
        data_headers = [
            'EAN', 'BC Code', 'Category', 'Article Description ',
            'Nail Paint Shade Number ', 'Product Classification',
            'HSN Code\n8 Digit', 'MRP', 'Retiler Margin',
            'Trade & Display Scheme', 'Ullage', 'QPS',
            'Qty In Case', 'Rate @ RLP', 'Amount @ RLP',
            'Order Qty', 'Order Amount', 'Tester Qty',
        ]
        critical_cols = {'EAN', 'BC Code', 'Order Qty', 'Tester Qty'}

        for ci, h in enumerate(data_headers, 1):
            cell = ws.cell(row=7, column=ci, value=h)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if h.strip() in critical_cols:
                cell.font = crit_hdr_font
                cell.fill = crit_hdr_fill
            else:
                cell.font = hdr_font
                cell.fill = hdr_fill

        # Sample row (row 8)
        sample = [
            8904473104307, 201238, 'Eye',
            'RENEE PURE BROWN KAJAL PEN WITH SHARPENER, 0.35GM',
            '-', 'Cosmetics', 33049990, 199, 1.2,
            '16.67% on RLP', '1.66 % on RLP', '4.81% on RLP',
            '', '', '', 72, '', 6,
        ]
        for ci, v in enumerate(sample, 1):
            ws.cell(row=8, column=ci, value=v).font = sample_font

        # Instructions
        ws.cell(row=10, column=1, value='⚠ INSTRUCTIONS:').font = Font(
            name='Aptos Display', size=11, bold=True, color='D32F2F'
        )
        for i, ins in enumerate([
            '1. Fill PO Number (Row 4, Col I) SO/GTM/####',
            '2. Fill Location (Row 6, Col I) AHD/BLR',
            '3. Fill Distributor, City, State',
            '4. Data from Row 8, delete sample',
            '5. BC Code numeric, Qty numeric',
            '6. RED = critical fields',
            '7. Save .xlsx → upload to generator',
        ]):
            ws.cell(row=11 + i, column=1, value=ins).font = note_font

        # Column widths
        for cl, w in {'A':16,'B':12,'C':12,'D':50,'E':12,'F':18,'G':14,'H':8,
                      'I':14,'J':20,'K':16,'L':14,'M':12,'N':12,'O':14,'P':12,'Q':14,'R':12}.items():
            ws.column_dimensions[cl].width = w

        ws.freeze_panes = 'A8'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output