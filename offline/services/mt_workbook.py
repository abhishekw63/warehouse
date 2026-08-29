"""
offline.services.mt_workbook
============================

Adapter that makes an MT (offline) run's downloadable workbook come out with
the SAME sheet structure as the Online B2B workbook — Headers (SO), Lines (SO),
Summary, Tracker, Validation, Rules & Exceptions, Warnings, Raw Data,
SKU Summary — so EVERY downloaded workbook is uniform.

Approach A (adapt → the ONLINE exporter): rather than duplicate any sheet
logic, this reshapes the frozen MT engine's parsed batch into the online
engine's :class:`ProcessingResult` (a list of :class:`SORow` + marketplace
metadata) and runs the SAME code path the online workbook uses — the online
``SOExporter`` + the two web-appended sheets (SKU Summary + Tracker). The frozen
engines are NEVER touched: this only READS ``POBatch`` / ``POFile`` / ``POLine``
fields and the engine's own ``build_offline_order_rows`` (for per-SO header
facts incl. PO/exp dates), then writes a fresh workbook.

MT is mapping-only (RL / MET / LS / SS-family carry no CP check), so every
resolved line's ``validation_status`` is 'OK' and the price-compare columns are
left blank — the true picture, not a silent drop. See the field map in
:func:`build_result`.
"""
from __future__ import annotations

from pathlib import Path

# ── SKU Summary + Tracker: reuse the SAME row-shape / formatting the online
#    Processor uses. These are standalone (module-level) versions of
#    ``Processor._sku_pivot`` / ``_append_sku_sheet`` / ``_append_tracker_sheet``
#    that read the SAME ``lines`` / ``headers`` dicts, so the two workbooks stay
#    byte-for-byte identical in structure. ────────────────────────────────────


def build_result(batch, channel, warehouse_display: str, warehouse_code: str,
                 marketplace_label: str, input_file_path: str = '',
                 warnings=None, testers=None):
    """Reshape a frozen MT ``POBatch`` into the online engine's
    :class:`ProcessingResult` so the online ``SOExporter`` can render it.

    One :class:`SORow` per resolved line of every cleanly-parsed PO that got an
    SO number — mirroring exactly what the frozen ``_write_lines_sheet`` emits
    (unresolved / SKIP lines are dropped from Lines by the engine and surfaced
    on Warnings instead, so we drop them here too and never silently lose them).

    Field map (MT POLine/POFile → SORow):
      * ``po_number``   = ``pf.so_number``  (matches Headers 'No.' — the D365 doc)
      * ``cust_no``     = ``channel.sell_to`` (the frozen Headers use sell_to,
                          NOT pf.cust_no — keep parity so D365 stays correct)
      * ``ship_to``     = ``pf.ship_to``    (e.g. LS '20044_N')
      * ``location`` / ``mapped_location`` = store name (raw) / del_location
      * ``item_no`` / ``ean`` / ``qty`` / ``description`` / ``gst_code`` / ``mrp``
      * ``vendor_mrp``  = the file's stated MRP (``line.mrp``)
      * ``amount``      = ``line.po_value`` (per-line value; sums to the PO total
                          ``input_po_value_total`` so Summary/Tracker match)
      * ``validation_status`` = 'OK' (mapping-only; no CP check)
      * price-compare fields (fob/ref_fob/calc/cost_price_ref/diffn) left blank —
        MT has no vendor CP to compare, so those columns are genuinely empty.
    """
    from online_po_processor.data.models import ProcessingResult, SORow

    rows: list = []
    for pf in batch.po_files:
        if pf.has_hard_errors or not pf.so_number:
            continue
        del_loc = pf.ship_to_entry.del_location if pf.ship_to_entry else ''
        store = str(pf.store_name or del_loc or '')
        for ln in pf.lines:
            if not ln.item_no or ln.status == 'SKIP':
                continue                       # dropped from Lines by the engine
            rows.append(SORow(
                po_number=str(pf.so_number),
                location=store,
                item_no=ln.item_no,
                qty=int(ln.quantity or 0),
                cust_no=str(channel.sell_to or ''),
                ship_to=str(pf.ship_to or ''),
                mapped=True,
                mapped_location=str(del_loc or ''),
                ean=str(ln.ean or ''),
                description=str(ln.items_master_desc or ln.sku_name or ''),
                vendor_mrp=_f(ln.mrp),
                mrp=_f(ln.items_master_mrp),
                gst_code=str(ln.gst_code or ''),
                amount=_f(ln.po_value),
                validation_status='OK',
            ))

    # ── Testers LAST — one SORow per tester line, AFTER every regular order,
    #    so the Headers (SO) / Lines (SO) sheets carry the tester SOs at the
    #    very end (clearly separated). forced_unit_price stamps the 0.54 into
    #    the Lines 'Unit Price' column; the External Doc (TESTER-<store>) is
    #    re-applied by _fix_external_doc after export. ──
    if testers is not None and getattr(testers, 'lines', None):
        for ln in testers.lines:
            rows.append(SORow(
                po_number=str(ln['po']),
                location=str(ln.get('location') or ''),
                item_no=str(ln['item_no']),
                qty=int(ln.get('qty') or 1),
                cust_no=str(channel.sell_to or ''),
                ship_to=str(ln.get('ship_to') or ''),
                mapped=True,
                mapped_location=str(ln.get('location') or ''),
                ean=str(ln.get('ean') or ''),
                description=str(ln.get('description') or ''),
                gst_code='',
                amount=_f(ln.get('unit_price')),
                forced_unit_price=_f(ln.get('unit_price')),
                validation_status='OK',
            ))

    result = ProcessingResult(
        rows=rows,
        warnings=[('', '', w) for w in (warnings or [])],
        marketplace=marketplace_label,       # e.g. 'Lifestyle' (channel display)
        input_file=Path(input_file_path).name if input_file_path else '',
        input_file_path=input_file_path,
        margin_pct=0.0,                      # no margin concept for mapping-only
        compare_basis='',                    # blanks the Validation price columns
        compare_label='',
        warehouse_code=warehouse_code or 'PICK',
        warehouse_display=warehouse_display or 'AHD',
        raw_df=_combined_raw_df(batch),
    )
    return result


def write_unified_workbook(batch, channel, warehouse_display, warehouse_code,
                           marketplace_label, output_path,
                           notes=None, warnings=None, testers=None):
    """Render the MT batch to a 9-sheet workbook at ``output_path`` (overwriting)
    using the online ``SOExporter`` + the SKU Summary + Tracker append helpers —
    the SAME path the Online B2B download uses. Returns the Path written.

    ``output_path`` is where the frozen engine already wrote its workbook; we
    overwrite in place so the existing download link is unchanged. The online
    exporter writes next to ``input_file_path``'s ``output/`` folder, so we point
    ``input_file_path`` at ``output_path`` itself to keep the file exactly there.
    """
    from online_po_processor.exporter.so_exporter import SOExporter

    out = Path(output_path)
    result = build_result(
        batch, channel, warehouse_display, warehouse_code, marketplace_label,
        input_file_path=str(out), warnings=warnings, testers=testers)

    # SOExporter derives its own filename inside ``<input_dir>/output/``. Point
    # it at a scratch input under the same directory so the produced file lands
    # in ``out.parent/output/``; then move it onto ``out`` (the download path).
    result.input_file_path = str(out.parent / out.name)
    written = SOExporter().export(result)
    if written is None:
        return None
    written = Path(written)
    # Normalise to the expected download path (overwrite the frozen 6-sheet file).
    if written.resolve() != out.resolve():
        try:
            if out.exists():
                out.unlink()
            written.replace(out)
        except OSError:
            out = written                    # fall back to whatever was written

    # Per-run 'SKU Summary' + 'Tracker' sheets (web post-process, additive — the
    # exporter's own sheets are untouched). Same shape as the online workbook.
    headers = _header_dicts(batch, channel, warehouse_display, str(out))
    lines = _line_dicts(batch, channel)
    try:
        _append_sku_sheet(str(out), lines)
    except Exception:  # noqa: BLE001 — never fail the export on a bonus sheet
        pass
    try:
        _append_tracker_sheet(str(out), headers, marketplace_label)
    except Exception:  # noqa: BLE001
        pass
    # The online Headers sheet hard-codes External Document No. = the doc 'No.'
    # (both from po_number). For MT the doc No. is the generated SO number, but
    # External Document No. must be the VENDOR PO (e.g. LS Order No 2052098) — so
    # re-stamp that column with the real PO, keyed by the SO number.
    try:
        _fix_external_doc(str(out), batch, testers)
    except Exception:  # noqa: BLE001
        pass
    return out


# ── helpers ─────────────────────────────────────────────────────────────────

def _fix_external_doc(out_path, batch, testers=None):
    """Re-stamp the Headers (SO) 'External Document No.' column with the VENDOR
    PO number (e.g. LS Order No ``2052098``) instead of the SO doc number. The
    frozen online Headers sheet writes External Document No. = the doc 'No.'
    (both from ``so_row.po_number``); for MT the doc No. is the generated SO
    number, so we correct the column here, keyed by the SO number in 'No.'.
    Tester SOs get ``TESTER-<store>``. Best-effort; a missing sheet/column is a
    no-op."""
    import openpyxl
    so2po = {str(pf.so_number): str(pf.po_no)
             for pf in batch.po_files
             if pf.so_number and getattr(pf, 'po_no', None)}
    if testers is not None:
        for h in getattr(testers, 'headers', []) or []:
            so2po[str(h['po'])] = str(h.get('external_doc') or f"TESTER-{h.get('store')}")
    if not so2po:
        return
    wb = openpyxl.load_workbook(out_path)
    if 'Headers (SO)' not in wb.sheetnames:
        return
    ws = wb['Headers (SO)']
    hdr = [str(c.value) for c in ws[1]]
    if 'No.' not in hdr or 'External Document No.' not in hdr:
        return
    c_no = hdr.index('No.') + 1
    c_ext = hdr.index('External Document No.') + 1
    for r in range(2, ws.max_row + 1):
        po = so2po.get(str(ws.cell(r, c_no).value or ''))
        if po:
            ws.cell(r, c_ext).value = po
    wb.save(out_path)


def _f(x):
    try:
        return round(float(x), 2) if x not in (None, '') else None
    except (TypeError, ValueError):
        return None


def _combined_raw_df(batch):
    """Concatenate every PO's source DataFrame for the Raw Data sheet (audit).
    None when no PO carried a raw_df."""
    import pandas as pd
    frames = [pf.raw_df for pf in batch.po_files
              if getattr(pf, 'raw_df', None) is not None
              and not getattr(pf.raw_df, 'empty', True)]
    if not frames:
        return None
    try:
        return pd.concat(frames, ignore_index=True)
    except Exception:  # noqa: BLE001 — mismatched columns → best-effort first
        return frames[0]


def _header_dicts(batch, channel, warehouse, output_file):
    """Per-SO header dicts (for the Tracker sheet) via the frozen engine's own
    ``build_offline_order_rows`` — so PO Date / Exp Date / location / value come
    from the SAME source the dashboard uses. Keyed 'po' = SO number (matches the
    Headers sheet). Reshaped to the keys the Tracker helper reads."""
    from . import mt_bridge
    eng = mt_bridge._engine()
    rows = eng.build_offline_order_rows(batch, channel, warehouse or '',
                                        output_file or '')
    return [{
        'po': str(r.get('po') or ''),
        'segment': 'Offline',
        'marketplace_label': r.get('marketplace_label') or channel.display_name,
        'location': r.get('location') or '',
        'po_date': r.get('po_date') or '',
        'exp_date': r.get('exp_date') or '',
        'qty': int(r.get('qty') or 0),
        'order_value': round(float(r.get('order_value') or 0), 2),
    } for r in rows]


def _line_dicts(batch, channel):
    """Per-line dicts (for the SKU Summary pivot) — the SAME shape the online
    ``Processor._lines`` produces (item_no / ean / description / qty / our_mrp /
    vendor_mrp / status / diff / po). Mapping-only, so status='OK' and there is
    no diff."""
    out: list = []
    for pf in batch.po_files:
        if pf.has_hard_errors or not pf.so_number:
            continue
        for ln in pf.lines:
            if not ln.item_no or ln.status == 'SKIP':
                continue
            out.append({
                'po': str(pf.so_number),
                'item_no': str(ln.item_no or ''),
                'ean': str(ln.ean or ''),
                'description': str(ln.items_master_desc or ln.sku_name or ''),
                'qty': int(ln.quantity or 0),
                'our_mrp': _f(ln.items_master_mrp),
                'vendor_mrp': _f(ln.mrp),
                'status': 'OK',
                'diff': None,
            })
    return out


def _sku_pivot(lines):
    """Per-SKU rollup grouped by (item_no, ean): qty per status, MRP comparison
    (+ varies flag), POs, worst diff. Mirror of
    ``online_b2b...Processor._sku_pivot`` so the two workbooks' SKU Summary
    sheets are identical."""
    agg: dict = {}
    for ln in lines:
        key = (ln.get('item_no') or '', ln.get('ean') or '')
        a = agg.get(key)
        if a is None:
            a = agg[key] = {'desc': ln.get('description') or '',
                            'our_mrp': ln.get('our_mrp'), 'vmrps': set(),
                            'tot': 0, 'ok': 0, 'mis': 0, 'nim': 0,
                            'pos': set(), 'diffs': []}
        q = int(ln.get('qty') or 0)
        a['tot'] += q
        st = ln.get('status') or 'OK'
        a['ok' if st == 'OK' else 'mis' if st == 'MISMATCH'
          else 'nim' if st == 'NOT_IN_MASTER' else 'ok'] += q
        a['pos'].add(ln.get('po'))
        if ln.get('vendor_mrp') is not None:
            a['vmrps'].add(round(float(ln['vendor_mrp']), 2))
        if ln.get('diff') is not None:
            a['diffs'].append(float(ln['diff']))
    rows = []
    for (item_no, ean), a in agg.items():
        rows.append([
            item_no, ean, a['desc'], a['our_mrp'],
            (max(a['vmrps']) if a['vmrps'] else None),
            'YES' if len(a['vmrps']) > 1 else '', a['tot'], a['ok'],
            a['mis'], a['nim'], len(a['pos']),
            (min(a['diffs']) if a['diffs'] else None)])
    rows.sort(key=lambda r: (-r[8], -r[9], -r[6]))   # mismatch, nim, tot qty
    return rows


def _append_sku_sheet(path, lines):
    """Append a per-run 'SKU Summary' sheet — identical layout/formatting to the
    online ``Processor._append_sku_sheet``."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    rows = _sku_pivot(lines)
    wb = load_workbook(path)
    if 'SKU Summary' in wb.sheetnames:
        del wb['SKU Summary']
    ws = wb.create_sheet('SKU Summary')
    hdr = ['Item No', 'EAN', 'Description', 'Our MRP', 'Their MRP',
           'MRP varies', 'Tot Qty', 'OK Qty', 'Mismatch Qty',
           'Not-in-Master Qty', '# POs', 'Worst Diff']
    ws.append(hdr)
    for r in rows:
        ws.append(r)
    navy = PatternFill('solid', fgColor='1A237E')
    hfont = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='E6E8EC')
    bd = Border(thin, thin, thin, thin)
    for cell in ws[1]:
        cell.font = hfont
        cell.fill = navy
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = bd
    widths = [11, 16, 46, 10, 10, 11, 9, 9, 13, 17, 8, 11]
    right_cols = {4, 5, 7, 8, 9, 10, 11, 12}
    center_cols = {6}
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = bd
            if cell.column in right_cols:
                cell.alignment = Alignment(horizontal='right')
            elif cell.column in center_cols:
                cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    wb.save(path)


def _tracker_date_val(v):
    """Coerce a date-like value to a real ``datetime.date`` so the Tracker cell
    is a GENUINE Excel date (groups by month in AutoFilter, sorts correctly, and
    survives a paste-into-the-WH-master), or ``None`` when it isn't a date."""
    import datetime as _dt
    if not v:
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v).strip()
    for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%b-%Y'):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_tracker_date(v) -> str:
    """Day-first ``dd-mm-YYYY`` string (fallback for un-coercible values)."""
    d = _tracker_date_val(v)
    if d is not None:
        return d.strftime('%d-%m-%Y')
    return '' if not v else str(v).strip()


def _append_tracker_sheet(path, headers, marketplace_label):
    """Append a per-PO 'Tracker' sheet (Segment · Market Place · PO · Location ·
    PO Date · Exp Date · PO Aging · Order Value · Order Qty · State · Zone ·
    Pincode), positioned 4th — identical to the online
    ``Processor._append_tracker_sheet`` so every workbook's Tracker is the same and
    online + offline rows paste into the org master with columns aligned. Segment
    'Offline' for MT."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    if not headers:
        return
    # Ship-to → State/Zone/Pincode map (SAME resolver the Online B2B tracker uses),
    # so MT rows paste into the org master tracker with columns lined up cell-for-cell
    # (the WH team pastes online + offline into ONE sheet). Best-effort — blanks on miss.
    try:
        from online_b2b.services.order_db import geo_for_location, location_geo_map
        geomap = location_geo_map()
    except Exception:  # noqa: BLE001
        geomap = {}
        def geo_for_location(_loc, _m=None):  # noqa: E306
            return {'pincode': '', 'state': '', 'zone': ''}
    wb = load_workbook(path)
    if 'Tracker' in wb.sheetnames:
        del wb['Tracker']
    ws = wb.create_sheet('Tracker')
    cols = ['Segment', 'Market Place', 'PO', 'Location', 'PO Date', 'Exp Date',
            'PO Aging For Exp', 'Order Value', 'Order Qty', 'State', 'Zone',
            'Pincode']
    ws.append(cols)
    for h in headers:
        pod_d = _tracker_date_val(h.get('po_date'))
        exd_d = _tracker_date_val(h.get('exp_date'))
        q = int(h.get('qty') or 0)
        v = round(float(h.get('order_value') or 0), 2)
        geo = geo_for_location(h.get('location'), geomap)   # State/Zone/Pincode from ship-to
        # Write REAL dates when coercible (so Excel groups them by month in the
        # WH team's filter) — fall back to the plain string only if un-parseable.
        ws.append([h.get('segment') or 'Offline',
                   h.get('marketplace_label') or marketplace_label or '',
                   str(h.get('po') or ''), h.get('location') or '',
                   pod_d if pod_d is not None else _fmt_tracker_date(h.get('po_date')),
                   exd_d if exd_d is not None else _fmt_tracker_date(h.get('exp_date')),
                   '', v, q,
                   geo.get('state') or '', geo.get('zone') or '',
                   geo.get('pincode') or ''])
        rr = ws.max_row
        if pod_d is not None:
            ws.cell(rr, 5).number_format = 'DD-MM-YYYY'
        if exd_d is not None:
            ws.cell(rr, 6).number_format = 'DD-MM-YYYY'
    navy = PatternFill('solid', fgColor='1A237E')
    hfont = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='E6E8EC')
    bd = Border(thin, thin, thin, thin)
    for cell in ws[1]:
        cell.font = hfont
        cell.fill = navy
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = bd
    widths = [13, 16, 18, 42, 13, 13, 16, 15, 11, 16, 11, 12]
    right_cols = {8, 9}
    center_cols = {5, 6, 7, 11, 12}
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = bd
            if cell.column in right_cols:
                cell.alignment = Alignment(horizontal='right')
            elif cell.column in center_cols:
                cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    idx = wb.sheetnames.index('Tracker')
    if len(wb.sheetnames) > 3 and idx != 3:
        wb.move_sheet('Tracker', offset=3 - idx)
    wb.save(path)
