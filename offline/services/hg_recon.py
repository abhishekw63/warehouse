"""
offline.services.hg_recon
=========================

**HG (Modern Trade) reconciliation & tester toolkit** — the reusable, parameterised
version of the ad-hoc scripts used on 2026-07-01 for the HG SO batches. Keep these
for every future HG dispatch. Nothing here is wired into the web request path; run
the functions from ``manage.py shell`` (they read/write Excel + the DB directly).

Contents
--------
* :func:`pdf_buyer_address`  — pull the BUYER (ship-to) address + pincode from a PO
  PDF, splitting the 2-column layout by x-position (supplier left / buyer right),
  so the store's real pincode isn't confused with the dispatch-warehouse one.
* :func:`pincode_map`        — ``{ship_to_code: postcode}`` for a party from the DB.
* :func:`cross_check`        — reconcile a generated SO workbook against its PO PDFs
  (Qty / SKU count / Value / ship-to pincode) and write a **"Cross Check"** sheet
  into the workbook. Returns counts.
* :func:`append_testers`     — append tester SOs to an EXISTING SO workbook, matched
  by (store, SKU) against a tester-requirement sheet (regulars untouched;
  ``SO/<ch>/TT/<counter>``, Ext Doc ``TESTERS``, qty 1 @ ``price``).
* :func:`mismatch_report`    — PO-wise "our ship-to vs PO-PDF" Excel (two sheets:
  Different Ship-to Address / Different Pincode) for corrective action.
* :func:`load_hg_master_dec25` — upsert an HG Master file (SKU→EAN, all sheets) into
  ``channel_sku_map`` (channel='HG'); MT reads this via
  ``mt_bridge.MTProcessor._apply_db_channel_master``.

Typical flow for a new HG batch::

    from offline.services import hg_recon as hr
    # 1) after the SO workbook is generated, append testers:
    hr.append_testers(WB, TESTER_REQ_XLSX, start_counter=NEXT_HG_COUNTER)
    # 2) verify every SO against its PO PDFs (writes the Cross Check sheet):
    hr.cross_check(WB, PDF_GLOB, title="HG <batch>")
    # 3) if pincodes/addresses differ, dump a PO-wise action list:
    hr.mismatch_report([("<batch>", WB, PDF_GLOB)], "C:/Users/.../Downloads/HG mismatches.xlsx")

**Reminder:** D365 is the ship-to source of truth, NOT the PO PDF — a PDF that
differs from a code whose D365 address matches ours is a false positive (see the
Rajaji Nagar case). Use :func:`mismatch_report` to *spot* differences, then confirm
against the D365 Ship-to Address List before changing any mapping.
"""
from __future__ import annotations

import glob
import os
import re

TESTER_PRICE = 0.54                         # HG tester sample unit price
# Dispatch-warehouse / RENEE pincodes that appear on the SUPPLIER side of the PO
# PDF — never the buyer. Extend if a new dispatch warehouse is used.
SUPPLIER_PINCODES = {'560067', '380009', '380015'}


def _cid(v) -> str:
    """Coerce an id like the engine — str, stripped, no trailing '.0'."""
    s = str(v if v is not None else '').strip()
    return s[:-2] if s.endswith('.0') else s


def _norm(s) -> str:
    return re.sub(r'\s+', ' ', str(s).strip().upper())


def _re1(text: str, pat: str) -> str:
    """First capture group of ``pat`` in ``text``, or ''."""
    m = re.search(pat, text)
    return m.group(1) if m else ''


# ── PO PDF buyer address / pincode ───────────────────────────────────────────
def pdf_buyer_address(path) -> tuple[str, str]:
    """Return ``(address, pincode)`` of the BUYER (ship-to) from a PO PDF.

    The PO PDF is a 2-column layout (supplier left, buyer right) that flattens
    into interleaved text, so we split by x-position and keep the right column.
    Pincode = the last 6-digit token in the buyer block that isn't a known
    supplier/warehouse pincode."""
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        pg = pdf.pages[0]
        mid = pg.width / 2
        words = [w for w in pg.extract_words() if w['x0'] >= mid]
    lines: dict = {}
    for w in words:
        lines.setdefault(round(w['top'] / 3), []).append((w['x0'], w['text']))
    txt = '\n'.join(' '.join(t for _, t in sorted(lines[k])) for k in sorted(lines))
    m = re.search(r'(Buyer Details.*?)(GSTIN|Sl\.?N)', txt, re.S)
    blk = m.group(1) if m else txt
    # pincode: from 'Pincode' lines / colon-prefixed 6-digits, excluding supplier
    cand = []
    for ln in blk.splitlines():
        if 'Pincode' in ln or ':' in ln:
            cand += re.findall(r':\s*(\d{6})(?!\d)', ln)
    if not cand:
        cand = re.findall(r'(?<!\d)(\d{6})(?!\d)', blk)
    pin = next((c for c in cand if c not in SUPPLIER_PINCODES),
               (cand[-1] if cand else ''))
    addr = re.sub(r'Buyer Details[^\n]*\n?', '', blk)
    addr = re.sub(r'Pincode\s*:?\s*\d*', '', addr).replace('Address :', '')
    addr = re.sub(r'\s+', ' ', addr).strip()
    return addr[:200], pin


def _po_from_pdf_name(path) -> str:
    m = re.search(r'_(\d{6,})_', os.path.basename(path))
    return m.group(1) if m else ''


# ── DB helpers ───────────────────────────────────────────────────────────────
def pincode_map(party: str = 'HG') -> dict:
    """``{ship_to_code: postcode}`` for ``party`` from ``ship_to_mapping``."""
    from online_b2b.services.order_db import _conn
    out: dict = {}
    with _conn() as (cur, d):
        ph = d['ph']
        cur.execute(f"SELECT ship_to, postcode FROM ship_to_mapping "
                    f"WHERE party={ph}", (party,))
        for st, pc in cur.fetchall():
            out[str(st)] = str(pc or '').split('.')[0]
    return out


# ── Reconciliation: SO workbook vs PO PDFs → "Cross Check" sheet ─────────────
def cross_check(workbook: str, pdf_glob: str, title: str = '',
                party: str = 'HG') -> dict:
    """Reconcile each regular SO in ``workbook`` against its PO PDF (Qty, SKU
    count, Value inc-GST, ship-to pincode) and (re)write a **"Cross Check"**
    sheet into the workbook. Returns ``{ok, review, total}``."""
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    def _num(s):
        try:
            return round(float(str(s).replace(',', '')), 2)
        except (TypeError, ValueError):
            return None

    pins = pincode_map(party)
    pdfs = glob.glob(pdf_glob)
    pdf = {}
    for f in pdfs:
        addr, pin = pdf_buyer_address(f)
        import pdfplumber
        with pdfplumber.open(f) as _p:
            tall = '\n'.join((pg.extract_text() or '') for pg in _p.pages)
        po = _po_from_pdf_name(f) or _re1(tall, r'PO No\s*:\s*(\d+)')
        pdf[po] = {'qty': _num(_re1(tall, r'Total Qty\s*:\s*(\d+)')),
                   'sku': _num(_re1(tall, r'SKU Count\s*:\s*(\d+)')),
                   'total': _num(_re1(tall, r'PO Total Value\s*:\s*([\d.,]+)')),
                   'pin': pin, 'addr': addr}
    rec = pd.read_excel(workbook, sheet_name='Reconciliation', dtype=str)
    so2po = {str(r['SO No']).strip(): str(r['PO No']).strip() for _, r in rec.iterrows()}
    summ = pd.read_excel(workbook, sheet_name='Summary', dtype=str)
    rows = []
    for _, s in summ.iterrows():
        so = str(s['PO']).strip()
        if not so.startswith('SO/'):
            continue
        po = so2po.get(so, '')
        p = pdf.get(po, {})
        sq = _num(s['Total Qty']); ssku = _num(s['Items'])
        sval = _num(s['Total Amount (Inc GST)']); code = str(s['Ship-to']).strip()
        mpin = pins.get(code, '')
        qok = 'Y' if p.get('qty') == sq else 'N'
        kok = 'Y' if p.get('sku') == ssku else 'N'
        pok = ('Y' if (p.get('pin') and mpin and p['pin'] == mpin)
               else ('N' if (p.get('pin') and mpin) else '?'))
        vok = 'Y' if (p.get('total') is not None and sval is not None
                      and abs(p['total'] - sval) <= 2) else 'N'
        iss = [x for x, ok in [('QTY', qok), ('SKU', kok), ('PIN', pok),
                               ('VALUE', vok)] if ok == 'N']
        verdict = 'OK' if not iss else 'CHECK: ' + ','.join(iss)
        rows.append([po, str(s['Location (Mapped)']).strip(), so, code,
                     p.get('qty'), sq, qok,
                     int(p['sku']) if p.get('sku') else None,
                     int(ssku) if ssku else None, kok,
                     p.get('pin'), mpin, pok, p.get('total'), sval, vok, verdict])
    wb = openpyxl.load_workbook(workbook)
    if 'Cross Check' in wb.sheetnames:
        del wb['Cross Check']
    ws = wb.create_sheet('Cross Check')
    navy = PatternFill('solid', fgColor='1F2A5A')
    okf = PatternFill('solid', fgColor='E6F4EA')
    badf = PatternFill('solid', fgColor='FCE8E6')
    white = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)
    nok = sum(1 for r in rows if r[16] == 'OK')
    ws.append([title or 'Cross-Check: Generated SO vs Source PO PDF'])
    ws.merge_cells('A1:Q1'); ws['A1'].font = Font(bold=True, size=13)
    ws.append([f'{len(rows)} orders checked | {nok} OK | {len(rows) - nok} to '
               f'review | Qty/SKU/Value & ship-to pincode vs PO PDF'])
    ws.merge_cells('A2:Q2'); ws['A2'].font = Font(italic=True, color='555555')
    ws.append([])
    cols = ['PO No', 'Store', 'SO No', 'Ship-to Code', 'Qty (PO)', 'Qty (SO)',
            'Qty OK', 'SKUs (PO)', 'SKUs (SO)', 'SKU OK', 'Pincode (PO)',
            'Pincode (Ours)', 'Pin OK', 'Value inc-GST (PO)',
            'Value inc-GST (SO)', 'Val OK', 'Verdict']
    ws.append(cols)
    for c in ws[ws.max_row]:
        c.fill = navy; c.font = white; c.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append(r); rr = ws.max_row
        cc = ws.cell(rr, 17); cc.fill = okf if r[16] == 'OK' else badf; cc.font = bold
        for ci in (7, 10, 13, 16):
            x = ws.cell(rr, ci); x.alignment = Alignment(horizontal='center')
            if x.value == 'N':
                x.fill = badf; x.font = bold
            elif x.value == 'Y':
                x.fill = okf
    ws.freeze_panes = 'A5'
    for i, w in enumerate([10, 26, 17, 13, 8, 8, 7, 9, 9, 7, 12, 13, 7, 16, 16,
                           7, 22], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(workbook)
    return {'ok': nok, 'review': len(rows) - nok, 'total': len(rows)}


# ── Append tester SOs to an existing workbook (regulars untouched) ───────────
def append_testers(workbook: str, tester_req: str, start_counter: int,
                   channel_code: str = 'HG', price: float = TESTER_PRICE) -> dict:
    """Append tester SOs to ``workbook`` for the (store, SKU) combos present in
    ``tester_req`` (any sheet with LOCATION_NAME + SKU_CODE [+ Tester Req]).
    One tester SO per regular SO with eligible lines — cloned header with
    ``No.=SO/<ch>/TT/<counter>``, Ext Doc ``TESTERS``, lines qty 1 @ ``price``.
    Returns ``{tester_sos, tester_lines, next_counter}``. Advancing/​persisting
    the shared ``mt_select_seq`` counter is the CALLER's responsibility."""
    import openpyxl
    import pandas as pd
    from openpyxl.styles import PatternFill

    elig = set()
    for sh in pd.ExcelFile(tester_req).sheet_names:
        raw = pd.read_excel(tester_req, sheet_name=sh, header=None, dtype=str)
        hdr = None
        for i in range(min(6, len(raw))):
            cells = [str(c).strip().upper() for c in raw.iloc[i] if c is not None]
            if 'LOCATION_NAME' in cells and 'SKU_CODE' in cells:
                hdr = i; break
        if hdr is None:
            continue
        df = pd.read_excel(tester_req, sheet_name=sh, header=hdr, dtype=str)
        low = {c.strip().upper(): c for c in df.columns}
        ln, sk = low.get('LOCATION_NAME'), low.get('SKU_CODE')
        tr = next((low[k] for k in low if k.startswith('TESTER')), None)
        if not ln or not sk:
            continue
        for _, r in df.iterrows():
            loc, sku = _norm(r[ln]), _cid(r[sk])
            flag = _cid(r[tr]) if tr else '1'
            if (loc and sku and loc != 'NAN' and sku != 'NAN'
                    and flag.lower() not in ('', '0', 'nan', 'none')):
                elig.add((loc, sku))

    rec = pd.read_excel(workbook, sheet_name='Reconciliation', dtype=str)
    summ = pd.read_excel(workbook, sheet_name='Summary', dtype=str)
    so_store = {str(r['PO']).strip(): _norm(r['Location (Mapped)'])
                for _, r in summ.iterrows() if str(r['PO']).startswith('SO/')}
    per_so: dict = {}
    for _, r in rec.iterrows():
        so = str(r['SO No']).strip(); sku = _cid(r['SKU Code'])
        item = _cid(r['Item No (Resolved)']); store = so_store.get(so)
        if store and item and item.lower() != 'nan' and (store, sku) in elig:
            per_so.setdefault(so, []).append(item)

    wb = openpyxl.load_workbook(workbook)
    wh, wl, ws = wb['Headers (SO)'], wb['Lines (SO)'], wb['Summary']
    regh = {}
    for row in range(4, wh.max_row + 1):
        so = _cid(wh.cell(row, 2).value)
        if so:
            regh[so] = [wh.cell(row, c).value for c in range(1, 19)]
    tint = PatternFill('solid', fgColor='DDEBF7')
    counter = start_counter; done = {}
    price_str = f'{price}'
    for so in sorted(per_so):
        if so not in regh:
            continue
        tso = f'SO/{channel_code}/TT/{counter:06d}'; counter += 1
        hdr = list(regh[so]); hdr[1] = tso; hdr[9] = 'TESTERS'
        wh.append(hdr)
        for c in wh[wh.max_row]:
            c.fill = tint
        lno = 10000
        for item in per_so[so]:
            wl.append(['Order', tso, lno, 'Item', item, 'PICK', 1, price_str])
            for c in wl[wl.max_row]:
                c.fill = tint
            lno += 10000
        done[so] = (tso, len(per_so[so]))
    hdrs = [c.value for c in ws[1]]; ci = {n: i + 1 for i, n in enumerate(hdrs)}
    totq = 0
    if 'Tester SO No' in ci:
        for r in range(2, ws.max_row + 1):
            so = _cid(ws.cell(r, ci['PO']).value)
            if so in done:
                ws.cell(r, ci['Tester SO No']).value = done[so][0]
                ws.cell(r, ci['Tester Qty']).value = done[so][1]
                ws.cell(r, ci['Tester Value']).value = round(done[so][1] * price, 2)
                totq += done[so][1]
            if ws.cell(r, ci['PO']).value == 'TOTAL':
                ws.cell(r, ci['Tester Qty']).value = totq
                ws.cell(r, ci['Tester Value']).value = round(totq * price, 2)
    wb.save(workbook)
    return {'tester_sos': len(done), 'tester_lines': totq, 'next_counter': counter}


# ── PO-wise "our ship-to vs PO PDF" mismatch report ─────────────────────────
def mismatch_report(batches: list, out_path: str, party: str = 'HG',
                    addr_diff_stores=()) -> dict:
    """Write a PO-wise Excel (sheets: *Different Ship-to Address* /
    *Different Pincode*) of ship-to pincode mismatches — our DB vs the PO PDF —
    for corrective action. ``batches`` = ``[(name, workbook_path, pdf_glob), …]``.
    ``addr_diff_stores`` = store names known to be a full-address (not just
    pincode) mismatch. REMEMBER: confirm against D365 before changing a mapping —
    the PDF can be stale (see Rajaji Nagar)."""
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    from online_b2b.services.order_db import _conn

    pins = pincode_map(party)
    addr_diff = {_norm(s) for s in addr_diff_stores}
    rows = []
    with _conn() as (cur, d):
        ph = d['ph']
        for bname, wb_path, pg in batches:
            rec = pd.read_excel(wb_path, sheet_name='Reconciliation', dtype=str)
            so2po = {str(r['SO No']).strip(): str(r['PO No']).strip()
                     for _, r in rec.iterrows()}
            summ = pd.read_excel(wb_path, sheet_name='Summary', dtype=str)
            pdfs = glob.glob(pg)
            for _, s in summ.iterrows():
                so = str(s['PO']).strip()
                if not so.startswith('SO/'):
                    continue
                store = str(s['Location (Mapped)']).strip()
                code = str(s['Ship-to']).strip()
                po = so2po.get(so, '')
                opin = pins.get(code, '')
                pf = next((x for x in pdfs if po and po in os.path.basename(x)), None)
                if not pf:
                    continue
                paddr, ppin = pdf_buyer_address(pf)
                if ppin and opin and ppin != opin:
                    cur.execute(f"SELECT address,address2,city FROM ship_to_mapping "
                                f"WHERE party={ph} AND del_location={ph}",
                                (party, store))
                    r = cur.fetchone()
                    oaddr = ', '.join(x for x in (r or ('', '', ''))
                                      if x and str(x).strip())[:200] if r else ''
                    cat = ('Different Ship-to Address' if _norm(store) in addr_diff
                           else 'Different Pincode')
                    rows.append([bname, po, so, store, code, opin, oaddr, ppin,
                                 paddr, cat])
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    navy = PatternFill('solid', fgColor='1F2A5A'); white = Font(bold=True, color='FFFFFF')
    cols = ['Batch', 'PO No', 'SO No', 'Store / Location', 'Our Ship-to Code',
            'Our Pincode', 'Our Address', 'PDF Pincode', 'PDF Address (delivery)',
            'Category']
    for cat in ['Different Ship-to Address', 'Different Pincode']:
        ws = wb.create_sheet(cat[:31]); ws.append(cols)
        for c in ws[1]:
            c.fill = navy; c.font = white; c.alignment = Alignment(horizontal='center')
        for r in sorted([r for r in rows if r[9] == cat], key=lambda r: (r[3], r[0], r[1])):
            ws.append(r)
        for i, w in enumerate([11, 10, 17, 26, 15, 10, 46, 10, 46, 22], 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'
    wb.save(out_path)
    return {'rows': len(rows),
            'addr_diff': sum(1 for r in rows if r[9] == 'Different Ship-to Address'),
            'pincode_diff': sum(1 for r in rows if r[9] == 'Different Pincode'),
            'out': out_path}


# ── Load an HG Master (SKU→EAN) file into channel_sku_map ────────────────────
def load_hg_master_dec25(path: str, channel: str = 'HG') -> dict:
    """Upsert SKU→EAN from an HG Master workbook (all sheets with sku_code + ENN
    code) into ``channel_sku_map`` (source='dec25'), resolving item_no via
    ``item_master``. MT then resolves these via
    ``mt_bridge.MTProcessor._apply_db_channel_master``. Returns counts."""
    import datetime as _dt

    import pandas as pd

    from online_b2b.services.order_db import _conn

    def _n(s):
        return str(s).strip().lower().replace(' ', '')

    m = {}
    for sh in pd.ExcelFile(path).sheet_names:
        raw = pd.read_excel(path, sheet_name=sh, header=None, dtype=str)
        hdr = None
        for i in range(min(6, len(raw))):
            cells = [_n(c) for c in raw.iloc[i] if c is not None and str(c).strip()]
            if 'sku_code' in cells and 'enncode' in cells:
                hdr = i; break
        if hdr is None:
            continue
        df = pd.read_excel(path, sheet_name=sh, header=hdr, dtype=str)
        low = {_n(c): c for c in df.columns}
        sc, en = low.get('sku_code'), low.get('enncode')
        if not sc or not en:
            continue
        for _, r in df.iterrows():
            k, e = _cid(r[sc]), _cid(r[en])
            if k and e and e.lower() != 'nan' and k not in m:
                m[k] = e
    now = _dt.datetime.now(); ins = upd = 0
    with _conn() as (cur, d):
        ph = d['ph']
        cur.execute('SELECT ean, item_no FROM item_master')
        e2i = {str(a): str(b) for a, b in cur.fetchall()}
        # Preload existing ids for this channel in ONE query, then bucket + two
        # executemany (was a SELECT + UPDATE/INSERT per SKU). Same result.
        cur.execute(f"SELECT sku_code, id FROM channel_sku_map WHERE channel={ph}",
                    (channel,))
        existing = {str(s): i for s, i in cur.fetchall()}
        upd_rows, ins_rows = [], []
        for sku, ean in m.items():
            item = e2i.get(ean, '')
            rid = existing.get(sku)
            if rid is not None:
                upd_rows.append((ean, item, now, rid))
            else:
                ins_rows.append((channel, sku, ean, item, now))
        if upd_rows:
            cur.executemany(f"UPDATE channel_sku_map SET ean={ph}, item_no={ph}, "
                            f"source='dec25', updated_at={ph} WHERE id={ph}", upd_rows)
        if ins_rows:
            cur.executemany(f"INSERT INTO channel_sku_map (channel,sku_code,ean,"
                            f"item_no,source,updated_at) VALUES ({ph},{ph},{ph},"
                            f"{ph},'dec25',{ph})", ins_rows)
        ins, upd = len(ins_rows), len(upd_rows)
        cur.connection.commit()
    return {'parsed': len(m), 'inserted': ins, 'updated': upd}
