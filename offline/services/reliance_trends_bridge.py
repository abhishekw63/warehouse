"""
offline.services.reliance_trends_bridge
=======================================

**Reliance Trends → ``renee_orders`` recorder** (Excel *BAP* PO format), mirroring
:mod:`offline.services.gt_mass_bridge`. Reliance Trends (D365 customer **20418**)
sends a Reliance-Retail SAP export ("Renee PO BAP.xlsx"), NOT the PDF the existing
'Reliance' channel parses — so this is a separate, additive channel. The frozen
engine is never touched.

Value basis (proven from the file): ``Net Value`` = ex-GST line value; ``Total CP``
= ``Net Value`` × 1.18 = inc-GST; so ``order_value`` (per PO) = Σ ``Total CP``
(inc-GST, consistent with every other channel) and line ``unit_price`` =
``Net Value`` ÷ ``PO Qty`` (ex-GST unit).

Ship-to: the Site code **``S0HZ`` is NOT unique** — it is on BOTH the Bangalore
(20418_1) and Bhiwandi (20418_2) rows. The *BAP* replenishment PO is the **Bhiwandi**
DC, so we map it to **``20418_2`` (Bhiwandi)**. How a Bangalore PO would be told
apart is unknown for now (the file has no per-line city) — flagged, not guessed.

STAGE 1 (this module, now): ``parse()`` — read + normalize + resolve items/value/
ship-to, read-only, for a preview. Recording (runs/order_headers/order_lines) is the
next stage, to be wired like ``gt_mass_bridge`` once the preview is confirmed.
"""
from __future__ import annotations

import re

from django.db import connections

MARKETPLACE = 'Reliance Trends'
SEGMENT = 'Offline'
CUST_NO = '20418'
GST_FALLBACK = 0.18

# Site → (ship_to, city). S0HZ is ambiguous (see module docstring); the BAP replen
# PO is Bhiwandi. Extend when Bangalore disambiguation is known.
SHIP_TO_BY_SITE = {
    'S0HZ': ('20418_2', 'Bhiwandi'),        # BAP = Bhiwandi DC
}
DEFAULT_SHIP_TO = ('20418_2', 'Bhiwandi')

# canonical field → accepted header spellings (lower-cased)
_HDR_ALIASES = {
    'po': ('purchasing document', 'po no', 'po number', 'purchase order no'),
    'ean': ('ean', 'gtin', 'barcode'),
    'article': ('article',),
    'short_text': ('short text', 'itemdescription', 'description'),
    'qty': ('po qty', 'scheduled quantity', 'order qty', 'quantity'),
    'net_value': ('net value', 'total cp ex gst'),
    'total_cp': ('total cp', 'total amount', 'gross value'),
    'tax_pct': ('tax %', 'tax percent', 'gst %'),
    'site': ('site',),
    'po_date': ('purchase order date', 'po date'),
    'exp_date': ('po expiry date', 'expiry date', 'stat.-rel. del. date'),
}


def _norm(s) -> str:
    return re.sub(r'\s+', ' ', str(s or '').strip()).lower()


def _map_headers(header_row) -> dict:
    idx = {}
    for j, cell in enumerate(header_row):
        n = _norm(cell)
        for canon, aliases in _HDR_ALIASES.items():
            if canon not in idx and n in aliases:
                idx[canon] = j
                break
    return idx


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_date(v):
    """Excel date / 'YYYY-MM-DD ...' string → date, else None. Day-first for
    ambiguous DD-MM-YYYY (Indian), never month-first."""
    import datetime as _dt
    if v is None or v == '':
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v).strip()[:10]
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y'):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _ean_map(eans) -> dict:
    """{ean: (item_no, description, gst_code)} from item_master for the file's EANs."""
    out = {}
    eans = sorted({str(e).strip() for e in eans if str(e or '').strip()})
    if not eans:
        return out
    with connections['orders'].cursor() as cur:
        fmt = ','.join(['%s'] * len(eans))
        cur.execute(f"SELECT ean, item_no, description, gst_code "
                    f"FROM item_master WHERE ean IN ({fmt})", tuple(eans))
        for ean, item_no, desc, gst in cur.fetchall():
            out[str(ean).strip()] = (str(item_no), str(desc or ''), str(gst or ''))
    return out


def _ship_to(site: str):
    return SHIP_TO_BY_SITE.get(str(site or '').strip().upper(), DEFAULT_SHIP_TO)


def parse(path) -> dict:
    """Read a Reliance Trends BAP Excel → normalized per-PO structure. Read-only.
    Returns ``{ok, error, pos, warnings, totals}`` where each PO carries its
    ship-to, lines (item/qty/value) and inc-GST value. Never raises."""
    import openpyxl
    out = {'ok': False, 'error': '', 'pos': {}, 'warnings': [],
           'totals': {'pos': 0, 'lines': 0, 'qty': 0, 'value': 0.0,
                      'unresolved': 0}}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            out['error'] = 'Empty sheet.'
            return out
        cmap = _map_headers(header)
        need = ('po', 'ean', 'qty')
        missing = [k for k in need if k not in cmap]
        if missing:
            out['error'] = ('Missing column(s): ' + ', '.join(missing)
                            + '. Is this a Reliance Trends BAP export?')
            return out

        def col(r, key):
            j = cmap.get(key)
            return r[j] if (j is not None and j < len(r)) else None

        data = [r for r in rows if r and col(r, 'po') not in (None, '')]
        emap = _ean_map([col(r, 'ean') for r in data])

        pos = out['pos']
        for r in data:
            po = str(col(r, 'po') or '').strip()
            if not po:
                continue
            ean = str(col(r, 'ean') or '').strip()
            qty = int(_num(col(r, 'qty')))
            net = _num(col(r, 'net_value'))
            total_inc = _num(col(r, 'total_cp'))
            if total_inc <= 0 and net > 0:
                total_inc = round(net * (1 + GST_FALLBACK), 2)
            gst_pct = _num(col(r, 'tax_pct')) or (GST_FALLBACK * 100)
            site = str(col(r, 'site') or '').strip()
            ship_to, city = _ship_to(site)
            item = emap.get(ean)
            item_no = item[0] if item else ''
            desc = (item[1] if item else '') or str(col(r, 'short_text') or '')
            gst_code = (item[2] if item else '') or str(int(gst_pct))
            resolved = bool(item_no)

            p = pos.get(po)
            if p is None:
                p = pos[po] = {'po': po, 'cust_no': CUST_NO, 'ship_to': ship_to,
                               'city': city, 'site': site, 'lines': [],
                               'value': 0.0, 'qty': 0, 'unresolved': 0,
                               'po_date': _to_date(col(r, 'po_date')),
                               'exp_date': _to_date(col(r, 'exp_date'))}
            p['lines'].append({
                'ean': ean, 'item_no': item_no, 'description': desc, 'qty': qty,
                'net_value': round(net, 2), 'total_inc': round(total_inc, 2),
                'unit_price': round(net / qty, 2) if qty else 0.0,
                'gst_code': gst_code, 'resolved': resolved})
            p['value'] += total_inc
            p['qty'] += qty
            if not resolved:
                p['unresolved'] += 1
                out['warnings'].append(f"PO {po}: EAN {ean} not in item master "
                                       f"(qty {qty}) — will need manual resolution.")

        t = out['totals']
        t['pos'] = len(pos)
        for p in pos.values():
            p['value'] = round(p['value'], 2)
            t['lines'] += len(p['lines'])
            t['qty'] += p['qty']
            t['value'] += p['value']
            t['unresolved'] += p['unresolved']
        t['value'] = round(t['value'], 2)
        wb.close()
        out['ok'] = True
    except Exception as e:  # noqa: BLE001
        out['error'] = f'{type(e).__name__}: {e}'
    return out


WH_CODE = {'AHD': 'PICK', 'BLR': 'DS_BL_OFF1'}
SO_CHANNEL = 'RT'          # SO numbers → SO/RT/{MM}/{counter:06d}


def _assign_so(po_keys, burn=True) -> dict:
    """Generate ``SO/RT/{MM}/{counter:06d}`` numbers (one per PO) from the SAME
    ``mt_select_seq.json`` counter the MT channels use — so RT SOs never collide
    with MT/tester SOs. ``burn=False`` snapshots (assigns without persisting) for
    a review-time preview. Returns ``{reliance_po: so_number}`` (insertion order)."""
    import datetime as _dt
    from . import mt_bridge
    eng = mt_bridge._engine()
    state = eng.load_seq_state()
    today_iso = _dt.date.today().isoformat()
    mm = _dt.date.today().strftime('%m')
    ch = state.get(SO_CHANNEL, {})
    if ch.get('date') != today_iso or 'next_counter' not in ch:
        ch = {'date': today_iso,
              'next_counter': int(_dt.date.today().strftime('%d%m%y'))}
    counter = int(ch['next_counter'])
    so_map = {}
    for po in po_keys:
        so_map[po] = f"SO/{SO_CHANNEL}/{mm}/{counter:06d}"
        counter += 1
    if burn and so_map:
        ch['next_counter'] = counter
        ch['date'] = today_iso
        state[SO_CHANNEL] = ch
        eng.save_seq_state(state)
    return so_map


def _restamp_external_doc(out_path, so_to_po: dict):
    """Overwrite Headers (SO) 'External Document No.' with the VENDOR PO (the
    Reliance Purchasing document), keyed by the generated SO in 'No.' — the online
    exporter writes External Doc = 'No.' by default. Best-effort."""
    import openpyxl
    if not so_to_po:
        return
    wb = openpyxl.load_workbook(out_path)
    if 'Headers (SO)' not in wb.sheetnames:
        return
    ws = wb['Headers (SO)']
    hdr = [str(c.value) for c in ws[1]]
    if 'No.' not in hdr or 'External Document No.' not in hdr:
        return
    c_no = hdr.index('No.') + 1
    c_ext = hdr.index('External Document No.') + 1
    for r in range(2, ws.max_row + 1):
        po = so_to_po.get(str(ws.cell(r, c_no).value or ''))
        if po:
            ws.cell(r, c_ext).value = po
    wb.save(out_path)


def build_workbook(path, warehouse='AHD', out_path=None, so_map=None,
                   burn_counter=False):
    """Render the unified **D365 SO workbook** (Headers (SO) · Lines (SO) · Summary ·
    Validation · Rules & Exceptions · Warnings · Raw Data · SKU Summary · Tracker) via
    the SAME online ``SOExporter`` + MT append helpers every other channel uses — so
    it matches [[unified-workbook-structure]]. Read-only (no DB write).

    Headers 'No.' = a generated **SO/RT/{MM}/{counter}** number (like MT); 'External
    Document No.' = the Reliance PO. Pass ``so_map`` (``{reliance_po: so}``) to reuse
    the SO numbers a confirm already burned; else they are assigned here
    (``burn_counter`` decides whether the counter is persisted).

    **Unit Price is left BLANK** (D365 auto-prices from the customer's price
    agreement) — the MT convention ([[mt-so-no-unit-price]]); the per-line inc-GST
    value still drives Summary/Tracker. cust 20418, ship-to per PO. Returns
    ``(Path, None)`` or ``(None, error)``."""
    import datetime as _dt
    from pathlib import Path
    from django.conf import settings
    from online_po_processor.data.models import ProcessingResult, SORow
    from online_po_processor.exporter.so_exporter import SOExporter
    from . import mt_workbook
    try:
        import pandas as pd
    except Exception:  # noqa: BLE001
        pd = None

    parsed = parse(path)
    if not parsed['ok']:
        return None, parsed['error']
    if not parsed['pos']:
        return None, 'No POs found in the file.'

    if so_map is None:
        so_map = _assign_so(list(parsed['pos'].keys()), burn=burn_counter)
    so_to_po = {so: po for po, so in so_map.items()}     # for External Doc re-stamp

    rows, line_dicts, header_dicts = [], [], []
    for po, p in parsed['pos'].items():
        so = so_map.get(po, po)                          # generated SO → 'No.'
        loc = f"{p['city']} ({p['ship_to']})"
        header_dicts.append({
            'po': so, 'segment': 'Offline', 'marketplace_label': MARKETPLACE,
            'location': loc, 'po_date': p.get('po_date') or '',
            'exp_date': p.get('exp_date') or '', 'qty': p['qty'],
            'order_value': p['value']})
        for ln in p['lines']:
            gst = ln['gst_code'] if str(ln['gst_code']).endswith('%') \
                else f"{ln['gst_code']}%"
            rows.append(SORow(
                po_number=so, location=loc, item_no=ln['item_no'], qty=ln['qty'],
                cust_no=CUST_NO, ship_to=p['ship_to'], mapped=True,
                mapped_location=p['city'], ean=ln['ean'],
                description=ln['description'], gst_code=gst,
                amount=ln['total_inc'], validation_status='OK'))  # unit_price=None → blank
            line_dicts.append({
                'po': so, 'item_no': ln['item_no'], 'ean': ln['ean'],
                'description': ln['description'], 'qty': ln['qty'],
                'our_mrp': None, 'vendor_mrp': None, 'status': 'OK', 'diff': None})

    if out_path is None:
        out_dir = Path(settings.MEDIA_ROOT) / 'reliance_trends_out'
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"reliance_trends_{_dt.datetime.now():%d%m%Y_%H%M%S}.xlsx"
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    raw_df = None
    if pd is not None:
        try:
            raw_df = pd.read_excel(path)
        except Exception:  # noqa: BLE001
            raw_df = None

    result = ProcessingResult(
        rows=rows, warnings=[('', '', w) for w in parsed['warnings']],
        marketplace=MARKETPLACE, input_file=Path(path).name,
        input_file_path=str(out.parent / out.name), margin_pct=0.0,
        compare_basis='', compare_label='',
        warehouse_code=WH_CODE.get(warehouse, 'PICK'), warehouse_display=warehouse,
        raw_df=raw_df)
    try:
        written = SOExporter().export(result)
    except Exception as e:  # noqa: BLE001
        return None, f'export failed: {e}'
    if written is None:
        return None, 'exporter produced no file'
    written = Path(written)
    if written.resolve() != out.resolve():
        try:
            if out.exists():
                out.unlink()
            written.replace(out)
        except OSError:
            out = written
    # External Document No. = the Reliance PO (not the SO 'No.') — re-stamp.
    try:
        _restamp_external_doc(str(out), so_to_po)
    except Exception:  # noqa: BLE001
        pass
    # bonus sheets (best-effort — never fail the export)
    try:
        mt_workbook._append_sku_sheet(str(out), line_dicts)
    except Exception:  # noqa: BLE001
        pass
    try:
        mt_workbook._append_tracker_sheet(str(out), header_dicts, MARKETPLACE)
    except Exception:  # noqa: BLE001
        pass
    return out, None


def record(path, warehouse='AHD', source_file='') -> dict:
    """Parse + WRITE to ``renee_orders`` (runs + order_headers + order_lines) with
    PO-level dedup — mirrors :mod:`gt_mass_bridge`. The frozen engine is untouched;
    all writes are web-owned. Returns ``{ok, recorded, run_id, recorded_pos,
    skipped, lines, value, error}``."""
    import datetime as _dt
    from online_b2b.services import lines_store
    from online_b2b.services.order_db import _conn

    parsed = parse(path)
    if not parsed['ok']:
        return {'ok': False, 'error': parsed['error']}
    pos = parsed['pos']
    if not pos:
        return {'ok': False, 'error': 'No POs found in the file.'}

    out_name = source_file or 'Renee PO BAP.xlsx'
    # Dedup by EXTERNAL DOC (the Reliance vendor PO) — the 'po' column now carries
    # the generated SO number (new each run), so it can't be the dedup key (MT does
    # the same). Also look up SOs already assigned to recorded POs, so the returned
    # so_map covers the whole file (recorder + workbook share one mapping).
    try:
        with _conn() as (cur, d):
            ph = d['ph']
            cur.execute(f"SELECT po, external_doc FROM order_headers "
                        f"WHERE marketplace={ph} AND external_doc IS NOT NULL",
                        (MARKETPLACE,))
            existing_so = {str(ext): str(so) for so, ext in cur.fetchall()}
    except Exception as e:  # noqa: BLE001
        return {'ok': False, 'error': f'DB read failed: {e}'}

    new = {po: p for po, p in pos.items() if po not in existing_so}
    skipped = len(pos) - len(new)
    # Assign SO numbers for the NEW POs (burns the counter once) + fold in the
    # already-recorded ones → full {reliance_po: so} map for the caller/workbook.
    so_new = _assign_so(list(new.keys()), burn=True)
    so_map = {po: (existing_so.get(po) or so_new.get(po)) for po in pos}

    if not new:
        return {'ok': True, 'recorded': False, 'skipped': skipped,
                'reason': 'all POs already recorded', 'recorded_pos': 0, 'lines': 0,
                'so_map': so_map}

    run_ts = _dt.datetime.now()
    run_ts_s = run_ts.strftime('%Y-%m-%d %H:%M:%S')
    total_pos = len(new)
    total_items = sum(len(p['lines']) for p in new.values())
    total_qty = sum(p['qty'] for p in new.values())
    total_value = round(sum(p['value'] for p in new.values()), 2)
    try:
        with _conn() as (cur, d):
            ph = d['ph']
            cur.execute(
                f"INSERT INTO runs (run_ts, mode, source, marketplaces, total_pos, "
                f"total_items, total_qty, total_value, consolidated_path, tracker_path)"
                f" VALUES ({ph},'MANUAL',{ph},1,{ph},{ph},{ph},{ph},'','')",
                (run_ts, f'OFFLINE RELIANCE TRENDS (web): {out_name}', total_pos,
                 total_items, total_qty, total_value))
            run_id = cur.lastrowid
            hcols = ('run_id, run_ts, mode, segment, marketplace, marketplace_label, '
                     'po, location, warehouse, po_date, exp_date, order_type, items, '
                     'qty, order_value, output_file, external_doc')
            marks = ', '.join([ph] * 17)
            for po, p in new.items():
                loc = f"{p['city']} ({p['ship_to']})"
                cur.execute(
                    f"INSERT INTO order_headers ({hcols}) VALUES ({marks})",
                    (run_id, run_ts, 'MANUAL', SEGMENT, MARKETPLACE, MARKETPLACE,
                     so_map[po], loc, warehouse, p.get('po_date'), p.get('exp_date'),
                     'SO', len(p['lines']), p['qty'], p['value'], out_name, po))
            cur.connection.commit()
    except Exception as e:  # noqa: BLE001
        return {'ok': False, 'error': f'DB write failed: {e}'}

    rows = []
    for po, p in new.items():
        loc = f"{p['city']} ({p['ship_to']})"
        for ln in p['lines']:
            gst = ln['gst_code'] if str(ln['gst_code']).endswith('%') else f"{ln['gst_code']}%"
            rows.append({
                'run_id': run_id, 'run_ts': run_ts_s, 'marketplace': MARKETPLACE,
                'po': so_map[po], 'location': loc, 'item_no': ln['item_no'],
                'ean': ln['ean'], 'description': ln['description'],
                'qty': ln['qty'], 'order_type': 'SO', 'gst_code': gst,
                'unit_price': ln['unit_price'], 'output_file': out_name,
                'our_mrp': None, 'vendor_mrp': None, 'our_landing': ln['unit_price'],
                'vendor_landing': None, 'our_cp': None, 'vendor_cp': None,
                'diff': None, 'margin_pct': None,
                'status': 'OK' if ln['resolved'] else 'NOT_IN_MASTER',
                'exception_label': '', 'received_ean': None, 'action': '',
                'remark': '' if ln['resolved'] else 'EAN not in item master',
            })
    try:
        lines_store.insert_lines(run_id, rows)
    except Exception as e:  # noqa: BLE001 — never block on the audit
        return {'ok': True, 'recorded': True, 'run_id': run_id,
                'recorded_pos': len(new), 'skipped': skipped, 'lines': 0,
                'value': total_value, 'reason': f'lines audit skipped: {e}',
                'so_map': so_map}
    return {'ok': True, 'recorded': True, 'run_id': run_id,
            'recorded_pos': len(new), 'skipped': skipped, 'lines': len(rows),
            'value': total_value, 'so_map': so_map}
