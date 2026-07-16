"""
offline.services.csd_processor
==============================

**CSD (Canteen Stores) order → D365 Sales Order builder.**

The customer (e.g. Indian Naval Canteen Service, Mumbai) sends an INCS order
workbook whose **``Shades``** sheet is the real demand — one row per individual
shade/SKU with its EAN, ``Demand (units)`` and current ``Inv`` (inventory). This
module turns that into the two D365 Sales Orders the team files by hand today:

  1. **Main SO**  — every demanded, in-stock line (``Demand>0 AND Inv>0``),
                    priced at the sheet's **Basic Price** (ex-GST).
  2. **Testers SO** — 1 unit of each **non-nail** main line at a nominal price,
                    external-doc ``TESTERS``.

Business rules (confirmed against a hand-built SO — reproduces it exactly):
  * drop **undemanded** rows (Demand = 0) and **zero-inventory** rows (Inv = 0);
  * **nails** (HSN ``33049920``) stay in the Main SO but are **excluded from
    Testers**;
  * line qty = Demand (units); unit price = Basic Price; Location = ``PICK``.

Pure/headless — no Django views here, so it's unit-testable and reusable by the
web flow. EAN→item_no resolves through the DB item master (never silent: an
unmapped EAN or a dropped line is reported in ``warnings``).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import openpyxl

# ── Fixed CSD parameters (confirmed from the hand-built SO) ──────────────────
LOCATION_CODE = 'PICK'
SUPPLY_TYPE = 'B2B'
DOC_TYPE = 'Order'
TESTER_UNIT_PRICE = 0.54          # nominal tester price (flat, per unit)
TESTER_QTY = 1                    # 1 unit of each non-nail item
TESTERS_EXT_DOC = 'TESTERS'
NAIL_HSN = '33049920'             # nail paint / enamel — excluded from testers

# The demand sheet + its column layout (0-based indices on the data rows).
DEMAND_SHEET = 'Shades'
_HDR_MARKER = 'Index No.'         # header row is the one whose 2nd cell is this
_C_INDEX, _C_HSN, _C_EAN, _C_DESC = 1, 3, 4, 5
_C_MRP, _C_BASIC, _C_GST, _C_INCS = 7, 8, 9, 10
_C_DEMAND, _C_INV = 11, 13


@dataclass
class CSDLine:
    ean: str
    item_no: str
    description: str
    hsn: str
    mrp: float
    basic_price: float
    gst: float
    demand: int
    inv: int

    @property
    def is_nail(self) -> bool:
        return str(self.hsn).strip() == NAIL_HSN


@dataclass
class CSDResult:
    lines: list[CSDLine] = field(default_factory=list)       # kept (main SO)
    warnings: list[str] = field(default_factory=list)
    dropped_zero_inv: list[str] = field(default_factory=list)
    dropped_no_demand: int = 0
    unmapped_eans: list[str] = field(default_factory=list)
    customer_name: str = ''
    ship_to_hint: str = ''

    @property
    def testers(self) -> list[CSDLine]:
        """Main lines that are NOT nails — 1 unit each on the Testers SO."""
        return [l for l in self.lines if not l.is_nail]


def _num(v, default=0):
    try:
        if v is None or v == '':
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _ean_to_item() -> dict:
    """``{ean: item_no}`` from the DB item master (the single source of truth)."""
    from online_b2b.services.order_db import _conn
    out = {}
    with _conn() as (cur, d):
        cur.execute('SELECT ean, item_no FROM item_master')
        for ean, item in cur.fetchall():
            if ean:
                out[str(ean).strip()] = str(item)
    return out


def _find_header_row(ws) -> int:
    """Row index (1-based) of the demand table header — the row whose 2nd cell is
    'Index No.'. Falls back to 13 (the observed layout)."""
    for r in range(1, 40):
        v = ws.cell(row=r, column=_C_INDEX + 1).value
        if v and _HDR_MARKER.lower() in str(v).strip().lower():
            return r
    return 13


def parse_demand(path: str) -> CSDResult:
    """Read the CSD workbook's demand (``Shades``) sheet → a :class:`CSDResult`
    of the kept Main-SO lines, plus warnings for everything dropped/unmapped."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if DEMAND_SHEET not in wb.sheetnames:
        raise ValueError(
            f"'{DEMAND_SHEET}' sheet not found — this doesn't look like a CSD "
            f"order file (sheets: {', '.join(wb.sheetnames)}).")
    ws = wb[DEMAND_SHEET]
    res = CSDResult()

    # Header meta (customer / ship-to) sits above the table — best-effort.
    for r in range(1, 12):
        row_txt = ' '.join(str(c.value) for c in ws[r] if c.value)
        if 'Customer name' in row_txt:
            res.customer_name = row_txt.split(':', 1)[-1].strip()
        elif row_txt.lower().startswith('ship to'):
            res.ship_to_hint = row_txt.split(':', 1)[-1].strip()

    e2i = _ean_to_item()
    hdr = _find_header_row(ws)
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if not row or len(row) <= _C_INV:
            continue
        idx = row[_C_INDEX]
        ean = str(row[_C_EAN]).strip() if row[_C_EAN] else ''
        if not idx or not ean:
            continue
        demand = int(_num(row[_C_DEMAND]))
        inv = int(_num(row[_C_INV]))
        desc = str(row[_C_DESC] or '').strip()
        if demand <= 0:                       # undemanded — skip (counted)
            res.dropped_no_demand += 1
            continue
        if inv <= 0:                          # zero inventory — DROP (named)
            res.dropped_zero_inv.append(f"{desc} (EAN {ean})")
            continue
        item_no = e2i.get(ean)
        if not item_no:                       # never silent — flag & skip
            res.unmapped_eans.append(f"{desc} (EAN {ean})")
            continue
        res.lines.append(CSDLine(
            ean=ean, item_no=item_no, description=desc,
            hsn=str(row[_C_HSN] or '').strip(),
            mrp=_num(row[_C_MRP]), basic_price=round(_num(row[_C_BASIC]), 6),
            gst=_num(row[_C_GST]), demand=demand, inv=inv))

    if res.dropped_zero_inv:
        res.warnings.append(
            f"{len(res.dropped_zero_inv)} demanded line(s) DROPPED for zero "
            f"inventory: {'; '.join(res.dropped_zero_inv)}")
    if res.unmapped_eans:
        res.warnings.append(
            f"{len(res.unmapped_eans)} demanded line(s) skipped — EAN not in the "
            f"item master: {'; '.join(res.unmapped_eans)}")
    return res


# ── Order building + D365 workbook export ───────────────────────────────────
# The hand-filed output is a 2-sheet Dynamics BC config-package workbook: table
# 36 = Sales Header, 37 = Sales Line. Reproduced exactly (see the sample SO).
_HEADER_COLS = [
    'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code', 'Posting Date',
    'Order Date', 'Document Date', 'Invoice From Date', 'Invoice To Date',
    'External Document No.', 'Location Code', 'Dimension Set ID', 'Supply Type',
    'Voucher Narration', 'Brand Code (Dimension)', 'Channel Code (Dimension)',
    'Catagory (Dimension)', 'Geography Code (Dimension)']
_LINE_COLS = [
    'Document Type', 'Document No.', 'Line No.', 'Type', 'No.', 'Location Code',
    'Quantity', 'Unit Price']
_TITLE = 'RENEE CSD - SO'


def _bump_so(no: str) -> str:
    """Next SO number — increment the trailing digits (SO/CSD/07/13726 → 13727)."""
    import re
    m = re.search(r'(\d+)(\D*)$', str(no))
    if not m:
        return str(no) + '-T'
    width = len(m.group(1))
    nxt = str(int(m.group(1)) + 1).zfill(width)
    return str(no)[:m.start(1)] + nxt + m.group(2)


def _header_row(no, cust, ship_to, order_date, ext_doc) -> list:
    return [DOC_TYPE, no, cust, ship_to, order_date, order_date, order_date,
            order_date, order_date, ext_doc, LOCATION_CODE, '', SUPPLY_TYPE,
            '', '', '', '', '']


def build_orders(res: CSDResult, so_main_no: str, cust_no: str, ship_to: str,
                 order_date: str, so_test_no: str = '') -> dict:
    """Turn the parsed demand into the two D365 SOs (headers + lines).

    ``order_date`` is a ``dd-mm-yyyy`` string (all five date columns use it).
    ``so_test_no`` defaults to ``so_main_no`` + 1. Returns a dict with
    ``headers`` (2 rows) and ``lines`` (main + tester rows), plus totals.
    """
    so_test_no = so_test_no or _bump_so(so_main_no)
    headers = [
        _header_row(so_main_no, cust_no, ship_to, order_date, so_main_no),
        _header_row(so_test_no, cust_no, ship_to, order_date, TESTERS_EXT_DOC)]

    lines = []
    ln = 10000
    for l in res.lines:                       # Main SO — every kept line
        lines.append([DOC_TYPE, so_main_no, ln, 'Item', l.item_no, LOCATION_CODE,
                      l.demand, l.basic_price])
        ln += 10000
    ln = 10000
    for l in res.testers:                     # Testers SO — 1 unit, non-nail only
        lines.append([DOC_TYPE, so_test_no, ln, 'Item', l.item_no, LOCATION_CODE,
                      TESTER_QTY, TESTER_UNIT_PRICE])
        ln += 10000

    return {
        'so_main_no': so_main_no, 'so_test_no': so_test_no,
        'cust_no': cust_no, 'ship_to': ship_to, 'order_date': order_date,
        'headers': headers, 'lines': lines,
        'main_lines': len(res.lines), 'tester_lines': len(res.testers),
        'main_qty': sum(l.demand for l in res.lines),
        'tester_qty': len(res.testers) * TESTER_QTY,
    }


def export_workbook(built: dict, out_path: str, title: str = _TITLE) -> str:
    """Write the 2-sheet (Sales Header / Sales Line) D365 config-package file."""
    from openpyxl import Workbook
    wb = Workbook()
    wsh = wb.active
    wsh.title = 'Sales Header'
    wsh.append([title, 'Sales Header', 36])
    wsh.append([])
    wsh.append(_HEADER_COLS)
    for h in built['headers']:
        wsh.append(h)

    wsl = wb.create_sheet('Sales Line')
    wsl.append([title, 'Sales Line', 37])
    wsl.append([])
    wsl.append(_LINE_COLS)
    for l in built['lines']:
        wsl.append(l)

    wb.save(out_path)
    return out_path
