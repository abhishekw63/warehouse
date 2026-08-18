"""Unit tests for online_b2b.services.common — the shared web helpers that Phase 4
consolidated (used app-wide, so locked down here). No DB needed → SimpleTestCase."""
import io
import tempfile
from pathlib import Path

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.http import Http404, HttpResponse
from django.test import RequestFactory, SimpleTestCase

from online_b2b.services import batch_flow as bf
from online_b2b.services import common
from online_b2b.services import marketplaces as reg


class TokenDirTests(SimpleTestCase):
    def setUp(self):
        self.root = Path(tempfile.mkdtemp())

    def test_normal_token_resolves_under_root(self):
        d = common.token_dir(self.root, 'abc123')
        self.assertEqual(d.name, 'abc123')
        self.assertEqual(d.parent.resolve(), self.root.resolve())

    def test_empty_token_returns_base(self):
        self.assertEqual(common.token_dir(self.root, '').resolve(), self.root.resolve())

    def test_path_traversal_raises_http404(self):
        for evil in ('../../etc', '..', '../secrets', 'a/../../b'):
            with self.assertRaises(Http404):
                common.token_dir(self.root, evil)


class SaveUploadTests(SimpleTestCase):
    def test_streams_file_to_dest_and_returns_path(self):
        dest = Path(tempfile.mkdtemp()) / 'out.bin'
        f = SimpleUploadedFile('x.bin', b'hello world' * 1000)
        out = common.save_upload(f, dest)
        self.assertEqual(out, dest)
        self.assertEqual(dest.read_bytes(), b'hello world' * 1000)


class IsAjaxTests(SimpleTestCase):
    def setUp(self):
        self.rf = RequestFactory()

    def test_xmlhttprequest_is_ajax(self):
        r = self.rf.get('/', HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertTrue(common.is_ajax(r))

    def test_no_header_is_not_ajax(self):
        self.assertFalse(common.is_ajax(self.rf.get('/')))

    def test_fetch_marker_is_not_this_helper(self):
        # the shell-nav 'fetch' marker is deliberately NOT treated as XHR here
        r = self.rf.get('/', HTTP_X_REQUESTED_WITH='fetch')
        self.assertFalse(common.is_ajax(r))


class PostDictTests(SimpleTestCase):
    def test_drops_csrf_token(self):
        r = RequestFactory().post('/', {'a': '1', 'b': '2', 'csrfmiddlewaretoken': 'xyz'})
        self.assertEqual(common.post_dict(r), {'a': '1', 'b': '2'})


class ValidDateTests(SimpleTestCase):
    def test_valid_and_invalid(self):
        self.assertEqual(common.valid_date('2026-08-18'), '2026-08-18')
        self.assertEqual(common.valid_date('  2026-08-18 '), '2026-08-18')
        for bad in ('bad', '2026-13-01', '', None, '18-08-2026'):
            self.assertEqual(common.valid_date(bad), '')


class ClampDaysTests(SimpleTestCase):
    def test_presets_and_default(self):
        for ok in (7, 30, 90, '7', '30', '90'):
            self.assertIn(common.clamp_days(ok), (7, 30, 90))
        for bad in (45, 0, -1, 'x', None, ''):
            self.assertEqual(common.clamp_days(bad), 30)

    def test_custom_allowed_and_default(self):
        self.assertEqual(common.clamp_days(5, allowed=(5, 10), default=10), 5)
        self.assertEqual(common.clamp_days(99, allowed=(5, 10), default=10), 10)


class XlsxResponseTests(SimpleTestCase):
    def _load(self, resp):
        return openpyxl.load_workbook(io.BytesIO(resp.content)).active

    def test_response_shape(self):
        resp = common.xlsx_response('Sheet', [('a', 'A')], [{'a': 1}], 'f.xlsx')
        self.assertIsInstance(resp, HttpResponse)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        self.assertEqual(resp['Content-Disposition'], 'attachment; filename="f.xlsx"')

    def test_header_values_fill_and_rows(self):
        cols = [('name', 'Name'), ('n', 'Count')]
        rows = [{'name': 'alpha', 'n': 3}, {'name': 'beta', 'n': None}]
        ws = self._load(common.xlsx_response('Data', cols, rows, 'd.xlsx'))
        self.assertEqual(ws.title, 'Data')
        self.assertEqual([ws.cell(1, c).value for c in (1, 2)], ['Name', 'Count'])
        self.assertEqual(ws.cell(1, 1).fill.fgColor.rgb[-6:], '1A237E')  # navy header
        self.assertTrue(ws.cell(1, 1).font.bold)
        self.assertEqual([ws.cell(2, 1).value, ws.cell(3, 1).value], ['alpha', 'beta'])

    def test_freeze_and_str_cols(self):
        ws = self._load(common.xlsx_response(
            'X', [('ts', 'When')], [{'ts': 20260818}], 'x.xlsx',
            freeze=True, str_cols=('ts',)))
        self.assertEqual(ws.freeze_panes, 'A2')
        self.assertEqual(ws.cell(2, 1).value, '20260818')  # coerced to str, not int


class MarketplaceRegistryTests(SimpleTestCase):
    """Locks the auto-detect resolution — incl. the BlinkMP header/line split fix."""

    def _resolve(self, marketplace, label):
        # mirrors daily_checklist._recorded_web: label (fine) first, then marketplace.
        dl, dk = reg.db_label_to_channel(), reg.db_key_to_channel()
        return dl.get(str(label or '')) or dk.get(str(marketplace))

    def test_blinkmp_header_autodetects(self):
        # BlinkMP records order_headers.marketplace/label='BlinkMP' → must resolve so
        # Daily Tasks auto-ticks (the bug: it only had db_key='Blink RO').
        self.assertEqual(self._resolve('BlinkMP', 'BlinkMP'), 'blinkmp')

    def test_blinkmp_line_key_still_folds(self):
        # order_lines record marketplace='Blink RO' — the board fold must still map it.
        self.assertEqual(reg.db_key_to_channel().get('Blink RO'), 'blinkmp')

    def test_blink_and_blinkmp_are_distinct(self):
        self.assertEqual(reg.get('blink').db_key, 'Blink')
        self.assertEqual(reg.get('blinkmp').db_key, 'Blink RO')
        self.assertEqual(reg.get('blinkmp').db_label, 'BlinkMP')
        self.assertEqual(self._resolve('Blink', 'Blink'), 'blink')  # not blinkmp

    def test_channel_keys_are_unique(self):
        keys = [c.key for c in reg.channels()]
        self.assertEqual(len(keys), len(set(keys)))

    def test_db_key_map_has_no_collisions(self):
        db_keys = [c.db_key for c in reg.channels() if c.db_key]
        self.assertEqual(len(db_keys), len(set(db_keys)))


class BatchDetectorTests(SimpleTestCase):
    """Locks the READ-ONLY batch-run MP detector's pure logic (no files/DB)."""

    def test_norm_folds_to_alnum(self):
        self.assertEqual(bf._norm('Supplier Unit Price'), 'supplierunitprice')
        self.assertEqual(bf._norm('PO No.'), 'pono')

    def test_flatten_cols_handles_str_list_dict_and_skips_placeholders(self):
        self.assertEqual(bf._flatten_cols('PO'), ['PO'])
        self.assertEqual(bf._flatten_cols(['PO', 'PO Number']), ['PO', 'PO Number'])
        self.assertEqual(bf._flatten_cols({'multiply': ['Landing Price', 'Quantity']}),
                         ['Landing Price', 'Quantity'])
        self.assertEqual(bf._flatten_cols('__po__'), [])   # engine placeholder skipped

    def test_filename_hint_disambiguates_lookalikes(self):
        self.assertEqual(bf._filename_hint('POItemExport_2026-08-18.xls'), 'RK')
        self.assertEqual(bf._filename_hint('purchase_order_FLS073C0F2B4.xlsx'), 'Flipkart')
        self.assertEqual(bf._filename_hint('Consignment_Details_204838801.csv'), 'Flipkart-TO')
        # Swiggy (all-digits) vs Zepto (hex) share PO_<id>.csv
        self.assertEqual(bf._filename_hint('PO_1786601384021.csv'), 'Swiggy')
        self.assertEqual(bf._filename_hint('PO_2683fb1ca1f301b0.csv'), 'Zepto')
        self.assertIsNone(bf._filename_hint('random_upload.xlsx'))

    def test_signatures_built_from_engine_configs(self):
        sigs = bf.signatures()
        self.assertIn('Myntra', sigs)
        # Myntra's GTIN column must be part of its signature
        self.assertIn('gtin', sigs['Myntra']['cols'])
