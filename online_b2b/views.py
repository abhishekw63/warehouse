"""
online_b2b.views — Phase 0 (Blink pilot)

Dashboard reads the order history straight from MySQL ``renee_orders``.
Upload runs the existing engine as a library and records the result back into
the same DB, so the dashboard reflects web runs immediately.
"""

import hashlib
import json
import os
import shutil
import uuid
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView

from .forms import UploadForm
from .services import engine_bridge, erp_import, order_db

# ── Central hub + branch dashboards (class-based) ───────────────────────────
# /b2b/ is the central Order-Management hub: compact overall KPIs + two group
# cards (Online B2B / Offline). Each group drills into its own branch dashboard.
# Channels differ a lot between the two worlds, so they stay distinct branches.

# Channels shown under each group card on the hub.
ONLINE_CHANNELS = [
    {'name': 'Blink', 'tag': 'live'},
    {'name': 'Flipkart', 'tag': 'live'},
    {'name': 'RK', 'tag': 'live'},
    {'name': 'DMart', 'tag': 'live'},
    {'name': 'Zepto', 'tag': 'live'},
    {'name': 'Flipkart Branch', 'tag': 'live'},
    {'name': 'Purplle', 'tag': 'live'},
    {'name': 'Swiggy', 'tag': 'live'},
    {'name': 'Nykaa', 'tag': 'live'},
    {'name': 'Myntra', 'tag': 'live'},
    {'name': 'Reliance', 'tag': 'live'},
    {'name': 'Meesho Branch', 'tag': 'live'},
    {'name': 'Big Basket', 'tag': 'live'},
    {'name': 'First Cry', 'tag': 'live'},
    # ── TO-DO · pending web integration (shown as "coming soon") ──
    {'name': 'BlinkMP', 'tag': 'soon'},
    {'name': 'Smytten', 'tag': 'soon'},
]
OFFLINE_CHANNELS = [
    {'name': 'MT (Modern Trade)', 'tag': 'live', 'url': '/offline/mt-flow/'},
    {'name': 'GT Mass', 'tag': 'live', 'url': '/offline/gt-mass-dump/'},
    {'name': 'GT Select', 'tag': 'live', 'url': '/b2b/gt-select/'},
    # ── TO-DO · pending web integration (shown as "coming soon") ──
    {'name': 'EKA', 'tag': 'soon'},
    {'name': 'CSD', 'tag': 'soon'},
    {'name': 'Off-Institutional', 'tag': 'soon'},
    {'name': 'Airport', 'tag': 'soon'},
    {'name': 'EBO / Kiosk', 'tag': 'soon'},
]


# Hub time-range selector — chip label per range (windowed KPIs). Default 30D.
_RANGE_LABELS = {
    'today': 'today', '7d': 'last 7 days', '30d': 'last 30 days',
    'mtd': 'this month', 'all': 'all-time',
}
_RANGE_DEFAULT = '30d'


class CentralHubView(LoginRequiredMixin, TemplateView):
    """`/b2b/` — central hub: compact overall KPIs + Online B2B / Offline groups.

    A global range switch (Today · 7D · 30D · MTD · All) scopes the VOLUME/VALUE
    KPI cards + the two segment cards by ``run_ts`` (default 30D). Genuinely
    cumulative metrics (Channels, Resolved, TAT) stay all-time. ``?range=`` picks
    the window; ``?partial=1`` (or an XHR) returns ONLY the KPI block for the
    no-refresh chip switch.
    """
    template_name = 'online_b2b/central.html'

    def _active_range(self):
        rng = self.request.GET.get('range', _RANGE_DEFAULT)
        return rng if rng in order_db.WINDOWS else _RANGE_DEFAULT

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        # AJAX chip switch — swap just the KPI block, no full-page chrome.
        if request.GET.get('partial') or (
                request.headers.get('x-requested-with') == 'XMLHttpRequest'):
            return render(request, 'online_b2b/_hub_kpis.html', context)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        rng = self._active_range()
        data = order_db.overview(segment='', window=rng)   # all channels, windowed
        ctx['ok'] = data.get('ok', False)
        ctx['error'] = data.get('error')
        ctx['backend'] = data.get('backend', '')
        ctx['kpis'] = data.get('kpis', {})
        ctx['issue_count'] = data.get('issue_count', 0)
        ctx['online_channels'] = ONLINE_CHANNELS
        ctx['offline_channels'] = OFFLINE_CHANNELS
        # Collapse the pending ones into a single "+N coming soon" chip (tooltip
        # lists them) so the hub card stays compact.
        ctx['online_soon'] = [c['name'] for c in ONLINE_CHANNELS if c.get('tag') == 'soon']
        ctx['offline_soon'] = [c['name'] for c in OFFLINE_CHANNELS if c.get('tag') == 'soon']
        ctx['online_kpis'] = order_db.segment_kpis('OnlineB2B', window=rng)
        ctx['offline_kpis'] = order_db.segment_kpis('Offline', window=rng)
        ctx['today'] = order_db.today_intake()
        ctx['recent'] = order_db.recent_orders(8)
        ctx['recent_runs'] = order_db.recent_runs(8)
        ctx['extra'] = order_db.hub_extra_kpis()
        from .services import tat_store
        tat_total = tat_store.breach_count()          # all TAT breaches (all-time)
        all_pos = ctx['extra'].get('all_pos') or 0    # all-time denominator
        ctx['tat_total'] = tat_total
        ctx['tat_rate'] = round(tat_total / all_pos * 100, 1) if all_pos else 0
        # Item-master staleness — 15-day refresh reminder.
        from .services import item_master_loader as iml
        ctx['im_status'] = iml.last_updated()
        # Range selector state for the chip switch + subtitle.
        ctx['active_range'] = rng
        ctx['range_label'] = _RANGE_LABELS.get(rng, 'last 30 days')
        ctx['ranges'] = [
            {'key': 'today', 'label': 'Today'}, {'key': '7d', 'label': '7D'},
            {'key': '30d', 'label': '30D'}, {'key': 'mtd', 'label': 'MTD'},
            {'key': 'all', 'label': 'All'},
        ]
        return ctx


# Branch descriptors — drive the shared overview template's header/actions.
ONLINE_BRANCH = {'kind': 'online', 'label': 'Online B2B'}
OFFLINE_BRANCH = {'kind': 'offline', 'label': 'Offline'}


# Code-level (behavioral) exceptions per marketplace — the ones enforced in the
# engine/bridge rather than the item_exceptions DB overlay. Keyed by the
# marketplace name (matches the config / template name).
_BEHAVIORAL_EXC = {
    'Swiggy': [
        {'title': 'Status filter', 'detail': 'Only CONFIRMED POs are punched; '
         'COMPLETED / EXPIRED / CANCELLED / PENDING are dropped from the run with '
         'a named notification per PO (never silent).',
         'ex_head': ['PO status', 'Action', 'Notification'],
         'ex_rows': [['CONFIRMED', '✅ Punched', '—'],
                     ['COMPLETED / EXPIRED', '🚫 Dropped',
                      '“PO STATUS … — IGNORED (not CONFIRMED)”']]},
        {'title': 'NFS → For-Sale', 'detail': 'A line landing on an NFS (Not-For-Sale) '
         'item is remapped to its <ean>_FS For-Sale twin so the negotiated deal '
         'price applies. Noted per line.',
         'ex_head': ['Input EAN', 'Resolves to', 'Item', 'Deal CP', 'Result'],
         'ex_rows': [['…674 (NFS)', '…674_FS (For Sale)', '200075', '₹35',
                      '✅ OK (was MISMATCH ₹134.92)']]},
    ],
    'Blink': [
        {'title': 'EPISENSE deal — unit-price override', 'detail': 'EPISENSE promo '
         'SKUs carry a negotiated deal price, so the engine applies a UNIT-PRICE '
         'override on those SKUs — the deal CP is written as the line’s Unit Price '
         'instead of the flat 70% landing, so the line is accepted rather than '
         'flagged as a mismatch. This is Blink’s only exception; MRP itself is not '
         'overridden and the standard MRP × margin ÷ GST rule (see Pricing rule) '
         'is unchanged.',
         'ex_head': ['Case', 'Standard basis', 'EPISENSE deal basis', 'Unit Price used'],
         'ex_rows': [['EPISENSE promo', '₹1099 @ 70%', '₹899 @ 76%', '₹579.02 ✅']]},
    ],
    'Myntra': [
        # NOTE: Goddess (Use Vendor CP) is NOT hard-coded here — it lives in the DB
        # 'Use Vendor CP' card (Master Exceptions overlay), so it's shown ONCE and
        # stays in sync. A "Use Vendor CP" SKU is distinct from a Deal SKU: it
        # accepts the vendor's stated CP as-is (no fixed price), whereas a Deal SKU
        # validates against an agreed transfer price.
        {'title': 'Compare on CP (+ Landing shown)', 'detail': 'Validation is on CP '
         '(List price, pre-GST); the with-GST Vendor Landing is shown alongside for '
         'reference only.',
         'ex_head': ['MRP', 'Vendor CP (no GST)', 'Vendor Landing (GST)', 'Validated on'],
         'ex_rows': [['₹199', '₹118.05', '₹139.30', 'CP']]},
    ],
    'Reliance': [
        {'title': 'GST-dependent margin', 'detail': "keep% isn't flat — it's "
         '1 − discount × (1 + GST), so it varies with each line’s GST rate.',
         'ex_head': ['GST rate', 'Keep %'],
         'ex_rows': [['18%', '63.42%'], ['12%', '65.28%'],
                     ['5%', '67.45%'], ['0%', '69.0%']]},
    ],
}


def _marketplace_exceptions(name: str) -> list:
    """Every exception applied to ONE marketplace, for its profile page: the
    code-level behavioral ones (``_BEHAVIORAL_EXC``) + the DB ``item_exceptions``
    overlay filtered to this marketplace (grouped by type with counts + a few
    examples). Read-only; never raises."""
    cards = [{**e, 'kind': 'behavioral'} for e in _BEHAVIORAL_EXC.get(name, [])]
    try:
        from .services import overrides_store
        allx = overrides_store.list_all()
    except Exception:  # noqa: BLE001
        return cards
    low = name.lower()
    def _clip(s, n=34):
        s = str(s or '')
        return s if len(s) <= n else s[:n - 1] + '…'

    if low == 'swiggy':
        deals = [r for r in allx if r.get('kind') == 'swiggy_deal']
        if deals:
            rows = [[r.get('item_id') or '—', r.get('source_code') or '—',
                     _clip(r.get('note')),
                     f"₹{r.get('override_mrp') or '—'}",
                     f"₹{r.get('cost_after_gst') or '—'}"] for r in deals]
            cards.append({'title': f'Deal SKUs ({len(deals)})', 'kind': 'deal',
                          'detail': 'Per-SKU negotiated deal prices — the agreed Cost '
                          'after GST is used as our CP (not MRP × margin). Full list:',
                          'ex_label': f'All {len(deals)}',
                          'ex_head': ['Item No', 'EAN', 'SKU', 'MRP', '→ CP (after GST)'],
                          'ex_rows': rows})
    # Myntra (and any marketplace) negotiated deal SKUs — the agreed transfer
    # price (Cost With GST) becomes the expected CP. List every one so the
    # operator can tally, same as the Swiggy deal card.
    mp_deals = [r for r in allx if r.get('kind') == 'myntra_deal'
                and str(r.get('marketplace') or '').strip().lower() == low]
    if mp_deals:
        rows = [[r.get('item_id') or '—', r.get('source_code') or '—',
                 _clip(r.get('note')),
                 f"₹{r.get('override_mrp') or '—'}",
                 f"₹{r.get('cost_with_gst') or '—'}"] for r in mp_deals]
        cards.append({'title': f'Deal SKUs ({len(mp_deals)})', 'kind': 'deal',
                      'detail': 'Per-SKU prices negotiated with the marketplace — the '
                      'agreed Cost With GST (transfer price) is used as the expected CP '
                      '(÷(1+GST)), not MRP × margin. Full list:',
                      'ex_label': f'All {len(mp_deals)}',
                      'ex_head': ['Item No', 'EAN', 'SKU', 'MRP', '→ CP (transfer, inc GST)'],
                      'ex_rows': rows})
    # DB overrides tagged to this marketplace, grouped by effect type. EVERY SKU is
    # listed (not just one example) so the operator can tally the full set.
    groups: dict = {}   # type -> {'head': [...], 'rows': [[...], ...]}
    for r in allx:
        if r.get('kind') in ('swiggy_deal', 'myntra_deal'):
            continue
        if str(r.get('marketplace') or '').strip().lower() != low:
            continue
        note = _clip(r.get('note') or r.get('source_code') or '')
        uc = (r.get('use_vendor_cp') or '').strip().upper()
        if uc.startswith('Y'):
            t = 'Use Vendor CP'
            head, row = ['SKU', 'Price used'], [note, 'vendor CP as-is']
        elif (r.get('maps_to') or '').strip():
            t = 'EAN remap'
            head, row = ['Punch EAN', '→ Master item'], [r.get('source_code'), r.get('maps_to')]
        elif (r.get('override_mrp') or '').strip():
            t = 'Override MRP'
            head, row = ['SKU', 'Deal MRP'], [note, f"₹{r.get('override_mrp')}"]
        elif (r.get('override_margin') or '').strip():
            t = 'Override margin'
            head, row = ['SKU', 'Margin'], [note, str(r.get('override_margin'))]
        else:
            t = 'Exception'
            head, row = ['SKU', 'Note'], [note, '—']
        # Blink's only real DB effect is the EPISENSE promo UNIT-PRICE override,
        # already shown as the 'EPISENSE deal' behavioral card above — so don't
        # also emit a redundant, mislabeled 'Override MRP' card for it. Scoped to
        # Blink's Override-MRP rows ONLY; every other marketplace's overlay cards
        # (Swiggy/Myntra deals, H&B, EAN remaps, Use Vendor CP, …) are untouched.
        if low == 'blink' and t == 'Override MRP':
            continue
        g = groups.setdefault(t, {'head': head, 'rows': []})
        g['rows'].append(row)
    for t, g in groups.items():
        n = len(g['rows'])
        cards.append({'title': f'{t} ({n})', 'kind': 'db',
                      'detail': 'From the Master Exceptions overlay — edit there and '
                      'it reflects here. Every SKU in this exception is listed:',
                      'ex_label': f'All {n}',
                      'ex_head': g['head'], 'ex_rows': g['rows']})
    return cards


def _mp_profile_context(name: str) -> dict:
    """Shared per-marketplace PROFILE context — the columns-read/ignored map
    (``tpl``) + the pricing ``rule`` + every ``exc_cards`` exception. The single
    source both the full "See full template" page and the Process-PO upload panel
    render, so they never drift. Read-only; ``tpl`` is ``None`` when no sample was
    captured for that marketplace."""
    tpl = engine_bridge.marketplace_template(name)
    nm = name
    if isinstance(tpl, dict):
        nm = tpl.get('name', name) or name
    try:
        rule = next((r for r in engine_bridge.marketplace_rules()
                     if str(r['name']).lower() == nm.lower()), None)
    except Exception:  # noqa: BLE001
        rule = None
    return {'tpl': tpl, 'rule': rule, 'exc_cards': _marketplace_exceptions(nm)}


def _group_exceptions(rows):
    """Group item_exceptions rows by marketplace for the Rules §4 live view —
    JSON-safe, one entry per marketplace with its lines (Swiggy deals grouped
    on their own). Type is derived so the operator sees WHAT each row does."""
    groups: dict = {}
    for r in rows:
        if r.get('kind') == 'swiggy_deal':
            mp, typ = 'Swiggy (deal SKUs)', 'Deal price'
        else:
            mp = (r.get('marketplace') or 'Any marketplace').strip() or 'Any marketplace'
            uc = (r.get('use_vendor_cp') or '').strip().upper()
            if uc.startswith('Y'):
                typ = 'Use Vendor CP'
            elif (r.get('maps_to') or '').strip():
                typ = 'EAN remap'
            elif (r.get('override_mrp') or '').strip():
                typ = 'Override MRP'
            elif (r.get('override_margin') or '').strip():
                typ = 'Override margin'
            else:
                typ = 'Exception'
        groups.setdefault(mp, []).append({
            'code': r.get('source_code', ''), 'maps_to': r.get('maps_to', '') or '',
            'type': typ, 'note': (r.get('note') or '')[:140],
            'cost': r.get('cost_after_gst') or r.get('cost_with_gst') or '',
            'mrp': r.get('override_mrp') or '', 'margin': r.get('override_margin') or ''})
    return [{'mp': k, 'rows': v} for k, v in sorted(groups.items())]


class RulesView(LoginRequiredMixin, TemplateView):
    """Marketplace Rules & Exceptions reference — margins, compare basis, item
    resolution, the engine's exception types + operator decisions, and Flipkart's
    location map. Reads the engine config so it never drifts."""
    template_name = 'online_b2b/rules.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        rules = engine_bridge.marketplace_rules()
        ctx['rules'] = rules
        # Marketplaces whose margin is GST-dependent (Reliance) — for the
        # elaborated exceptions section.
        ctx['gst_margin_rules'] = [r for r in rules if r.get('gst_margin')]
        ctx['formats'] = engine_bridge.marketplace_formats()
        # Marketplaces that have a full "See full template" preview available.
        ctx['template_names'] = list(engine_bridge.marketplace_templates().keys())
        ctx['locations'] = engine_bridge.location_rules()
        # Actual Swiggy deal SKUs (name + agreed prices) for accuracy on the card.
        try:
            from .services import overrides_store
            allx = overrides_store.list_all()
            ctx['swiggy_deals'] = [r for r in allx if r.get('kind') == 'swiggy_deal']
            ctx['zepto_deals'] = [r for r in allx if r.get('kind') == 'zepto_deal']
            ctx['exc_groups'] = _group_exceptions(allx)
            ctx['exc_total'] = len(allx)
        except Exception:  # noqa: BLE001
            ctx['swiggy_deals'] = []
            ctx['zepto_deals'] = []
            ctx['exc_groups'] = []
            ctx['exc_total'] = 0
        # MT (Modern Trade) child channels + their per-channel input requirements
        # (data-driven, so new channels e.g. Reliance auto-appear on the Rules page).
        try:
            from offline.services import mt_bridge
            eng = mt_bridge._engine()
            mt = []
            for code in mt_bridge.WEB_CHANNELS:
                cfg = eng.CHANNELS.get(code)
                if not cfg:
                    continue
                req = mt_bridge.channel_requirements(code) or {}
                mt.append({
                    'code': code, 'name': cfg.display_name,
                    'sell_to': getattr(cfg, 'sell_to', ''),
                    'lookup': getattr(cfg, 'lookup_via', ''),
                    'required': req.get('required', ''),
                    'optional': req.get('optional', ''),
                    'if_absent': req.get('if_absent', ''),
                })
            ctx['mt_channels'] = mt
        except Exception:  # noqa: BLE001
            ctx['mt_channels'] = []
        return ctx


class MarketplaceTemplateView(LoginRequiredMixin, TemplateView):
    """Rules → "See full template": the full column list + a few real sample rows
    for one marketplace, with the columns the engine actually reads highlighted
    (by role) and the rest dulled — a visual, drift-proof successor to the desktop
    "download template"."""
    template_name = 'online_b2b/template.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        name = self.kwargs.get('slug', '')
        # Full per-marketplace PROFILE on this page: the columns map + the pricing
        # RULE + every EXCEPTION we apply to THIS marketplace, built by the shared
        # helper (the same one the upload panel uses) so the two never drift.
        prof = _mp_profile_context(name)
        if prof['tpl'] is None:
            raise Http404(f'No template captured for “{name}”.')
        ctx.update(prof)
        return ctx


@login_required
def b2b_mp_profile(request, mp):
    """READ-ONLY: the shared marketplace-profile partial for ONE marketplace —
    powers the Process-PO upload page's live panel (fetched on the marketplace
    <select> change). Reuses the exact same context builder as the full template
    page; renders nothing new. Never writes; never touches the engine run."""
    prof = _mp_profile_context(mp)
    if prof['tpl'] is None:
        from django.utils.html import escape
        return HttpResponse(
            '<div class="mp-empty">No column template has been captured for '
            f'<b>{escape(mp)}</b> yet. Its pricing rule and exceptions still apply '
            'during processing.</div>')
    return render(request, 'online_b2b/_mp_profile.html', prof)


def _sku_filters(request):
    """Shared SKU-demand filters (marketplace + upload-date range). Defaults to
    the *last 30 days* of uploads on first load; validates the dates."""
    import datetime as _dt
    sf, st = request.GET.get('sku_from'), request.GET.get('sku_to')
    smp = (request.GET.get('sku_mp') or '').strip()
    if sf is None and st is None:
        st = _dt.date.today().isoformat()
        sf = (_dt.date.today() - _dt.timedelta(days=29)).isoformat()

    def _ok(v):
        v = (v or '').strip()
        if not v:
            return ''
        try:
            _dt.date.fromisoformat(v)
            return v
        except ValueError:
            return ''
    return _ok(sf), _ok(st), smp


def _daily_ctx(request):
    """Daily-Intake tab context (chart + breakdown tree). Its own filter —
    independent of the SKU tab. Either a quick preset (?days=7/30/90) or an
    explicit range (?start=&end=). A single day (start==end, or legacy ?date=)
    spotlights that bar on the chart. Defaults to the last 30 days."""
    import datetime as _dt

    def _ok(v):
        v = (v or '').strip()
        if not v:
            return ''
        try:
            _dt.date.fromisoformat(v)
            return v
        except ValueError:
            return ''

    try:
        days = int(request.GET.get('days') or 30)
    except (TypeError, ValueError):
        days = 30
    days = days if days in (7, 30, 90) else 30
    start, end = _ok(request.GET.get('start')), _ok(request.GET.get('end'))
    date = _ok(request.GET.get('date'))
    if date and not (start and end):                 # legacy single-date → range of one
        start = end = date
    if start and end and start > end:
        start, end = end, start

    if start and end:
        daily = order_db.daily_intake(start=start, end=end)
        hier = order_db.intake_hierarchy(start=start, end=end)
    else:
        daily = order_db.daily_intake(days)
        hier = order_db.intake_hierarchy(days)
    if start and end and start == end:               # spotlight the single day
        try:
            daily['focus'] = _dt.date.fromisoformat(start).strftime('%d %b')
        except ValueError:
            pass
    return {'days': days, 'start': start, 'end': end,
            'hier': hier, 'daily': daily}


def _sku_ctx(request):
    """SKU-demand tab context — full SKU list, own marketplace + upload-date
    filter (defaults to today's uploads)."""
    sf, st, smp = _sku_filters(request)
    return {'sku': order_db.sku_analytics(sf, st, smp, full=True),
            'sku_from': sf, 'sku_to': st, 'sku_mp': smp}


def _trends_ctx(request):
    """Trends & Momentum tab context — current window vs the previous equal
    window, with per-marketplace movers. Own ?days=7/30/90 filter."""
    try:
        days = int(request.GET.get('days') or 30)
    except (TypeError, ValueError):
        days = 30
    days = days if days in (7, 30, 90) else 30
    return {'days': days, 'trends': order_db.intake_trends(days)}


def _fulfil_ctx(request):
    """Fulfilment-Risk tab context — period demand vs current inventory, at-risk
    SKUs ranked by unfulfillable value. Reuses the SKU filter (marketplace +
    upload-date range; defaults to last 30 days)."""
    from .services import availability
    sf, st, smp = _sku_filters(request)
    return {'risk': availability.fulfilment_risk(sf, st, smp),
            'sku_from': sf, 'sku_to': st, 'sku_mp': smp}


def _exc_ctx(request):
    """Exceptions & Quality tab context — clean rate, mismatches, exceptions by
    marketplace + type. Reuses the SKU filter (marketplace + upload-date range;
    defaults to last 30 days)."""
    sf, st, smp = _sku_filters(request)
    return {'exc': order_db.exceptions_quality(sf, st, smp),
            'sku_from': sf, 'sku_to': st, 'sku_mp': smp}


def _geo_ctx(request):
    """Geography & Concentration tab context — demand by state/city + Pareto/ABC
    SKU concentration. Reuses the SKU filter (marketplace + upload-date range;
    defaults to last 30 days)."""
    sf, st, smp = _sku_filters(request)
    return {'geo': order_db.geography(sf, st, smp),
            'pareto': order_db.value_concentration(sf, st, smp),
            'sku_from': sf, 'sku_to': st, 'sku_mp': smp}


def _otif_ctx(request):
    """OTIF / Readiness tab — projected OTIF over OPEN orders. Own filter:
    marketplace + due-horizon (0 = all open, else due within N days)."""
    from .services import availability
    smp = (request.GET.get('sku_mp') or '').strip()
    try:
        horizon = int(request.GET.get('horizon') or 0)
    except (TypeError, ValueError):
        horizon = 0
    horizon = horizon if horizon in (0, 7, 30) else 0
    return {'rd': availability.fulfilment_readiness(smp, horizon),
            'sku_mp': smp, 'horizon': horizon}


def _dos_ctx(request):
    """Inventory Days-of-Supply tab — stock cover per SKU (on-hand ÷ avg daily
    demand). Own filter: demand window (?days=7/30/90) + marketplace."""
    from .services import inventory_store
    smp = (request.GET.get('sku_mp') or '').strip()
    try:
        days = int(request.GET.get('days') or 30)
    except (TypeError, ValueError):
        days = 30
    days = days if days in (7, 30, 90) else 30
    return {'dos': inventory_store.days_of_supply(days, smp),
            'days': days, 'sku_mp': smp}


class AnalyticsView(LoginRequiredMixin, TemplateView):
    """Management analytics — two AJAX tabs under one page, each with its own
    filter (no page refresh):
      • Daily Intake — daily stacked chart + segment→marketplace→child tree.
      • SKU Demand   — every SKU's demanded qty/value (lazy-loaded on first open).
    ``?partial=daily|sku`` returns just that tab body for the fetch() calls."""
    template_name = 'online_b2b/analytics.html'

    def get(self, request, *args, **kwargs):
        partial = request.GET.get('partial')
        if partial == 'daily':
            return render(request, 'online_b2b/_analytics_daily.html', _daily_ctx(request))
        if partial == 'sku':
            return render(request, 'online_b2b/_analytics_sku.html', _sku_ctx(request))
        if partial == 'trends':
            return render(request, 'online_b2b/_analytics_trends.html', _trends_ctx(request))
        if partial == 'fulfil':
            return render(request, 'online_b2b/_analytics_fulfil.html', _fulfil_ctx(request))
        if partial == 'exc':
            return render(request, 'online_b2b/_analytics_exc.html', _exc_ctx(request))
        if partial == 'geo':
            return render(request, 'online_b2b/_analytics_geo.html', _geo_ctx(request))
        if partial == 'otif':
            return render(request, 'online_b2b/_analytics_otif.html', _otif_ctx(request))
        if partial == 'dos':
            return render(request, 'online_b2b/_analytics_dos.html', _dos_ctx(request))
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_daily_ctx(self.request))         # Daily tab rendered on first paint
        return ctx


class SkuDemandView(LoginRequiredMixin, TemplateView):
    """Full view — every SKU's demanded qty + value for the chosen marketplace +
    upload-date range (defaults to today). Sortable; CSV export."""
    template_name = 'online_b2b/sku_demand.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        sf, st, smp = _sku_filters(self.request)
        ctx['sku'] = order_db.sku_analytics(sf, st, smp, full=True)
        ctx['sku_from'], ctx['sku_to'], ctx['sku_mp'] = sf, st, smp
        return ctx


@login_required
def sku_demand_export(request):
    """CSV of the full SKU-demand list for the current filters."""
    import csv
    import io
    sf, st, smp = _sku_filters(request)
    data = order_db.sku_analytics(sf, st, smp, full=True)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['Item No', 'Description', 'Qty', 'Value', 'POs', 'Marketplaces'])
    for r in data.get('rows', []):
        w.writerow([r['item_no'], r['description'], r['qty'], r['value'],
                    r['pos'], r['mps']])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="sku_demand.csv"'
    return resp


class TrackerView(LoginRequiredMixin, TemplateView):
    """Consolidated order tracker — the single source of truth across BOTH
    segments (Online B2B + Offline). One row per order (latest run per PO) with
    Dept · WH · Marketplace · PO · External Doc · Location · Pincode · Zone ·
    dates · value · qty · upload · file source. Server-side filters; CSV export."""
    template_name = 'online_b2b/tracker.html'
    SEG_MAP = {'Online B2B': 'OnlineB2B', 'Offline': 'Offline'}

    def _filters(self, request):
        seg_label = (request.GET.get('segment') or '').strip()
        return {
            'segment': self.SEG_MAP.get(seg_label, ''), 'seg_label': seg_label,
            'marketplace': (request.GET.get('marketplace') or '').strip(),
            'warehouse': (request.GET.get('warehouse') or '').strip(),
            'q': (request.GET.get('q') or '').strip(),
        }

    def _ctx(self, request):
        f = self._filters(request)
        return {'t': order_db.consolidated_tracker(
            f['segment'], f['marketplace'], f['warehouse'], f['q']), 'f': f}

    def get(self, request, *args, **kwargs):
        if request.GET.get('partial'):        # AJAX: KPIs + table only, no reload
            return render(request, 'online_b2b/_tracker_body.html', self._ctx(request))
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(self._ctx(self.request))
        return ctx


class TrackerExportView(LoginRequiredMixin, View):
    """CSV of the consolidated tracker honoring the current filters."""

    def get(self, request):
        import csv
        import io
        seg = TrackerView.SEG_MAP.get((request.GET.get('segment') or '').strip(), '')
        data = order_db.consolidated_tracker(
            seg, (request.GET.get('marketplace') or '').strip(),
            (request.GET.get('warehouse') or '').strip(),
            (request.GET.get('q') or '').strip(), limit=100000, display_limit=100000)
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['Dept', 'WH', 'Marketplace', 'PO', 'External Doc No', 'Location',
                    'PO Date', 'Exp Date', 'Order Qty', 'Order Value', 'Pincode',
                    'Zone', 'Uploaded', 'OMT'])
        for r in data.get('rows', []):
            w.writerow([r['dept'], r['wh'], r['marketplace'], r['po'], r['external_doc'],
                        r['location'], r['po_date'] or '', r['exp_date'] or '',
                        r['qty'], r['order_value'], r['pincode'], r['zone'],
                        r['uploaded'] or '', r.get('omt', '')])
        resp = HttpResponse(buf.getvalue(), content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="consolidated_tracker.csv"'
        return resp


class TrackerAddView(LoginRequiredMixin, View):
    """Add a MANUAL tracker row — for a PO that can't be uploaded via the app but
    still needs tracking. Writes only the isolated tracker_manual table."""

    def post(self, request):
        from .services import tracker_store
        res = tracker_store.add(request.POST.dict(), user=request.user.get_username())
        if not res.get('ok') and (request.headers.get('X-Requested-With') == 'fetch'):
            return JsonResponse(res, status=400)
        return redirect(request.META.get('HTTP_REFERER') or 'b2b_tracker')


class TrackerDeleteView(LoginRequiredMixin, View):
    """Delete a MANUAL tracker row (auto rows have no id and can't be deleted)."""

    def post(self, request, row_id):
        from .services import tracker_store
        tracker_store.delete(row_id)
        return redirect(request.META.get('HTTP_REFERER') or 'b2b_tracker')


class OfflineBranchView(LoginRequiredMixin, TemplateView):
    """`/b2b/offline/` — the Offline branch: SAME rich dashboard as Online B2B
    (KPIs + charts + marketplace mix), scoped to the Offline segment."""
    template_name = 'online_b2b/overview.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['d'] = order_db.overview(segment='Offline')
        ctx['branch'] = OFFLINE_BRANCH
        return ctx

_MEDIA = Path(settings.MEDIA_ROOT)
_RUNS_INDEX = _MEDIA / 'b2b_runs'          # run_id -> {output_path,...} sidecars


def _save_run_index(run_id, payload: dict) -> None:
    if run_id is None:
        return
    _RUNS_INDEX.mkdir(parents=True, exist_ok=True)
    (_RUNS_INDEX / f"{run_id}.json").write_text(
        json.dumps(payload), encoding='utf-8')


def _load_run_index(run_id) -> dict:
    p = _RUNS_INDEX / f"{run_id}.json"
    if p.exists():
        try:
            return json.loads(p.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}


def _persist_d365(run_id, marketplace, paths, warehouse, margin_pct, actions,
                  ean_fixes=None):
    """Build the decided D365 dump once at lock time and store it as a web-owned
    sidecar next to the run index. Returns the saved path, or None on any
    failure — a D365 hiccup (e.g. every line Excluded) must never block the lock.
    Pure file output: no DB connection is opened here."""
    try:
        _RUNS_INDEX.mkdir(parents=True, exist_ok=True)
        out = _RUNS_INDEX / f"{run_id}_d365.xlsx"
        res = engine_bridge.generate_d365(
            marketplace, paths, str(out),
            warehouse=warehouse, margin_pct=margin_pct, actions=actions,
            ean_fixes=ean_fixes)
        if res.get('ok') and os.path.exists(res.get('d365_path', '')):
            return str(out)
    except Exception:
        pass
    return None


def _int(request, name):
    try:
        return int(request.GET.get(name) or 0)
    except (TypeError, ValueError):
        return 0


def _filters(request):
    g = request.GET
    return {
        'segment': g.get('segment', '').strip(),
        'marketplace': g.get('marketplace', '').strip(),
        'days': _int(request, 'days'),
        'q': g.get('q', '').strip(),
        'warehouse': g.get('warehouse', '').strip(),
        'order_type': g.get('order_type', '').strip(),
        'date_from': g.get('date_from', '').strip(),
        'date_to': g.get('date_to', '').strip(),
        'sort': g.get('sort', 'date').strip() or 'date',
        'direction': g.get('dir', 'desc').strip() or 'desc',
    }


def _is_ajax(request):
    return bool(request.GET.get('partial')) or request.headers.get(
        'x-requested-with') == 'XMLHttpRequest'


@login_required
def dashboard(request):
    """Online B2B branch dashboard — KPIs + charts + marketplace summary for the
    online marketplaces (Blink/Flipkart/RK/DMart/BlinkMP). Reached from the central hub."""
    data = order_db.overview(segment='OnlineB2B')
    return render(request, 'online_b2b/overview.html',
                  {'d': data, 'branch': ONLINE_BRANCH})


@login_required
def orders(request):
    """Full order list — filter bar + table + load-more + export."""
    f = _filters(request)
    data = order_db.dashboard(offset=_int(request, 'offset'), **f)
    if _is_ajax(request):
        return render(request, 'online_b2b/_orders_results.html', {'d': data})
    return render(request, 'online_b2b/orders.html', {'d': data})


@login_required
def orders_more(request):
    """'Load more' → next page of order rows only."""
    f = _filters(request)
    page = order_db.orders_page(offset=_int(request, 'offset'), **f)
    return render(request, 'online_b2b/_orders_rows.html',
                  {'orders': page['orders'], 'page': page})


def _line_filters(request):
    g = request.GET
    return {
        'marketplace': g.get('marketplace', '').strip(),
        'status': g.get('status', '').strip(),
        'po': g.get('po', '').strip(),
        'q': g.get('q', '').strip(),
    }


@login_required
def lines(request):
    """Browsable line-items explorer (order_lines, all lines)."""
    data = order_db.line_items(offset=_int(request, 'offset'),
                               **_line_filters(request))
    if _is_ajax(request):
        return render(request, 'online_b2b/_lines_results.html', {'d': data})
    return render(request, 'online_b2b/lines.html', {'d': data})


@login_required
def lines_more(request):
    """'Load more' → next page of line rows only."""
    page = order_db.line_items_page(offset=_int(request, 'offset'),
                                    **_line_filters(request))
    return render(request, 'online_b2b/_lines_rows.html',
                  {'rows': page['rows']})


def _issue_filters(request) -> dict:
    """Shared Issues filters (used by the page + the export)."""
    return {
        'marketplace': request.GET.get('marketplace', '').strip(),
        'q': request.GET.get('q', '').strip(),
        'status': request.GET.get('status', '').strip(),
        'resolution': request.GET.get('resolution', 'pending').strip() or 'pending',
        'date_from': request.GET.get('date_from', '').strip(),
        'date_to': request.GET.get('date_to', '').strip(),
    }


@login_required
def issues(request):
    filters = _issue_filters(request)
    # Fresh open (no query params) → default to TODAY · All resolution, so the
    # page lands on today's issues. Any interaction (AJAX/Reset sends params)
    # keeps the operator's explicit choice — Reset still clears to all-time.
    if not request.GET:
        import datetime as _dt
        _today = _dt.date.today().isoformat()
        filters['resolution'] = 'all'
        filters['date_from'] = _today
        filters['date_to'] = _today
    data = order_db.issues(**filters)
    # Value KPIs (excluded/included qty + value, the loss) — same source as the
    # email, date-filtered — so the page cards match the mail and change with the
    # filter. Best-effort; the page still renders if this fails.
    try:
        from .services.issue_email import IssuesEmailReport
        data['kpi'] = IssuesEmailReport(filters).summary
    except Exception:  # noqa: BLE001
        data['kpi'] = {}
    if _is_ajax(request):
        return render(request, 'online_b2b/_issues_table.html', {'d': data})
    return render(request, 'online_b2b/issues.html',
                  {'d': data, 'eanfix': order_db.ean_corrections()})


@login_required
def issues_export(request):
    """Download the currently-filtered issue lines as .xlsx (respects every
    filter, including the upload-date window). No row cap on the export."""
    import datetime as _dt
    import io as _io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    data = order_db.issues(limit=0, **_issue_filters(request))
    rows = data.get('rows', []) if data.get('ok') else []
    cols = [('run_ts', 'Upload Date'), ('marketplace', 'Marketplace'),
            ('po', 'PO'), ('item_no', 'Item No'), ('ean', 'EAN'),
            ('received_ean', 'Received EAN'), ('description', 'Description'),
            ('qty', 'Qty'), ('vendor_mrp', 'Vendor MRP'), ('our_mrp', 'Our MRP'),
            ('vendor_cp', 'Vendor CP'), ('our_cp', 'Our CP'),
            ('vendor_landing', 'Vendor Landing'), ('our_landing', 'Our Landing'),
            ('diff', 'Diff'), ('status', 'Status'),
            ('exception_label', 'Exception'), ('action', 'Action'),
            ('remark', 'Remark')]
    wb = Workbook(); ws = wb.active; ws.title = 'Issues'
    hf = Font(bold=True, color='FFFFFF'); navy = PatternFill('solid', fgColor='1A237E')
    for c, (_k, h) in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.font = hf; cell.fill = navy
        cell.alignment = Alignment(horizontal='center')
    for r, row in enumerate(rows, 2):
        for c, (k, _h) in enumerate(cols, 1):
            v = row.get(k)
            ws.cell(r, c, str(v) if k == 'run_ts' and v is not None else v)
    for col in ws.columns:
        L = col[0].column_letter
        w = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[L].width = min(w + 2, 48)
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    f = _issue_filters(request)
    scope = f['date_from'] or 'all'
    if f['date_to'] and f['date_to'] != f['date_from']:
        scope = f"{f['date_from'] or 'start'}_to_{f['date_to']}"
    stamp = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f"issues_{f['resolution']}_{scope}_{stamp}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def _email_extras(request):
    """Pull the operator's note + recipient overrides off the request (modal
    posts them; preview may send them via GET for a live re-render). Returns
    ``(note, to, cc)`` where ``to``/``cc`` are ``None`` when the field was not
    supplied at all (⇒ report falls back to the config defaults)."""
    src = request.POST if request.method == 'POST' else request.GET
    note = src.get('note', '')
    to = src.get('to') if 'to' in src else None
    cc = src.get('cc') if 'cc' in src else None
    return note, to, cc


@login_required
def issues_email_preview(request):
    """Render (NO send) the Issues email for the current filter — the modal
    shows subject + recipients + the exact HTML that will be emailed. Accepts an
    optional ``note`` / ``to`` / ``cc`` so the preview reflects the operator's
    edits live."""
    from .services.issue_email import IssuesEmailReport
    note, to, cc = _email_extras(request)
    rep = IssuesEmailReport(_issue_filters(request), note=note, to=to, cc=cc)
    p = rep.preview()
    return JsonResponse({'ok': True, 'subject': p['subject'], 'html': p['html'],
                         'to': p['to'], 'cc': p['cc'], 'count': len(rep.rows)})


@login_required
@require_POST
def issues_email_send(request):
    """Send the Issues email for the current filter (filters come via the query
    string, same as the preview/export). The operator's note + edited To/Cc come
    via POST. Returns the delivery result as JSON — never sends silently on
    error."""
    from .services.issue_email import IssuesEmailReport
    note, to, cc = _email_extras(request)
    rep = IssuesEmailReport(_issue_filters(request), note=note, to=to, cc=cc)
    if not rep.rows:
        return JsonResponse(
            {'ok': False,
             'error': 'No issue lines in the current filter — nothing to send.'})
    # Never send with an empty To. If the operator supplied a To field (non-None)
    # that cleaned to nothing (all blank/invalid), refuse — do NOT silently fall
    # back to the config default and mail the wrong people. ``rep._to`` is the
    # cleaned override (None ⇒ To not supplied ⇒ config defaults apply).
    if to is not None and not rep._to:
        return JsonResponse(
            {'ok': False, 'error': 'No valid recipient in "To" — add at least '
             'one valid stakeholder email before sending.'})
    if not rep.recipients().get('to'):
        return JsonResponse(
            {'ok': False, 'error': 'No recipient configured — add a "To" address '
             'before sending.'})
    ok, reason = rep.send()
    return JsonResponse({'ok': ok, 'error': reason, 'count': len(rep.rows)})


@login_required
def daily_tasks(request):
    """Daily Activity Checklist — per-day grid of channels × workflow steps.
    Dual-render: JSON for AJAX (API-ready), template otherwise."""
    from .services import daily_checklist as dc
    from django.urls import reverse
    data = dc.get_day(request.GET.get('day'))
    # Per-channel "process URL" so clicking an MP name jumps straight to its upload
    # page with the marketplace pre-selected (skip Process-PO's manual pick). Online
    # web-integrated channels only; others get '' (plain label).
    _upl = reverse('b2b_upload')

    def _set_proc(seg_name, ch):
        ch['process_url'] = (_upl + '?mp=' + ch['db_key']
                             if seg_name == 'Online' and ch.get('live') and ch.get('db_key')
                             else '')
    for _seg in data.get('segments', []):
        for _c in _seg.get('channels', []):
            if _c.get('is_parent'):
                for _kid in _c.get('children', []):
                    _set_proc(_seg['segment'], _kid)
            else:
                _set_proc(_seg['segment'], _c)
    if _is_ajax(request):
        return JsonResponse({'ok': True, 'data': data, 'adhoc': dc.adhoc_list()})
    return render(request, 'online_b2b/daily_tasks.html',
                  {'d': data, 'adhoc': dc.adhoc_list()})


@login_required
@require_POST
def daily_tasks_toggle(request):
    """Tick/untick one cell (channel × step) for a day — records timestamp+user."""
    from .services import daily_checklist as dc
    res = dc.toggle(
        request.POST.get('day'),
        request.POST.get('channel', ''),
        request.POST.get('step', ''),
        request.POST.get('checked') in ('1', 'true', 'True', 'on'),
        user=request.user.get_username(),
        remark=request.POST.get('remark', ''))
    return JsonResponse(res)


@login_required
@require_POST
def daily_hold_reason(request):
    """Edit the Hold reason on a channel already on hold (no un-hold)."""
    from .services import daily_checklist as dc
    return JsonResponse(dc.set_hold_reason(
        request.POST.get('day'), request.POST.get('channel', ''),
        request.POST.get('remark', ''), user=request.user.get_username()))


@login_required
def daily_email_preview(request):
    """Render (NO send) the Daily Activity email for a day — the modal shows
    subject + recipients + the exact HTML that will be sent. Accepts optional
    ``note`` / ``to`` / ``cc`` so the preview reflects the operator's edits."""
    from .services.daily_email import DailyTasksEmailReport
    note, to, cc = _email_extras(request)
    rep = DailyTasksEmailReport(request.GET.get('day'), note=note, to=to, cc=cc)
    p = rep.preview()
    return JsonResponse({'ok': True, 'subject': p['subject'], 'html': p['html'],
                         'to': p['to'], 'cc': p['cc'],
                         'count': len(rep.active)})


@login_required
@require_POST
def daily_email_send(request):
    """Send the Daily Activity email for a day. The day comes via the query
    string; the note + edited To/Cc via POST. Returns JSON — never sends
    silently on error, and never falls back to config defaults if the operator
    supplied an empty/invalid To."""
    from .services.daily_email import DailyTasksEmailReport
    note, to, cc = _email_extras(request)
    rep = DailyTasksEmailReport(request.GET.get('day'), note=note, to=to, cc=cc)
    if to is not None and not rep._to:
        return JsonResponse(
            {'ok': False, 'error': 'No valid recipient in "To" — add your '
             'senior\'s email before sending.'})
    if not rep.recipients().get('to'):
        return JsonResponse(
            {'ok': False, 'error': 'No recipient configured — add a "To" address '
             'before sending.'})
    ok, reason = rep.send()
    return JsonResponse({'ok': ok, 'error': reason, 'count': len(rep.active)})


# ── Consolidated Summary email (dedicated Email page: review → send) ──────
def _summary_report(request):
    """Build a :class:`SummaryEmailReport` from the request — the day + optional
    subject come via the query string / POST; the note + edited To/Cc via
    ``_email_extras`` (same contract the Issues/Daily emails use)."""
    from .services.summary_email import SummaryEmailReport
    note, to, cc = _email_extras(request)
    src = request.POST if request.method == 'POST' else request.GET
    # Cockpit defaults to the ONLINE tab (Online/Offline are separate tabs now);
    # an explicit ?seg=offline / ?seg=both still overrides.
    return SummaryEmailReport(day=src.get('day') or None,
                              note=note, subject=src.get('subject', ''),
                              to=to, cc=cc, seg_filter=(src.get('seg') or 'online'))


@login_required
def email_page(request):
    """The dedicated **Email** page — lands on a composed CONSOLIDATED SUMMARY
    (received board + per-MP details) for REVIEW: subject, recipients and the
    summary are all shown/editable; nothing is sent until the operator clicks
    Send (which posts to :func:`email_send`). Read-only compose."""
    # Landing = a chooser (Online / Offline). We only build + show a segment's board
    # once the operator picks one (?seg=online|offline). Keeps the entry light.
    if not (request.GET.get('seg') or '').strip():
        import datetime as _dt
        return render(request, 'online_b2b/email.html', {
            'chooser': True, 'day': (request.GET.get('day') or _dt.date.today().isoformat())})
    rep = _summary_report(request)
    r = rep.recipients()
    return render(request, 'online_b2b/email.html', {
        'd': rep.data, 'subject': rep.subject(),
        'to': ', '.join(r['to']), 'cc': ', '.join(r['cc']),
    })


@method_decorator(login_required, name='dispatch')
class CockpitPOSkusView(View):
    """Lazy 2nd-level drill-down: the SKU rows for ONE po, fetched on click so the
    cockpit board never renders thousands of SKU rows upfront (that froze the page).
    GET ?po=&day=&pgid=[&mp=] → the ``_sku_rows.html`` partial. Read-only."""

    def get(self, request):
        from .services import summary_email as _se
        po = (request.GET.get('po') or '').strip()
        day = (request.GET.get('day') or '').strip() or None
        mp = (request.GET.get('mp') or '').strip()
        seg = (request.GET.get('seg') or '').strip()
        pgid = (request.GET.get('pgid') or '').strip()
        skus = _se.po_skus(day=day, marketplace=mp, po=po, segment=seg) if po else []
        return render(request, 'online_b2b/_cockpit_po_skus.html',
                      {'skus': skus, 'pgid': pgid})


@login_required
def email_preview(request):
    """Re-render the summary email for the current day/subject/note/recipients
    (AJAX) — NO send. Returns the exact subject + HTML that would be mailed, so
    the review pane stays live as the operator edits."""
    rep = _summary_report(request)
    p = rep.preview()
    return JsonResponse({'ok': True, 'subject': p['subject'], 'html': p['html'],
                         'to': p['to'], 'cc': p['cc'],
                         'received': rep.data['grand']['received_count'],
                         'total': rep.data['grand']['total_count']})


@login_required
@require_POST
def email_send(request):
    """Send the consolidated summary via the SAME shared send helper the
    Issues/Daily emails use. Guards match those flows — refuse an empty/invalid
    To rather than silently mailing the config default. Returns JSON."""
    rep = _summary_report(request)
    note, to, cc = _email_extras(request)
    if to is not None and not rep._to:
        return JsonResponse(
            {'ok': False, 'error': 'No valid recipient in "To" — add at least '
             'one valid stakeholder email before sending.'})
    if not rep.recipients().get('to'):
        return JsonResponse(
            {'ok': False, 'error': 'No recipient configured — add a "To" address '
             'before sending.'})
    ok, reason = rep.send()
    return JsonResponse({'ok': ok, 'error': reason,
                         'received': rep.data['grand']['received_count']})


@login_required
@require_POST
def daily_adhoc_add(request):
    """Add a personal ad-hoc task (random / Outlook item) to the Daily page."""
    from .services import daily_checklist as dc
    return JsonResponse(dc.adhoc_add(
        request.POST.get('title', ''), request.POST.get('note', ''),
        request.POST.get('due', ''), user=request.user.get_username()))


@login_required
@require_POST
def daily_adhoc_toggle(request):
    """Tick / untick an ad-hoc task (records done time + user)."""
    from .services import daily_checklist as dc
    return JsonResponse(dc.adhoc_toggle(
        request.POST.get('id'),
        request.POST.get('done') in ('1', 'true', 'True', 'on'),
        user=request.user.get_username()))


@login_required
@require_POST
def daily_adhoc_delete(request):
    """Delete an ad-hoc task."""
    from .services import daily_checklist as dc
    return JsonResponse(dc.adhoc_delete(request.POST.get('id')))


@login_required
@require_POST
def issues_save(request):
    """Save the operator's Action + Remark on one flagged line (Issues page)."""
    from .services import lines_store
    res = lines_store.update_action(
        request.POST.get('line_id'),
        request.POST.get('action', ''),
        request.POST.get('remark', ''))
    return JsonResponse(res)


@login_required
@require_POST
def issues_save_bulk(request):
    """Apply one Action + Remark to many flagged lines at once (Issues page)."""
    from .services import lines_store
    res = lines_store.update_action_bulk(
        request.POST.getlist('line_ids'),
        request.POST.get('action', ''),
        request.POST.get('remark', ''))
    return JsonResponse(res)


@login_required
@require_POST
def issues_fix_ean(request):
    """Post-lock EAN correction on a NOT_IN_MASTER line (Issues page).

    The lock-first model: the wrong EAN is already recorded. Operator types the
    correct EAN here → the line re-resolves against the item master, OUR pricing
    is recomputed with the engine's own helpers, the line flips to OK/MISMATCH,
    and the wrong EAN is kept as ``received_ean`` for the vendor-escalation audit.
    """
    from .services import lines_store
    res = lines_store.apply_issue_ean_fix(
        request.POST.get('line_id'),
        (request.POST.get('correct_ean', '') or '').strip())
    return JsonResponse(res, status=200 if res.get('ok') else 400)


# ── TAT (24h SLA) — breaches + reasons ──────────────────────────────────────

def _tat_filters(request) -> dict:
    return {
        'marketplace': request.GET.get('marketplace', '').strip(),
        'segment': request.GET.get('segment', '').strip(),
        'q': request.GET.get('q', '').strip(),
        'status': request.GET.get('status', 'pending').strip() or 'pending',
        'date_from': request.GET.get('date_from', '').strip(),
        'date_to': request.GET.get('date_to', '').strip(),
        'run': request.GET.get('run', '').strip(),
    }


@login_required
def tat(request):
    """TAT / SLA page — orders uploaded later than 1 working day after the PO
    date. Operator records a breach reason here (reactive). Filterable run-wise."""
    from .services import tat_store
    data = tat_store.breaches(**_tat_filters(request))
    data['reasons'] = tat_store.REASONS
    if _is_ajax(request):
        return render(request, 'online_b2b/_tat_rows.html', {'d': data})
    data['runs'] = order_db.recent_runs(50)
    return render(request, 'online_b2b/tat.html', {'d': data})


@login_required
@require_POST
def tat_save(request):
    """Set / clear the TAT breach reason for one order."""
    from .services import tat_store
    res = tat_store.set_reason(
        request.POST.get('order_id'),
        request.POST.get('reason_code', ''),
        request.POST.get('note', ''),
        by=getattr(request.user, 'username', ''))
    return JsonResponse(res, status=200 if res.get('ok') else 400)


@login_required
def tat_export(request):
    """Download the filtered TAT breaches as .xlsx (respects all filters)."""
    import datetime as _dt
    import io as _io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from .services import tat_store
    f = _tat_filters(request)
    data = tat_store.breaches(limit=100000, **f)
    rows = data.get('rows', []) if data.get('ok') else []
    cols = [('po_date', 'PO Date'), ('run_ts', 'Uploaded'),
            ('segment', 'Segment'), ('marketplace', 'Channel'), ('po', 'PO / SO'),
            ('location', 'Location'), ('qty', 'Qty'), ('order_value', 'Value'),
            ('wd_late', 'Working days taken'), ('days_over', 'Days over TAT'),
            ('reason_code', 'Reason'), ('note', 'Note'),
            ('reason_by', 'By'), ('reason_at', 'Reason at')]
    wb = Workbook(); ws = wb.active; ws.title = 'TAT breaches'
    hf = Font(bold=True, color='FFFFFF'); navy = PatternFill('solid', fgColor='1A237E')
    for c, (_k, h) in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.font = hf; cell.fill = navy; cell.alignment = Alignment(horizontal='center')
    for r, row in enumerate(rows, 2):
        for c, (k, _h) in enumerate(cols, 1):
            v = row.get(k)
            ws.cell(r, c, str(v) if k in ('run_ts', 'po_date', 'reason_at') and v is not None else v)
    for col in ws.columns:
        L = col[0].column_letter
        w = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[L].width = min(w + 2, 48)
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    stamp = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f"tat_breaches_{f['status']}_{stamp}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def sku_summary(request):
    """SKU-wise validation rollup — group by (Item No, EAN): qty + line-counts
    per status (OK / Mismatch / Not-in-master), MRP comparison, drill-down.
    READ-ONLY aggregation over the order_lines_full view (no DB changes)."""
    data = order_db.sku_summary(
        marketplace=request.GET.get('marketplace', '').strip(),
        q=request.GET.get('q', '').strip(),
        date_from=request.GET.get('from', '').strip(),
        date_to=request.GET.get('to', '').strip(),
        issues_only=request.GET.get('issues') == '1')
    if _is_ajax(request):
        return render(request, 'online_b2b/_sku_rows.html', {'d': data})
    return render(request, 'online_b2b/sku_summary.html', {'d': data})


@login_required
def sku_summary_lines(request):
    """Drill-down: the individual PO-lines behind one SKU (AJAX partial)."""
    data = order_db.sku_lines(request.GET.get('item_no', '').strip(),
                              request.GET.get('ean', '').strip())
    return render(request, 'online_b2b/_sku_drill.html', {'d': data})


@login_required
def export(request):
    """Stream the current filtered Orders view as an .xlsx."""
    import io
    from datetime import datetime

    from openpyxl import Workbook

    f = _filters(request)
    rows = order_db.orders_for_export(**f)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Online B2B Orders'
    cols = ['Run', 'Run Time', 'Marketplace', 'PO', 'Location', 'Warehouse',
            'PO Date', 'Exp Date', 'Type', 'Items', 'Qty', 'Order Value',
            'Days to Expiry']
    ws.append(cols)
    for r in rows:
        ws.append([
            r.get('run_id'), str(r.get('run_ts') or ''), r.get('marketplace'),
            r.get('po'), r.get('location'), r.get('warehouse'),
            str(r.get('po_date') or ''), str(r.get('exp_date') or ''),
            r.get('order_type'), r.get('items'), r.get('qty'), r.get('value'),
            r.get('days_to_expiry'),
        ])
    ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"online_b2b_orders_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    resp = FileResponse(buf, as_attachment=True, filename=fname,
                        content_type='application/vnd.openxmlformats-'
                        'officedocument.spreadsheetml.sheet')
    return resp


_UPLOADS = _MEDIA / 'b2b_uploads'
_ERP = _MEDIA / 'b2b_erp'


def _token_dir(token: str) -> Path:
    """Resolve a token folder, guarding against path traversal."""
    base = _UPLOADS.resolve()
    d = (_UPLOADS / token).resolve()
    if d != base and base not in d.parents:
        raise Http404()
    return d


def _load_meta(token: str):
    d = _token_dir(token)
    p = d / 'meta.json'
    if not p.exists():
        return None, d
    try:
        return json.loads(p.read_text(encoding='utf-8')), d
    except Exception:
        return None, d


# Per-token locks serialise the read-modify-write of meta.json so concurrent
# save-decision POSTs (Apply-to-selected fires several at once) don't clobber
# each other or read a half-written file.
import threading as _threading  # noqa: E402
_META_LOCKS: dict = {}
_META_LOCKS_GUARD = _threading.Lock()


def _meta_lock(token: str):
    with _META_LOCKS_GUARD:
        lk = _META_LOCKS.get(token)
        if lk is None:
            lk = _META_LOCKS[token] = _threading.Lock()
        return lk


def _save_meta(d: Path, meta: dict) -> None:
    """Persist a token's meta.json ATOMICALLY (temp + os.replace) so a concurrent
    reader never sees a partially-written file — fixes the save-decision race /
    intermittent 404. Callers that read-modify-write should hold ``_meta_lock``."""
    import tempfile
    p = d / 'meta.json'
    fd, tmp = tempfile.mkstemp(dir=str(d), suffix='.tmp')
    try:
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            json.dump(meta, f)
        os.replace(tmp, p)   # atomic on the same filesystem (incl. Windows)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


def _preview_sig(meta) -> str:
    """Cache key for a token's processed preview — everything that changes the
    engine output. Notably includes ``ean_fixes`` so an EAN-fix re-validate
    busts the cache automatically."""
    import hashlib
    key = json.dumps({
        'mp': meta.get('marketplace'), 'wh': meta.get('warehouse'),
        'mg': meta.get('margin_pct'), 'files': sorted(meta.get('files') or []),
        'fixes': meta.get('ean_fixes') or {},
    }, sort_keys=True)
    return hashlib.md5(key.encode()).hexdigest()


def _cached_preview(token, d, meta, force=False):
    """Run the engine preview ONCE per (files + EAN-fixes) and cache its JSON
    result on the token, so the review page — and every reload — renders
    instantly instead of re-running the engine. Re-runs only when the signature
    changes (new files / EAN fix) or ``force``."""
    cache = d / 'preview.json'
    sig = _preview_sig(meta)
    if not force and cache.exists():
        try:
            blob = json.loads(cache.read_text(encoding='utf-8'))
            if blob.get('sig') == sig and (blob.get('res') or {}).get('ok'):
                return blob['res']
        except Exception:  # noqa: BLE001 — corrupt cache → just re-run
            pass
    res = engine_bridge.preview(
        meta['marketplace'], [str(d / n) for n in meta['files']],
        warehouse=meta['warehouse'], margin_pct=meta['margin_pct'] / 100.0,
        ean_fixes=meta.get('ean_fixes'))
    if res.get('ok'):
        try:
            cache.write_text(json.dumps({'sig': sig, 'res': res}, default=str),
                             encoding='utf-8')
        except Exception:  # noqa: BLE001 — caching is best-effort
            pass
    return res


@login_required
def upload(request):
    """Phase 1: stash the uploaded PO(s) under a token and go to Review.
    Nothing is processed or written here."""
    if request.method == 'POST':
      # Wrap the WHOLE POST path: multipart/file parsing (request.FILES),
      # saving, and the engine preview can each raise. For AJAX we must never
      # 500 to an HTML page (the client does r.json() and shows only a vague
      # "Network error") — catch everything, log the traceback, and return the
      # REAL reason as JSON so the operator (and the logs) see what failed.
      try:
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            token = uuid.uuid4().hex[:12]
            up_dir = _UPLOADS / token
            up_dir.mkdir(parents=True, exist_ok=True)
            saved = []
            for f in form.cleaned_data['po_files']:
                dest = up_dir / Path(f.name).name
                with open(dest, 'wb') as out:
                    for chunk in f.chunks():
                        out.write(chunk)
                saved.append(dest.name)
            marketplace = form.cleaned_data['marketplace']
            # Blank margin → the marketplace's configured default landing rate
            # (Flipkart 77, Blink 70) — same as the Tkinter app's pre-fill.
            margin = (form.cleaned_data.get('margin_pct')
                      or engine_bridge.default_margin_pct(marketplace))
            meta = {
                'marketplace': marketplace,
                'warehouse': form.cleaned_data['warehouse'],
                'margin_pct': margin,
                'files': saved,
            }
            (up_dir / 'meta.json').write_text(json.dumps(meta), encoding='utf-8')
            # AJAX: do the import HERE (engine preview, cached) so the upload
            # page can show a real progress overlay + a definitive "✓ Imported"
            # completion before navigating — and the review page then loads from
            # the cache instantly. Non-JS clients fall back to the plain redirect
            # (review runs the same cached preview).
            if _is_ajax(request):
                from django.urls import reverse
                res = _cached_preview(token, up_dir, meta)
                if not res.get('ok'):
                    return JsonResponse(
                        {'ok': False, 'error': res.get('error', 'Import failed.')},
                        status=200)
                s = res.get('summary', {})
                return JsonResponse({
                    'ok': True, 'review_url': reverse('b2b_review', args=[token]),
                    'pos': s.get('pos', 0), 'lines': s.get('lines', 0),
                    'affected': s.get('affected', 0),
                    'warnings': len(res.get('warnings') or []),
                })
            return redirect('b2b_review', token=token)
        elif _is_ajax(request):
            errs = '; '.join(f"{k}: {v.as_text()}" for k, v in form.errors.items())
            return JsonResponse({'ok': False, 'error': errs or 'Invalid form.'},
                                status=200)
      except Exception as e:  # noqa: BLE001
        import logging
        logging.getLogger(__name__).exception('upload POST crashed')
        if _is_ajax(request):
            return JsonResponse(
                {'ok': False,
                 'error': f'Upload crashed: {type(e).__name__}: {e}'},
                status=200)
        raise
    else:
        # Pre-select the marketplace from ?mp= (Daily Tasks "click a channel → its
        # upload page"); fall back to the first pilot option.
        _mp = (request.GET.get('mp') or '').strip()
        initial_mp = (_mp if _mp in engine_bridge.PILOT_MARKETPLACES
                      else (engine_bridge.PILOT_MARKETPLACES[0]
                            if engine_bridge.PILOT_MARKETPLACES else ''))
        form = UploadForm(initial={'marketplace': initial_mp})
    # initial_mp (set above) drives both the <select> and the server-rendered
    # profile panel so it shows with no flash (the panel swaps live on change).
    return render(request, 'online_b2b/upload.html',
                  {'form': form,
                   'initial_mp': initial_mp,
                   **_mp_profile_context(initial_mp),
                   'margin_defaults': json.dumps(engine_bridge.margin_defaults()),
                   # {marketplace: "what file(s) this MP needs"} — drives the
                   # dynamic per-marketplace upload hint (updates on select change).
                   'mp_hints': json.dumps({
                       f['name']: f['note']
                       for f in engine_bridge.marketplace_formats() if f.get('note')}),
                   # marketplaces that have a "See full template" page → the hint
                   # shows a "Full detail →" link only for those.
                   'mp_templates': json.dumps(
                       list(engine_bridge.marketplace_templates().keys()))})


@login_required
def review(request, token):
    """Phase 2: process in memory (no DB write) and show the review page."""
    meta, d = _load_meta(token)
    if not meta:
        raise Http404("Upload not found or expired.")
    # Load the cached preview (the AJAX upload already ran + cached it → instant);
    # falls back to a fresh run for non-JS uploads or a busted cache. Reopening a
    # 'Review Later' draft passes ?revalidate=1 → force a fresh run so the team's
    # master correction (CP fix / new deal SKU) is picked up and the MISMATCH clears.
    force = request.GET.get('revalidate') == '1'
    res = _cached_preview(token, d, meta, force=force)
    has_preview = bool(res.get('output_path') and os.path.exists(res['output_path']))
    # Re-attach any saved decisions so the Affected rows show them (esp. when locked).
    decisions = meta.get('decisions') or {}
    for ln in res.get('affected') or []:
        key = f"{ln.get('po', '')}|{ln.get('item_no', '')}|{ln.get('ean', '')}"
        ln['decision'] = decisions.get(key, {})
    # Make the Line Items tab the FINAL ready-to-go view: attach each line's
    # disposition (the action taken on affected ones) so the operator sees OK /
    # Included / Override / Excluded per line before locking — same as the
    # post-lock Issues page, just earlier.
    for ln in res.get('lines') or []:
        key = f"{ln.get('po', '')}|{ln.get('item_no', '')}|{ln.get('ean', '')}"
        ln['decision'] = decisions.get(key, {})
    # NOT_IN_MASTER lines (wrong / unknown EAN) — the operator corrects the EAN
    # here, then re-validate resolves them against the DB master.
    nim_lines = [ln for ln in (res.get('affected') or [])
                 if ln.get('status') == 'NOT_IN_MASTER']
    # Auto-corrected lines: a previously-seen wrong EAN resolved through the
    # historical map. Flag them (TEMPORARY fix) + how many times that wrong EAN
    # was received before, so the team can escalate to the vendor.
    recv_counts = order_db.ean_correction_counts()
    auto_fixed = []
    for ln in (res.get('lines') or []):
        re = ln.get('received_ean')
        if re:
            ln['recv_count'] = recv_counts.get(re, 0)
            auto_fixed.append(ln)
    # Engine exceptions (vendor deals / overrides / remaps) — poke the operator
    # that these lines deviate from the flat marketplace rule.
    exc_count = sum(1 for ln in (res.get('lines') or [])
                    if (ln.get('exception_label') or '').strip())
    # KPI: qty on affected (MISMATCH / NOT_IN_MASTER) lines, and the share of qty
    # that is clean (OK). ok% = OK-qty / total-qty — e.g. 98% means only 2% of
    # the ordered units are flagged. Qty-weighted, not line-weighted.
    _AFF = {'MISMATCH', 'NOT_IN_MASTER'}
    total_qty = affected_qty = 0
    for ln in (res.get('lines') or []):
        q = int(ln.get('qty') or 0)
        total_qty += q
        if (ln.get('status') or '') in _AFF:
            affected_qty += q
    ok_qty_pct = round((total_qty - affected_qty) * 100 / total_qty, 1) if total_qty else 100.0
    # Per-PO CLEAN / AFFECTED for the Orders tab (mirrors the workbook Summary
    # 'Status' column): a PO is AFFECTED if any line is dropped — EXCLUDEd, or an
    # unresolved MISMATCH / NOT_IN_MASTER; else 100% goes to D365 → CLEAN.
    affected_pos = set()
    for ln in (res.get('lines') or []):
        act = str((ln.get('decision') or {}).get('action') or '').upper()
        if act == 'EXCLUDE' or (ln.get('status') in _AFF and act not in ('INCLUDE', 'OVERRIDE')):
            affected_pos.add(str(ln.get('po') or ''))
    n_clean_po = 0
    for h in (res.get('headers') or []):
        h['clean'] = str(h.get('po') or '') not in affected_pos
        if h['clean']:
            n_clean_po += 1
    res['n_clean_po'] = n_clean_po
    res['n_affected_po'] = len(res.get('headers') or []) - n_clean_po
    return render(request, 'online_b2b/review.html',
                  {'token': token, 'meta': meta, 'r': res,
                   'has_preview': has_preview,
                   'is_draft': bool(meta.get('draft')),
                   'locked': bool(meta.get('locked')),
                   'run_id': meta.get('run_id'),
                   'exc_count': exc_count, 'nim_lines': nim_lines,
                   'auto_fixed': auto_fixed,
                   'affected_qty': affected_qty, 'ok_qty_pct': ok_qty_pct,
                   'margin': meta.get('margin_pct')})


def _collect_drafts() -> list[dict]:
    """All parked 'Review Later' runs as API-ready dicts (token, marketplace,
    when, note, PO count, undecided-affected count, file count). Read-only; the
    fat data layer behind :class:`DraftsView` and the JSON endpoint."""
    rows: list[dict] = []
    if not _UPLOADS.exists():
        return rows
    for d in _UPLOADS.iterdir():
        if not d.is_dir():
            continue
        mp = d / 'meta.json'
        if not mp.exists():
            continue
        try:
            meta = json.loads(mp.read_text(encoding='utf-8'))
        except Exception:  # noqa: BLE001
            continue
        if not meta.get('draft') or meta.get('locked'):
            continue
        npos = undecided = 0
        cache = d / 'preview.json'
        if cache.exists():
            try:
                res = (json.loads(cache.read_text(encoding='utf-8'))
                       .get('res') or {})
                npos = len(res.get('headers') or [])
                dec = meta.get('decisions') or {}
                for ln in (res.get('affected') or []):
                    k = (f"{ln.get('po', '')}|{ln.get('item_no', '')}"
                         f"|{ln.get('ean', '')}")
                    if not (dec.get(k) or {}).get('action'):
                        undecided += 1
            except Exception:  # noqa: BLE001
                pass
        rows.append({
            'token': d.name,
            'marketplace': meta.get('marketplace', ''),
            'draft_at': meta.get('draft_at', ''),
            'note': meta.get('draft_note', ''),
            'pos': npos, 'undecided': undecided,
            'files': len(meta.get('files') or []),
        })
    rows.sort(key=lambda r: r['draft_at'], reverse=True)
    return rows


class SaveReviewLaterView(LoginRequiredMixin, View):
    """Park the WHOLE run as a 'Review Later' draft — kept intact (raw file +
    parsed result), NOT locked/recorded. Use when a CP issue can't be decided yet
    (needs the team to correct the master). The operator later reopens it from
    Drafts, re-validates (picks up the correction), and finalizes — never
    re-uploaded. API: returns ``{ok, redirect}`` / ``{ok:false, error}``."""

    def post(self, request, token):
        import datetime as _dt
        meta, d = _load_meta(token)
        if not meta:
            return JsonResponse({'ok': False, 'error': 'Upload not found or expired.'},
                                status=404)
        if meta.get('locked'):
            return JsonResponse({'ok': False,
                                 'error': 'Already recorded — nothing to defer.'})
        meta['draft'] = True
        meta['draft_at'] = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        meta['draft_note'] = (request.POST.get('note') or '')[:300]
        _save_meta(d, meta)
        # Parking a run = an unresolved CP issue → auto-HOLD that channel on
        # today's Daily Tasks so it's not chased as pending until it's resolved
        # (finalizing the run un-holds it). The draft note rides along as the
        # hold reason. Best-effort — never blocks the park.
        try:
            from .services import daily_checklist as dc
            from .services import marketplaces as reg
            ch = reg.db_key_to_channel().get(str(meta.get('marketplace')))
            if ch:
                dc.toggle(None, ch, 'hold', True,
                          getattr(request.user, 'username', '') or 'system',
                          remark=meta.get('draft_note', '') or 'Parked for review')
        except Exception:  # noqa: BLE001
            pass
        # NOTE: the CP-issue email is sent MANUALLY (a button on the review page),
        # NOT automatically on park — see `draft_email_send`.
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'redirect': '/b2b/drafts/'})
        return redirect('b2b_drafts')


@login_required
@require_POST
def draft_email_send(request, token):
    """Send (or re-send) a run's **CP-issue email** to the stakeholders — the
    ecom team sees which item has which issue + affected qty. Triggered manually
    from the Review page or the Review-Later list (parking also auto-sends once).
    Works for any run with flagged lines, parked or not."""
    meta, d = _load_meta(token)
    if not meta:
        return JsonResponse({'ok': False, 'error': 'Run not found or expired.'}, status=404)
    try:
        res = _cached_preview(token, d, meta)
    except Exception as e:  # noqa: BLE001
        return JsonResponse({'ok': False, 'error': f'Could not load the run: {e}'},
                            status=400)
    affected = res.get('affected') or []
    if not affected:
        return JsonResponse({'ok': False,
                             'error': 'No flagged lines on this run — nothing to email.'})
    # Same KPI row as the review page (POs/lines/qty/value/affected/affected-qty/OK-qty).
    summary = res.get('summary') or {}
    _AFF = {'MISMATCH', 'NOT_IN_MASTER'}
    _tq = _aq = 0
    for _ln in (res.get('lines') or []):
        try:
            _q = int(float(_ln.get('qty') or 0))
        except (TypeError, ValueError):
            _q = 0
        _tq += _q
        if (_ln.get('status') or '') in _AFF:
            _aq += _q
    kpis = {'pos': summary.get('pos'), 'lines': summary.get('lines'),
            'qty': summary.get('qty'), 'value': summary.get('value'),
            'affected': summary.get('affected'), 'affected_qty': _aq,
            'ok_qty_pct': round((_tq - _aq) * 100 / _tq, 1) if _tq else 100.0}
    from .services.issue_email import ReviewLaterEmailReport
    note, to, cc = _email_extras(request)
    rep = ReviewLaterEmailReport(
        marketplace=meta.get('marketplace', ''), affected=affected,
        note=note or meta.get('draft_note', ''), draft_at=meta.get('draft_at', ''),
        to=to, cc=cc, kpis=kpis)
    if to is not None and not rep._to:
        return JsonResponse({'ok': False, 'error': 'No valid recipient in "To".'})
    ok, reason = rep.send()
    return JsonResponse({'ok': ok, 'error': None if ok else reason,
                         'to': rep.recipients()['to'], 'lines': len(affected)})


class DraftsView(LoginRequiredMixin, TemplateView):
    """'Review Later' list — parked runs the operator reopens (instead of
    re-uploading) once the team corrects the CP. Each row deep-links to its review
    with ?revalidate=1 so it re-checks against the current master. Dual-render:
    ``?format=json`` (or an AJAX request) returns ``{ok, data}``."""
    template_name = 'online_b2b/drafts.html'

    def get(self, request, *args, **kwargs):
        drafts = _collect_drafts()
        if (request.GET.get('format') == 'json'
                or request.headers.get('x-requested-with') == 'XMLHttpRequest'):
            return JsonResponse({'ok': True, 'data': drafts})
        return self.render_to_response({'drafts': drafts})


@login_required
@require_POST
def confirm(request, token):
    """Phase 3: re-process and PUSH headers + lines to MySQL.

    Answers JSON when called via fetch (``X-Requested-With``) so the review page
    can lock in place with a progress bar — no full reload. Falls back to the
    classic redirect flow when JS is off."""
    ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    meta, d = _load_meta(token)
    if not meta:
        raise Http404("Upload not found or expired.")
    paths = [str(d / n) for n in meta['files']]

    # Per-affected-line operator decision (Action + Override CP + Remark),
    # keyed po|item|ean. Action ∈ INCLUDE / OVERRIDE / EXCLUDE.
    keys = request.POST.getlist('aff_key')
    acts = request.POST.getlist('aff_action')
    ocps = request.POST.getlist('aff_override_cp')
    rems = request.POST.getlist('aff_remark')
    actions = {}
    for i, k in enumerate(keys):
        a = (acts[i] if i < len(acts) else '').strip()
        ocp = (ocps[i] if i < len(ocps) else '').strip()
        r = (rems[i] if i < len(rems) else '').strip()
        if k and (a or r):
            actions[k] = {'action': a, 'remark': r, 'override_cp': ocp}

    # A finalized "Review-later" draft belongs to the day it was PARKED, not
    # today — back-date the whole recorded run (run_ts + created_at on headers &
    # lines) to draft_at so Daily Tasks / analytics credit the park day. Normal
    # runs pass as_of=None → stamped now. Bad/absent draft_at → falls back to now.
    as_of = None
    if meta.get('draft') and meta.get('draft_at'):
        import datetime as _dt
        try:
            as_of = _dt.datetime.strptime(meta['draft_at'][:19], '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            as_of = None

    res = engine_bridge.confirm(
        meta['marketplace'], paths, warehouse=meta['warehouse'],
        margin_pct=meta['margin_pct'] / 100.0, actions=actions,
        ean_fixes=meta.get('ean_fixes'), as_of=as_of)

    if not res.get('ok'):
        err = res.get('error', 'Push failed.')
        if ajax:
            return JsonResponse({'ok': False, 'error': err}, status=400)
        messages.error(request, err)
        return redirect('b2b_review', token=token)

    run_id = res.get('run_id')
    if run_id is None:
        msg = "All POs were already uploaded — nothing new to push."
        if ajax:
            from django.urls import reverse
            return JsonResponse({'ok': True, 'run_id': None, 'message': msg,
                                 'redirect': reverse('b2b_dashboard')})
        messages.info(request, msg)
        return redirect('b2b_dashboard')

    # Build the D365 dump ONCE at lock time and keep it next to the run sidecar
    # (web-owned, survives upload-temp cleanup) so it's retrievable later from
    # Orders → Run #N even if the review tab is closed. Never blocks the lock.
    d365_path = _persist_d365(
        run_id, meta['marketplace'], paths,
        meta['warehouse'], meta['margin_pct'] / 100.0, actions,
        ean_fixes=meta.get('ean_fixes'))
    _save_run_index(run_id, {
        'output_path': res.get('output_path'),
        'marketplace': meta['marketplace'],
        'summary': res.get('summary'),
        'warnings': res.get('warnings'),
        'd365_path': d365_path,
    })
    # Lock the decisions on the token so the review page now offers Generate D365.
    was_draft = bool(meta.get('draft'))
    meta['decisions'] = actions
    meta['run_id'] = run_id
    meta['locked'] = True
    (d / 'meta.json').write_text(json.dumps(meta), encoding='utf-8')
    # A parked ("Review Later") run just got finalized → the CP issue is resolved,
    # so auto-UN-HOLD that channel on today's Daily Tasks (mirrors the auto-hold on
    # park). Best-effort. (Uploaded-web will also auto-tick from the new record.)
    if was_draft:
        try:
            from .services import daily_checklist as dc
            from .services import marketplaces as reg
            ch = reg.db_key_to_channel().get(str(meta.get('marketplace')))
            if ch:
                # The hold was placed on the PARK day, so un-hold THAT day (not
                # today) — matches the back-dated record. Fall back to today.
                hold_day = as_of.date().isoformat() if as_of else None
                dc.toggle(hold_day, ch, 'hold', False,
                          getattr(request.user, 'username', '') or 'system')
        except Exception:  # noqa: BLE001
            pass
    pos = res['summary']['pos']
    lines = res.get('lines_recorded', 0)
    if ajax:
        from django.urls import reverse
        return JsonResponse({
            'ok': True, 'run_id': run_id, 'pos': pos, 'lines': lines,
            'has_d365': bool(d365_path),
            'run_url': reverse('b2b_run_detail', args=[run_id]),
            'd365_url': reverse('b2b_generate_d365', args=[token]),
            'message': f"Locked & recorded {pos} PO(s), {lines} line(s).",
        })
    messages.success(
        request, f"🔒 Locked & recorded {pos} PO(s), {lines} line(s). "
        "Now generate the D365 dump.")
    return redirect('b2b_review', token=token)


@login_required
@require_POST
def generate_d365(request, token):
    """Build + download the ERP D365 dump from the LOCKED decisions (Excludes
    dropped, Overrides repriced). Engine + full SO Workbook untouched."""
    meta, d = _load_meta(token)
    if not meta:
        raise Http404("Upload not found or expired.")
    if not meta.get('locked'):
        messages.error(request, "Lock the decisions first, then generate D365.")
        return redirect('b2b_review', token=token)
    paths = [str(d / n) for n in meta['files']]
    out_path = d / f"{str(meta['marketplace']).lower()}_d365_decided.xlsx"
    res = engine_bridge.generate_d365(
        meta['marketplace'], paths, str(out_path),
        warehouse=meta['warehouse'], margin_pct=meta['margin_pct'] / 100.0,
        actions=meta.get('decisions') or {}, ean_fixes=meta.get('ean_fixes'))
    if not res.get('ok') or not os.path.exists(res.get('d365_path', '')):
        messages.error(request, res.get('error', 'D365 generation failed.'))
        return redirect('b2b_review', token=token)
    return FileResponse(
        open(res['d365_path'], 'rb'), as_attachment=True,
        filename=f"{meta['marketplace']}_D365_import.xlsx")


@login_required
@require_POST
def save_decision(request, token):
    """Persist ONE affected line's decision (Action / Override CP / Remark) to
    the upload session as the operator sets it — so decisions survive an EAN
    re-validate and the final lock just commits what's saved. Returns JSON.

    The whole read-modify-write is serialised per token (``_meta_lock``) +
    written atomically, so the several near-simultaneous POSTs from
    Apply-to-selected neither clobber each other nor 404 on a half-written file."""
    key = (request.POST.get('key') or '').strip()
    if not key:
        return JsonResponse({'ok': False, 'error': 'no key'})
    action = (request.POST.get('action') or '').strip()
    ocp = (request.POST.get('override_cp') or '').strip()
    remark = (request.POST.get('remark') or '').strip()
    with _meta_lock(token):
        meta, d = _load_meta(token)
        if not meta:
            return JsonResponse({'ok': False, 'error': 'expired'}, status=404)
        if meta.get('locked'):
            return JsonResponse({'ok': False, 'error': 'locked'})
        decisions = dict(meta.get('decisions') or {})
        if action or remark or ocp:
            decisions[key] = {'action': action, 'remark': remark, 'override_cp': ocp}
        else:
            decisions.pop(key, None)
        meta['decisions'] = decisions
        _save_meta(d, meta)
    return JsonResponse({'ok': True, 'saved': len(decisions)})


@login_required
@require_POST
def fix_ean(request, token):
    """Correct the EAN on NOT_IN_MASTER lines: validate each 'correct EAN'
    against the DB master, stash the fixes on the upload session, and re-run the
    review so they resolve. The wrong EAN is kept (becomes ``received_ean`` at
    lock). Persistent: a repeat wrong EAN auto-resolves on future POs."""
    meta, d = _load_meta(token)
    if not meta:
        raise Http404("Upload not found or expired.")
    if meta.get('locked'):
        messages.error(request, "Already locked — EANs can't be changed.")
        return redirect('b2b_review', token=token)
    from .services import item_master_loader as iml
    wrongs = request.POST.getlist('nim_ean')
    fixes = request.POST.getlist('nim_fix')
    ean_fixes = dict(meta.get('ean_fixes') or {})
    applied, errors = 0, []
    for i, w in enumerate(wrongs):
        w = (w or '').strip()
        c = (fixes[i] if i < len(fixes) else '').strip()
        if not w or not c:
            continue
        hit = iml.resolve_in_master(c)
        if not hit:
            errors.append(f"'{c}' is not in the item master — can't use it to "
                          f"correct EAN {w}.")
            continue
        ean_fixes[w] = hit['ean'] or hit['item_no']
        applied += 1
    if applied:
        meta['ean_fixes'] = ean_fixes
        (d / 'meta.json').write_text(json.dumps(meta), encoding='utf-8')
        messages.success(request, f"✓ Corrected {applied} EAN(s) — re-validating "
                         "against the master.")
    for e in errors:
        messages.error(request, e)
    return redirect('b2b_review', token=token)


@login_required
@require_POST
def discard(request, token):
    d = _token_dir(token)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
    messages.info(request, "Upload discarded.")
    return redirect('b2b_dashboard')


def _full_name(name: str) -> str:
    """Clear download name for the REVIEW workbook (every line, all sheets) so
    it's never confused with the post-lock **Completed** workbook (accepted-only)
    or the headers-only ``*_d365.xlsx`` import package."""
    for s in ('_so_', '_to_'):
        if s in name:
            return name.replace(s, '_Review_')
    return 'Review_' + name


def _full_workbook(outdir: Path):
    """The full data workbook in a folder = the newest .xlsx that is NOT the
    ``*_d365.xlsx`` D365 import sibling."""
    if not outdir.exists():
        return None
    files = sorted([p for p in outdir.glob('*.xlsx')
                    if not p.stem.endswith('_d365')],
                   key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def _count_pos(path) -> int:
    """POs in a workbook = data rows of its 'Headers (SO)' / 'Headers (TO)'
    sheet(s) (one row per document)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        n = 0
        for sh in ('Headers (SO)', 'Headers (TO)'):
            if sh in wb.sheetnames:
                n += sum(1 for r in wb[sh].iter_rows(min_row=2, values_only=True)
                         if r and r[0] not in (None, ''))
        wb.close()
        return n
    except Exception:  # noqa: BLE001 — naming is best-effort, never block a download
        return 0


def _lot_name(marketplace: str, path, kind: str) -> str:
    """Self-describing download name: ``{Mp}_{N}po_{dd-mm-YYYY_HHMMSS}_{kind}`` so
    the lot size + run timestamp are obvious and Review vs Completed never clash."""
    import re as _re
    import time as _time
    name = os.path.basename(str(path))
    m = _re.search(r'(\d{2}-\d{2}-\d{4}_\d{6})', name)
    ts = m.group(1) if m else _time.strftime(
        '%d-%m-%Y_%H%M%S', _time.localtime(os.path.getmtime(path)))
    mp = str(marketplace or 'SO').replace(' ', '')
    return f"{mp}_{_count_pos(path)}po_{ts}_{kind}.xlsx"


@login_required
def review_download(request, token):
    """Download the FULL preview/review workbook — every line (Summary /
    Validation / Raw Data / Headers / Lines). NOT the *_d365.xlsx package, and
    NOT decision-filtered (use review_download_completed for accepted-only)."""
    meta, d = _load_meta(token)
    if not meta:
        raise Http404("Upload not found or expired.")
    f = _full_workbook(d / 'output')
    if not f:
        raise Http404("Preview workbook not found.")
    return FileResponse(open(f, 'rb'), as_attachment=True,
                        filename=_lot_name(meta.get('marketplace', ''), f, 'review'))


def _completed_cache_key(meta: dict) -> str:
    """Signature for the LOCKED completed workbook — covers every input to
    ``export_decided_workbook``. Post-lock the run is frozen, so a matching key
    guarantees byte-identical output → safe to re-serve instead of rebuilding."""
    basis = {'files': meta.get('files', []), 'wh': meta.get('warehouse', ''),
             'margin': meta.get('margin_pct', ''), 'dec': meta.get('decisions') or {},
             'ean': meta.get('ean_fixes'), 'run': meta.get('run_id'),
             'locked': bool(meta.get('locked')),
             # bump when the exporter output changes so stale caches rebuild.
             # v2: Summary Included/Excluded now reads the real line action
             # (INCLUDE/OVERRIDE no longer shown as dropped).
             'exp_v': 2}
    return hashlib.md5(
        json.dumps(basis, sort_keys=True, default=str).encode('utf-8')).hexdigest()


def _completed_cache_get(d: Path, meta: dict):
    """Cached completed-workbook path if the marker matches the current signature
    and the file exists; else None. Fully guarded — never raises, so a bad cache
    simply falls through to a normal rebuild."""
    try:
        marker = d / 'completed.cache.json'
        if not marker.exists():
            return None
        blob = json.loads(marker.read_text(encoding='utf-8'))
        if blob.get('key') != _completed_cache_key(meta):
            return None
        p = d / blob.get('name', '')
        return str(p) if p.exists() else None
    except Exception:  # noqa: BLE001 — cache is best-effort, never block a download
        return None


def _completed_cache_put(d: Path, meta: dict, built_path: str) -> None:
    """Copy the freshly built completed workbook into the token dir + write the
    signature marker. Guarded — a failure just means no caching, never blocks."""
    try:
        import shutil
        dest = d / os.path.basename(built_path)
        if os.path.abspath(built_path) != os.path.abspath(str(dest)):
            shutil.copy2(built_path, dest)
        (d / 'completed.cache.json').write_text(
            json.dumps({'key': _completed_cache_key(meta), 'name': dest.name}),
            encoding='utf-8')
    except Exception:  # noqa: BLE001
        pass


@login_required
def review_download_completed(request, token):
    """Download the SO Workbook with LOCKED decisions applied — accepted lines
    only, Overrides repriced (post-lock companion to review_download). Same full
    multi-sheet workbook, just the decided/accepted set."""
    meta, d = _load_meta(token)
    if not meta:
        raise Http404("Upload not found or expired.")
    if not meta.get('locked'):
        messages.error(request, "Lock & Record first, then download the completed workbook.")
        return redirect('b2b_review', token=token)
    # Post-lock the run is frozen, so the completed workbook is deterministic.
    # Serve a cached copy on repeat downloads instead of regenerating (this was
    # the slowest endpoint). Pure view-level cache — export_decided_workbook is
    # UNCHANGED, and any cache miss/failure falls straight through to a rebuild.
    _cached = _completed_cache_get(d, meta)
    if _cached is None:
        paths = [str(d / n) for n in meta['files']]
        res = engine_bridge.export_decided_workbook(
            meta['marketplace'], paths,
            warehouse=meta['warehouse'], margin_pct=meta['margin_pct'] / 100.0,
            actions=meta.get('decisions') or {}, ean_fixes=meta.get('ean_fixes'),
            # Drop POs already uploaded in an EARLIER run from the import file so
            # D365 doesn't get duplicate SOs (this run's new POs stay). The Review
            # download is unaffected. [[completed-dedup]]
            exclude_uploaded_run_id=meta.get('run_id'))
        if not res.get('ok') or not os.path.exists(res.get('path', '')):
            messages.error(request, res.get('error', 'Completed workbook generation failed.'))
            return redirect('b2b_review', token=token)
        _cached = res['path']
        _completed_cache_put(d, meta, _cached)
    # Downloading the Completed workbook is a real workflow milestone → auto-tick
    # the Daily Tasks "Workbook downloaded" step for this channel. A parked
    # ("Review Later") run credits its PARK day (draft_at) — the record,
    # uploaded-web tick and un-hold are all back-dated there, so this step must
    # land there too, not on the day the CP issue was finally resolved. Normal
    # runs → today. Never blocks the download (the helper swallows its errors).
    from .services import daily_checklist as _dc
    _wb_day = meta['draft_at'][:10] if (meta.get('draft') and meta.get('draft_at')) else None
    _dc.mark_workbook_downloaded(meta['marketplace'],
                                 user=request.user.get_username(), day=_wb_day)
    return FileResponse(open(_cached, 'rb'), as_attachment=True,
                        filename=_lot_name(meta['marketplace'], _cached, 'completed'))


# ── Bulk import (ERP Sales Orders) ──────────────────────────────────────

def _erp_file(token: str):
    base = _ERP.resolve()
    d = (_ERP / token).resolve()
    if d != base and base not in d.parents:
        raise Http404()
    nm = d / 'name.txt'
    if not nm.exists():
        return None
    return d / nm.read_text(encoding='utf-8').strip()


@login_required
def bulk_upload(request):
    """Upload an ERP 'Sales Orders' header export → stash → go to review."""
    if request.method == 'POST' and request.FILES.get('erp_file'):
        token = uuid.uuid4().hex[:12]
        up = _ERP / token
        up.mkdir(parents=True, exist_ok=True)
        f = request.FILES['erp_file']
        dest = up / Path(f.name).name
        with open(dest, 'wb') as out:
            for chunk in f.chunks():
                out.write(chunk)
        (up / 'name.txt').write_text(dest.name, encoding='utf-8')
        return redirect('b2b_bulk_review', token=token)
    return render(request, 'online_b2b/bulk_upload.html')


@login_required
def bulk_review(request, token):
    """Preview the parsed ERP rows (new vs already-imported) — no DB write."""
    fp = _erp_file(token)
    if not fp or not fp.exists():
        raise Http404("Upload not found or expired.")
    res = erp_import.preview(str(fp))
    return render(request, 'online_b2b/bulk_review.html',
                  {'token': token, 'r': res})


@login_required
@require_POST
def bulk_confirm(request, token):
    """Import the NEW ERP rows into the order DB (segment=Offline)."""
    fp = _erp_file(token)
    if not fp or not fp.exists():
        raise Http404("Upload not found or expired.")
    res = erp_import.do_import(str(fp))
    if not res.get('ok'):
        messages.error(request, res.get('error', 'Import failed.'))
        return redirect('b2b_bulk_review', token=token)
    if not res.get('imported'):
        messages.info(request, "All orders were already imported — nothing new.")
        return redirect('b2b_orders')
    messages.success(
        request, f"Imported {res['imported']} order(s) "
        f"({res.get('skipped', 0)} already present).")
    return redirect('b2b_orders')


@login_required
def run_detail(request, run_id):
    data = order_db.run_detail(int(run_id))
    idx = _load_run_index(run_id)
    has_file = bool(idx.get('output_path') and os.path.exists(idx['output_path']))
    has_d365 = bool(idx.get('d365_path') and os.path.exists(idx['d365_path']))
    return render(request, 'online_b2b/run_detail.html', {
        'run_id': run_id, 'd': data, 'idx': idx, 'has_file': has_file,
        'has_d365': has_d365,
    })


@login_required
def download_d365(request, run_id):
    """Re-download the decided D365 dump saved for this run at lock time
    (Excludes dropped, Overrides repriced). A derived static file — serving it
    reads nothing from and writes nothing to the DB."""
    idx = _load_run_index(run_id)
    path = idx.get('d365_path')
    if not path or not os.path.exists(path):
        raise Http404("No D365 dump saved for this run.")
    mkt = idx.get('marketplace', 'D365')
    return FileResponse(open(path, 'rb'), as_attachment=True,
                        filename=f"{mkt}_D365_import.xlsx")


@login_required
def download(request, run_id):
    """Download the FULL SO workbook for a run (all sheets) — the engine's main
    output, not the headers-only *_d365.xlsx package."""
    idx = _load_run_index(run_id)
    path = idx.get('output_path')
    if not path or not os.path.exists(path):
        raise Http404("Output file not found for this run.")
    return FileResponse(open(path, 'rb'), as_attachment=True,
                        filename=_full_name(os.path.basename(path)))


def _delete_run_files(run_id) -> list:
    """Remove the web-owned file sidecars of a run (SO workbook, D365 dump, and
    the run-index json). Returns the list of paths actually removed. Best-effort:
    a missing/locked file never blocks the DB delete."""
    removed = []
    idx = _load_run_index(run_id)
    for key in ('output_path', 'd365_path'):
        p = idx.get(key)
        if p and os.path.exists(p):
            try:
                os.remove(p); removed.append(p)
            except OSError:
                pass
    for sc in (_RUNS_INDEX / f"{run_id}.json", _RUNS_INDEX / f"{run_id}_d365.xlsx"):
        if sc.exists():
            try:
                sc.unlink(); removed.append(str(sc))
            except OSError:
                pass
    return removed


@login_required
@require_POST
def run_delete(request, run_id):
    """HARD-DELETE a whole run — its runs row, order_headers, order_lines +
    validation, and the file sidecars. Destructive & irreversible; POST-only and
    guarded by a typed confirmation on the run page. No other run is touched."""
    before = order_db.run_summary(run_id)
    res = order_db.delete_run(run_id)
    if not res.get('ok'):
        messages.error(request, f"Could not delete run {run_id}: "
                                f"{res.get('error', 'unknown error')}")
        return redirect('b2b_run_detail', run_id=run_id)
    files = _delete_run_files(run_id)
    mp = (before.get('marketplaces') or '—') if before.get('ok') else '—'
    messages.success(
        request,
        f"🗑 Run {run_id} deleted — {mp}: {res['headers']} PO header(s), "
        f"{res['lines']} line(s), {res['validation']} validation row(s), "
        f"{len(files)} file(s) removed.")
    return redirect('b2b_dashboard')


# ── Item Master (DB) ────────────────────────────────────────────────────────
# Upload the two ERP exports (Items + Item M.R.P.) → preview → rebuild the
# `item_master` table. The engine reads this DB master via DBMasterLoader — no
# manual Excel. The page also browses the live master (overview + search).

_IM_UPLOADS = _MEDIA / 'b2b_item_master'


def _im_token_dir(token: str) -> Path:
    base = _IM_UPLOADS.resolve()
    d = (_IM_UPLOADS / token).resolve()
    if d != base and base not in d.parents:
        raise Http404()
    return d


class ItemMasterView(LoginRequiredMixin, TemplateView):
    """`/b2b/item-master/` — status + upload/refresh + browsable overview."""
    template_name = 'online_b2b/item_master.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import item_master_loader as iml
        ctx['status'] = iml.status()
        ctx['overview'] = iml.list_items(self.request.GET.get('q', ''), limit=100)
        return ctx


@login_required
@require_POST
def item_master_upload(request):
    """Stash the two uploaded files under a token → preview. No DB write."""
    items_f = request.FILES.get('items_file')
    mrp_f = request.FILES.get('mrp_file')
    if not items_f or not mrp_f:
        messages.error(request, "Choose BOTH files — the Items export and the "
                                "Item M.R.P. export.")
        return redirect('b2b_item_master')
    token = uuid.uuid4().hex[:12]
    d = _IM_UPLOADS / token
    d.mkdir(parents=True, exist_ok=True)
    items_path = d / ('items' + Path(items_f.name).suffix)
    mrp_path = d / ('mrp' + Path(mrp_f.name).suffix)
    for f, dest in ((items_f, items_path), (mrp_f, mrp_path)):
        with open(dest, 'wb') as out:
            for chunk in f.chunks():
                out.write(chunk)
    (d / 'meta.json').write_text(json.dumps({
        'items_name': items_f.name, 'mrp_name': mrp_f.name,
        'items_path': items_path.name, 'mrp_path': mrp_path.name,
    }), encoding='utf-8')
    return redirect('b2b_item_master_preview', token=token)


@login_required
def item_master_preview(request, token):
    """Compute the master from the two files in-memory (no write) and show the
    stats / warnings / sample for the operator to confirm."""
    d = _im_token_dir(token)
    mp = d / 'meta.json'
    if not mp.exists():
        raise Http404("Upload not found or expired.")
    meta = json.loads(mp.read_text(encoding='utf-8'))
    from .services import item_master_loader as iml
    try:
        rows, stats, warnings = iml.build_rows(
            str(d / meta['items_path']), str(d / meta['mrp_path']))
    except Exception as e:  # noqa: BLE001
        messages.error(request, f"Could not read the files: {type(e).__name__}: {e}")
        return redirect('b2b_item_master')
    return render(request, 'online_b2b/item_master_preview.html', {
        'token': token, 'meta': meta, 'stats': stats, 'warnings': warnings,
        'sample': rows[:15], 'current': iml.status(),
        'diff': iml.diff_against_current(rows),
    })


@login_required
@require_POST
def item_master_confirm(request, token):
    """Rebuild item_master from the two files in one transaction (full replace).
    Seeds the durable Swiggy map once if empty. No engine change, no order data
    touched."""
    d = _im_token_dir(token)
    mp = d / 'meta.json'
    if not mp.exists():
        raise Http404("Upload not found or expired.")
    meta = json.loads(mp.read_text(encoding='utf-8'))
    from .services import item_master_loader as iml
    try:
        if not iml.load_swiggy_map():
            from online_po_processor.config.paths import get_bundled_master_path
            bm = get_bundled_master_path()
            if bm:
                iml.seed_swiggy_from_excel(str(bm))
        rows, stats, _ = iml.build_rows(
            str(d / meta['items_path']), str(d / meta['mrp_path']))
        res = iml.replace_item_master(rows)
    except Exception as e:  # noqa: BLE001
        messages.error(request, f"Rebuild failed: {type(e).__name__}: {e}")
        return redirect('b2b_item_master_preview', token=token)
    shutil.rmtree(d, ignore_errors=True)
    messages.success(
        request, f"✓ Item master rebuilt — {res['rows']} items now live "
        f"({stats['swiggy_mapped']} Swiggy-mapped · batch {res['batch_id']}).")
    return redirect('b2b_item_master')


@login_required
@require_POST
def item_master_discard(request, token):
    d = _im_token_dir(token)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
    messages.info(request, "Upload discarded.")
    return redirect('b2b_item_master')


def _im_row_json(r: dict) -> dict:
    """JSON-safe item_master row (dates → str, Decimal → float)."""
    def s(v):
        return '' if v is None else str(v)
    return {
        'item_no': s(r.get('item_no')), 'ean': s(r.get('ean')),
        'description': s(r.get('description')), 'gst_code': s(r.get('gst_code')),
        'hsn': s(r.get('hsn')),
        'mrp': None if r.get('mrp') is None else float(r['mrp']),
        'mrp_start': s(r.get('mrp_start')), 'mrp_end': s(r.get('mrp_end')),
    }


@login_required
def item_master_search(request):
    """As-you-type search over item_master → JSON (item no / EAN / description /
    Swiggy). Read-only."""
    from .services import item_master_loader as iml
    res = iml.list_items(request.GET.get('q', ''), limit=100)
    return JsonResponse({
        'total': res['total'], 'shown': res['shown'], 'q': res['q'],
        'rows': [_im_row_json(r) for r in res['rows']],
    })


@login_required
def item_master_export(request):
    """Download the item master as .xlsx — ALL rows (no cap), honouring the same
    `?q=` search as the on-page browser. Read-only."""
    import datetime as _dt
    import io as _io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from .services import item_master_loader as iml
    q = (request.GET.get('q', '') or '').strip()
    rows = iml.export_rows(q)
    cols = [('item_no', 'Item No'), ('ean', 'EAN'),
            ('description', 'Description'), ('mrp', 'MRP'),
            ('gst_code', 'GST Code'), ('hsn', 'HSN'),
            ('mrp_start', 'MRP Start'), ('mrp_end', 'MRP End')]
    wb = Workbook(); ws = wb.active; ws.title = 'Item Master'
    hf = Font(bold=True, color='FFFFFF'); navy = PatternFill('solid', fgColor='1A237E')
    for c, (_k, h) in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.font = hf; cell.fill = navy
        cell.alignment = Alignment(horizontal='center')
    for r, row in enumerate(rows, 2):
        for c, (k, _h) in enumerate(cols, 1):
            v = row.get(k)
            ws.cell(r, c, str(v) if k in ('mrp_start', 'mrp_end') and v is not None else v)
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        L = col[0].column_letter
        w = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[L].width = min(w + 2, 60)
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    stamp = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    scope = 'filtered' if q else 'all'
    fname = f"item_master_{scope}_{len(rows)}items_{stamp}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


# ── GT Select (Offline parent — D365-finalised; import headers + lines) ─────

_GTS_UPLOADS = _MEDIA / 'b2b_gt_select'


def _gts_token_dir(token: str) -> Path:
    base = _GTS_UPLOADS.resolve()
    d = (_GTS_UPLOADS / token).resolve()
    if d != base and base not in d.parents:
        raise Http404()
    return d


class OfflineProcessView(LoginRequiredMixin, TemplateView):
    """Offline 'Process PO' — one entry that lists the Offline **parent**
    marketplaces (and MT's **children**), each routing to that channel's
    upload/import. Mirrors the online Process-PO marketplace picker so the two
    segments feel the same; replaces the scattered per-channel links."""
    template_name = 'online_b2b/offline_process.html'

    def get_context_data(self, **kwargs):
        from django.urls import reverse
        ctx = super().get_context_data(**kwargs)
        ctx['parents'] = [
            {'key': 'MT', 'name': 'MT · Modern Trade', 'tag': 'live',
             'desc': 'Retail chains — master-driven SO generation. Pick a child.',
             'url': reverse('mt_flow_upload'),
             'children': [
                 {'key': 'SS', 'name': 'Shoppers Stop (SS)', 'tag': 'live',
                  'url': reverse('mt_flow_upload')},
                 {'key': 'HG', 'name': 'Health & Glow (HG)', 'tag': 'live',
                  'url': reverse('mt_flow_upload')},
                 {'key': 'NT', 'name': 'Naturals (NT)', 'tag': 'live',
                  'url': reverse('mt_flow_upload')},
                 {'key': 'LL', 'name': 'Lulu (LL)', 'tag': 'live',
                  'url': reverse('mt_flow_upload')},
                 {'key': 'BN', 'name': 'Apollo (BN)', 'tag': 'live',
                  'url': reverse('mt_flow_upload')},
                 {'key': 'HB', 'name': 'HB', 'tag': 'soon'},
             ]},
            {'key': 'GT Mass', 'name': 'GT Mass', 'tag': 'live',
             'url': reverse('index'), 'desc': 'General trade mass — dump processing.',
             'children': []},
            {'key': 'GT Select', 'name': 'GT Select', 'tag': 'live',
             'url': reverse('b2b_gt_select'),
             'desc': 'D365-finalised — import Sales Orders + Lines.', 'children': []},
            {'key': 'EKA', 'name': 'EKA', 'tag': 'soon', 'desc': 'Coming soon.',
             'children': []},
            {'key': 'CSD', 'name': 'CSD', 'tag': 'soon', 'desc': 'Coming soon.',
             'children': []},
        ]
        return ctx


class GtSelectView(LoginRequiredMixin, TemplateView):
    """`/b2b/gt-select/` — upload the D365 Sales Orders + Sales Lines exports."""
    template_name = 'online_b2b/gt_select.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['kpis'] = order_db.segment_kpis('Offline')
        return ctx


@login_required
@require_POST
def gt_select_upload(request):
    """Stash the two D365 exports under a token → preview."""
    hf = request.FILES.get('headers_file')
    lf = request.FILES.get('lines_file')
    if not hf or not lf:
        messages.error(request, "Choose BOTH files — Sales Orders (headers) and "
                                "Sales Lines.")
        return redirect('b2b_gt_select')
    token = uuid.uuid4().hex[:12]
    d = _GTS_UPLOADS / token
    d.mkdir(parents=True, exist_ok=True)
    hp = d / ('headers' + Path(hf.name).suffix)
    lp = d / ('lines' + Path(lf.name).suffix)
    for f, dest in ((hf, hp), (lf, lp)):
        with open(dest, 'wb') as out:
            for chunk in f.chunks():
                out.write(chunk)
    (d / 'meta.json').write_text(json.dumps({
        'headers_name': hf.name, 'lines_name': lf.name,
        'headers_path': hp.name, 'lines_path': lp.name,
    }), encoding='utf-8')
    return redirect('b2b_gt_select_preview', token=token)


@login_required
def gt_select_preview(request, token):
    """Parse + join + dedup (no DB write) for the review page."""
    d = _gts_token_dir(token)
    mp = d / 'meta.json'
    if not mp.exists():
        raise Http404("Upload not found or expired.")
    meta = json.loads(mp.read_text(encoding='utf-8'))
    from .services import gt_select_import as gts
    pv = gts.preview(str(d / meta['headers_path']), str(d / meta['lines_path']))
    if not pv.get('ok'):
        messages.error(request, pv.get('error', 'Could not read the files.'))
        return redirect('b2b_gt_select')
    return render(request, 'online_b2b/gt_select_preview.html', {
        'token': token, 'meta': meta, 'r': pv,
    })


@login_required
@require_POST
def gt_select_confirm(request, token):
    """Import the NEW GT Select orders + lines under one IMPORT run."""
    d = _gts_token_dir(token)
    mp = d / 'meta.json'
    if not mp.exists():
        raise Http404("Upload not found or expired.")
    meta = json.loads(mp.read_text(encoding='utf-8'))
    from .services import gt_select_import as gts
    res = gts.do_import(str(d / meta['headers_path']), str(d / meta['lines_path']))
    if not res.get('ok'):
        messages.error(request, res.get('error', 'Import failed.'))
        return redirect('b2b_gt_select_preview', token=token)
    shutil.rmtree(d, ignore_errors=True)
    if not res.get('imported'):
        messages.info(request, "All orders were already imported — nothing new.")
        return redirect('b2b_orders')
    messages.success(
        request, f"✓ Imported {res['imported']} GT Select order(s) · "
        f"{res['lines']} line(s) ({res.get('skipped', 0)} already present).")
    if res.get('run_id'):
        return redirect('b2b_run_detail', run_id=res['run_id'])
    return redirect('b2b_orders')


@login_required
@require_POST
def gt_select_discard(request, token):
    d = _gts_token_dir(token)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
    messages.info(request, "Upload discarded.")
    return redirect('b2b_gt_select')


@login_required
@require_POST
def item_master_add(request):
    """Add/overwrite ONE item by hand — written to item_master (flagged
    batch_id='manual', durable across rebuilds); a typed Swiggy SKU goes to
    channel_sku_map."""
    from .services import item_master_loader as iml
    res = iml.upsert_manual_item({
        'item_no': request.POST.get('item_no', ''),
        'ean': request.POST.get('ean', ''),
        'description': request.POST.get('description', ''),
        'gst_code': request.POST.get('gst_code', ''),
        'hsn': request.POST.get('hsn', ''),
        'mrp': request.POST.get('mrp', ''),
        'mrp_start': request.POST.get('mrp_start', ''),
        'mrp_end': request.POST.get('mrp_end', ''),
        'swiggy_sku_code': request.POST.get('swiggy_sku_code', ''),
    })
    if not res.get('ok'):
        messages.error(request, res.get('error', 'Could not add the item.'))
    else:
        messages.success(request, f"✓ Item {res['item_no']} saved to the master "
                         "(kept across rebuilds until the ERP export includes it).")
    return redirect('b2b_item_master')


# ── Ship-To Mapping (DB-backed; the bundled Ship to B2B.xlsx is retired) ─────
_STM_UPLOADS = _MEDIA / 'b2b_ship_to'


def _stm_token_dir(token: str) -> Path:
    base = _STM_UPLOADS.resolve()
    d = (_STM_UPLOADS / token).resolve()
    if d != base and base not in d.parents:
        raise Http404()
    return d


class ShipToView(LoginRequiredMixin, TemplateView):
    """`/b2b/ship-to/` — status + upload/replace + browse/search + add/delete."""
    template_name = 'online_b2b/ship_to.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import mapping_store as ms
        ctx['status'] = ms.status()
        ctx['overview'] = ms.list_mappings(
            self.request.GET.get('party', ''), self.request.GET.get('q', ''),
            limit=300)
        return ctx


@login_required
@require_POST
def ship_to_upload(request):
    """Stash an uploaded Ship-To B2B Excel under a token → preview. No DB write."""
    f = request.FILES.get('mapping_file')
    if not f:
        messages.error(request, "Choose the Ship-To B2B Excel to upload.")
        return redirect('b2b_ship_to')
    token = uuid.uuid4().hex[:12]
    d = _STM_UPLOADS / token
    d.mkdir(parents=True, exist_ok=True)
    dest = d / ('mapping' + Path(f.name).suffix)
    with open(dest, 'wb') as out:
        for chunk in f.chunks():
            out.write(chunk)
    (d / 'meta.json').write_text(json.dumps(
        {'name': f.name, 'path': dest.name}), encoding='utf-8')
    return redirect('b2b_ship_to_preview', token=token)


@login_required
def ship_to_preview(request, token):
    """Parse the uploaded Excel in-memory (no write) and show counts to confirm."""
    d = _stm_token_dir(token)
    mp = d / 'meta.json'
    if not mp.exists():
        raise Http404("Upload not found or expired.")
    meta = json.loads(mp.read_text(encoding='utf-8'))
    from .services import mapping_store as ms
    rows, stats, warnings = ms.build_rows(str(d / meta['path']))
    if not rows:
        messages.error(request, '; '.join(warnings) or "No rows parsed.")
        return redirect('b2b_ship_to')
    return render(request, 'online_b2b/ship_to_preview.html', {
        'token': token, 'meta': meta, 'stats': stats, 'warnings': warnings,
        'sample': rows[:15], 'current': ms.status(),
    })


@login_required
@require_POST
def ship_to_confirm(request, token):
    """Replace the Excel-sourced rows from the uploaded file (manual rows kept)."""
    d = _stm_token_dir(token)
    mp = d / 'meta.json'
    if not mp.exists():
        raise Http404("Upload not found or expired.")
    meta = json.loads(mp.read_text(encoding='utf-8'))
    from .services import mapping_store as ms
    try:
        rows, stats, _ = ms.build_rows(str(d / meta['path']))
        res = ms.replace_mapping(rows)
    except Exception as e:  # noqa: BLE001
        messages.error(request, f"Replace failed: {type(e).__name__}: {e}")
        return redirect('b2b_ship_to_preview', token=token)
    shutil.rmtree(d, ignore_errors=True)
    messages.success(request, f"✓ Ship-To mapping replaced — {res['rows']} rows "
                     f"across {stats['parties']} parties now live.")
    return redirect('b2b_ship_to')


@login_required
@require_POST
def ship_to_discard(request, token):
    d = _stm_token_dir(token)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
    messages.info(request, "Upload discarded.")
    return redirect('b2b_ship_to')


@login_required
@require_POST
def ship_to_seed(request):
    """Re-seed the table from the bundled Ship to B2B.xlsx (one-click)."""
    from .services import mapping_store as ms
    res = ms.seed_from_bundled()
    if res.get('ok'):
        messages.success(request, f"✓ Seeded {res['rows']} rows from the bundled "
                         f"Ship-To B2B Excel.")
    else:
        messages.error(request, res.get('error', 'Seed failed.'))
    return redirect('b2b_ship_to')


@login_required
def ship_to_search(request):
    """As-you-type filter (party + text) → the table partial."""
    from .services import mapping_store as ms
    data = ms.list_mappings(request.GET.get('party', ''),
                            request.GET.get('q', ''), limit=300)
    return render(request, 'online_b2b/_ship_to_rows.html', {'overview': data})


@login_required
@require_POST
def ship_to_add(request):
    """Add ONE mapping row (durable manual). Passes every posted field through so
    custom ``cf_*`` values are captured too."""
    from .services import mapping_store as ms
    data = {k: v for k, v in request.POST.items() if k != 'csrfmiddlewaretoken'}
    res = ms.add_mapping(data)
    return JsonResponse(res, status=200 if res.get('ok') else 400)


@login_required
@require_POST
def ship_to_field_add(request):
    """Personalization — define a new custom column from a human label."""
    from .services import mapping_store as ms
    res = ms.add_custom_field(request.POST.get('label', ''))
    return JsonResponse(res, status=200 if res.get('ok') else 400)


@login_required
@require_POST
def ship_to_field_delete(request):
    """Remove a custom column (per-row values are left dormant in ``extra``)."""
    from .services import mapping_store as ms
    res = ms.delete_custom_field(request.POST.get('name', ''))
    return JsonResponse(res, status=200 if res.get('ok') else 400)


@login_required
def ship_to_export(request):
    """Download the Ship-To mapping as Excel — FULL columns, honoring the current
    party/search filter (so 'export what I'm viewing' just works). One 'Ship-To
    Mapping' sheet: Party · Del Location · Cust No · Ship-to · Name · Address ·
    Address 2 · Postcode · City · Source."""
    import datetime as _dt
    import io as _io
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from .services import mapping_store as ms
    party = request.GET.get('party', '')
    q = request.GET.get('q', '')
    cols, rows = ms.export_rows(party, q)
    heads = ['Party', 'Del Location', 'Cust No', 'Ship-to', 'Name', 'Address',
             'Address 2', 'Postcode', 'City', 'Source']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ship-To Mapping'
    ws.append(heads)
    for r in rows:
        ws.append([r.get(c, '') for c in cols])
    navy = PatternFill('solid', fgColor='1A237E')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = navy
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for i, w in enumerate([10, 34, 10, 14, 40, 44, 34, 11, 18, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{ws.max_row}"
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tag = (party or 'all').replace(' ', '') + (f"_{q}" if q else '')
    fname = (f"ship_to_mapping_{tag}_"
             f"{_dt.datetime.now():%d-%m-%Y_%H%M%S}.xlsx")
    return FileResponse(
        buf, as_attachment=True, filename=fname,
        content_type='application/vnd.openxmlformats-officedocument.'
                     'spreadsheetml.sheet')


@login_required
@require_POST
def ship_to_edit(request, row_id):
    """Edit ONE mapping row. Only the posted keys are updated (partial), so an
    inline City/Postcode tweak never blanks the enriched Address. Custom ``cf_*``
    values ride along."""
    from .services import mapping_store as ms
    data = {k: v for k, v in request.POST.items() if k != 'csrfmiddlewaretoken'}
    res = ms.update_mapping(row_id, data)
    return JsonResponse(res, status=200 if res.get('ok') else 400)


@login_required
@require_POST
def ship_to_delete(request, row_id):
    from .services import mapping_store as ms
    return JsonResponse(ms.delete_mapping(row_id))


class AvailabilityView(LoginRequiredMixin, TemplateView):
    """`/b2b/availability/` — Order Availability Checker. Paste order number(s)
    from the tracker → check each recorded line against current inventory in the
    mapped warehouse (auto + override). Read-only compose; the check is AJAX."""
    template_name = 'online_b2b/availability.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import inventory_store as inv
        ctx['warehouses'] = inv.WAREHOUSES
        return ctx


@login_required
@require_POST
def availability_check(request):
    """Paste blob → parsed order nos → availability dict (JSON). Read-only."""
    from .services import availability as av
    nos = av.parse_order_nos(request.POST.get('orders', ''))
    if not nos:
        return JsonResponse({'ok': False,
                             'error': 'Paste at least one order number.'})
    nos = nos[:500]                       # sane cap for a paste-and-check
    return JsonResponse(av.check_orders(nos, request.POST.get('warehouse', '')))


@login_required
@require_POST
def availability_export(request):
    """Same check → styled multi-sheet .xlsx (Summary · PO Summary · By Order
    Lines · By SKU · Not Found). Qty AND value fill at every angle."""
    import datetime as _dt
    from .services import availability as av
    nos = av.parse_order_nos(request.POST.get('orders', ''))
    if not nos:
        return JsonResponse({'ok': False,
                             'error': 'Paste at least one order number.'})
    data = av.check_orders(nos[:500], request.POST.get('warehouse', ''))
    # Enrich with per-item bins (bulk per WH) for the 'SKU Bins' sheet.
    from collections import defaultdict
    from .services import inventory_store as inv
    by_wh = defaultdict(set)
    for k in data.get('skus', []):
        by_wh[k['wh']].add(k['item_no'])
    data['sku_bins'] = {wh: inv.item_bins_bulk(wh, items)
                        for wh, items in by_wh.items()}
    buf = av.to_workbook(data)
    fname = f"availability_{_dt.datetime.now():%d-%m-%Y_%H%M%S}.xlsx"
    return FileResponse(
        buf, as_attachment=True, filename=fname,
        content_type='application/vnd.openxmlformats-officedocument.'
                      'spreadsheetml.sheet')


@login_required
def availability_bins(request):
    """Per-item bin breakdown for the lazy expand on an SKU row — ONLY the
    INCLUDED (sellable) bins, so the drill-down sums to the item's available qty.
    Excluded return/QC/block bins are deliberately not shown here.
    GET ?wh=<code>&item=<item_no> → {ok, item, wh, bins:[{bin,zone,decision,qty}]}."""
    from .services import inventory_store as inv
    wh = (request.GET.get('wh') or '').strip()
    item = (request.GET.get('item') or '').strip()
    if not wh or not item:
        return JsonResponse({'ok': False, 'error': 'wh and item required.'})
    bins = [b for b in inv.item_bins(wh, item) if b.get('decision') == 'INCLUDED']
    return JsonResponse({'ok': True, 'item': item, 'wh': inv.wh_short(wh),
                         'bins': bins})


# ── Sales Validation — REMOVED 2026-07-20 (superseded by Triangular Validation).
#    Its service (services/sales_validation.py) + template were deleted; the
#    3 URL routes + sidebar link are gone. Full Validation stays (Triangular
#    depends on its service); only its sidebar link is hidden.


# ── UI Lab (learning surface: htmx · Alpine · animation) ────────────────────
# Isolated, staff-facing page that demonstrates the newer front-end tech on
# REAL data without touching any production flow. htmx/Alpine load ONLY here
# (see the template), so the blast radius is this one page.
@login_required
def ui_lab(request):
    """Render the UI Lab page. Thin view — reuses the existing item_master
    search service (no new query), so htmx has real data to fetch."""
    from .services import item_master_loader as iml
    ctx = {'result': iml.list_items('', limit=12)}   # initial rows for first paint
    return render(request, 'online_b2b/ui_lab.html', ctx)


@login_required
def ui_lab_search(request):
    """htmx endpoint: return ONLY the results partial for a search term.
    htmx swaps this fragment into the page — no full reload, no hand-written
    fetch/JSON. Reuses iml.list_items (same code path as the Item Master page)."""
    from .services import item_master_loader as iml
    res = iml.list_items(request.GET.get('q', ''), limit=25)
    return render(request, 'online_b2b/_ui_lab_rows.html', {'result': res})


# ── SKU price/CP exceptions — operator-managed (feeds the engine's exception
#    overlay; additive, the engine already auto-reads item_exceptions) ────────
@login_required
def exceptions_page(request):
    """Manage per-marketplace SKU exceptions (Use Vendor CP / Override MRP /
    Override Margin / EAN remap). Rows land in `item_exceptions` (source=manual)
    and are applied automatically on the next run — no engine change.

    Presented marketplace-wise: each MP shows its own SKU exceptions; MPs with
    none show 'all flat <margin>%'. Rows enriched with item name/no from master."""
    import datetime as _dt
    from collections import OrderedDict

    from .services import overrides_store as ov
    ov.ensure_tables()
    rows = ov.list_with_ids()

    # Enrich each row with the item name + item no from the DB master (lookup by
    # EAN or by item no — source_code can be either).
    by_ean, by_item, gst_of = {}, {}, {}
    try:
        with order_db._conn() as (cur, d):
            cur.execute('SELECT ean, item_no, description, gst_code FROM item_master')
            for ean, it, desc, gst in cur.fetchall():
                if ean:
                    by_ean[str(ean).strip()] = (str(it), desc)
                    gst_of[str(ean).strip()] = gst
                if it:
                    by_item[str(it).strip()] = desc
                    gst_of[str(it).strip()] = gst
    except Exception:  # noqa: BLE001
        pass

    def _f(x):
        try:
            return float(str(x).strip())
        except (TypeError, ValueError):
            return None

    def _gst_div(code):
        c = str(code or '').upper()
        if '18' in c:
            return 1.18
        if '12' in c:
            return 1.12
        if '5' in c and '15' not in c and '25' not in c:
            return 1.05
        if '3' in c:
            return 1.03
        if '0' in c:
            return 1.00
        return 1.18

    def _eff(r, gdiv=1.18):
        # Two operator kinds only: EAN Remap, or Override CP. For CP we show the
        # ACTUAL overridden unit price (₹) — never an internal margin %.
        if r.get('maps_to'):
            return ('remap', 'EAN Remap', f"→ {r['maps_to']}")
        # Deal SKUs carry a negotiated cost. 'Cost after GST' is already the
        # pre-GST CP; 'Cost With GST' (Myntra transfer price) → ÷(1+GST), exactly
        # as the engine writes it to the D365 Lines unit price (engine_bridge
        # expected_cp = transfer ÷ gst_div).
        price = _f(r.get('cost_after_gst'))
        if price is None:
            cwg = _f(r.get('cost_with_gst'))
            if cwg:
                price = round(cwg / gdiv, 2)
        if price is None:
            mrp, mgn = _f(r.get('override_mrp')), _f(r.get('override_margin'))
            if mrp and mgn:
                price = round(mrp * mgn / 100 / gdiv, 2)   # CP = MRP×keep% ÷ (1+GST)
            elif mrp:
                price = mrp
        if price:
            return ('cp', 'Override CP', f"₹{price:g}")
        if r.get('use_vendor_cp'):
            return ('cp', 'Override CP', 'vendor CP')
        return ('cp', 'Override CP', '—')

    def _ago(v):
        """'2 days ago' style relative string + absolute for the tooltip."""
        if not v:
            return '', ''
        try:
            dtv = v if hasattr(v, 'year') else _dt.datetime.fromisoformat(str(v)[:19])
        except (ValueError, TypeError):
            return str(v), str(v)
        delta = _dt.datetime.now() - dtv
        s = int(delta.total_seconds())
        if s < 3600:
            rel = f"{max(s // 60, 0)}m ago"
        elif s < 86400:
            rel = f"{s // 3600}h ago"
        elif s < 86400 * 30:
            rel = f"{s // 86400}d ago"
        else:
            rel = dtv.strftime('%d %b %Y')
        return rel, dtv.strftime('%d %b %Y, %H:%M')

    # Deal-SKU rows carry no Marketplace column, so attribute them by kind
    # (swiggy_deal → Swiggy, myntra_deal → Myntra) — else they'd read as
    # '(unassigned)' and the channel would wrongly look flat.
    _KIND_MP = {'swiggy_deal': 'Swiggy', 'myntra_deal': 'Myntra',
                'zepto_deal': 'Zepto'}
    eff_counts: dict = {}
    for r in rows:
        code = str(r.get('source_code') or '').strip()
        item_no, desc = by_ean.get(code, (code if code in by_item else '', by_item.get(code, '')))
        r['item_no'] = item_no
        r['item_name'] = desc or ''
        r['eff_kind'], r['eff_label'], r['eff_detail'] = _eff(r, _gst_div(gst_of.get(code)))
        # Internal sub-type — kept in the remark column, NOT a top-level label
        # (the two operator categories stay Override CP / EAN Remap only).
        if r.get('kind') == 'swiggy_deal':
            r['subtype'] = 'Swiggy deal SKU'
        elif r.get('kind') == 'myntra_deal':
            r['subtype'] = 'Myntra deal SKU'
        elif r.get('kind') == 'zepto_deal':
            r['subtype'] = 'Zepto deal SKU'
        elif r.get('use_vendor_cp'):
            r['subtype'] = 'Vendor CP'
        elif r.get('maps_to'):
            r['subtype'] = 'EAN remap'
        elif r.get('override_margin'):
            r['subtype'] = 'Margin override'
        elif r.get('override_mrp'):
            r['subtype'] = 'MRP override'
        else:
            r['subtype'] = ''
        r['created_rel'], r['created_abs'] = _ago(r.get('created_at'))
        r['updated_rel'], r['updated_abs'] = _ago(r.get('updated_at'))
        r['mp_eff'] = ((r.get('marketplace') or '').strip()
                       or _KIND_MP.get(r.get('kind'), ''))
        eff_counts[r['eff_kind']] = eff_counts.get(r['eff_kind'], 0) + 1

    # De-dup twins: some SKUs carry BOTH a legacy 'use vendor CP' row AND a deal
    # row with the actual fixed transfer price (e.g. Myntra matte-lock ↔ ₹63.35).
    # Show the fixed price once — drop the redundant vendor-CP twin from the view
    # (display only; the engine + DB are untouched). SKUs that are genuinely
    # vendor-CP with no fixed number (e.g. Goddess) keep their row.
    def _has_fixed(r):
        return bool(_f(r.get('cost_after_gst')) or _f(r.get('cost_with_gst'))
                    or (_f(r.get('override_mrp')) and _f(r.get('override_margin'))))
    _fixed_keys = {(r['mp_eff'], str(r.get('source_code') or '').strip())
                   for r in rows if _has_fixed(r)}
    rows = [r for r in rows
            if not (r.get('use_vendor_cp') and not r.get('maps_to') and not _has_fixed(r)
                    and (r['mp_eff'], str(r.get('source_code') or '').strip()) in _fixed_keys)]
    # Recount effects after the de-dup so the effect-lens totals stay honest.
    eff_counts = {}
    for r in rows:
        eff_counts[r['eff_kind']] = eff_counts.get(r['eff_kind'], 0) + 1

    # Group by (effective) marketplace (order: most exceptions first).
    groups: dict = OrderedDict()
    for r in rows:
        mp = r['mp_eff'] or '(unassigned)'
        groups.setdefault(mp, []).append(r)
    grouped = sorted(
        ({'mp': mp, 'rows': rs, 'count': len(rs),
          'n_manual': sum(1 for x in rs if x.get('source') == 'manual'),
          'margin': engine_bridge.default_margin_pct(mp) if mp != '(unassigned)' else None}
         for mp, rs in groups.items()),
        key=lambda g: (-g['count'], g['mp']))

    # Marketplaces with NO exceptions → 'all flat <margin>%'.
    flat = [{'mp': mp, 'margin': engine_bridge.default_margin_pct(mp)}
            for mp in engine_bridge.PILOT_MARKETPLACES if mp not in groups]

    # Effect lens — the TWO operator categories only (must match _eff()'s
    # eff_kind: 'cp' / 'remap'), ordered, with a label + count.
    _EFF_ORDER = [('cp', 'Override CP'), ('remap', 'EAN Remap')]
    by_effect = [{'kind': k, 'label': lbl, 'count': eff_counts.get(k, 0)}
                 for k, lbl in _EFF_ORDER if eff_counts.get(k)]

    return render(request, 'online_b2b/exceptions.html', {
        'rows': rows, 'grouped': grouped, 'flat_mps': flat, 'by_effect': by_effect,
        'marketplaces': engine_bridge.PILOT_MARKETPLACES,
        'counts': ov.table_counts(),
        'n_manual': sum(1 for r in rows if r.get('source') == 'manual'),
        'n_mp_with': len(grouped), 'n_flat': len(flat)})


@login_required
@require_POST
def exception_add(request):
    from .services import overrides_store as ov
    res = ov.add_manual(
        marketplace=request.POST.get('marketplace', ''),
        source_code=request.POST.get('source_code', ''),
        maps_to=request.POST.get('maps_to', ''),
        override_mrp=request.POST.get('override_mrp', ''),
        override_margin=request.POST.get('override_margin', ''),
        use_vendor_cp=request.POST.get('use_vendor_cp', ''),
        note=request.POST.get('note', ''))
    return JsonResponse(res)


@login_required
@require_POST
def exception_update(request, row_id):
    from .services import overrides_store as ov
    fields = {k: request.POST[k] for k in
              ('marketplace', 'source_code', 'maps_to', 'override_mrp',
               'override_margin', 'use_vendor_cp', 'note') if k in request.POST}
    return JsonResponse(ov.update_manual(row_id, **fields))


@login_required
@require_POST
def exception_delete(request, row_id):
    from .services import overrides_store as ov
    return JsonResponse(ov.delete_manual(row_id))
