"""
online_b2b.services.eka_data
============================

**EKA location registry — shifted Excel → DB** (mirrors the item-master and
ship-to-mapping moves). The desktop EKA constructor reads ``EKA_DATA.xlsx`` (one
row per airport / EBO / kiosk store: Bill-to, Ship-to, Location code, prefixes,
SO-number examples, margin, active flag). We relocate that whole sheet into the
``eka_data`` table so the web app owns the data; the desktop stays untouched for
now and we migrate the reader to the DB later, one step at a time.

Web-owned table in ``renee_orders`` — never touches the business order tables.
"""
from __future__ import annotations

from .order_db import _conn

_TABLE = 'eka_data'

# Excel header (as it appears in EKA_DATA.xlsx) → DB column. Order-independent;
# matched case/space-insensitively so a stray trailing space ('Bill to ') is fine.
_COL_MAP = {
    'desc': 'descr',
    'billto': 'bill_to',
    'shipto': 'ship_to',
    'location': 'location_code',
    'genbizpostinggroup': 'posting_group',
    'shortname': 'short_name',
    'prefix': 'prefix',
    'shortcode': 'short_code',
    'transfercode': 'transfer_code',
    'type': 'kind',
    'exampleregular': 'example_regular',
    'exampletester': 'example_tester',
    'status': 'status',
    'mrgnpct': 'margin_pct',
}

_COLS = ['descr', 'bill_to', 'ship_to', 'location_code', 'posting_group',
         'short_name', 'prefix', 'short_code', 'transfer_code', 'kind',
         'example_regular', 'example_tester', 'status', 'margin_pct']

_CREATE = f"""
CREATE TABLE IF NOT EXISTS {_TABLE} (
    id             INT AUTO_INCREMENT PRIMARY KEY,
    descr          VARCHAR(255),
    bill_to        VARCHAR(20),
    ship_to        VARCHAR(20),
    location_code  VARCHAR(60),
    posting_group  VARCHAR(40),
    short_name     VARCHAR(80),
    prefix         VARCHAR(10),
    short_code     VARCHAR(30),
    transfer_code  VARCHAR(60),
    kind           VARCHAR(20),
    example_regular VARCHAR(60),
    example_tester  VARCHAR(60),
    status         VARCHAR(12),
    margin_pct     DECIMAL(6,3),
    updated_at     DATETIME,
    INDEX idx_eka_short (short_name),
    INDEX idx_eka_loc (location_code)
)"""


def ensure_table() -> None:
    with _conn() as (cur, d):
        cur.execute(_CREATE)
        cur.connection.commit()


def _norm(h) -> str:
    return ''.join(ch for ch in str(h or '').lower() if ch.isalnum())


def load_from_excel(path: str, replace: bool = True) -> dict:
    """Read ``EKA_DATA.xlsx`` and (re)load every store row into ``eka_data``.

    ``replace=True`` truncates first so the table mirrors the sheet exactly
    (the sheet is the authoritative desktop registry today). Returns counts.
    """
    import datetime as _dt

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # position of each DB column in the sheet (via the normalised header map)
    pos = {}
    for i, h in enumerate(header):
        key = _COL_MAP.get(_norm(h))
        if key:
            pos[key] = i

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # skip fully-blank rows; a store row must have a Short Name or Desc
        sn = r[pos['short_name']] if 'short_name' in pos and pos['short_name'] < len(r) else None
        ds = r[pos['descr']] if 'descr' in pos and pos['descr'] < len(r) else None
        if not sn and not ds:
            continue
        vals = []
        for c in _COLS:
            idx = pos.get(c)
            v = r[idx] if idx is not None and idx < len(r) else None
            if isinstance(v, str):
                v = v.strip()
            # Store margin as a WHOLE PERCENT (60 = 60%). A sheet may carry it as a
            # fraction (0.60); normalise anything <= 1 to whole percent so the value
            # the team sees/edits and the engine's (÷100) agree.
            if c == 'margin_pct' and v not in (None, ''):
                try:
                    fv = float(v)
                    v = fv * 100 if fv <= 1 else fv
                except (TypeError, ValueError):
                    v = None
            vals.append(v)
        rows.append(vals)

    ensure_table()
    now = _dt.datetime.now()
    with _conn() as (cur, d):
        ph = d['ph']
        if replace:
            cur.execute(f"DELETE FROM {_TABLE}")
        marks = ', '.join([ph] * (len(_COLS) + 1))
        cur.executemany(
            f"INSERT INTO {_TABLE} ({', '.join(_COLS)}, updated_at) VALUES ({marks})",
            [tuple(v) + (now,) for v in rows])
        cur.connection.commit()
    return {'loaded': len(rows), 'replaced': replace}


def all_rows() -> list[dict]:
    """Every EKA store as a JSON-safe dict (for the web page / future engine).
    Includes the row ``id`` so the page can edit a single row."""
    ensure_table()
    cols = ['id'] + _COLS
    with _conn() as (cur, d):
        cur.execute(f"SELECT {', '.join(cols)} FROM {_TABLE} ORDER BY short_name")
        out = []
        for r in cur.fetchall():
            row = dict(zip(cols, r))
            if row.get('margin_pct') is not None:
                row['margin_pct'] = float(row['margin_pct'])
            out.append(row)
        return out


def active_rows() -> list[dict]:
    return [r for r in all_rows() if str(r.get('status', '')).lower() != 'inactive']


# ── web CRUD (edit margin / Type / status / mappings on the EKA Data page) ────
# Editable via the web; the engine reads margin_pct + kind (SO/TO) from here.
_EDITABLE = ['descr', 'bill_to', 'ship_to', 'location_code', 'posting_group',
             'short_name', 'prefix', 'short_code', 'transfer_code', 'kind',
             'example_regular', 'example_tester', 'status', 'margin_pct']


def get_row(row_id) -> dict | None:
    ensure_table()
    cols = ['id'] + _COLS
    with _conn() as (cur, d):
        cur.execute(f"SELECT {', '.join(cols)} FROM {_TABLE} WHERE id={d['ph']}", (int(row_id),))
        r = cur.fetchone()
        return dict(zip(cols, r)) if r else None


def update_row(row_id, fields: dict) -> dict:
    """Edit ONE EKA store row — only the standard editable fields present in
    ``fields`` are written (partial). ``margin_pct`` is coerced to a number."""
    import datetime as _dt
    ensure_table()
    present = [k for k in _EDITABLE if k in fields]
    if not present:
        return {'ok': False, 'error': 'Nothing to update.'}
    vals = []
    for k in present:
        v = fields.get(k)
        if k == 'margin_pct':
            try:
                v = float(v) if str(v).strip() != '' else None
            except (TypeError, ValueError):
                return {'ok': False, 'error': 'Margin must be a number (e.g. 60).'}
        elif isinstance(v, str):
            v = v.strip()
        vals.append(v)
    with _conn() as (cur, d):
        ph = d['ph']
        sets = present + ['updated_at']
        cur.execute(
            f"UPDATE {_TABLE} SET {', '.join(f'{c}={ph}' for c in sets)} WHERE id={ph}",
            vals + [_dt.datetime.now(), int(row_id)])
        n = cur.rowcount
        cur.connection.commit()
    # rowcount is 0 when the values are unchanged (MySQL/TiDB counts CHANGED rows,
    # not matched) — that's still a successful save, so don't report it as failed.
    return {'ok': True, 'updated': n or 0}


def add_row(fields: dict) -> dict:
    """Insert a new EKA store row (source = web)."""
    import datetime as _dt
    ensure_table()
    if not str(fields.get('short_name') or '').strip() and not str(fields.get('descr') or '').strip():
        return {'ok': False, 'error': 'Short Name or Desc is required.'}
    vals = []
    for k in _COLS:
        v = fields.get(k)
        if k == 'margin_pct':
            try:
                v = float(v) if str(v).strip() != '' else None
            except (TypeError, ValueError):
                v = None
        elif isinstance(v, str):
            v = v.strip()
        vals.append(v)
    with _conn() as (cur, d):
        ph = d['ph']
        marks = ', '.join([ph] * (len(_COLS) + 1))
        cur.execute(
            f"INSERT INTO {_TABLE} ({', '.join(_COLS)}, updated_at) VALUES ({marks})",
            tuple(vals) + (_dt.datetime.now(),))
        new_id = cur.lastrowid
        cur.connection.commit()
    return {'ok': True, 'id': new_id}


def status() -> dict:
    ensure_table()
    rows = all_rows()
    active = [r for r in rows if str(r.get('status', '')).lower() != 'inactive']
    return {'count': len(rows), 'active': len(active),
            'inactive': len(rows) - len(active)}
