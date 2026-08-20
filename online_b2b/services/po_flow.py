"""
online_b2b.services.po_flow
===========================

Reusable, channel-agnostic **PO flow scaffold**: upload → review → confirm,
shared by every segment so new channels/marketplaces plug in instead of
duplicating the page.

ADDITIVE — the existing online_b2b and offline backends are NOT modified. A
channel wires in through a :class:`FlowSpec` (a processor factory + capability
flags + URL names + base template + optional channel-specific slots). The flow
stores the upload under ``MEDIA/<dirname>/<token>/`` (files + ``meta.json`` +
a signature-keyed ``preview.json`` cache), exactly mirroring the proven online
token model.

**Capabilities & slots keep ONE template working across mismatched channels.**
Not every channel is fully compatible (online has vendor-price compare +
Override + D365; GT Mass has none of those but has its own file-level
exceptions like *PO-number-missing* / *template-mismatch*). So:

* every optional block renders only when its capability flag is on
  (``caps`` = ``{'warehouse','margin','marketplace','vendor_cols','override',
  'ean_fix','exclude','d365'}``), and
* ``extra_partial`` is a **null-by-default** include where a channel injects its
  own panel — dark for every other channel.

A processor (the ``FlowSpec.processor`` factory result) must expose:

    preview() -> payload                 # no DB writes
    confirm(actions: dict) -> result     # writes, returns {'ok', 'run_id', ...}

``payload`` (and ``result`` where noted) is the unified review shape::

    {ok, error?, summary:{pos,lines,qty,value,affected,skipped},
     headers:[{po,location,order_type,items,qty,order_value}],
     lines:[{po,item_no,ean,description,qty,unit_price,our_mrp,
             status,exception_label, key}],
     affected:[ <subset of lines, status != OK> ],
     file_issues:[{file,problem,detail,kind}],   # channel-specific exceptions
     skipped:[{po,location,qty,order_value,marketplace_label}],
     warnings:[...], output_path?}
"""
from __future__ import annotations

import hashlib
import json
import shutil
import uuid
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path

from django.conf import settings


@dataclass(frozen=True)
class FlowSpec:
    """Per-channel configuration for the shared flow. One instance per channel."""
    key: str                       # 'gt_mass' — stable slug
    title: str                     # 'GT Mass' — shown in the UI
    segment: str                   # 'Offline' / 'Online B2B'
    base_template: str             # e.g. 'core/base.html' / 'online_b2b/base_b2b.html'
    upload_dirname: str            # MEDIA subdir for this channel's uploads
    processor: Callable[[dict], object]   # (meta) -> processor with preview()/confirm()
    urls: dict                     # name map: upload/review/confirm/decision/discard/download/back/dashboard
    caps: frozenset = frozenset()  # capability flags (see module docstring)
    warehouses: tuple = ()         # ((code, label), ...) — only if 'warehouse' cap
    marketplaces: tuple = ()       # ((value, label), ...) — only if 'marketplace' cap
    default_margin: float | None = None
    extra_partial: str | None = None       # back-compat alias for slots['after_kpis']
    # Named channel-specific slots — partial templates injected at fixed anchors so
    # each channel can add its own bits where its inconsistency needs them. Every
    # slot is null by default (dark for channels that don't set it). Anchors:
    #   upload:  'fields'        (extra form fields)
    #   review:  'top'           (panel above the KPI cards)
    #            'after_kpis'    (panel under the KPI cards — e.g. file exceptions)
    #            'tabs'          (extra tab button)  + 'panes' (its tab pane)
    #            'actions'       (extra confirm-bar button)
    slots: dict = field(default_factory=dict)
    intro: str = ''                # one-line page subtitle
    accept: str = '.xlsx,.xls,.xlsm,.csv,.pdf'

    def slot_map(self) -> dict:
        """Effective slots (``extra_partial`` folds into ``after_kpis``)."""
        s = dict(self.slots)
        if self.extra_partial and 'after_kpis' not in s:
            s['after_kpis'] = self.extra_partial
        return s

    def caps_map(self) -> dict:
        """``{cap: True}`` for template ``{% if caps.exclude %}`` checks."""
        return {c: True for c in self.caps}


# ── token / upload store ─────────────────────────────────────────────────
def _root(spec: FlowSpec) -> Path:
    return Path(settings.MEDIA_ROOT) / spec.upload_dirname


def _dir(spec: FlowSpec, token: str) -> Path:
    return _root(spec) / token


def new_token() -> str:
    return uuid.uuid4().hex[:12]


def save_upload(spec: FlowSpec, files, extra: dict | None = None) -> str:
    """Persist uploaded file(s) under a fresh token + write ``meta.json``."""
    token = new_token()
    d = _dir(spec, token)
    d.mkdir(parents=True, exist_ok=True)
    paths = []
    for f in files:
        name = Path(getattr(f, 'name', 'upload')).name
        dest = d / name
        with open(dest, 'wb') as out:
            for chunk in f.chunks():
                out.write(chunk)
        paths.append(str(dest))
    meta = {'files': paths, 'decisions': {}, 'locked': False, 'run_id': None}
    meta.update(extra or {})
    _write_meta(spec, token, meta)
    return token


def _meta_path(spec: FlowSpec, token: str) -> Path:
    return _dir(spec, token) / 'meta.json'


def load_meta(spec: FlowSpec, token: str) -> dict | None:
    p = _meta_path(spec, token)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding='utf-8'))
    except Exception:  # noqa: BLE001
        return None


def _write_meta(spec: FlowSpec, token: str, meta: dict) -> None:
    _meta_path(spec, token).write_text(
        json.dumps(meta, ensure_ascii=False, indent=1), encoding='utf-8')


def _sig(meta: dict) -> str:
    """Signature that invalidates the preview cache when inputs change."""
    basis = {'files': meta.get('files', []),
             'warehouse': meta.get('warehouse', ''),
             'margin': meta.get('margin_pct', ''),
             'ean_fixes': meta.get('ean_fixes', {})}
    return hashlib.md5(
        json.dumps(basis, sort_keys=True).encode('utf-8')).hexdigest()


def preview(spec: FlowSpec, token: str, meta: dict) -> dict:
    """Cached preview — runs the processor once per (inputs) signature."""
    cache = _dir(spec, token) / 'preview.json'
    sig = _sig(meta)
    if cache.exists():
        try:
            blob = json.loads(cache.read_text(encoding='utf-8'))
            if blob.get('sig') == sig:
                return blob['payload']
        except Exception:  # noqa: BLE001
            pass
    payload = spec.processor(meta).preview()
    try:
        cache.write_text(json.dumps({'sig': sig, 'payload': payload},
                                    ensure_ascii=False), encoding='utf-8')
    except Exception:  # noqa: BLE001
        pass
    return payload


def _invalidate(spec: FlowSpec, token: str) -> None:
    c = _dir(spec, token) / 'preview.json'
    if c.exists():
        try:
            c.unlink()
        except Exception:  # noqa: BLE001
            pass


# Public alias — reopen-from-Drafts re-validates a parked run against fresh masters.
invalidate = _invalidate


def set_decision(spec: FlowSpec, token: str, key: str, action: str,
                 override_cp: str = '', remark: str = '') -> int:
    """Persist one per-line decision onto ``meta.json``. Returns total decided."""
    meta = load_meta(spec, token) or {}
    dec = meta.setdefault('decisions', {})
    if action:
        dec[key] = {'action': action,
                    'override_cp': override_cp or None,
                    'remark': remark or ''}
    else:
        dec.pop(key, None)
    _write_meta(spec, token, meta)
    return len(dec)


def discard(spec: FlowSpec, token: str) -> None:
    d = _dir(spec, token)
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)


def _overlay_decisions(payload: dict, meta: dict) -> None:
    """Attach saved per-line decisions to lines/affected so the review page
    shows prior operator choices (and so confirm can replay them)."""
    dec = meta.get('decisions', {})
    for bucket in ('lines', 'affected'):
        for ln in payload.get(bucket, []) or []:
            ln['decision'] = dec.get(ln.get('key', ''), {})


_AFF_STATUSES = {'MISMATCH', 'NOT_IN_MASTER'}


def _sku_rows(lines) -> list:
    """Per-SKU demand rollup (parity with the Online B2B 'SKU' tab) — grouped by
    (item_no, ean): qty, # POs, value (Σ unit price × qty), the unit price, and an
    overridden flag. Computed purely from the already-resolved lines — no engine
    touch. Value basis follows the channel's unit_price (channel-native)."""
    agg: dict = {}
    for ln in lines or []:
        key = (ln.get('item_no') or '', ln.get('ean') or '')
        a = agg.get(key)
        if a is None:
            a = agg[key] = {'item_no': ln.get('item_no') or '', 'ean': ln.get('ean') or '',
                            'description': ln.get('description') or '', 'qty': 0, 'pos': set(),
                            'value': 0.0, 'ups': set(), 'overridden': False, 'labels': set()}
        q = int(ln.get('qty') or 0)
        a['qty'] += q
        a['pos'].add(ln.get('po'))
        up = ln.get('unit_price')
        if up not in (None, ''):
            try:
                upf = float(up)
                a['ups'].add(round(upf, 2))
                a['value'] += upf * q
            except (TypeError, ValueError):
                pass
        lbl = (ln.get('exception_label') or '').strip()
        act = str((ln.get('decision') or {}).get('action') or '').upper()
        if lbl:
            a['labels'].add(lbl)
            a['overridden'] = True
        if act in ('OVERRIDE', 'EXCLUDE', 'INCLUDE'):
            a['overridden'] = True
    rows = []
    for a in agg.values():
        ups = sorted(a['ups'])
        # No per-line price on this channel (e.g. MT) → leave value/price blank
        # rather than a misleading ₹0.
        val = round(a['value'], 2) if ups else None
        rows.append({'item_no': a['item_no'], 'ean': a['ean'], 'description': a['description'],
                     'qty': a['qty'], 'pos': len(a['pos']), 'value': val,
                     'unit_price': (ups[-1] if ups else None), 'unit_price_varies': len(ups) > 1,
                     'overridden': a['overridden'], 'note': ', '.join(sorted(a['labels']))})
    rows.sort(key=lambda r: (-(r['value'] or 0), -r['qty']))
    return rows


def _mark_clean(headers, lines):
    """Flag each header CLEAN / AFFECTED (parity with the Online B2B Orders tab +
    workbook Summary Status): a PO is AFFECTED if any of its lines is EXCLUDEd or
    an unresolved MISMATCH / NOT_IN_MASTER; else 100% goes through → CLEAN."""
    affected = set()
    for ln in lines or []:
        act = str((ln.get('decision') or {}).get('action') or '').upper()
        if act == 'EXCLUDE' or ((ln.get('status') or '') in _AFF_STATUSES
                                and act not in ('INCLUDE', 'OVERRIDE')):
            affected.add(str(ln.get('po') or ''))
    n_clean = 0
    for h in headers or []:
        h['clean'] = str(h.get('po') or '') not in affected
        n_clean += 1 if h['clean'] else 0
    return n_clean, len(headers or []) - n_clean


def _tag_warehouse(payload, meta):
    """Tag each header with the fulfilment warehouse SHORT (AHD / BLR / North) it
    maps to — the SAME resolution the tracker + availability use — so the Orders
    tab can show a 'WH' column (parity with Online B2B). Uses the header's own
    warehouse if present, else the run-level one; blank resolves to the default."""
    from . import inventory_store as inv
    summ = payload.get('summary') or {}
    run_wh = summ.get('warehouse') or meta.get('warehouse') or ''
    mp = summ.get('marketplace') or meta.get('marketplace') or ''
    for h in payload.get('headers') or []:
        raw = h.get('warehouse') or run_wh
        h['wh'] = inv.wh_short(inv.resolve_order_wh(raw, mp, mp))


def _mapping_report(headers):
    """Per-PO ship-to resolution for a Mapping tab (parity with Online B2B):
    MAPPED (ship-to resolved) vs UNMAPPED (blank ship-to → the SO can't reach
    D365). Only meaningful when the channel exposes ``ship_to`` on the header.
    Returns ``(rows, n_unmapped, has_mapping)``."""
    rows, n_unmapped, has_mapping = [], 0, False
    for h in headers or []:
        if 'ship_to' not in h:
            continue
        has_mapping = True
        ship = str(h.get('ship_to') or '').strip()
        mapped = bool(ship)
        if not mapped:
            n_unmapped += 1
        h['mapped'] = mapped                       # for the Orders-tab chip too
        rows.append({'po': h.get('po'),
                     'location': str(h.get('raw_location') or h.get('location') or ''),
                     'ship_to': ship, 'qty': h.get('qty'),
                     'del_location': '', 'name': '',
                     'city': '', 'state': '', 'postcode': '',
                     'match_type': 'MAPPED' if mapped else 'UNMAPPED'})
    # Enrich with the RESOLVED D365 delivery location + address (del location /
    # name / city / state / pin) per ship-to code, so the Mapping tab reads
    # "location (from file) → ship-to → the resolved delivery location + address"
    # exactly like the Online B2B ship-to page.
    codes = {r['ship_to'] for r in rows if r['ship_to']}
    if codes:
        try:
            from .order_db import _conn
            with _conn() as (cur, d):
                ph = d['ph']
                ins = ','.join([ph] * len(codes))
                cur.execute(f"SELECT ship_to, del_location, name, city, state, postcode "
                            f"FROM ship_to_mapping WHERE ship_to IN ({ins})", tuple(codes))
                addr = {}
                for st, dl, nm, city, state, pin in cur.fetchall():
                    addr.setdefault(str(st), (str(dl or ''), str(nm or ''),
                                              str(city or ''), str(state or ''), str(pin or '')))
            for r in rows:
                a = addr.get(r['ship_to'])
                if a:
                    r['del_location'], r['name'], r['city'], r['state'], r['postcode'] = a
        except Exception:  # noqa: BLE001 — enrichment is best-effort
            pass
    rows.sort(key=lambda r: (r['match_type'] != 'UNMAPPED', str(r['po'])))
    return rows, n_unmapped, has_mapping


def save_draft(spec: FlowSpec, token: str, note: str = '') -> bool:
    """Park the WHOLE run as a 'Review Later' draft — kept intact (raw file(s) +
    cached preview + any per-line decisions), NOT locked/recorded. Use when a
    line can't be decided yet (e.g. a CP/master needs correcting first). The
    operator later reopens it from Drafts, re-validates (picks up the fix) and
    confirms — never re-uploaded. Mirrors Online B2B's ``meta['draft']`` scheme
    (no separate table). Returns False if already locked/missing."""
    import datetime as _dt
    meta = load_meta(spec, token)
    if not meta or meta.get('locked'):
        return False
    meta['draft'] = True
    meta['draft_at'] = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta['draft_note'] = str(note or '')[:300]
    _write_meta(spec, token, meta)
    return True


def collect_drafts(spec: FlowSpec) -> list[dict]:
    """All parked 'Review Later' runs for this flow as API-ready dicts (token,
    marketplace, warehouse, when, note, PO count, file count). Read-only — the
    fat data layer behind the Drafts list. Sorted newest-first."""
    rows: list[dict] = []
    base = _root(spec)
    if not base.exists():
        return rows
    for d in base.iterdir():
        if not d.is_dir():
            continue
        meta = load_meta(spec, d.name)
        if not meta or not meta.get('draft') or meta.get('locked'):
            continue
        npos = 0
        cache = d / 'preview.json'
        if cache.exists():
            try:
                cached = json.loads(cache.read_text(encoding='utf-8'))
                res = cached.get('payload', cached)
                npos = len(res.get('headers') or [])
            except Exception:  # noqa: BLE001
                pass
        rows.append({
            'token': d.name,
            'marketplace': meta.get('marketplace', ''),
            'warehouse': meta.get('warehouse', ''),
            'draft_at': meta.get('draft_at', ''),
            'note': meta.get('draft_note', ''),
            'pos': npos,
            'files': len(meta.get('files') or []),
        })
    rows.sort(key=lambda r: r['draft_at'], reverse=True)
    return rows


def review_context(spec: FlowSpec, token: str, meta: dict) -> dict:
    """Build the full template context for the shared ``review.html``."""
    payload = preview(spec, token, meta)
    _overlay_decisions(payload, meta)
    # Parity with Online B2B (additive, from the resolved payload only):
    # a per-SKU rollup + per-PO CLEAN/AFFECTED status + ship-to mapping report.
    payload['sku_rows'] = _sku_rows(payload.get('lines'))
    n_clean_po, n_affected_po = _mark_clean(payload.get('headers'), payload.get('lines'))
    _tag_warehouse(payload, meta)          # WH each PO maps to → Orders-tab column
    mapping_report, n_unmapped, has_mapping = _mapping_report(payload.get('headers'))
    payload['mapping_report'] = mapping_report
    # Qty-weighted KPIs (parity with the Online B2B review): total qty on
    # affected lines, and the share of qty that is clean (OK). Derived from the
    # already-computed lines — no processor/engine change.
    _AFF = {'MISMATCH', 'NOT_IN_MASTER'}
    total_qty = affected_qty = 0
    for ln in (payload.get('lines') or []):
        q = int(ln.get('qty') or 0)
        total_qty += q
        if (ln.get('status') or '') in _AFF:
            affected_qty += q
    ok_qty_pct = round((total_qty - affected_qty) * 100 / total_qty, 1) if total_qty else 100.0
    return {
        'affected_qty': affected_qty, 'ok_qty_pct': ok_qty_pct,
        'n_clean_po': n_clean_po, 'n_affected_po': n_affected_po,
        'n_unmapped': n_unmapped, 'has_mapping': has_mapping,
        'spec': spec,
        'title': spec.title,
        'segment': spec.segment,
        'base_template': spec.base_template,
        'intro': spec.intro,
        'caps': spec.caps_map(),
        'slots': spec.slot_map(),
        'token': token,
        'meta': meta,
        'r': payload,
        's': payload.get('summary', {}),
        'locked': bool(meta.get('locked')),
        # 'Review Later' — this run was parked as a draft. Shows a banner + lets
        # the reopen flow re-validate (pick up any master fix) before confirm.
        'is_draft': bool(meta.get('draft')) and not meta.get('locked'),
        'draft_note': meta.get('draft_note', ''),
        'run_id': meta.get('run_id'),
        'has_download': bool(meta.get('output_path')),
        'warehouses': spec.warehouses,
        'marketplaces': spec.marketplaces,
        # URL names (used as `{% url u_confirm token %}` — variable name form)
        'u_upload': spec.urls['upload'], 'u_review': spec.urls['review'],
        'u_confirm': spec.urls['confirm'], 'u_decision': spec.urls['decision'],
        'u_discard': spec.urls['discard'], 'u_download': spec.urls['download'],
        'u_back': spec.urls['back'], 'u_dashboard': spec.urls['dashboard'],
        # Optional 'Export review to Excel' (no SO numbers) — only if the spec
        # wired an 'export' URL. Available pre- AND post-lock (it's just the
        # on-screen review data, for eyeballing in Excel before you commit).
        'u_export': spec.urls.get('export'),
        # Optional 'Save for Review Later' + Drafts list — only when the spec
        # wired them (MT / GT Mass). Absent → the button/link simply don't render.
        'u_save_later': spec.urls.get('save_later'),
        'u_drafts': spec.urls.get('drafts'),
    }


def confirm(spec: FlowSpec, token: str, meta: dict, actions: dict | None = None
            ) -> dict:
    """Run the processor's confirm (DB write) and lock the token on success."""
    result = spec.processor(meta).confirm(actions or meta.get('decisions', {}))
    if result.get('ok') and result.get('run_id'):
        meta['locked'] = True
        meta['run_id'] = result['run_id']
        if result.get('output_path'):
            meta['output_path'] = result['output_path']
        _write_meta(spec, token, meta)
    return result


def download_path(spec: FlowSpec, token: str) -> Path | None:
    """The workbook to serve on the review 'download' link. Prefers the workbook
    already produced at confirm; otherwise, if the channel's processor exposes a
    ``workbook()`` method (``'download'`` cap), generates it on demand so the
    operator can download during review — before confirm."""
    meta = load_meta(spec, token) or {}
    p = meta.get('output_path')
    if p and Path(p).exists():
        return Path(p)
    gen = getattr(spec.processor(meta), 'workbook', None)
    if callable(gen):
        try:
            out = gen()
            return Path(out) if out else None
        except Exception:  # noqa: BLE001
            return None
    return None


def _count_pos(path) -> int:
    """POs in a workbook = data rows of its 'Headers (SO)' / 'Headers (TO)'
    sheet(s), one row per document. Best-effort — never blocks a download."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        n = 0
        for sh in ('Headers (SO)', 'Headers (TO)'):
            if sh in wb.sheetnames:
                n += sum(1 for r in wb[sh].iter_rows(min_row=2, values_only=True)
                         if r and r[0] not in (None, ''))
        wb.close()
        return n
    except Exception:  # noqa: BLE001
        return 0


def download_name(spec: FlowSpec, meta: dict, path) -> str:
    """Uniform, self-describing download filename — SAME scheme as Online B2B's
    ``_lot_name``: ``{Mp}_{N}po_{dd-mm-YYYY_HHMMSS}_{review|completed}.xlsx`` so
    the lot size + run timestamp are obvious and Review never clashes with
    Completed. ``Mp`` = the channel code (MT, e.g. RBL/HG) or the flow key (GT
    Mass). Replaces the raw ``tmpXXXX_…`` temp name on the review download."""
    import re
    import time
    p = Path(path)
    kind = 'completed' if meta.get('locked') else 'review'
    m = re.search(r'(\d{2}-\d{2}-\d{4}_\d{6})', p.name)
    ts = (m.group(1) if m else
          time.strftime('%d-%m-%Y_%H%M%S', time.localtime(p.stat().st_mtime)))
    mp = str(meta.get('marketplace') or spec.key or 'SO').replace(' ', '')
    return f"{mp}_{_count_pos(p)}po_{ts}_{kind}.xlsx"


def export_review_xlsx(spec: FlowSpec, token: str, meta: dict) -> Path | None:
    """Build a plain **review** workbook (Orders + Line items exactly as shown on
    the review page) so the operator can eyeball it in Excel BEFORE confirming.

    Deliberately carries **no SO numbers** — SO numbers are assigned only at
    Confirm & Record (the real SO workbook is the post-lock ⬇ download). This is
    just the on-screen data in a sheet, available any time. Channel-agnostic —
    every flow gets it for free once its spec wires an 'export' URL."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    payload = preview(spec, token, meta)
    _overlay_decisions(payload, meta)
    headers = payload.get('headers', []) or []
    lines = payload.get('lines', []) or []

    wb = Workbook()
    band = PatternFill('solid', fgColor='1f2a5a')
    white = Font(bold=True, color='FFFFFF')
    right = Alignment(horizontal='right')

    def _sheet(ws, cols, rows, keys):
        ws.append(cols)
        for c in ws[1]:
            c.fill = band
            c.font = white
        for r in rows:
            ws.append([r.get(k, '') for k in keys])
        for i, w in enumerate(_widths(cols), start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'

    def _widths(cols):
        return [max(10, min(48, len(c) + 6)) for c in cols]

    ws1 = wb.active
    ws1.title = 'Orders'
    _sheet(ws1, ['PO', 'Location', 'Type', 'Items', 'Qty', 'Value'], headers,
           ['po', 'location', 'order_type', 'items', 'qty', 'order_value'])

    ws2 = wb.create_sheet('Line items')
    _sheet(ws2,
           ['PO', 'Item No', 'EAN', 'Description', 'Qty', 'Unit price', 'MRP',
            'Status', 'Exception', 'Decision'],
           [{**ln, 'decision': (ln.get('decision') or {}).get('action', '')}
            for ln in lines],
           ['po', 'item_no', 'ean', 'description', 'qty', 'unit_price',
            'our_mrp', 'status', 'exception_label', 'decision'])
    # right-align the numeric columns on both sheets
    for col in ('D', 'E', 'F'):
        for cell in ws1[col][1:]:
            cell.alignment = right
    for col in ('E', 'F', 'G'):
        for cell in ws2[col][1:]:
            cell.alignment = right

    out = _dir(spec, token) / f"{spec.key}_review_{token}.xlsx"
    try:
        wb.save(str(out))
    except Exception:  # noqa: BLE001
        return None
    return out
