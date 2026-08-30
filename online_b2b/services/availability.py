"""
online_b2b.services.availability
================================

**Order Availability Checker** — paste order number(s) straight from the Excel
tracker → for each, pull its recorded line items from the DB → check every SKU
against the CURRENT inventory snapshot in the *mapped* warehouse (auto-resolved
from the order's warehouse/marketplace, with a manual override).

Read-only. Reuses the existing building blocks — NO duplication of stock or
order logic:
  * :func:`order_db._conn` + the ``order_lines_full`` view (recorded lines) and
    ``order_headers`` (warehouse + marketplace).
  * :mod:`inventory_store` — ``current_stock_map`` (available qty per item for a
    warehouse), ``resolve_order_wh`` (WH auto-map incl. MP overrides), warehouse
    metadata.
"""

from __future__ import annotations

import logging
import re

from . import inventory_store as inv
from .order_db import _conn

_log = logging.getLogger(__name__)

# Split pasted text into tokens: Excel copy is tab/space/newline separated; also
# tolerate commas and semicolons. Order numbers themselves keep their internal
# '/' and '-' (e.g. 'SO/RL/07/280728'), so we only break on whitespace + , ;.
_SPLIT = re.compile(r'[\s,;]+')


def parse_order_nos(text) -> list[str]:
    """Pasted blob → de-duplicated, order-preserving list of order numbers."""
    seen: set[str] = set()
    out: list[str] = []
    for tok in _SPLIT.split(str(text or '').strip()):
        t = tok.strip()
        if t and t not in seen:
            seen.add(t)
            out.append(t)
    return out


def _q(x):
    """Qty display — whole numbers as int, else 1-dp float."""
    x = float(x or 0)
    return int(x) if x == int(x) else round(x, 1)


def _line_status(found: bool, ordered: float, available: float) -> str:
    # OOS covers BOTH "no sellable stock at all" and "pick face empty/negative" —
    # one label, since both mean nothing can be filled (avoids the confusing
    # NO STOCK vs OOS split). ``found`` kept for signature compatibility.
    if available <= 0:
        return 'OOS'               # nothing available (empty or not stocked)
    if available < ordered:
        return 'SHORT'             # partial cover
    return 'OK'                    # fully coverable


def check_orders(order_nos, wh_override: str = '') -> dict:
    """For each order number: resolve its warehouse (override wins, else the
    order's own mapped WH) and compare each recorded line's qty to the available
    stock there. Returns a render-ready dict::

        {ok, orders:[{po, marketplace, wh, wh_short, wh_auto, overridden,
                      lines:[{item_no, ean, description, ordered, available,
                              fillable, short, status}],
                      ord_qty, fillable_qty, short_qty, fill_pct, skus}],
         not_found:[po,...], override, wh_options, summary}
    """
    override_code = inv.wh_normalize(wh_override) if (wh_override or '').strip() else ''

    # One stock map per distinct warehouse actually used, cached per snapshot so
    # repeat checks (and the tracker drawer) skip the ~1s reload — see _cached_stock_map.
    _snaps = inv.current_snapshots()
    _stock: dict[str, dict] = {}

    def stock_for(wh: str) -> dict:
        if wh not in _stock:
            _stock[wh] = _cached_stock_map(wh, _snaps)
        return _stock[wh]

    # Manual per-PO warehouse overrides (loaded once; beat the auto map).
    _ov = inv.wh_override_map()

    def snap_ts(wh: str) -> str:
        s = _snaps.get(wh)
        return str(s['captured_at']) if s and s.get('captured_at') else ''

    orders: list[dict] = []
    not_found: list[str] = []
    # SKU-wise aggregate across ALL pasted orders, keyed by (warehouse, item) so
    # cumulative demand for one SKU (spanning several POs) is netted against the
    # single stock figure for its warehouse.
    sku_agg: dict = {}

    with _conn() as (cur, d):
        ph, ot = d['ph'], d['orders']
        # ── Batched header + line fetch (was 2 queries PER pasted PO — a Singapore
        #    round-trip each). Two chunked queries total now, same semantics:
        #    'most recent run per PO supersedes' + lines for that exact run. ──
        uniq_pos = list(dict.fromkeys(order_nos))          # de-dup, keep paste order
        hdr_by_po: dict = {}
        for i in range(0, len(uniq_pos), 900):
            chunk = uniq_pos[i:i + 900]
            marks = ','.join([ph] * len(chunk))
            cur.execute(
                f"SELECT h.po, h.run_id, h.warehouse, h.marketplace_label "
                f"FROM {ot} h JOIN (SELECT po, MAX(run_ts) mx FROM {ot} "
                f"WHERE po IN ({marks}) GROUP BY po) t "
                f"ON h.po=t.po AND h.run_ts=t.mx WHERE h.po IN ({marks})",
                tuple(chunk + chunk))
            for po_v, run_id, wh_raw, mp_label in cur.fetchall():
                hdr_by_po.setdefault(po_v, (run_id, wh_raw, mp_label))   # first wins on tie
        pairs = [(po, hdr_by_po[po][0]) for po in uniq_pos if po in hdr_by_po]
        lines_by_po: dict = {}
        for i in range(0, len(pairs), 900):
            chunk = pairs[i:i + 900]
            marks = ','.join([f'({ph},{ph})'] * len(chunk))
            flat = [v for pair in chunk for v in pair]
            cur.execute(
                f"SELECT po, item_no, ean, description, qty, unit_price, our_landing "
                f"FROM order_lines_full WHERE (po, run_id) IN ({marks}) "
                f"ORDER BY po, item_no", tuple(flat))
            for row in cur.fetchall():
                lines_by_po.setdefault(row[0], []).append(row[1:])

        for po in order_nos:
            hdr = hdr_by_po.get(po)
            if not hdr:
                not_found.append(po)
                continue
            run_id, wh_raw, mp_label = hdr
            # auto map → manual per-PO override → page-level WH override (widest wins)
            wh_auto = inv.effective_order_wh(po, wh_raw, mp_label, mp_label, _ov)
            wh = override_code or wh_auto
            sm = stock_for(wh)

            lrows: list[dict] = []
            ord_qty = fill_qty = short_qty = 0.0
            ord_val = fill_val = short_val = 0.0
            for item_no, ean, desc, qty, unit_price, our_landing in lines_by_po.get(po, []):
                key = str(item_no or '').strip()
                q = float(qty or 0)
                # per-unit value: inc-GST landing preferred, else unit price (CP).
                uv = float(our_landing or 0) or float(unit_price or 0)
                found = key in sm
                avail = float(sm.get(key, 0) or 0)
                avail_eff = avail if avail > 0 else 0.0   # oversold (<0) → 0 fillable
                fillable = min(q, avail_eff)
                short = q - fillable                      # ≤ ordered, always
                lo_v, lf_v, ls_v = q * uv, fillable * uv, short * uv
                lrows.append({
                    'item_no': key, 'ean': str(ean or ''),
                    'description': str(desc or ''),
                    'ordered': _q(q),
                    'available': _q(avail), 'fillable': _q(fillable), 'short': _q(short),
                    'unit_value': round(uv, 2),
                    'ordered_value': round(lo_v, 2), 'fillable_value': round(lf_v, 2),
                    'short_value': round(ls_v, 2),
                    'status': _line_status(found, q, avail),
                })
                ord_qty += q; fill_qty += fillable; short_qty += short
                ord_val += lo_v; fill_val += lf_v; short_val += ls_v
                # accumulate SKU-wise (demand summed; availability captured once)
                a = sku_agg.get((wh, key))
                if a is None:
                    a = sku_agg[(wh, key)] = {
                        'item_no': key, 'ean': str(ean or ''),
                        'description': str(desc or ''), 'wh': wh,
                        'wh_short': inv.wh_short(wh), 'ordered': 0.0,
                        'ordered_value': 0.0, 'available': avail, 'found': found,
                        'pos': set()}
                a['ordered'] += q
                a['ordered_value'] += lo_v
                a['pos'].add(po)
                if not a['ean'] and ean:
                    a['ean'] = str(ean)
                if not a['description'] and desc:
                    a['description'] = str(desc)
            orders.append({
                'po': po, 'marketplace': str(mp_label or ''),
                'wh': wh, 'wh_short': inv.wh_short(wh), 'stock_as_of': snap_ts(wh),
                'wh_auto': wh_auto, 'wh_auto_short': inv.wh_short(wh_auto),
                'overridden': bool(override_code) and override_code != wh_auto,
                'lines': lrows, 'skus': len(lrows),
                'ord_qty': _q(ord_qty), 'fillable_qty': _q(fill_qty), 'short_qty': _q(short_qty),
                'fill_pct': round(fill_qty / ord_qty * 100, 1) if ord_qty else 0.0,
                'ord_value': round(ord_val, 2), 'fillable_value': round(fill_val, 2),
                'short_value': round(short_val, 2),
                'fill_val_pct': round(fill_val / ord_val * 100, 1) if ord_val else 0.0,
                'has_value': ord_val > 0,
                'fully': short_qty <= 0 and len(lrows) > 0,
            })

    # SKU-wise fill rate — demand netted against stock (worst-case truth when the
    # same SKU is pulled by multiple pasted POs from one warehouse).
    skus: list[dict] = []
    for a in sku_agg.values():
        o = a['ordered']
        av = a['available']
        ov = a['ordered_value']
        avail_eff = av if av > 0 else 0.0                 # oversold (<0) → 0 fillable
        fillable = min(o, avail_eff)
        short = o - fillable                              # ≤ ordered, always
        uv = (ov / o) if o else 0.0                       # avg per-unit value
        fv, sv = fillable * uv, short * uv
        skus.append({
            'item_no': a['item_no'], 'ean': a['ean'],
            'description': a['description'], 'wh': a['wh'], 'wh_short': a['wh_short'],
            'pos': len(a['pos']), 'ordered': _q(o), 'available': _q(av),
            'fillable': _q(fillable), 'short': _q(short),
            'ordered_value': round(ov, 2), 'fillable_value': round(fv, 2),
            'short_value': round(sv, 2),
            'fill_pct': round(fillable / o * 100, 1) if o else 0.0,
            'fill_val_pct': round(fv / ov * 100, 1) if ov else 0.0,
            'status': _line_status(a['found'], o, av),
        })
    skus.sort(key=lambda x: (-x['short'], -x['ordered']))

    tot_ord = sum(o['ord_qty'] for o in orders)
    tot_fill = sum(o['fillable_qty'] for o in orders)
    tot_ordv = sum(o['ord_value'] for o in orders)
    tot_fillv = sum(o['fillable_value'] for o in orders)
    # "inventory as of" — per warehouse actually used (short tag → captured_at).
    used_wh = {(o['wh_short'], snap_ts(o['wh'])) for o in orders if snap_ts(o['wh'])}
    wh_stock_as_of = {short: ts for short, ts in used_wh}
    summary = {
        'orders': len(orders), 'not_found': len(not_found),
        'skus': len(skus),
        'ord_qty': _q(tot_ord), 'fillable_qty': _q(tot_fill),
        'short_qty': _q(sum(o['short_qty'] for o in orders)),
        'fill_pct': round(tot_fill / tot_ord * 100, 1) if tot_ord else 0.0,
        'ord_value': round(tot_ordv, 2), 'fillable_value': round(tot_fillv, 2),
        'short_value': round(sum(o['short_value'] for o in orders), 2),
        'fill_val_pct': round(tot_fillv / tot_ordv * 100, 1) if tot_ordv else 0.0,
        'has_value': tot_ordv > 0,
        'fully': sum(1 for o in orders if o['fully']),
        'wh_stock_as_of': wh_stock_as_of,
        'stock_as_of': next(iter(wh_stock_as_of.values()), '') if len(wh_stock_as_of) == 1 else '',
    }
    # Bin classification for the warehouse(s) touched — which bins are INCLUDED
    # vs EXCLUDED (so the WH team can see WHY an item reads short: its stock may
    # sit in an excluded return/QC bin, an unclassified 'new' bin, or a negative
    # pick face). Aggregated per bin (not per item — that detail isn't stored).
    bins: dict = {}
    _DEC = {'include': 'INCLUDED', 'exclude': 'EXCLUDED', 'new': 'NEW (unclassified)'}
    for wh in {o['wh'] for o in orders}:
        snap = _snaps.get(wh)
        if not snap:
            continue
        try:
            rows = inv.bin_audit(snap['snapshot_id'])
        except Exception:  # noqa: BLE001
            _log.exception('bin_audit(%s) failed — no bin drill-down for this WH',
                           snap.get('snapshot_id'))
            rows = []
        bins[inv.wh_short(wh)] = [{
            'bin': r.get('bin_code', ''), 'zone': r.get('zone_code', ''),
            'decision': _DEC.get(r.get('decision', ''), str(r.get('decision', '')).upper()),
            'lines': r.get('lines', 0), 'qty': _q(r.get('qty', 0)),
        } for r in rows]

    return {'ok': True, 'orders': orders, 'skus': skus, 'not_found': not_found,
            'bins': bins, 'override': override_code, 'wh_options': inv.WAREHOUSES,
            'summary': summary}


# Per-process stock-map cache for the Tracker billing column, keyed by
# (warehouse, snapshot_id). A new bin upload makes a NEW snapshot_id → new key →
# automatic refresh (stays fully dynamic), so this only skips the redundant ~1s
# reload of the SAME current snapshot across repeated filter changes. At most one
# entry per warehouse (stale snapshots are evicted), so it stays tiny.
_BILL_STOCK_CACHE: dict = {}


def _cached_stock_map(wh, snaps) -> dict:
    """``inv.current_stock_map(wh)`` cached per ``(warehouse, snapshot_id)`` for
    this process. A new bin upload changes the snapshot_id → new key → automatic
    refresh (stays dynamic); the SAME current snapshot is reused across calls,
    skipping the ~1s reload. Shared by ``tracker_billing`` + ``check_orders`` (so
    the tracker column, the drill drawer, and the Availability Checker all warm
    the same cache). At most one entry per warehouse."""
    snap = snaps.get(wh)
    sid = snap.get('snapshot_id') if snap else None
    key = (wh, sid)
    sm = _BILL_STOCK_CACHE.get(key)
    if sm is None:
        sm = inv.current_stock_map(wh)
        for k in [k for k in _BILL_STOCK_CACHE if k[0] == wh]:
            del _BILL_STOCK_CACHE[k]          # evict this WH's stale snapshot
        _BILL_STOCK_CACHE[key] = sm
    return sm


def tracker_billing(pos) -> dict:
    """Batched **estimated billing** for many POs at once — powers the Tracker's
    'Est. Billing' column. For each PO: the ₹ value its mapped warehouse's CURRENT
    inventory snapshot can fill right now (billable-from-stock). Same fill logic as
    the Availability Checker — ``min(ordered, available) × inc-GST landing`` — the
    same WH resolution (auto map → manual shift override) and the same current
    snapshots. Gross per-PO (each PO sees full stock; no cross-PO allocation), so a
    row's number matches what the Availability Checker shows for that PO alone.

    Returns ``{po: {est_value, ord_value, short_value, fill_pct, wh_short,
    stock_as_of, no_stock}}`` (``no_stock=True`` when the PO's WH has no current
    snapshot → the UI shows '—', never a misleading ₹0). Read-only; never raises."""
    pos = [str(p).strip() for p in (pos or []) if str(p or '').strip()]
    out: dict = {}
    if not pos:
        return out
    try:
        from .triangular_validation import _line_val, _is_dropped
        snaps = inv.current_snapshots()
        ov = inv.wh_override_map()
        stock_cache: dict = {}

        def stock_for(wh):
            if wh not in stock_cache:
                stock_cache[wh] = _cached_stock_map(wh, snaps)
            return stock_cache[wh]

        po_wh: dict = {}          # po → resolved inventory WH code
        hfull: dict = {}          # po → (header qty, header order_value) = the FULL totals
        acc: dict = {}            # po → {est, ord, short, excl_q, excl_v}
        CHUNK = 900               # keep the IN(...) list well under driver limits
        with _conn() as (cur, d):
            ph = d['ph']
            for i in range(0, len(pos), CHUNK):
                batch = pos[i:i + CHUNK]
                marks = ','.join([ph] * len(batch))
                # latest run per PO in this batch (+ its raw WH / marketplace)
                cur.execute(
                    f"SELECT h.po, h.run_id, h.warehouse, h.marketplace_label, "
                    f"h.qty, h.order_value FROM order_headers h JOIN ("
                    f"  SELECT po, MAX(run_ts) mx FROM order_headers "
                    f"  WHERE po IN ({marks}) GROUP BY po) t "
                    f"ON h.po=t.po AND h.run_ts=t.mx", tuple(batch))
                hdr = {}
                for po, run_id, wh_raw, mp, hqty, hval in cur.fetchall():
                    po = str(po)
                    hdr[po] = str(run_id)
                    hfull[po] = (int(hqty or 0), float(hval or 0))
                    po_wh[po] = inv.effective_order_wh(
                        po, wh_raw, str(mp or ''), str(mp or ''), ov)
                if not hdr:
                    continue
                # recorded lines for exactly those POs' latest runs (+ the fields
                # for the Full → Uploaded(D365) → Excluded qty/value breakdown)
                cur.execute(
                    f"SELECT po, run_id, item_no, qty, unit_price, our_landing, "
                    f"gst_code, status, action FROM order_lines_full "
                    f"WHERE po IN ({marks})", tuple(batch))
                for (po, run_id, item_no, qty, unit_price, our_landing,
                     gst_code, status, action) in cur.fetchall():
                    po = str(po)
                    if hdr.get(po) != str(run_id):
                        continue                       # superseded run → skip
                    sm = stock_for(po_wh.get(po, ''))
                    q = float(qty or 0)
                    uv = float(our_landing or 0) or float(unit_price or 0)
                    avail = float(sm.get(str(item_no or '').strip(), 0) or 0)
                    fillable = min(q, avail if avail > 0 else 0.0)
                    a = acc.setdefault(po, {'est': 0.0, 'ord': 0.0, 'short': 0.0,
                                            'excl_q': 0.0, 'excl_v': 0.0})
                    a['est'] += fillable * uv
                    a['ord'] += q * uv
                    a['short'] += (q - fillable) * uv
                    # Excluded (dropped) qty/value only — Full/Total comes from the
                    # header (authoritative; matches the existing columns), and
                    # Uploaded = Full − Excluded, so the three always tie out.
                    if _is_dropped(status, action):
                        a['excl_q'] += q
                        a['excl_v'] += _line_val(our_landing, unit_price, gst_code, int(q))
        for po, a in acc.items():
            wh = po_wh.get(po, '')
            snap = snaps.get(wh)
            fq, fv = hfull.get(po, (0, 0.0))
            eq, ev = int(a['excl_q']), round(a['excl_v'], 2)
            out[po] = {
                'est_value': round(a['est'], 2),
                'ord_value': round(a['ord'], 2),
                'short_value': round(a['short'], 2),
                'fill_pct': round(a['est'] / a['ord'] * 100, 1) if a['ord'] else 0.0,
                'wh_short': inv.wh_short(wh),
                'stock_as_of': (snap.get('captured_at') if snap else None),  # datetime → template formats
                'no_stock': snap is None,     # WH has no current snapshot → show '—'
                # Full → Uploaded(D365) → Excluded breakdown, anchored to the header
                # totals so it ties out to the existing Qty / Value columns.
                'full_qty': fq, 'excl_qty': eq, 'upl_qty': fq - eq,
                'full_value': round(fv, 2), 'excl_value': ev, 'upl_value': round(fv - ev, 2),
            }
    except Exception:  # noqa: BLE001 — additive; a billing hiccup must never break the tracker
        return out
    return out


def wh_scenarios(order_nos) -> dict:
    """**Best-warehouse comparison.** For the pasted orders, work out what the
    fill rate WOULD be if the whole batch shipped from each warehouse (AHD / BLR /
    North), so the highest-fill facility can be chosen. Read-only; never raises.

    Demand is gathered ONCE (latest run per PO), then netted against *each*
    warehouse's current stock independently. Returns three lenses::

        {ok, warehouses:[short,...], best_wh, best_wh_code,
         overall:[{wh, wh_short, code, fill_pct, fill_val_pct, fillable_qty,
                   short_qty, oos_skus, skus, as_of, best}],
         po_wise:[{po, mp, ord_qty, ord_value, by_wh:{AHD:{fill_pct,..}}, best_wh}],
         sku_wise:[{item_no, ean, description, ordered, pos,
                    by_wh:{AHD:{available, fillable, short, oos}}, best_wh}],
         total_qty, total_value, n_skus, not_found}

    ``overall`` and ``sku_wise`` net cumulative demand (worst-case truth when one
    SKU is pulled by several POs); ``po_wise`` scores each PO on its own lines.
    """
    order_nos = parse_order_nos(order_nos) if isinstance(order_nos, str) else list(order_nos or [])
    if not order_nos:
        return {'ok': False, 'error': 'Paste at least one order number.'}

    whs = inv.WAREHOUSES                       # [{code, name, short}, ...]
    stock = {w['code']: inv.current_stock_map(w['code']) for w in whs}
    snaps = inv.current_snapshots()
    ovmap = inv.wh_override_map()              # per-PO manual shifts

    def _avail(code: str, item: str) -> float:
        v = float(stock[code].get(item, 0) or 0)
        return v if v > 0 else 0.0             # oversold (<0) → 0 fillable

    # ── gather demand ONCE: per-PO lines + per-SKU netted across POs ──
    po_demand: list[dict] = []                 # [{po, mp, lines:[{item, uv, qty}], ord_qty, ord_val}]
    sku: dict = {}                             # item -> {item, ean, desc, qty, val, pos:set}
    not_found: list[str] = []
    with _conn() as (cur, d):
        ph, ot = d['ph'], d['orders']
        # ── Batched header + line fetch (was 2 queries PER pasted PO). Same
        #    'latest run per PO' semantics as check_orders — 2 chunked queries. ──
        uniq_pos = list(dict.fromkeys(order_nos))
        hdr_by_po: dict = {}
        for i in range(0, len(uniq_pos), 900):
            chunk = uniq_pos[i:i + 900]
            marks = ','.join([ph] * len(chunk))
            cur.execute(
                f"SELECT h.po, h.run_id, h.marketplace_label, h.warehouse "
                f"FROM {ot} h JOIN (SELECT po, MAX(run_ts) mx FROM {ot} "
                f"WHERE po IN ({marks}) GROUP BY po) t "
                f"ON h.po=t.po AND h.run_ts=t.mx WHERE h.po IN ({marks})",
                tuple(chunk + chunk))
            for po_v, run_id, mp_label, wh_raw in cur.fetchall():
                hdr_by_po.setdefault(po_v, (run_id, mp_label, wh_raw))
        pairs = [(po, hdr_by_po[po][0]) for po in uniq_pos if po in hdr_by_po]
        lines_by_po: dict = {}
        for i in range(0, len(pairs), 900):
            chunk = pairs[i:i + 900]
            marks = ','.join([f'({ph},{ph})'] * len(chunk))
            flat = [v for pair in chunk for v in pair]
            cur.execute(
                f"SELECT po, item_no, ean, description, qty, unit_price, our_landing "
                f"FROM order_lines_full WHERE (po, run_id) IN ({marks}) "
                f"ORDER BY po, item_no", tuple(flat))
            for row in cur.fetchall():
                lines_by_po.setdefault(row[0], []).append(row[1:])

        for po in order_nos:
            hdr = hdr_by_po.get(po)
            if not hdr:
                not_found.append(po)
                continue
            run_id, mp_label, wh_raw = hdr
            lines: list[dict] = []
            oq = ov = 0.0
            for item_no, ean, desc, qty, unit_price, our_landing in lines_by_po.get(po, []):
                key = str(item_no or '').strip()
                q = float(qty or 0)
                uv = float(our_landing or 0) or float(unit_price or 0)
                lines.append({'item': key, 'qty': q, 'uv': uv})
                oq += q; ov += q * uv
                s = sku.get(key)
                if s is None:
                    s = sku[key] = {'item': key, 'ean': str(ean or ''),
                                    'desc': str(desc or ''), 'qty': 0.0, 'val': 0.0,
                                    'pos': set()}
                s['qty'] += q; s['val'] += q * uv; s['pos'].add(po)
                if not s['ean'] and ean:
                    s['ean'] = str(ean)
                if not s['desc'] and desc:
                    s['desc'] = str(desc)
            cur_code = inv.effective_order_wh(po, wh_raw, mp_label, mp_label, ovmap)
            ovr = ovmap.get(str(po))
            po_demand.append({
                'po': po, 'mp': str(mp_label or ''), 'lines': lines,
                'ord_qty': oq, 'ord_val': ov,
                'cur_code': cur_code, 'cur_short': inv.wh_short(cur_code),
                'shifted': bool(ovr),
                'orig_short': inv.wh_short(ovr['orig_warehouse']) if ovr and ovr.get('orig_warehouse') else '',
                'shifted_by': ovr.get('actor', '') if ovr else '',
                'shifted_at': ovr.get('updated_at', '') if ovr else ''})

    if not po_demand:
        return {'ok': False, 'error': 'None of those orders were found in recorded runs.',
                'not_found': not_found}

    total_q = sum(p['ord_qty'] for p in po_demand)
    total_v = sum(p['ord_val'] for p in po_demand)

    def _as_of(code: str) -> str:
        s = snaps.get(code)
        return str(s['captured_at']) if s and s.get('captured_at') else ''

    # ── OVERALL per WH (SKU-netted) ──
    overall: list[dict] = []
    for w in whs:
        code = w['code']; fq = fv = 0.0; oos = 0
        for it, s in sku.items():
            av = _avail(code, it)
            f = min(s['qty'], av)
            fq += f
            fv += f * (s['val'] / s['qty'] if s['qty'] else 0.0)
            if av <= 0:
                oos += 1
        overall.append({
            'wh': w['name'], 'wh_short': w['short'], 'code': code,
            'fill_pct': round(fq / total_q * 100, 1) if total_q else 0.0,
            'fill_val_pct': round(fv / total_v * 100, 1) if total_v else 0.0,
            'fillable_qty': _q(fq), 'short_qty': _q(total_q - fq),
            'ord_qty': _q(total_q), 'oos_skus': oos, 'skus': len(sku),
            'as_of': _as_of(code),
        })
    # best = highest qty fill%, tie-break on value fill%
    best = max(overall, key=lambda x: (x['fill_pct'], x['fill_val_pct'])) if overall else None
    best_wh = best['wh_short'] if best else ''
    best_code = best['code'] if best else ''
    for o in overall:
        o['best'] = (o['code'] == best_code)

    # ── PO-wise: each PO scored independently in each WH ──
    po_wise: list[dict] = []
    for p in po_demand:
        by: dict = {}
        for w in whs:
            code = w['code']; fq = fv = 0.0
            for l in p['lines']:
                f = min(l['qty'], _avail(code, l['item']))
                fq += f; fv += f * l['uv']
            by[w['short']] = {
                'fill_pct': round(fq / p['ord_qty'] * 100, 1) if p['ord_qty'] else 0.0,
                'fill_val_pct': round(fv / p['ord_val'] * 100, 1) if p['ord_val'] else 0.0,
                'fillable_qty': _q(fq), 'short_qty': _q(p['ord_qty'] - fq),
            }
        bw = max(by, key=lambda k: by[k]['fill_pct']) if by else ''
        po_wise.append({'po': p['po'], 'mp': p['mp'], 'ord_qty': _q(p['ord_qty']),
                        'ord_value': round(p['ord_val'], 2), 'by_wh': by, 'best_wh': bw,
                        'cur_wh': p['cur_short'], 'cur_code': p['cur_code'],
                        'shifted': p['shifted'], 'orig_wh': p['orig_short'],
                        'shifted_by': p['shifted_by'], 'shifted_at': p['shifted_at'],
                        # a shift is worth suggesting only if the best WH beats the
                        # current one by a real margin (avoid noise on ties)
                        'can_improve': bw != p['cur_short'] and by.get(bw, {}).get('fill_pct', 0)
                                       > by.get(p['cur_short'], {}).get('fill_pct', -1) + 0.5})

    # ── SKU-wise: availability + fillable in each WH ──
    sku_wise: list[dict] = []
    for it, s in sku.items():
        by = {}
        for w in whs:
            av = _avail(w['code'], it)
            by[w['short']] = {'available': _q(av), 'fillable': _q(min(s['qty'], av)),
                              'short': _q(s['qty'] - min(s['qty'], av)), 'oos': av <= 0}
        bw = max(by, key=lambda k: by[k]['fillable']) if by else ''
        sku_wise.append({'item_no': it, 'ean': s['ean'], 'description': s['desc'],
                         'ordered': _q(s['qty']), 'pos': len(s['pos']),
                         'by_wh': by, 'best_wh': bw})
    sku_wise.sort(key=lambda x: -x['ordered'])

    return {'ok': True, 'warehouses': [w['short'] for w in whs],
            'wh_codes': [w['code'] for w in whs], 'not_found': not_found,
            'best_wh': best_wh, 'best_wh_code': best_code,
            'overall': overall, 'po_wise': po_wise, 'sku_wise': sku_wise,
            'total_qty': _q(total_q), 'total_value': round(total_v, 2),
            'n_skus': len(sku), 'n_orders': len(po_demand)}


def fulfilment_risk(date_from='', date_to='', marketplace='') -> dict:
    """**Fulfilment risk over a period.** Aggregates demand (the *latest* run per
    PO within the upload-date window) per (mapped warehouse, item), then nets it
    against the CURRENT inventory snapshot in that warehouse. Ranks the at-risk
    SKUs (OOS / short) by unfulfillable value, and rolls up a per-warehouse and
    overall fill rate. Same WH-resolution + stock logic as :func:`check_orders`
    (no duplication). Read-only; never raises.

    Note: demand is historical (the chosen window); inventory is *right now* — so
    this answers "of this period's demand, how much could we ship with today's
    stock, and which SKUs are the biggest gaps?"."""
    out = {'ok': False, 'date_from': date_from, 'date_to': date_to,
           'marketplace': marketplace, 'marketplaces': [], 'rows': [], 'by_wh': [],
           'stock_as_of': {},
           'summary': {'skus': 0, 'at_risk': 0, 'oos': 0, 'short': 0,
                       'demand_qty': 0, 'fill_qty': 0, 'short_qty': 0,
                       'demand_value': 0.0, 'fill_value': 0.0, 'short_value': 0.0,
                       'fill_pct': 0.0, 'fill_val_pct': 0.0}}
    try:
        _snaps = inv.current_snapshots()
        _stock: dict[str, dict] = {}

        def stock_for(wh: str) -> dict:
            if wh not in _stock:
                _stock[wh] = inv.current_stock_map(wh)
            return _stock[wh]

        with _conn() as (cur, d):
            ph = d['ph']
            cur.execute(
                "SELECT DISTINCT marketplace_label FROM order_headers "
                "WHERE marketplace_label IS NOT NULL AND marketplace_label<>'' "
                "ORDER BY marketplace_label")
            out['marketplaces'] = [r[0] for r in cur.fetchall()]

            where, args = [], []
            if date_from:
                where.append(f"DATE(run_ts) >= {ph}"); args.append(date_from)
            if date_to:
                where.append(f"DATE(run_ts) <= {ph}"); args.append(date_to)
            if marketplace:
                where.append(f"marketplace_label={ph}"); args.append(marketplace)
            wsql = ' AND '.join(where) if where else '1=1'

            # latest run per PO in the window → its lines' demand, grouped by
            # (item, PO, raw warehouse, marketplace). PO is in the key so a manual
            # per-PO warehouse shift moves that order's demand to its new WH.
            cur.execute(
                "SELECT l.item_no, MAX(l.description), hh.po, hh.warehouse, "
                "hh.marketplace_label, SUM(l.qty), "
                "SUM(l.qty*COALESCE(NULLIF(l.our_landing,0), l.unit_price, 0)) "
                "FROM order_lines_full l JOIN ("
                "  SELECT hd.po, hd.run_id, hd.warehouse, hd.marketplace_label "
                "  FROM order_headers hd JOIN ("
                f"    SELECT po, MAX(run_ts) mx FROM order_headers WHERE {wsql} GROUP BY po"
                "  ) t ON hd.po=t.po AND hd.run_ts=t.mx"
                ") hh ON l.po=hh.po AND l.run_id=hh.run_id "
                "GROUP BY l.item_no, hh.po, hh.warehouse, hh.marketplace_label",
                tuple(args))
            raw = cur.fetchall()

        # aggregate demand per (resolved WH, item) — override wins per PO
        _ov = inv.wh_override_map()
        agg: dict = {}
        for item_no, desc, po, wh_raw, mp_label, qty, val in raw:
            item = str(item_no or '').strip()
            if not item:
                continue
            wh = inv.effective_order_wh(po, wh_raw, mp_label, mp_label, _ov)
            a = agg.get((wh, item))
            if a is None:
                a = agg[(wh, item)] = {'item_no': item, 'description': str(desc or ''),
                                       'wh': wh, 'wh_short': inv.wh_short(wh),
                                       'qty': 0.0, 'value': 0.0}
            a['qty'] += float(qty or 0)
            a['value'] += float(val or 0)

        rows: list[dict] = []
        by_wh: dict = {}
        tot_q = tot_v = fill_q = fill_v = short_q = short_v = 0.0
        oos = short_n = 0
        for a in agg.values():
            stock = stock_for(a['wh'])
            avail = float(stock.get(a['item_no'], 0) or 0)
            q, v = a['qty'], a['value']
            avail_eff = avail if avail > 0 else 0.0
            fillable = min(q, avail_eff)
            short = q - fillable
            uv = (v / q) if q else 0.0
            fv, sv = fillable * uv, short * uv
            st = _line_status(a['item_no'] in stock, q, avail)
            tot_q += q; tot_v += v; fill_q += fillable; fill_v += fv
            short_q += short; short_v += sv
            w = by_wh.setdefault(a['wh'], {'wh': a['wh'], 'wh_short': a['wh_short'],
                                           'skus': 0, 'demand_qty': 0.0, 'demand_value': 0.0,
                                           'short_qty': 0.0, 'short_value': 0.0})
            w['skus'] += 1; w['demand_qty'] += q; w['demand_value'] += v
            w['short_qty'] += short; w['short_value'] += sv
            if st == 'OK':
                continue                      # at-risk table lists only OOS / SHORT
            (oos, short_n) = (oos + 1, short_n) if st == 'OOS' else (oos, short_n + 1)
            rows.append({
                'item_no': a['item_no'], 'description': a['description'],
                'wh': a['wh'], 'wh_short': a['wh_short'],
                'demand': _q(q), 'available': _q(avail), 'fillable': _q(fillable),
                'short': _q(short), 'demand_value': round(v, 2),
                'short_value': round(sv, 2),
                'fill_pct': round(fillable / q * 100, 1) if q else 0.0,
                'status': st, '_sv': sv, '_short': short})
        rows.sort(key=lambda r: (-r['_sv'], -r['_short']))
        for r in rows:
            r.pop('_sv', None); r.pop('_short', None)

        wh_list = []
        for w in by_wh.values():
            dq, dv = w['demand_qty'], w['demand_value']
            w['demand_qty'] = _q(dq); w['short_qty'] = _q(w['short_qty'])
            sv = round(w['short_value'], 2)
            w['demand_value'] = round(dv, 2); w['short_value'] = sv
            w['fill_val_pct'] = round((dv - sv) / dv * 100, 1) if dv else 0.0
            s = _snaps.get(w['wh'])
            w['stock_as_of'] = str(s['captured_at']) if s and s.get('captured_at') else ''
            wh_list.append(w)
        wh_list.sort(key=lambda x: -x['short_value'])

        out['rows'] = rows
        out['by_wh'] = wh_list
        out['stock_as_of'] = {w['wh_short']: w['stock_as_of'] for w in wh_list if w['stock_as_of']}
        out['summary'] = {
            'skus': len(agg), 'at_risk': len(rows), 'oos': oos, 'short': short_n,
            'demand_qty': _q(tot_q), 'fill_qty': _q(fill_q), 'short_qty': _q(short_q),
            'demand_value': round(tot_v, 2), 'fill_value': round(fill_v, 2),
            'short_value': round(short_v, 2),
            'fill_pct': round(fill_q / tot_q * 100, 1) if tot_q else 0.0,
            'fill_val_pct': round(fill_v / tot_v * 100, 1) if tot_v else 0.0}
        out['ok'] = True
    except Exception as e:  # noqa: BLE001
        out['error'] = f"{type(e).__name__}: {e}"
    return out


def fulfilment_readiness(marketplace='', horizon=0) -> dict:
    """**Projected OTIF / Fulfilment Readiness** — self-contained (no delivery
    feed). Scoped to **OPEN orders** (exp_date >= today) so it isn't time-
    degenerate: a readiness score only means something for orders still to be
    shipped, compared to CURRENT stock. Per open order, at ORDER level:

      * In-Full (projected) — every line fully coverable from current stock in the
        order's mapped warehouse.
      * Accurate            — no MISMATCH / NOT_IN_MASTER line on the PO.
      * Ready               — In-Full AND Accurate (set up to be a perfect order).

    On-time risk is carried by **due-date urgency**: at-risk orders (not in-full)
    due soonest are the fire. ``horizon`` (days) optionally limits to orders due
    within N days (0 = all open). NOT actual OTIF (that needs a delivery feed).
    Reuses WH resolution + current_stock_map + the standard zone map. Read-only;
    never raises."""
    import datetime as _dt
    from .order_db import _IN_STATES, _IN_ZONES

    out = {'ok': False, 'marketplace': marketplace, 'horizon': horizon,
           'marketplaces': [],
           'summary': {'orders': 0, 'in_full': 0, 'accurate': 0, 'ready': 0,
                       'at_risk': 0, 'urgent': 0, 'in_full_pct': 0.0,
                       'accurate_pct': 0.0, 'ready_pct': 0.0},
           'by_channel': [], 'by_zone': [], 'at_risk_orders': []}

    def pct(a, b):
        return round(a / b * 100, 1) if b else 0.0

    def nk(s):
        return re.sub(r'[^a-z0-9]', '', str(s or '').lower())

    try:
        today = _dt.date.today()
        _stock: dict[str, dict] = {}

        def stock_for(wh):
            if wh not in _stock:
                _stock[wh] = inv.current_stock_map(wh)
            return _stock[wh]

        with _conn() as (cur, d):
            ph = d['ph']
            cur.execute("SELECT DISTINCT marketplace_label FROM order_headers "
                        "WHERE marketplace_label IS NOT NULL AND marketplace_label<>'' "
                        "ORDER BY marketplace_label")
            out['marketplaces'] = [r[0] for r in cur.fetchall()]

            cur.execute('SELECT del_location,ship_to,name,state FROM ship_to_mapping')
            loc2st = {}
            for dl, shp, nm, st in cur.fetchall():
                if not st:
                    continue
                for kk in (dl, shp, nm):
                    if kk:
                        loc2st.setdefault(nk(kk), str(st).strip())

            # OPEN orders only: exp_date >= today (+ optional due-horizon)
            where = [f"hd.exp_date >= {ph}"]
            args = [today.isoformat()]
            if horizon and int(horizon) > 0:
                where.append(f"hd.exp_date <= {ph}")
                args.append((today + _dt.timedelta(days=int(horizon))).isoformat())
            if marketplace:
                where.append(f"hd.marketplace_label={ph}")
                args.append(marketplace)
            wsql = ' AND '.join(where)

            cur.execute(
                "SELECT l.po, l.item_no, l.qty, l.status, hh.warehouse, "
                "hh.marketplace_label, hh.exp_date, hh.location "
                "FROM order_lines_full l JOIN ("
                "  SELECT hd.po, hd.run_id, hd.warehouse, hd.marketplace_label, "
                "         hd.exp_date, hd.location FROM order_headers hd JOIN ("
                "    SELECT po, MAX(run_ts) mx FROM order_headers GROUP BY po"
                "  ) t ON hd.po=t.po AND hd.run_ts=t.mx "
                f"  WHERE {wsql}"
                ") hh ON l.po=hh.po AND l.run_id=hh.run_id", tuple(args))

            _ov = inv.wh_override_map()
            orders: dict = {}
            for po, item, qty, status, wh_raw, mp, exp, loc in cur.fetchall():
                o = orders.get(po)
                if o is None:
                    o = orders[po] = {'mp': mp or 'Other',
                                      'wh': inv.effective_order_wh(po, wh_raw, mp, mp, _ov),
                                      'exp': exp, 'loc': loc, 'lines': [],
                                      'accurate': True}
                o['lines'].append((str(item or '').strip(), float(qty or 0)))
                if str(status or '') in ('MISMATCH', 'NOT_IN_MASTER'):
                    o['accurate'] = False

        ch, zo = {}, {}
        S = out['summary']
        at_risk = []
        for po, o in orders.items():
            stock = stock_for(o['wh'])
            in_full = True
            short_lines = 0
            for item, q in o['lines']:
                av = float(stock.get(item, 0) or 0)
                if (av if av > 0 else 0) < q:
                    in_full = False
                    short_lines += 1
            accurate = o['accurate']
            ready = in_full and accurate
            ed = o['exp'].date() if hasattr(o['exp'], 'date') else o['exp']
            days_left = (ed - today).days if ed else None
            urgent = (not in_full) and days_left is not None and days_left <= 2
            stname = _IN_STATES.get((loc2st.get(nk(o['loc'])) or '').upper(),
                                    loc2st.get(nk(o['loc'])))
            zone = _IN_ZONES.get(stname, '(Unzoned)') if stname else '(Unzoned)'

            S['orders'] += 1
            S['in_full'] += in_full
            S['accurate'] += accurate
            S['ready'] += ready
            if not ready:
                S['at_risk'] += 1
            if urgent:
                S['urgent'] += 1
            c = ch.setdefault(o['mp'], {'name': o['mp'], 'orders': 0, 'ready': 0})
            c['orders'] += 1; c['ready'] += ready
            z = zo.setdefault(zone, {'zone': zone, 'orders': 0, 'ready': 0})
            z['orders'] += 1; z['ready'] += ready
            if not ready:
                at_risk.append({'po': po, 'mp': o['mp'], 'wh_short': inv.wh_short(o['wh']),
                                'zone': zone, 'days_left': days_left,
                                'exp': str(ed) if ed else '',
                                'why': ('short' if not in_full else '') +
                                       ('+exception' if not accurate else '') if (not in_full or not accurate) else '',
                                'short_lines': short_lines, 'lines': len(o['lines'])})

        N = S['orders']
        S['in_full_pct'] = pct(S['in_full'], N)
        S['accurate_pct'] = pct(S['accurate'], N)
        S['ready_pct'] = pct(S['ready'], N)
        for c in ch.values():
            c['ready_pct'] = pct(c['ready'], c['orders'])
        for z in zo.values():
            z['ready_pct'] = pct(z['ready'], z['orders'])
        at_risk.sort(key=lambda x: (x['days_left'] if x['days_left'] is not None else 9999))
        out['by_channel'] = sorted(ch.values(), key=lambda x: -x['orders'])
        out['by_zone'] = sorted(zo.values(), key=lambda x: -x['orders'])
        out['at_risk_orders'] = at_risk[:50]
        out['ok'] = True
    except Exception as e:  # noqa: BLE001
        out['error'] = f"{type(e).__name__}: {e}"
    return out


# ── Styled Excel export (same look as our other workbook downloads) ──────────

def to_workbook(data: dict):
    """Render an availability result (from :func:`check_orders`) into a styled
    multi-sheet .xlsx — Summary · By Order (PO-SKU line items) · By SKU · Not
    Found — matching our standard workbook styling (navy header, frozen header,
    auto-filter). Returns a ``BytesIO`` positioned at 0."""
    import datetime as _dt
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    NAVY = PatternFill('solid', fgColor='1A237E')
    HEADF = Font(bold=True, color='FFFFFF')
    CENTER = Alignment(horizontal='center', vertical='center')
    OK = PatternFill('solid', fgColor='DCFCE7')
    SHORT = PatternFill('solid', fgColor='FEF3C7')
    OOS = PatternFill('solid', fgColor='FEE2E2')
    NOST = PatternFill('solid', fgColor='EAECEF')
    STFILL = {'OK': OK, 'SHORT': SHORT, 'OOS': OOS, 'NO STOCK': NOST}
    CUR = '[$₹-4009]#,##,##0.00'         # Indian-grouped rupee currency format

    def _cur_cols(ws, cols):
        """Apply the rupee format to `cols` (1-indexed) for every data row."""
        for row in range(2, ws.max_row + 1):
            for col in cols:
                ws.cell(row=row, column=col).number_format = CUR

    def _sheet(ws, heads, widths):
        ws.append(heads)
        for c in ws[1]:
            c.font = HEADF; c.fill = NAVY; c.alignment = CENTER
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    def _finish(ws, ncols):
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    s = data.get('summary', {})
    wb = openpyxl.Workbook()

    # 1) Summary
    ws = wb.active; ws.title = 'Summary'
    ws['A1'] = 'AVAILABILITY CHECK'; ws['A1'].font = Font(bold=True, size=14, color='1A237E')
    asof = s.get('wh_stock_as_of') or {}
    pairs = [
        ('Generated', f"{_dt.datetime.now():%d-%b-%Y %H:%M}"),
        ('Orders checked', s.get('orders', 0)),
        ('Not found', s.get('not_found', 0)),
        ('Distinct SKUs', s.get('skus', 0)),
        ('Fully coverable orders', s.get('fully', 0)),
        ('— Quantity —', ''),
        ('Ordered qty', s.get('ord_qty', 0)),
        ('Fillable qty', s.get('fillable_qty', 0)),
        ('Short qty', s.get('short_qty', 0)),
        ('Fill rate % (qty)', s.get('fill_pct', 0)),
        ('— Value (₹) —', ''),
        ('Ordered value', s.get('ord_value', 0)),
        ('Fillable value', s.get('fillable_value', 0)),
        ('Short value', s.get('short_value', 0)),
        ('Fill rate % (value)', s.get('fill_val_pct', 0)),
        ('Inventory as of', ' | '.join(f"{k}: {v}" for k, v in asof.items()) or '—'),
    ]
    _cur_labels = {'Ordered value', 'Fillable value', 'Short value'}
    for i, (k, v) in enumerate(pairs, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=v)
        if k in _cur_labels:
            cell.number_format = CUR
    ws.column_dimensions['A'].width = 24; ws.column_dimensions['B'].width = 40

    # 2) PO Summary — one row per order, fill rate qty AND value
    ws = wb.create_sheet('PO Summary')
    _sheet(ws, ['Order No', 'Marketplace', 'Warehouse', 'SKUs', 'Ordered Qty',
                'Fillable Qty', 'Short Qty', 'Fill % (Qty)', 'Ordered ₹',
                'Fillable ₹', 'Short ₹', 'Fill % (Val)', 'Fully'],
           [20, 18, 12, 7, 11, 11, 10, 11, 14, 14, 13, 11, 8])
    for o in data.get('orders', []):
        ws.append([o['po'], o['marketplace'], o['wh_short'], o['skus'],
                   o['ord_qty'], o['fillable_qty'], o['short_qty'], o['fill_pct'],
                   o['ord_value'], o['fillable_value'], o['short_value'],
                   o['fill_val_pct'], 'YES' if o['fully'] else 'NO'])
    _finish(ws, 13)
    _cur_cols(ws, [9, 10, 11])          # Ordered ₹ · Fillable ₹ · Short ₹

    # 3) By Order — PO-SKU line items (qty + value)
    ws = wb.create_sheet('By Order Lines')
    _sheet(ws, ['Order No', 'Marketplace', 'Warehouse', 'Item No', 'EAN',
                'Description', 'Ordered', 'Available', 'Fillable', 'Short',
                'Unit ₹', 'Ordered ₹', 'Fillable ₹', 'Short ₹', 'Status'],
           [20, 18, 12, 12, 16, 40, 9, 10, 9, 8, 10, 13, 13, 12, 11])
    for o in data.get('orders', []):
        for l in o['lines']:
            ws.append([o['po'], o['marketplace'], o['wh_short'], l['item_no'],
                       l['ean'], l['description'], l['ordered'], l['available'],
                       l['fillable'], l['short'], l['unit_value'],
                       l['ordered_value'], l['fillable_value'], l['short_value'],
                       l['status']])
            fill = STFILL.get(l['status'])
            if fill:
                ws.cell(row=ws.max_row, column=15).fill = fill
    _finish(ws, 15)
    _cur_cols(ws, [11, 12, 13, 14])     # Unit ₹ · Ordered ₹ · Fillable ₹ · Short ₹

    # 4) By SKU — aggregated across pasted orders (qty + value)
    ws = wb.create_sheet('By SKU')
    _sheet(ws, ['Item No', 'EAN', 'Description', 'Warehouse', 'POs', 'Ordered',
                'Available', 'Fillable', 'Short', 'Fill % (Qty)', 'Ordered ₹',
                'Fillable ₹', 'Short ₹', 'Fill % (Val)', 'Status'],
           [12, 16, 40, 12, 6, 10, 11, 10, 9, 11, 13, 13, 12, 11, 11])
    for k in data.get('skus', []):
        ws.append([k['item_no'], k['ean'], k['description'], k['wh_short'],
                   k['pos'], k['ordered'], k['available'], k['fillable'],
                   k['short'], k['fill_pct'], k['ordered_value'],
                   k['fillable_value'], k['short_value'], k['fill_val_pct'],
                   k['status']])
        fill = STFILL.get(k['status'])
        if fill:
            ws.cell(row=ws.max_row, column=15).fill = fill
    _finish(ws, 15)
    _cur_cols(ws, [11, 12, 13])         # Ordered ₹ · Fillable ₹ · Short ₹

    # 5) Bin Classification — which bins we INCLUDE vs EXCLUDE per warehouse, so
    #    the WH team can see why an item reads short (stock in an excluded
    #    return/QC bin, an unclassified 'new' bin, or a negative pick face).
    bins = data.get('bins') or {}
    if bins:
        ws = wb.create_sheet('Bin Classification')
        _sheet(ws, ['Warehouse', 'Bin', 'Zone', 'Decision', 'Lines', 'Qty'],
               [12, 32, 18, 20, 8, 12])
        INC = PatternFill('solid', fgColor='DCFCE7')
        EXC = PatternFill('solid', fgColor='FEE2E2')
        NEWF = PatternFill('solid', fgColor='FEF3C7')
        for wh_short, rows in bins.items():
            for b in rows:
                ws.append([wh_short, b['bin'], b['zone'], b['decision'],
                           b['lines'], b['qty']])
                dec = b['decision']
                f = (INC if dec.startswith('INCLUDED')
                     else EXC if dec.startswith('EXCLUDED') else NEWF)
                ws.cell(row=ws.max_row, column=4).fill = f
        _finish(ws, 6)

    # 6) SKU Bins — per-item bin breakdown (where each SKU's stock sits: INCLUDED
    #    pick faces vs EXCLUDED return/QC bins) — explains a short SKU bin-by-bin.
    sku_bins = data.get('sku_bins') or {}
    if sku_bins:
        ws = wb.create_sheet('SKU Bins')
        _sheet(ws, ['Item No', 'Description', 'Warehouse', 'Bin', 'Zone',
                    'Decision', 'Qty'], [12, 40, 12, 30, 16, 14, 12])
        INC = PatternFill('solid', fgColor='DCFCE7')
        EXC = PatternFill('solid', fgColor='FEE2E2')
        NEWF = PatternFill('solid', fgColor='FEF3C7')
        for k in data.get('skus', []):
            blist = (sku_bins.get(k['wh']) or {}).get(k['item_no'], [])
            for b in blist:
                ws.append([k['item_no'], k['description'], k['wh_short'],
                           b['bin'], b['zone'], b['decision'], b['qty']])
                dec = b['decision']
                f = (INC if dec.startswith('INCLUDED')
                     else EXC if dec.startswith('EXCLUDED') else NEWF)
                ws.cell(row=ws.max_row, column=6).fill = f
        _finish(ws, 7)

    # 7) Best-Warehouse comparison — overall + PO-wise + SKU-wise fill in EACH
    #    warehouse (so the best-fill facility, and any manual shifts, are visible).
    sc = data.get('scenarios') or {}
    if sc.get('ok'):
        whs = sc.get('warehouses', [])
        BEST = PatternFill('solid', fgColor='DCFCE7')     # green = the best WH
        boldc = Font(bold=True)

        # 7a) Overall — one row per warehouse, best highlighted
        ws = wb.create_sheet('WH Comparison')
        _sheet(ws, ['Warehouse', 'Fill % (Qty)', 'Fill % (Val)', 'Fillable Qty',
                    'Short Qty', 'OOS SKUs', 'Total SKUs', 'Inventory as of'],
               [14, 12, 12, 12, 11, 10, 11, 22])
        for o in sc.get('overall', []):
            ws.append([o['wh_short'], o['fill_pct'], o['fill_val_pct'],
                       o['fillable_qty'], o['short_qty'], o['oos_skus'],
                       o['skus'], o['as_of']])
            if o.get('best'):
                for c in ws[ws.max_row]:
                    c.fill = BEST; c.font = boldc
        ws.append([])
        ws.append(['Best warehouse (whole batch)', sc.get('best_wh', ''),
                   'Ordered qty', sc.get('total_qty', 0), 'SKUs', sc.get('n_skus', 0)])
        ws.cell(row=ws.max_row, column=1).font = boldc
        _finish(ws, 8)

        # 7b) PO by Warehouse — each PO's fill% in every WH + current/shift state,
        #     with an OVERALL (whole-batch) footer row.
        ws = wb.create_sheet('PO by Warehouse')
        heads = (['Order No', 'Marketplace', 'Ordered Qty', 'Current WH']
                 + [f'{w} Fill%' for w in whs] + ['Best WH', 'Shifted From'])
        _sheet(ws, heads, [20, 18, 11, 11] + [10] * len(whs) + [10, 12])
        best_col = {w: 5 + i for i, w in enumerate(whs)}   # WH% columns start at col 5
        for p in sc.get('po_wise', []):
            row = ([p['po'], p['mp'], p['ord_qty'], p['cur_wh']]
                   + [(p['by_wh'].get(w) or {}).get('fill_pct', 0) for w in whs]
                   + [p['best_wh'], p['orig_wh'] if p.get('shifted') else ''])
            ws.append(row)
            if p['best_wh'] in best_col:
                ws.cell(row=ws.max_row, column=best_col[p['best_wh']]).fill = BEST
        # OVERALL footer (SKU-netted whole-batch fill per WH)
        ov_by = {o['wh_short']: o for o in sc.get('overall', [])}
        foot = (['OVERALL', '', sc.get('total_qty', 0), '']
                + [(ov_by.get(w) or {}).get('fill_pct', 0) for w in whs]
                + [sc.get('best_wh', ''), ''])
        ws.append(foot)
        for c in ws[ws.max_row]:
            c.font = boldc
        if sc.get('best_wh') in best_col:
            ws.cell(row=ws.max_row, column=best_col[sc['best_wh']]).fill = BEST
        _finish(ws, len(heads))

        # 7c) SKU by Warehouse — fillable + available per WH (numeric, sortable)
        ws = wb.create_sheet('SKU by Warehouse')
        heads = (['Item No', 'Description', 'Ordered', 'POs']
                 + [h for w in whs for h in (f'{w} Fillable', f'{w} Avail')]
                 + ['Best WH'])
        _sheet(ws, heads, [12, 40, 9, 6] + [11] * (2 * len(whs)) + [10])
        for k in sc.get('sku_wise', []):
            row = [k['item_no'], k['description'], k['ordered'], k['pos']]
            for w in whs:
                d = k['by_wh'].get(w) or {}
                row += [d.get('fillable', 0), d.get('available', 0)]
            row.append(k['best_wh'])
            ws.append(row)
        _finish(ws, len(heads))

    # 8) Not Found (only if any)
    nf = data.get('not_found', [])
    if nf:
        ws = wb.create_sheet('Not Found')
        _sheet(ws, ['Order No (not in system)'], [30])
        for po in nf:
            ws.append([po])
        _finish(ws, 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
