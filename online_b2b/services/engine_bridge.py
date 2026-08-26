"""
online_b2b.services.engine_bridge
=================================

Class-based bridge that runs the EXISTING ``online_po_processor`` engine as a
library (Option A — sibling on ``sys.path``, wired in settings). The engine is
the frozen backup and is NEVER modified — these classes only orchestrate it.

  * :class:`Processor`        — base: loads masters, runs the engine (single or
    multi-file), applies dedup, and builds the web preview/confirm payloads.
  * :class:`FlipkartProcessor`— Flipkart: always multi-file (one
    ``purchase_order_*.xlsx`` per PO), plus the optional ``purchase-orders-*.csv``
    header → Tracker sheet (``result.flipkart_tracker_rows``). FK Grocery /
    hyperlocal ship-to resolution is master-driven (engine handles it).

Module-level :func:`preview` / :func:`confirm` keep their old signatures (views
call them) but delegate to the right Processor subclass via :func:`processor_for`.
"""

from __future__ import annotations

import os
import re
import time
from pathlib import Path

PILOT_MARKETPLACES = ['Blink', 'Flipkart', 'RK', 'Dmart', 'Zepto', 'Flipkart-TO',
                      'Purplle', 'Swiggy', 'Nykaa', 'Myntra', 'Reliance',
                      'Meesho-TO', 'Bigbasket', 'Firstcry', 'BlinkMP']
# Friendly labels for the upload dropdown where the engine key isn't operator-
# facing. The engine key (value) is unchanged — only the shown text differs.
PILOT_LABELS = {'Flipkart-TO': 'Flipkart Branch', 'Meesho-TO': 'Meesho Branch',
                'Bigbasket': 'Big Basket', 'Firstcry': 'First Cry'}
# Web-side config overrides — applied to a COPY of the marketplace config at run
# time, so the frozen engine config file is never touched. Myntra: compare on CP
# (cost) instead of the landing rate (operator, 2026-07-01).
_WEB_CONFIG_OVERRIDES = {
    # Compare on CP (pre-GST cost), not landing. The vendor's pre-GST CP is the
    # 'List price(FOB+Transport-Excise)' column → fob_col (drives the validation).
    # NOTE: do NOT add ref_fob_col here — it flips the engine into a dual
    # landing+cost check that flags every line. The Vendor Landing (with-GST
    # 'Landing Price') is instead read web-side in build_lines for display only.
    # loc_match='address': Myntra's ship-to is a full postal address, and two
    # warehouse PAIRS share a pincode (Binola/Gurgaon 122413, Bangalore/Hoskote
    # 560067) — so the generic name/substring tiers mis-resolve (e.g. the token
    # 'Haryana' matched BOTH Binola & Gurgaon → wrong ship-to code). Address
    # matching (pincode-gated word-overlap vs the D365 ship_to_mapping addresses)
    # disambiguates them correctly. Requires the Myntra ship_to_mapping to carry
    # the full D365 addresses (rebuilt from the D365 Ship-to Address List).
    'Myntra': {'compare_basis': 'cost', 'compare_label': 'CP',
               'fob_col': 'List price(FOB+Transport-Excise)',
               'loc_match': 'address'},
}
# Per-marketplace column whose value should populate vendor_landing for DISPLAY
# only (read web-side; never fed to the frozen engine's validation). Myntra's
# 'Landing Price' (with GST) — so the tracker shows Vendor CP (List price, no GST)
# AND Vendor Landing (Landing Price, with GST) without breaking the CP validation.
_WEB_VENDOR_LANDING_COL = {'Myntra': 'Landing Price'}
_ISSUE_STATUSES = {'MISMATCH', 'NOT_IN_MASTER'}


def pilot_choices() -> list:
    """``[(engine_key, label)]`` for the upload form — value stays the engine
    marketplace key, label is the operator-friendly name where one is defined."""
    return [(m, PILOT_LABELS.get(m, m)) for m in PILOT_MARKETPLACES]


def _engine_imports():
    """Import the engine (frozen backup) lazily and return its handles in a
    dict. Explicit dict (not locals()) so linters never strip the imports."""
    from online_po_processor.config.marketplaces import (
        DEFAULT_WAREHOUSE,
        MARKETPLACE_CONFIGS,
        WAREHOUSE_CODES,
        WAREHOUSE_DISPLAY_NAMES,
    )
    from online_po_processor.config.paths import (
        get_bundled_mapping_path,
        get_bundled_master_path,
    )
    from online_po_processor.data.mapping_loader import MappingLoader
    from online_po_processor.data.master_loader import MasterLoader
    from online_po_processor.engine.marketplace_engine import MarketplaceEngine
    from online_po_processor.exporter.so_exporter import SOExporter
    return {
        'MARKETPLACE_CONFIGS': MARKETPLACE_CONFIGS,
        'WAREHOUSE_CODES': WAREHOUSE_CODES,
        'DEFAULT_WAREHOUSE': DEFAULT_WAREHOUSE,
        'WAREHOUSE_DISPLAY_NAMES': WAREHOUSE_DISPLAY_NAMES,
        'get_bundled_master_path': get_bundled_master_path,
        'get_bundled_mapping_path': get_bundled_mapping_path,
        'MappingLoader': MappingLoader,
        'MasterLoader': MasterLoader,
        'MarketplaceEngine': MarketplaceEngine,
        'SOExporter': SOExporter,
    }


def warehouse_choices() -> list[str]:
    try:
        return list(_engine_imports()['WAREHOUSE_DISPLAY_NAMES'])
    except Exception:
        return ['AHD']


def default_warehouse() -> str:
    try:
        return _engine_imports()['DEFAULT_WAREHOUSE']
    except Exception:
        return 'AHD'


def default_margin_pct(marketplace: str) -> int:
    try:
        cfg = _engine_imports()['MARKETPLACE_CONFIGS'][marketplace]
        return int(cfg.get('default_margin', 70))
    except Exception:
        return 70


def marketplace_rules() -> list:
    """Per-marketplace rule summary straight from the engine config — drives the
    Rules & Exceptions reference page (so it can never drift from the engine).
    Read-only; never raises."""
    out = []
    try:
        cfgs = _engine_imports()['MARKETPLACE_CONFIGS']
        for name, c in cfgs.items():
            basis = c.get('compare_basis', 'cost')
            parser = (c.get('file_parser') or c.get('pdf_parser')
                      or c.get('source_format') or '')
            gmd = c.get('gst_margin_discount')
            # GST-dependent margin (Reliance): keep% = 1 − discount × (1+GST).
            # Computed from the config so the table can never drift.
            gst_margin = None
            if gmd is not None:
                gst_margin = [{'gst': g, 'pct': round((1 - gmd * (1 + g / 100.0)) * 100, 2)}
                              for g in (18, 12, 5, 0)]
            # Honest margin label: a flat "70%" for normal channels, but a RANGE
            # ("63.42–69% · by GST") for GST-dependent ones (Reliance) so the
            # table never reads as a single flat rate.
            if gst_margin:
                _p = [g['pct'] for g in gst_margin]
                margin_label = f"{min(_p):g}–{max(_p):g}%"
            else:
                margin_label = f"{c.get('default_margin', 70)}%"
            # Category/rule-based margin tiers (e.g. Nykaa: Perfume/Fragrance 69%,
            # Cosmetics 66%) — surfaced so the profile page renders the REAL rule
            # instead of a single flat margin. Read straight from the engine config.
            mr = c.get('margin_rules') or {}
            margin_tiers = None
            if isinstance(mr, dict) and mr.get('rules'):
                tiers = []
                for rr in mr.get('rules', []):
                    crit = []
                    if rr.get('contains'):
                        crit.append('name has ' + ' / '.join(rr['contains']))
                    if rr.get('hsn_prefix'):
                        crit.append('HSN ' + ' / '.join(str(h) for h in rr['hsn_prefix']))
                    if rr.get('excludes'):
                        crit.append('excl. ' + ' / '.join(rr['excludes']))
                    tiers.append({'label': rr.get('label', ''),
                                  'keep_pct': rr.get('keep_pct'),
                                  'criteria': '; '.join(crit)})
                ov = mr.get('item_keep_overrides') or {}
                margin_tiers = {
                    'tiers': tiers,
                    'default_pct': mr.get('default_keep_pct'),
                    'default_label': mr.get('default_label', 'Default'),
                    'overrides': [{'item': k, 'pct': v} for k, v in ov.items()],
                    'override_count': len(ov),
                    'flag_hsn_conflicts': bool(mr.get('flag_hsn_conflicts')),
                }
                # honest compact label: "66/69% · by category"
                pcts = sorted({mr.get('default_keep_pct')} | {t['keep_pct'] for t in tiers})
                margin_label = '/'.join(f'{p:g}' for p in pcts if p is not None) + '% · by category'
            out.append({
                'name': name,
                'party': c.get('party_name', ''),
                'margin': c.get('default_margin', 70),
                'margin_label': margin_label,
                'margin_tiers': margin_tiers,
                'basis': ('Landing — MRP × margin% (pre-GST)' if basis == 'landing'
                          else 'Cost — MRP × margin% ÷ GST (post-GST)'),
                'compare_label': c.get('compare_label', ''),
                'item': c.get('item_resolution', '') or 'mapping',
                'gst_discount': gmd,
                'gst_margin': gst_margin,
                'parser': parser,
                'pilot': name in PILOT_MARKETPLACES,
            })
        out.sort(key=lambda x: (not x['pilot'], x['name']))
    except Exception:  # noqa: BLE001
        pass
    return out


# Curated, operator-facing note on each marketplace's FILE SHAPE (what the team
# actually uploads). The columns themselves come from the engine config below, so
# only the shape (multi-file / extra report / how PO+location are found) is here.
_FMT_NOTES = {
    'Blink': 'One consolidated punch Excel — one row per ordered line.',
    'Flipkart': 'Many purchase_order_*.xlsx (one per PO) + an optional header .csv '
                'that classifies each PO as FK Hyperlocal / FK Grocery.',
    'Flipkart-TO': 'Per-PO Consignment_Details_<PO>.csv files + an optional '
                   'Consignment Visibility Report .csv for the destinations.',
    'Meesho-TO': 'One order-line-items-<PO>[_<city>].csv per order; the destination '
                 'comes from a city token in the filename (MS_BLR → Bengaluru…).',
    'Dmart': 'One or more Avenue/DMart PO PDFs; PO & store read from the PDF.',
    'Reliance': 'One or more Reliance PO PDFs; GST rate is read per line.',
    'Firstcry': 'A FirstCry PO PDF (bordered line-item table).',
    'Swiggy': 'A Swiggy consignment .csv; items arrive as a Swiggy SKU code that '
              'is resolved to an EAN via the channel map.',
    'Myntra': 'A compiled dump .xlsx (a Myntra PO PDF is a fallback).',
    'Purplle': 'A tab-separated .xls / .csv export.',
    'Bigbasket': 'A custom-layout BigBasket Excel.',
    'Nykaa': 'A Nykaa PO Excel (.xlsx).',
    'RK': 'An RK PO Excel (.xlsx).',
    'Zepto': 'A Zepto PO Excel (.xlsx).',
}


def marketplace_formats() -> list:
    """Per-marketplace **file format** reference for the Rules page — the file
    type, how the PO / location / item are found, and the key columns read.
    Columns come straight from the engine config (so they never drift); the
    file-shape note is curated (``_FMT_NOTES``). Read-only; never raises."""
    out = []
    try:
        cfgs = _engine_imports()['MARKETPLACE_CONFIGS']

        def _col(v):
            if isinstance(v, list):
                return ' / '.join(str(x) for x in v)
            if isinstance(v, dict):
                m = v.get('multiply')
                return ('computed (' + ' × '.join(m) + ')') if m else 'computed'
            if v and str(v).startswith('__'):
                return '(from the file)'
            return v or '—'

        for name, c in cfgs.items():
            ir = c.get('item_resolution', '')
            if ir == 'from_ean':
                item_by = f"EAN → master · col “{c.get('ean_col', '')}”"
            elif ir == 'from_swiggy_sku':
                item_by = f"Swiggy SKU → EAN → master · col “{c.get('sku_col') or c.get('ean_col') or 'SkuCode'}”"
            elif ir == 'from_column':
                item_by = f"Item No · col “{c.get('item_col', '')}”"
            else:
                item_by = _col(c.get('ean_col') or c.get('item_col'))
            exts = c.get('accepted_extensions')
            if not exts:                      # some (Flipkart-TO/Meesho) nest it
                for v in c.values():
                    if isinstance(v, dict) and v.get('accepted_extensions'):
                        exts = v['accepted_extensions']
                        break
            if c.get('source_format') == 'pdf':
                ftype = 'PDF'
            elif exts:
                ftype = ' / '.join(exts)
            else:
                ftype = '.xlsx'
            out.append({
                'name': name,
                'file_type': ftype,
                'note': _FMT_NOTES.get(name, ''),
                'po_by': _col(c.get('po_col')),
                'loc_by': _col(c.get('loc_col')),
                'item_by': item_by,
                'qty_by': _col(c.get('qty_col')),
                'cost_by': _col(c.get('fob_col')),
                'mrp_by': _col(c.get('mrp_col')) if c.get('mrp_col') else '—',
                'pilot': name in PILOT_MARKETPLACES,
            })
        out.sort(key=lambda x: (not x['pilot'], x['name']))
    except Exception:  # noqa: BLE001
        pass
    return out


_TEMPLATES_PATH = Path(__file__).with_name('template_samples.json')
# config key → (role label shown in the UI, CSS role class for colour)
_ROLE_KEYS = [
    ('po_col', 'PO', 'po'), ('loc_col', 'Destination', 'loc'),
    ('ean_col', 'Item · EAN', 'item'), ('item_col', 'Item', 'item'),
    ('sku_col', 'Item · SKU', 'item'), ('qty_col', 'Quantity', 'qty'),
    ('fob_col', 'Vendor cost', 'cost'), ('mrp_col', 'MRP', 'mrp'),
    ('po_date_col', 'PO date', 'date'), ('exp_date_col', 'Expiry', 'exp'),
    ('hsn_col', 'HSN', 'hsn'), ('amount_col', 'Amount', 'amt'),
]


def _role_names(v) -> list:
    """Real column name(s) from a config value (str / list); skip synthetic
    ``__…__`` keys and computed dicts."""
    if isinstance(v, str) and not v.startswith('__'):
        return [v]
    if isinstance(v, list):
        return [x for x in v if isinstance(x, str) and not x.startswith('__')]
    return []


def _role_map(cfg: dict) -> dict:
    """``{lower-cased header → (role label, role class)}`` for every column the
    engine actually reads — including columns nested in sub-configs
    (Flipkart-TO / Meesho-TO)."""
    out: dict = {}
    dicts = [cfg] + [v for v in cfg.values() if isinstance(v, dict)]
    for d in dicts:
        for key, role, cls in _ROLE_KEYS:
            for nm in _role_names(d.get(key)):
                out.setdefault(nm.strip().lower(), (role, cls))
    return out


# ── Raw-header role fallback for parser-translated channels ──────────────
# Some marketplaces feed the engine through a bespoke ``file_parser`` that
# TRANSLATES the raw vendor file's headers into the flat column names the
# config consumes (e.g. Flipkart's portal 'Supplier Unit Price' → the config's
# 'COST PRICE'). The template preview shows the REAL raw headers, so to keep
# the highlighting honest we must resolve those raw headers to the SAME roles.
# The parser itself owns that raw→field map in a module-level ``_LABELS`` dict;
# we READ it (never modify the frozen parser) and bridge each canonical parser
# field to a role label/class via ``_ROLE_KEYS`` (the single source of truth).
# This is a FALLBACK ONLY — a fixture column that already matches a config
# name (Blink, Zepto, …) is resolved by ``_role_map`` first and never reaches
# here, so those profiles stay byte-identical.
#
# canonical parser field (a ``_LABELS`` value) → the ``_ROLE_KEYS`` config key
# that defines its role. Fields with no reviewer-facing role (e.g.
# 'description') are omitted → their column stays dulled.
_PARSER_FIELD_ROLE_KEY = {
    'ean': 'ean_col', 'qty': 'qty_col', 'cost': 'fob_col',
    'mrp': 'mrp_col', 'total_amount': 'amount_col', 'amount': 'amount_col',
    'po': 'po_col', 'loc': 'loc_col', 'address': 'loc_col',
    'fsn': 'item_col', 'item': 'item_col', 'sku': 'sku_col',
    'hsn': 'hsn_col',
}
_ROLE_BY_CFG_KEY = {k: (role, cls) for k, role, cls in _ROLE_KEYS}


def _norm_header(s) -> str:
    """Whitespace-free, lower-cased header — mirrors the Flipkart parser's own
    ``_norm`` so raw portal headers line up with its ``_LABELS`` keys."""
    return re.sub(r'\s+', '', str(s if s is not None else '')).lower()


def _parser_label_roles(cfg: dict) -> dict:
    """``{normalized raw header → (role label, role class)}`` for a marketplace
    whose ``file_parser`` module exposes a ``_LABELS`` raw→field map. A
    read-only reflection of the frozen parser; returns ``{}`` for MPs without
    such a parser (so the fallback is a genuine no-op for them → unchanged)."""
    key = cfg.get('file_parser') or cfg.get('pdf_parser')
    if not key:
        return {}
    try:
        import sys
        from online_po_processor.engine.marketplace_engine import PDF_PARSERS
        fn = PDF_PARSERS.get(key)
        mod = sys.modules.get(getattr(fn, '__module__', '') or '')
        labels = getattr(mod, '_LABELS', None)
        if not isinstance(labels, dict):
            return {}
    except Exception:  # noqa: BLE001 — never break the profile on reflection
        return {}
    out: dict = {}
    for raw_norm, field in labels.items():
        ck = _PARSER_FIELD_ROLE_KEY.get(str(field).strip().lower())
        role = _ROLE_BY_CFG_KEY.get(ck) if ck else None
        if role:
            out[str(raw_norm).strip().lower()] = role
    return out


def marketplace_templates() -> dict[str, dict]:
    """Per-marketplace **full file template** for the Rules “See full template”
    page: every column of a real sample file plus a few sample rows, each column
    tagged with the role the engine reads it as (``role`` / ``role_class``) or
    left blank (unused → dulled in the UI). Columns and sample rows are a frozen
    fixture (``template_samples.json``, captured from real files); the used/role
    tagging is computed **live** from the engine config so highlighting never
    drifts from what the parser actually reads. Read-only; never raises."""
    import json
    out: dict = {}
    try:
        samples = json.loads(_TEMPLATES_PATH.read_text(encoding='utf-8'))
    except Exception:  # noqa: BLE001
        return out
    try:
        cfgs = _engine_imports()['MARKETPLACE_CONFIGS']
    except Exception:  # noqa: BLE001
        cfgs = {}
    for name, s in samples.items():
        cfg = cfgs.get(name, {})
        roles = _role_map(cfg)
        # Fallback for parser-translated channels (Flipkart): resolve the raw
        # portal headers via the parser's own _LABELS. Empty for every other
        # MP, so their columns are computed exactly as before (byte-identical).
        label_roles = _parser_label_roles(cfg)
        cols = []
        for col in s.get('columns', []):
            role, cls = roles.get(str(col).strip().lower(), ('', ''))
            if not role and label_roles:
                role, cls = label_roles.get(_norm_header(col), ('', ''))
            cols.append({'name': col, 'role': role, 'role_class': cls,
                         'used': bool(role)})
        # Pre-align sample rows to the column order as cell dicts, so the template
        # can render + style each cell without dynamic dict-key lookup.
        grid = [
            [{'value': r.get(c['name'], ''), 'used': c['used'],
              'role_class': c['role_class']} for c in cols]
            for r in s.get('rows', [])
        ]
        legend, seen = [], set()
        for c in cols:
            if c['used'] and c['role'] not in seen:
                seen.add(c['role'])
                legend.append({'role': c['role'], 'role_class': c['role_class']})
        out[name] = {
            'name': name,
            'file_type': s.get('file_type', ''),
            'sample_file': s.get('sample_file', ''),
            'columns': cols,
            'grid': grid,
            'legend': legend,
            'used': sum(1 for c in cols if c['used']),
            'total': len(cols),
            'pilot': name in PILOT_MARKETPLACES,
            # Optional header block for files that carry a PO/address preamble
            # ABOVE the line-item table (e.g. Flipkart's new portal PO). A list
            # of ``{label, value, used?, role_class?}`` rendered as a labelled
            # key/value band. MPs whose fixture omits it get ``[]`` → the panel
            # renders no header block (unchanged).
            'header_fields': s.get('header_fields', []),
            'header_title': s.get('header_title', ''),
            'header_note': s.get('header_note', ''),
            # Optional EXACT-DOCUMENT-REPLICA mode (currently Flipkart): when the
            # fixture sets ``doc_layout: true`` the profile renders the real PO
            # document (banner / supplier|retailer blocks / payment / line
            # table) from the structured ``doc`` object instead of the generic
            # header+table. Absent/false for every other MP → generic render
            # (byte-identical). Reverting = drop the flag; nothing else changes.
            'doc_layout': bool(s.get('doc_layout')),
            'doc': s.get('doc', {}),
            # EXACT DUMP REPLICA — the file's real cell grid (merge-aware),
            # captured per-MP from a sample; renders verbatim so the panel
            # mirrors the actual received dump (NOT a Flipkart-shaped copy).
            'raw_replica': s.get('raw_replica'),
            # PDF dump facsimile — the real page-1 text lines of THIS MP's PO
            # PDF, rendered as-received (per-MP; not a Flipkart copy).
            'pdf_replica': s.get('pdf_replica'),
        }
    return out


def marketplace_template(name: str):
    """Single marketplace template by engine key, or ``None`` if unknown."""
    return marketplace_templates().get(name)


def location_rules() -> list:
    """Flipkart Origin-Warehouse → sub-marketplace (FK Hyperlocal / FK Grocery)
    locked map — for the Rules page. Read-only."""
    try:
        from online_po_processor.engine.flipkart_tracker import LOCATION_MARKETPLACE
        return sorted(({'loc': k, 'mkt': v} for k, v in LOCATION_MARKETPLACE.items()),
                      key=lambda r: (r['mkt'], r['loc']))
    except Exception:  # noqa: BLE001
        return []


def margin_defaults() -> dict:
    """``{marketplace: default_margin%}`` for every pilot marketplace — so the
    upload form can auto-fill the right margin when the marketplace changes
    (e.g. Flipkart → 77, Blink → 70). Mirrors the desktop GUI's per-marketplace
    pre-fill."""
    return {m: default_margin_pct(m) for m in PILOT_MARKETPLACES}


# ── Processor (base) ────────────────────────────────────────────────────

class Processor:
    """Run the engine for one marketplace and build web payloads. Subclasses
    override the small hooks (``engine_files`` / ``use_multi`` / ``post_process``)."""

    def __init__(self, marketplace, po_paths, warehouse=None, margin_pct=None,
                 ean_fixes=None):
        self.marketplace = marketplace
        self.po_paths = [str(p) for p in (po_paths or [])]
        self.warehouse = warehouse
        self.margin_pct = margin_pct
        # operator's pending EAN corrections for THIS upload ({wrong → correct})
        self.ean_fixes = dict(ean_fixes or {})
        self.warnings: list[str] = []
        # (msg, po) for each ENGINE warning, so preview can drop warnings that
        # belong to already-uploaded (deduped) POs — see _run's post-dedup filter.
        self._engine_warn: list[tuple[str, str]] = []
        # Informational notes (successful, non-problem messages e.g. "compiled N
        # per-PO files") — shown separately from warnings so a clean run doesn't
        # read as "1 warning".
        self.notes: list[str] = []
        self.skipped: list = []
        # Ship-to/locations the engine could NOT map (cust_no + ship_to left
        # blank) — surfaced on the review page's "Mapping" tab so the operator
        # can add the missing mapping. [[never-skip-silently]]
        self.unmapped: list = []
        self.env = None
        self.config = None
        self.result = None

    # ── overridable hooks ───────────────────────────────────────────────
    def engine_files(self) -> list[str]:
        """Files actually fed to the engine."""
        return self.po_paths

    def use_multi(self, config) -> bool:
        """Whether to call ``process_multi`` (mirrors the engine's
        ``_supports_multi_file``: pdf / pdf_parser / file_parser)."""
        return bool(config.get('source_format') == 'pdf'
                    or config.get('pdf_parser') or config.get('file_parser'))

    def post_process(self, result, env) -> None:
        """Hook after a successful engine run (e.g. Flipkart tracker)."""
        # Honor the marketplace config's ``override_unit_price`` in the WEB flow.
        # Since v2.1.3 that flag is only a GUI-checkbox hint; the runtime decision
        # lives on ``result.override_unit_price`` (set from the desktop checkbox),
        # which the headless web path never sets. Without this, BlinkMP — which
        # shares BCPL's 70% vendor record but runs at 75% — posts a BLANK Unit
        # Price and D365 back-fills the wrong 70% cost. Setting the flag makes the
        # D365 exporter stamp our computed 75% CP (MRP × margin% ÷ GST) into col H.
        # Only BlinkMP has the config flag True, so no other channel is affected.
        try:
            if self.config.get('override_unit_price'):
                result.override_unit_price = True
        except Exception:  # noqa: BLE001 — never block the run on this
            pass

    def _accept_deal_exceptions(self, result) -> None:
        """A **deal SKU is the REVISED agreed price** — so accept it as an applied
        exception (same idea as the Blink EPISENSE deal), NOT a mismatch: override
        the CP into the D365 dump and mark the line **OK + labelled '<MP> deal'**,
        even when the marketplace's PO still carries the old/standard price. The
        operator sees "exception applied" (not red MISMATCH); D365 bills the
        negotiated price, not the flat margin. never-skip-silently: the count is
        warned and the CP gap stays visible on the row (+ the deal audit)."""
        accepted = 0
        for so in getattr(result, 'rows', None) or []:
            lbl = str(getattr(so, 'exception_label', '') or '')
            if 'deal' not in lbl.lower():
                continue
            if getattr(so, 'validation_status', '') != 'MISMATCH':
                continue
            deal_cp = getattr(so, 'cost_price_ref', None)
            so.validation_status = 'OK'
            # normalise the 'Deal ≠ vendor' label back to the clean MP deal label
            if '≠' in lbl or ' vs ' in lbl.lower():
                so.exception_label = f'{self.marketplace} deal'
            if not getattr(so, 'forced_unit_price', None) and deal_cp:
                try:
                    so.forced_unit_price = float(deal_cp)
                except (TypeError, ValueError):
                    pass
            accepted += 1
        if accepted:
            self.warnings.append(
                f"Deal SKUs: {accepted} line(s) accepted at the revised negotiated "
                f"price (exception applied → D365 uses the deal price, not the flat "
                f"margin) — shown as an exception, not a mismatch. Verify values on "
                f"the deal audit.")

    def run_engine(self, engine, files, config):
        """Call the engine for this marketplace. Default = standard single-file
        ``process`` (or ``process_multi`` for pdf/multi-file). Subclasses with a
        bespoke ingest (e.g. Flipkart-TO consignments) override this."""
        if self.use_multi(config):
            return engine.process_multi(files, config, margin_pct=self.margin_pct)
        if len(files) > 1:
            self.warnings.append(
                f"{self.marketplace} expects a single file; processing "
                f"'{Path(files[0]).name}', ignored {len(files) - 1} extra.")
        return engine.process(files[0], config, margin_pct=self.margin_pct)

    # ── core run ────────────────────────────────────────────────────────
    def _run(self, skip_dedup=False):
        """Process → ``self.result`` (dedup applied unless ``skip_dedup``).
        Returns an error dict on failure, or ``None`` on success. ``skip_dedup``
        is used when re-running purely to EXPORT (e.g. the D365 package) — the
        ERP file must carry the full upload, not a DB-deduped subset."""
        env = self.env = _engine_imports()
        configs = env['MARKETPLACE_CONFIGS']
        if self.marketplace not in configs:
            return {'ok': False, 'error': f"Unknown marketplace '{self.marketplace}'."}
        files = self.engine_files()
        if not files:
            return {'ok': False, 'error': "No PO file uploaded."}

        config = configs[self.marketplace]
        _ov = _WEB_CONFIG_OVERRIDES.get(self.marketplace)
        if _ov:
            config = {**config, **_ov}   # copy — never mutate the frozen config
        self.config = config
        self.warehouse = self.warehouse or env['DEFAULT_WAREHOUSE']
        if self.margin_pct is None:
            self.margin_pct = config.get('default_margin', 70) / 100.0

        # ── Single source of truth: the DB. The web no longer reads ANY bundled
        # Excel — item master, Ship-To mapping, AND the pricing-override overlays
        # (exceptions + Swiggy deal SKUs) all come from MySQL. (The desktop app
        # keeps its Excel; only the web is retired off it.) Empty table → a clear
        # "seed it first" error, never a silent Excel fallback.
        party = config['party_name']
        from . import item_master_loader as iml
        from . import mapping_store as mstore

        if mstore.table_count() == 0:
            return {'ok': False, 'error': "Ship-To mapping DB is empty — seed it "
                    "on the Ship-To Mapping page first."}
        mapping = mstore.DBMappingLoader()
        loc_count = mapping.load(None, party, [])
        if loc_count == 0:
            return {'ok': False, 'error': f"No mapping locations for "
                    f"'{self.marketplace}' (party '{party}'). Add it on the "
                    f"Ship-To Mapping page."}

        # ── Retire provisional warehouse aliases that now map EXACT ──────────
        # TO channels (Flipkart-TO / Meesho-TO) decode a raw FC code to a
        # friendly Ship-To name via a config ``warehouse_aliases`` bridge, which
        # the engine flags FUZZY (raw ≠ friendly). Once the operator adds the RAW
        # code as an exact Ship-To B2B row, skip the alias so the match reads
        # EXACT. Data-driven + generic — the alias auto-retires per code; the
        # frozen config is never mutated (we prune a deep copy). [[fk-alias-exact]]
        cmode = (self.config.get('consignment_mode') or {})
        if cmode.get('warehouse_aliases'):
            import copy as _copy
            from .order_db import _conn
            with _conn() as (cur, dd):
                ph = dd['ph']
                cur.execute(
                    f"SELECT del_location FROM ship_to_mapping WHERE party={ph}",
                    (party,))
                exact_locs = {str(r[0]).strip().lower() for r in cur.fetchall()}
            retire = {k for k in cmode['warehouse_aliases']
                      if str(k).strip().lower() in exact_locs}
            if retire:
                cm2 = _copy.deepcopy(cmode)
                for k in retire:
                    cm2['warehouse_aliases'].pop(k, None)
                self.config = {**self.config, 'consignment_mode': cm2}
                self.warnings.append(
                    f"{len(retire)} warehouse alias(es) now resolve EXACT from "
                    f"Ship-To B2B — provisional alias retired: "
                    f"{', '.join(sorted(retire))}.")

        if iml.table_count() == 0:
            return {'ok': False, 'error': "Item master DB is empty — upload it on "
                    "the Item Master page first."}
        try:
            master = iml.DBMasterLoader().load_from_db()
        except Exception as e:  # noqa: BLE001
            return {'ok': False,
                    'error': f"Item master load failed: {type(e).__name__}: {e}"}

        # Apply this upload's pending EAN corrections so the wrong EANs resolve
        # on re-validation (DB-backed master exposes add_session_aliases;
        # historical fixes are already merged at load).
        if self.ean_fixes and hasattr(master, 'add_session_aliases'):
            master.add_session_aliases(self.ean_fixes)

        engine = env['MarketplaceEngine'](mapping, master=master)
        # Use self.config (NOT the local `config`) — the warehouse-alias retire
        # block above prunes self.config; passing the stale local would keep the
        # retired aliases and the promoted codes would still read FUZZY.
        result = self.run_engine(engine, files, self.config)

        result.margin_pct = self.margin_pct
        result.warehouse_display = self.warehouse
        result.warehouse_code = env['WAREHOUSE_CODES'].get(self.warehouse, 'PICK')
        import re as _re
        for _po, _loc, msg in result.warnings:
            self.warnings.append(msg)
            self._engine_warn.append((msg, str(_po or '')))
            if 'not found in mapping' in msg:
                m = _re.search(r"Location '(.*?)' not found", msg)
                loc = _loc or (m.group(1) if m else '')
                self.unmapped.append({'location': loc, 'po': str(_po or '')})

        if not result.rows:
            # Surface WHY no rows came out — the engine's warnings hold the real
            # cause (wrong sheet, renamed/missing column, empty file). Without them
            # the operator only sees a vague "No valid rows" and can't fix the file.
            detail = '; '.join(dict.fromkeys(w for w in self.warnings if w))  # de-dup, keep order
            err = "No valid rows extracted from the PO file(s)."
            if detail:
                err += " Likely cause — " + detail[:400]
            else:
                err += (" The sheet/columns didn't match this marketplace's expected "
                        "format — check it's the right marketplace's PO file (see the "
                        "template's exact columns) and that no sheet/column was renamed.")
            return {'ok': False, 'error': err, 'warnings': self.warnings}

        if not skip_dedup:
            try:
                from . import lines_store
                self.skipped = lines_store.web_dedup(result, self.marketplace) or []
            except Exception as e:  # noqa: BLE001
                self.warnings.append(f"Dedup check skipped ({type(e).__name__}: {e}).")
            # Per-PO engine warnings (e.g. "Cost mismatch: Item …") for POs that
            # were ALREADY uploaded in an earlier run are noise here — those POs
            # aren't in Line items (they sit under 'Already uploaded'), so the
            # operator can't act on them. Drop such warnings so the review shows
            # only issues on the NEW POs. Preview only (dedup ran); the full
            # every-line export keeps all warnings (skip_dedup=True there).
            skipped_pos = {str(s.get('po', '')) for s in (self.skipped or [])
                           if s.get('po')}
            if skipped_pos and self._engine_warn:
                self.warnings = [msg for (msg, wpo) in self._engine_warn
                                 if not (wpo and wpo in skipped_pos)]

        self.post_process(result, env)
        # Deal SKUs (Swiggy/Myntra) → accept the revised negotiated price as an
        # applied exception (EPISENSE-style), not a mismatch. Runs for every
        # channel; no-op unless a line carries a '<MP> deal' label. [[deals]]
        self._accept_deal_exceptions(result)
        # Per-PO info notes for already-uploaded (deduped) POs are misleading —
        # that PO isn't in Line items (it sits under 'Already uploaded'), so a
        # note like "Swiggy deal price applied: PO X" points at a line the
        # operator can't see. Drop such notes so notifications cover NEW POs
        # only. Preview only (dedup ran); the every-line export keeps all notes.
        if not skip_dedup and self.notes:
            skipped_pos = {str(s.get('po', '')) for s in (self.skipped or [])
                           if s.get('po')}
            if skipped_pos:
                self.notes = [n for n in self.notes
                              if not any(po in n for po in skipped_pos)]
        self.result = result
        return None

    # ── D365 package from the operator's locked decisions ───────────────
    def _vendor_cp(self, so):
        """The line's vendor CP ('Their CP') for the run's compare basis — the
        value that **Include (their CP)** must stamp into Unit Price. Mirrors
        lines_store: fob_price on a cost basis, else ref_fob_price."""
        basis = getattr(self.result, 'compare_basis', None) or 'landing'
        v = so.fob_price if basis == 'cost' else so.ref_fob_price
        try:
            return round(float(v), 2) if v is not None else None
        except (TypeError, ValueError):
            return None

    def _apply_decisions(self, actions):
        """Copy of ``self.result`` with operator decisions applied to Unit Price
        (via the engine's own ``forced_unit_price``, which the D365 package reads):
        EXCLUDE rows dropped; **INCLUDE (their CP)** stamped to the vendor CP;
        **OVERRIDE (our CP)** stamped to the operator's CP. Originals are NOT
        mutated; the engine is untouched. ``actions`` is keyed ``po|item_no|ean``.
        Inclusion means the line stays in Lines(SO) at the CHOSEN price — their CP
        for INCLUDE, our CP for OVERRIDE."""
        import copy
        actions = actions or {}
        rows = []
        for so in self.result.rows:
            key = f"{so.po_number}|{so.item_no or ''}|{so.ean or ''}"
            dec = actions.get(key) or {}
            act = str(dec.get('action') or '').upper()
            if act == 'EXCLUDE':
                continue
            if act == 'OVERRIDE':
                try:
                    ocp = float(dec.get('override_cp'))
                except (TypeError, ValueError):
                    ocp = None
                if ocp is not None:
                    so = copy.copy(so)
                    so.forced_unit_price = round(ocp, 2)
            elif act == 'INCLUDE':
                vcp = self._vendor_cp(so)      # their CP → Lines(SO) unit price
                if vcp is not None:
                    so = copy.copy(so)
                    so.forced_unit_price = vcp
            rows.append(so)
        r2 = copy.copy(self.result)
        r2.rows = rows
        return r2

    def export_decided_workbook(self, actions=None,
                                exclude_uploaded_run_id=None) -> dict:
        """The full multi-sheet SO Workbook, but with the operator's LOCKED
        decisions applied — EXCLUDE rows dropped, OVERRIDE rows repriced (same
        filter the D365 dump uses). Powers the post-lock **"Download SO Workbook
        (Completed)"**; the review download stays the every-line workbook.

        ``exclude_uploaded_run_id`` (the current run) → drop POs that were
        already uploaded in an EARLIER run from the WHOLE workbook, so the file
        you import to D365 never re-creates a duplicate SO. Only prior-run POs
        drop (keyed on run_id ≠ current), so this run's new POs stay; the dropped
        set is listed on an **'Already Uploaded'** sheet (never silent). The
        Review download is unaffected (it serves the stored full file)."""
        err = self._run(skip_dedup=True)
        if err:
            return err
        # Dedup for the IMPORT file (Completed only): remove POs uploaded in a
        # PRIOR run so D365 doesn't get duplicate SOs. [[completed-dedup]]
        already_rows = []
        if exclude_uploaded_run_id is not None:
            already_pos = self._already_uploaded_pos(exclude_uploaded_run_id)
            if already_pos:
                try:
                    from online_po_processor.auto.history_db import build_tracker_rows
                    trk = {str(t['po']): t for t in build_tracker_rows(self.result)}
                except Exception:  # noqa: BLE001
                    trk = {}
                present = {str(so.po_number) for so in self.result.rows}
                already_rows = [trk.get(p, {'po': p})
                                for p in sorted(already_pos & present)]
                self.result.rows = [so for so in self.result.rows
                                    if str(so.po_number) not in already_pos]
        # Build the FULL workbook (of the remaining rows) so Headers / Summary /
        # Validation / Raw Data stay consistent; then finalize the **Lines (SO)
        # sheet ONLY** to the accepted set (the operator's decisions).
        path = self._export(actions)
        if not path:
            return {'ok': False, 'error': 'Completed workbook export failed.'}
        try:
            self._finalize_lines_so(path, actions or {})
        except Exception as e:  # noqa: BLE001 — never lose the workbook over this
            self.warnings.append(f"Lines (SO) finalize skipped ({type(e).__name__}).")
        if already_rows:
            try:
                self._append_already_uploaded_sheet(path, already_rows)
            except Exception as e:  # noqa: BLE001
                self.warnings.append(f"'Already Uploaded' sheet skipped ({type(e).__name__}).")
        return {'ok': True, 'path': str(path), 'already_uploaded': len(already_rows)}

    def _already_uploaded_pos(self, current_run_id) -> set:
        """PO numbers already recorded for THIS marketplace in some OTHER run
        (run_id ≠ current) — i.e. genuine prior uploads. Read-only; never raises."""
        from .order_db import _conn
        try:
            with _conn() as (cur, d):
                ph = d['ph']
                cur.execute(
                    f"SELECT DISTINCT po FROM order_headers WHERE marketplace={ph} "
                    f"AND run_id <> {ph}", (self.marketplace, current_run_id))
                return {str(r[0]) for r in cur.fetchall() if r[0]}
        except Exception:  # noqa: BLE001
            return set()

    def _append_already_uploaded_sheet(self, path, rows) -> None:
        """Append an 'Already Uploaded' sheet listing the prior-run POs dropped
        from this import file (PO · Marketplace · Location · Qty · Value)."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = openpyxl.load_workbook(path)
        if 'Already Uploaded' in wb.sheetnames:
            del wb['Already Uploaded']
        ws = wb.create_sheet('Already Uploaded')
        cols = [('po', 'PO'), ('market_place', 'Marketplace'),
                ('location', 'Location'), ('order_qty', 'Qty'),
                ('order_value', 'Value')]
        hf = Font(bold=True, color='FFFFFF'); fill = PatternFill('solid', fgColor='B45309')
        for c, (_k, h) in enumerate(cols, 1):
            cell = ws.cell(1, c, h); cell.font = hf; cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
        for r, row in enumerate(rows, 2):
            for c, (k, _h) in enumerate(cols, 1):
                ws.cell(r, c, row.get(k) if isinstance(row, dict) else None)
        ws.cell(len(rows) + 3, 1,
                f"{len(rows)} PO(s) were already uploaded in an earlier run — "
                f"EXCLUDED from this import file to avoid duplicate SOs in D365.")
        for col in ws.columns:
            L = col[0].column_letter
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[L].width = min(w + 2, 60)
        wb.save(path)

    def _finalize_lines_so(self, path, actions) -> None:
        """In-place on the **'Lines (SO)' sheet ONLY** (matched on Document No. +
        item No.): drop EXCLUDEd rows, reprice OVERRIDE rows to the operator's CP,
        and reprice INCLUDE rows to the vendor's CP ('Include their CP'). Every
        other sheet is left exactly as the review workbook, so completed and review
        differ solely in Lines (SO)."""
        excl, over, incl_keys = set(), {}, set()
        for key, dec in (actions or {}).items():
            parts = str(key).split('|')
            if len(parts) < 2:
                continue
            po, item = parts[0], parts[1]
            act = str((dec or {}).get('action') or '').upper()
            if act == 'EXCLUDE':
                excl.add((po, item))
            elif act == 'OVERRIDE':
                try:
                    over[(po, item)] = round(float(dec.get('override_cp')), 2)
                except (TypeError, ValueError):
                    pass
            elif act == 'INCLUDE':
                incl_keys.add((po, item))
        # INCLUDE → stamp the vendor CP (their CP), sourced from the run's rows.
        incl = {}
        if incl_keys and getattr(self, 'result', None) is not None:
            for so in self.result.rows:
                k = (str(so.po_number), str(so.item_no or ''))
                if k in incl_keys:
                    vcp = self._vendor_cp(so)
                    if vcp is not None:
                        incl[k] = vcp
        if not excl and not over and not incl:
            return
        import openpyxl
        wb = openpyxl.load_workbook(path)
        try:
            # Transfer-Order channels (Meesho-TO / Flipkart-TO) write a 'Lines (TO)'
            # sheet with 'Item No.' + 'Transfer Price' columns — NOT 'Lines (SO)' /
            # 'No.' / 'Unit Price'. Handle BOTH; otherwise EXCLUDE/OVERRIDE silently
            # never dropped/repriced the TO lines (the excluded line stayed in the
            # workbook + D365 dump). Document No. = the TO/order id = the decision po.
            sheet = next((s for s in ('Lines (SO)', 'Lines (TO)')
                          if s in wb.sheetnames), None)
            if not sheet:
                return
            ws = wb[sheet]
            hdr = {str(c.value).strip(): i for i, c in enumerate(ws[1], 1)}
            c_po = hdr.get('Document No.')
            c_item = hdr.get('No.') or hdr.get('Item No.')
            c_up = hdr.get('Unit Price') or hdr.get('Transfer Price')
            if not c_po or not c_item:
                return
            drop = []
            for r in range(2, ws.max_row + 1):
                k = (str(ws.cell(r, c_po).value or '').strip(),
                     str(ws.cell(r, c_item).value or '').strip())
                if k in excl:
                    drop.append(r)
                elif k in over and c_up:
                    ws.cell(r, c_up).value = over[k]
                elif k in incl and c_up:
                    ws.cell(r, c_up).value = incl[k]
            for r in reversed(drop):
                ws.delete_rows(r, 1)
            wb.save(path)
        finally:
            wb.close()

    def generate_d365(self, out_path, actions=None) -> dict:
        """Build the ERP-uploadable D365 package at ``out_path`` from LOCKED
        decisions. Reuses the engine's ``export_d365_package`` — engine + the full
        SO Workbook stay untouched. No DB write, no dedup."""
        err = self._run(skip_dedup=True)
        if err:
            return err
        try:
            from online_po_processor.exporter.d365_package import (
                export_d365_package,
            )
            from online_po_processor.exporter.so_exporter import SOExporter
        except Exception as e:  # noqa: BLE001
            return {'ok': False, 'error': f"D365 exporter unavailable: {e}"}
        filtered = self._apply_decisions(actions)
        if not filtered.rows:
            return {'ok': False, 'error': "No lines left after Exclude decisions."}
        is_to = getattr(filtered, 'output_type', 'so') == 'to'
        template = (SOExporter._D365_TO_TEMPLATE if is_to
                    else SOExporter._D365_SO_TEMPLATE)
        if not Path(template).exists():
            return {'ok': False, 'error': "D365 template not bundled on this host."}
        try:
            export_d365_package(filtered, Path(template), Path(out_path))
        except Exception as e:  # noqa: BLE001
            return {'ok': False, 'error': f"D365 export failed: "
                    f"{type(e).__name__}: {e}"}
        excluded = len(self.result.rows) - len(filtered.rows)
        return {'ok': True, 'd365_path': str(out_path),
                'lines': len(filtered.rows), 'excluded': excluded}

    # ── payload builders ────────────────────────────────────────────────
    def _amountless_to(self) -> bool:
        """True for a Transfer-Order marketplace whose dump carries NO vendor
        amount (``amount_col`` omitted) — e.g. Flipkart Branch (Flipkart-TO).
        Its file value is zero, so we derive the inc-GST transfer value from OUR
        master pricing instead."""
        cfg = self.config or {}
        return (getattr(self.result, 'output_type', 'so') == 'to'
                and not cfg.get('amount_col'))

    def _to_value_by_po(self, lines) -> dict:
        """``{po: Σ (our Landing × qty)}`` — the inc-GST transfer value computed
        from our pricing (Landing = MRP × margin% = Cost × (1+GST), so Landing ×
        qty is the GST-inclusive amount). Used to fill order_value for
        amount-less TOs where the dump total is zero."""
        vals: dict = {}
        for ln in lines:
            land = ln.get('our_landing')
            if land is None:
                continue
            po = str(ln.get('po', ''))
            vals[po] = vals.get(po, 0.0) + float(land) * int(ln.get('qty') or 0)
        return {p: round(v, 2) for p, v in vals.items()}

    def _headers(self):
        from online_po_processor.auto.history_db import order_rows_from_result
        headers = order_rows_from_result(
            self.result, self.marketplace, self.warehouse, '')
        # Amount-less TO (Flipkart Branch): the dump carries no price, so the
        # transfer value is OUR calculated pricing (Landing × qty = calculated
        # CP inc-GST), not the vendor's. Fill it so preview, the summary, and the
        # locked DB show the real value — and log it (never silent: the operator
        # must know the amount is computed, not received).
        if self._amountless_to():
            vals = self._to_value_by_po(self._lines())
            filled = 0
            for h in headers:
                v = vals.get(str(h.get('po', '')))
                if v is not None:
                    h['order_value'] = v
                    filled += 1
            if filled:
                total = round(sum(vals.values()), 2)
                label = PILOT_LABELS.get(self.marketplace, self.marketplace)
                self.warnings.append(
                    f"{label}: dump carries no price — order value is COMPUTED "
                    f"from our master pricing (calculated CP inc-GST = Landing × "
                    f"qty) for {filled} PO(s), total ₹{total:,.2f}.")
        return headers

    _EAN_MAX = 20   # matches the order_lines.ean VARCHAR(20) column

    def _lines(self, run_id=None, output_file='', actions=None, as_of=None):
        from . import lines_store
        # Combine historical + this-session EAN fixes so build_lines can swap
        # the wrong EAN → correct on the line and stamp received_ean (audit).
        combined = dict(lines_store.ean_alias_map())
        combined.update(self.ean_fixes)
        rows = lines_store.build_lines(self.result, run_id=run_id,
                                       output_file=output_file, actions=actions,
                                       ean_fixes=combined, as_of=as_of)
        # Defensive: a PDF parser (e.g. Myntra) can occasionally emit a
        # malformed over-long EAN — two EANs concatenated on a page wrap. The DB
        # ean column is VARCHAR(20), so one bad row would crash the WHOLE lock.
        # Cap it (DB-safe) + warn (never silent) — such a line is NOT_IN_MASTER
        # anyway, so the operator corrects it on the Affected/Issues tab.
        for r in rows:
            e = r.get('ean')
            if e and len(str(e)) > self._EAN_MAX:
                msg = (
                    f"Malformed EAN on PO {r.get('po')} item {r.get('item_no')}: "
                    f"'{e}' ({len(str(e))} chars > {self._EAN_MAX}) — looks like two "
                    f"EANs merged on a page wrap. Stored truncated; line is "
                    f"NOT_IN_MASTER — set the correct EAN on the Affected tab.")
                if msg not in self.warnings:   # _lines runs >once per flow
                    self.warnings.append(msg)
                r['ean'] = str(e)[:self._EAN_MAX]
        return rows

    def _summary(self, lines, headers):
        mism = sum(1 for l in lines if l['status'] == 'MISMATCH')
        nim = sum(1 for l in lines if l['status'] == 'NOT_IN_MASTER')
        return {
            'marketplace': self.marketplace, 'warehouse': self.warehouse,
            'margin_pct': round(self.margin_pct * 100, 2),
            'pos': len({l['po'] for l in lines}),
            'lines': len(lines),
            'qty': sum(int(l['qty'] or 0) for l in lines),
            'value': sum(float(h.get('order_value') or 0) for h in headers),
            'mismatch': mism, 'not_in_master': nim, 'affected': mism + nim,
            'skipped': len(self.skipped),
        }

    def _export(self, actions=None):
        # The engine writes the workbook to ``input_file_path.parent/output``.
        # Web uploads already live UNDER MEDIA (b2b_uploads/<token>/), so the
        # engine writes output/ there — web-owned, per-token, and where
        # review_download reads it. ONLY when the input is OUTSIDE media (a
        # direct/script run against the operator's source folder) do we redirect
        # to a web-owned exports dir, so the workbook never lands next to source.
        # ``input_file_path`` is used solely to locate that folder (never embedded
        # in the workbook), so the swap is safe; we restore it afterwards.
        orig = getattr(self.result, 'input_file_path', '') or ''
        redirected = False
        try:
            from django.conf import settings
            media = Path(settings.MEDIA_ROOT).resolve()
            op = Path(orig).resolve() if orig else None
            under_media = bool(op and (op == media or media in op.parents))
            if orig and not under_media:
                exports = media / 'b2b_exports'
                exports.mkdir(parents=True, exist_ok=True)
                self.result.input_file_path = str(exports / Path(orig).name)
                redirected = True
        except Exception:  # noqa: BLE001 — fall back to engine default on any issue
            pass
        path = None
        try:
            path = self.env['SOExporter']().export(self.result, start_time=time.time())
        except Exception as e:  # noqa: BLE001
            self.warnings.append(f"Workbook export issue ({type(e).__name__}: {e}).")
        finally:
            if redirected:
                self.result.input_file_path = orig
        # Append a per-run 'SKU Summary' sheet (web post-process; the engine's
        # workbook + all its sheets are untouched). Best-effort — never fails the
        # export.
        if path:
            try:
                self._append_sku_sheet(str(path))
            except Exception as e:  # noqa: BLE001
                self.warnings.append(f"SKU Summary sheet skipped ({type(e).__name__}).")
            try:
                self._append_tracker_sheet(str(path))
            except Exception as e:  # noqa: BLE001
                self.warnings.append(f"Tracker sheet skipped ({type(e).__name__}).")
            try:
                self._append_excluded_to_summary(str(path), self._lines(actions=actions))
            except Exception as e:  # noqa: BLE001
                self.warnings.append(f"Summary excluded-qty skipped ({type(e).__name__}).")
        return path

    def _append_excluded_to_summary(self, path, lines) -> None:
        """Augment the SO Workbook 'Summary' sheet (per-PO) with three columns:
        **Included Qty**, **Excluded/Dropped Qty** and **Final Qty (to D365)**.
        * Excluded = lines the operator EXCLUDEd, plus still-affected lines
          (MISMATCH / NOT_IN_MASTER not resolved) — qty that WON'T reach D365.
        * Included = flagged lines the operator chose to INCLUDE / OVERRIDE — KEPT
          in the order (part of Final Qty), shown separately so an inclusion never
          looks like a drop.  Final Qty = Total − Excluded (unchanged).
        Lets the operator see, right in the downloaded workbook, exactly what was
        included-by-decision vs dropped (then check it in Validation). Frozen
        exporter + every existing sheet/column stay untouched — we only APPEND."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        excl: dict = {}
        incl: dict = {}   # NEW — flagged (MISMATCH / NOT_IN_MASTER) lines the
                          # operator chose to INCLUDE / OVERRIDE: KEPT in the order,
                          # never dropped. Shown so "we included this" reads clearly
                          # instead of the qty silently sitting inside Final Qty.
        pushed: dict = {}  # NEW — inc-GST VALUE of the lines that actually reach D365
                           # (everything NOT dropped), so the Summary can show the
                           # final pushed amount next to Final Qty. Same value basis
                           # as the SKU Summary sheet: our-CP × GST × qty, or the
                           # already-inc-GST landing when no unit price.
        for l in lines:
            # build_lines() stamps the operator decision as a TOP-LEVEL 'action'
            # (INCLUDE/OVERRIDE/EXCLUDE); keep the legacy 'decision' dict as a
            # fallback. Reading only 'decision' made every INCLUDE/OVERRIDE line
            # show as Excluded (it never reached D365 in the Summary, though the
            # Lines correctly kept it) — the phantom "1 dropped" bug.
            act = (l.get('action') or (l.get('decision') or {}).get('action') or '').upper()
            st = l.get('status')
            dropped = act == 'EXCLUDE' or (st in _ISSUE_STATUSES
                                           and act not in ('INCLUDE', 'OVERRIDE'))
            po = str(l.get('po') or '')
            if dropped:
                excl[po] = excl.get(po, 0) + int(l.get('qty') or 0)
            elif st in _ISSUE_STATUSES and act in ('INCLUDE', 'OVERRIDE'):
                incl[po] = incl.get(po, 0) + int(l.get('qty') or 0)
            if not dropped:                       # value that actually reaches D365
                q = int(l.get('qty') or 0)
                up = l.get('unit_price')
                try:
                    if up not in (None, ''):
                        v = float(up) * self._gst_mult(l.get('gst_code')) * q
                    elif l.get('our_landing') is not None:
                        v = float(l['our_landing']) * q   # landing already inc-GST
                    else:
                        v = 0.0
                except (TypeError, ValueError):
                    v = 0.0
                pushed[po] = pushed.get(po, 0.0) + v
        wb = openpyxl.load_workbook(path)
        if 'Summary' not in wb.sheetnames:
            return
        ws = wb['Summary']
        hdr = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
        if 'PO' not in hdr:
            return
        c_po = hdr.index('PO') + 1
        c_qty = hdr.index('Total Qty') + 1 if 'Total Qty' in hdr else None
        c_amt = hdr.index('Total Amount') + 1 if 'Total Amount' in hdr else None
        amt_fmt = ws.cell(2, c_amt).number_format if c_amt else '#,##0.00'
        _base = ws.max_column
        c_inc, c_exc, c_fin, c_famt, c_stat = (_base + 1, _base + 2, _base + 3,
                                               _base + 4, _base + 5)
        ws.cell(1, c_inc, 'Included Qty')
        ws.cell(1, c_exc, 'Excluded/Dropped Qty')
        ws.cell(1, c_fin, 'Final Qty (to D365)')
        # Final Amount = value (inc-GST) of the lines that actually reach D365 —
        # the "final pushed amount", the money analog of Final Qty.
        ws.cell(1, c_famt, 'Final Amount (to D365)')
        # CLEAN = nothing dropped, 100% of the PO goes to D365 as-is; AFFECTED =
        # some qty was excluded. Lets the operator scan the Summary for clean POs.
        ws.cell(1, c_stat, 'Status')
        for cc in (c_inc, c_exc, c_fin, c_famt, c_stat):
            h = ws.cell(1, cc)
            h.font = Font(bold=True, color='FFFFFF')
            h.fill = PatternFill('solid', fgColor='B45309')
            h.alignment = Alignment(horizontal='center', wrap_text=True)
        # Included Qty is KEPT qty (a decision, not a problem) → green header.
        ws.cell(1, c_inc).fill = PatternFill('solid', fgColor='15803D')
        green = PatternFill('solid', fgColor='E7F6EC')
        red = PatternFill('solid', fgColor='FDE7E7')
        total_e = sum(excl.values())
        total_i = sum(incl.values())
        green_i = PatternFill('solid', fgColor='EAF7EE')
        n_clean = n_aff = 0
        for r in range(2, ws.max_row + 1):
            po = str(ws.cell(r, c_po).value or '')
            tot = ws.cell(r, c_qty).value if c_qty else None
            try:
                totf = int(float(tot))
            except (ValueError, TypeError):
                continue                       # metadata/footer row → leave blank
            if po.upper().startswith('TOTAL'):
                ws.cell(r, c_inc, total_i).font = Font(bold=True)
                ws.cell(r, c_exc, total_e)
                ws.cell(r, c_fin, totf - total_e)
                fa = ws.cell(r, c_famt, round(sum(pushed.values()), 2))
                fa.font = Font(bold=True); fa.number_format = amt_fmt
                sc = ws.cell(r, c_stat, f'{n_clean} CLEAN · {n_aff} AFFECTED')
                sc.font = Font(bold=True)
                continue
            if not po:
                continue
            i = incl.get(po, 0)
            ic = ws.cell(r, c_inc, i)
            if i:
                ic.font = Font(bold=True, color='0A7D33')
                ic.fill = green_i
            e = excl.get(po, 0)
            ws.cell(r, c_exc, e)
            ws.cell(r, c_fin, totf - e)
            fa = ws.cell(r, c_famt, round(pushed.get(po, 0.0), 2))
            fa.number_format = amt_fmt
            clean = (e == 0)
            sc = ws.cell(r, c_stat, 'CLEAN' if clean else 'AFFECTED')
            sc.font = Font(bold=True, color='0A7D33' if clean else 'B91C1C')
            sc.alignment = Alignment(horizontal='center')
            sc.fill = green if clean else red
            if clean:
                n_clean += 1
            else:
                n_aff += 1
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_inc)].width = 15
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_exc)].width = 18
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_fin)].width = 18
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_famt)].width = 22
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_stat)].width = 20
        # Centre-align the Summary DATA cells (what the operator otherwise does by
        # hand). Header row is left as the exporter styled it; the meta/footer note
        # row (non-numeric Total Qty, not the TOTAL row) is skipped so it stays put.
        for r in range(2, ws.max_row + 1):
            is_total = str(ws.cell(r, c_po).value or '').upper().startswith('TOTAL')
            if not is_total:
                try:
                    float(ws.cell(r, c_qty).value if c_qty else None)
                except (TypeError, ValueError):
                    continue
            for cc in range(1, c_stat + 1):
                cell = ws.cell(r, cc)
                keep = bool(cell.alignment and cell.alignment.wrap_text)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=keep)
        wb.save(path)

    @staticmethod
    def _gst_mult(code) -> float:
        """GST multiplier (1+rate) from a GST group code (e.g. 'G-18-S' → 1.18)."""
        c = str(code or '').upper()
        if '28' in c:
            return 1.28
        if '18' in c:
            return 1.18
        if '12' in c:
            return 1.12
        if '5' in c and '15' not in c and '25' not in c:
            return 1.05
        if '3' in c:
            return 1.03
        if '0' in c:
            return 1.0
        return 1.18

    def sku_rows(self, lines=None) -> list[dict]:
        """Per-SKU demand rollup — the **single source** for BOTH the review
        'SKU' tab and the workbook 'SKU Summary' sheet, so one change renders on
        both sides. Grouped by (item_no, ean): total qty demanded, # of POs it
        appears on, inc-GST value (Σ unit CP × (1+GST) × qty), the unit price
        (CP written to D365), and **deal-SKU / overridden** flags (→ highlight)."""
        lines = self._lines() if lines is None else lines
        agg: dict = {}
        for ln in lines:
            key = (ln.get('item_no') or '', ln.get('ean') or '')
            a = agg.get(key)
            if a is None:
                a = agg[key] = {'item_no': ln.get('item_no') or '',
                                'ean': ln.get('ean') or '',
                                'description': ln.get('description') or '',
                                'qty': 0, 'pos': set(), 'value': 0.0,
                                'unit_prices': set(), 'deal': False,
                                'overridden': False, 'labels': set(),
                                # earlier MRP/diff rollup (kept for the workbook)
                                'our_mrp': ln.get('our_mrp'), 'vmrps': set(),
                                'ok': 0, 'mis': 0, 'nim': 0, 'diffs': []}
            q = int(ln.get('qty') or 0)
            a['qty'] += q
            a['pos'].add(ln.get('po'))
            up = ln.get('unit_price')
            if up is not None:
                a['unit_prices'].add(round(float(up), 2))
                a['value'] += float(up) * self._gst_mult(ln.get('gst_code')) * q
            elif ln.get('our_landing') is not None:      # landing is already inc-GST
                a['value'] += float(ln['our_landing']) * q
            # MRP comparison + status split + worst diff (the earlier columns)
            st = (ln.get('status') or 'OK').upper()
            a['ok' if st == 'OK' else 'mis' if st == 'MISMATCH'
              else 'nim' if st == 'NOT_IN_MASTER' else 'ok'] += q
            if ln.get('vendor_mrp') is not None:
                try:
                    a['vmrps'].add(round(float(ln['vendor_mrp']), 2))
                except (TypeError, ValueError):
                    pass
            if ln.get('diff') is not None:
                try:
                    a['diffs'].append(float(ln['diff']))
                except (TypeError, ValueError):
                    pass
            lbl = (ln.get('exception_label') or '').strip()
            act = ((ln.get('action') or '')
                   or (ln.get('decision') or {}).get('action') or '').upper()
            if lbl:
                a['labels'].add(lbl)
                a['overridden'] = True
            if act == 'OVERRIDE':
                a['overridden'] = True
            if 'deal' in lbl.lower():
                a['deal'] = True
        rows: list[dict] = []
        for a in agg.values():
            ups = sorted(a['unit_prices'])
            label = ', '.join(sorted(a['labels']))
            # Note carries WHY it's overridden (deal SKU / vendor CP / manual …);
            # the operator only needs Overridden Yes/No + this detail.
            note_full = label or ('Deal SKU' if a['deal'] else '')
            rows.append({
                'item_no': a['item_no'], 'ean': a['ean'],
                'description': a['description'], 'qty': a['qty'],
                'pos': len(a['pos']), 'value_incgst': round(a['value'], 2),
                'unit_price': (ups[-1] if ups else None),
                'unit_price_varies': len(ups) > 1,
                'deal': a['deal'], 'overridden': a['overridden'],
                'label': label, 'note_full': note_full,
                # earlier SKU-level difference columns (workbook only)
                'our_mrp': a['our_mrp'],
                'their_mrp': (max(a['vmrps']) if a['vmrps'] else None),
                'mrp_varies': 'YES' if len(a['vmrps']) > 1 else '',
                'ok_qty': a['ok'], 'mismatch_qty': a['mis'], 'nim_qty': a['nim'],
                'worst_diff': (min(a['diffs']) if a['diffs'] else None),
            })
        rows.sort(key=lambda r: (-r['value_incgst'], -r['qty']))
        return rows

    def _sku_pivot(self, lines):
        """Per-run SKU rollup grouped by (item_no, ean): qty per status, MRP
        comparison (+ varies flag), POs, worst diff. Returns sorted rows."""
        agg: dict = {}
        for ln in lines:
            key = (ln.get('item_no') or '', ln.get('ean') or '')
            a = agg.get(key)
            if a is None:
                a = agg[key] = {'desc': ln.get('description') or '',
                                'our_mrp': ln.get('our_mrp'), 'vmrps': set(),
                                'tot': 0, 'ok': 0, 'mis': 0, 'nim': 0,
                                'pos': set(), 'diffs': []}
            q = int(ln.get('qty') or 0)
            a['tot'] += q
            st = ln.get('status') or 'OK'
            a['ok' if st == 'OK' else 'mis' if st == 'MISMATCH'
              else 'nim' if st == 'NOT_IN_MASTER' else 'ok'] += q
            a['pos'].add(ln.get('po'))
            if ln.get('vendor_mrp') is not None:
                a['vmrps'].add(round(float(ln['vendor_mrp']), 2))
            if ln.get('diff') is not None:
                a['diffs'].append(float(ln['diff']))
        rows = []
        for (item_no, ean), a in agg.items():
            rows.append([
                item_no, ean, a['desc'], a['our_mrp'],
                (max(a['vmrps']) if a['vmrps'] else None),
                'YES' if len(a['vmrps']) > 1 else '', a['tot'], a['ok'],
                a['mis'], a['nim'], len(a['pos']),
                (min(a['diffs']) if a['diffs'] else None)])
        rows.sort(key=lambda r: (-r[8], -r[9], -r[6]))   # mismatch, nim, tot qty
        return rows

    def _append_sku_sheet(self, path):
        """Write the per-run **SKU Summary** sheet from the shared
        :meth:`sku_rows` (same data the review 'SKU' tab shows): qty demanded,
        # POs, inc-GST value, unit price (CP), Deal SKU + Overridden flags. Rows
        that are overridden (deal / vendor CP / manual) are **yellow-highlighted**
        so they stand out. Additive post-process; the engine's sheets are untouched."""
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        rows = self.sku_rows(self._lines())
        wb = load_workbook(path)
        if 'SKU Summary' in wb.sheetnames:
            del wb['SKU Summary']
        ws = wb.create_sheet('SKU Summary')
        hdr = ['Item No', 'EAN', 'Description', 'Qty Demanded', '# POs',
               'Value (inc GST)', 'Unit Price (CP)', 'Overridden', 'Note',
               # earlier SKU-level difference columns, kept at the end
               'Our MRP', 'Their MRP', 'MRP varies', 'OK Qty', 'Mismatch Qty',
               'Not-in-Master Qty', 'Worst Diff']
        ws.append(hdr)
        for r in rows:
            ws.append([
                r['item_no'], r['ean'], r['description'], r['qty'], r['pos'],
                r['value_incgst'],
                (r['unit_price'] if not r['unit_price_varies'] else
                 f"{r['unit_price']} (varies)"),
                'Yes' if r['overridden'] else 'No',
                r['note_full'],
                r['our_mrp'], r['their_mrp'], r['mrp_varies'],
                r['ok_qty'], r['mismatch_qty'], r['nim_qty'], r['worst_diff']])
        # ── formatting: header band, widths, alignment, borders, freeze ──
        navy = PatternFill('solid', fgColor='1A237E')
        yellow = PatternFill('solid', fgColor='FFF3C4')     # overridden highlight
        hfont = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='E6E8EC')
        bd = Border(thin, thin, thin, thin)
        for cell in ws[1]:
            cell.font = hfont
            cell.fill = navy
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            cell.border = bd
        widths = [11, 16, 44, 12, 7, 15, 13, 10, 24,
                  10, 10, 10, 8, 12, 15, 10]
        right_cols = {4, 5, 6, 7, 10, 11, 13, 14, 15, 16}
        center_cols = {8, 12}                       # Overridden / MRP varies
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 0):
            hot = bool(rows[ri]['overridden']) if ri < len(rows) else False
            for cell in row:
                cell.border = bd
                if hot:
                    cell.fill = yellow
                if cell.column in right_cols:
                    cell.alignment = Alignment(horizontal='right')
                elif cell.column in center_cols:
                    cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        wb.save(path)

    @staticmethod
    def _tracker_date_val(v):
        """Coerce a date-like value to a real ``datetime.date`` so the Tracker
        cell is a GENUINE Excel date — it then groups by month in the WH team's
        AutoFilter (not listed as flat text), sorts right, and survives a paste
        into the org master. ``None`` when the value isn't a date."""
        import datetime as _dt
        if not v:
            return None
        if isinstance(v, _dt.datetime):
            return v.date()
        if isinstance(v, _dt.date):
            return v
        s = str(v).strip()
        for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%b-%Y'):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @classmethod
    def _fmt_tracker_date(cls, v) -> str:
        """Day-first ``dd-mm-YYYY`` string (fallback for un-coercible values)."""
        d = cls._tracker_date_val(v)
        if d is not None:
            return d.strftime('%d-%m-%Y')
        return '' if not v else str(v).strip()

    def _append_tracker_sheet(self, path):
        """Append a per-PO **Tracker** sheet to the SO workbook (all marketplaces):
        Platform · PO/RO No · Location · PO Date · Expiry Date · Order Type ·
        Items · Total Qty · Total Amount (inc GST). Dates come from the engine or
        the PDF-date backfill (``_source_dates_by_po``). Additive post-process;
        the engine's own sheets are untouched. Best-effort — never fails export."""
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        headers = self._headers()
        if not headers:
            return
        try:
            dts = self._source_dates_by_po() or {}
        except Exception:  # noqa: BLE001
            dts = {}
        try:
            short = self._source_location_by_po() or {}   # {po: short warehouse code}
        except Exception:  # noqa: BLE001
            short = {}
        # Ship-to → State/Zone map (same resolution the web tracker uses), loaded
        # once. Resolved from the REAL ship-to location, not the WH-code override.
        try:
            from .order_db import geo_for_location, location_geo_map
            geomap = location_geo_map()
        except Exception:  # noqa: BLE001
            geomap = {}
            def geo_for_location(_loc, _m=None):  # noqa: E306
                return {'pincode': '', 'state': '', 'zone': ''}
        wb = load_workbook(path)
        if 'Tracker' in wb.sheetnames:
            del wb['Tracker']
        ws = wb.create_sheet('Tracker')
        # Columns MATCH the org master tracker ('New PO format.xlsx') EXACTLY and
        # in order for A2:I; State + Zone are appended (J, K) as extra columns —
        # same for EVERY marketplace (this runs in the base Processor). 'Order
        # Receive Date' + 'Picklist Qty' are left blank (filled by hand in the
        # master). No TOTAL row → the whole block pastes cleanly.
        cols = ['Segment', 'Market Place', 'PO', 'Location', 'PO Date', 'Exp Date',
                'PO Aging For Exp', 'Order Value', 'Order Qty', 'State', 'Zone']
        ws.append(cols)
        for h in headers:
            po = str(h.get('po') or '')
            d = dts.get(po) or {}
            pod_d = self._tracker_date_val(d.get('po_date') or h.get('po_date'))
            exd_d = self._tracker_date_val(d.get('exp_date') or h.get('exp_date'))
            q = int(h.get('qty') or 0)
            v = round(float(h.get('order_value') or 0), 2)
            geo = geo_for_location(h.get('location'), geomap)   # State/Zone from ship-to
            # Write REAL Excel dates when coercible (so the WH master's filter
            # groups them by month) — fall back to a plain string only if the
            # value can't be parsed as a date.
            ws.append([h.get('segment') or 'OnlineB2B',
                       h.get('marketplace_label') or h.get('marketplace') or '',
                       po, short.get(po) or h.get('location') or '',
                       pod_d if pod_d is not None else self._fmt_tracker_date(d.get('po_date') or h.get('po_date')),
                       exd_d if exd_d is not None else self._fmt_tracker_date(d.get('exp_date') or h.get('exp_date')),
                       '', v, q,          # PO Aging For Exp = filled manually
                       geo.get('state') or '', geo.get('zone') or ''])
            rr = ws.max_row
            if pod_d is not None:
                ws.cell(rr, 5).number_format = 'DD-MM-YYYY'
            if exd_d is not None:
                ws.cell(rr, 6).number_format = 'DD-MM-YYYY'
        # ── formatting ──
        navy = PatternFill('solid', fgColor='1A237E')
        hfont = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='E6E8EC')
        bd = Border(thin, thin, thin, thin)
        for cell in ws[1]:
            cell.font = hfont
            cell.fill = navy
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            cell.border = bd
        widths = [13, 16, 18, 42, 13, 13, 16, 15, 11, 16, 11]
        right_cols = {8, 9}          # Order Value, Order Qty
        center_cols = {5, 6, 7, 11}  # PO Date, Exp Date, PO Aging For Exp, Zone
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = bd
                if cell.column in right_cols:
                    cell.alignment = Alignment(horizontal='right')
                elif cell.column in center_cols:
                    cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        # Position the Tracker as the 4th sheet (after Headers / Lines / Summary).
        idx = wb.sheetnames.index('Tracker')
        if len(wb.sheetnames) > 3 and idx != 3:
            wb.move_sheet('Tracker', offset=3 - idx)
        wb.save(path)

    def _grouped_unmapped(self) -> list:
        """Unmapped ship-tos grouped by location → one row per missing mapping,
        with the affected PO list. Feeds the review page's "Mapping" tab."""
        by_loc: dict = {}
        for u in self.unmapped:
            loc = u.get('location') or '(blank)'
            g = by_loc.setdefault(loc, {'location': loc, 'marketplace': self.marketplace,
                                        'pos': [], 'count': 0})
            g['count'] += 1
            po = u.get('po')
            if po and po not in g['pos']:
                g['pos'].append(po)
        return sorted(by_loc.values(), key=lambda g: -g['count'])

    def _mapping_report(self) -> list:
        """Per-PO ship-to resolution for the review **Mapping** tab: every PO →
        the location it sent → the mapped Cust/Ship-to, classified **EXACT /
        FUZZY / UNMAPPED**. Lets the operator confirm each match and promote fuzzy
        ones to exact rows ([[fuzzy→exact goal]]). One row per PO."""
        if not self.result or not getattr(self.result, 'rows', None):
            return []
        seen: dict = {}
        for row in self.result.rows:
            po = getattr(row, 'po_number', '') or ''
            if not po or po in seen:
                continue
            raw = (getattr(row, 'location', '') or '').strip()
            mapped = (getattr(row, 'mapped_location', '') or '').strip()
            ship = getattr(row, 'ship_to', '') or ''
            cust = getattr(row, 'cust_no', '') or ''
            if not ship and not bool(getattr(row, 'mapped', False)):
                mt = 'UNMAPPED'
                remark = 'No Ship-To match — Cust/Ship-to left blank. Add an exact mapping.'
            elif mapped and raw and raw == mapped:
                mt, remark = 'EXACT', 'Exact match on Del Location.'
            else:
                mt = 'FUZZY'
                remark = (f'Fuzzy match — PO sent "{raw or "(blank)"}" → resolved to '
                          f'"{mapped or "?"}". Confirm it\'s right; promote to an exact row.')
            seen[po] = {'po': po, 'location': raw, 'mapped_location': mapped,
                        'cust_no': cust, 'ship_to': ship, 'match_type': mt, 'remark': remark}
        rank = {'UNMAPPED': 0, 'FUZZY': 1, 'EXACT': 2}
        return sorted(seen.values(), key=lambda x: (rank.get(x['match_type'], 3), x['po']))

    # ── phase 1: preview (no DB write) ──────────────────────────────────
    def preview(self) -> dict:
        err = self._run()
        if err:
            return err
        output_path = self._export()
        lines = self._lines()
        headers = self._headers()
        affected = [l for l in lines if l['status'] in _ISSUE_STATUSES]
        return {
            'ok': True, 'summary': self._summary(lines, headers),
            'headers': headers, 'lines': lines, 'affected': affected,
            'sku_rows': self.sku_rows(lines),
            'skipped': self.skipped, 'warnings': self.warnings,
            'unmapped': self._grouped_unmapped(),
            'mapping_report': self._mapping_report(),
            'notes': self.notes,
            'output_path': str(output_path) if output_path else None,
        }

    # ── phase 2: confirm (push headers + lines) ─────────────────────────
    def confirm(self, actions=None, as_of=None, recorded_by=None) -> dict:
        err = self._run()
        if err:
            return err
        output_path = self._export(actions)   # decisions → Summary Excluded/Final cols
        if output_path is None:
            return {'ok': False, 'error': "Workbook export failed.",
                    'warnings': self.warnings}
        lines = self._lines()
        out: dict = {
            'ok': True, 'output_path': str(output_path), 'run_id': None,
            'warnings': self.warnings,
            'summary': self._summary(lines, self._headers()),
        }
        from . import lines_store
        # ── CORE record: runs + order_headers + order_lines in ONE TRANSACTION ──
        # All-or-nothing. If ANYTHING fails (or the process is interrupted), the
        # whole thing rolls back — the DB is 100% written or completely untouched,
        # never a partial run. On failure we report cleanly (nothing recorded).
        try:
            # Build the line rows first — the run_id is stamped inside the atomic
            # write. _lines() applies EAN fixes (correct ean + received_ean audit).
            rows = self._lines(output_file=str(output_path), actions=actions,
                               as_of=as_of)
            rec = lines_store.record_run_atomic(
                self.result, self.marketplace, self.warehouse, str(output_path),
                rows, as_of=as_of, recorded_by=recorded_by)
            out['run_id'] = rec.get('run_id')
            out['new_orders'] = rec.get('new_orders', 0)
            out['lines_recorded'] = rec.get('lines_recorded', 0)
        except Exception as e:  # noqa: BLE001 — rolled back; NOTHING was recorded
            out['ok'] = False
            out['error'] = (f"Nothing was recorded — the run was rolled back "
                            f"({type(e).__name__}). No partial data was written; "
                            f"please retry.")
            return out
        # ── post-record ENRICHMENT (best-effort) ──────────────────────────────
        # These polish the ALREADY-COMMITTED run (backfill amount-less TO value,
        # PDF po/exp dates, friendly location). A failure here must NOT flip the
        # run to "not recorded" — the record is safe; just warn.
        try:
            if self._amountless_to():
                upd = lines_store.set_order_value(
                    out['run_id'], self._to_value_by_po(rows))
                out['value_backfilled'] = upd.get('updated', 0)
            dts = self._source_dates_by_po()
            if dts and out.get('run_id'):
                out['dates_backfilled'] = lines_store.set_po_dates(
                    out['run_id'], dts, force=self._dates_force).get('updated', 0)
            locs = self._source_location_by_po()
            if locs and out.get('run_id'):
                out['locations_relabelled'] = lines_store.set_location(
                    out['run_id'], locs).get('updated', 0)
        except Exception as e:  # noqa: BLE001 — enrichment failure ≠ un-recorded run
            self.warnings.append(
                f"Run recorded OK, but a post-record backfill was skipped "
                f"({type(e).__name__}).")
        return out

    #: When True, ``_source_dates_by_po`` OVERWRITES the engine's po_date/exp_date
    #: (not COALESCE) — for marketplaces whose engine date parse is WRONG, not
    #: merely blank (Swiggy day-first timestamp). Base = additive/blank-fill only.
    _dates_force: bool = False

    def _source_dates_by_po(self) -> dict:
        """``{po: {'po_date': date, 'exp_date': date}}`` from the source file(s),
        for marketplaces whose parser carries the dates in the header (not a row
        column). Base = none; PDF processors override. Used to backfill the
        tracker (po_date/exp_date) so TAT works for them."""
        return {}

    def _source_location_by_po(self) -> dict:
        """``{po: 'short location'}`` to re-stamp onto ``order_headers.location``
        after recording — for marketplaces that must feed the engine the RAW
        address but want a friendly short name on the tracker. Base = none;
        Myntra / Purplle override."""
        return {}

    def _mapped_city_by_po(self) -> dict:
        """``{po: mapped city}`` from ``ship_to_mapping`` (this marketplace's rows)
        — for channels that feed the engine the RAW full ship-to address (kept as
        the del_location key) but want the friendly CITY on the tracker (e.g.
        Purplle's single DC → 'Mumbai'). Matches each PO's raw location to a mapped
        del_location that carries a city. Read-only; ``{}`` on any error."""
        try:
            from .order_db import _conn
            with _conn() as (cur, d):
                ph = d['ph']
                cur.execute(f"SELECT del_location, city FROM ship_to_mapping "
                            f"WHERE party={ph} AND city<>''", (self.marketplace,))
                city_by_loc = {str(l).strip().lower(): str(c).strip()
                               for l, c in cur.fetchall()}
        except Exception:  # noqa: BLE001
            import logging
            logging.getLogger(__name__).exception(
                '_mapped_city_by_po: ship_to city lookup failed — no city overlay')
            return {}
        out = {}
        for h in self._headers():
            po = str(h.get('po') or '')
            c = city_by_loc.get(str(h.get('location') or '').strip().lower())
            if po and c:
                out[po] = c
        return out


# ── Flipkart ────────────────────────────────────────────────────────────

class FlipkartProcessor(Processor):
    """Flipkart: many ``purchase_order_*.xlsx`` (multi-file, PO from filename,
    address→ship-to) + an optional ``purchase-orders-*.csv`` header that drives
    the Tracker sheet. FK Grocery / hyperlocal resolution is master-driven."""

    def engine_files(self) -> list[str]:
        # The engine gets the PO xlsx; a uploaded .csv is the tracker header.
        return [p for p in self.po_paths if p.lower().endswith('.xlsx')]

    def use_multi(self, config) -> bool:
        return True

    # ── Tracker-driven sub-marketplace labels (FK Hyperlocal / FK Grocery) ──
    def _tracker_labels(self) -> dict:
        """``{PO: 'FK Hyperlocal'|'FK Grocery'|…}`` from the tracker rows. Empty
        when no header CSV was uploaded (then labels stay the engine default)."""
        tracker = getattr(self.result, 'flipkart_tracker_rows', None) or []
        return {str(t.get('PO', '')).strip(): t.get('Market Place', '')
                for t in tracker if t.get('Market Place')}

    def _headers(self):
        """Override the blanket 'Flipkart Alpha' label with the per-PO tracker
        classification so the dashboard/preview show FK Hyperlocal / FK Grocery."""
        rows = super()._headers()
        by_po = self._tracker_labels()
        for h in rows:
            label = by_po.get(str(h.get('po', '')).strip())
            if label:
                h['marketplace_label'] = label
        return rows

    def confirm(self, actions=None, as_of=None, recorded_by=None) -> dict:
        out = super().confirm(actions, as_of=as_of, recorded_by=recorded_by)
        # The engine's record_manual wrote 'Flipkart Alpha'; re-stamp the
        # web-owned display column per PO from the (latest) tracker mapping.
        if out.get('ok') and out.get('run_id'):
            by_po = self._tracker_labels()
            if by_po:
                try:
                    from .order_db import _conn_tx
                    payload = [(label, out['run_id'], po, 'Flipkart')
                               for po, label in by_po.items()]
                    # one atomic batch (executemany in _conn_tx) instead of a per-PO
                    # UPDATE loop on autocommit — all labels re-stamp together or none.
                    with _conn_tx() as (cur, d):
                        ph = d['ph']
                        cur.executemany(
                            f"UPDATE order_headers SET marketplace_label={ph} "
                            f"WHERE run_id={ph} AND po={ph} AND marketplace={ph}",
                            payload)
                except Exception as e:  # noqa: BLE001
                    out.setdefault('warnings', []).append(
                        f"Tracker re-label skipped ({type(e).__name__}: {e}).")
        return out

    def post_process(self, result, env) -> None:
        csv = next((p for p in self.po_paths if p.lower().endswith('.csv')), None)
        if not csv:
            self.warnings.append(
                "Flipkart Tracker: no 'purchase-orders-*.csv' header uploaded — "
                "Tracker sheet skipped (SO/values/locations are unaffected).")
            return
        try:
            from online_po_processor.engine.flipkart_tracker import (
                build_flipkart_tracker,
            )
            rows = build_flipkart_tracker(csv)
            # The header CSV lists EVERY open Flipkart PO on the portal. Keep ONLY
            # the POs actually uploaded this run (the purchase_order_*.xlsx we fed
            # the SO), so the Tracker reflects THIS upload — not the whole list.
            uploaded = set()
            for p in self.engine_files():
                stem = os.path.splitext(os.path.basename(p))[0]
                m = re.match(r'(?i)purchase[_-]?order[_-](.+)', stem)
                uploaded.add((m.group(1) if m else stem).strip().upper())
            before = len(rows)
            if uploaded:
                rows = [r for r in rows
                        if str(r.get('PO', '')).strip().upper() in uploaded]
            # Our-layer promotion: lift any Origin-Warehouse the operator has
            # promoted (e.g. a new FC that the frozen map still reads as
            # 'FK (review)') to its confirmed label. Frozen engine untouched.
            try:
                from .flipkart_wh_override import apply as _apply_wh_ov
                _changed = _apply_wh_ov(rows)
                if _changed:
                    self.warnings.append(
                        f"Flipkart: {_changed} PO(s) re-labelled from the "
                        f"promoted warehouse-override map.")
            except Exception:  # noqa: BLE001 — label refinement never blocks
                pass
            result.flipkart_tracker_rows = rows
            if uploaded and before != len(rows):
                self.warnings.append(
                    f"Flipkart Tracker: kept the {len(rows)} uploaded PO(s); "
                    f"the header CSV listed {before} — the rest were left out.")
        except Exception as e:  # noqa: BLE001
            self.warnings.append(f"Flipkart Tracker skipped ({type(e).__name__}: {e}).")

    def _source_location_by_po(self) -> dict:
        """``{PO: 'origin warehouse code'}`` — the SHORT internal warehouse code
        (e.g. ``ahm_sh_wh_nl_02nl``) from the header CSV's 'Origin Warehouse'
        column, keyed by 'Purchase Order ID'. Used to re-stamp the recorded
        headers + the Tracker Location with the short code instead of the
        resolved full ship-to address. No CSV / columns → no-op."""
        import pandas as pd
        csv = next((p for p in self.po_paths if p.lower().endswith('.csv')), None)
        if not csv:
            return {}
        try:
            df = pd.read_csv(csv, dtype=str)
        except Exception:  # noqa: BLE001
            return {}
        df.columns = [str(c).strip() for c in df.columns]
        if 'Purchase Order ID' not in df.columns or 'Origin Warehouse' not in df.columns:
            return {}
        out: dict = {}
        for _, r in df.iterrows():
            po = str(r.get('Purchase Order ID', '') or '').strip()
            ow = str(r.get('Origin Warehouse', '') or '').strip()
            if po and ow:
                out[po] = ow
        return out


# ── Flipkart Branch (Flipkart-TO) ───────────────────────────────────────

class FlipkartTOProcessor(Processor):
    """Flipkart Branch (Transfer Orders). The operator hands us Flipkart's raw
    per-PO exports ``Consignment_Details_<PO>_<date>.csv`` (one per PO) plus an
    optional ``Consignment_Visibility_Report*.csv`` for destination Locations.
    The engine's ``process_consignments`` parses the PO from each filename,
    joins the visibility report, and runs the standard TO pipeline.

    The dump carries no order amount, so the inc-GST transfer value is derived
    from our master pricing (Landing × qty) in ``_headers`` / ``confirm`` — see
    ``Processor._amountless_to``. Falls back to the single consolidated dump
    when one .xlsx is uploaded instead."""

    _CONSIGNMENT_RE = re.compile(r'Consignment_Details_', re.I)
    _VISIBILITY_RE = re.compile(r'Consignment_Visibility_Report', re.I)
    #: The visibility report's dates are day-first ('02-07-2026 18:13') but the
    #: frozen engine parses them month-first (2 Jul → Feb 7 for days 1–12), so we
    #: re-read them here and OVERWRITE the recorded po_date/exp_date.
    _dates_force = True

    def use_multi(self, config) -> bool:
        return True

    def _source_dates_by_po(self) -> dict:
        """``{PO: {po_date, exp_date}}`` read DAY-FIRST from the Consignment
        Visibility Report — 'Creation Date' → po_date, 'Scheduled Pick Up Date'
        → exp_date, keyed by 'Consignment Id' (== PO). Fixes the engine's
        month-first swap for days 1–12."""
        import pandas as pd
        vis = self._visibility_file()
        if not vis:
            return {}
        try:
            df = pd.read_csv(vis, dtype=str)
        except Exception:  # noqa: BLE001
            return {}
        df.columns = [str(c).strip() for c in df.columns]
        if 'Consignment Id' not in df.columns:
            return {}
        out: dict = {}
        for _, r in df.iterrows():
            po = str(r.get('Consignment Id', '') or '').strip()
            if not po:
                continue
            pod = _parse_dayfirst(r.get('Creation Date'))
            exd = _parse_dayfirst(r.get('Scheduled Pick Up Date'))
            if pod or exd:
                out[po] = {'po_date': pod, 'exp_date': exd}
        return out

    def _consignment_files(self) -> list:
        return [p for p in self.po_paths
                if self._CONSIGNMENT_RE.search(Path(p).name)]

    def _visibility_file(self):
        return next((p for p in self.po_paths
                     if self._VISIBILITY_RE.search(Path(p).name)), None)

    def run_engine(self, engine, files, config):
        cons = self._consignment_files()
        if cons:
            vis = self._visibility_file()
            if not vis:
                self.warnings.append(
                    "Flipkart Branch: no Consignment Visibility Report uploaded "
                    "— destination Locations may be blank (Transfer-to Code "
                    "unresolved). Add it to fill them.")
            return engine.process_consignments(
                cons, config, margin_pct=self.margin_pct,
                visibility_report_path=vis)
        # Fallback: a single pre-consolidated 7-column dump (.xlsx).
        return engine.process(files[0], config, margin_pct=self.margin_pct)


# ── Meesho Branch (Meesho-TO) ───────────────────────────────────────────

class MeeshoTOProcessor(Processor):
    """Meesho Branch (Transfer Orders). Bulk-consignment-ONLY: Meesho exports one
    CSV per order, ``order-line-items-<PO>[_<city>].csv`` (no consolidated dump,
    no visibility report). PO comes from the filename; Location from a city token
    in the filename (config ``filename_loc_from_shipto``: ``MS_BLR`` → 'blr', …),
    resolved to the Transfer-to Code via Ship-To B2B. The CSV carries no order
    amount (``sellingPricePerUnit`` is a selling price, deliberately ignored), so
    the inc-GST transfer value is derived from OUR master pricing (Landing × qty)
    in ``_headers`` / ``confirm`` — see ``Processor._amountless_to``."""

    _ITEMS_RE = re.compile(r'order-line-items', re.I)

    def use_multi(self, config) -> bool:
        return True

    def run_engine(self, engine, files, config):
        csvs = [p for p in self.po_paths if p.lower().endswith('.csv')]
        if not csvs:                      # tolerate any uploaded csv naming
            csvs = [p for p in files if str(p).lower().endswith('.csv')] or files
        # Location is filename-token driven for Meesho → no visibility report.
        result = engine.process_consignments(
            csvs, config, margin_pct=self.margin_pct,
            visibility_report_path=None)
        # The frozen engine appends a Flipkart-specific per-PO line
        # ("Amount reference — … Flipkart portal 'Amount' … = ₹0.00") for EVERY
        # TO, because Meesho Branch reuses the same process_consignments pipeline.
        # For Meesho the dump carries NO vendor price, so the portal total is
        # always ₹0.00 AND the "Flipkart" label is wrong — the line is mislabelled
        # and meaningless here. Drop it; the accurate "dump carries no price —
        # order value COMPUTED from our master pricing" warning (from _headers)
        # already states the real basis, so nothing is lost.
        try:
            result.warnings = [w for w in result.warnings
                               if "Flipkart portal 'Amount'" not in str(w[2])]
        except Exception:  # noqa: BLE001
            pass
        return result


# ── DMart (Avenue) — PDF dates for the tracker ──────────────────────────

def _parse_ddmmyyyy(s):
    """'17.06.2026' / '17-06-2026' / '17/06/2026' → date, else None."""
    import datetime as _d
    s = (s or '').strip()
    for fmt in ('%d.%m.%Y', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return _d.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


class DmartProcessor(Processor):
    """DMart (Avenue PO PDFs). Processing is the base flow — this subclass adds:

    1. **Correct ship-to (FC)** — the frozen ``avenue_pdf_parser`` hardcodes the
       ship-to to 'Bhiwandi' and silently defaults every other FC to it. We
       resolve the REAL FC from the PDF (``dmart_shipto``) and inject it into the
       engine's ``__loc__`` at runtime, so the right ``ship_to`` code flows to the
       D365 output AND the record. Unresolved/ambiguous/pincode-mismatch → the run
       is BLOCKED with a clear message (never routed to a default).
    2. **Tracker dates** — the parser reads ``Purchase Order Date`` / ``PO
       Validity`` from the PDF header, backfilled per PO after recording.
    """

    def _run(self, skip_dedup=False):
        # Confirm every DMart PO's FC before the engine runs. Split the failures
        # so a ship-to *nitpick* never reads like a parse failure:
        #   • UNROUTABLE (no ship_to resolved at all — PDF unreadable, FC not in
        #     the mapping / likely a non-DMart or wrong-format file, or an
        #     ambiguous multi-FC match): HARD-block the import, because the frozen
        #     parser would otherwise silently default the ship-to to 'Bhiwandi'.
        #     [[never-skip-silently]] [[dmart-shipto-fix]]
        #   • ROUTABLE-BUT-UNVERIFIED (a single FC WAS identified but a soft
        #     cross-check failed, e.g. the PDF pincode differs from the mapping):
        #     the ship_to code is known and correct, so DON'T kill the run —
        #     inject the FC and surface the issue as a review WARNING. The operator
        #     opens the page, verifies / fixes the mapping pincode, then locks.
        from . import dmart_shipto
        res = dmart_shipto.resolve_paths(self.po_paths)
        unroutable = {po: r for po, r in res.items()
                      if not r.get('ok') and not r.get('ship_to')}
        if unroutable:
            details = "\n• ".join(f"PO {po}: {r['reason']}"
                                  for po, r in unroutable.items())
            return {'ok': False, 'error':
                    "DMart ship-to could not be identified — nothing was recorded "
                    "(no PO was routed to a default warehouse). This usually means "
                    "the FC isn't in the Ship-To Mapping yet, or a non-DMart / "
                    "wrong-format file was picked:\n• " + details}
        # Soft issues (FC known, only a cross-check failed) → proceed + warn on
        # the review page so the operator can confirm rather than being blocked.
        for po, r in res.items():
            if not r.get('ok') and r.get('ship_to'):
                self.warnings.append(
                    f"DMart PO {po}: {r['reason']} Routed to '{r['fc']}' "
                    f"({r['ship_to']}) — verify the ship-to; if the mapping "
                    f"pincode is wrong, fix it on the Ship-To Mapping page.")
        loc_by_po = {po: r['fc'] for po, r in res.items() if r.get('fc')}
        if not loc_by_po:
            return super()._run(skip_dedup=skip_dedup)   # no PDFs → base flow

        # Runtime injection: override the frozen parser's __loc__ with the
        # confirmed FC, keyed by PO. Restored in finally (never left patched).
        import online_po_processor.engine.avenue_pdf_parser as _ap
        _orig = _ap.avenue_po_to_dataframe

        def _patched(po):
            df = _orig(po)
            fc = loc_by_po.get(str(po.header.po_number))
            if fc:
                df['__loc__'] = fc
            return df

        _ap.avenue_po_to_dataframe = _patched
        try:
            return super()._run(skip_dedup=skip_dedup)
        finally:
            _ap.avenue_po_to_dataframe = _orig

    def _source_dates_by_po(self) -> dict:
        try:
            from online_po_processor.engine.avenue_pdf_parser import (
                parse_avenue_pdf,
            )
        except Exception:  # noqa: BLE001
            return {}
        out: dict = {}
        for p in self.po_paths:
            if not str(p).lower().endswith('.pdf'):
                continue
            try:
                po = parse_avenue_pdf(p)
                h = po.header
                if h.po_number:
                    out[str(h.po_number)] = {
                        'po_date': _parse_ddmmyyyy(h.po_date),
                        'exp_date': _parse_ddmmyyyy(h.validity_to),
                    }
            except Exception:  # noqa: BLE001 — never block on date extraction
                continue
        return out


class FirstcryProcessor(Processor):
    """FirstCry (bordered PO PDFs). Processing is the base flow — this subclass
    only adds the tracker dates: the FirstCry parser reads ``PO Date`` and
    ``PO Expiry Date`` from each PDF header (not a row column), so we backfill
    po_date/exp_date per PO after recording (so FirstCry shows on the TAT page),
    mirroring :class:`DmartProcessor`."""

    def _source_dates_by_po(self) -> dict:
        try:
            from online_po_processor.engine.firstcry_pdf_parser import (
                parse_firstcry_pdf,
            )
        except Exception:  # noqa: BLE001
            return {}
        out: dict = {}
        for p in self.po_paths:
            if not str(p).lower().endswith('.pdf'):
                continue
            try:
                po = parse_firstcry_pdf(p)
                h = po.header
                if h.po_number:
                    out[str(h.po_number)] = {
                        'po_date': _parse_ddmmyyyy(h.po_date),
                        'exp_date': _parse_ddmmyyyy(h.po_expiry),
                    }
            except Exception:  # noqa: BLE001 — never block on date extraction
                continue
        return out


class MyntraProcessor(Processor):
    """Myntra now sends ONE ``PO_<id>_PO-MYNJ-*.xlsx`` PER PO (a title + a
    header block, with the line-item table a few rows down) instead of a single
    compiled dump. The operator used to copy-paste them into one sheet by hand;
    this compiles them automatically into the flat dump the engine's Myntra
    reader expects — ``PO`` (from 'PO Barcode') + ``Location`` (from 'Ship To')
    + the line columns — then runs the engine on that. A file that is ALREADY a
    compiled dump (has a ``PO`` column) passes through unchanged. Frozen engine
    untouched."""

    _COMPILED_NAME = '_myntra_compiled.xlsx'
    #: {po: short location} built during _compile, re-stamped after recording.
    _short_loc: dict = {}
    #: {(po, ean): Landing Price} captured during _compile, stamped onto
    #: so.ref_fob_price in post_process for DISPLAY only (never validation).
    _landing_by_key: dict = {}

    @staticmethod
    def _norm_ean(v) -> str:
        s = str(v if v is not None else '').strip()
        return s[:-2] if s.endswith('.0') else s

    def post_process(self, result, env) -> None:
        """After the engine runs: (1) stamp the with-GST 'Landing Price' onto
        ``so.ref_fob_price`` for the tracker's Vendor Landing display (never the
        CP validation — see the ref_fob_col note); (2) build the tracker's SHORT
        location per PO from the engine's RESOLVED ship-to code (so.ship_to →
        D365 short name), guaranteeing the tracker label matches the ship-to code
        actually sent to D365 (e.g. 20011_4 → 'Gurgaon', 20011_1 → 'Binola')."""
        import pandas as pd
        if '_short_loc' not in self.__dict__:
            self._short_loc = {}   # instance dict (compile may not have run)
        # (2) code → short name (D365 'name' col), for the tracker relabel.
        code2name: dict = {}
        try:
            from .order_db import _conn
            with _conn() as (cur, d):
                cur.execute("SELECT ship_to, name FROM ship_to_mapping "
                            "WHERE party='Myntra' AND name IS NOT NULL")
                code2name = {str(a): str(b) for a, b in cur.fetchall() if b}
        except Exception:  # noqa: BLE001
            code2name = {}
        # (3) Myntra negotiated deal SKUs: expected CP = agreed transfer price
        # ÷(1+GST). The frozen engine's deal path is Swiggy-ONLY (gated at
        # marketplace_engine.py, so it never leaks), so we apply Myntra's here.
        try:
            from . import overrides_store
            deal_map = overrides_store.myntra_deal_map()
        except Exception as e:  # noqa: BLE001
            deal_map = {}
            self.warnings.append(
                f"Myntra deal-SKU prices couldn't load ({type(e).__name__}) — deal "
                f"lines will fall back to flat margin, NOT the negotiated price. Verify prices.")
        applied = mismatched = 0
        for so in result.rows:
            po = str(so.po_number)
            # (1) Vendor Landing display value.
            if self._landing_by_key:
                v = self._landing_by_key.get((po, self._norm_ean(so.ean)))
                if v is not None and pd.notna(v):
                    so.ref_fob_price = float(v)
            # (2) short tracker location from resolved ship-to code.
            code = str(getattr(so, 'ship_to', '') or '')
            if code and code in code2name:
                self._short_loc[po] = code2name[code]
            # (3) negotiated deal price override.
            transfer = deal_map.get(self._norm_ean(so.ean))
            if transfer is None:
                continue
            # Use Vendor CP WINS: if the engine already accepted the vendor's
            # stated CP for this SKU (e.g. Goddess → exception_label 'Vendor CP
            # (deal)'), do NOT clobber it with a deal price. The two rules are
            # mutually exclusive per SKU; vendor-CP takes precedence.
            if str(getattr(so, 'exception_label', '') or '').startswith('Vendor CP'):
                continue
            vendor_cp = getattr(so, 'fob_price', None)   # Myntra basis='cost'
            # Per-item GST factor from the vendor's own landing/CP ratio (both
            # scale with the GST rate), clamped; default 18%.
            gst_div = 1.18
            try:
                ratio = float(so.ref_fob_price) / float(vendor_cp)
                if 1.01 <= ratio <= 1.40:
                    gst_div = ratio
            except (TypeError, ValueError, ZeroDivisionError):
                pass
            expected_cp = round(float(transfer) / gst_div, 2)
            so.cost_price_ref = expected_cp
            note_status = 'OK'
            try:
                diff = round(float(vendor_cp) - expected_cp, 2)
            except (TypeError, ValueError):
                diff = None
            if diff is not None and abs(diff) <= 1.0:
                so.validation_status = 'OK'
                # Flag it so an OK-by-deal is distinguishable from a natural OK
                # (drives the ⚑ EXCEPTION column + row highlight), and force the
                # agreed cost into the D365 Unit Price so ERP uses the negotiated
                # price, NOT the flat Myntra margin — same as 'Vendor CP (deal)'.
                so.exception_label = 'Myntra deal'
                try:
                    so.forced_unit_price = float(vendor_cp)
                except (TypeError, ValueError):
                    pass
                applied += 1
            else:
                so.validation_status = 'MISMATCH'   # vendor differs from the deal
                so.exception_label = 'Deal ≠ vendor'
                note_status = 'MISMATCH vs deal'
                mismatched += 1
            try:
                result.exceptions_applied.append({
                    'type': 'price_override', 'po': po, 'ean': str(so.ean),
                    'item_no': str(so.item_no),
                    'detail': f"Myntra deal SKU — expected CP {expected_cp} "
                              f"(transfer {transfer} inc GST) [{note_status}]"})
            except Exception:  # noqa: BLE001
                pass
        if applied or mismatched:
            self.warnings.append(
                f"Myntra deal SKUs: {applied} line(s) matched the negotiated "
                f"transfer price (expected CP set, marked OK); {mismatched} still "
                f"differ from the agreed price (kept MISMATCH). Never silent.")

    def engine_files(self) -> list[str]:
        xlsx = [p for p in self.po_paths if p.lower().endswith('.xlsx')]
        if not xlsx:
            return self.po_paths
        passthrough = [p for p in xlsx if self._is_compiled_dump(p)]
        per_po = [p for p in xlsx if p not in passthrough]
        if not per_po:
            return xlsx
        compiled = self._compile(per_po)
        return ([compiled] + passthrough) if compiled else (passthrough or xlsx)

    @staticmethod
    def _is_compiled_dump(path) -> bool:
        import pandas as pd
        try:
            cols = [str(c).strip().lower() for c in pd.read_excel(path, nrows=0).columns]
        except Exception:  # noqa: BLE001
            return False
        return 'po' in cols or 'po number' in cols

    def _compile(self, files):
        import pandas as pd
        # The engine reads the RAW Ship-To address for its own ship-to
        # resolution, so keep it as-is here. Separately we remember the resolved
        # SHORT mapped name per PO (``_short_loc``) to backfill onto the tracker
        # after recording, so order_headers.location shows 'Mumbai'/'West bengal'
        # rather than the full address.
        self._short_loc = {}
        self._landing_by_key = {}
        lcol = _WEB_VENDOR_LANDING_COL.get(self.marketplace)  # 'Landing Price'
        frames = []
        for f in files:
            try:
                raw = pd.read_excel(f, sheet_name=0, header=None)
            except Exception as e:  # noqa: BLE001
                self.warnings.append(
                    f"[{os.path.basename(f)}] Myntra compile skipped: {e}")
                continue
            po = self._label_value(raw, 'PO Barcode') or self._po_from_name(f)
            location = (self._label_value(raw, 'Ship To', multiline=True)
                        or self._label_value(raw, 'Bill To', multiline=True) or '')
            hrow = self._header_row(raw)
            if hrow is None:
                self.warnings.append(
                    f"[{os.path.basename(f)}] Myntra: line-item header "
                    f"(SKU Code / GTIN) not found — file skipped.")
                continue
            tbl = pd.read_excel(f, sheet_name=0, header=hrow)
            tbl.columns = [str(c).strip() for c in tbl.columns]
            key = 'GTIN' if 'GTIN' in tbl.columns else (
                'SKU Code' if 'SKU Code' in tbl.columns else None)
            if key:
                tbl = tbl[tbl[key].notna()]
            tbl = tbl.dropna(how='all')
            if tbl.empty:
                # NEVER silent: a PO whose line table filters to empty (blank GTIN/
                # SKU column, or a header variant) would otherwise vanish unrecorded.
                self.warnings.append(
                    f"[{os.path.basename(f)}] Myntra{f' PO {po}' if po else ''}: no "
                    f"line rows after filtering (GTIN/SKU column blank or renamed) — "
                    f"PO skipped, nothing recorded from it. Check the file.")
                continue
            # Remember the with-GST Landing Price per (po, ean) for the tracker's
            # Vendor Landing display column (stamped in post_process; the engine
            # itself keeps comparing on CP = List price).
            if lcol and lcol in tbl.columns and 'GTIN' in tbl.columns:
                for _, rr in tbl.iterrows():
                    g = self._norm_ean(rr.get('GTIN'))
                    lv = rr.get(lcol)
                    if g and pd.notna(lv):
                        try:
                            self._landing_by_key[(str(po), g)] = float(lv)
                        except (TypeError, ValueError):
                            pass
            tbl.insert(0, 'Location', location)
            tbl.insert(0, 'PO', po)
            frames.append(tbl)
        if not frames:
            self.warnings.append(
                "Myntra: no per-PO files could be compiled (nothing extracted).")
            return None
        big = pd.concat(frames, ignore_index=True)
        out = os.path.join(os.path.dirname(files[0]), self._COMPILED_NAME)
        big.to_excel(out, index=False, sheet_name='Sheet1')
        self.notes.append(
            f"Myntra: compiled {len(frames)} per-PO file(s) into one dump "
            f"({len(big)} line(s)).")
        return out

    @staticmethod
    def _label_value(raw, label, multiline=False):
        """Value to the RIGHT of a header-block label cell (e.g. 'PO Barcode',
        'Ship To'). With ``multiline``, append continuation rows (merged address
        cells) until a new label or a blank appears."""
        import pandas as pd
        lab = label.strip().lower()
        nrows, ncols = raw.shape
        for r in range(min(nrows, 12)):
            for c in range(ncols):
                v = raw.iat[r, c]
                if isinstance(v, str) and v.strip().lower() == lab:
                    vc = next((cc for cc in range(c + 1, ncols)
                               if pd.notna(raw.iat[r, cc])
                               and str(raw.iat[r, cc]).strip()), None)
                    if vc is None:
                        return ''
                    parts = [str(raw.iat[r, vc]).strip()]
                    if multiline:
                        rr = r + 1
                        while rr < nrows and (pd.isna(raw.iat[rr, c])
                                              or not str(raw.iat[rr, c]).strip()):
                            cell = raw.iat[rr, vc]
                            if pd.isna(cell) or not str(cell).strip():
                                break
                            parts.append(str(cell).strip())
                            rr += 1
                    return ' '.join(parts).strip()
        return ''

    @staticmethod
    def _header_row(raw):
        import pandas as pd
        nrows, ncols = raw.shape
        for r in range(min(nrows, 20)):
            vals = {str(raw.iat[r, c]).strip().lower()
                    for c in range(ncols) if pd.notna(raw.iat[r, c])}
            if 'sku code' in vals or 'gtin' in vals or 'sku id' in vals:
                return r
        return None

    @staticmethod
    def _po_from_name(path):
        name = os.path.splitext(os.path.basename(path))[0]
        m = re.search(r'(MYNJ[-\w]*)', name, re.IGNORECASE)
        return m.group(1) if m else name

    def _source_dates_by_po(self) -> dict:
        """Myntra's per-PO files carry ``PO Approved Date`` + ``Estimated
        Delivery Date`` in the header block (not a line column), so backfill
        po_date/exp_date per PO after recording — so Myntra shows on the TAT
        page. Mirrors DMart / FirstCry."""
        import pandas as pd
        out: dict = {}
        for p in self.po_paths:
            if not str(p).lower().endswith('.xlsx'):
                continue
            try:
                raw = pd.read_excel(p, sheet_name=0, header=None)
            except Exception:  # noqa: BLE001
                continue
            po = self._label_value(raw, 'PO Barcode') or self._po_from_name(p)
            pod = _parse_ddmmyyyy(self._label_value(raw, 'PO Approved Date'))
            exd = _parse_ddmmyyyy(self._label_value(raw, 'Estimated Delivery Date'))
            if po and (pod or exd):
                out[str(po)] = {'po_date': pod, 'exp_date': exd}
        return out

    def _source_location_by_po(self) -> dict:
        """Short mapped location per PO (built in :meth:`_compile`), so the
        tracker shows 'Mumbai'/'West bengal' while the engine still reads the raw
        ship-to address for its own resolution."""
        return dict(self._short_loc)


# ── Swiggy ──────────────────────────────────────────────────────────────

class SwiggyProcessor(Processor):
    """Swiggy: flat ``PO_<id>.csv``. The engine reads ``PoCreatedAt`` (po_date)
    which is a **day-first timestamp WITH a time** ('01-07-2026 13:38'). The
    engine's tracker date formatter tries date-only day-first patterns, they fail
    on the time, and it falls back to ``pd.to_datetime`` (month-first) — so a PO
    created on the 1st–12th gets its day/month SWAPPED (1-Jul → 7-Jan), which
    then reads as a huge false TAT breach. Engine is frozen, so we re-read the
    source dates day-first here and OVERWRITE (``_dates_force``).

    Status filter: Swiggy must punch ONLY ``CONFIRMED`` POs. The frozen engine
    (``_flag_po_status``) flags-and-KEEPS the rest for manual review; the operator
    instead wants non-CONFIRMED POs IGNORED (dropped) with a notification. So we
    drop them here (``run_engine`` → ``_drop_non_confirmed``) and name every
    dropped PO on Warnings — never silent (golden rule)."""

    _dates_force = True

    def _status_by_po(self) -> dict:
        """``{PoNumber: set(UPPER statuses)}`` from the source dump(s)
        (CSV or XLSX). Empty if the dump has no ``Status`` column."""
        import pandas as pd
        out: dict = {}
        for p in self.po_paths:
            low = str(p).lower()
            try:
                if low.endswith('.csv'):
                    df = pd.read_csv(p, dtype=str)
                elif low.endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(p, dtype=str)
                else:
                    continue
            except Exception:  # noqa: BLE001
                continue
            df.columns = [str(c).strip() for c in df.columns]
            if 'PoNumber' not in df.columns or 'Status' not in df.columns:
                continue
            for _, row in df.iterrows():
                po = str(row.get('PoNumber', '')).strip()
                st = str(row.get('Status', '')).strip().upper()
                if po and st and st != 'NAN':
                    out.setdefault(po, set()).add(st)
        return out

    def run_engine(self, engine, files, config):
        self._remap_nfs_to_forsale(engine)
        result = super().run_engine(engine, files, config)
        self._drop_non_confirmed(result, config)
        self._note_nfs_remaps(result)
        return result

    def _nfs_forsale_map(self) -> dict:
        """``{base_ean: base_ean+'_FS'}`` for master items whose description
        contains 'NFS' (Not-For-Sale) AND which have a ``<ean>_FS`` (For-Sale)
        twin. Swiggy sells the FOR-SALE variant and the negotiated deal price is
        registered on the ``_FS`` EAN — so a Swiggy line that resolves to the NFS
        item must be redirected to its For-Sale twin (then the deal applies)."""
        out: dict = {}
        try:
            from .order_db import _conn
            with _conn() as (cur, d):
                cur.execute("SELECT ean FROM item_master "
                            "WHERE UPPER(description) LIKE '%NFS%'")
                nfs = {str(r[0]) for r in cur.fetchall()}
                cur.execute("SELECT ean FROM item_master WHERE ean LIKE '%\\_FS'")
                fs = {str(r[0]) for r in cur.fetchall()}
        except Exception:  # noqa: BLE001
            return {}
        for base in nfs:
            if base + '_FS' in fs:
                out[base] = base + '_FS'
        return out

    def _remap_nfs_to_forsale(self, engine) -> None:
        """Redirect Swiggy SkuCode→EAN entries that land on an NFS item to the
        item's ``_FS`` For-Sale twin, BEFORE the engine resolves items (so item
        resolution + deal-SKU override both key off the For-Sale EAN). Mutates
        only this run's in-memory master; the frozen engine is untouched."""
        m = getattr(engine, 'master', None)
        self._nfs_remaps = self._nfs_forsale_map() if m and getattr(
            m, 'swiggy_sku', None) else {}
        if not self._nfs_remaps:
            return
        for sku, ean in list(m.swiggy_sku.items()):
            if str(ean) in self._nfs_remaps:
                m.swiggy_sku[sku] = self._nfs_remaps[str(ean)]

    def _note_nfs_remaps(self, result) -> None:
        """NEVER SILENT: one note per PO whose NFS line was remapped to its
        For-Sale twin (so the Swiggy deal applied)."""
        fs_set = set(getattr(self, '_nfs_remaps', {}).values())
        if not fs_set:
            return
        seen = set()
        for so in result.rows:
            e = str(getattr(so, 'ean', '') or '')
            key = (str(so.po_number), e)
            if e in fs_set and key not in seen:
                seen.add(key)
                self.notes.append(
                    f"Swiggy NFS→For-Sale: PO {so.po_number} — EAN {e[:-3]} "
                    f"(Not-For-Sale) remapped to For-Sale variant {e} "
                    f"(item {so.item_no}); Swiggy deal price applied.")

    def _drop_non_confirmed(self, result, config) -> None:
        """KEEP only ``status_keep`` (CONFIRMED) POs; DROP every other state
        (EXPIRED / COMPLETED / CANCELLED / PENDING) with ONE named warning per
        dropped PO. Overrides the frozen engine's flag-and-keep: we suppress its
        now-inaccurate "KEPT in output / pasted as-is" flags and replace them with
        "IGNORED" notices. A PO is kept only when ALL its lines are CONFIRMED."""
        keep = {str(s).strip().upper()
                for s in (config.get('status_keep') or ['CONFIRMED'])}
        st_by_po = self._status_by_po()
        if not st_by_po:
            return
        drop_pos = {po for po, sts in st_by_po.items() if not (sts <= keep)}
        if not drop_pos:
            return
        # Suppress the engine's flag-and-keep status-review warnings — we DROP.
        result.warnings = [w for w in result.warnings
                           if not ('pasted as-is' in str(w[2])
                                   or 'KEPT in output' in str(w[2]))]
        before = len(result.rows)
        result.rows = [so for so in result.rows
                       if str(so.po_number).strip() not in drop_pos]
        dropped = before - len(result.rows)
        for po in sorted(drop_pos):
            sts = ', '.join(sorted(st_by_po[po] - keep)) or ', '.join(sorted(st_by_po[po]))
            result.warnings.append((
                po, '',
                f"PO STATUS {sts} — IGNORED (not {sorted(keep)}); dropped from "
                f"this run. Re-upload if it should be punched."))
        result.warnings.append((
            '', '',
            f"Swiggy status filter: dropped {dropped} line(s) across "
            f"{len(drop_pos)} non-CONFIRMED PO(s); only {sorted(keep)} punched."))
        import logging
        logging.info("Swiggy status filter: dropped %d line(s) across %d PO(s) %s",
                     dropped, len(drop_pos), sorted(drop_pos))

    def _source_dates_by_po(self) -> dict:
        import pandas as pd
        out: dict = {}
        for p in self.po_paths:
            if not str(p).lower().endswith('.csv'):
                continue
            try:
                df = pd.read_csv(p, dtype=str)
            except Exception:  # noqa: BLE001
                continue
            df.columns = [str(c).strip() for c in df.columns]
            if 'PoNumber' not in df.columns:
                continue
            for _, row in df.iterrows():
                po = str(row.get('PoNumber', '')).strip()
                if not po:
                    continue
                pod = _parse_dayfirst(row.get('PoCreatedAt'))
                exd = _parse_dayfirst(row.get('PoExpiryDate'))
                if po and (pod or exd):
                    out[po] = {'po_date': pod, 'exp_date': exd}
        return out


def _parse_dayfirst(v):
    """Parse a day-first date/timestamp ('01-07-2026 13:38', '12-07-2026') to a
    ``date``, day BEFORE month always. Returns None on blank/unparseable."""
    import datetime as _dt

    import pandas as pd
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'nat', 'none'):
        return None
    # A leading 4-digit year means it's already year-first (ISO) — don't apply
    # dayfirst (which would swap YYYY-DD-MM). Everything else is day-first.
    iso = bool(re.match(r'^\d{4}-\d{1,2}-\d{1,2}', s))
    try:
        ts = pd.to_datetime(s, dayfirst=not iso, errors='raise')
        if pd.notna(ts):
            return ts.date()
    except Exception:  # noqa: BLE001
        pass
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%d.%m.%Y', '%Y-%m-%d'):
        try:
            return _dt.datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


# ── Factory + module entry points (views call these) ────────────────────

class BlinkMPProcessor(Processor):
    """BlinkMP (BCPL Reorder channel). Each RO arrives as a **PAIR**: a per-RO
    ORDER **Excel** (the line items) + a per-RO ORDER **PDF** (RO date, expiry
    date, delivery location) — paired by the RO number in the filename. The tabular
    dump the engine reads has no dates and no location column, so we build one:
    read each RO's Excel lines, prepend ``ro_number`` + ``location`` (from the PDF),
    concat into the flat dump the frozen BlinkMP config reads, and backfill
    po_date/exp_date from the PDF. Frozen engine untouched.

    Aligns with the other online marketplaces (mirrors :class:`MyntraProcessor` —
    ``engine_files``→compile, ``_source_dates_by_po``→date backfill)."""

    _COMPILED_NAME = '_blinkmp_compiled.xlsx'
    #: {ro_number: {'po_date': date, 'exp_date': date}} built during compile.
    _dates: dict = {}

    def engine_files(self) -> list[str]:
        pairs = self._pair_files()
        if not pairs:
            return self.po_paths     # let the engine raise a clear column error
        compiled = self._compile(pairs)
        return [compiled] if compiled else self.po_paths

    def _expanded_paths(self) -> list:
        """The uploaded paths, with any ``.zip`` extracted — BlinkMP's raw
        download is two zips (``…ORDER_PDF.zip`` + ``…ORDER_XLS.zip``), so accept
        them directly and pull the per-RO ``.pdf`` / ``.xls`` out."""
        import tempfile
        import zipfile
        out = []
        for p in self.po_paths:
            if str(p).lower().endswith('.zip'):
                try:
                    dst = tempfile.mkdtemp(suffix='_bmp_zip')
                    with zipfile.ZipFile(p) as z:
                        z.extractall(dst)
                    for root, _dirs, files in os.walk(dst):
                        out += [os.path.join(root, f) for f in files
                                if f.lower().endswith(('.pdf', '.xls', '.xlsx'))]
                except Exception as e:  # noqa: BLE001
                    self.warnings.append(
                        f"[{os.path.basename(str(p))}] zip extract failed: {e}")
            else:
                out.append(p)
        return out

    def _pair_files(self) -> list[dict]:
        """Pair the uploaded files by the RO number (leading digits) in the
        filename — ``<ro>_ORDER_XLS.xls`` ↔ ``<ro>_ORDER_PDF.pdf`` (zips expanded)."""
        excel: dict = {}
        pdf: dict = {}
        for p in self._expanded_paths():
            base = os.path.basename(str(p))
            m = re.search(r'\d{5,}', base)
            if not m:
                continue
            ro = m.group()
            low = str(p).lower()
            if low.endswith(('.xlsx', '.xls')):
                excel.setdefault(ro, p)
            elif low.endswith('.pdf'):
                pdf.setdefault(ro, p)
        return [{'ro': ro, 'excel': excel[ro], 'pdf': pdf.get(ro)} for ro in excel]

    def _compile(self, pairs):
        import pandas as pd
        self._dates = {}
        frames = []
        for pr in pairs:
            ro = str(pr['ro'])
            meta = self._pdf_meta(pr['pdf']) if pr.get('pdf') else {}
            if not pr.get('pdf'):
                self.warnings.append(
                    f"[RO {ro}] no matching ORDER PDF — location + PO/expiry dates "
                    f"missing for this RO (upload the PDF too).")
            try:
                df = pd.read_excel(pr['excel'])
            except Exception as e:  # noqa: BLE001
                self.warnings.append(
                    f"[RO {ro}] {os.path.basename(str(pr['excel']))}: Excel read "
                    f"failed: {e}")
                continue
            if 'Item Code' in df.columns:      # drop Total/Net + blank-item rows
                df = df[df['Item Code'].notna()]
                df = df[~df['Item Code'].astype(str).str.contains(
                    'Total|Net', case=False, na=False)]
            df = df.dropna(how='all')
            if df.empty:
                # NEVER silent: an RO that filters to empty (all Total/Net rows or a
                # blank Item Code column) would otherwise vanish unrecorded.
                self.warnings.append(
                    f"[RO {ro}] {os.path.basename(str(pr['excel']))}: no item rows "
                    f"after filtering — RO skipped, nothing recorded from it. Check the file.")
                continue
            if 'Product UPC' in df.columns:    # float/sci-notation → clean EAN
                df['Product UPC'] = df['Product UPC'].map(self._clean_ean)
            df.insert(0, 'location', meta.get('location', ''))
            df.insert(0, 'ro_number', ro)
            frames.append(df)
            if meta.get('po_date') or meta.get('exp_date'):
                self._dates[ro] = {'po_date': meta.get('po_date'),
                                   'exp_date': meta.get('exp_date')}
        if not frames:
            self.warnings.append(
                "BlinkMP: no RO pairs could be compiled (nothing extracted).")
            return None
        big = pd.concat(frames, ignore_index=True)
        # Write the compiled dump into the UPLOAD token dir (where po_paths live,
        # under MEDIA) — NOT the zip's temp-extract dir — so the engine's workbook
        # export lands in <token>/output where review_download reads it.
        base_dir = (os.path.dirname(str(self.po_paths[0])) if self.po_paths
                    else os.path.dirname(str(pairs[0]['excel'])))
        out = os.path.join(base_dir, self._COMPILED_NAME)
        big.to_excel(out, index=False, sheet_name='Sheet1')
        self.notes.append(
            f"BlinkMP: compiled {len(frames)} RO pair(s) into one dump "
            f"({len(big)} line(s)).")
        return out

    @staticmethod
    def _clean_ean(v) -> str:
        import pandas as pd
        try:
            if pd.isna(v):
                return ''
        except (TypeError, ValueError):
            pass
        try:                                    # 8.904473e+12 / 8904473012345.0 → int
            f = float(v)
            if f == int(f):
                return str(int(f))
        except (TypeError, ValueError):
            pass
        s = str(v).strip()
        return s[:-2] if s.endswith('.0') else s

    @staticmethod
    def _pdf_meta(pdf_path) -> dict:
        """Delivery location + RO date + expiry date from the ORDER PDF. Dates are
        parsed with the standard ``dateutil`` parser (robust to 'June 8, 2026' /
        'Jul. 8, 2025' / full-or-abbreviated month), NOT a brittle month-map."""
        import pdfplumber
        from dateutil import parser as _dp
        try:
            with pdfplumber.open(pdf_path) as pdf:
                t = re.sub(r'\s+', ' ',
                           ' '.join(pg.extract_text() or '' for pg in pdf.pages))
        except Exception:  # noqa: BLE001
            return {}
        loc = re.search(
            r'(BCPL\s*-\s*[A-Za-z]+(?:\s+[A-Za-z0-9]+)*(?:\s*-\s*[A-Za-z0-9]+)?'
            # allow an optional ' - ' right before the terminal word, so a store
            # like "BCPL - Mumbai M12 - Feeder" (dash before a bare 'Feeder' with
            # no trailing 'Warehouse') is captured, not just "… Feeder Warehouse".
            r'\s*(?:-\s*)?(?:SR\s+Feeder|Feeder\s+Warehouse|Feeder|Warehouse)'
            r'(?:\s*Warehouse)?)', t, re.I)

        def _date(pat):
            m = re.search(pat, t, re.I)
            if not m:
                return None
            try:
                return _dp.parse(m.group(1), dayfirst=False).date()
            except (ValueError, OverflowError):
                return None
        return {
            'location': loc.group(1).strip() if loc else '',
            # "Date :June 8, 2026, 6:47 p.m." / "R.O. expiry :July 8, 2026, …"
            'po_date': _date(r'\bDate\s*:\s*([A-Za-z]+\.?\s+\d{1,2},\s*\d{4})'),
            'exp_date': _date(r'expiry\s*:?\s*([A-Za-z]+\.?\s+\d{1,2},\s*\d{4})'),
        }

    def _source_dates_by_po(self) -> dict:
        """RO date → po_date, expiry → exp_date, keyed by ``ro_number`` (== the
        engine's po_number). The tabular dump has no dates, so this fills them
        (COALESCE — blanks only) so BlinkMP shows on the TAT tracker."""
        return dict(self._dates or {})


class RelianceProcessor(Processor):
    """Reliance (**online**, cust 20015) PO PDFs — base flow + one our-layer fix.

    The frozen ``reliance_pdf_parser`` sets ``__loc__`` to the delivery **city**,
    so Gurgaon-district DCs collide: ``GURGAON`` substring-matches
    ``'Reliance Retail Limited-Gurgaon'`` (20015_5) and never reaches Farukhnagar
    (20015_6). We read the delivery **pincode** and, for the deliberate pincode
    splits in :mod:`reliance_shipto`, override ``__loc__`` so the correct
    ``ship_to`` flows to the D365 output, the record AND the tracker Location
    (e.g. 122506 → 'FARUKHNAGAR' → 20015_6). Scoped to Reliance; unknown pincodes
    and all other channels are untouched. [[dmart-shipto-fix]]
    """

    def _run(self, skip_dedup=False):
        from . import reliance_shipto
        # Guard (block-on-ambiguous): a PO in a multi-DC city whose delivery
        # pincode is not a known DC would fall back to a city-only guess and
        # could ship to the WRONG DC. Block the run (nothing recorded, no PO
        # routed to a default) rather than guess. Single-DC cities and every
        # mapped pincode pass untouched. Mirrors DmartProcessor. [[never-skip-silently]]
        bad = {po: reason for po, (ok, reason)
               in reliance_shipto.confirm_paths(self.po_paths).items() if not ok}
        if bad:
            details = "\n• ".join(f"PO {po}: {r}" for po, r in bad.items())
            return {'ok': False, 'error':
                    "Reliance ship-to could not be confirmed — nothing was recorded "
                    "(no PO was routed to a default DC):\n• " + details}
        try:
            import online_po_processor.engine.reliance_pdf_parser as _rp
        except Exception:  # noqa: BLE001 — no parser → base flow, no override
            return super()._run(skip_dedup=skip_dedup)
        _orig = _rp.reliance_po_to_dataframe

        def _patched(po):
            df = _orig(po)
            new_loc = reliance_shipto.loc_override_for_pin(
                getattr(po.header, 'delivery_pin', ''))
            if new_loc:                       # only known pincode splits act
                df['__loc__'] = new_loc
            return df

        _rp.reliance_po_to_dataframe = _patched
        try:
            return super()._run(skip_dedup=skip_dedup)
        finally:                              # never leave the parser patched
            _rp.reliance_po_to_dataframe = _orig


class ZeptoProcessor(Processor):
    """Zepto has per-SKU negotiated **deal prices** (a flat 'Unit Base Cost',
    already net of GST). The frozen engine's deal path is Swiggy-ONLY, so — exactly
    like :class:`MyntraProcessor` — we apply Zepto's deal SKUs here and let the
    shared :meth:`_accept_deal_exceptions` mark them **OK + 'Zepto deal'** (the ⚑
    EXCEPTION column + yellow row highlight), with the deal price forced into the
    D365 unit price. AS-IS: the base cost is written straight through (no ÷(1+GST),
    unlike Myntra's with-GST transfer price). Zepto-ONLY; no other channel touched."""

    @staticmethod
    def _norm_ean(v) -> str:
        s = str(v if v is not None else '').strip()
        return s[:-2] if s.endswith('.0') else s

    def post_process(self, result, env) -> None:
        super().post_process(result, env)
        try:
            from . import overrides_store
            deal_map = overrides_store.zepto_deal_map()
        except Exception as e:  # noqa: BLE001
            deal_map = {}
            self.warnings.append(
                f"Zepto deal-SKU prices couldn't load ({type(e).__name__}) — deal "
                f"lines will fall back to flat margin, NOT the negotiated price. Verify prices.")
        if not deal_map:
            return
        applied = mismatched = 0
        for so in getattr(result, 'rows', None) or []:
            cp = deal_map.get(self._norm_ean(so.ean))
            if cp is None:
                continue
            # Vendor-CP acceptance (if any) wins, same rule as Myntra.
            if str(getattr(so, 'exception_label', '') or '').startswith('Vendor CP'):
                continue
            expected_cp = round(float(cp), 2)     # AS-IS base cost → D365 unit price
            so.cost_price_ref = expected_cp
            so.forced_unit_price = expected_cp    # deal price authoritative
            vendor_cp = getattr(so, 'fob_price', None)
            note_status = 'OK'
            try:
                diff = round(float(vendor_cp) - expected_cp, 2)
            except (TypeError, ValueError):
                diff = None
            if diff is not None and abs(diff) <= 1.0:
                so.validation_status = 'OK'
                so.exception_label = 'Zepto deal'
                applied += 1
            else:
                # _accept_deal_exceptions() normalises this back to 'Zepto deal' + OK
                so.validation_status = 'MISMATCH'
                so.exception_label = 'Deal ≠ vendor'
                note_status = 'MISMATCH vs deal'
                mismatched += 1
            try:
                result.exceptions_applied.append({
                    'type': 'price_override', 'po': str(so.po_number),
                    'ean': str(so.ean), 'item_no': str(so.item_no),
                    'detail': f"Zepto deal SKU — CP {expected_cp} "
                              f"(Unit Base Cost, as-is) [{note_status}]"})
            except Exception:  # noqa: BLE001
                pass
        if applied or mismatched:
            self.warnings.append(
                f"Zepto deal SKUs: {applied} line(s) at the negotiated base cost "
                f"(CP set, marked OK); {mismatched} differ from the PO price (kept "
                f"MISMATCH → accepted as exception, D365 uses the deal price). "
                f"Never silent.")


class PurplleProcessor(Processor):
    """Purplle: a single DC. The engine reads the RAW full ship-to address (kept as
    the ship_to_mapping del_location key), but the tracker shows the friendly mapped
    CITY ('Mumbai') instead of the whole address. Same idea as MyntraProcessor —
    resolution is unchanged, only the tracker's Location display is friendlier."""

    def _source_location_by_po(self) -> dict:
        return self._mapped_city_by_po()


_PROCESSORS = {'Flipkart': FlipkartProcessor, 'Flipkart-TO': FlipkartTOProcessor,
               'Meesho-TO': MeeshoTOProcessor, 'Dmart': DmartProcessor,
               'Firstcry': FirstcryProcessor, 'Myntra': MyntraProcessor,
               'Swiggy': SwiggyProcessor, 'BlinkMP': BlinkMPProcessor,
               'Reliance': RelianceProcessor, 'Zepto': ZeptoProcessor,
               'Purplle': PurplleProcessor}


def processor_for(marketplace, po_paths, warehouse=None, margin_pct=None,
                  ean_fixes=None) -> Processor:
    cls = _PROCESSORS.get(marketplace, Processor)
    return cls(marketplace, po_paths, warehouse, margin_pct, ean_fixes)


def preview(marketplace: str, po_paths, warehouse=None, margin_pct=None,
            ean_fixes=None) -> dict:
    return processor_for(marketplace, po_paths, warehouse, margin_pct,
                         ean_fixes).preview()


def confirm(marketplace: str, po_paths, warehouse=None, margin_pct=None,
            actions=None, ean_fixes=None, as_of=None, recorded_by=None) -> dict:
    return processor_for(marketplace, po_paths, warehouse, margin_pct,
                         ean_fixes).confirm(actions, as_of=as_of,
                                            recorded_by=recorded_by)


def generate_d365(marketplace: str, po_paths, out_path, warehouse=None,
                  margin_pct=None, actions=None, ean_fixes=None) -> dict:
    """Build the ERP D365 package reflecting the operator's locked Include/
    Override/Exclude decisions. Engine + full SO Workbook untouched."""
    return processor_for(marketplace, po_paths, warehouse, margin_pct,
                         ean_fixes).generate_d365(out_path, actions)


def export_decided_workbook(marketplace: str, po_paths, warehouse=None,
                            margin_pct=None, actions=None, ean_fixes=None,
                            exclude_uploaded_run_id=None) -> dict:
    """Full SO Workbook with the operator's locked decisions applied (accepted
    lines only, overrides repriced) — the post-lock "Completed" download. Pass
    ``exclude_uploaded_run_id`` (this run) to drop prior-run (already-uploaded)
    POs from the import file. The review download stays the full workbook."""
    return processor_for(marketplace, po_paths, warehouse, margin_pct,
                         ean_fixes).export_decided_workbook(
        actions, exclude_uploaded_run_id=exclude_uploaded_run_id)
