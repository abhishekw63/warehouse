"""
online_b2b.services.sales_validation — the ★ Sales Validator
============================================================

A pure, testable reconcile engine that checks a **D365 Sales export**
(Sales Order *Headers* + *Lines*) against the **recorded run** for that
dispatch — proving that what D365 booked equals what we recorded.

Design
------
* Channel-agnostic. LS (Lifestyle) is the first channel; the D365 SO links
  to a recorded order via ``External Document No.`` (= the marketplace PO) and
  the ``Ship-to Code`` (→ store / delivery location).
* Two possible *targets* to reconcile against, defaulting to the DB run:
    - a **recorded run** in our order DB (``order_headers`` / ``order_lines``),
      matched by parsing an SO's PO from the upload; OR
    - a **source file** (the raw marketplace ``.xlsb``) when no DB run exists.
* One public entry point, ``validate()``, returning a **JSON-safe dict**
  ``{ok, summary, checks, line_summary, line_findings, rows, excel_path}`` so
  the view stays thin and the same result renders on-page + drives the Excel.

Nothing here writes to the business DB — it only ``SELECT``s the item master
and ship-to mapping to resolve stores / expected items.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

# ── small cleaners (mirrors the proven reference scripts) ────────────────────


def _cln(v) -> str:
    return str(v if v is not None else '').strip()


def _ean(v) -> str:
    """Normalise a barcode: strip a trailing ``.0`` (Excel float) + leading 0s."""
    return _cln(v).replace('.0', '').lstrip('0')


def _int(v) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ── robust column resolution (D365 header names vary a little) ───────────────

def _resolve(header: list[str], *aliases: str) -> int | None:
    """Return the index of the first alias present (case/space-insensitive),
    else None. Aliases are matched loosely so ``Total Qty`` finds
    ``Total Quantity`` etc."""
    norm = {re.sub(r'[^a-z0-9]', '', h.lower()): i for i, h in enumerate(header)}
    for a in aliases:
        key = re.sub(r'[^a-z0-9]', '', a.lower())
        if key in norm:
            return norm[key]
    return None


def _read_sheet(path: str) -> tuple[list[str], list[tuple]]:
    """Read the first worksheet of an .xlsx/.xls → (header, rows)."""
    import openpyxl
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True).active
    it = ws.iter_rows(values_only=True)
    header = [_cln(c) for c in next(it)]
    rows = [r for r in it]
    return header, rows


# ── SO number → PO parsing (identifies the recorded run) ─────────────────────

_SO_PO_RE = re.compile(r'(\d{5,})')


def _po_from_so(external_doc: str, so_no: str) -> str:
    """The recorded PO is the D365 ``External Document No.`` (already the
    marketplace PO). We keep the SO only for display / run lookup."""
    return _cln(external_doc)


# ── loaders for the resolvers (item master + ship-to), read-only ─────────────

def _load_master() -> dict[str, tuple[str, str]]:
    """EAN → (item_no, description) from ``item_master``."""
    from .order_db import _conn
    out: dict[str, tuple[str, str]] = {}
    with _conn() as (cur, _d):
        cur.execute("SELECT ean, item_no, description FROM item_master")
        for ean, item_no, desc in cur.fetchall():
            out[_ean(ean)] = (_cln(item_no), _cln(desc))
    return out


def _load_ship_to(party: str) -> dict[str, dict]:
    """Ship-to Code → {store, postcode, address} from ``ship_to_mapping``."""
    from .order_db import _conn
    out: dict[str, dict] = {}
    with _conn() as (cur, _d):
        cur.execute(
            "SELECT ship_to, del_location, address2, postcode "
            "FROM ship_to_mapping WHERE party=%s"
            if _conn_is_mysql() else
            "SELECT ship_to, del_location, address2, postcode "
            "FROM ship_to_mapping WHERE party=?",
            (party,),
        )
        for shp, dl, a2, pin in cur.fetchall():
            m = re.search(r'(\d{3,5})\s*$', _cln(dl))
            out[_cln(shp)] = {
                'store': m.group(1) if m else '',
                'postcode': _cln(pin),
                'address': f"{_cln(a2)} - {_cln(pin)}".strip(' -'),
            }
    return out


def _conn_is_mysql() -> bool:
    from .order_db import _backend
    kind, _ = _backend()
    return kind == 'mysql'


# ── parse the two uploaded D365 files into normalised dicts ──────────────────

def _parse_headers(path: str) -> dict[str, dict]:
    header, rows = _read_sheet(path)
    ix = {
        'so': _resolve(header, 'No.'),
        'ext': _resolve(header, 'External Document No.'),
        'sell': _resolve(header, 'Sell-to Customer No.'),
        'ship': _resolve(header, 'Ship-to Code'),
        'name': _resolve(header, 'Ship-to Name'),
        'pin': _resolve(header, 'Ship-to Postcode'),
        'tqty': _resolve(header, 'Total Quantity', 'Total Qty'),
        'tamt': _resolve(header, 'Total Amount Incl. GST',
                         'Total Amount Including GST'),
        'loc': _resolve(header, 'Location Code'),
        'dd': _resolve(header, 'Document Date'),
    }
    if ix['so'] is None:
        raise ValueError("Headers file has no 'No.' (SO number) column.")
    out: dict[str, dict] = {}

    def g(r, k):
        i = ix[k]
        return r[i] if (i is not None and i < len(r)) else None

    for r in rows:
        so = _cln(g(r, 'so'))
        if not so:
            continue
        out[so] = {
            'ext': _cln(g(r, 'ext')),
            'sell': _cln(g(r, 'sell')),
            'ship': _cln(g(r, 'ship')),
            'name': _cln(g(r, 'name')),
            'pin': _cln(g(r, 'pin')),
            'tqty': _int(g(r, 'tqty')),
            'tamt': round(_num(g(r, 'tamt')), 2),
            'loc': _cln(g(r, 'loc')),
            'dd': _cln(g(r, 'dd'))[:10],
        }
    return out


def _parse_lines(path: str) -> list[dict]:
    header, rows = _read_sheet(path)
    ix = {
        'so': _resolve(header, 'Document No.'),
        'gtin': _resolve(header, 'GTIN', 'EAN'),
        'item': _resolve(header, 'No.'),
        'desc': _resolve(header, 'Description'),
        'qty': _resolve(header, 'Quantity', 'Qty'),
        'loc': _resolve(header, 'Location Code'),
        'amt': _resolve(header, 'Line Amount Excl. VAT'),
        'type': _resolve(header, 'Type'),
    }
    if ix['so'] is None or ix['qty'] is None:
        raise ValueError("Lines file missing 'Document No.' or 'Quantity'.")
    out: list[dict] = []

    def g(r, k):
        i = ix[k]
        return r[i] if (i is not None and i < len(r)) else None

    for r in rows:
        so = _cln(g(r, 'so'))
        if not so:
            continue
        out.append({
            'so': so,
            'ean': _ean(g(r, 'gtin')),
            'item': _cln(g(r, 'item')),
            'desc': _cln(g(r, 'desc')),
            'qty': _int(g(r, 'qty')),
            'loc': _cln(g(r, 'loc')),
            'amt': round(_num(g(r, 'amt')), 2),
            'type': _cln(g(r, 'type')),
        })
    return out


# ── build the TARGET (recorded run) — from DB or from a source .xlsb ──────────
#
# A target is keyed by (po, store) and carries per-line (ean → qty/item/desc)
# so we can reconcile both aggregate and line-level.

def _target_from_db(run_id: int) -> dict:
    """Assemble the recorded run from ``order_headers`` + ``order_lines``.

    Keyed by (external_doc/po, store). Store is derived from the header
    ``location`` (trailing digits) so it aligns with the ship-to mapping."""
    from .order_db import _conn
    ph = '%s' if _conn_is_mysql() else '?'
    groups: dict[tuple, dict] = {}
    with _conn() as (cur, _d):
        cur.execute(
            f"SELECT po, external_doc, location, qty, order_value "
            f"FROM order_headers WHERE run_id={ph}", (run_id,))
        hdr_rows = cur.fetchall()
        cur.execute(
            f"SELECT po, location, ean, item_no, description, qty "
            f"FROM order_lines WHERE run_id={ph}", (run_id,))
        line_rows = cur.fetchall()
    hdr_val = {}
    for po, ext, loc, _qty, val in hdr_rows:
        m = re.search(r'(\d{3,5})\s*$', _cln(loc))
        store = m.group(1) if m else ''
        key = (_cln(ext or po), store)
        hdr_val[key] = _num(val)
    for po, loc, ean, item_no, desc, qty in line_rows:
        m = re.search(r'(\d{3,5})\s*$', _cln(loc))
        store = m.group(1) if m else ''
        key = (_cln(po), store)
        g = groups.setdefault(key, {'q': 0, 'n': 0, 'val': 0.0, 'lines': {}})
        e = _ean(ean)
        g['q'] += _int(qty)
        g['n'] += 1
        g['lines'][e] = {'qty': _int(qty) + g['lines'].get(e, {}).get('qty', 0),
                         'item': _cln(item_no), 'desc': _cln(desc)}
    for key, g in groups.items():
        g['val'] = hdr_val.get(key, 0.0)
    return groups


def _target_from_source(xlsb_path: str) -> dict:
    """Assemble the target from the raw marketplace .xlsb (LS layout).

    Keyed by (Order No, Plant ID) → the recorded PO + store."""
    import pandas as pd
    df = pd.read_excel(xlsb_path, sheet_name='Sheet1', engine='pyxlsb')
    df = df[df['Order No'].notna()]
    groups: dict[tuple, dict] = {}
    for _, r in df.iterrows():
        po = str(_int(r['Order No']))
        store = str(_int(r['Plant ID']))
        ean = _ean(r['EAN/UPC'])
        qty = _int(pd.to_numeric(r['Final Order Qty'], errors='coerce'))
        val = _num(pd.to_numeric(r['Total Order value'], errors='coerce'))
        key = (po, store)
        g = groups.setdefault(key, {'q': 0, 'n': 0, 'val': 0.0, 'lines': {}})
        g['q'] += qty
        g['n'] += 1
        g['val'] += val
        prev = g['lines'].get(ean, {'qty': 0})
        g['lines'][ean] = {'qty': prev['qty'] + qty, 'item': '', 'desc': ''}
    for g in groups.values():
        g['val'] = round(g['val'], 2)
    return groups


# ── line-level reconciliation (the per-line tally) ───────────────────────────

def _reconcile_lines(hdr: dict, lines: list[dict], ship_to: dict,
                     target: dict, master: dict) -> tuple[list[dict], dict]:
    """For EVERY D365 line, match to the target by (PO, store) + EAN and mark
    OK / QTY_MISMATCH / ITEM_MISMATCH / MISSING_IN_TARGET / EXTRA_IN_D365.
    Also emit MISSING rows for target EANs with no D365 counterpart."""
    findings: list[dict] = []
    # aggregate D365 lines to (so, ean) so split lines compare cleanly
    d365_agg: dict[tuple, dict] = {}
    for ln in lines:
        k = (ln['so'], ln['ean'])
        g = d365_agg.setdefault(
            k, {'qty': 0, 'item': ln['item'], 'desc': ln['desc']})
        g['qty'] += ln['qty']
    seen_target_keys: set[tuple] = set()
    for (so, ean), g in d365_agg.items():
        h = hdr.get(so, {})
        store = ship_to.get(h.get('ship', ''), {}).get('store', '')
        tkey = (h.get('ext', ''), store)
        tgt = target.get(tkey)
        expected_item = master.get(ean, ('', ''))[0]
        rec = {
            'so': so, 'store': store, 'ean': ean,
            'item_no': g['item'], 'description': g['desc'],
            'd365_qty': g['qty'], 'target_qty': None,
            'status': '', 'detail': '',
        }
        if not tgt:
            rec['status'] = 'MISSING_IN_TARGET'
            rec['detail'] = f"SO {so}: no recorded (PO {tkey[0]}, store {store})"
        elif ean not in tgt['lines']:
            rec['status'] = 'EXTRA_IN_D365'
            rec['detail'] = f"EAN {ean} not in recorded (PO {tkey[0]}, store {store})"
        else:
            seen_target_keys.add((tkey, ean))
            tq = tgt['lines'][ean]['qty']
            rec['target_qty'] = tq
            if g['qty'] != tq:
                rec['status'] = 'QTY_MISMATCH'
                rec['detail'] = f"D365 qty {g['qty']} vs recorded {tq}"
            elif expected_item and g['item'] and g['item'] != expected_item:
                rec['status'] = 'ITEM_MISMATCH'
                rec['detail'] = (f"item {g['item']} vs item_master "
                                 f"{expected_item} for EAN {ean}")
            else:
                rec['status'] = 'OK'
                rec['detail'] = ''
        findings.append(rec)
    # target lines with no D365 counterpart → MISSING (never-silent)
    # reverse map (ext, store) -> so for display
    key_to_so = {}
    for so, h in hdr.items():
        store = ship_to.get(h.get('ship', ''), {}).get('store', '')
        key_to_so.setdefault((h.get('ext', ''), store), so)
    for tkey, tg in target.items():
        for ean, tl in tg['lines'].items():
            if (tkey, ean) in seen_target_keys:
                continue
            findings.append({
                'so': key_to_so.get(tkey, ''), 'store': tkey[1], 'ean': ean,
                'item_no': tl.get('item', ''), 'description': tl.get('desc', ''),
                'd365_qty': None, 'target_qty': tl['qty'],
                'status': 'MISSING',
                'detail': f"recorded EAN {ean} (PO {tkey[0]}, store {tkey[1]}) "
                          f"absent in D365",
            })
    counts = Counter(f['status'] for f in findings)
    summary = {
        'total': len(findings),
        'ok': counts.get('OK', 0),
        'qty_mismatch': counts.get('QTY_MISMATCH', 0),
        'item_mismatch': counts.get('ITEM_MISMATCH', 0),
        'missing': counts.get('MISSING_IN_TARGET', 0) + counts.get('MISSING', 0),
        'extra': counts.get('EXTRA_IN_D365', 0),
    }
    return findings, summary


# ── per-SO side-by-side rows (drives the on-page table + Excel sheet) ─────────

def _side_by_side(hdr: dict, lines: list[dict], ship_to: dict,
                  target: dict) -> list[dict]:
    dl = defaultdict(lambda: {'q': 0, 'n': 0})
    for ln in lines:
        dl[ln['so']]['q'] += ln['qty']
        dl[ln['so']]['n'] += 1
    rows: list[dict] = []
    for so in sorted(hdr, key=lambda s: (hdr[s]['ext'],
                     ship_to.get(hdr[s]['ship'], {}).get('store', ''))):
        h = hdr[so]
        store = ship_to.get(h['ship'], {}).get('store', '?')
        s = target.get((h['ext'], store), {'q': 0, 'n': 0, 'val': 0.0})
        d = dl.get(so, {'q': 0, 'n': 0})
        po_ok = bool(h['ext'])
        qty_ok = d['q'] == s['q']
        ln_ok = d['n'] == s['n']
        vdiff = round(h['tamt'] - round(s.get('val', 0.0), 2), 2)
        verdict = 'OK' if (qty_ok and ln_ok and po_ok) else 'CHECK'
        rows.append({
            'store': store, 'po': h['ext'], 'so': so,
            'ship': h['ship'], 'name': h['name'], 'pin': h['pin'],
            'address': ship_to.get(h['ship'], {}).get('address', ''),
            'src_qty': s['q'], 'd365_qty': d['q'], 'qty_ok': qty_ok,
            'src_lines': s['n'], 'd365_lines': d['n'], 'lines_ok': ln_ok,
            'src_value': round(s.get('val', 0.0), 2), 'd365_value': h['tamt'],
            'value_diff': vdiff, 'po_ok': po_ok, 'verdict': verdict,
        })
    return rows


# ── the reconciliation Excel (side-by-side sheet + line-detail sheet) ─────────

def _write_excel(rows: list[dict], line_findings: list[dict], out_path: str, *,
                 summary: dict, checks: list[dict],
                 line_summary: dict) -> str:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    G = PatternFill('solid', fgColor='C6EFCE')
    R = PatternFill('solid', fgColor='FFC7CE')
    A = PatternFill('solid', fgColor='FFEB9C')
    Hf = PatternFill('solid', fgColor='1F4E78')
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary (overall verdict + check list + metrics) ──
    wss = wb.active
    wss.title = 'Summary'
    val = summary.get('value', {})
    wss.append(['SALES VALIDATION — Summary'])
    wss['A1'].font = Font(bold=True, size=13, color='1F4E78')
    wss.append([])
    verdict = 'ALL CHECKS PASS' if summary.get('all_pass') else 'CHECK — FAILURES'
    wss.append(['Overall', verdict])
    wss.cell(wss.max_row, 2).fill = G if summary.get('all_pass') else R
    wss.cell(wss.max_row, 2).font = Font(bold=True)
    metrics = [
        ('Target', summary.get('target_kind', '')),
        ('D365 headers', summary.get('headers')),
        ('D365 lines', summary.get('lines')),
        ('Total qty', summary.get('total_qty')),
        ('Recorded units', summary.get('recorded_units')),
        ('Checks passed', f"{summary.get('checks_pass')}/{summary.get('checks_total')}"),
        ('D365 value (incl GST)', val.get('d365_value')),
        ('Recorded value', val.get('target_value')),
        ('Value gap (flagged)', f"{val.get('gap')} ({val.get('gap_pct')}%)"),
        ('Lines OK', line_summary.get('ok')),
        ('Lines qty-mismatch', line_summary.get('qty_mismatch')),
        ('Lines item-mismatch', line_summary.get('item_mismatch')),
        ('Lines missing', line_summary.get('missing')),
        ('Lines extra', line_summary.get('extra')),
    ]
    for k, v in metrics:
        wss.append([k, v])
    wss.append([])
    wss.append(['Check', 'Result', 'Detail'])
    hdr_row = wss.max_row
    for c in wss[hdr_row]:
        c.fill = Hf
        c.font = Font(bold=True, color='FFFFFF', size=9)
    for ch in checks:
        wss.append([ch['label'], 'PASS' if ch['pass'] else 'FAIL', ch['detail']])
        wss.cell(wss.max_row, 2).fill = G if ch['pass'] else R
    # value row (flagged, amber)
    wss.append(['Value (flagged, non-failing)', 'FLAG', val.get('note', '')])
    wss.cell(wss.max_row, 2).fill = A
    for i, w in enumerate([34, 16, 60], 1):
        wss.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Headers Reco (per-SO side-by-side) ──
    ws = wb.create_sheet('Headers Reco')
    cols = ['Store #', 'PO (recorded)', 'PO (D365 Ext Doc)', 'PO ✓',
            'SO No. (D365)', 'Ship-to Code', 'Ship-to Name', 'Postcode',
            'Delivery Address (our)', 'Rec Qty', 'D365 Qty', 'Qty ✓',
            'Rec Lines', 'D365 Lines', 'Lines ✓', 'Rec Value',
            'D365 Value(incGST)', 'Value Δ', 'VERDICT']
    ws.append(cols)
    for c in ws[1]:
        c.fill = Hf
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.alignment = Alignment(wrap_text=True, vertical='center')
    allok = 0
    for r in rows:
        if r['verdict'] == 'OK':
            allok += 1
        ws.append([r['store'], r['po'], r['po'], 'Y' if r['po_ok'] else 'N',
                   r['so'], r['ship'], r['name'], r['pin'], r['address'],
                   r['src_qty'], r['d365_qty'], 'Y' if r['qty_ok'] else 'N',
                   r['src_lines'], r['d365_lines'], 'Y' if r['lines_ok'] else 'N',
                   r['src_value'], r['d365_value'], r['value_diff'], r['verdict']])
        row = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(row, c).font = Font(size=9)
            ws.cell(row, c).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row, cols.index('Qty ✓') + 1).fill = G if r['qty_ok'] else R
        ws.cell(row, cols.index('Lines ✓') + 1).fill = G if r['lines_ok'] else R
        ws.cell(row, cols.index('PO ✓') + 1).fill = G if r['po_ok'] else R
        ws.cell(row, cols.index('Value Δ') + 1).fill = A
        ws.cell(row, cols.index('VERDICT') + 1).fill = G if r['verdict'] == 'OK' else R
    ws.append(['TOTAL', '', '', '', f'{len(rows)} SOs', '', '', '', '',
               sum(r['src_qty'] for r in rows), sum(r['d365_qty'] for r in rows),
               '', sum(r['src_lines'] for r in rows),
               sum(r['d365_lines'] for r in rows), '',
               round(sum(r['src_value'] for r in rows), 2),
               round(sum(r['d365_value'] for r in rows), 2), '',
               f'{allok}/{len(rows)} OK'])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    widths = [8, 13, 14, 6, 16, 11, 26, 9, 40, 9, 9, 7, 9, 10, 7, 12, 14, 10, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # ── Sheet 3: Lines Reco (the full per-line tally) ──
    ws2 = wb.create_sheet('Lines Reco')
    lcols = ['SO', 'Store', 'EAN', 'Item No.', 'Description', 'D365 Qty',
             'Target Qty', 'Status', 'Note']
    ws2.append(lcols)
    for c in ws2[1]:
        c.fill = Hf
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.alignment = Alignment(wrap_text=True, vertical='center')
    # OK first-or-not: keep upload order but colour by status
    for f in line_findings:
        ws2.append([f['so'], f['store'], f['ean'], f['item_no'],
                    f['description'], f['d365_qty'], f['target_qty'],
                    f['status'], f['detail']])
        row = ws2.max_row
        ok = f['status'] == 'OK'
        for c in range(1, len(lcols) + 1):
            ws2.cell(row, c).font = Font(size=9)
        ws2.cell(row, lcols.index('Status') + 1).fill = G if ok else R
    lwidths = [16, 8, 16, 12, 42, 9, 10, 18, 40]
    for i, w in enumerate(lwidths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ── the aggregate check list (green/red per check) ───────────────────────────

def _build_checks(hdr: dict, lines: list[dict], ship_to: dict, target: dict,
                  master: dict, line_summary: dict) -> tuple[list[dict], dict]:
    checks: list[dict] = []

    def ck(key, label, ok, detail=''):
        checks.append({'key': key, 'label': label, 'pass': bool(ok),
                       'detail': detail})

    n_hdr = len(hdr)
    n_units = len(target)
    ck('count_headers', 'Header count == distinct recorded SOs/units',
       n_hdr == n_units, f"D365 headers {n_hdr} vs recorded units {n_units}")

    n_lines = len(lines)
    tgt_lines = sum(g['n'] for g in target.values())
    ck('count_lines', 'Line count matches recorded',
       n_lines == tgt_lines, f"D365 lines {n_lines} vs recorded {tgt_lines}")

    d365_qty = sum(ln['qty'] for ln in lines)
    tgt_qty = sum(g['q'] for g in target.values())
    ck('qty_total', 'Total quantity exact',
       d365_qty == tgt_qty, f"D365 qty {d365_qty} vs recorded {tgt_qty}")

    hdr_qty = sum(h['tqty'] for h in hdr.values())
    ck('qty_header_sum', 'Σ header Total Qty == recorded total qty',
       hdr_qty == tgt_qty, f"Σ header {hdr_qty} vs recorded {tgt_qty}")

    # per-line (store, EAN) → qty parity
    d365_se: dict[tuple, int] = defaultdict(int)
    for ln in lines:
        st = ship_to.get(hdr.get(ln['so'], {}).get('ship', ''), {}).get('store', '')
        if st:
            d365_se[(st, ln['ean'])] += ln['qty']
    tgt_se: dict[tuple, int] = defaultdict(int)
    for (_po, store), g in target.items():
        for ean, tl in g['lines'].items():
            tgt_se[(store, ean)] += tl['qty']
    se_ok = dict(d365_se) == dict(tgt_se)
    ck('qty_line_level', 'Per-line (store, EAN) → qty matches recorded', se_ok,
       f"{len(d365_se)} D365 keys vs {len(tgt_se)} recorded keys")

    # PO: every header External Doc resolves to a recorded (po, store)
    po_bad = 0
    for h in hdr.values():
        store = ship_to.get(h['ship'], {}).get('store', '')
        if (h['ext'], store) not in target:
            po_bad += 1
    ck('po_match', 'Every header External Doc matches a recorded PO+store',
       po_bad == 0, f"{po_bad} unmatched")

    # ITEM↔EAN via item_master
    item_bad = 0
    for ln in lines:
        m = master.get(ln['ean'])
        if m and m[0] and ln['item'] and m[0] != ln['item']:
            item_bad += 1
    ck('item_ean', "Every line's item No. matches item_master for its EAN",
       item_bad == 0, f"{item_bad} mismatched")

    # ship-to resolves to a store
    noship = sum(1 for h in hdr.values()
                 if not ship_to.get(h['ship'], {}).get('store'))
    ck('ship_to', 'Every header Ship-to resolves to a store',
       noship == 0, f"{noship} unresolved")

    # postcode alignment (soft where mapping has no pin)
    pin_bad = []
    for so, h in hdr.items():
        mp = ship_to.get(h['ship'], {}).get('postcode', '')
        if mp and h['pin'] and mp != h['pin']:
            pin_bad.append(so)
    ck('postcode', 'Header postcode matches ship_to_mapping',
       not pin_bad, f"{len(pin_bad)} mismatched")

    # header Total Qty == Σ its line qty
    hdr_lineqty: dict[str, int] = defaultdict(int)
    for ln in lines:
        hdr_lineqty[ln['so']] += ln['qty']
    hdr_line_ok = all(hdr[so]['tqty'] == hdr_lineqty[so] for so in hdr)
    ck('header_lines', 'Each header Total Qty == Σ its line qty', hdr_line_ok,
       f"{sum(1 for so in hdr if hdr[so]['tqty'] != hdr_lineqty[so])} off")

    # INTEGRITY — orphan lines, dup item in SO, blank item/EAN
    orphan = sum(1 for ln in lines if ln['so'] not in hdr)
    ck('no_orphan', 'No orphan lines (every line Document No. in headers)',
       orphan == 0, f"{orphan} orphans")
    dup = defaultdict(Counter)
    for ln in lines:
        dup[ln['so']][ln['item']] += 1
    dupes = [(so, it, c) for so, cc in dup.items() for it, c in cc.items()
             if it and c > 1]
    ck('no_dup_item', 'No duplicate item within an SO', not dupes,
       f"{len(dupes)} dup item/SO")
    blank = sum(1 for ln in lines if not ln['ean'] or not ln['item'])
    ck('no_blank', 'No line with blank item / EAN', blank == 0,
       f"{blank} blank")

    # line-level reconcile roll-up (never-silent structural gate)
    line_clean = (line_summary['qty_mismatch'] == 0
                  and line_summary['item_mismatch'] == 0
                  and line_summary['missing'] == 0
                  and line_summary['extra'] == 0)
    ck('line_reconcile',
       'Per-line reconciliation clean (no qty/item/missing/extra)',
       line_clean,
       f"ok {line_summary['ok']} · qty {line_summary['qty_mismatch']} · "
       f"item {line_summary['item_mismatch']} · missing {line_summary['missing']} "
       f"· extra {line_summary['extra']}")

    # VALUE — flagged, non-failing
    d365_val = round(sum(h['tamt'] for h in hdr.values()), 2)
    tgt_val = round(sum(g['val'] for g in target.values()), 2)
    gap = round(d365_val - tgt_val, 2)
    gap_pct = round((gap / tgt_val * 100), 3) if tgt_val else 0.0
    value = {
        'd365_value': d365_val, 'target_value': tgt_val,
        'gap': gap, 'gap_pct': gap_pct,
        'note': ('D365 prices from its own master — a small gap is expected '
                 'for mapping-only channels; flagged, not failed.'),
    }
    return checks, value


# ── PUBLIC ENTRY POINT ───────────────────────────────────────────────────────

def validate(headers_path: str, lines_path: str, *, run_id: int | None = None,
             source_path: str | None = None, party: str = 'LS',
             excel_out: str | None = None) -> dict:
    """Reconcile a D365 Sales export against the recorded run.

    Parameters
    ----------
    headers_path, lines_path : str
        The two uploaded D365 files (Sales Order Headers + Sales Lines).
    run_id : int, optional
        Recorded run to reconcile against (DB target). Preferred when present.
    source_path : str, optional
        Raw marketplace .xlsb to reconcile against when no DB run exists.
    party : str
        Channel key for ship-to resolution (default 'LS').
    excel_out : str, optional
        Where to write the reconciliation workbook. Defaults to a temp file
        next to the headers upload.

    Returns
    -------
    dict — JSON-safe:
        ``{ok, summary, checks, value, line_summary, line_findings, rows,
           excel_path, error?}``
    """
    try:
        hdr = _parse_headers(headers_path)
        lines = _parse_lines(lines_path)
        if not hdr:
            return {'ok': False, 'error': 'No headers parsed from the file.'}
        master = _load_master()
        ship_to = _load_ship_to(party)

        if run_id is not None:
            target = _target_from_db(int(run_id))
            target_kind = f'DB run #{run_id}'
        elif source_path:
            target = _target_from_source(source_path)
            target_kind = f'source file {Path(source_path).name}'
        else:
            return {'ok': False,
                    'error': 'No target: pass run_id or source_path.'}
        if not target:
            return {'ok': False,
                    'error': f'Target ({target_kind}) has no recorded lines.'}

        line_findings, line_summary = _reconcile_lines(
            hdr, lines, ship_to, target, master)
        checks, value = _build_checks(
            hdr, lines, ship_to, target, master, line_summary)
        rows = _side_by_side(hdr, lines, ship_to, target)

        n_pass = sum(1 for c in checks if c['pass'])
        d365_qty = sum(ln['qty'] for ln in lines)
        summary = {
            'target_kind': target_kind,
            'headers': len(hdr),
            'lines': len(lines),
            'total_qty': d365_qty,
            'recorded_units': len(target),
            'checks_pass': n_pass,
            'checks_total': len(checks),
            'all_pass': n_pass == len(checks),
            'value': value,
        }

        if excel_out is None:
            excel_out = str(Path(headers_path).with_name(
                'reconciliation.xlsx'))
        excel_path = _write_excel(
            rows, line_findings, excel_out,
            summary=summary, checks=checks, line_summary=line_summary)

        return {
            'ok': True,
            'summary': summary,
            'checks': checks,
            'value': value,
            'line_summary': line_summary,
            'line_findings': line_findings,
            'rows': rows,
            'excel_path': excel_path,
        }
    except Exception as e:  # noqa: BLE001
        return {'ok': False, 'error': f'{type(e).__name__}: {e}'}


# ── run discovery for the UI dropdown / auto-match ───────────────────────────

def recent_runs(limit: int = 30) -> list[dict]:
    """Recent runs for the picker (id, ts, marketplaces, POs, qty)."""
    from .order_db import _conn
    out: list[dict] = []
    try:
        with _conn() as (cur, _d):
            cur.execute(
                "SELECT run_id, run_ts, marketplaces, total_pos, total_qty "
                "FROM runs ORDER BY run_id DESC LIMIT %s"
                if _conn_is_mysql() else
                "SELECT run_id, run_ts, marketplaces, total_pos, total_qty "
                "FROM runs ORDER BY run_id DESC LIMIT ?", (limit,))
            for rid, ts, mps, pos, qty in cur.fetchall():
                out.append({'run_id': rid, 'run_ts': str(ts),
                            'marketplaces': _cln(mps), 'pos': pos, 'qty': qty})
    except Exception:  # noqa: BLE001
        pass
    return out


def match_run(headers_path: str) -> list[int]:
    """Auto-match: from the uploaded headers' External Docs (recorded POs),
    find run_ids in ``order_headers`` carrying those POs. Returns candidate
    run_ids (empty if none)."""
    from .order_db import _conn
    try:
        hdr = _parse_headers(headers_path)
    except Exception:  # noqa: BLE001
        return []
    pos = {h['ext'] for h in hdr.values() if h['ext']}
    if not pos:
        return []
    ph = '%s' if _conn_is_mysql() else '?'
    placeholders = ','.join([ph] * len(pos))
    runs: list[int] = []
    try:
        with _conn() as (cur, _d):
            cur.execute(
                f"SELECT run_id, COUNT(*) c FROM order_headers "
                f"WHERE external_doc IN ({placeholders}) "
                f"GROUP BY run_id ORDER BY c DESC", tuple(pos))
            runs = [r[0] for r in cur.fetchall()]
    except Exception:  # noqa: BLE001
        pass
    return runs
