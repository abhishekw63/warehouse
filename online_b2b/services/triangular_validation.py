"""
online_b2b.services.triangular_validation  —  3-way triangular check (WRAPPER)
==============================================================================

Adds the **DUMP leg** and a **file-upload audit** ON TOP of the existing,
untouched :mod:`online_b2b.services.full_validation` (which already reconciles
D365 ⟷ system for all marketplaces, with excluded/final accounting).

    LEG 1  D365   (Sales Orders headers + Sales Lines)   ← full_validation reads
    LEG 3  System (our DB)                               ← full_validation reads
    LEG 2  Dump   (raw *_completed.xlsx, one per MP)      ← THIS module adds

ZERO-TOUCH ISOLATION — this module only *imports and calls*
``full_validation.validate()`` and reads the raw dumps. It modifies NOTHING in
full_validation or anywhere else, and writes NOTHING to the business DB (read
only). Delete this file and the app is byte-identical.

It layers on:
  • the **DUMP** leg — dump final/full qty + ship-to per PO, cross-checked
    against D365 and system (a true 3-way triangle per PO);
  • the **INCLUDED value delta** — full_validation reports the *full*-value gap;
    here we subtract the deducted (EXCLUDED) line value so the value ties to ~0
    apples-to-apples against D365 (the phantom-gap fix from the daily CLI);
  • a **file audit** — every uploaded file, the POs it carried, and which POs
    are in D365 but in no dump (a **missing dump/file**) or in a dump but not in
    D365 (recorded but **not pushed**) — so a missed file/PO is never invisible.

Public entry: ``validate(headers_path, lines_path, dumps, excel_out=None)``
  → ``{ok, error, data}`` where ``data`` carries the full_validation result plus
    ``{value_legs, triangle, dump_issues, audit}``.
"""
from __future__ import annotations

import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

from . import full_validation as _fv
from .order_db import _conn


# ── tiny coercions ───────────────────────────────────────────────────────────
def _s(x):
    return str(x).strip() if x is not None else ''


def _gm(c):
    c = str(c or '').upper()
    return (1.28 if '28' in c else 1.18 if '18' in c else 1.12 if '12' in c
            else 1.05 if ('5' in c and '15' not in c and '25' not in c)
            else 1.03 if '3' in c else 1.0 if '0' in c else 1.18)


def _line_val(oland, up, gst, q):
    """inc-GST value of a line: our_landing × qty (landing is already inc-GST),
    else unit_price × qty × (1+GST). NOT our_cp (that's the cost price, a
    different basis that does NOT compare to D365's booked value)."""
    if oland not in (None, 0, '0'):
        return float(oland) * q
    if up not in (None, ''):
        return float(up) * q * _gm(gst)
    return 0.0


_AFF = {'MISMATCH', 'NOT_IN_MASTER'}


def _is_dropped(status, action):
    act = _s(action).upper()
    return act == 'EXCLUDE' or (_s(status) in _AFF and act not in ('INCLUDE', 'OVERRIDE'))


def _db_legs(pos):
    """Per-PO value legs on the inc-GST our_landing basis (the basis that tied to
    26 paise on 2026-07-18), keeping only each PO's LATEST run (re-run safe).
    Returns {po: {mp, full_val, ded_val, incl_val, dropped:[...]}}. Read-only."""
    pos = sorted({_s(p) for p in pos if _s(p)})
    out = {}
    if not pos:
        return out
    with _conn() as (cur, d):
        fmt = ','.join([d['ph']] * len(pos)) if isinstance(d, dict) and 'ph' in d \
            else ','.join(['%s'] * len(pos))
        cur.execute(f"""SELECT run_id, marketplace, po, item_no, ean, description, qty,
                               our_landing, unit_price, gst_code, status, action
                        FROM order_lines_full WHERE po IN ({fmt})""", tuple(pos))
        keys = [c[0] for c in cur.description]
        rows = [dict(zip(keys, r)) for r in cur.fetchall()]
    latest = {}
    for r in rows:
        po = _s(r['po'])
        if po not in latest or r['run_id'] > latest[po]:
            latest[po] = r['run_id']
    for r in rows:
        po = _s(r['po'])
        if r['run_id'] != latest[po]:
            continue
        o = out.setdefault(po, {'mp': _disp(_s(r['marketplace'])), 'full_val': 0.0,
                                'ded_val': 0.0, 'incl_val': 0.0,
                                'incl_qty': 0, 'drop_qty': 0, 'dropped': []})
        q = int(r['qty'] or 0)
        v = _line_val(r['our_landing'], r['unit_price'], r['gst_code'], q)
        o['full_val'] += v
        if _is_dropped(r['status'], r['action']):
            o['ded_val'] += v
            o['drop_qty'] += q            # qty matters even when the line has no price
            reason = ('EXCLUDE' if _s(r['action']).upper() == 'EXCLUDE'
                      else f"{_s(r['status'])} (unresolved)")
            o['dropped'].append({'po': po, 'mp': o['mp'], 'item': _s(r['item_no']),
                                 'ean': _s(r['ean']), 'desc': _s(r['description']),
                                 'qty': q, 'val': round(v, 2), 'reason': reason})
        else:
            o['incl_val'] += v
            o['incl_qty'] += q
    return out


def _f(x):
    try:
        return float(str(x).strip())
    except (TypeError, ValueError):
        return 0.0


def _sheet(wb, name):
    return list(wb[name].iter_rows(values_only=True)) if name in wb.sheetnames else []


def _hdr(rows):
    return {_s(c): i for i, c in enumerate(rows[0])} if rows else {}


def _raw_sheet_confirms(wb, summary_full):
    """Does the dump's verbatim ``Raw Data`` sheet (the engine's untouched capture
    of the marketplace PO) confirm the unified ``Summary Total Qty``? Tries every
    qty-like column (channel formats differ) and returns (True, col) if ANY totals
    exactly to ``summary_full``; (False, '') if a Raw Data sheet exists but none
    match; (None, '') if there is no Raw Data sheet. This proves the raw total we
    reconcile against equals the untouched raw — without guessing the column."""
    if 'Raw Data' not in wb.sheetnames:
        return None, ''
    rows = _sheet(wb, 'Raw Data')
    if not rows:
        return None, ''
    for ci, cname in enumerate(rows[0]):
        low = _s(cname).lower()
        if 'qty' not in low and 'quantity' not in low:
            continue
        tot = sum(int(_f(r[ci])) for r in rows[1:]
                  if r and ci < len(r) and r[ci] is not None)
        if tot == summary_full and summary_full:
            return True, _s(cname)
    return False, ''


# ── LEG 2: load the raw dump(s) — per-PO + per-file (for the audit) ───────────
def _load_dumps(paths):
    """Return (per_po, per_item, files, issues).

    per_po[po]  = {ship, raw, final_qty, full_qty, dump_file}
    per_item    = {(po, item): qty}   (dump final lines)
    files       = [{name, pos:[...], skipped:bool, reason}]  (audit trail)
    issues      = [str]  (never-silent: skipped/odd files)
    """
    per_po, per_item, files, issues = {}, defaultdict(int), [], []
    per_warn = {}          # (po, item) -> Warnings text (cost mismatch etc.); item -> text fallback
    for wbp in paths:
        base = os.path.basename(_s(wbp))
        entry = {'name': base, 'pos': [], 'skipped': False, 'reason': ''}
        if 'Flipkart-TO' in base:
            entry.update(skipped=True, reason='Flipkart-TO transfer order (not an SO)')
            issues.append(f"{base}: skipped (transfer order, not an SO).")
            files.append(entry); continue
        try:
            wb = openpyxl.load_workbook(wbp, data_only=True, read_only=True)
        except Exception as e:  # noqa: BLE001
            entry.update(skipped=True, reason=f'open failed ({type(e).__name__})')
            issues.append(f"{base}: could not open ({type(e).__name__}).")
            files.append(entry); continue
        if not {'Headers (SO)', 'Summary', 'Lines (SO)'} <= set(wb.sheetnames):
            entry.update(skipped=True, reason='missing Headers(SO)/Summary/Lines(SO)')
            issues.append(f"{base}: skipped (missing standard sheets).")
            files.append(entry); continue
        hd = _sheet(wb, 'Headers (SO)'); hh = _hdr(hd)
        so2po = {}
        for r in hd[1:]:
            if r and hh.get('No.') is not None and _s(r[hh['No.']]):
                ext = hh.get('External Document No.', hh.get('No.'))
                so2po[_s(r[hh['No.']])] = _s(r[ext])
        sm = _sheet(wb, 'Summary'); sh = _hdr(sm)
        fq = sh.get('Final Qty (to D365)', sh.get('Total Qty')); tq = sh.get('Total Qty')
        seen = set(); dump_full = 0; dup_pos = set()
        for r in sm[1:]:
            po_so = _s(r[sh['PO']]) if 'PO' in sh else ''
            if not po_so or po_so.upper().startswith('TOTAL') or 'Marketplace:' in po_so:
                continue
            po = so2po.get(po_so, po_so)
            if po in per_po:                 # this PO already came from an EARLIER dump
                dup_pos.add(po)              # → duplicate; keep the first, skip this one
                continue
            full_q = int(_f(r[tq])) if tq is not None else 0
            per_po[po] = {'ship': _s(r[sh.get('Ship-to', -1)]),
                          'raw': _s(r[sh.get('Location (Raw)', -1)]),
                          'final_qty': int(_f(r[fq])) if fq is not None else 0,
                          'full_qty': full_q,
                          'dump_file': base}
            seen.add(po); dump_full += full_q
        if dup_pos:
            issues.append(f"{base}: {len(dup_pos)} PO(s) already in an earlier dump — "
                          f"skipped as duplicates (kept the first) so nothing is "
                          f"double-counted: {', '.join(sorted(dup_pos)[:6])}"
                          f"{'…' if len(dup_pos) > 6 else ''}.")
        # verify the untouched Raw Data sheet confirms this dump's raw total
        confirmed, rcol = _raw_sheet_confirms(wb, dump_full)
        entry['raw_confirmed'] = confirmed
        entry['raw_qty_col'] = rcol
        entry['full_qty'] = dump_full
        # 'unconfirmed' is informational (the raw sheet may carry extra POs so a
        # whole-sheet total won't match) — NOT a warning; the Summary raw is used
        # and the raw=deducted+included identity still holds. Surfaced in raw_status.
        ln = _sheet(wb, 'Lines (SO)'); lh = _hdr(ln)
        for r in ln[1:]:
            if not r or lh.get('No.') is None or _s(r[lh['No.']]) == '':
                continue
            po = so2po.get(_s(r[lh['Document No.']]), _s(r[lh['Document No.']]))
            if po in dup_pos:                # duplicate PO — its lines already counted
                continue
            it = _s(r[lh['No.']]); q = int(_f(r[lh['Quantity']]))
            if it:
                per_item[(po, it)] += q
        # Warnings sheet → per-line reason (cost mismatch etc.), keyed (po, item)
        # with an item-only fallback (same SKU flagged across POs). Additive.
        wn = _sheet(wb, 'Warnings'); wh = _hdr(wn)
        c_wpo, c_wloc, c_wtxt = wh.get('PO'), wh.get('Location'), wh.get('Warning')
        if c_wtxt is not None:
            for r in wn[1:]:
                if not r or c_wtxt >= len(r):
                    continue
                txt = _s(r[c_wtxt])
                if not txt:
                    continue
                wpo = so2po.get(_s(r[c_wpo]), _s(r[c_wpo])) if c_wpo is not None else ''
                wit = _s(r[c_wloc]) if c_wloc is not None else ''
                if wit:
                    per_warn.setdefault(wit, txt)              # item-only fallback
                    if wpo:
                        per_warn[(wpo, wit)] = txt             # precise (po,item)
        entry['pos'] = sorted(seen)
        files.append(entry)
    return per_po, per_item, files, issues, per_warn


def _load_d365_lines(so_path, lines_path):
    """D365 per-(PO, item) qty for the line-level D365 ⟷ dump edge — scoped to the
    day's SOs (a broad Sales Lines export must not pollute). Returns {(po,it): qty}.
    PO is the External-Doc (falls back to SO 'No.'), matching the dump/system keys."""
    out = defaultdict(int)
    if not lines_path:
        return out
    wso = openpyxl.load_workbook(so_path, data_only=True, read_only=True)
    sr = _sheet(wso, wso.sheetnames[0]); h = _hdr(sr)
    c_no, c_ext = h.get('No.'), h.get('External Document No.')
    so2po = {}
    for r in sr[1:]:
        if not r or c_no is None or _s(r[c_no]) == '':
            continue
        no = _s(r[c_no]); po = _s(r[c_ext]) if c_ext is not None else no
        so2po[no] = po or no
    valid = set(so2po) | set(so2po.values())
    wl = openpyxl.load_workbook(lines_path, data_only=True, read_only=True)
    lr = _sheet(wl, wl.sheetnames[0]); lh = _hdr(lr)
    c_doc, c_it, c_q = lh.get('Document No.'), lh.get('No.'), lh.get('Quantity')
    if c_doc is None or c_it is None or c_q is None:
        return out
    for r in lr[1:]:
        if not r or _s(r[c_it]) == '':
            continue
        doc = _s(r[c_doc])
        if doc not in valid:                          # not one of today's SOs → skip
            continue
        out[(so2po.get(doc, doc), _s(r[c_it]))] += int(_f(r[c_q]))
    return out


# online-B2B posting group(s) in the D365 SO; everything else (GT Select / GT Mass
# / MT Select / Subsidiary) is a DIFFERENT segment the online check must ignore.
_ONLINE_POSTING_GROUPS = {'ON-B2B'}


def _load_d365_segments(so_path):
    """{PO (External Doc, else SO 'No.'): Gen. Bus. Posting Group} from the D365
    Sales Orders. Used to keep only ON-B2B POs and bucket the rest (GT Select,
    GT Mass, MT, Subsidiary) as 'other segment — ignored'. Empty if the column is
    absent (then no filtering happens — backward compatible)."""
    wso = openpyxl.load_workbook(so_path, data_only=True, read_only=True)
    sr = _sheet(wso, wso.sheetnames[0]); h = _hdr(sr)
    c_no, c_ext = h.get('No.'), h.get('External Document No.')
    c_g = h.get('Gen. Bus. Posting Group') or h.get('Gen Bus Posting Group')
    out = {}
    if c_g is None or c_no is None:
        return out
    for r in sr[1:]:
        if not r or _s(r[c_no]) == '':
            continue
        no = _s(r[c_no]); po = _s(r[c_ext]) if c_ext is not None else no
        out[po or no] = _s(r[c_g])
    return out


# ── marketplace identity by name (1st) then content (2nd) ────────────────────
# Labels align to order_headers.marketplace_label so the name guess can be
# cross-checked against the DB-derived MP. Both the dump names (Blinkit_…,
# RK_…, Bigbasket_…) and the RAW input names (bulk_po_csv, POItemExport, 4126…)
# resolve to the same label. First matching pattern wins.
_MP_NAME_PATTERNS = [
    ('Blinkit',    [r'^blinkit', r'bulk_po_csv', r'\bblink']),
    ('RK',         [r'^rk[_\- ]', r'poitemexport', r'reliancekart']),
    ('Big Basket', [r'^bigbasket', r'big\s*basket', r'^\d{7,}\.xlsx$', r'\bira\d']),
    ('Swiggy',     [r'^swiggy', r'purchase\s*orders_allpo']),
    ('Zepto',      [r'^zepto', r'^po_\d+\.csv$']),
    ('Flipkart',   [r'^flipkart', r'\bfk[_\- ]', r'flipkart', r'order-line-items']),
    ('Nykaa',      [r'^nykaa', r'nykaa']),
    ('Apollo',     [r'^apollo', r'apollo']),
    ('Zepto',      [r'^p\d{6,}']),   # Zepto POs are P-numbers (last, weakest)
]


# Operator-facing label fixes for the web view (e.g. BlinkMP records under its
# party_name 'Blink RO' in the DB — show 'BlinkMP' like the dump Tracker does).
_MP_DISPLAY = {'Blink RO': 'BlinkMP'}


def _disp(mp):
    return _MP_DISPLAY.get(str(mp).strip(), mp)


def _mp_from_name(basename):
    """Marketplace label guessed from a filename (dump OR raw input), '' if none."""
    n = str(basename).lower()
    for mp, pats in _MP_NAME_PATTERNS:
        for p in pats:
            if re.search(p, n):
                return mp
    return ''


def _mp_matches(name_mp, db_mps):
    """Loose label match (alnum-only, substring either way) so 'Big Basket' ==
    'Bigbasket'. Used to confirm the name guess against the DB-derived MP."""
    a = re.sub(r'[^a-z0-9]', '', str(name_mp).lower())
    for m in db_mps or []:
        b = re.sub(r'[^a-z0-9]', '', str(m).lower())
        if a and (a in b or b in a):
            return True
    return False


# ── smart upload bifurcation (drop everything, we sort it) ───────────────────
def classify_file(path):
    """Peek at a file and name its role: 'headers' (D365 Sales Orders) · 'lines'
    (D365 Sales Lines) · 'dump' (raw *_completed.xlsx) · 'source' (.xlsb) ·
    'unknown'. By CONTENT (columns / sheets), robust to filenames."""
    name = os.path.basename(str(path)).lower()
    if name.endswith('.xlsb'):
        return 'source'
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return 'unknown'
    sheets = set(wb.sheetnames)

    def _row1(ws):
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return {_s(c).lower() for c in row if c is not None}
        return set()

    # ONLINE completed dump — the '_completed' marker, OR the 3 SO sheets WITH the
    # online Summary's own column (so a GT-Mass / other output with lookalike
    # sheets isn't mistaken for a dump). Strict, because a whole-folder upload
    # mixes many workbooks.
    if '_completed' in name:
        return 'dump'
    if {'Headers (SO)', 'Summary', 'Lines (SO)'} <= sheets:
        try:
            scol = _row1(wb['Summary'])
            if 'final qty (to d365)' in scol or 'total amount (inc gst)' in scol:
                return 'dump'
        except Exception:  # noqa: BLE001
            pass
    # Any OTHER 'Headers (SO)'/'Lines (SO)' workbook (e.g. a GT-Mass dump) mimics the
    # D365 columns but is NOT a raw D365 export — the real export is a plain
    # single-sheet file. Exclude dump-family workbooks from the D365 match.
    if 'Headers (SO)' in sheets or 'Lines (SO)' in sheets:
        return 'unknown'
    # D365 exports — by a DISTINCTIVE multi-column signature (a single column like
    # 'External Document No.' also appears in dump sheets, so require a pair).
    try:
        hdr = _row1(wb[wb.sheetnames[0]])
    except Exception:  # noqa: BLE001
        return 'unknown'
    if 'document no.' in hdr and ('gtin' in hdr or 'line amount excl. vat' in hdr):
        return 'lines'                                 # D365 Sales Lines
    if 'external document no.' in hdr and 'ship-to code' in hdr:
        return 'headers'                               # D365 Sales Orders
    return 'unknown'


def _newest_key(path):
    """Sort key to pick the NEWEST D365 export among duplicates: prefer the ISO
    timestamp in the D365 name ('Sales Orders - 2026-07-20T113006'), else a
    trailing '(N)' ('Sales Lines (7)'), else the basename."""
    n = os.path.basename(str(path))
    m = re.search(r'(\d{4}-\d{2}-\d{2}T\d{6})', n)
    if m:
        return (2, m.group(1))
    m = re.search(r'\((\d+)\)', n)
    if m:
        return (1, f'{int(m.group(1)):09d}')
    return (0, n)


def classify_files(paths):
    """Bifurcate a dropped batch → {headers, lines, dumps[], source[], unknown[],
    superseded[], summary:[{name, role, marketplace, mp_by}]}. The marketplace is
    guessed by NAME first. If MORE THAN ONE D365 Sales Orders / Sales Lines is
    present (an old + a new export), the **newest** wins (by the timestamp/(N) in
    the filename) and the older one(s) go to ``superseded`` (never silently
    picked)."""
    out = {'headers': None, 'lines': None, 'dumps': [], 'source': [],
           'unknown': [], 'superseded': [], 'summary': []}
    hdr_c, lines_c, rows = [], [], []
    for p in paths:
        base = os.path.basename(str(p))
        role = classify_file(p)
        mp = _mp_from_name(base)
        display = 'raw' if (role in ('unknown', 'source') and mp) else role
        rows.append({'name': base, 'role': display, 'marketplace': mp,
                     'mp_by': 'name' if mp else ''})
        if role == 'headers':
            hdr_c.append(str(p))
        elif role == 'lines':
            lines_c.append(str(p))
        elif role == 'dump':
            out['dumps'].append(str(p))
        elif role == 'source':
            out['source'].append(str(p))
        else:
            out['unknown'].append(str(p))

    def _pick(cands, label):
        if not cands:
            return None
        best = max(cands, key=_newest_key)
        for c in cands:
            if c != best:
                out['superseded'].append({'name': os.path.basename(c), 'role': label})
        return best
    out['headers'] = _pick(hdr_c, 'headers')
    out['lines'] = _pick(lines_c, 'lines')
    # relabel the older (superseded) exports in the summary so the UI shows them
    _sup = {s['name'] for s in out['superseded']}
    for r in rows:
        if r['name'] in _sup and r['role'] in ('headers', 'lines'):
            r['role'] = r['role'] + '·old'
    out['summary'] = rows
    return out


# ── the wrapper entry point ──────────────────────────────────────────────────
def validate(headers_path, lines_path, dumps, excel_out=None):
    """3-way triangular validation. Calls the untouched full_validation for the
    D365 ⟷ system reconciliation, then layers the dump leg + included-value
    delta + file audit. Returns ``{ok, error, data}``."""
    try:
        base = _fv.validate(headers_path, lines_path, excel_out=None)
        if not base.get('ok'):
            return {'ok': False, 'error': base.get('error', 'full_validation failed.'),
                    'data': None}

        # normalize operator-facing MP labels (e.g. 'Blink RO' → 'BlinkMP')
        for r in base.get('headers', []):
            r['mp'] = _disp(r.get('mp'))
        for r in base.get('lines', []):
            r['mp'] = _disp(r.get('mp'))
        for m in base.get('marketplaces', []):
            m['mp'] = _disp(m.get('mp'))

        # ── keep only ON-B2B POs; bucket the other D365 segments (GT Select /
        #    GT Mass / MT / Subsidiary) as 'ignored' so they never appear as
        #    UNKNOWN or inflate the online Δ. Uses D365's own Gen. Bus. Posting
        #    Group — no hardcoded customer list. No column → no filtering. ──
        seg = _load_d365_segments(headers_path)
        online_pos = ({p for p, g in seg.items() if g in _ONLINE_POSTING_GROUPS}
                      if seg else None)
        other_segments = []
        if online_pos is not None:
            keep, drop = [], []
            for r in base['headers']:
                (keep if r['po'] in online_pos else drop).append(r)
            base['headers'] = keep
            base['lines'] = [ln for ln in base['lines'] if ln['po'] in online_pos]
            oseg = defaultdict(lambda: {'pos': 0, 'qty': 0, 'val': 0.0})
            for r in drop:
                g = seg.get(r['po'], '') or '(blank)'
                oseg[g]['pos'] += 1
                oseg[g]['qty'] += int(r.get('d365_qty') or 0)
                oseg[g]['val'] += float(r.get('d365_val') or 0.0)
            other_segments = [{'group': g, 'pos': v['pos'], 'qty': v['qty'],
                               'value': round(v['val'], 2)}
                              for g, v in sorted(oseg.items())]

        d_po, d_item, files, dump_issues, d_warn = _load_dumps(list(dumps or []))

        # per-PO MP map. Start from the D365 headers, then fill dump-only POs
        # (e.g. BlinkMP — present in a dump / our DB but NOT in this D365 export)
        # from the DB so the marketplace is never blank anywhere on the page.
        po_mp = {r['po']: r['mp'] for r in base['headers']}
        d365_pos = set(po_mp)
        dump_pos = set(d_po)
        legs = _db_legs(d365_pos | dump_pos)            # our_landing basis (26-paise fix)
        po_mp_full = {po: L['mp'] for po, L in legs.items()}   # DB label (BlinkMP…)
        po_mp_full.update({p: m for p, m in po_mp.items() if m})  # D365 label wins

        # enrich each dropped line's reason with the dump's Warnings text (the
        # cost-mismatch numbers) — precise (po,item) first, else item-only. Additive.
        for L in legs.values():
            for dl in L['dropped']:
                w = d_warn.get((dl['po'], dl['item'])) or d_warn.get(dl['item'])
                if w:
                    dl['reason'] = w

        # ── LINE-LEVEL D365 ⟷ dump edge (the "foolproof, line-wise" check the
        #    per-PO triangle alone can't catch — right PO total, wrong SKU split). ──
        d365_lines = _load_d365_lines(headers_path, lines_path)
        if online_pos is not None:
            d365_lines = {k: v for k, v in d365_lines.items() if k[0] in online_pos}
        sku_rows = []
        for k in sorted(set(d365_lines) | set(d_item)):
            dq, oq = d365_lines.get(k, 0), d_item.get(k, 0)
            if dq != oq:
                sku_rows.append({'po': k[0], 'item': k[1],
                                 'mp': po_mp_full.get(k[0]) or '—',
                                 'd365_qty': dq, 'dump_qty': oq, 'delta': dq - oq})
        sku_check = {'keys': len(set(d365_lines) | set(d_item)),
                     'mismatches': len(sku_rows), 'rows': sku_rows}

        # ── value legs on the inc-GST our_landing basis (from the DB). ──
        full_v = defaultdict(float); ded_v = defaultdict(float); incl_v = defaultdict(float)
        d365_v = defaultdict(float)
        for po, L in legs.items():
            # Group by the HEADER marketplace label (same as the D365 leg below),
            # not the order_lines_full.marketplace column — the DB stores the label
            # two ways ('Big Basket' vs 'Bigbasket'), which would split one MP.
            mp = po_mp_full.get(po) or L['mp'] or 'UNKNOWN'
            full_v[mp] += L['full_val']; ded_v[mp] += L['ded_val']; incl_v[mp] += L['incl_val']
        for r in base['headers']:
            d365_v[r['mp']] += (r['d365_val'] or 0.0)
        value_legs = []
        tot = {'full': 0.0, 'deducted': 0.0, 'included': 0.0, 'd365': 0.0}
        for mp in sorted(set(full_v) | set(d365_v)):
            full = round(full_v[mp], 2); ded = round(ded_v[mp], 2)
            incl = round(incl_v[mp], 2); d3 = round(d365_v[mp], 2)
            value_legs.append({'mp': mp, 'full': full, 'deducted': ded,
                               'included': incl, 'd365': d3,
                               'delta': round(d3 - incl, 2)})
            tot['full'] += full; tot['deducted'] += ded; tot['included'] += incl; tot['d365'] += d3
        value_legs.append({'mp': 'TOTAL', 'full': round(tot['full'], 2),
                           'deducted': round(tot['deducted'], 2),
                           'included': round(tot['included'], 2),
                           'd365': round(tot['d365'], 2),
                           'delta': round(tot['d365'] - tot['included'], 2)})
        dropped = [ln for L in legs.values() for ln in L['dropped']]

        # ── QTY legs: RAW (dump 'Total Qty', confirmed against the verbatim Raw
        #    Data sheet) = DEDUCTED + INCLUDED, and INCLUDED == D365. The explicit
        #    "we carried the whole raw order to D365 minus the logged drops" proof.
        #    raw_ties flags any MP where the dump raw ≠ system (included+deducted). ──
        raw_q = defaultdict(int); inc_q = defaultdict(int); drp_q = defaultdict(int); d3_q = defaultdict(int)
        for po, info in d_po.items():
            raw_q[po_mp_full.get(po) or 'UNKNOWN'] += info.get('full_qty', 0)
        for po, L in legs.items():
            mp = po_mp_full.get(po) or L['mp'] or 'UNKNOWN'
            inc_q[mp] += L['incl_qty']; drp_q[mp] += L['drop_qty']
        for r in base['headers']:
            d3_q[r['mp']] += r['d365_qty']
        qty_legs = []
        qt = {'raw': 0, 'deducted': 0, 'included': 0, 'd365': 0}
        qt_val = 0.0
        for mp in sorted(set(raw_q) | set(d3_q) | set(inc_q)):
            raw, ded, inc, d3 = raw_q[mp], drp_q[mp], inc_q[mp], d3_q[mp]
            val = round(d365_v.get(mp, 0.0), 2)               # D365 value (inc GST) per MP
            qty_legs.append({'mp': mp, 'raw': raw, 'deducted': ded, 'included': inc,
                             'd365': d3, 'delta': d3 - inc, 'raw_ties': raw == inc + ded,
                             'value': val})
            qt['raw'] += raw; qt['deducted'] += ded; qt['included'] += inc; qt['d365'] += d3
            qt_val += val
        qty_legs.append({'mp': 'TOTAL', 'raw': qt['raw'], 'deducted': qt['deducted'],
                         'included': qt['included'], 'd365': qt['d365'],
                         'delta': qt['d365'] - qt['included'],
                         'raw_ties': qt['raw'] == qt['included'] + qt['deducted'],
                         'value': round(qt_val, 2)})
        raw_unconfirmed = [f['name'] for f in files if f.get('raw_confirmed') is False]
        raw_status = {
            'all_confirmed': not raw_unconfirmed,
            'unconfirmed': raw_unconfirmed,
            'confirmed_files': [{'name': f['name'], 'col': f.get('raw_qty_col', '')}
                                for f in files if f.get('raw_confirmed') is True],
            'qty_ties': all(r['raw_ties'] for r in qty_legs),
        }

        # ── per-PO triangle: D365 qty ⟷ system final qty ⟷ dump final qty ──
        triangle = []
        for r in base['headers']:
            po = r['po']
            dqd = d_po.get(po, {})
            d365_q = r['d365_qty']
            sys_final = r['final']            # system: our_qty − excluded
            dump_final = dqd.get('final_qty')
            agree = (dump_final is None) or (d365_q == sys_final == dump_final)
            flags = []
            if dump_final is None:
                flags.append('no dump for this PO')
            elif not (d365_q == sys_final == dump_final):
                flags.append(f'QTY d365 {d365_q} · system {sys_final} · dump {dump_final}')
            if dqd and r.get('ship_d365') and dqd.get('ship') and r['ship_d365'] != dqd['ship']:
                flags.append(f"SHIP dump {dqd['ship']} vs d365 {r['ship_d365']}")
            # per-PO VALUE (Total Amount inc GST): D365 header vs our INCLUDED value
            d365_val = round(float(r.get('d365_val') or 0), 2)
            our_val = round(float(legs.get(po, {}).get('incl_val', 0.0)), 2)
            val_delta = round(d365_val - our_val, 2)
            val_ok = abs(val_delta) < max(2.0, our_val * 0.005)   # ≤ ₹2 or 0.5%
            if not val_ok and our_val:
                flags.append(f"VALUE d365 ₹{d365_val:.0f} vs ours ₹{our_val:.0f}")
            # ── standard format: Raw = Filled + Excluded (per PO) ──
            #   raw      = the untouched marketplace order (dump Total Qty)
            #   filled   = what reached D365
            #   excluded = system-logged drops for this PO
            raw_q = dqd.get('full_qty')            # None if no dump for this PO
            drop_q = int(legs.get(po, {}).get('drop_qty', 0))
            ident_ok = (raw_q is None) or (raw_q == d365_q + drop_q)
            addr_ok = not (dqd and r.get('ship_d365') and dqd.get('ship')
                           and r['ship_d365'] != dqd['ship'])
            triangle.append({'mp': r['mp'], 'po': po, 'd365_qty': d365_q,
                             'system_final': sys_final, 'dump_final': dump_final,
                             'raw_qty': raw_q, 'filled_qty': d365_q, 'excluded_qty': drop_q,
                             'ident_ok': ident_ok, 'addr_ok': addr_ok,
                             'd365_val': d365_val, 'our_val': our_val,
                             'val_delta': val_delta, 'val_ok': val_ok,
                             'dump_ship': dqd.get('ship', ''), 'd365_ship': r.get('ship_d365', ''),
                             'agree': agree and not flags, 'note': ' · '.join(flags)})

        # ── file audit: which file carried which POs + missing/extra ──
        #    2nd check: the dump's NAME-guessed MP vs its DB (content) MP — flag
        #    a name-says-X / content-says-Y mismatch (never silent).
        mp_conflicts = []
        for fe in files:
            mps = sorted({po_mp_full.get(p) or 'UNKNOWN' for p in fe['pos']})
            fe['marketplaces'] = mps
            fe['po_count'] = len(fe['pos'])
            fe['name_mp'] = _mp_from_name(fe['name'])
            db_mps = [m for m in mps if m and m != 'UNKNOWN']
            fe['mp_mismatch'] = bool(fe['name_mp']) and bool(db_mps) and not _mp_matches(fe['name_mp'], db_mps)
            if fe['mp_mismatch']:
                mp_conflicts.append(f"{fe['name']}: name says '{fe['name_mp']}' but content is "
                                    f"{', '.join(db_mps)}.")
        missing_dump = sorted(d365_pos - dump_pos)     # in D365, no dump uploaded
        not_pushed = sorted(dump_pos - d365_pos)       # in a dump, not in D365
        audit = {
            'files': files,
            'd365_pos': len(d365_pos),
            'dump_pos': len(dump_pos),
            'mp_conflicts': mp_conflicts,
            'missing_dump': [{'po': p, 'mp': po_mp_full.get(p) or 'UNKNOWN'} for p in missing_dump],
            'not_pushed': [{'po': p, 'dump_file': d_po[p]['dump_file']} for p in not_pushed],
        }

        # ── COVERAGE: every D365 order AND every D365 line must have a matching
        #    raw — so nothing is skipped. Two gaps, both surfaced loudly:
        #      • PO-level  — a D365 PO with NO raw uploaded  (raw file missing)
        #      • line-level — a D365 (PO,SKU) line absent from the raw
        #    Grouped by marketplace so the operator knows WHICH raw file to add. ──
        pos_missing_raw = audit['missing_dump']
        lines_missing_raw = [{'po': r['po'], 'mp': r['mp'], 'item': r['item'],
                              'd365_qty': r['d365_qty']}
                             for r in sku_check['rows']
                             if r['dump_qty'] == 0 and r['d365_qty'] > 0]
        miss_by_mp = defaultdict(lambda: {'pos': 0, 'lines': 0, 'qty': 0})
        for m in pos_missing_raw:
            miss_by_mp[m['mp']]['pos'] += 1
        for l in lines_missing_raw:
            miss_by_mp[l['mp']]['lines'] += 1
            miss_by_mp[l['mp']]['qty'] += l['d365_qty']
        coverage = {
            'd365_pos': len(d365_pos), 'd365_lines': len(d365_lines),
            'pos_missing_raw': pos_missing_raw,
            'lines_missing_raw': lines_missing_raw,
            'by_mp': [{'mp': k, **v} for k, v in sorted(miss_by_mp.items())],
            'full': not pos_missing_raw and not lines_missing_raw,
        }

        # ── connect the dots: every check → one verdict + the review headline ──
        vt = value_legs[-1]; qtot = qty_legs[-1]
        ship_flags = [t for t in triangle if 'SHIP' in t['note']]
        qty_flags = [t for t in triangle if not t['agree'] and 'SHIP' not in t['note']]
        n_mp = len([q for q in qty_legs if q['mp'] != 'TOTAL'])
        checks = [
            {'key': 'cover', 'label': 'Full coverage — nothing skipped',
             'sub': 'every D365 order & line has raw', 'ok': coverage['full'],
             'detail': (f"{coverage['d365_pos']} POs · {coverage['d365_lines']} lines covered"
                        if coverage['full'] else
                        f"{len(coverage['pos_missing_raw'])} PO · "
                        f"{len(coverage['lines_missing_raw'])} line(s) missing raw")},
            {'key': 'raw', 'label': 'Raw order fully accounted',
             'sub': 'Raw = Deducted + Included', 'ok': raw_status['qty_ties'],
             'detail': f"{qtot['raw']:,} = {qtot['deducted']:,} dropped + {qtot['included']:,} pushed"},
            {'key': 'incl', 'label': 'Included qty reached D365',
             'sub': 'Included = D365', 'ok': qtot['delta'] == 0,
             'detail': ('exact' if qtot['delta'] == 0 else f"Δ {qtot['delta']} units")},
            {'key': 'qty', 'label': 'Qty triangle',
             'sub': 'D365 = system = dump', 'ok': not qty_flags,
             'detail': f"{len(triangle)} POs · {len(qty_flags)} flagged"},
            {'key': 'sku', 'label': 'SKU lines',
             'sub': 'D365 = raw dump, line-wise', 'ok': sku_check['mismatches'] == 0,
             'detail': f"{sku_check['keys']} keys · {sku_check['mismatches']} mismatch"},
            {'key': 'value', 'label': 'Value',
             'sub': 'Included value vs D365', 'ok': abs(vt['delta']) < 2,
             'detail': f"Δ ₹{vt['delta']}"},
            {'key': 'addr', 'label': 'Ship-to address',
             'sub': 'dump = D365', 'ok': not ship_flags,
             'detail': ('all match' if not ship_flags else f"{len(ship_flags)} mismatch")},
            {'key': 'audit', 'label': 'File coverage',
             'sub': 'every PO has a dump, none unpushed', 'ok': not audit['missing_dump'] and not audit['not_pushed'],
             'detail': f"{len(audit['missing_dump'])} missing · {len(audit['not_pushed'])} not-pushed"},
            {'key': 'mp', 'label': 'Marketplace ID',
             'sub': 'name ↔ content', 'ok': not audit['mp_conflicts'],
             'detail': ('confirmed' if not audit['mp_conflicts'] else f"{len(audit['mp_conflicts'])} conflict")},
        ]
        n_fail = sum(1 for c in checks if not c['ok'])
        verdict = {
            'all_clean': n_fail == 0, 'n_fail': n_fail, 'n_checks': len(checks),
            'raw': qtot['raw'], 'included': qtot['included'], 'deducted': qtot['deducted'],
            'd365': qtot['d365'], 'n_mp': n_mp, 'n_pos': len(triangle),
            'value_delta': vt['delta'],
            'headline': (f"{qtot['raw']:,} units ordered → {qtot['included']:,} pushed to D365 "
                         f"+ {qtot['deducted']:,} logged drops · {n_mp} marketplace(s) · "
                         f"{len(triangle)} POs"),
        }

        data = dict(base)                              # full_validation result, intact
        data.update({'value_legs': value_legs, 'qty_legs': qty_legs,
                     'raw_status': raw_status, 'triangle': triangle,
                     'sku_check': sku_check, 'checks': checks, 'verdict': verdict,
                     'dropped': dropped, 'dump_issues': dump_issues, 'audit': audit,
                     'coverage': coverage, 'other_segments': other_segments,
                     'included_delta': value_legs[-1]['delta']})
        return {'ok': True, 'error': None, 'data': data}
    except Exception as e:  # noqa: BLE001
        return {'ok': False, 'error': f'{type(e).__name__}: {e}', 'data': None}


# ── the 360° triangular workbook (Value Legs · Qty Triangle · Audit · Dropped) ─
def build_workbook(data, out_path):
    """Write the 360° triangular Excel and return out_path. Four sheets:
    Value Legs (full = deducted + included, Δ vs D365) · Qty Triangle (D365 ⟷
    system ⟷ dump) · File Audit (files → POs → MP, missing/extra) · Dropped Lines."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    NAVY = PatternFill('solid', fgColor='1D2433'); GR = PatternFill('solid', fgColor='E7F6EF')
    AM = PatternFill('solid', fgColor='FDF1E3'); RD = PatternFill('solid', fgColor='FDECEC')
    HF = Font(bold=True, color='FFFFFF'); thin = Side(style='thin', color='D5D9E2')
    BD = Border(thin, thin, thin, thin)
    wb = openpyxl.Workbook()

    def head(ws, cols, widths):
        ws.append(cols)
        for c in ws[1]:
            c.fill = NAVY; c.font = HF
            c.alignment = Alignment('center', 'center', wrap_text=True); c.border = BD
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    # 1 — Qty Legs (RAW = Deducted + Included ; Included == D365)
    ws0 = wb.active; ws0.title = 'Qty Legs (Raw)'
    head(ws0, ['Marketplace', 'RAW (full order)', 'Deducted', 'Included (pushed)',
               'D365', 'Δ (D365 − Incl)', 'Raw = Ded+Incl'], [18, 16, 12, 17, 12, 16, 15])
    for q in data.get('qty_legs', []):
        ws0.append([q['mp'], q['raw'], q['deducted'], q['included'], q['d365'],
                    q['delta'], 'YES' if q['raw_ties'] else 'NO'])
        rr = ws0.max_row
        for c in ws0[rr]:
            c.border = BD
            if q['mp'] == 'TOTAL':
                c.font = Font(bold=True)
        ws0.cell(rr, 6).fill = GR if q['delta'] == 0 else AM
        ws0.cell(rr, 7).fill = GR if q['raw_ties'] else RD

    # 2 — Value Legs
    ws = wb.create_sheet('Value Legs')
    head(ws, ['Marketplace', 'Full (inc GST)', 'Deducted', 'Included (pushed)',
              'D365', 'Δ (D365 − Included)'], [18, 16, 14, 17, 16, 20])
    for v in data.get('value_legs', []):
        ws.append([v['mp'], v['full'], v['deducted'], v['included'], v['d365'], v['delta']])
        rr = ws.max_row
        for c in ws[rr]:
            c.border = BD
            if v['mp'] == 'TOTAL':
                c.font = Font(bold=True)
        ws.cell(rr, 6).fill = GR if abs(v['delta']) < 1 else AM

    # 2 — Qty Triangle
    ws2 = wb.create_sheet('Qty Triangle')
    head(ws2, ['MP', 'PO', 'D365 Qty', 'System Final', 'Dump Final', 'Agree', 'Note'],
         [13, 18, 10, 12, 11, 8, 44])
    for t in data.get('triangle', []):
        ws2.append([t['mp'], t['po'], t['d365_qty'], t['system_final'],
                    t['dump_final'], 'Y' if t['agree'] else 'N', t['note']])
        for c in ws2[ws2.max_row]:
            c.border = BD
        ws2.cell(ws2.max_row, 6).fill = GR if t['agree'] else RD

    # 2b — SKU: D365 ⟷ dump (line-level; lists only mismatches)
    sc = data.get('sku_check', {}) or {}
    wsk = wb.create_sheet('SKU D365 vs Dump')
    head(wsk, ['MP', 'PO', 'Item No', 'D365 Qty', 'Dump Qty', 'Δ'],
         [13, 18, 12, 10, 10, 8])
    if sc.get('rows'):
        for r_ in sc['rows']:
            wsk.append([r_['mp'], r_['po'], r_['item'], r_['d365_qty'], r_['dump_qty'], r_['delta']])
            for c in wsk[wsk.max_row]:
                c.border = BD; c.fill = RD
    else:
        wsk.append(['— all lines match —', '', '', '', '', ''])
        wsk.cell(wsk.max_row, 1).fill = GR
    wsk.append([])
    wsk.append([f"{sc.get('keys', 0)} (PO,SKU) keys checked · {sc.get('mismatches', 0)} mismatch(es)"])
    wsk.cell(wsk.max_row, 1).font = Font(bold=True)

    # 3 — File Audit
    ws3 = wb.create_sheet('File Audit')
    head(ws3, ['File', 'Marketplace(s)', '#PO', 'POs', 'Skipped / note'],
         [46, 18, 7, 60, 26])
    a = data.get('audit', {})
    for fe in a.get('files', []):
        ws3.append([fe['name'], ', '.join(fe.get('marketplaces', [])), fe['po_count'],
                    ', '.join(fe['pos']), fe['reason'] if fe['skipped'] else ''])
        for c in ws3[ws3.max_row]:
            c.border = BD
            if fe['skipped']:
                c.fill = AM
    ws3.append([])
    ws3.append([f"MISSING DUMP (in D365, no dump uploaded): {len(a.get('missing_dump', []))}"])
    ws3.cell(ws3.max_row, 1).font = Font(bold=True)
    for m in a.get('missing_dump', []):
        ws3.append([f"   {m['mp']}   PO {m['po']}"]); ws3.cell(ws3.max_row, 1).fill = RD
    ws3.append([f"NOT PUSHED (in dump, not in D365): {len(a.get('not_pushed', []))}"])
    ws3.cell(ws3.max_row, 1).font = Font(bold=True)
    for m in a.get('not_pushed', []):
        ws3.append([f"   PO {m['po']}  ({m['dump_file']})"]); ws3.cell(ws3.max_row, 1).fill = RD

    # 4 — Dropped Lines
    ws4 = wb.create_sheet('Dropped Lines')
    head(ws4, ['PO', 'Marketplace', 'Item No', 'EAN', 'Description', 'Dropped Qty',
               'Value (inc GST)', 'Reason'], [16, 12, 12, 16, 36, 12, 15, 22])
    dropped = data.get('dropped', [])
    dq = dv = 0
    for r_ in sorted(dropped, key=lambda x: (x['po'], x['item'], x['ean'])):
        ws4.append([r_['po'], r_['mp'], r_['item'], r_['ean'], r_['desc'],
                    r_['qty'], round(r_['val'], 2), r_['reason']])
        for c in ws4[ws4.max_row]:
            c.border = BD; c.fill = AM
        dq += r_['qty']; dv += r_['val']
    ws4.append(['TOTAL', '', '', '', f"{len(dropped)} lines", dq, round(dv, 2), ''])
    for c in ws4[ws4.max_row]:
        c.font = Font(bold=True)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
