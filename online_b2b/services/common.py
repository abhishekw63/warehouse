"""online_b2b.services.common — small shared web helpers.

Behavior-preserving consolidation of tiny utilities that were copy-pasted across
the views (path-traversal guard, upload save, AJAX detection, POST-dict). Pure and
side-effect-free except where noted; import and delegate to these instead of
re-implementing. See [[page-asset-separation]]/[[dry-skeleton-first]] spirit for the
Python side.
"""
from __future__ import annotations

from pathlib import Path

from django.http import Http404


def token_dir(root: Path, token: str) -> Path:
    """Resolve ``root/token`` and refuse anything that escapes ``root`` (path
    traversal). Returns the resolved Path or raises Http404 — identical to the
    per-feature ``_tok_dir``/``_token_dir`` guards it replaces."""
    root = Path(root)
    base = root.resolve()
    d = (root / token).resolve()
    if d != base and base not in d.parents:
        raise Http404()
    return d


def save_upload(fileobj, dest) -> Path:
    """Stream an uploaded file to ``dest`` in chunks (never loads it all into
    memory). Returns the destination Path. Caller ensures the parent dir exists."""
    dest = Path(dest)
    with open(dest, 'wb') as out:
        for chunk in fileobj.chunks():
            out.write(chunk)
    return dest


def is_ajax(request) -> bool:
    """True for an XMLHttpRequest fetch (the app's AJAX convention). Matches the
    existing ``_is_ajax`` exactly (header ``X-Requested-With: XMLHttpRequest``)."""
    return request.headers.get('x-requested-with') == 'XMLHttpRequest'


def post_dict(request) -> dict:
    """``request.POST`` as a plain dict minus the CSRF token — the exact pattern
    used by the ship-to / eka CRUD endpoints."""
    return {k: v for k, v in request.POST.items() if k != 'csrfmiddlewaretoken'}


def xlsx_response(sheet_title, columns, rows, filename,
                  width_cap=48, str_cols=(), freeze=False):
    """Build a downloadable .xlsx HttpResponse with the app's standard export
    styling — navy header (bold white text, centred), autofit column widths capped
    at ``width_cap``, ``str()``-coerced values for the keys in ``str_cols`` (e.g.
    timestamps/dates that must not be reformatted), optional frozen header row.

    ``columns`` = ``[(row_key, 'Header'), …]``; ``rows`` = list of dicts. Byte-for-
    byte identical to the per-export recipes it replaces (issues / tat / item-master).
    """
    import io

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    hf = Font(bold=True, color='FFFFFF')
    navy = PatternFill('solid', fgColor='1A237E')
    for c, (_k, h) in enumerate(columns, 1):
        cell = ws.cell(1, c, h)
        cell.font = hf
        cell.fill = navy
        cell.alignment = Alignment(horizontal='center')
    strset = set(str_cols)
    for r, row in enumerate(rows, 2):
        for c, (k, _h) in enumerate(columns, 1):
            v = row.get(k)
            ws.cell(r, c, str(v) if k in strset and v is not None else v)
    if freeze:
        ws.freeze_panes = 'A2'
    for col in ws.columns:
        L = col[0].column_letter
        w = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[L].width = min(w + 2, width_cap)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
