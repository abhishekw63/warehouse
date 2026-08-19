"""
online_b2b.services.full_validation  —  STANDALONE, self-contained.

Multi-marketplace D365 reconciliation: upload the full D365 **Sales Orders**
(headers) + **Sales Lines** exports (all marketplaces in one file), reconcile
against our records (``order_headers`` / ``order_lines_full``), and produce a
3-tier result — Marketplace Summary · Headers Reco · Lines Reco.

Deliberately ISOLATED (own module, own views file, own token dir, own template)
so it can be deleted in one go without touching the rest of the app. Read-only:
never writes to the business DB.

``validate(headers_path, lines_path, *, excel_out=None) -> dict``
  → {ok, summary, marketplaces:[...], headers:[...], lines:[...], excel_path}
"""
from __future__ import annotations

from collections import defaultdict

TOL_PCT, TOL_ABS = 0.005, 1.0   # value tolerance band: <=0.5% or <=Rs1


def _num(s):
    try:
        return float(str(s).replace(',', ''))
    except (ValueError, TypeError):
        return 0.0


def _vmatch(a, b) -> bool:
    return b == 0 or abs(a - b) <= TOL_ABS or abs(a - b) / max(abs(b), 1) <= TOL_PCT


def _read(path):
    import pandas as pd
    df = pd.read_excel(path, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def validate(headers_path, lines_path, *, excel_out=None) -> dict:
    """Reconcile a full-check D365 Headers + Lines pair against our records."""
    import pandas as pd
    try:
        h = _read(headers_path)
        lf = _read(lines_path)
    except Exception as e:  # noqa: BLE001
        return {'ok': False, 'error': f"Could not read the files: {type(e).__name__}: {e}"}

    ext = next((c for c in h.columns if c.lower() == 'external document no.'), None)
    if ext is None:
        return {'ok': False, 'error': "Headers file has no 'External Document No.' column — is it the D365 Sales Orders export?"}
    docno = next((c for c in lf.columns if c.lower() == 'document no.'), None)
    if docno is None:
        return {'ok': False, 'error': "Lines file has no 'Document No.' column — is it the D365 Sales Lines export?"}

    def col(df, *names):
        for n in names:
            for c in df.columns:
                if n.lower() == c.lower():
                    return c
        return None

    hq, hv = col(h, 'Total Quantity'), col(h, 'Total Amount Incl. GST')
    hpin, hst = col(h, 'Ship-to Postcode'), col(h, 'Ship-to Code')
    lq, lamt = col(lf, 'Quantity'), col(lf, 'Line Amount Excl. VAT')
    lgtin, lno = col(lf, 'GTIN'), col(lf, 'No.')
    ldesc = col(lf, 'Description')

    h = h[h[ext].notna() & (h[ext].astype(str).str.strip() != '')].copy()
    # PO numbers are case-invariant identifiers, but D365 exports them UPPERCASE
    # while some channels record them lowercase (e.g. Firstcry 'pin…' vs D365
    # 'PIN…'). Fold BOTH sides to upper so the match is case-insensitive — else a
    # recorded PO falsely reads EXTERNAL. Applied uniformly (D365 + our SQL + line
    # keys), so same-case pairs are unaffected; only case-variant pairs now match.
    h[ext] = h[ext].astype(str).str.strip().str.upper()
    d365h = h.groupby(ext).agg(
        qty=(hq, lambda s: sum(_num(x) for x in s)),
        val=(hv, lambda s: sum(_num(x) for x in s)),
        pin=(hpin, 'first'), st=(hst, 'first')).to_dict('index')

    lf[docno] = lf[docno].astype(str).str.strip().str.upper()   # case-insensitive PO match
    d365L = defaultdict(lambda: {'qty': 0.0, 'amt': 0.0, 'ean': '', 'desc': ''})
    for _, r in lf.iterrows():
        k = (r[docno], str(r[lno]).strip())
        g = d365L[k]
        g['qty'] += _num(r[lq]); g['amt'] += _num(r[lamt])
        g['ean'] = str(r.get(lgtin) or ''); g['desc'] = str(r.get(ldesc) or '')[:60]

    pos = tuple(d365h.keys())
    if not pos:
        return {'ok': False, 'error': "No POs (External Document No.) found in the headers file."}

    from .order_db import _conn
    with _conn() as (cur, d):
        ph = ','.join([d['ph']] * len(pos))
        # UPPER(po) both sides → the match is case-insensitive regardless of the
        # column's collation (pos are already upper from d365h).
        cur.execute(f"SELECT UPPER(po), marketplace_label, SUM(qty), SUM(order_value), MAX(location) "
                    f"FROM order_headers WHERE UPPER(po) IN ({ph}) GROUP BY UPPER(po), marketplace_label", pos)
        ourh = {str(r[0]): {'mp': r[1], 'qty': float(r[2] or 0), 'val': float(r[3] or 0),
                            'loc': r[4]} for r in cur.fetchall()}
        cur.execute(f"SELECT UPPER(po),item_no,ean,description,qty,our_cp,status,action,exception_label "
                    f"FROM order_lines_full WHERE UPPER(po) IN ({ph})", pos)
        ourL = {}
        for po, it, ean, desc, qty, cp, st, act, exc in cur.fetchall():
            ourL[(str(po), str(it))] = {'ean': str(ean or ''), 'desc': str(desc or '')[:60],
                                        'qty': int(qty or 0), 'cp': float(cp or 0), 'status': st,
                                        'action': (act or ''), 'exc': (exc or '')}

    hdr_rows, line_rows = [], []
    roll = defaultdict(lambda: {'pos': 0, 'qok': 0, 'vok': 0, 'excl': 0, 'qd': 0, 'qo': 0, 'vd': 0,
                                'vo': 0, 'ln_ok': 0, 'ln_excl': 0, 'ln_miss': 0, 'ln_extra': 0, 'ln_qty': 0})
    for po, D in d365h.items():
        o = ourh.get(po); mp = (o or {}).get('mp', 'UNKNOWN'); R = roll[mp]
        excl_qty = sum(v['qty'] for (p, it), v in ourL.items()
                       if p == po and str(v['action']).upper() == 'EXCLUDE')
        excl_val = sum(v['cp'] * v['qty'] for (p, it), v in ourL.items()
                       if p == po and str(v['action']).upper() == 'EXCLUDE')
        R['pos'] += 1; R['qd'] += D['qty']; R['vd'] += D['val']
        qok = bool(o) and abs(D['qty'] - o['qty']) < 0.5
        q_expl = bool(o) and abs((o['qty'] - excl_qty) - D['qty']) < 0.5
        vok = bool(o) and _vmatch(D['val'], o['val'])
        if o:
            R['qo'] += o['qty']; R['vo'] += o['val']
        R['excl'] += excl_qty
        if qok or q_expl:
            R['qok'] += 1
        if vok:
            R['vok'] += 1
        verd = 'OK' if qok else (f'OK — {int(excl_qty)} excluded' if q_expl else 'QTY MISMATCH')
        hdr_rows.append({'mp': mp, 'po': po, 'our_qty': int(o['qty']) if o else None,
                         'd365_qty': int(D['qty']), 'excluded': int(excl_qty),
                         'excl_val': round(excl_val, 2),
                         'final': int(o['qty'] - excl_qty) if o else None,
                         'qty_ok': bool(qok or q_expl),
                         'our_val': round(o['val'], 2) if o else None, 'd365_val': round(D['val'], 2),
                         'val_diff': round(D['val'] - (o['val'] if o else 0), 2), 'val_ok': bool(vok),
                         'ship_our': (o['loc'] if o else ''), 'ship_d365': D['st'], 'pin_d365': D['pin'],
                         'verdict': verd})
        our_keys = {(p, it) for (p, it) in ourL if p == po}
        d_keys = {(p, it) for (p, it) in d365L if p == po}
        for k in (our_keys | d_keys):
            it = k[1]; ov = ourL.get(k); dv = d365L.get(k)
            oq = ov['qty'] if ov else None
            dq = int(dv['qty']) if dv else None
            oval = round(ov['cp'] * ov['qty'], 2) if ov else None
            dval = round(dv['amt'], 2) if dv else None
            ean = (ov or dv or {}).get('ean', ''); desc = (ov or dv or {}).get('desc', '')
            if ov and dv:
                status, reason = ('OK', '') if oq == dq else ('QTY_MISMATCH', f'our {oq} vs D365 {dq}')
            elif ov and not dv:
                if str(ov['action']).upper() == 'EXCLUDE':
                    status, reason = 'EXCLUDED', f"CP-exclude ({ov['exc'] or 'MISMATCH'}) — intentional"
                else:
                    status, reason = 'MISSING_IN_D365', f"in our record, not in D365 (status {ov['status']})"
            else:
                status, reason = 'EXTRA_IN_D365', 'in D365, not in our record'
            val_ok = ('YES' if (ov and dv and _vmatch(dval or 0, oval or 0))
                      else ('n/a' if status in ('EXCLUDED', 'MISSING_IN_D365', 'EXTRA_IN_D365') else 'NO'))
            line_rows.append({'mp': mp, 'po': po, 'item': it, 'ean': ean, 'desc': desc,
                              'our_qty': oq, 'd365_qty': dq,
                              'qty_ok': ('YES' if (ov and dv and oq == dq) else ('n/a' if not (ov and dv) else 'NO')),
                              'our_val': oval, 'd365_val': dval, 'val_ok': val_ok,
                              'status': status, 'reason': reason})
            R['ln_ok'] += status == 'OK'
            R['ln_excl'] += status == 'EXCLUDED'
            R['ln_miss'] += status == 'MISSING_IN_D365'
            R['ln_extra'] += status == 'EXTRA_IN_D365'
            R['ln_qty'] += status == 'QTY_MISMATCH'

    markets = []
    tot = defaultdict(float)
    for mp, R in sorted(roll.items()):
        clean = (R['qok'] == R['pos'] and R['ln_miss'] == 0 and R['ln_extra'] == 0 and R['ln_qty'] == 0)
        markets.append({'mp': mp, 'pos': R['pos'], 'qok': R['qok'], 'vok': R['vok'],
                        'd365_qty': int(R['qd']), 'our_qty': int(R['qo']), 'excluded': int(R['excl']),
                        'd365_val': round(R['vd']), 'our_val': round(R['vo']), 'val_gap': round(R['vd'] - R['vo']),
                        'ln_ok': R['ln_ok'], 'ln_excl': R['ln_excl'], 'ln_miss': R['ln_miss'],
                        'ln_extra': R['ln_extra'], 'ln_qty': R['ln_qty'],
                        'verdict': 'CLEAN' if clean else 'REVIEW'})
        for k in ('pos', 'qd', 'qo', 'excl', 'vd', 'vo', 'ln_ok', 'ln_excl', 'ln_miss', 'ln_extra', 'ln_qty'):
            tot[k] += R[k]

    summary = {'pos': int(tot['pos']), 'lines': len(line_rows), 'line_ok': int(tot['ln_ok']),
               'excluded': int(tot['ln_excl']), 'missing': int(tot['ln_miss']), 'extra': int(tot['ln_extra']),
               'qty_mismatch': int(tot['ln_qty']), 'value_gap': round(tot['vd'] - tot['vo']),
               'marketplaces': len(markets),
               'clean': int(tot['ln_miss']) == 0 and int(tot['ln_extra']) == 0 and int(tot['ln_qty']) == 0}

    excel_path = None
    if excel_out:
        try:
            excel_path = _write_excel(excel_out, markets, hdr_rows, line_rows, summary)
        except Exception as e:  # noqa: BLE001
            summary['excel_error'] = f"{type(e).__name__}: {e}"

    return {'ok': True, 'summary': summary, 'marketplaces': markets,
            'headers': hdr_rows, 'lines': line_rows, 'excel_path': excel_path}


def _write_excel(path, markets, hdr_rows, line_rows, summary) -> str:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    HF = PatternFill('solid', fgColor='1D2433'); GR = PatternFill('solid', fgColor='E7F6EF')
    AM = PatternFill('solid', fgColor='FDF1E3'); RD = PatternFill('solid', fgColor='FDECEC')
    thin = Side(style='thin', color='D5D9E2'); BD = Border(thin, thin, thin, thin)

    def head(ws):
        for c in ws[1]:
            c.fill = HF; c.font = Font(bold=True, color='FFFFFF', size=10)
            c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center'); c.border = BD
        ws.freeze_panes = 'A2'

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Marketplace Summary'
    ws.append(['Marketplace', 'POs', 'Qty OK', 'Value OK', 'D365 Qty', 'Our Qty', 'Excluded Qty',
               'D365 Value', 'Our Value', 'Value Gap', 'Line OK', 'Excluded', 'MISSING', 'EXTRA',
               'Qty-mismatch', 'Verdict'])
    head(ws)
    for m in markets:
        ws.append([m['mp'], m['pos'], f"{m['qok']}/{m['pos']}", f"{m['vok']}/{m['pos']}", m['d365_qty'],
                   m['our_qty'], m['excluded'], m['d365_val'], m['our_val'], m['val_gap'], m['ln_ok'],
                   m['ln_excl'], m['ln_miss'], m['ln_extra'], m['ln_qty'], m['verdict']])
        rr = ws.max_row
        for c in ws[rr]:
            c.border = BD
        ws.cell(rr, 16).fill = (RD if (m['ln_miss'] or m['ln_extra'] or m['ln_qty'])
                                else (GR if m['verdict'] == 'CLEAN' else AM))
    ws.append(['TOTAL', summary['pos'], '', '', '', '', summary['excluded'], '', '', summary['value_gap'],
               summary['line_ok'], summary['excluded'], summary['missing'], summary['extra'],
               summary['qty_mismatch'], ''])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for i, w in enumerate([16, 5, 8, 8, 10, 10, 11, 14, 14, 12, 8, 9, 9, 8, 12, 10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws2 = wb.create_sheet('Headers Reco')
    ws2.append(['Marketplace', 'PO', 'Our Qty', 'D365 Qty', 'Excluded', 'Final', 'Qty OK', 'Our Value',
                'D365 Value', 'Value Diff', 'Value OK', 'Ship-to (our)', 'Ship-to (D365)', 'Pincode', 'Verdict'])
    head(ws2)
    for r in sorted(hdr_rows, key=lambda x: (x['mp'], x['po'])):
        ws2.append([r['mp'], r['po'], r['our_qty'], r['d365_qty'], r['excluded'], r['final'],
                    'YES' if r['qty_ok'] else 'NO', r['our_val'], r['d365_val'], r['val_diff'],
                    'YES' if r['val_ok'] else 'NO', r['ship_our'], r['ship_d365'], r['pin_d365'], r['verdict']])
        rr = ws2.max_row
        for c in ws2[rr]:
            c.border = BD
        ws2.cell(rr, 7).fill = GR if r['qty_ok'] else RD
        ws2.cell(rr, 11).fill = GR if r['val_ok'] else AM
    for i, w in enumerate([14, 16, 9, 9, 9, 9, 8, 13, 13, 11, 9, 22, 16, 12, 20], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws3 = wb.create_sheet('Lines Reco')
    ws3.append(['Marketplace', 'PO', 'Item No', 'EAN', 'Description', 'Our Qty', 'D365 Qty', 'Qty OK',
                'Our Val (ex-GST)', 'D365 Val (ex-VAT)', 'Value OK', 'Status', 'Reason'])
    head(ws3)
    cfill = {'OK': GR, 'EXCLUDED': AM, 'QTY_MISMATCH': RD, 'MISSING_IN_D365': RD, 'EXTRA_IN_D365': RD}
    for r in sorted(line_rows, key=lambda x: (x['mp'], x['po'], str(x['item']))):
        ws3.append([r['mp'], r['po'], r['item'], r['ean'], r['desc'], r['our_qty'], r['d365_qty'],
                    r['qty_ok'], r['our_val'], r['d365_val'], r['val_ok'], r['status'], r['reason']])
        rr = ws3.max_row
        for c in ws3[rr]:
            c.border = BD
        ws3.cell(rr, 12).fill = cfill.get(r['status'], GR)
    for i, w in enumerate([14, 16, 9, 16, 42, 8, 9, 8, 15, 16, 9, 16, 40], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(path)
    return str(path)
