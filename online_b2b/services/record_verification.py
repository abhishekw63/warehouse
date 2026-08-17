"""Record Verification — reconcile our recorded data (DB: recorded + excluded)
against the D365 Headers + Lines export, per PO, and flag any MISMATCH.

Purpose: catch cases where our recorded data disagrees with D365 (the ERP truth)
— qty, value, ship-to pincode, and line-level. It REUSES ``full_validation`` for
the heavy DB↔D365 reconciliation (which already nets out EXCLUDED lines), then:
  • resolves OUR ship-to pincode from ``ship_to_mapping`` (del-location already in
    the DB) and compares it to D365's Ship-to Postcode,
  • computes a per-PO delta (qty / value),
  • persists a checked-PO LOG (one row per PO, latest check wins) so you can see
    which POs are verified / mismatched / not-yet-checked and their deltas.

Standalone + removable — nothing else imports this. [[sales-validation-procedure]]
"""
from __future__ import annotations

import datetime as _dt
import re as _re

from . import full_validation as _fv
from .order_db import _conn

_LOG_TABLE = 'record_verification_log'

_MYSQL_DDL = f"""
CREATE TABLE IF NOT EXISTS {_LOG_TABLE} (
    id            BIGINT AUTO_INCREMENT PRIMARY KEY,
    po            VARCHAR(120) NOT NULL,
    marketplace   VARCHAR(80),
    status        VARCHAR(32),
    our_qty       INT, d365_qty INT, excluded_qty INT, qty_delta INT,
    our_val       DECIMAL(16,2), d365_val DECIMAL(16,2), val_delta DECIMAL(16,2),
    our_pin       VARCHAR(20), d365_pin VARCHAR(20), pin_ok TINYINT,
    mismatch_fields VARCHAR(255),
    checked_by    VARCHAR(150),
    checked_at    DATETIME,
    UNIQUE KEY uq_rv_po (po)
)"""
_SQLITE_DDL = f"""
CREATE TABLE IF NOT EXISTS {_LOG_TABLE} (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    po TEXT NOT NULL, marketplace TEXT, status TEXT,
    our_qty INTEGER, d365_qty INTEGER, excluded_qty INTEGER, qty_delta INTEGER,
    our_val REAL, d365_val REAL, val_delta REAL,
    our_pin TEXT, d365_pin TEXT, pin_ok INTEGER,
    mismatch_fields TEXT, checked_by TEXT, checked_at TEXT,
    UNIQUE (po)
)"""


def _nk(s):
    return _re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def _ensure_table():
    with _conn() as (cur, d):
        cur.execute(_SQLITE_DDL if d.get('sqlite') else _MYSQL_DDL)
        cur.connection.commit()


def _our_pin_map():
    """nk(del_location / ship_to / name / city) -> postcode, from ship_to_mapping —
    so we can resolve OUR recorded ship-to location to its pincode."""
    out = {}
    with _conn() as (cur, d):
        cur.execute("SELECT del_location, ship_to, name, city, postcode "
                    "FROM ship_to_mapping WHERE postcode IS NOT NULL AND postcode<>''")
        for dl, st, nm, ci, pc in cur.fetchall():
            pc = str(pc or '').strip()
            for kk in (dl, st, nm, ci):
                if kk:
                    out.setdefault(_nk(kk), pc)
    return out


_LOG_COLS = ['po', 'marketplace', 'status', 'our_qty', 'd365_qty', 'excluded_qty',
             'qty_delta', 'our_val', 'd365_val', 'val_delta', 'our_pin', 'd365_pin',
             'pin_ok', 'mismatch_fields', 'checked_by', 'checked_at']


def preview(headers_path, lines_path) -> dict:
    """PHASE 1 — compute the DB↔D365 reconciliation per PO. Writes NOTHING (review
    first, like the PO review page). :func:`confirm` persists it. Returns
    ``{ok, error, data}``; each header row gets status / delta / pincode / lines_ok.

    Ordered checks per PO — **SKU** (same SKU set on both sides) → **qty** (header +
    per-line, EXCLUDED netted) → **value** (our value net of excluded vs D365, within
    tolerance; only evaluated once SKU & qty align) → **pincode** (our ship-to vs
    D365). Status is **OK** only when all applicable checks pass; otherwise MISMATCH,
    and ``mismatch_fields`` names exactly which of SKU / qty / value / pincode failed."""
    base = _fv.validate(headers_path, lines_path, excel_out=None)
    if not base.get('ok'):
        return {'ok': False, 'error': base.get('error', 'Reconciliation failed.'), 'data': None}

    pin_map = _our_pin_map()
    rows = base.get('headers', [])          # per-PO rows from full_validation
    # Split line problems into SKU-set vs line-qty (EXCLUDED lines are intentional
    # and never count against a PO):
    #   • SKU mismatch  = a SKU is in ours-not-D365 or D365-not-ours (MISSING/EXTRA)
    #   • line-qty      = same SKU on both sides but the quantity differs
    sku_bad, lqty_bad = set(), set()
    for ln in base.get('lines', []):
        s = ln.get('status')
        if s in ('MISSING_IN_D365', 'EXTRA_IN_D365'):
            sku_bad.add(ln.get('po'))
        elif s == 'QTY_MISMATCH':
            lqty_bad.add(ln.get('po'))

    n_ok = n_mismatch = n_external = 0
    for r in rows:
        d365_pin = str(r.get('pin_d365') or '').strip()

        # ── EXTERNAL: in D365 but NOT in our records (GT Select / Testers etc. are
        #    uploaded outside this app) — recorded, but beyond our cross-check ──
        if r.get('our_qty') is None:
            r['our_pin'] = ''; r['pin_ok'] = True
            r['lines_ok'] = None; r['mismatch_fields'] = []; r['status'] = 'EXTERNAL'
            n_external += 1
            continue

        our_pin = pin_map.get(_nk(r.get('ship_our'))) or ''
        pin_ok = (not our_pin or not d365_pin) or (our_pin == d365_pin)
        sku_ok = r['po'] not in sku_bad                       # same SKU set on both sides
        qty_ok = bool(r.get('qty_ok')) and r['po'] not in lqty_bad   # header + per-line qty
        r['our_pin'] = our_pin; r['pin_ok'] = bool(pin_ok)
        r['sku_ok'] = sku_ok; r['lines_ok'] = sku_ok and qty_ok

        # VALUE against RAW, netted for excluded: our recorded value INCLUDES the
        # EXCLUDED lines (never pushed to D365), so subtract them before comparing —
        # else every PO with exclusions falsely flags a value mismatch. Compare the
        # netted "ours" to D365 within tolerance; show the netted value side by side.
        net_our_val = round((r.get('our_val') or 0) - (r.get('excl_val') or 0), 2)
        r['our_val'] = net_our_val
        r['val_diff'] = round((r.get('d365_val') or 0) - net_our_val, 2)
        val_ok = _fv._vmatch(r.get('d365_val') or 0, net_our_val)
        r['val_ok'] = bool(val_ok)

        # Ordered checks: SKU match → qty match → (only then) value → pincode.
        # Value is only meaningful once the SKU set + quantities line up, so it's
        # flagged ONLY when SKU & qty are already OK (no noisy value flag on a PO
        # whose SKUs/qty are wrong). OK requires all applicable checks to pass.
        mism = []
        if not sku_ok:
            mism.append('SKU')
        if not qty_ok:
            mism.append('qty')
        if sku_ok and qty_ok and not val_ok:
            mism.append('value')
        if not pin_ok:
            mism.append('pincode')
        r['mismatch_fields'] = mism
        verified = not mism
        r['status'] = 'OK' if verified else 'MISMATCH'
        if verified:
            n_ok += 1
        else:
            n_mismatch += 1

    _ord = {'MISMATCH': 0, 'EXTERNAL': 1, 'OK': 2}
    rows.sort(key=lambda x: _ord.get(x.get('status'), 3))
    base['verify_summary'] = {'checked': len(rows), 'ok': n_ok,
                              'mismatch': n_mismatch, 'external': n_external}
    return {'ok': True, 'error': None, 'data': base}


def confirm(rows, checked_by='', only_pos=None) -> dict:
    """PHASE 2 — persist the reviewed verification to the checked-PO log (upsert,
    one row per PO; latest confirm wins). ``rows`` = the previewed header rows.

    ``only_pos`` (optional): iterable of PO numbers the operator TICKED on the
    review table (review-page style). When given, only those POs are recorded —
    this is how EXTERNAL (GT Select / Testers) POs are "pushed" without a full
    cross-check: the operator ticks the ones to keep a record of. When omitted,
    every reviewed row is recorded (back-compat)."""
    _ensure_table()
    now = _dt.datetime.now()
    pick = None if only_pos is None else {str(p) for p in only_pos}
    payload = []
    for r in rows:
        if pick is not None and str(r.get('po')) not in pick:
            continue
        st = r.get('status')
        oq = r.get('final'); dq = r.get('d365_qty')
        qd = (int(dq) - int(oq)) if (oq is not None and dq is not None) else None
        fields = (','.join(r.get('mismatch_fields') or [])
                  or ('not-uploaded-by-us' if st == 'EXTERNAL' else ''))
        payload.append((
            r.get('po'), r.get('mp'), st, r.get('our_qty'), dq, r.get('excluded'), qd,
            r.get('our_val'), r.get('d365_val'), r.get('val_diff'),
            r.get('our_pin') or '', str(r.get('pin_d365') or ''),
            1 if r.get('pin_ok') else 0, fields, checked_by or 'system', now))
    if not payload:
        return {'ok': True, 'confirmed': 0}
    with _conn() as (cur, d):
        ph = d['ph']
        marks = ','.join([ph] * len(_LOG_COLS))
        popos = [p[0] for p in payload]
        pm = ','.join([ph] * len(popos))
        cur.execute(f"DELETE FROM {_LOG_TABLE} WHERE po IN ({pm})", tuple(popos))
        cur.executemany(
            f"INSERT INTO {_LOG_TABLE} ({', '.join(_LOG_COLS)}) VALUES ({marks})", payload)
        cur.connection.commit()
    return {'ok': True, 'confirmed': len(payload)}


def checked_log(limit=500, status='') -> list:
    """The persisted checked-PO log (latest check per PO), newest first — for the
    coverage view (which POs are verified / mismatched, and their deltas)."""
    _ensure_table()
    out = []
    with _conn() as (cur, d):
        ph = d['ph']
        where, args = '', []
        if status:
            where = f" WHERE status={ph}"; args = [status]
        cur.execute(
            f"SELECT po, marketplace, status, our_qty, d365_qty, excluded_qty, qty_delta, "
            f"our_val, d365_val, val_delta, our_pin, d365_pin, pin_ok, mismatch_fields, "
            f"checked_by, checked_at FROM {_LOG_TABLE}{where} "
            f"ORDER BY checked_at DESC LIMIT {int(limit)}", tuple(args))
        cols = ['po', 'marketplace', 'status', 'our_qty', 'd365_qty', 'excluded_qty',
                'qty_delta', 'our_val', 'd365_val', 'val_delta', 'our_pin', 'd365_pin',
                'pin_ok', 'mismatch_fields', 'checked_by', 'checked_at']
        for r in cur.fetchall():
            out.append(dict(zip(cols, r)))
    return out


def coverage() -> dict:
    """Headline coverage: how many distinct POs exist vs how many have been checked,
    and how many currently mismatch."""
    _ensure_table()
    out = {'total_pos': 0, 'checked': 0, 'ok': 0, 'mismatch': 0, 'external': 0, 'unchecked': 0}
    try:
        with _conn() as (cur, d):
            cur.execute("SELECT COUNT(DISTINCT po) FROM order_headers")
            out['total_pos'] = int((cur.fetchone() or [0])[0] or 0)
            cur.execute(f"SELECT status, COUNT(*) FROM {_LOG_TABLE} GROUP BY status")
            for st, c in cur.fetchall():
                out['checked'] += int(c)
                if st == 'OK':
                    out['ok'] += int(c)
                elif st == 'EXTERNAL':
                    out['external'] += int(c)
                else:
                    out['mismatch'] += int(c)
        # external POs aren't "ours", so don't count them against our coverage
        out['unchecked'] = max(0, out['total_pos'] - out['ok'] - out['mismatch'])
    except Exception:  # noqa: BLE001
        pass
    return out


def build_workbook(data, out_path) -> str:
    """Write the side-by-side comparison Excel for a previewed run and return the
    path. One row per PO: our (netted for excluded) vs D365 for qty / value /
    pincode, the three-check status, and which fields mismatched. Row-tinted by
    status (green OK · red MISMATCH · blue EXTERNAL)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparison'
    cols = [
        ('PO', 16), ('Marketplace', 14), ('Status', 12), ('Mismatch', 16),
        ('Qty ours', 10), ('Qty D365', 10), ('Excl.', 8), ('Qty Δ', 9),
        ('Value ours (net)', 15), ('Value D365', 14), ('Val Δ', 11),
        ('Our pin', 10), ('D365 pin', 10), ('Pincode OK', 10),
    ]
    head_fill = PatternFill('solid', fgColor='1F2A5A')
    head_font = Font(bold=True, color='FFFFFF', size=10)
    ok_fill = PatternFill('solid', fgColor='E7F6EF')
    bad_fill = PatternFill('solid', fgColor='FDECEC')
    ext_fill = PatternFill('solid', fgColor='EEF0FE')
    for c, (title, w) in enumerate(cols, 1):
        cell = ws.cell(1, c, title)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = w
    ws.freeze_panes = 'A2'

    for i, r in enumerate(data.get('headers', []), start=2):
        st = r.get('status')
        oq = r.get('our_qty'); dq = r.get('d365_qty')
        qd = (int(dq) - int(r.get('final'))) if (r.get('final') is not None and dq is not None) else None
        vals = [
            r.get('po'), r.get('mp') or '', st,
            ', '.join(r.get('mismatch_fields') or []) or ('not-uploaded-by-us' if st == 'EXTERNAL' else ''),
            oq if oq is not None else '', dq, r.get('excluded') or 0, qd if qd is not None else '',
            r.get('our_val') if r.get('our_val') is not None else '', r.get('d365_val'), r.get('val_diff'),
            r.get('our_pin') or '', str(r.get('pin_d365') or ''),
            ('' if st == 'EXTERNAL' else ('YES' if r.get('pin_ok') else 'NO')),
        ]
        fill = ext_fill if st == 'EXTERNAL' else (ok_fill if st == 'OK' else bad_fill)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.fill = fill
            if c >= 5:
                cell.alignment = Alignment(horizontal='right')

    wb.save(out_path)
    return str(out_path)
