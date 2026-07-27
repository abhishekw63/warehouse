"""
online_b2b.services.inventory_store
===================================

**Inventory — Fill-Rate cockpit** data layer. We UPLOAD a D365 *Bin Contents*
export per warehouse (a point-in-time stock snapshot) and record the sellable
on-hand per item, so the dashboard can compute fill-rate / OOS / tentative
billing against the recorded orders.

Design (mirrors the other web-owned stores — additive, never touches the frozen
engine or the business tables):

* **Snapshot, not live** — "Available Qty. to Take" is *deducted* as pickers
  generate Picking Lists (PL), so every upload is timestamped (``captured_at``)
  and the page always says *"stock as per <that time>"*. Latest snapshot per
  warehouse is flagged ``is_current``.
* **Bin filtering** — not every bin is sellable. A bin is classified by editable
  rules (:func:`classify_bin`): default **INCLUDE** prefixes ``ON-`` / ``OFF-`` /
  ``PS2-`` (online + offline pick faces + reserve storage), a **known-EXCLUDE**
  list of virtual bins (QC / RETURN / RTV / REJECT / EXPIRY / SHORTAGE / PACK …),
  and anything matching **neither → NEW** (flagged for classification, its qty
  held OUT of the sellable total — never silently swept in or dropped).
* Three warehouses (D365 Location codes): ``PICK`` = Ahmedabad, ``DS_BL_OFF1`` =
  Bangalore, ``NORTH WH-0`` = North.

Tables (created on demand, dual MySQL/SQLite DDL, in the ``orders`` DB):
``inventory_snapshot`` · ``inventory_stock`` · ``inventory_bin_audit`` ·
``inventory_bin_rule``.
"""
from __future__ import annotations

import datetime as _dt
import re

from .order_db import _conn

# ── warehouses (D365 Location code → friendly name / short tag) ──────────────
WAREHOUSES = [
    {'code': 'PICK', 'name': 'RENEE Warehouse Ahmedabad', 'short': 'AHD'},
    {'code': 'DS_BL_OFF1', 'name': 'DirectShelf Bangalore Offline', 'short': 'BLR'},
    {'code': 'NORTH WH-0', 'name': 'North Warehouse', 'short': 'North'},
]
WH_BY_CODE = {w['code']: w for w in WAREHOUSES}


def wh_name(code: str) -> str:
    w = WH_BY_CODE.get(str(code or '').strip())
    return w['name'] if w else str(code or '')


def wh_short(code: str) -> str:
    w = WH_BY_CODE.get(str(code or '').strip())
    return w['short'] if w else str(code or '')


# ── order-warehouse → inventory-warehouse resolution ────────────────────────
# order_headers.warehouse holds a MIX of display names ('AHD'/'BLR'), D365 codes
# ('PICK'/'DS_BL_OFF1'), or the occasional stray ship-to code. Normalize any of
# them to the canonical inventory Location code so demand lines up with stock.
# (Engine map is WAREHOUSE_CODES = {'AHD':'PICK', 'BLR':'DS_BL_OFF1'}.)
_WH_ALIASES = {
    'AHD': 'PICK', 'PICK': 'PICK',
    'BLR': 'DS_BL_OFF1', 'DS_BL_OFF1': 'DS_BL_OFF1', 'DS': 'DS_BL_OFF1',
    'NORTH': 'NORTH WH-0', 'NORTH WH': 'NORTH WH-0', 'NORTH WH-0': 'NORTH WH-0',
}
DEFAULT_WH = 'PICK'

# Marketplace → forced fulfilment warehouse. Overrides the stored per-order WH
# for channels whose stock actually ships from a different warehouse than the one
# recorded. BlinkMP is fulfilled from Bangalore (DS_BL_OFF1), NOT Ahmedabad (PICK),
# even though its orders were recorded under 'AHD'. Keyed by lower-cased
# marketplace OR marketplace_label. Edit here to add channels.
MP_WAREHOUSE_OVERRIDE = {
    'blinkmp': 'DS_BL_OFF1',
}


def wh_normalize(raw) -> str:
    """Any stored warehouse token → canonical inventory Location code."""
    s = str(raw or '').strip().upper()
    if s in _WH_ALIASES:
        return _WH_ALIASES[s]
    if str(raw or '').strip() in WH_BY_CODE:      # already a known code
        return str(raw).strip()
    return DEFAULT_WH


def resolve_order_wh(warehouse_raw, marketplace='', label='') -> str:
    """Fulfilment warehouse (inventory code) for an order line: a marketplace
    override wins (e.g. BlinkMP → DS_BL_OFF1), else the normalized stored WH."""
    for key in (str(marketplace or '').strip().lower(),
                str(label or '').strip().lower()):
        if key and key in MP_WAREHOUSE_OVERRIDE:
            return MP_WAREHOUSE_OVERRIDE[key]
    return wh_normalize(warehouse_raw)


# ── default bin-classification rules (seeded once; editable on the page) ─────
# match_type: 'prefix' = bin code starts with pattern (case-insensitive);
#             'segment' = bin code's FIRST token (split on space / _ / - / .)
#                         equals pattern; 'exact' = whole bin code equals pattern.
_DEFAULT_RULES = (
    # INCLUDE — the real sellable pick faces + reserve storage
    [('ON-', 'prefix', 'include'), ('OFF-', 'prefix', 'include'),
     ('PS2-', 'prefix', 'include')]
    # EXCLUDE — known virtual / non-saleable bins (first-token match)
    + [(t, 'segment', 'exclude') for t in (
        'QC', 'CONSUMABLE', 'CC', 'APRIL', 'REWORK', 'RET', 'REJECT', 'ORDER',
        'EXP', 'NEAR', 'OLD', 'RETURN', 'RT', 'RTZ', 'R', 'IB', 'V', 'BAD',
        'DESTROY', 'DEC', 'PALLET', 'PACK', 'D2', 'SHIPMENT', 'RECEIPT', 'HDR')]
)

_MYSQL_DDL = [
    """
    CREATE TABLE IF NOT EXISTS inventory_snapshot (
        snapshot_id     BIGINT AUTO_INCREMENT PRIMARY KEY,
        warehouse       VARCHAR(40),
        warehouse_name  VARCHAR(120),
        captured_at     DATETIME,
        source_file     VARCHAR(255),
        uploaded_by     VARCHAR(80),
        total_lines     INT DEFAULT 0,
        included_lines  INT DEFAULT 0,
        excluded_lines  INT DEFAULT 0,
        new_lines       INT DEFAULT 0,
        item_count      INT DEFAULT 0,
        included_qty    DECIMAL(16,2) DEFAULT 0,
        excluded_qty    DECIMAL(16,2) DEFAULT 0,
        new_qty         DECIMAL(16,2) DEFAULT 0,
        is_current      TINYINT DEFAULT 1,
        created_at      DATETIME,
        INDEX idx_invsnap_wh (warehouse),
        INDEX idx_invsnap_cur (warehouse, is_current)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS inventory_stock (
        id            BIGINT AUTO_INCREMENT PRIMARY KEY,
        snapshot_id   BIGINT,
        warehouse     VARCHAR(40),
        item_no       VARCHAR(60),
        ean           VARCHAR(40),
        description   VARCHAR(255),
        uom           VARCHAR(20),
        available_qty DECIMAL(16,2) DEFAULT 0,
        INDEX idx_invstk_snap (snapshot_id),
        INDEX idx_invstk_wh_item (warehouse, item_no)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS inventory_bin_audit (
        id            BIGINT AUTO_INCREMENT PRIMARY KEY,
        snapshot_id   BIGINT,
        warehouse     VARCHAR(40),
        bin_code      VARCHAR(120),
        zone_code     VARCHAR(60),
        decision      VARCHAR(10),
        n_lines       INT DEFAULT 0,
        qty           DECIMAL(16,2) DEFAULT 0,
        INDEX idx_invbin_snap (snapshot_id)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS inventory_bin_rule (
        id          BIGINT AUTO_INCREMENT PRIMARY KEY,
        pattern     VARCHAR(120),
        match_type  VARCHAR(10) DEFAULT 'prefix',
        decision    VARCHAR(10) DEFAULT 'exclude',
        warehouse   VARCHAR(40) DEFAULT '',
        note        VARCHAR(255),
        updated_by  VARCHAR(80),
        updated_at  DATETIME,
        UNIQUE KEY uq_invrule (pattern, match_type, warehouse)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
]
_SQLITE_DDL = [
    """CREATE TABLE IF NOT EXISTS inventory_snapshot (
        snapshot_id INTEGER PRIMARY KEY AUTOINCREMENT, warehouse TEXT,
        warehouse_name TEXT, captured_at TEXT, source_file TEXT, uploaded_by TEXT,
        total_lines INTEGER DEFAULT 0, included_lines INTEGER DEFAULT 0,
        excluded_lines INTEGER DEFAULT 0, new_lines INTEGER DEFAULT 0,
        item_count INTEGER DEFAULT 0, included_qty REAL DEFAULT 0,
        excluded_qty REAL DEFAULT 0, new_qty REAL DEFAULT 0,
        is_current INTEGER DEFAULT 1, created_at TEXT)""",
    """CREATE TABLE IF NOT EXISTS inventory_stock (
        id INTEGER PRIMARY KEY AUTOINCREMENT, snapshot_id INTEGER, warehouse TEXT,
        item_no TEXT, ean TEXT, description TEXT, uom TEXT,
        available_qty REAL DEFAULT 0)""",
    """CREATE TABLE IF NOT EXISTS inventory_bin_audit (
        id INTEGER PRIMARY KEY AUTOINCREMENT, snapshot_id INTEGER, warehouse TEXT,
        bin_code TEXT, zone_code TEXT, decision TEXT, n_lines INTEGER DEFAULT 0,
        qty REAL DEFAULT 0)""",
    """CREATE TABLE IF NOT EXISTS inventory_bin_rule (
        id INTEGER PRIMARY KEY AUTOINCREMENT, pattern TEXT, match_type TEXT DEFAULT 'prefix',
        decision TEXT DEFAULT 'exclude', warehouse TEXT DEFAULT '', note TEXT,
        updated_by TEXT, updated_at TEXT,
        UNIQUE (pattern, match_type, warehouse))""",
]


def ensure_tables() -> None:
    with _conn() as (cur, d):
        for ddl in (_MYSQL_DDL if d['kind'] == 'mysql' else _SQLITE_DDL):
            cur.execute(ddl)
        # ── migration: warehouse-scope column on pre-existing installs ──
        # (bin rules can be scoped to ONE warehouse, e.g. QC counts as sellable in
        # BLR but not in AHD; blank warehouse = applies to all). Idempotent.
        try:
            if d['kind'] == 'mysql':
                cur.execute(
                    "SELECT COUNT(*) FROM information_schema.columns "
                    "WHERE table_schema=DATABASE() AND table_name='inventory_bin_rule' "
                    "AND column_name='warehouse'")
                if not cur.fetchone()[0]:
                    cur.execute("ALTER TABLE inventory_bin_rule "
                                "ADD COLUMN warehouse VARCHAR(40) DEFAULT ''")
                    for stmt in ("ALTER TABLE inventory_bin_rule DROP INDEX uq_invrule",
                                 "ALTER TABLE inventory_bin_rule ADD UNIQUE KEY "
                                 "uq_invrule (pattern, match_type, warehouse)"):
                        try:
                            cur.execute(stmt)
                        except Exception:  # noqa: BLE001 — index may already be shaped
                            pass
            else:
                cur.execute("PRAGMA table_info(inventory_bin_rule)")
                if 'warehouse' not in [r[1] for r in cur.fetchall()]:
                    cur.execute("ALTER TABLE inventory_bin_rule "
                                "ADD COLUMN warehouse TEXT DEFAULT ''")
        except Exception:  # noqa: BLE001 — never let a migration break reads
            pass
        # seed default rules once (idempotent — UNIQUE(pattern, match_type))
        ph = d['ph']
        now = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        verb = 'INSERT IGNORE INTO' if d['kind'] == 'mysql' else 'INSERT OR IGNORE INTO'
        for pat, mt, dec in _DEFAULT_RULES:
            cur.execute(
                f"{verb} inventory_bin_rule "
                f"(pattern, match_type, decision, note, updated_by, updated_at) "
                f"VALUES ({ph},{ph},{ph},{ph},{ph},{ph})",
                (pat, mt, dec, 'seed default', 'system', now))
        cur.connection.commit()


# ── bin classification ──────────────────────────────────────────────────────
def _first_segment(bin_code: str) -> str:
    m = re.match(r'^\s*([A-Za-z0-9]+)', str(bin_code or ''))
    return m.group(1).upper() if m else ''


def load_rules() -> list[dict]:
    """All bin rules (editable list). Warehouse-specific first, then global; within
    each scope include-first then exclude — the exact order :func:`classify_bin`
    walks (first match wins)."""
    ensure_tables()
    out = []
    with _conn() as (cur, d):
        cur.execute("SELECT id, pattern, match_type, decision, warehouse, note, "
                    "updated_by, updated_at FROM inventory_bin_rule "
                    "ORDER BY (warehouse<>'') DESC, warehouse, decision DESC, pattern")
        cols = ['id', 'pattern', 'match_type', 'decision', 'warehouse', 'note',
                'updated_by', 'updated_at']
        out = [dict(zip(cols, r)) for r in cur.fetchall()]
    return out


def _ordered(rules: list[dict]):
    """Ordered [(pattern_up, match_type, decision)] — includes before excludes so a
    bin that could match both counts as sellable (preserves the include-wins rule)."""
    inc = [(str(r['pattern']).upper(), r['match_type'], 'include') for r in rules
           if r['decision'] == 'include']
    exc = [(str(r['pattern']).upper(), r['match_type'], 'exclude') for r in rules
           if r['decision'] == 'exclude']
    return inc + exc


def _compile(rules: list[dict], warehouse: str = ''):
    """Compile rules into the ordered match-list for ONE warehouse: that
    warehouse's own rules FIRST (they win — e.g. QC=include in BLR beats the
    global QC=exclude), then the global (blank-warehouse) rules. ``warehouse=''``
    → global rules only."""
    wh = str(warehouse or '').strip()
    wh_rules = [r for r in rules if str(r.get('warehouse') or '').strip() == wh and wh]
    g_rules = [r for r in rules if not str(r.get('warehouse') or '').strip()]
    return _ordered(wh_rules) + _ordered(g_rules)


def _match(bin_up: str, seg: str, pat: str, mt: str) -> bool:
    if mt == 'prefix':
        return bin_up.startswith(pat)
    if mt == 'segment':
        return seg == pat
    return bin_up == pat            # exact


def classify_bin(bin_code: str, compiled=None) -> str:
    """→ 'include' | 'exclude' | 'new'. Walks the compiled rule list in priority
    order (warehouse-specific → global, include → exclude); FIRST match wins. A bin
    matching nothing is 'new' (unknown → flagged, held out of the sellable total).
    ``compiled`` is the list from :func:`_compile` (defaults to global rules)."""
    bin_up = str(bin_code or '').strip().upper()
    if not bin_up:
        return 'new'
    seg = _first_segment(bin_up)
    if compiled is None:
        compiled = _compile(load_rules())
    for pat, mt, decision in compiled:
        if _match(bin_up, seg, pat, mt):
            return decision
    return 'new'


# ── rule CRUD (editable include/exclude list) ───────────────────────────────
def add_rule(pattern, match_type, decision, note='', user='', warehouse='') -> dict:
    pattern = str(pattern or '').strip()
    match_type = match_type if match_type in ('prefix', 'segment', 'exact') else 'prefix'
    decision = 'include' if str(decision).lower() == 'include' else 'exclude'
    warehouse = str(warehouse or '').strip()   # '' = applies to every warehouse
    if not pattern:
        return {'ok': False, 'error': 'Pattern is required.'}
    ensure_tables()
    now = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _conn() as (cur, d):
        ph = d['ph']
        # backend-agnostic upsert on (pattern, match_type, warehouse) — a bin rule
        # can now exist per-warehouse, so we key on all three.
        cur.execute(
            f"DELETE FROM inventory_bin_rule WHERE pattern={ph} AND match_type={ph} "
            f"AND warehouse={ph}", (pattern.upper(), match_type, warehouse))
        cur.execute(
            f"INSERT INTO inventory_bin_rule "
            f"(pattern, match_type, decision, warehouse, note, updated_by, updated_at) "
            f"VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph})",
            (pattern.upper(), match_type, decision, warehouse, note, user, now))
        cur.connection.commit()
    return {'ok': True}


def delete_rule(rule_id) -> dict:
    ensure_tables()
    with _conn() as (cur, d):
        ph = d['ph']
        cur.execute(f"DELETE FROM inventory_bin_rule WHERE id={ph}", (rule_id,))
        cur.connection.commit()
    return {'ok': True}


# ── bin-content parse (D365 export) ─────────────────────────────────────────
# canonical field → accepted header spellings (lower-cased, stripped)
_HDR_ALIASES = {
    'warehouse': ('location filter', 'location', 'location code'),
    'bin_code': ('bin code', 'bin'),
    'zone_code': ('zone code', 'zone'),
    'ean': ('gtin', 'ean', 'barcode'),
    'item_no': ('item no.', 'item no', 'item number', 'item'),
    'uom': ('unit of measure code', 'uom', 'unit of measure'),
    'available_qty': ('available qty. to take', 'available qty to take',
                      'available quantity to take'),
    'description': ('itemdescription', 'item description', 'description'),
}


def _norm(s) -> str:
    return re.sub(r'\s+', ' ', str(s or '').strip()).lower()


def _map_headers(header_row) -> dict:
    """Header cell → column index, resolving to canonical field names. Returns
    {canonical: col_index}. Robust to column re-ordering."""
    idx = {}
    for j, cell in enumerate(header_row):
        n = _norm(cell)
        for canon, aliases in _HDR_ALIASES.items():
            if canon in idx:
                continue
            if n in aliases:
                idx[canon] = j
                break
    return idx


def parse_bin_content(path) -> dict:
    """Read a Bin Contents .xlsx → per-warehouse aggregation. Returns
    ``{ok, error, warehouses:{code:{stock, bins, totals}}, headers_ok}``.

    * ``stock``  = {item_no: {item_no, ean, description, uom, qty}} (INCLUDE bins)
    * ``bins``   = {bin_code: {bin_code, zone, decision, lines, qty}}
    * ``totals`` = counts + qty split by decision, item_count
    Read-only; never writes. Never raises (returns ok=False)."""
    import openpyxl
    out = {'ok': False, 'error': '', 'warehouses': {}, 'headers_ok': False,
           'file_rows': 0}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            out['error'] = 'Empty sheet.'
            return out
        cmap = _map_headers(header)
        need = ('warehouse', 'item_no', 'available_qty', 'bin_code')
        missing = [k for k in need if k not in cmap]
        if missing:
            out['error'] = ('Could not find column(s): '
                            + ', '.join(missing) + '. Is this a Bin Contents export?')
            return out
        out['headers_ok'] = True
        # Compile per-warehouse (a rule may be scoped to one WH). Cache by code so
        # each warehouse's ordered rule-list is built once.
        _all_rules = load_rules()
        _compiled_cache: dict = {}

        def _compiled_for(code):
            code = str(code or '').strip()
            if code not in _compiled_cache:
                _compiled_cache[code] = _compile(_all_rules, code)
            return _compiled_cache[code]

        def col(r, key):
            j = cmap.get(key)
            return r[j] if (j is not None and j < len(r)) else None

        whs = out['warehouses']

        def _wh(code):
            code = str(code or '').strip() or '(blank)'
            w = whs.get(code)
            if w is None:
                w = whs[code] = {
                    'stock': {}, 'bins': {},
                    'totals': {'total_lines': 0, 'included_lines': 0,
                               'excluded_lines': 0, 'new_lines': 0,
                               'included_qty': 0.0, 'excluded_qty': 0.0,
                               'new_qty': 0.0}}
            return w

        for r in rows:
            item = col(r, 'item_no')
            if item in (None, ''):
                continue
            out['file_rows'] += 1
            wh_code = str(col(r, 'warehouse') or '').strip()
            wh = _wh(wh_code)
            t = wh['totals']
            bin_code = str(col(r, 'bin_code') or '').strip()
            zone = str(col(r, 'zone_code') or '').strip()
            try:
                qty = float(col(r, 'available_qty') or 0)
            except (TypeError, ValueError):
                qty = 0.0
            decision = classify_bin(bin_code, _compiled_for(wh_code))
            word = {'include': 'included', 'exclude': 'excluded',
                    'new': 'new'}[decision]
            t['total_lines'] += 1
            t[f'{word}_lines'] += 1
            t[f'{word}_qty'] += qty
            # bin audit
            b = wh['bins'].get(bin_code)
            if b is None:
                b = wh['bins'][bin_code] = {'bin_code': bin_code, 'zone': zone,
                                            'decision': decision, 'lines': 0,
                                            'qty': 0.0}
            b['lines'] += 1
            b['qty'] += qty
            # sellable stock — INCLUDE bins only
            if decision == 'include':
                key = str(item).strip()
                s = wh['stock'].get(key)
                if s is None:
                    s = wh['stock'][key] = {
                        'item_no': key, 'ean': str(col(r, 'ean') or '').strip(),
                        'description': str(col(r, 'description') or '').strip(),
                        'uom': str(col(r, 'uom') or '').strip(), 'qty': 0.0}
                s['qty'] += qty
                if not s['ean']:
                    s['ean'] = str(col(r, 'ean') or '').strip()
                if not s['description']:
                    s['description'] = str(col(r, 'description') or '').strip()
        for w in whs.values():
            w['totals']['item_count'] = len(w['stock'])
        wb.close()
        out['ok'] = True
    except Exception as e:  # noqa: BLE001
        out['error'] = f'{type(e).__name__}: {e}'
    return out


# ── save a parsed snapshot (per warehouse) ──────────────────────────────────
def save_snapshot(warehouse, parsed_wh, source_file='', user='',
                  captured_at=None) -> dict:
    """Persist ONE warehouse's parsed block as the new current snapshot; the
    previous current snapshot for that WH is demoted (kept for history)."""
    ensure_tables()
    now = _dt.datetime.now()
    cap = captured_at or now
    if isinstance(cap, _dt.datetime):
        cap = cap.strftime('%Y-%m-%d %H:%M:%S')
    t = parsed_wh['totals']
    with _conn() as (cur, d):
        ph = d['ph']
        cur.execute(f"UPDATE inventory_snapshot SET is_current=0 "
                    f"WHERE warehouse={ph} AND is_current=1", (warehouse,))
        cur.execute(
            f"INSERT INTO inventory_snapshot (warehouse, warehouse_name, captured_at,"
            f" source_file, uploaded_by, total_lines, included_lines, excluded_lines,"
            f" new_lines, item_count, included_qty, excluded_qty, new_qty, is_current,"
            f" created_at) VALUES ({','.join([ph]*15)})",
            (warehouse, wh_name(warehouse), cap, source_file, user,
             t['total_lines'], t['included_lines'], t['excluded_lines'],
             t['new_lines'], t.get('item_count', len(parsed_wh['stock'])),
             round(t['included_qty'], 2), round(t['excluded_qty'], 2),
             round(t['new_qty'], 2), 1, now.strftime('%Y-%m-%d %H:%M:%S')))
        snap_id = cur.lastrowid
        stock_rows = [
            (snap_id, warehouse, str(s['item_no'])[:60], str(s['ean'])[:40],
             s['description'][:255], str(s['uom'])[:20], round(s['qty'], 2))
            for s in parsed_wh['stock'].values()]
        if stock_rows:
            cur.executemany(
                f"INSERT INTO inventory_stock (snapshot_id, warehouse, item_no, ean,"
                f" description, uom, available_qty) VALUES ({','.join([ph]*7)})",
                stock_rows)
        bin_rows = [
            (snap_id, warehouse, b['bin_code'][:120], b['zone'][:60],
             b['decision'], b['lines'], round(b['qty'], 2))
            for b in parsed_wh['bins'].values()]
        if bin_rows:
            cur.executemany(
                f"INSERT INTO inventory_bin_audit (snapshot_id, warehouse, bin_code,"
                f" zone_code, decision, n_lines, qty) VALUES ({','.join([ph]*7)})",
                bin_rows)
        cur.connection.commit()
    return {'ok': True, 'snapshot_id': snap_id, 'warehouse': warehouse,
            'item_count': len(stock_rows)}


# ── reads ───────────────────────────────────────────────────────────────────
def current_snapshots() -> dict:
    """{warehouse_code: snapshot dict} for the latest snapshot of each WH."""
    ensure_tables()
    out = {}
    with _conn() as (cur, d):
        cur.execute(
            "SELECT snapshot_id, warehouse, warehouse_name, captured_at, source_file,"
            " uploaded_by, total_lines, included_lines, excluded_lines, new_lines,"
            " item_count, included_qty, excluded_qty, new_qty, created_at "
            "FROM inventory_snapshot WHERE is_current=1")
        cols = ['snapshot_id', 'warehouse', 'warehouse_name', 'captured_at',
                'source_file', 'uploaded_by', 'total_lines', 'included_lines',
                'excluded_lines', 'new_lines', 'item_count', 'included_qty',
                'excluded_qty', 'new_qty', 'created_at']
        for r in cur.fetchall():
            row = dict(zip(cols, r))
            out[row['warehouse']] = row
    return out


def current_stock_map(warehouse='') -> dict:
    """{item_no: total available_qty} from CURRENT snapshots. If ``warehouse`` is
    given → that WH only; else summed across all warehouses' current snapshots."""
    ensure_tables()
    out: dict = {}
    with _conn() as (cur, d):
        ph = d['ph']
        sql = ("SELECT s.item_no, s.available_qty FROM inventory_stock s "
               "JOIN inventory_snapshot p ON p.snapshot_id=s.snapshot_id "
               "WHERE p.is_current=1")
        params: list = []
        if warehouse:
            sql += f" AND s.warehouse={ph}"
            params.append(warehouse)
        cur.execute(sql, tuple(params))
        for item_no, qty in cur.fetchall():
            k = str(item_no)
            out[k] = out.get(k, 0.0) + float(qty or 0)
    return out


def stock_by_item() -> list[dict]:
    """Per-item available stock across CURRENT snapshots — one row per item with a
    per-warehouse qty map (+ total), for the Inventory stock view. Read-only.
    Returns [{item_no, ean, description, uom, wh:{code:qty}, total}], qty-desc."""
    ensure_tables()
    rows: dict = {}
    with _conn() as (cur, d):
        cur.execute(
            "SELECT s.item_no, s.warehouse, s.ean, s.description, s.uom, s.available_qty "
            "FROM inventory_stock s JOIN inventory_snapshot p ON p.snapshot_id=s.snapshot_id "
            "WHERE p.is_current=1")
        for item_no, wh, ean, desc, uom, qty in cur.fetchall():
            k = str(item_no or '').strip()
            if not k:
                continue
            r = rows.get(k)
            if r is None:
                r = rows[k] = {'item_no': k, 'ean': str(ean or ''),
                               'description': str(desc or ''), 'uom': str(uom or ''),
                               'wh': {}, 'total': 0.0}
            code = str(wh or '')
            q = float(qty or 0)
            r['wh'][code] = r['wh'].get(code, 0.0) + q
            r['total'] += q
            if not r['ean'] and ean:
                r['ean'] = str(ean)
            if not r['description'] and desc:
                r['description'] = str(desc)
    out = list(rows.values())
    out.sort(key=lambda x: -x['total'])
    return out


def bin_audit(snapshot_id) -> list[dict]:
    """Per-bin audit rows for a snapshot (for the bin-classification view)."""
    ensure_tables()
    with _conn() as (cur, d):
        ph = d['ph']
        cur.execute(
            f"SELECT bin_code, zone_code, decision, n_lines, qty "
            f"FROM inventory_bin_audit WHERE snapshot_id={ph} "
            f"ORDER BY (decision='new') DESC, qty DESC", (snapshot_id,))
        cols = ['bin_code', 'zone_code', 'decision', 'lines', 'qty']
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def new_bins(snapshot_id) -> list[dict]:
    """Unknown/new bins in a snapshot (need classification) — the alert list."""
    return [b for b in bin_audit(snapshot_id) if b['decision'] == 'new']
