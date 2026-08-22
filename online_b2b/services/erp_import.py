"""
online_b2b.services.erp_import
==============================

Bulk import of the **ERP "Sales Orders" header export** (Business Central /
D365) so manually-created orders — the ones that can't go through the package
engine — still reflect on the dashboard.

Web-owned (the engine is the frozen backup): we parse the xlsx, map each Sales
Order header to an ``order_headers`` row tagged ``segment='Offline'`` with the
channel derived from the SO No prefix (``SO/CSD/06/…`` → ``CSD``), and insert
under a new ``runs`` row (``mode='MANUAL'``). Dedup is on the SO No.
Header-only — no line items (these orders have none in this export).
"""

from __future__ import annotations

import datetime as _dt
import os
import re

from .order_db import _conn, _conn_tx

# ERP header name (lower, stripped) → our field. Matched fuzzily by substring.
_FIELD_BY_HEADER = [
    ('no.', 'so_no'),
    ('external document', 'external_doc'),
    ('ship-to name', 'location'),
    ('sell-to customer name', 'customer'),
    ('gen. bus. posting group', 'posting_group'),
    ('total amount incl', 'order_value'),
    ('total quantity', 'qty'),
    ('location code', 'warehouse'),
    ('document date', 'po_date'),
    ('invoice to date', 'exp_date'),
    ('status', 'status'),
]

# order_headers insert columns (raw) — external_doc added by ensure_columns().
_HDR_COLS = [
    'run_id', 'run_ts', 'mode', 'segment', 'marketplace', 'marketplace_label',
    'po', 'location', 'warehouse', 'po_date', 'exp_date', 'order_type', 'items',
    'qty', 'order_value', 'output_file', 'external_doc',
]

SEGMENT_OFFLINE = 'Offline'


def _channel_from_so(so_no: str) -> str:
    """'SO/CSD/06/22626' → 'CSD'. Falls back to '' if no clear prefix."""
    parts = [p for p in re.split(r'[\\/]', str(so_no or '')) if p]
    # parts like ['SO', 'CSD', '06', '22626'] → channel is the 2nd segment
    if len(parts) >= 2 and parts[0].upper() in ('SO', 'TO'):
        return parts[1].strip().upper()
    return ''


def _num(v):
    if v is None or v == '':
        return None
    try:
        return float(str(v).replace(',', ''))
    except (TypeError, ValueError):
        return None


def _date(v):
    if v is None or v == '':
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v)[:10]
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── Parse ────────────────────────────────────────────────────────────────

def parse(filepath: str) -> dict:
    """Read the ERP xlsx → list of mapped header rows. Returns
    {ok, rows, error, headers_found}."""
    out: dict = {'ok': False, 'rows': [], 'headers_found': []}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return {'ok': False, 'error': 'Empty sheet.'}

        header = [str(c).strip().lower() if c is not None else ''
                  for c in all_rows[0]]
        out['headers_found'] = [str(c) for c in all_rows[0] if c]

        col = {}
        for idx, h in enumerate(header):
            for needle, field in _FIELD_BY_HEADER:
                if needle in h and field not in col:
                    col[field] = idx
                    break

        if 'so_no' not in col:
            return {'ok': False, 'error':
                    "Couldn't find the 'No.' (Sales Order number) column — "
                    "is this the ERP 'Sales Orders' export?"}

        def g(r, key):
            i = col.get(key)
            return r[i] if i is not None and i < len(r) else None

        rows: list[dict] = []
        for r in all_rows[1:]:
            so_no = g(r, 'so_no')
            if not so_no or not str(so_no).strip():
                continue
            channel = _channel_from_so(so_no)
            rows.append({
                'so_no': str(so_no).strip(),
                'external_doc': str(g(r, 'external_doc') or '').strip(),
                'channel': channel or 'ERP',
                'location': str(g(r, 'location') or g(r, 'customer') or '').strip(),
                'warehouse': str(g(r, 'warehouse') or '').strip(),
                'po_date': _date(g(r, 'po_date')),
                'exp_date': _date(g(r, 'exp_date')),
                'qty': int(_num(g(r, 'qty')) or 0),
                'order_value': _num(g(r, 'order_value')) or 0.0,
                'status': str(g(r, 'status') or '').strip(),
                'order_type': 'TO' if str(so_no).upper().startswith('TO/') else 'SO',
            })
        out['rows'] = rows
        out['ok'] = True
    except Exception as e:  # noqa: BLE001
        out['error'] = f"{type(e).__name__}: {e}"
    return out


# ── Schema + dedup helpers (web-owned, additive) ──────────────────────────

def ensure_columns() -> None:
    """Add the nullable external_doc column to order_headers if absent
    (additive; the engine's inserts omit it → NULL, so it's unaffected)."""
    with _conn() as (cur, d):
        if d['kind'] == 'mysql':
            cur.execute(
                "SELECT COUNT(*) FROM information_schema.columns WHERE "
                "table_schema=DATABASE() AND table_name='order_headers' "
                "AND column_name='external_doc'")
            if cur.fetchone()[0] == 0:
                cur.execute("ALTER TABLE order_headers "
                            "ADD COLUMN external_doc VARCHAR(100) NULL")
        else:
            cols = [r[1] for r in cur.execute(
                "PRAGMA table_info(order_headers)").fetchall()]
            if 'external_doc' not in cols:
                cur.execute("ALTER TABLE order_headers "
                            "ADD COLUMN external_doc TEXT")
        cur.connection.commit()


def existing_pos(pos: list[str]) -> set:
    """Subset of the given SO Nos already present in order_headers."""
    if not pos:
        return set()
    with _conn() as (cur, d):
        ot, ph = d['orders'], d['ph']
        marks = ','.join([ph] * len(pos))
        cur.execute(f"SELECT po FROM {ot} WHERE po IN ({marks})", tuple(pos))
        return {r[0] for r in cur.fetchall()}


# ── Preview + import ──────────────────────────────────────────────────────

def preview(filepath: str) -> dict:
    """Parse + dedup-classify (new vs already-imported) for the review page."""
    p = parse(filepath)
    if not p['ok']:
        return p
    ensure_columns()
    rows = p['rows']
    seen = existing_pos([r['so_no'] for r in rows])
    for r in rows:
        r['is_new'] = r['so_no'] not in seen
    new_rows = [r for r in rows if r['is_new']]
    channels = sorted({r['channel'] for r in rows})
    return {
        'ok': True, 'rows': rows,
        'summary': {
            'total': len(rows), 'new': len(new_rows),
            'dup': len(rows) - len(new_rows),
            'channels': channels,
            'qty': sum(r['qty'] for r in new_rows),
            'value': sum(r['order_value'] for r in new_rows),
        },
    }


def do_import(filepath: str) -> dict:
    """Insert the NEW headers under one ERP_IMPORT run. Returns
    {ok, run_id, imported, skipped, error}."""
    pv = preview(filepath)
    if not pv['ok']:
        return pv
    new_rows = [r for r in pv['rows'] if r['is_new']]
    if not new_rows:
        return {'ok': True, 'run_id': None, 'imported': 0,
                'skipped': len(pv['rows'])}

    run_ts = _dt.datetime.now()
    src = os.path.basename(filepath)
    total_qty = sum(r['qty'] for r in new_rows)
    total_val = sum(r['order_value'] for r in new_rows)
    n_channels = len({r['channel'] for r in new_rows})

    try:
        with _conn_tx() as (cur, d):   # atomic: runs row + its headers commit together (no orphan run)
            ph = d['ph']
            # runs row
            cur.execute(
                f"INSERT INTO runs (run_ts, mode, source, marketplaces, "
                f"total_pos, total_items, total_qty, total_value, "
                f"consolidated_path, tracker_path) VALUES "
                f"({ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph})",
                (run_ts, 'MANUAL', f"ERP import: {src}", n_channels,
                 len(new_rows), len(new_rows), total_qty, total_val, '', ''))
            run_id = cur.lastrowid

            cols = ', '.join(_HDR_COLS)
            phs = ', '.join([ph] * len(_HDR_COLS))
            payload = []
            for r in new_rows:
                payload.append((
                    run_id, run_ts, 'MANUAL', SEGMENT_OFFLINE,
                    r['channel'], r['channel'], r['so_no'], r['location'],
                    r['warehouse'], r['po_date'], r['exp_date'],
                    r['order_type'], None, r['qty'], r['order_value'], src,
                    r['external_doc'],
                ))
            cur.executemany(
                f"INSERT INTO order_headers ({cols}) VALUES ({phs})", payload)
        return {'ok': True, 'run_id': run_id, 'imported': len(new_rows),
                'skipped': len(pv['rows']) - len(new_rows)}
    except Exception as e:  # noqa: BLE001
        return {'ok': False, 'error': f"{type(e).__name__}: {e}"}
