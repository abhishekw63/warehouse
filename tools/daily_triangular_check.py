#!/usr/bin/env python
"""
Daily D365 triangular reconciliation  —  D365  vs  our system/dump  vs  source.

WHAT IT CHECKS (per the daily SOP):
  • PO / SO count          D365 headers  vs  our completed dumps
  • Quantity               header total + per-SKU line qty (exact, unit-level)
  • Ship-to (address)      D365 Ship-to Code  vs  our dump's Ship-to Code
  • Pincode                D365 Ship-to Postcode  vs  ship_to_mapping (when we store it)
  • Value (inc GST)        D365 Total Amount Incl. GST  vs  our INCLUDED (pushed) value
                           — compared apples-to-apples on the post-exclusion lines
                           (NOT the full PO), so dropped-line value is not a phantom Δ
  • Dropped lines          2nd sheet: every unit NOT pushed (EXCLUDE / unresolved
                           mismatch) PO-wise · line-wise · value-wise — 360° safety net

WHY IT'S SAFE: read-only. Nothing is written to the business DB. It only reads
the D365 export, our *_completed.xlsx dumps, and ship_to_mapping.

USAGE (from the repo root, venv active):
  python tools/daily_triangular_check.py "<folder with the D365 + completed files>"
  python tools/daily_triangular_check.py "<folder>" --so "Sales Orders*.xlsx" --lines "Sales Lines*.xlsx"

The folder must contain:
  - the D365  Sales Orders*.xlsx  (headers)                [required]
  - the D365  Sales Lines*.xlsx   (lines)                  [optional; enables per-SKU]
  - our  *_completed.xlsx  dumps for that day              [required]
A reconciliation workbook  Triangular_Reconciliation_<today>.xlsx  is saved beside them.

Notes:
  - Flipkart-TO (transfer orders) dumps are skipped — they are not Sales Orders.
  - D365 exports sometimes include other customers; join is by PO so that's harmless.
  - Online B2B compares against the FINAL (post-exclusion) dump, so the qty = what
    actually reached D365 (excluded/issue lines are the difference from the raw order).
"""
import argparse
import datetime as _dt
import glob
import os
import sys
from collections import Counter, defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'renee_cosmetics.settings')
import django  # noqa: E402
django.setup()
from online_b2b.services.order_db import _conn  # noqa: E402


def s(x):
    return str(x).strip() if x is not None else ''


def f(x):
    try:
        return float(str(x).strip())
    except (TypeError, ValueError):
        return 0.0


def _sheet(wb, name):
    return list(wb[name].iter_rows(values_only=True)) if name in wb.sheetnames else []


def _hdr(rows):
    return {s(c): i for i, c in enumerate(rows[0])} if rows else {}


# ── D365 side ────────────────────────────────────────────────────────────────
def load_d365(so_path, lines_path):
    """Return (per_po, per_line, so2po). per_po keyed by SO 'No.'; per_line keyed
    by (PO, item) where PO is the External Doc (falls back to 'No.')."""
    wso = openpyxl.load_workbook(so_path, data_only=True, read_only=True)
    sr = _sheet(wso, wso.sheetnames[0]); h = _hdr(sr)

    def H(*n):
        for x in n:
            if x in h:
                return h[x]
        return None
    c_no, c_ext = H('No.'), H('External Document No.')
    per_po, so2po = {}, {}
    for r in sr[1:]:
        if not r or s(r[c_no]) == '':
            continue
        no = s(r[c_no]); po = s(r[c_ext]) if c_ext is not None else no
        so2po[no] = po or no
        per_po[po or no] = {
            'so': no, 'cust': s(r[H('Sell-to Customer No.')]) if H('Sell-to Customer No.') is not None else '',
            'ship': s(r[H('Ship-to Code')]) if H('Ship-to Code') is not None else '',
            'pin': s(r[H('Ship-to Postcode')]) if H('Ship-to Postcode') is not None else '',
            'qty': int(f(r[H('Total Quantity')])) if H('Total Quantity') is not None else 0,
            'incl': f(r[H('Total Amount Incl. GST')]) if H('Total Amount Incl. GST') is not None else 0.0,
        }
    per_line = defaultdict(int)
    item_q = defaultdict(int)
    # Valid scope = only the SOs/POs in TODAY's Sales Orders header — the Sales
    # Lines export is often broader (other SOs/customers), which must not pollute
    # the per-SKU totals.
    valid = set(so2po) | set(so2po.values())
    if lines_path:
        wl = openpyxl.load_workbook(lines_path, data_only=True, read_only=True)
        lr = _sheet(wl, wl.sheetnames[0]); lh = _hdr(lr)

        def L(*n):
            for x in n:
                if x in lh:
                    return lh[x]
            return None
        for r in lr[1:]:
            if not r or L('No.') is None or s(r[L('No.')]) == '':
                continue
            doc = s(r[L('Document No.')]) if L('Document No.') is not None else ''
            if doc not in valid:                         # line not in today's SOs → skip
                continue
            po = so2po.get(doc, doc)                      # map SO -> PO
            it = s(r[L('No.')]); q = int(f(r[L('Quantity')]))
            per_line[(po, it)] += q
            item_q[it] += q
    return per_po, per_line, item_q


# ── Our side (completed dumps) ───────────────────────────────────────────────
def load_ours(folder):
    per_po, per_line, item_q = {}, defaultdict(int), defaultdict(int)
    for wbp in glob.glob(os.path.join(folder, '*_completed.xlsx')):
        base = os.path.basename(wbp)
        if 'Flipkart-TO' in base:                        # transfer orders, not SOs
            continue
        wb = openpyxl.load_workbook(wbp, data_only=True, read_only=True)
        if not {'Headers (SO)', 'Summary', 'Lines (SO)'} <= set(wb.sheetnames):
            continue
        hd = _sheet(wb, 'Headers (SO)'); hh = _hdr(hd)
        so2po = {}
        for r in hd[1:]:
            if r and hh.get('No.') is not None and s(r[hh['No.']]):
                ext = hh.get('External Document No.', hh.get('No.'))
                so2po[s(r[hh['No.']])] = s(r[ext])
        sm = _sheet(wb, 'Summary'); sh = _hdr(sm)
        fq = sh.get('Final Qty (to D365)', sh.get('Total Qty'))
        for r in sm[1:]:
            po_so = s(r[sh['PO']]) if 'PO' in sh else ''
            if not po_so or po_so.upper().startswith('TOTAL') or 'Marketplace:' in po_so:
                continue                                  # skip TOTAL + footer meta rows
            po = so2po.get(po_so, po_so)
            per_po[po] = {'ship': s(r[sh.get('Ship-to', -1)]),
                          'raw': s(r[sh.get('Location (Raw)', -1)]),
                          'qty': int(f(r[fq])) if fq is not None else 0,
                          'incl': f(r[sh.get('Total Amount (Inc GST)', sh.get('Total Amount', -1))])}
        ln = _sheet(wb, 'Lines (SO)'); lh = _hdr(ln)
        for r in ln[1:]:
            if not r or lh.get('No.') is None or s(r[lh['No.']]) == '':
                continue
            po = so2po.get(s(r[lh['Document No.']]), s(r[lh['Document No.']]))
            it = s(r[lh['No.']]); q = int(f(r[lh['Quantity']]))
            if it:
                per_line[(po, it)] += q
                item_q[it] += q
    return per_po, per_line, item_q


def pincodes():
    with _conn() as (cur, d):
        cur.execute("SELECT ship_to, postcode FROM ship_to_mapping WHERE ship_to<>''")
        out = {}
        for a, b in cur.fetchall():
            out.setdefault(s(a), s(b))
    return out


def _gm(c):
    c = str(c or '').upper()
    return (1.28 if '28' in c else 1.18 if '18' in c else 1.12 if '12' in c
            else 1.05 if ('5' in c and '15' not in c and '25' not in c)
            else 1.03 if '3' in c else 1.0 if '0' in c else 1.18)


def _line_val(oland, up, gst, q):
    """inc-GST value of a line: our_landing x qty (landing is already inc-GST),
    else unit_price x qty x (1+GST)."""
    if oland not in (None, 0, '0'):
        return float(oland) * q
    if up not in (None, ''):
        return float(up) * q * _gm(gst)
    return 0.0


_AFF = {'MISMATCH', 'NOT_IN_MASTER'}


def _dropped(status, action):
    """A line is dropped from the D365 push if it's EXCLUDE-actioned, or an
    unresolved MISMATCH/NOT_IN_MASTER (not INCLUDE/OVERRIDE-decided)."""
    act = s(action).upper()
    return act == 'EXCLUDE' or (s(status) in _AFF and act not in ('INCLUDE', 'OVERRIDE'))


def db_lines_for_pos(pos):
    """Every order_lines_full row for these POs, keeping only each PO's LATEST run
    (re-run safe — a superseded earlier run must not double-count)."""
    pos = sorted({s(p) for p in pos if s(p)})
    if not pos:
        return []
    with _conn() as (cur, d):
        fmt = ','.join(['%s'] * len(pos))
        cur.execute(f"""SELECT run_id, marketplace, po, item_no, ean, description, qty,
                               our_landing, unit_price, gst_code, status, action
                        FROM order_lines_full WHERE po IN ({fmt})""", pos)
        keys = [c[0] for c in cur.description]
        rows = [dict(zip(keys, r)) for r in cur.fetchall()]
    latest = {}
    for r in rows:
        po = s(r['po'])
        if po not in latest or r['run_id'] > latest[po]:
            latest[po] = r['run_id']
    return [r for r in rows if r['run_id'] == latest[s(r['po'])]]


def split_included_dropped(db_rows):
    """Latest-run rows → (included inc-GST value per PO, list of dropped lines).
    'included' is the honest apples-to-apples partner for D365's Total Amount
    Incl. GST — it ties to what actually reached D365, not the full PO."""
    inc = defaultdict(float)
    dropped = []
    for r in db_rows:
        q = int(r['qty'] or 0)
        v = _line_val(r['our_landing'], r['unit_price'], r['gst_code'], q)
        if _dropped(r['status'], r['action']):
            reason = ('EXCLUDE' if s(r['action']).upper() == 'EXCLUDE'
                      else f"{s(r['status'])} (unresolved)")
            dropped.append({'po': s(r['po']), 'mp': s(r['marketplace']), 'item': s(r['item_no']),
                            'ean': s(r['ean']), 'desc': s(r['description']), 'qty': q,
                            'val': v, 'reason': reason})
        else:
            inc[s(r['po'])] += v
    return inc, dropped


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('folder')
    ap.add_argument('--so', default='Sales Orders*.xlsx')
    ap.add_argument('--lines', default='Sales Lines*.xlsx')
    a = ap.parse_args()
    so = next(iter(glob.glob(os.path.join(a.folder, a.so))), None)
    lines = next(iter(glob.glob(os.path.join(a.folder, a.lines))), None)
    if not so:
        sys.exit('No D365 "Sales Orders*.xlsx" found in the folder.')
    print('D365 SO   :', os.path.basename(so))
    print('D365 Lines:', os.path.basename(lines) if lines else '(none — per-SKU skipped)')

    d_po, d_line, d_item = load_d365(so, lines)
    o_po, o_line, o_item = load_ours(a.folder)
    pin_of = pincodes()

    # our INCLUDED (pushed-to-D365) value + the dropped lines, from the DB latest
    # run per PO. Value MUST be compared included-vs-D365 (not full PO), else the
    # dropped lines' value shows up as a phantom Δ.
    db_rows = db_lines_for_pos(set(d_po) | set(o_po))
    inc_val, dropped = split_included_dropped(db_rows)

    # ── per-PO: qty · ship-to · pincode · value ──
    recs = []
    for po in sorted(set(d_po) | set(o_po)):
        dd, oo = d_po.get(po), o_po.get(po)
        flags = []
        if not dd:
            flags.append('missing in D365')
        if not oo:
            flags.append('missing in ours')
        if dd and oo and dd['qty'] != oo['qty']:
            flags.append(f"QTY {dd['qty']}vs{oo['qty']}")
        if dd and oo and dd['ship'] and oo['ship'] and dd['ship'] != oo['ship']:
            flags.append(f"SHIP {oo['ship']}vs{dd['ship']}")
        opin = pin_of.get((oo or {}).get('ship', ''), '') or pin_of.get((dd or {}).get('ship', ''), '')
        if dd and opin and dd['pin'] and opin != dd['pin']:
            flags.append(f"PIN {opin}vs{dd['pin']}")
        # our value = INCLUDED (pushed) value from the DB; Summary full amount is a
        # fallback only when the PO isn't in the DB.
        ov = inc_val.get(po, (oo or {}).get('incl', 0.0))
        recs.append({'po': po, 'dq': (dd or {}).get('qty', 0), 'oq': (oo or {}).get('qty', 0),
                     'dv': (dd or {}).get('incl', 0.0), 'ov': ov,
                     'ocode': (oo or {}).get('ship', ''), 'dcode': (dd or {}).get('ship', ''),
                     'opin': opin, 'dpin': (dd or {}).get('pin', ''),
                     'status': 'OK' if not flags else ' · '.join(flags)})
    nok = sum(1 for r in recs if r['status'] == 'OK')
    print(f"\n=== PER-PO ===  D365={len(d_po)}  ours={len(o_po)}  OK={nok}  FLAGGED={len(recs)-nok}")
    print(f"Qty (final):  D365={sum(r['dq'] for r in recs):,}  ours={sum(r['oq'] for r in recs):,}")
    print(f"Value(incGST, INCLUDED):D365={sum(r['dv'] for r in recs):,.2f}  "
          f"ours={sum(r['ov'] for r in recs):,.2f}  Δ={sum(r['dv']-r['ov'] for r in recs):,.2f}")
    for r in recs:
        if r['status'] != 'OK':
            print(f"  {r['po']:<16} {r['status']}")
    dq = sum(r['qty'] for r in dropped); dv = sum(r['val'] for r in dropped)
    print(f"\n=== DROPPED (not pushed) ===  lines={len(dropped)}  qty={dq:,}  value(incGST)={dv:,.2f}")

    # ── per-SKU: item-level aggregate (key-independent, unit-exact) ──
    if lines:
        allit = set(d_item) | set(o_item); mis = 0
        for it in sorted(allit):
            if d_item[it] != o_item.get(it, 0):
                mis += 1
                if mis <= 25:
                    print(f"  ITEM {it}: D365={d_item[it]} ours={o_item.get(it,0)}")
        print(f"\n=== PER-SKU (item qty) ===  items={len(allit)}  mismatches={mis}"
              + ('  → ALL MATCH' if mis == 0 else ''))
        print(f"  D365 line qty={sum(d_item.values()):,}  ours line qty={sum(o_item.values()):,}")

    # ── pincode summary ──
    pm = pmm = pnm = 0
    for po, dd in d_po.items():
        opin = pin_of.get(dd['ship'], '')
        if not opin:
            pnm += 1
        elif opin == dd['pin']:
            pm += 1
        else:
            pmm += 1
    print(f"\n=== POSTCODE ===  matched={pm}  mismatched={pmm}  not-stored-our-side={pnm}")
    if pnm and not pmm:
        print("  (online B2B stores ship-to CODE, not pincode — code match above IS the address check)")

    # ── save workbook ──
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    NAVY = PatternFill('solid', fgColor='1A237E'); RED = PatternFill('solid', fgColor='FDE7E7')
    HF = Font(bold=True, color='FFFFFF'); thin = Side(style='thin', color='D9DEE8'); BD = Border(thin, thin, thin, thin)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Per PO'
    cols = ['PO', 'D365 Qty', 'Our Qty', 'D365 Value(incGST)', 'Our Value', 'Δ Value',
            'Our Ship-to', 'D365 Ship-to', 'Our Pin', 'D365 Pin', 'Status']
    ws.append(cols)
    for c in ws[1]:
        c.fill = NAVY; c.font = HF; c.alignment = Alignment('center', 'center', wrap_text=True); c.border = BD
    for r in recs:
        ws.append([r['po'], r['dq'], r['oq'], round(r['dv'], 2), round(r['ov'], 2), round(r['dv'] - r['ov'], 2),
                   r['ocode'], r['dcode'], r['opin'], r['dpin'], r['status']])
        for c in ws[ws.max_row]:
            c.border = BD
            if r['status'] != 'OK':
                c.fill = RED
    for i, w in enumerate([16, 9, 9, 17, 15, 11, 13, 13, 9, 9, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # ── sheet 2: DROPPED / EXCLUDED lines — the 360° safety net ──
    # Every unit we did NOT push (EXCLUDE or unresolved mismatch), PO-wise +
    # line-wise + value-wise, so a drop is never invisible after the fact.
    ws2 = wb.create_sheet('Dropped Lines')
    AMBER = PatternFill('solid', fgColor='FFF3CD')
    dcols = ['PO', 'Marketplace', 'Item No', 'EAN', 'Description', 'Dropped Qty',
             'Value (inc GST)', 'Reason']
    ws2.append(dcols)
    for c in ws2[1]:
        c.fill = NAVY; c.font = HF; c.alignment = Alignment('center', 'center', wrap_text=True); c.border = BD
    for r in sorted(dropped, key=lambda x: (x['po'], x['item'], x['ean'])):
        ws2.append([r['po'], r['mp'], r['item'], r['ean'], r['desc'], r['qty'],
                    round(r['val'], 2), r['reason']])
        for c in ws2[ws2.max_row]:
            c.border = BD; c.fill = AMBER
    ws2.append(['TOTAL', '', '', '', f"{len(dropped)} lines", dq, round(dv, 2), ''])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True); c.border = BD
    for i, w in enumerate([16, 12, 12, 16, 36, 12, 15, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    out = os.path.join(a.folder, f"Triangular_Reconciliation_{_dt.date.today():%d-%m-%Y}.xlsx")
    wb.save(out)
    print('\nSAVED:', out)


if __name__ == '__main__':
    main()
