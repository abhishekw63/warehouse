"""core.py — everything the offline app does, in one file.

    paths     where to read bundled assets / write settings (works frozen)
    pricing   LR = MRP x margin% ;  CP = LR / (1 + GST)
    master    Item Master + Ship-To, loaded from files the user picks
    channels  read a PO file into plain line dicts
    run       the pipeline: PO file -> priced rows -> Excel

Deliberately one module: the whole app is small enough to read top to bottom.
"""
from __future__ import annotations

import csv
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

APP_NAME = 'OMT Offline'


# ══════════════════════════════════════════════════════════════════════════
# PATHS — survive being frozen into an .exe
# ══════════════════════════════════════════════════════════════════════════
# PyInstaller unpacks a one-file build into a temporary READ-ONLY folder and
# points __file__ at it. So bundled assets come from _MEIPASS, and anything we
# write must go somewhere real or it vanishes when the app closes.

def is_frozen() -> bool:
    return getattr(sys, 'frozen', False)


def resource_path(*parts) -> Path:
    """A read-only asset shipped with the app (the icon)."""
    base = (Path(getattr(sys, '_MEIPASS', Path(sys.executable).parent))
            if is_frozen() else Path(__file__).resolve().parent)
    return base.joinpath(*parts)


def data_dir() -> Path:
    """Where we may WRITE. Beside the .exe when possible (keeps it portable),
    else %LOCALAPPDATA% (an install under Program Files is not writable)."""
    if not is_frozen():
        return Path(__file__).resolve().parent
    beside = Path(sys.executable).resolve().parent
    try:
        probe = beside / '.write_test'
        probe.write_text('', encoding='utf-8')
        probe.unlink()
        return beside
    except OSError:
        d = Path(os.environ.get('LOCALAPPDATA') or Path.home()) / APP_NAME.replace(' ', '')
        d.mkdir(parents=True, exist_ok=True)
        return d


def settings_file() -> Path:
    return data_dir() / 'settings.json'


def reference_dir() -> Path:
    """Where the app keeps ITS OWN copy of the two master files.

    The point: you load them once. The originals live in OneDrive/Downloads and
    may move, be renamed, or be offline — so we copy them in and read the copy
    from then on. Browse is only for *updating*, never a per-run chore.
    """
    d = data_dir() / 'reference'
    d.mkdir(parents=True, exist_ok=True)
    return d


def cache_reference(src, kind) -> Path:
    """Copy a picked master file into the app. -> the cached path.

    `kind` is 'item_master' or 'ship_to'. The extension is preserved so the same
    reader handles it. Falls back to the original path if the copy fails.
    """
    import shutil
    src = Path(src)
    try:
        dest = reference_dir() / f'{kind}{src.suffix.lower()}'
        # Clear a stale copy in the other format (xlsx -> csv and back).
        for old in reference_dir().glob(f'{kind}.*'):
            if old != dest:
                try:
                    old.unlink()
                except OSError:
                    pass
        if src.resolve() != dest.resolve():
            shutil.copy2(src, dest)          # copy2 keeps the source mtime
        return dest
    except OSError:
        return src


def file_stamp(path) -> str:
    """'01-09-2026 15:08' — when the underlying data was last written."""
    try:
        return datetime.fromtimestamp(Path(path).stat().st_mtime)\
                       .strftime('%d-%m-%Y %H:%M')
    except OSError:
        return ''


# ══════════════════════════════════════════════════════════════════════════
# PRICING — the only price rule there is
# ══════════════════════════════════════════════════════════════════════════
# One visible margin per channel. No per-SKU exceptions hiding behind it.

DEFAULT_MARGINS = {'Blinkit': 70, 'RK': 70, 'Swiggy': 80, 'GT Mass': 70}


def gst_rate(code):
    """'G-18-S' -> 0.18, 'G-5' -> 0.05, '0-G' -> 0.0.  None if unrecognised.

    A port of the engine's gst_divisor cascade — the ORDER matters ('G-5' must
    not be caught by the 18/12 tests), so it stays a sequence of string tests
    rather than a dict or a regex.
    """
    if code is None:
        return None
    s = str(code).strip().upper()
    if not s or s == 'NAN':
        return 0.0
    if s in ('0-G', 'G-0', 'G-0-S', '0'):
        return 0.0
    if s in ('G-3', 'G-3-S'):
        return 0.03
    if '5' in s and '18' not in s and '12' not in s:
        return 0.05
    if '12' in s:
        return 0.12
    if '18' in s:
        return 0.18
    return None


def compute(mrp, code, margin_pct):
    """Price one line -> (lr, cp, issue). Never guesses silently: anything we
    cannot price honestly comes back with an issue instead of a number."""
    try:
        mrp_v = float(mrp)
    except (TypeError, ValueError):
        return None, None, 'MRP missing'
    if mrp_v <= 0:
        return None, None, 'MRP is zero'
    try:
        m = float(margin_pct)
    except (TypeError, ValueError):
        return None, None, 'Margin invalid'
    if m <= 0:
        return None, None, 'Margin invalid'

    lr = round(mrp_v * (m / 100.0), 2)
    rate, issue = gst_rate(code), ''
    if rate is None:
        rate, issue = 0.18, 'GST code missing (assumed 18%)'
    return lr, round(lr / (1.0 + rate), 2), issue


# ══════════════════════════════════════════════════════════════════════════
# MASTER DATA — Item Master + Ship-To, from files the USER picks
# ══════════════════════════════════════════════════════════════════════════
# Nothing is baked in. The two boxes are how the app stays current.
# Accepts the web app's exports (/b2b/item-master/export/, /b2b/ship-to/export/)
# and the desktop workbooks (Items March.xlsx, Ship to B2B.xlsx).

def _norm(name) -> str:
    """'Item No' / 'item_no' / 'ITEMNO' -> 'itemno'."""
    return ''.join(ch for ch in str(name).lower() if ch.isalnum())


def _pick(df, *aliases):
    """The real column label matching the first alias present."""
    have = {}
    for c in df.columns:
        have.setdefault(_norm(c), c)
    for a in aliases:
        if a in have:
            return have[a]
    return None


def _read_any(path, prefer_sheets=()):
    """csv / xlsx / xls -> DataFrame of strings, preferring a named sheet."""
    p = Path(path)
    if p.suffix.lower() == '.csv':
        return pd.read_csv(p, dtype=str, keep_default_na=False)
    xl = pd.ExcelFile(p)
    want = {_norm(x) for x in prefer_sheets}
    sheet = next((s for s in xl.sheet_names if _norm(s) in want), xl.sheet_names[0])
    return xl.parse(sheet, dtype=str).fillna('')


def clean_code(v) -> str:
    """Fold an EAN/item code to one key. The same code arrives as '8906...',
    '8906....0' or '8.906e+12'; all must match or a lookup silently misses."""
    if v is None:
        return ''
    s = str(v).strip()
    if not s:
        return ''
    low = s.lower()
    if low.endswith('.0') or 'e+' in low:
        try:
            s = str(int(float(s)))
        except (TypeError, ValueError):
            pass
    return s.strip().upper()


class ItemMaster:
    """Lookup by EAN (what PO files carry) or by our item number."""

    def __init__(self, rows=None):
        self.rows = rows or []
        self._by_ean, self._by_item = {}, {}
        for r in self.rows:
            if r['ean']:
                self._by_ean.setdefault(clean_code(r['ean']), r)
            if r['item_no']:
                self._by_item.setdefault(clean_code(r['item_no']), r)

    def __len__(self):
        return len(self.rows)

    def lookup(self, ean=None, item_no=None):
        if ean:
            hit = self._by_ean.get(clean_code(ean))
            if hit:
                return hit
        return self._by_item.get(clean_code(item_no)) if item_no else None


def load_items(path) -> ItemMaster:
    df = _read_any(path, ('Item Master', 'ItemMaster', 'Items'))
    c_item = _pick(df, 'itemno', 'itemnumber', 'itemcode', 'item')
    c_ean = _pick(df, 'ean', 'eancode', 'barcode', 'upc')
    c_mrp = _pick(df, 'mrp', 'mrpvalue')
    c_gst = _pick(df, 'gstcode', 'gst', 'gstgroup', 'gstgroupcode')
    c_desc = _pick(df, 'description', 'itemdescription', 'name')

    missing = [n for n, c in (('Item No', c_item), ('MRP', c_mrp),
                              ('GST Code', c_gst)) if c is None]
    if missing:
        raise ValueError('Item Master is missing: ' + ', '.join(missing) +
                         '\nColumns found: ' +
                         ', '.join(str(c) for c in df.columns[:14]))

    rows = []
    for rec in df.to_dict('records'):
        item_no = str(rec.get(c_item) or '').strip()
        ean = str(rec.get(c_ean) or '').strip() if c_ean else ''
        if not item_no and not ean:
            continue
        rows.append({'item_no': item_no, 'ean': ean,
                     'description': str(rec.get(c_desc) or '').strip() if c_desc else '',
                     'mrp': rec.get(c_mrp),
                     'gst_code': str(rec.get(c_gst) or '').strip()})
    return ItemMaster(rows)


class ShipTo:
    """Delivery location -> customer no + ship-to code."""

    def __init__(self, rows=None):
        self.rows = rows or []
        self._by_loc = {}
        for r in self.rows:
            k = self._key(r['del_location'])
            if k:
                self._by_loc.setdefault(k, r)

    @staticmethod
    def _key(v) -> str:
        """Drop spaces AND hyphens, like the engine, so 'BCPL - Bengaluru B3'
        matches 'BCPL Bengaluru B3'."""
        return ''.join(ch for ch in str(v or '').lower() if ch.isalnum())

    def __len__(self):
        return len(self.rows)

    def parties(self):
        return sorted({r['party'] for r in self.rows if r['party']})

    def lookup(self, location):
        k = self._key(location)
        if not k:
            return None
        hit = self._by_loc.get(k)
        if hit:
            return hit
        for key, row in self._by_loc.items():
            if k in key or key in k:
                return row
        return None


class SkuMap:
    """Vendor SKU code -> EAN / Item No.

    Swiggy punches carry ``SkuCode`` and no EAN at all, so without this map every
    Swiggy line reads as 'Not in item master'. Chain: SkuCode -> EAN -> item.
    """

    def __init__(self, rows=None):
        self.rows = rows or []
        self._by_code = {}
        for r in self.rows:
            k = clean_code(r['sku_code'])
            if k:
                self._by_code.setdefault(k, r)

    def __len__(self):
        return len(self.rows)

    def lookup(self, sku_code):
        return self._by_code.get(clean_code(sku_code))


def load_sku_map(path) -> SkuMap:
    """Load a vendor SKU map (Swiggy). Needs a code column plus EAN or Item No."""
    df = _read_any(path, ('Swiggy', 'SKU Map', 'Channel SKU Map', 'Sku Map'))
    c_code = _pick(df, 'skucode', 'code', 'vendorskucode', 'vendorcode',
                   'channelsku', 'sku')
    c_ean = _pick(df, 'ean', 'eancode', 'barcode', 'upc')
    c_item = _pick(df, 'itemno', 'itemnumber', 'itemcode')
    if c_code is None or (c_ean is None and c_item is None):
        raise ValueError('SKU map needs a SKU Code column plus an EAN or Item No.'
                         '\nColumns found: ' +
                         ', '.join(str(c) for c in df.columns[:14]))
    rows = []
    for rec in df.to_dict('records'):
        code = str(rec.get(c_code) or '').strip()
        if not code:
            continue
        rows.append({'sku_code': code,
                     'ean': str(rec.get(c_ean) or '').strip() if c_ean else '',
                     'item_no': str(rec.get(c_item) or '').strip() if c_item else ''})
    return SkuMap(rows)


def load_shipto(path) -> ShipTo:
    df = _read_any(path, ('Ship To', 'ShipTo', 'Ship to B2B', 'Mapping'))
    c_party = _pick(df, 'party', 'partyname', 'marketplace', 'channel')
    c_loc = _pick(df, 'dellocation', 'deliverylocation', 'location', 'facility')
    c_cust = _pick(df, 'custno', 'customerno', 'selltocustomerno')
    c_ship = _pick(df, 'shipto', 'shiptocode')
    if c_loc is None or c_ship is None:
        raise ValueError('Ship-To file needs a Del Location and a Ship-to column.'
                         '\nColumns found: ' +
                         ', '.join(str(c) for c in df.columns[:14]))
    rows = []
    for rec in df.to_dict('records'):
        loc = str(rec.get(c_loc) or '').strip()
        if not loc:
            continue
        rows.append({'party': str(rec.get(c_party) or '').strip() if c_party else '',
                     'del_location': loc,
                     'cust_no': str(rec.get(c_cust) or '').strip() if c_cust else '',
                     'ship_to': str(rec.get(c_ship) or '').strip()})
    return ShipTo(rows)


# ══════════════════════════════════════════════════════════════════════════
# CHANNELS — read a PO file into line dicts
# ══════════════════════════════════════════════════════════════════════════
# Alias matching rather than fixed columns: real PO files drift, and a renamed
# header should not crash the app.

CHANNELS = ['Blinkit', 'RK', 'Swiggy', 'GT Mass']

_ALIASES = {
    'po':          ['ponumber', 'po', 'pono', 'purchaseorder', 'ordernumber'],
    # Blinkit -> 'upc' · RK -> 'External ID' · Swiggy sometimes 'EAN'
    'ean':         ['ean', 'upc', 'externalid', 'eancode', 'barcode', 'gtin'],
    # GT Mass -> 'BC Code' (that IS our item no) · Swiggy -> 'SkuCode'
    'item_code':   ['bccode', 'itemno', 'itemcode', 'sku', 'skucode', 'articlecode'],
    # RK calls it 'Product name'; Blinkit 'name'; Swiggy 'SkuDescription'.
    'description': ['skudescription', 'description', 'productname', 'name',
                    'itemname', 'articledescription', 'title'],
    # RK counts 'Accepted quantity', NOT 'Requested quantity'
    'qty':         ['acceptedquantity', 'orderedqty', 'unitsordered', 'orderqty',
                    'qty', 'quantity', 'poqty'],
    # The vendor's own cost on the punch — drives 'Diffn with Cost'.
    # Blinkit -> 'cost_price' · RK -> 'Cost' · Swiggy -> 'UnitBasedCost'
    'vendor_cost': ['costprice', 'unitbasedcost', 'cost', 'supplierunitprice',
                    'landingrate', 'basicprice', 'clp', 'unitprice'],
    # Delivery point — matched against the Ship-To mapping.
    # Blinkit -> 'facility_name' · RK -> 'Ship-to location' · Swiggy -> 'FacilityName'
    'location':    ['facilityname', 'shiptolocation', 'location', 'facility',
                    'deliverylocation', 'sitename', 'storename', 'warehouse'],
}

_STATUS_FILTER = {'Swiggy': ('status', 'CONFIRMED')}


def _read_po(path, channel) -> pd.DataFrame:
    p = Path(path)
    ext = p.suffix.lower()
    if ext == '.csv':
        return pd.read_csv(p, dtype=str, keep_default_na=False)
    if ext == '.xlsb':
        return pd.read_excel(p, dtype=str, engine='pyxlsb')
    if channel == 'GT Mass':
        return _read_gt_mass(p)
    return pd.read_excel(p, dtype=str).fillna('')


def _read_gt_mass(p: Path) -> pd.DataFrame:
    """GT Mass sheets put distributor/city/PO metadata ABOVE the real header,
    so the header row is found, not assumed."""
    raw = pd.read_excel(p, header=None, dtype=str).fillna('')
    for i in range(min(len(raw), 40)):
        cells = {_norm(v) for v in raw.iloc[i].tolist() if str(v).strip()}
        if 'bccode' in cells and any('orderqty' in c or 'orderquantity' in c
                                     for c in cells):
            df = raw.iloc[i + 1:].copy()
            df.columns = [str(c).strip() for c in raw.iloc[i].tolist()]
            return df.reset_index(drop=True)
    raise ValueError("Not a GT Mass sheet — no header row with 'BC Code' and "
                     "'Order Qty' in the first 40 rows.")


def read_po(path, channel):
    """-> (lines, warnings). Nothing is dropped silently."""
    warnings = []
    df = _read_po(path, channel)
    if df.empty:
        return [], ['File has no rows.']

    if channel in _STATUS_FILTER:
        col, keep = _STATUS_FILTER[channel]
        actual = next((c for c in df.columns if _norm(c) == col), None)
        if actual is not None:
            before = len(df)
            df = df[df[actual].astype(str).str.strip().str.upper() == keep]
            if before - len(df):
                warnings.append(f'{before - len(df)} row(s) skipped — not {keep}.')

    cols = {}
    have = {}
    for c in df.columns:
        have.setdefault(_norm(c), c)
    for field, names in _ALIASES.items():
        for want in names:
            if want in have:
                cols[field] = have[want]
                break

    # Only the identifier is essential. PO / qty / description are context, not
    # output columns — a file without them is fine, so don't nag about them.
    if 'ean' not in cols and 'item_code' not in cols:
        raise ValueError('No EAN/barcode or item-code column found.\nColumns: ' +
                         ', '.join(str(c) for c in df.columns[:12]))

    lines = []
    for rec in df.to_dict('records'):
        get = lambda f: rec.get(cols[f]) if f in cols else None   # noqa: E731
        ean, code = get('ean'), get('item_code')
        if not str(ean or '').strip() and not str(code or '').strip():
            continue
        try:
            qty = int(float(str(get('qty')).strip()))
        except (TypeError, ValueError):
            qty = None
        try:
            vcost = float(str(get('vendor_cost')).strip())
        except (TypeError, ValueError):
            vcost = None
        lines.append({'po': str(get('po') or '').strip(),
                      'ean': str(ean or '').strip(),
                      'item_code': str(code or '').strip(),
                      'description': str(get('description') or '').strip(),
                      'location': str(get('location') or '').strip(),
                      'qty': qty, 'vendor_cost': vcost})

    if len(df) - len(lines) > 0:
        warnings.append(f'{len(df) - len(lines)} row(s) skipped — no EAN or item code.')
    return lines, warnings


# ══════════════════════════════════════════════════════════════════════════
# RUN — the pipeline
# ══════════════════════════════════════════════════════════════════════════

# Output columns — the same block the web app's Raw Data sheet shows, so the
# offline sheet reads identically to the one people already know. The Landing
# heading carries the margin actually used, exactly like the web app.
_KEYS = ['ean', 'item_no', 'mrp', 'lr', 'gst_code', 'cp', 'diffn', 'issue']


def columns(margin_pct):
    """[(key, heading)] for the given margin."""
    m = int(margin_pct) if float(margin_pct) == int(float(margin_pct)) else margin_pct
    return list(zip(_KEYS, [
        'EAN', 'Item No (Master)', 'MRP', f'Landing ({m}%)', 'GST Code',
        'Cost Price', 'Diffn with Cost', 'Issue',
    ]))


def process(path, channel, margin_pct, items: ItemMaster, sku_map: SkuMap = None):
    """Run one PO file. -> {rows, warnings, counts}"""
    lines, warnings = read_po(path, channel)
    if items is None or not len(items):
        warnings.append('No Item Master loaded — nothing can be priced.')
        items = ItemMaster([])
    # Swiggy punches identify items ONLY by SkuCode, so without the map nothing
    # will resolve. Say so up front rather than emitting 200 identical issues.
    if channel == 'Swiggy' and (sku_map is None or not len(sku_map)) \
            and lines and not any(ln['ean'] for ln in lines):
        warnings.append('This Swiggy file has no EAN column — load the Swiggy SKU '
                        'Map to resolve SkuCode to items.')

    rows = []
    for ln in lines:
        hit = items.lookup(ean=ln['ean'], item_no=ln['item_code'])
        mapped = None
        if hit is None and sku_map is not None and ln['item_code']:
            # Vendor SKU code -> EAN / Item No -> master.
            mapped = sku_map.lookup(ln['item_code'])
            if mapped:
                hit = items.lookup(ean=mapped['ean'], item_no=mapped['item_no'])
        if hit is None:
            # Keep unmatched lines visible — never drop them. Distinguish "the
            # code isn't mapped" from "the mapped item isn't in the master".
            if ln['item_code'] and sku_map is not None and mapped is None \
                    and not ln['ean']:
                why = 'SKU code not in SKU map'
            else:
                why = 'Not in item master'
            rows.append({'ean': ln['ean'] or (mapped or {}).get('ean', ''),
                         'item_no': ln['item_code'],
                         'description': ln['description'], 'mrp': None,
                         'gst_code': '', 'lr': None, 'cp': None, 'diffn': None,
                         'issue': why})
            continue
        lr, cp, issue = compute(hit.get('mrp'), hit.get('gst_code'), margin_pct)
        try:
            mrp = float(hit.get('mrp'))
        except (TypeError, ValueError):
            mrp = None
        # Diffn with Cost — what the vendor put on the punch vs what we compute.
        # Blank when the file carries no cost column; that is not an error.
        vcost, diffn = ln.get('vendor_cost'), None
        if vcost is not None and cp is not None:
            diffn = round(vcost - cp, 2)
            if issue == '' and abs(diffn) > 1.0:
                issue = f'Cost differs by {diffn:+.2f}'
        rows.append({'ean': ln['ean'] or hit.get('ean', ''),
                     'item_no': hit.get('item_no', ''),
                     'description': hit.get('description') or ln['description'],
                     'mrp': mrp, 'gst_code': hit.get('gst_code', ''),
                     'lr': lr, 'cp': cp, 'diffn': diffn, 'issue': issue})

    return {'rows': rows, 'lines': lines, 'warnings': warnings,
            'counts': {'total': len(rows),
                       'priced': sum(1 for r in rows if not r['issue']),
                       'issues': sum(1 for r in rows if r['issue'])}}


# The web app's own palette, so the offline sheet looks like the online one.
CALC_GREEN = '1B5E20'          # the green calc-block header
ISSUE_PINK = 'FDE7E9'


def _style_header(ws, ncols, fill_hex):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    fill = PatternFill('solid', fgColor=fill_hex)
    thin = Side(style='thin', color='FFFFFF')
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _autofit(ws):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 55)
    ws.freeze_panes = 'A2'


def default_name(channel, margin_pct) -> str:
    stamp = datetime.now().strftime('%d_%m_%Y__%H_%M_%S')
    return f"{channel.replace(' ', '_')}_{int(float(margin_pct))}pct_{stamp}.xlsx"


def to_excel(rows, channel, margin_pct, out_path):
    """Write the priced rows to a one-sheet workbook at out_path. -> Path

    Pass a full file path (a directory is also accepted, and gets a stamped
    default name)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    out = Path(out_path)
    if out.is_dir():
        out = out / default_name(channel, margin_pct)
    cols = columns(margin_pct)
    wb = Workbook()
    ws = wb.active
    ws.title = channel[:31]
    ws.append([h for _, h in cols])
    _style_header(ws, len(cols), CALC_GREEN)

    issue_fill = PatternFill('solid', fgColor=ISSUE_PINK)
    for r in rows:
        ws.append([r.get(k) for k, _ in cols])
        if r.get('issue'):
            for c in range(1, len(cols) + 1):
                ws.cell(ws.max_row, c).fill = issue_fill

    for cell in ws['A'][1:]:                                # EAN stays text
        cell.number_format = '@'
    for letter in ('C', 'D', 'F', 'G'):                     # MRP, Landing, CP, Diffn
        for cell in ws[letter][1:]:
            cell.number_format = '0.00'
    _autofit(ws)
    wb.save(out)
    return out


# ══════════════════════════════════════════════════════════════════════════
# THE 7-SHEET WORKBOOK — same shape and styling as the web app
# ══════════════════════════════════════════════════════════════════════════
#   1 Headers (SO)   2 Lines (SO)   3 Summary   4 Validation
#   5 SKU Summary    6 Raw Data     7 Warnings
#
# Golden rule carried over from the web app: Lines (SO) column H (Unit Price) is
# left BLANK. D365 prices from the vendor master; writing our own number there
# is how wrong prices get posted. Our computed CP lives on Validation instead.

HEADER_NAVY = '1A237E'
WARN_ORANGE = 'E65100'
LINE_NO_STEP = 10_000

HEADERS_SO_COLS = [
    'Document Type', 'No.', 'Sell-to Customer No.', 'Ship-to Code',
    'Posting Date', 'Order Date', 'Document Date', 'Invoice From Date',
    'Invoice To Date', 'External Document No.', 'Location Code',
    'Dimension Set ID', 'Supply Type', 'Voucher Narration',
    'Brand (Dimension)', 'Channel (Dimension)', 'Catagory (Dimension)',
    'Geography Code (Dimension)',
]
LINES_SO_COLS = ['Document Type', 'Document No.', 'Line No.', 'Type', 'No.',
                 'Location Code', 'Quantity', 'Unit Price']


def _sheet(wb, title, cols, fill, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(cols)
    _style_header(ws, len(cols), fill)
    return ws


def build_workbook(rows, lines, channel, margin_pct, out_path,
                   warehouse='PICK', shipto: 'ShipTo' = None, warnings=()):
    """Write the full 7-sheet workbook. -> Path

    `rows`  — priced rows from process()
    `lines` — the parsed PO lines (carry po / location / qty)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    out = Path(out_path)
    if out.is_dir():
        out = out / default_name(channel, margin_pct)
    today = datetime.now().strftime('%d-%m-%Y')
    wb = Workbook()

    # Pair each priced row with its source line (same order, 1:1).
    paired = list(zip(rows, lines)) if len(rows) == len(lines) else \
        [(r, {}) for r in rows]

    # ── PO grouping, first-seen order preserved ──
    pos, order = {}, []
    for r, ln in paired:
        po = (ln.get('po') or '').strip() or '(no PO)'
        if po not in pos:
            pos[po] = []
            order.append(po)
        pos[po].append((r, ln))

    def ship_for(ln):
        """Resolve cust no + ship-to from the loaded mapping (blank if none)."""
        if shipto is None:
            return '', ''
        hit = shipto.lookup(ln.get('location') or ln.get('po_location') or '')
        return (hit['cust_no'], hit['ship_to']) if hit else ('', '')

    # ── 1. Headers (SO) — one row per PO ──
    ws = _sheet(wb, 'Headers (SO)', HEADERS_SO_COLS, HEADER_NAVY, first=True)
    for po in order:
        _, ln0 = pos[po][0]
        cust, sto = ship_for(ln0)
        ws.append(['Order', po, cust, sto, today, today, today, today, today,
                   po, warehouse, '', 'B2B', '', '', '', '', ''])
    _autofit(ws)

    # ── 2. Lines (SO) — line no. resets per PO; Unit Price stays BLANK ──
    ws = _sheet(wb, 'Lines (SO)', LINES_SO_COLS, HEADER_NAVY)
    for po in order:
        n = 0
        for r, ln in pos[po]:
            n += LINE_NO_STEP
            ws.append(['Order', po, n, 'Item', r.get('item_no', ''),
                       warehouse, ln.get('qty'), ''])
    _autofit(ws)

    # ── 3. Summary — one row per PO ──
    ws = _sheet(wb, 'Summary', ['PO', 'Location', 'Cust No', 'Ship-to', 'Items',
                                'Total Qty', 'Total Amount', 'Status'], HEADER_NAVY)
    g_qty = g_amt = 0
    for po in order:
        items_n = len(pos[po])
        qty = sum((ln.get('qty') or 0) for _, ln in pos[po])
        amt = sum((r.get('cp') or 0) * (ln.get('qty') or 0) for r, ln in pos[po])
        cust, sto = ship_for(pos[po][0][1])
        bad = any(r.get('issue') for r, _ in pos[po])
        ws.append([po, pos[po][0][1].get('location', ''), cust, sto, items_n,
                   qty, round(amt, 2), 'CHECK' if bad else 'OK'])
        g_qty += qty
        g_amt += amt
    ws.append(['TOTAL', '', '', '', len(rows), g_qty, round(g_amt, 2), ''])
    for c in range(1, 9):
        ws.cell(ws.max_row, c).font = Font(bold=True)
    _autofit(ws)

    # ── 4. Validation — the green calc block (what the screen shows) ──
    cols = columns(margin_pct)
    ws = _sheet(wb, 'Validation', [h for _, h in cols], CALC_GREEN)
    pink = PatternFill('solid', fgColor=ISSUE_PINK)
    for r in rows:
        ws.append([r.get(k) for k, _ in cols])
        if r.get('issue'):
            for c in range(1, len(cols) + 1):
                ws.cell(ws.max_row, c).fill = pink
    for cell in ws['A'][1:]:
        cell.number_format = '@'
    for letter in ('C', 'D', 'F', 'G'):
        for cell in ws[letter][1:]:
            cell.number_format = '0.00'
    _autofit(ws)

    # ── 5. SKU Summary — per item across the whole file ──
    ws = _sheet(wb, 'SKU Summary', ['Item No', 'EAN', 'Description',
                                    'Qty Demanded', '# POs', 'Cost Price',
                                    'Value'], CALC_GREEN)
    agg = {}
    for r, ln in paired:
        k = r.get('item_no') or r.get('ean') or '?'
        a = agg.setdefault(k, {'ean': r.get('ean', ''),
                               'desc': r.get('description', ''),
                               'qty': 0, 'pos': set(), 'cp': r.get('cp')})
        a['qty'] += (ln.get('qty') or 0)
        a['pos'].add((ln.get('po') or '').strip())
    for k, a in sorted(agg.items(), key=lambda kv: -kv[1]['qty']):
        val = round((a['cp'] or 0) * a['qty'], 2) if a['cp'] else None
        ws.append([k, a['ean'], a['desc'], a['qty'], len(a['pos']), a['cp'], val])
    _autofit(ws)

    # ── 6. Raw Data — the punch as read, plus what we computed from it ──
    ws = _sheet(wb, 'Raw Data', ['PO', 'Location', 'EAN', 'SKU / Item Code',
                                 'Description', 'Qty', 'Vendor Cost',
                                 'Item No (Master)', 'MRP',
                                 f'Landing ({int(float(margin_pct))}%)',
                                 'GST Code', 'Cost Price', 'Diffn with Cost'],
                 CALC_GREEN)
    for r, ln in paired:
        ws.append([ln.get('po', ''), ln.get('location', ''), ln.get('ean', ''),
                   ln.get('item_code', ''), ln.get('description', ''),
                   ln.get('qty'), ln.get('vendor_cost'), r.get('item_no', ''),
                   r.get('mrp'), r.get('lr'), r.get('gst_code', ''),
                   r.get('cp'), r.get('diffn')])
    _autofit(ws)

    # ── 7. Warnings — every flagged line, named. Never a silent drop. ──
    ws = _sheet(wb, 'Warnings', ['PO', 'EAN', 'Item No', 'Description', 'Issue'],
                WARN_ORANGE)
    for r, ln in paired:
        if r.get('issue'):
            ws.append([ln.get('po', ''), r.get('ean', ''), r.get('item_no', ''),
                       r.get('description', ''), r['issue']])
    for w in warnings:
        ws.append(['', '', '', '', w])
    if ws.max_row == 1:
        ws.append(['', '', '', '', 'No issues — every line priced cleanly.'])
    _autofit(ws)

    wb.save(out)
    return out


# ══════════════════════════════════════════════════════════════════════════
# TEMPLATES — blank files in the exact shape the boxes expect
# ══════════════════════════════════════════════════════════════════════════
# So an operator can maintain their own Item Master / Ship-To without having to
# guess the column names, and without a web login.

ITEM_MASTER_COLUMNS = ['Item No', 'EAN', 'Description', 'MRP', 'GST Code', 'HSN']
ITEM_MASTER_SAMPLE = [
    ['100504', '8904473105649', 'PRINCESS BY RENEE TROLLEY BAG', 750, 'G-18-S', '42021290'],
    ['200001', '8906121646917', 'DISNEY FROZEN BUBBLE BATH', 250, 'G-18-S', '33073090'],
]

SHIPTO_COLUMNS = ['Party', 'Del Location', 'Cust No', 'Ship-to', 'Name',
                  'Address', 'Postcode', 'City']
SHIPTO_SAMPLE = [
    ['Blink', 'Bhiwandi', '20647', '20647_27', 'BCPL Bhiwandi', 'Warehouse Rd', '421302', 'Bhiwandi'],
    ['Swiggy', 'Bangalore FC', '20003', '20003_1', 'Swiggy BLR', 'FC Road', '560068', 'Bangalore'],
]

SKU_MAP_COLUMNS = ['SKU Code', 'EAN', 'Item No']
SKU_MAP_SAMPLE = [
    ['SWG-100504', '8904473105649', '100504'],
    ['SWG-200001', '8906121646917', '200001'],
]


def _write_template(path, headers, sample, sheet_title, note):
    """Data sheet + a separate 'How to use' sheet.

    The note deliberately does NOT sit under the data: anything in column A gets
    read back as a row, so an in-sheet note would load as a phantom item.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    _style_header(ws, len(headers), CALC_GREEN)
    for row in sample:
        ws.append(row)
    _autofit(ws)

    hlp = wb.create_sheet('How to use')
    hlp['A1'] = f'{sheet_title} — template'
    hlp['A1'].font = Font(bold=True, size=12)
    hlp['A3'] = note
    hlp['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    hlp['A5'] = 'Keep the header row exactly as it is. Extra columns are ignored.'
    hlp['A6'] = f'Keep the data on the "{sheet_title}" sheet.'
    for r in ('A3', 'A5', 'A6'):
        hlp[r].font = Font(color='6C6C72')
    hlp.column_dimensions['A'].width = 90
    hlp.row_dimensions[3].height = 34

    wb.save(path)
    return Path(path)


def write_item_master_template(path):
    return _write_template(
        path, ITEM_MASTER_COLUMNS, ITEM_MASTER_SAMPLE, 'Item Master',
        'Replace the two sample rows with your items. Item No, MRP and GST Code '
        'are required; EAN is what PO files are matched on.')


def write_shipto_template(path):
    return _write_template(
        path, SHIPTO_COLUMNS, SHIPTO_SAMPLE, 'Ship To',
        'Replace the two sample rows with your locations. Del Location and '
        'Ship-to are required — Del Location is matched against the PO file.')


def write_sku_map_template(path):
    return _write_template(
        path, SKU_MAP_COLUMNS, SKU_MAP_SAMPLE, 'Swiggy',
        'Replace the two sample rows with your channel SKU codes. SKU Code is '
        'what the Swiggy punch carries; give it an EAN (or an Item No) so the '
        'line can be matched to the Item Master.')


# ── Standard PO templates, one per channel ────────────────────────────────
# The real column names each channel's punch uses, so an operator can hand-build
# or fix up a PO file the app will read first time.

PO_TEMPLATES = {
    'Blinkit': {
        'cols': ['po_number', 'upc', 'name', 'units_ordered', 'cost_price',
                 'facility_name', 'order_date', 'expiry_date', 'total_amount'],
        'rows': [['PO12345', '8904473105649', 'PRINCESS BY RENEE TROLLEY BAG',
                  10, 444.92, 'Bhiwandi', '01-09-2026', '15-09-2026', 4449.20]],
        'note': 'Blinkit punch (.csv). EAN goes in "upc"; quantity in '
                '"units_ordered"; the vendor cost in "cost_price".',
    },
    'RK': {
        'cols': ['PO', 'Product name', 'External ID', 'Accepted quantity', 'Cost',
                 'Ship-to location', 'Order date', 'Cancellation deadline'],
        'rows': [['5L4LRIND', 'RENEE Midnight Combo', '8904473102495', 48, 829.92,
                  'HBA4', '01-09-2026', '19-09-2026']],
        'note': 'RK POItemExport (.xls). EAN goes in "External ID". Quantity is '
                'read from "Accepted quantity" — NOT "Requested quantity".',
    },
    'Swiggy': {
        'cols': ['PoNumber', 'SkuCode', 'EAN', 'SkuDescription', 'OrderedQty',
                 'UnitBasedCost', 'FacilityName', 'Status', 'PoCreatedAt',
                 'PoExpiryDate'],
        # cost = 750 x 80% / 1.18 — consistent with Swiggy's 80% default, so the
        # sample row comes back clean rather than flagging itself.
        'rows': [['SW-1001', 'SWG-100504', '8904473105649', 'TROLLEY BAG', 12,
                  508.47, 'Bangalore FC', 'CONFIRMED', '01-09-2026', '15-09-2026']],
        'note': 'Swiggy punch (.csv). Only rows with Status = CONFIRMED are '
                'priced. If the file has no EAN column, load the Swiggy SKU Map '
                'so SkuCode can be resolved.',
    },
    'GT Mass': {
        'cols': ['BC Code', 'Article Description', 'EAN', 'Order Qty', 'Tester Qty'],
        'rows': [['100504', 'PRINCESS BY RENEE TROLLEY BAG', '8904473105649', 20, 1],
                 ['200002', 'DISNEY FROZEN BUBBLE BATH', '8906121646856', 8, 0]],
        'note': 'GT Mass sheet. The metadata rows above the header are part of '
                'the format — keep them. BC Code is the item number.',
        'meta': [['Distributor Name', 'ACME TRADERS'],
                 ['City', 'Ahmedabad'],
                 ['State', 'Gujarat'],
                 ['PO Number', 'SO/GTM/1234', '', 'Location', 'AHD'],
                 []],
    },
}


def write_po_template(channel, path):
    """Write a standard PO-file template for one channel. -> Path"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    spec = PO_TEMPLATES[channel]
    wb = Workbook()
    ws = wb.active
    ws.title = channel[:31]

    # GT Mass carries distributor/PO metadata ABOVE the header — reproduce it,
    # because the reader finds the header row by scanning past exactly this.
    meta = spec.get('meta') or []
    for row in meta:
        ws.append(row)
    for r in range(1, len(meta) + 1):
        ws.cell(r, 1).font = Font(bold=True)

    head_at = len(meta) + 1
    ws.append(spec['cols'])
    for c in range(1, len(spec['cols']) + 1):
        cell = ws.cell(head_at, c)
        cell.font = Font(bold=True, color='FFFFFF')
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill('solid', fgColor=CALC_GREEN)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in spec['rows']:
        ws.append(row)
    _autofit(ws)
    ws.freeze_panes = f'A{head_at + 1}'

    hlp = wb.create_sheet('How to use')
    hlp['A1'] = f'{channel} — PO file template'
    hlp['A1'].font = Font(bold=True, size=12)
    hlp['A3'] = spec['note']
    hlp['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    hlp['A5'] = 'Column names are matched loosely — case and spacing do not matter.'
    hlp['A6'] = 'Extra columns are ignored, so a full vendor export works as-is.'
    for r in ('A3', 'A5', 'A6'):
        hlp[r].font = Font(color='6C6C72')
    hlp.column_dimensions['A'].width = 92
    hlp.row_dimensions[3].height = 46
    wb.save(path)
    return Path(path)
